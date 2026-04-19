# UFC fight outcome predictor using a stacked ensemble: LightGBM + CatBoost +
# RandomForest + ExtraTrees, with a simplex-constrained log-loss meta-combiner. Training pipeline:
#   1. load_data         -> parse and clean fight_data.csv
#   2. fix_data_leakage  -> chronologically recompute all fighter stats so no future data leaks
#   3. build_all_features -> 25+ tiers of engineered features (ELO, Glicko-2, SVD, etc.)
#   4. train             -> antisymmetric decomposition, Stability Selection, Optuna tuning, OOF stacking
# At inference, UFCApp drives the tkinter GUI and calls UFCPredictor.predict_matchup().

from tkinter import ttk, filedialog, messagebox
import tkinter as tk
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import openpyxl
from sklearn.metrics import (
    accuracy_score,
    log_loss,
    precision_score,
    recall_score,
    f1_score,
    brier_score_loss,
    roc_auc_score,
    confusion_matrix,
    average_precision_score,
    matthews_corrcoef,
    balanced_accuracy_score,
    classification_report,
)
from sklearn.cluster import KMeans
from sklearn.decomposition import TruncatedSVD
from sklearn.preprocessing import StandardScaler, LabelEncoder
from sklearn.ensemble import RandomForestClassifier
from sklearn.base import BaseEstimator, ClassifierMixin
import pandas as pd
import numpy as np
import joblib
from collections import defaultdict
from datetime import datetime, date
import multiprocessing as mp
import traceback
import threading
import warnings
import hashlib
import copy
import math
import time
import random
import sys
import os

os.environ["PYTHONWARNINGS"] = "ignore"


warnings.filterwarnings("ignore")


# All boosting libraries and Optuna are optional. The pipeline degrades gracefully:
# missing libraries simply aren't added to the ensemble.
try:
    import xgboost as xgb

    HAS_XGB = True
except ImportError:
    HAS_XGB = False

try:
    import lightgbm as lgb

    HAS_LGB = True
except ImportError:
    HAS_LGB = False

try:
    import catboost as cb

    HAS_CAT = True
except ImportError:
    HAS_CAT = False

try:
    import optuna

    optuna.logging.set_verbosity(optuna.logging.WARNING)
    HAS_OPTUNA = True
except ImportError:
    HAS_OPTUNA = False


RANDOM_SEED = 42

# Seed every randomness source so the full training run is deterministic on CPU.
# XGBoost with device='cuda' still has non-deterministic GPU float-reduction order;
# switch to device='cpu' in XGB params to fully eliminate that last source of variance.
random.seed(RANDOM_SEED)
np.random.seed(RANDOM_SEED)
os.environ["PYTHONHASHSEED"] = str(RANDOM_SEED)
os.environ["TF_DETERMINISTIC_OPS"] = "1"

# Use half of available CPUs to avoid starving the OS during long training runs.
SAFE_N_JOBS = max(1, mp.cpu_count() // 2)


if getattr(sys, "frozen", False):
    SCRIPT_DIR = os.path.dirname(sys.executable)
else:
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

DEFAULT_DATA_PATH = os.path.join(SCRIPT_DIR, "fight_data.csv")
MODEL_CACHE_DIR = os.path.join(SCRIPT_DIR, ".model_cache")

# Bump a stage's version when you change code that affects that stage's output.
# Any change invalidates that stage AND all downstream stages.
STAGE_VERSIONS = {
    # WC route prior: chrono timelines (no final-count snapshot leakage)
    "data_features": "12",  # removed noisy composites (ring_rust, peak_score, uncertainty, etc.)
    "preprocessing": "5",  # sample weight half-life 2.5y, floor 0.08
    "winner_model": "26",  # corr dedup, MLP early stop, no GNB, HGB no random ES
}
STAGE_ORDER = [
    "data_features",
    "preprocessing",
    "winner_model",
]

# Explicit dependency DAG — each stage lists its direct parents.
STAGE_DEPS = {
    "data_features": [],
    "preprocessing": ["data_features"],
    "winner_model": ["preprocessing"],
}


def _stage_ancestors(stage_name):
    """Return all ancestors of a stage (including itself) in topological order."""
    visited = []
    stack = [stage_name]
    seen = set()
    while stack:
        s = stack.pop()
        if s in seen:
            continue
        seen.add(s)
        visited.append(s)
        for dep in STAGE_DEPS.get(s, []):
            stack.append(dep)
    # Return in topological order (reverse of DFS post-order)
    return [s for s in STAGE_ORDER if s in seen]


def _stage_cache_key(stage_name, data_hash):
    """Compute a cache key for a stage: hash of data + ancestor versions + own version."""
    ancestors = _stage_ancestors(stage_name)
    parts = [data_hash]
    for s in ancestors:
        parts.append(f"{s}={STAGE_VERSIONS[s]}")
    return hashlib.sha256("|".join(parts).encode()).hexdigest()[:24]


def _save_stage_cache(stage_name, data_hash, payload):
    """Save a stage's output to disk. Removes stale files for this stage first."""
    os.makedirs(MODEL_CACHE_DIR, exist_ok=True)
    key = _stage_cache_key(stage_name, data_hash)
    path = os.path.join(MODEL_CACHE_DIR, f"{stage_name}_{key}.joblib")
    # Clean up old cache files for this stage (different keys = stale)
    import glob as _glob_mod

    for old in _glob_mod.glob(os.path.join(MODEL_CACHE_DIR, f"{stage_name}_*.joblib")):
        if old != path:
            os.remove(old)
    joblib.dump(payload, path, compress=3)
    print_step(f"  Cached stage '{stage_name}' -> {os.path.basename(path)}")


def _load_stage_cache(stage_name, data_hash):
    """Load a stage's cached output. Returns None if cache miss or invalid."""
    key = _stage_cache_key(stage_name, data_hash)
    path = os.path.join(MODEL_CACHE_DIR, f"{stage_name}_{key}.joblib")
    if not os.path.isfile(path):
        return None
    try:
        payload = joblib.load(path)
        print_step(
            f"  Cache hit for stage '{stage_name}' <- {os.path.basename(path)}")
        return payload
    except Exception as e:
        print_step(f"  Cache load failed for '{stage_name}': {e}")
        return None


def _stage_descendants(stage_name):
    """Return this stage and all stages that transitively depend on it."""
    desc = {stage_name}
    changed = True
    while changed:
        changed = False
        for s, deps in STAGE_DEPS.items():
            if s not in desc and any(d in desc for d in deps):
                desc.add(s)
                changed = True
    return desc


def _clear_downstream_caches(stage_name, data_hash):
    """Remove cached files for this stage and all downstream stages."""
    to_clear = _stage_descendants(stage_name)
    for s in to_clear:
        pattern = os.path.join(MODEL_CACHE_DIR, f"{s}_*.joblib")
        import glob as _glob_mod

        for f in _glob_mod.glob(pattern):
            os.remove(f)


def _hash_file(path):
    """Compute SHA-256 hash of a file."""
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def _clean_estimator_for_retrain(name, est, fixed_iters=None):
    from sklearn.base import clone as _clone

    if name == "xgb":
        params = est.get_params(deep=False).copy()
        params.pop("early_stopping_rounds", None)
        params.pop("callbacks", None)
        params.pop("eval_metric", None)
        if fixed_iters is not None:
            params["n_estimators"] = fixed_iters
        return xgb.XGBClassifier(**params)

    if name == "cat":
        params = est.get_params(deep=False).copy()
        params.pop("early_stopping_rounds", None)
        params.pop("use_best_model", None)
        params.pop("eval_set", None)
        if fixed_iters is not None:
            params["iterations"] = fixed_iters
        return cb.CatBoostClassifier(**params)

    if name == "lgb":
        new_est = _clone(est)
        if fixed_iters is not None:
            new_est.set_params(n_estimators=fixed_iters)
        return new_est

    return _clone(est)


def detect_gpu():
    gpu_info = {"xgb": False, "lgb": False, "cat": False}
    X_dummy = np.random.rand(2, 4)
    y_dummy = np.array([0, 1])

    if HAS_XGB:
        try:
            m = xgb.XGBClassifier(
                device="cuda",
                n_estimators=1,
                verbosity=0,
                random_state=RANDOM_SEED,
                eval_metric="logloss",
            )
            m.fit(X_dummy, y_dummy)
            gpu_info["xgb"] = True
        except Exception:
            pass

    # LightGBM is forced to CPU regardless of hardware — GPU mode caused out-of-resources errors.
    gpu_info["lgb"] = False

    if HAS_CAT:
        try:
            m = cb.CatBoostClassifier(
                task_type="GPU",
                iterations=1,
                verbose=0,
                random_seed=RANDOM_SEED,
                allow_writing_files=False,
            )
            m.fit(X_dummy, y_dummy)
            gpu_info["cat"] = True
        except Exception:
            pass

    return gpu_info


def print_section(title):
    width = 70
    print("\n" + "=" * width)
    print(f"  {title}")
    print("=" * width)


def print_step(msg):
    print(f"  >> {msg}")


def print_metric(label, value):
    val_str = str(value)
    padding = max(68 - 2 - 42, len(val_str))
    print(f"  {label:<42}{val_str:>{padding}}")


def print_divider():
    print("  " + "-" * 66)


# FeatureEngineer maintains all stateful rating systems that must be computed
# incrementally in chronological fight order to prevent data leakage:
#   - ELO and Glicko-2 skill ratings (updated fight-by-fight)
#   - Weight-class Z-score normalization (per weight class per year)
#   - Common opponent win/loss matrix
#   - KMeans style clustering (striker / grappler / etc.) and cluster-vs-cluster win rates
class _StabilitySelector:
    """Drop-in replacement that exposes the same interface the rest of the pipeline uses.

    estimator_.feature_importances_ returns the mean selection frequency of the kept
    features, which serves as a stable proxy for relative importance.
    Defined at module level so it is picklable for stage caching.
    """

    def __init__(self, support, sel_freq):
        self.support_ = support
        self.n_features_ = int(support.sum())
        self._sel_freq = sel_freq

    class _FakeEstimator:
        def __init__(self, importances):
            self.feature_importances_ = importances

    @property
    def estimator_(self):
        return self._FakeEstimator(self._sel_freq[self.support_])

    def transform(self, X):
        return X[:, self.support_]


def _defaultdict_list():
    """Picklable factory for defaultdict(lambda: defaultdict(list))."""
    return defaultdict(list)


class FeatureEngineer:
    def __init__(self):

        self.elo_ratings = {}

        self.glicko_ratings = {}
        self.GLICKO_MU = 1500.0
        self.GLICKO_PHI = 200.0
        self.GLICKO_SIGMA = 0.06
        self.GLICKO_TAU = 0.5
        self.GLICKO_SCALE = 173.7178

        self.weight_class_stats = defaultdict(_defaultdict_list)

        self.fighter_opponents = defaultdict(set)
        self.fight_outcomes = {}

        self.kmeans = None
        self.cluster_scaler = StandardScaler()
        self.n_clusters = 8
        self.style_features = [
            "SLpM",
            "rd1_slpm_avg",
            "cardio_index",
            "TD",
            "rd1_td_rate",
            "Sub",
            "Finish",
        ]
        self.fighter_cluster = {}

        self.style_performance = defaultdict(_defaultdict_list)
        # Chronological finish rate for ordered style cluster matchups (updated in Tier 7)
        self.style_matchup_finish = defaultdict(_defaultdict_list)

    def elo_initial(self, fighter):
        if fighter not in self.elo_ratings:
            self.elo_ratings[fighter] = 1500.0
        return self.elo_ratings[fighter]

    def elo_expected(self, rA, rB):
        return 1.0 / (1.0 + 10 ** ((rB - rA) / 400.0))

    # K-factor scales how much a single fight moves a fighter's ELO.
    # Modifiers applied in order: experience bracket -> title bout -> finish method/round ->
    # win streak -> high-ELO opponent bonus -> upset bonus -> loss-streak penalty.
    def elo_k_factor(
        self,
        fighter_fights,
        is_title,
        method,
        result,
        finish_round=0,
        winner_streak=0,
        opponent_elo=1500,
        loser_streak=0,
    ):
        k = 32
        if fighter_fights < 5:
            k = 48
        elif fighter_fights > 20:
            k = 24
        if is_title:
            k *= 2
        if method in ("KO/TKO", "Submission") and result == "win":
            k *= 1.2

            if finish_round == 1:
                k *= 1.5
            elif finish_round == 2:
                k *= 1.3
            elif finish_round == 3:
                k *= 1.2

        # Split decisions are contested wins — reward and penalise less than dominant outcomes.
        # Unanimous decisions signal clear dominance — modest bonus for the winner.
        if method == "Split Decision":
            k *= 0.85
        elif method == "Unanimous Decision" and result == "win":
            k *= 1.05

        if winner_streak >= 3 and result == "win":
            k *= 1.1

        if opponent_elo > 1600 and result == "win":
            k *= 1.15

        if result == "win" and opponent_elo > 1600:
            upset_factor = min((opponent_elo - 1600) / 400.0, 0.5)
            k *= 1.0 + upset_factor

        if result == "loss" and loser_streak <= -3:
            k *= 1.25
        return k

    def elo_update(
        self,
        r_fighter,
        b_fighter,
        winner,
        is_title,
        method,
        r_fights,
        b_fights,
        finish_round=0,
        winner_streak=0,
        opponent_elo=1500,
        r_loss_streak=0,
        b_loss_streak=0,
    ):
        rA = self.elo_initial(r_fighter)
        rB = self.elo_initial(b_fighter)
        eA = self.elo_expected(rA, rB)
        eB = 1.0 - eA

        if winner == "Red":
            sA, sB = 1.0, 0.0
        elif winner == "Blue":
            sA, sB = 0.0, 1.0
        else:
            sA, sB = 0.5, 0.5

        kA = self.elo_k_factor(
            r_fights,
            is_title,
            method,
            "win" if winner == "Red" else "loss",
            finish_round=finish_round,
            winner_streak=winner_streak if winner == "Red" else 0,
            opponent_elo=opponent_elo if winner == "Red" else rA,
            loser_streak=r_loss_streak,
        )
        kB = self.elo_k_factor(
            b_fights,
            is_title,
            method,
            "win" if winner == "Blue" else "loss",
            finish_round=finish_round,
            winner_streak=winner_streak if winner == "Blue" else 0,
            opponent_elo=opponent_elo if winner == "Blue" else rB,
            loser_streak=b_loss_streak,
        )

        pre_rA = rA
        pre_rB = rB
        self.elo_ratings[r_fighter] = rA + kA * (sA - eA)
        self.elo_ratings[b_fighter] = rB + kB * (sB - eB)
        return pre_rA, pre_rB

    def elo_get(self, fighter):
        return self.elo_ratings.get(fighter, 1500.0)

    def glicko2_initial(self, fighter):
        if fighter not in self.glicko_ratings:
            self.glicko_ratings[fighter] = (
                self.GLICKO_MU,
                self.GLICKO_PHI,
                self.GLICKO_SIGMA,
            )
        return self.glicko_ratings[fighter]

    def _g(self, phi):
        return 1.0 / math.sqrt(1 + 3 * phi**2 / math.pi**2)

    def _E(self, mu, mu_j, phi_j):
        return 1.0 / (1.0 + math.exp(-self._g(phi_j) * (mu - mu_j)))

    def glicko2_update(self, fighter, opponents):
        mu, phi, sigma = self.glicko2_initial(fighter)
        mu_s = (mu - self.GLICKO_MU) / self.GLICKO_SCALE
        phi_s = phi / self.GLICKO_SCALE

        if not opponents:
            phi_star = math.sqrt(phi_s**2 + sigma**2)
            self.glicko_ratings[fighter] = (
                mu, phi_star * self.GLICKO_SCALE, sigma)
            return

        v_inv = 0.0
        delta_sum = 0.0
        for opp_r, opp_rd, score in opponents:
            mu_j = (opp_r - self.GLICKO_MU) / self.GLICKO_SCALE
            phi_j = opp_rd / self.GLICKO_SCALE
            g_j = self._g(phi_j)
            E_j = self._E(mu_s, mu_j, phi_j)
            v_inv += g_j**2 * E_j * (1 - E_j)
            delta_sum += g_j * (score - E_j)

        v = 1.0 / v_inv if v_inv > 0 else 1e6
        delta = v * delta_sum

        # Illinois algorithm (iterative bisection) to find the new volatility sigma
        # as described in Glickman's 2012 paper. Converges when |B - A| < 1e-6.
        a = math.log(sigma**2)
        tau = self.GLICKO_TAU

        def f(x):
            ex = math.exp(x)
            num = ex * (delta**2 - phi_s**2 - v - ex)
            den = 2 * (phi_s**2 + v + ex) ** 2
            return num / den - (x - a) / (tau**2)

        A = a
        if delta**2 > phi_s**2 + v:
            B = math.log(delta**2 - phi_s**2 - v)
        else:
            k = 1
            while f(a - k * tau) < 0:
                k += 1
            B = a - k * tau

        fA, fB = f(A), f(B)
        for _ in range(100):
            C = A + (A - B) * fA / (fB - fA)
            fC = f(C)
            if fC * fB < 0:
                A, fA = B, fB
            else:
                fA /= 2
            B, fB = C, fC
            if abs(B - A) < 1e-6:
                break
        new_sigma = math.exp(A / 2)

        phi_star = math.sqrt(phi_s**2 + new_sigma**2)
        new_phi_s = 1.0 / math.sqrt(1.0 / phi_star**2 + 1.0 / v)
        new_mu_s = mu_s + new_phi_s**2 * delta_sum

        self.glicko_ratings[fighter] = (
            new_mu_s * self.GLICKO_SCALE + self.GLICKO_MU,
            new_phi_s * self.GLICKO_SCALE,
            new_sigma,
        )

    def glicko2_get(self, fighter):
        r, rd, vol = self.glicko2_initial(fighter)
        return r, rd, vol

    def update_weight_class_stats(self, weight_class, year, stats_dict):
        for feat, val in stats_dict.items():
            if val is not None and not math.isnan(float(val)):
                self.weight_class_stats[(
                    weight_class, year)][feat].append(float(val))

    def get_z_score(self, weight_class, year, feat, value):
        key = (weight_class, year)
        if key not in self.weight_class_stats:
            return 0.0
        vals = self.weight_class_stats[key].get(feat, [])
        if len(vals) < 2:
            return 0.0
        mu = np.mean(vals)
        std = np.std(vals)
        if std < 1e-9:
            return 0.0
        return (value - mu) / std

    def update_common_opponents(self, r_fighter, b_fighter, winner):
        self.fighter_opponents[r_fighter].add(b_fighter)
        self.fighter_opponents[b_fighter].add(r_fighter)
        if winner == "Red":
            self.fight_outcomes[(r_fighter, b_fighter)] = 1
            self.fight_outcomes[(b_fighter, r_fighter)] = 0
        elif winner == "Blue":
            self.fight_outcomes[(r_fighter, b_fighter)] = 0
            self.fight_outcomes[(b_fighter, r_fighter)] = 1
        else:
            self.fight_outcomes[(r_fighter, b_fighter)] = 0.5
            self.fight_outcomes[(b_fighter, r_fighter)] = 0.5

    def get_common_opponent_features(self, r_fighter, b_fighter):
        common = self.fighter_opponents[r_fighter] & self.fighter_opponents[b_fighter]
        n_common = len(common)
        r_wins_common = 0
        b_wins_common = 0
        for opp in common:
            r_wins_common += self.fight_outcomes.get((r_fighter, opp), 0.5)
            b_wins_common += self.fight_outcomes.get((b_fighter, opp), 0.5)
        return {
            "n_common_opponents": n_common,
            "r_wins_vs_common": r_wins_common,
            "b_wins_vs_common": b_wins_common,
            "common_opp_edge": r_wins_common - b_wins_common,
        }

    def fit_clusters(self, fighter_stats):
        if len(fighter_stats) < self.n_clusters:
            return
        rows = []
        fighters = []
        for f, s in fighter_stats.items():
            row = [s.get(k, 0.0) for k in self.style_features]
            rows.append(row)
            fighters.append(f)
        X = np.array(rows, dtype=float)
        X = np.nan_to_num(X)
        X_scaled = self.cluster_scaler.fit_transform(X)
        self.kmeans = KMeans(
            n_clusters=self.n_clusters, random_state=RANDOM_SEED, n_init=10
        )
        labels = self.kmeans.fit_predict(X_scaled)
        for f, lbl in zip(fighters, labels):
            self.fighter_cluster[f] = int(lbl)

    def get_fighter_cluster(self, fighter):
        return self.fighter_cluster.get(fighter, -1)

    def predict_cluster(self, feature_values):
        if self.kmeans is None:
            return -1
        x = np.array(feature_values, dtype=float).reshape(1, -1)
        x = np.nan_to_num(x)
        x_scaled = self.cluster_scaler.transform(x)
        return int(self.kmeans.predict(x_scaled)[0])

    def update_style_performance(self, cluster, opponent_cluster, won):
        self.style_performance[cluster][opponent_cluster].append(
            1 if won else 0)

    def update_style_matchup_finish(self, cluster, opponent_cluster, was_finish):
        if cluster < 0 or opponent_cluster < 0:
            return
        self.style_matchup_finish[cluster][opponent_cluster].append(
            1.0 if was_finish else 0.0
        )

    def get_style_matchup_finish_rate(self, cluster, opponent_cluster):
        v = self.style_matchup_finish[cluster].get(opponent_cluster, [])
        return float(np.mean(v)) if v else 0.5

    def get_style_matchup_features(self, r_cluster, b_cluster):
        r_vs_b = self.style_performance[r_cluster].get(b_cluster, [])
        b_vs_r = self.style_performance[b_cluster].get(r_cluster, [])
        r_winrate = np.mean(r_vs_b) if r_vs_b else 0.5
        b_winrate = np.mean(b_vs_r) if b_vs_r else 0.5
        return {
            "r_style_win_vs_opp_cluster": r_winrate,
            "b_style_win_vs_opp_cluster": b_winrate,
            "style_matchup_edge": r_winrate - b_winrate,
            "r_cluster": r_cluster,
            "b_cluster": b_cluster,
        }


# TimeSeriesSplit that inserts a gap between train and test folds to prevent leakage
# caused by fights that share overlapping feature computation windows (e.g. rolling stats).
# purge_days is converted to an approximate fight-count gap using a ~1 fight/week assumption.
class PurgedTimeSeriesSplit:
    def __init__(self, n_splits=5, purge_days=30):
        self.n_splits = n_splits
        self.purge_days = purge_days

    def get_n_splits(self, X=None, y=None, groups=None):
        return self.n_splits

    def split(self, X, y=None, groups=None):
        n = len(X)
        if groups is not None:
            # Split on unique groups so augmented pairs (fight + its corner-swap)
            # always land in the same fold, preventing OOF leakage.
            groups = np.asarray(groups)
            unique_groups = np.unique(groups)
            n_g = len(unique_groups)
            fold_size = n_g // (self.n_splits + 1)
            for i in range(self.n_splits):
                g_tr_end = fold_size * (i + 1)
                g_te_start = g_tr_end + max(1, self.purge_days // 7)
                g_te_end = min(g_te_start + fold_size, n_g)
                if g_te_end <= g_te_start:
                    continue
                train_idx = np.where(
                    np.isin(groups, unique_groups[:g_tr_end]))[0]
                test_idx = np.where(
                    np.isin(groups, unique_groups[g_te_start:g_te_end])
                )[0]
                yield train_idx, test_idx
        else:
            fold_size = n // (self.n_splits + 1)
            for i in range(self.n_splits):
                train_end = fold_size * (i + 1)
                test_start = train_end + max(1, self.purge_days // 7)
                test_end = min(test_start + fold_size, n)
                if test_end <= test_start:
                    continue
                train_idx = np.arange(0, train_end)
                test_idx = np.arange(test_start, test_end)
                yield train_idx, test_idx


# Manual OOF stacking because sklearn's StackingClassifier produces NaN meta-features
# for some base models (RF, MLP, CatBoost), poisoning the LogisticRegression meta-learner.
# This class explicitly clones each estimator per fold, catches per-model per-fold errors,
# and falls back to uniform probabilities (1/n_classes) rather than propagating NaN.
# Base models are passed already-fitted; clones are used for OOF, originals for test inference.
class _ManualStackingEnsemble(BaseEstimator, ClassifierMixin):
    def __init__(
        self,
        estimators,
        meta_C=0.05,
        n_splits=3,
        random_state=42,
        passthrough_indices=None,
        verbose=True,
    ):

        self.estimators = estimators
        self.meta_C = meta_C
        self.n_splits = n_splits
        self.random_state = random_state
        self.passthrough_indices = passthrough_indices
        self.verbose = verbose
        self.classes_ = None
        self.final_estimator_ = None

    def fit(self, X, y, groups=None, sample_weight=None):
        from sklearn.base import clone as _clone

        y = np.asarray(y)
        self.classes_ = np.unique(y)
        nc = len(self.classes_)
        n = X.shape[0]
        nm = len(self.estimators)

        # Initialize OOF meta-features to uniform prior; unfilled folds (due to errors) stay at 1/k.
        meta_X_oof = np.full((n, nm * nc), 1.0 / nc)

        # Use PurgedTimeSeriesSplit for chronological OOF generation so the
        # meta-learner never sees predictions from models trained on future data.
        splitter = PurgedTimeSeriesSplit(n_splits=self.n_splits, purge_days=30)
        split_iter = splitter.split(X, groups=groups)
        for tr_idx, va_idx in split_iter:
            X_tr_f, X_va_f = X[tr_idx], X[va_idx]
            y_tr_f = y[tr_idx]
            sw_f = sample_weight[tr_idx] if sample_weight is not None else None
            for ei, (name, est) in enumerate(self.estimators):
                col = ei * nc
                try:
                    fold_est = _clone(est)

                    # Fix MLP temporal leakage: disable internal early-stopping
                    # validation split which randomly samples future fights.
                    if name in ("mlp", "mlp2") and hasattr(fold_est, "early_stopping"):
                        fold_est.set_params(early_stopping=False)

                    fit_kw = {}
                    X_fit, y_fit = X_tr_f, y_tr_f

                    # MLPs don't accept sample_weight — resample training
                    # data proportional to weights so recency is respected.
                    if name in ("mlp", "mlp2") and sw_f is not None:
                        _rng_sw = np.random.RandomState(self.random_state)
                        _probs = sw_f / sw_f.sum()
                        _resample_idx = _rng_sw.choice(
                            len(X_tr_f), size=len(X_tr_f), replace=True, p=_probs
                        )
                        X_fit = X_tr_f[_resample_idx]
                        y_fit = y_tr_f[_resample_idx]
                    elif sw_f is not None:
                        fit_kw["sample_weight"] = sw_f

                    # Pass eval_set to boosting models so OOF folds use early
                    # stopping, matching the behaviour of the final retrained models.
                    if name == "xgb":
                        fit_kw["eval_set"] = [(X_va_f, y[va_idx])]
                        fit_kw["verbose"] = False
                    elif name == "lgb":
                        fit_kw["eval_set"] = [(X_va_f, y[va_idx])]
                        fit_kw["callbacks"] = [
                            lgb.early_stopping(50, verbose=False),
                            lgb.log_evaluation(-1),
                        ]
                    elif name == "cat":
                        fit_kw["eval_set"] = (X_va_f, y[va_idx])

                    try:
                        fold_est.fit(X_fit, y_fit, **fit_kw)
                    except TypeError:
                        fold_est.fit(X_fit, y_fit)
                    proba = fold_est.predict_proba(X_va_f)
                    if proba.shape[1] != nc or np.any(np.isnan(proba)):
                        raise ValueError("bad proba shape or NaN")
                    meta_X_oof[va_idx, col: col + nc] = proba
                except Exception:
                    meta_X_oof[va_idx, col: col + nc] = 1.0 / nc

        # Append passthrough raw features to meta-feature matrix
        if self.passthrough_indices is not None and len(self.passthrough_indices) > 0:
            pt_cols = X[:, self.passthrough_indices]
            meta_X_oof = np.hstack([meta_X_oof, pt_cols])

        # Simplex-constrained log-loss meta-combiner (super learner formulation):
        # Weights >= 0, sum to 1; minimizes weighted log-loss on OOF positive-class
        # probabilities. Enforces additive combination — no sign flips possible.
        # Passthrough raw features are excluded; base model OOF probs already encode them.
        from scipy.optimize import minimize as _sp_minimize

        _pos_idx = np.array([_ei * nc + 1 for _ei in range(nm)])
        _P_oof = meta_X_oof[:, _pos_idx]  # (n, nm) positive-class OOF probs
        _y_f = y.astype(float)
        _sw = (
            (sample_weight / sample_weight.mean())
            if sample_weight is not None
            else np.ones(len(y))
        )

        def _simplex_logloss(_w):
            _p = np.clip(_P_oof @ _w, 1e-9, 1 - 1e-9)
            return -np.dot(_sw, _y_f * np.log(_p) + (1 - _y_f) * np.log(1 - _p)) / len(
                _y_f
            )

        _w0 = np.ones(nm) / nm
        _bounds = [(0.0, 1.0)] * nm
        _cons = {"type": "eq", "fun": lambda _w: _w.sum() - 1.0}
        _res = _sp_minimize(
            _simplex_logloss,
            _w0,
            method="SLSQP",
            bounds=_bounds,
            constraints=_cons,
            options={"ftol": 1e-12, "maxiter": 2000},
        )
        self._simplex_weights = _res.x
        self._nm_simplex = nm
        self._nc_simplex = nc
        self._oof_pos_proba = np.clip(_P_oof @ _res.x, 1e-9, 1 - 1e-9)
        self.final_estimator_ = None
        if self.verbose:
            print(
                f"    Simplex meta-learner weights (OOF log-loss={_res.fun:.4f}):")
            for (_sname, _), _sw_w in zip(self.estimators, self._simplex_weights):
                print(f"      {_sname:<6s}  {_sw_w:.4f}")
        return self

    def _meta_features(self, X):
        nc = len(self.classes_)
        nm = len(self.estimators)
        meta_X = np.full((X.shape[0], nm * nc), 1.0 / nc)
        for ei, (name, est) in enumerate(self.estimators):
            col = ei * nc
            try:
                proba = est.predict_proba(X)
                if proba.shape[1] != nc or np.any(np.isnan(proba)):
                    raise ValueError("bad proba")
                meta_X[:, col: col + nc] = proba
            except Exception:
                meta_X[:, col: col + nc] = 1.0 / nc
        if self.passthrough_indices is not None and len(self.passthrough_indices) > 0:
            pt_cols = X[:, self.passthrough_indices]
            meta_X = np.hstack([meta_X, pt_cols])
        return meta_X

    def predict(self, X):
        return self.classes_[(self.predict_proba(X)[:, 1] >= 0.5).astype(int)]

    def predict_proba(self, X):
        _meta_X = self._meta_features(X)
        _pos_idx = np.array(
            [_ei * self._nc_simplex + 1 for _ei in range(self._nm_simplex)]
        )
        _P_pos = np.clip(_meta_X[:, _pos_idx] @
                         self._simplex_weights, 1e-9, 1 - 1e-9)
        return np.column_stack([1.0 - _P_pos, _P_pos])


# Central model class. Owns the full ML pipeline from raw CSV to predictions.
# Key attributes after training:
#   self.stacking_clf   -> trained _ManualStackingEnsemble
#   self.feature_cols   -> ordered list of feature column names the model expects
#   self.scaler         -> StandardScaler fitted on training data
#   self.feature_engineer -> FeatureEngineer holding all fighter rating state
#   self._wc_finish_chrono -> per-(WC,gender,rounds) cumulative KO/sub counts by fight index (no snapshot leakage)
class UFCPredictor:
    def __init__(self, data_path=None, status_callback=None):
        self.data_path = data_path or DEFAULT_DATA_PATH
        self.status_callback = status_callback or (lambda msg: None)
        self.df = None
        self.feature_engineer = FeatureEngineer()
        self.feature_cols = []
        self.scaler = StandardScaler()
        self.label_encoder = LabelEncoder()
        self.models = {}
        self.stacking_clf = None

        self.is_trained = False
        self.gpu_info = {}

        self.svd_striking = TruncatedSVD(
            n_components=5, random_state=RANDOM_SEED)
        self.svd_grappling = TruncatedSVD(
            n_components=5, random_state=RANDOM_SEED)
        self.svd_physical = TruncatedSVD(
            n_components=3, random_state=RANDOM_SEED)
        self.svd_form = TruncatedSVD(n_components=5, random_state=RANDOM_SEED)
        self.svd_fitted = False

        self.svd_striking_cols = []
        self.svd_grappling_cols = []
        self.svd_physical_cols = []
        self.svd_form_cols = []
        self._selected_d_indices = []

        self.weight_classes = []
        self.all_fighters = []
        self.predictions = []
        self._wc_finish_chrono = None
        self._wc_chrono_total_fights = 0

    def _finalize_wc_finish_chrono_timelines(self):
        """Build searchable cumulative KO/sub counts per WC bucket ordered by fight index."""
        raw = getattr(self, "_wc_finish_raw_events", None) or {}
        ch = {}
        for wck, evs in raw.items():
            if not evs:
                continue
            evs_s = sorted(evs, key=lambda t: t[0])
            seqs, cko_l, csu_l = [], [], []
            ck = cs = 0
            for s, dk, ds in evs_s:
                ck += int(dk)
                cs += int(ds)
                seqs.append(int(s))
                cko_l.append(ck)
                csu_l.append(cs)
            ch[wck] = {
                "seq": np.asarray(seqs, dtype=np.int64),
                "cum_ko": np.asarray(cko_l, dtype=np.int64),
                "cum_sub": np.asarray(csu_l, dtype=np.int64),
            }
        self._wc_finish_chrono = ch

    def _wc_route_prior_before_chrono_seq(
        self, weight_class, gender, total_rounds, chrono_seq
    ):
        """Sub-share minus KO-share among finishes in bucket, using only fights with index < chrono_seq."""
        try:
            tr = int(total_rounds) if total_rounds is not None else 3
        except (TypeError, ValueError):
            tr = 3
        if tr <= 0:
            tr = 3
        key = (
            str(weight_class or "").strip(),
            str(gender or "").lower().strip(),
            tr,
        )
        blk = (getattr(self, "_wc_finish_chrono", None) or {}).get(key)
        if blk is None or len(blk["seq"]) == 0:
            return 0.0
        seq_arr = blk["seq"]
        try:
            q = int(chrono_seq)
        except (TypeError, ValueError):
            q = 2**62
        pos = int(np.searchsorted(seq_arr, q, side="left")) - 1
        if pos < 0:
            return 0.0
        ko = int(blk["cum_ko"][pos])
        su = int(blk["cum_sub"][pos])
        fin = ko + su
        if fin <= 0:
            return 0.0
        return (su / fin) - (ko / fin)

    def _apply_wc_finish_route_prior_chrono_safe(self, df):
        """Fill wc_finish_route_prior using only fights before this row's chrono index (no leakage)."""
        if not getattr(self, "_wc_finish_chrono", None):
            return
        if "weight_class" not in df.columns:
            return
        n_tot = int(getattr(self, "_wc_chrono_total_fights", 0))

        def _seq_for_row(row):
            v = row.get("_wc_chrono_seq", None)
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return n_tot
            try:
                return int(v)
            except (TypeError, ValueError):
                return n_tot

        def _one(row):
            return self._wc_route_prior_before_chrono_seq(
                row.get("weight_class", ""),
                row.get("gender", ""),
                row.get("total_rounds", 3),
                _seq_for_row(row),
            )

        if "wc_finish_route_prior" not in df.columns:
            df["wc_finish_route_prior"] = [_one(r) for _, r in df.iterrows()]
            return
        for idx, row in df.iterrows():
            v = df.at[idx, "wc_finish_route_prior"]
            if pd.isna(v):
                df.at[idx, "wc_finish_route_prior"] = _one(row)

    def _log(self, msg):
        self.status_callback(msg)
        print_step(msg)

    # ── SVD helpers ─────────────────────────────────────────────────
    _SVD_GROUPS = ("striking", "grappling", "physical", "form")

    def _expected_svd_output_cols(self):
        """Return {prefix: [col_names]} for every SVD group that has been fitted."""
        out = {}
        for prefix in self._SVD_GROUPS:
            src_cols = getattr(self, f"svd_{prefix}_cols", [])
            svd_obj = getattr(self, f"svd_{prefix}", None)
            if not src_cols or svd_obj is None:
                continue
            if not hasattr(svd_obj, "components_"):
                continue
            n = svd_obj.components_.shape[0]
            out[prefix] = [f"{prefix}_svd_{i}" for i in range(n)]
        return out

    def _materialize_svd_columns(self, df):
        """Transform *df* with the fitted SVD objects, adding output columns.

        Safe to call on any DataFrame that contains the source columns — it
        overwrites existing SVD output columns or creates them from scratch.
        """
        for prefix in self._SVD_GROUPS:
            src_cols = getattr(self, f"svd_{prefix}_cols", [])
            svd_obj = getattr(self, f"svd_{prefix}", None)
            if not src_cols or svd_obj is None or not hasattr(svd_obj, "components_"):
                continue
            try:
                X = np.column_stack(
                    [
                        df[c].fillna(
                            0).values if c in df.columns else np.zeros(len(df))
                        for c in src_cols
                    ]
                )
                X_svd = svd_obj.transform(X)
                for i in range(X_svd.shape[1]):
                    df[f"{prefix}_svd_{i}"] = X_svd[:, i]
            except Exception as e:
                raise RuntimeError(
                    f"SVD materialization failed for '{prefix}': {e}"
                ) from e

    def _validate_svd_state(self, label=""):
        """Assert every expected SVD output column exists in self.df.

        Call after fresh feature building *and* after cache restore.
        Raises RuntimeError with details if columns are missing.
        """
        if not self.svd_fitted:
            return
        expected = self._expected_svd_output_cols()
        missing_all = {}
        for prefix, cols in expected.items():
            missing = [c for c in cols if c not in self.df.columns]
            if missing:
                missing_all[prefix] = missing
        if missing_all:
            # Attempt re-materialization before giving up
            print_step(
                f"SVD columns missing after {label or 'load'}, re-materializing..."
            )
            self._materialize_svd_columns(self.df)
            # Re-check
            still_missing = {}
            for prefix, cols in expected.items():
                m = [c for c in cols if c not in self.df.columns]
                if m:
                    still_missing[prefix] = m
            if still_missing:
                raise RuntimeError(
                    f"SVD columns missing from self.df after re-materialization: {still_missing}"
                )

    def run_with_cache(self):
        """Run the full pipeline with per-stage caching.

        Each stage checks for a valid cache before running. If the cache is
        valid (data unchanged + all upstream stage versions match), the stage
        is skipped and its outputs are restored from disk. Otherwise the stage
        runs and its outputs are cached for next time.

        Bump the version string in STAGE_VERSIONS when you change code that
        affects a stage's output — that invalidates that stage and all
        downstream stages automatically.
        """
        data_hash = _hash_file(self.data_path)
        print_section("PIPELINE CACHE CHECK")
        print_metric("Data hash:", data_hash[:16] + "...")
        print_metric(
            "Stage versions:",
            ", ".join(f"{s}=v{STAGE_VERSIONS[s]}" for s in STAGE_ORDER),
        )

        # ── Stage 1: data_features ──────────────────────────────────────
        cached = _load_stage_cache("data_features", data_hash)
        if cached is not None:
            self.df = cached["df"]
            self.feature_engineer = cached["feature_engineer"]
            self.weight_classes = cached["weight_classes"]
            self.all_fighters = cached["all_fighters"]
            self.feature_cols = cached.get("feature_cols", [])
            self.svd_striking = cached.get("svd_striking", self.svd_striking)
            self.svd_grappling = cached.get(
                "svd_grappling", self.svd_grappling)
            self.svd_physical = cached.get("svd_physical", self.svd_physical)
            self.svd_form = cached.get("svd_form", self.svd_form)
            self.svd_fitted = cached.get("svd_fitted", False)
            self.svd_striking_cols = cached.get("svd_striking_cols", [])
            self.svd_grappling_cols = cached.get("svd_grappling_cols", [])
            self.svd_physical_cols = cached.get("svd_physical_cols", [])
            self.svd_form_cols = cached.get("svd_form_cols", [])
            self._wc_finish_chrono = cached.get("wc_finish_chrono")
            self._wc_chrono_total_fights = int(
                cached.get("wc_chrono_total_fights", 0) or 0
            )
        else:
            self.load_data()
            self.fix_data_leakage()
            self.build_all_features()
            _save_stage_cache(
                "data_features",
                data_hash,
                {
                    "df": self.df,
                    "feature_engineer": self.feature_engineer,
                    "weight_classes": self.weight_classes,
                    "all_fighters": self.all_fighters,
                    "feature_cols": self.feature_cols,
                    "svd_striking": self.svd_striking,
                    "svd_grappling": self.svd_grappling,
                    "svd_physical": self.svd_physical,
                    "svd_form": self.svd_form,
                    "svd_fitted": self.svd_fitted,
                    "svd_striking_cols": self.svd_striking_cols,
                    "svd_grappling_cols": self.svd_grappling_cols,
                    "svd_physical_cols": self.svd_physical_cols,
                    "svd_form_cols": self.svd_form_cols,
                    "wc_finish_chrono": self._wc_finish_chrono,
                    "wc_chrono_total_fights": self._wc_chrono_total_fights,
                },
            )

        self._validate_svd_state(label="data_features stage")

        # ── Stages 2-5 are inside train() ───────────────────────────────
        self.train(data_hash=data_hash)

    def load_data(self):
        print_section("LOADING DATA")
        self._log(f"Reading: {self.data_path}")
        self.df = pd.read_csv(self.data_path, low_memory=False)
        self._log(
            f"Loaded {len(self.df):,} rows x {len(self.df.columns)} columns")

        self.df["event_date"] = pd.to_datetime(
            self.df["event_date"], format="%m/%d/%Y", errors="coerce"
        )
        self.df = self.df.sort_values("event_date").reset_index(drop=True)

        # Recompute current_age from DOB so it reflects fight-day age (not scrape-day).
        # For rows without an event_date (upcoming fights) fall back to today.
        from datetime import datetime as _dt

        def _calc_age(born, ref):
            if pd.isna(born) or pd.isna(ref):
                return None
            try:
                return (
                    ref.year
                    - born.year
                    - ((ref.month, ref.day) < (born.month, born.day))
                )
            except Exception:
                return None

        _today = _dt.now().date()
        for corner in ("r", "b"):
            dob_col = f"{corner}_date_of_birth"
            if dob_col in self.df.columns:
                dob = pd.to_datetime(self.df[dob_col], errors="coerce")
                evt = self.df["event_date"]
                self.df[f"{corner}_current_age"] = pd.to_numeric(
                    pd.Series(
                        [
                            _calc_age(d, e.date() if pd.notna(e) else _today)
                            for d, e in zip(dob, evt)
                        ],
                        index=self.df.index,
                    ),
                    errors="coerce",
                )
        if "r_current_age" in self.df.columns and "b_current_age" in self.df.columns:
            self.df["current_age_diff"] = (
                pd.to_numeric(self.df["r_current_age"], errors="coerce")
                - pd.to_numeric(self.df["b_current_age"], errors="coerce")
            ).round(2)

        # Prefix women's weight classes so they're distinct from men's divisions.
        # The source data only labels "Women's Strawweight"; the rest are bare.
        _women_mask = self.df["gender"].fillna(
            "").str.lower().str.strip() == "women"
        _wc = self.df["weight_class"].fillna("")
        _needs_prefix = _women_mask & ~_wc.str.startswith("Women")
        self.df.loc[_needs_prefix, "weight_class"] = (
            "Women's " + self.df.loc[_needs_prefix, "weight_class"]
        )

        self.df["winner"] = self.df["winner"].str.strip()
        self.df["method"] = self.df["method"].str.strip()

        method_map = {
            "Decision - Unanimous": "Unanimous Decision",
            "Decision - Split": "Split Decision",
            "Decision - Majority": "Majority Decision",
            "TKO - Doctor's Stoppage": "KO/TKO",
            "Could Not Continue": None,
            "Overturned": None,
            "DQ": None,
            "No Contest": None,
        }
        self.df["method"] = self.df["method"].apply(
            lambda m: (
                method_map.get(str(m).strip(), str(
                    m).strip()) if pd.notna(m) else m
            )
        )
        self.df = self.df[
            self.df["method"].notna()
            & self.df["method"].isin(
                [
                    "Unanimous Decision",
                    "Majority Decision",
                    "Split Decision",
                    "KO/TKO",
                    "Submission",
                    "Draw",
                ]
            )
        ]

        self.df = self.df[self.df["winner"].isin(["Red", "Blue", "Draw"])].reset_index(
            drop=True
        )

        self.weight_classes = sorted(
            self.df["weight_class"].dropna().unique().tolist())
        self.all_fighters = sorted(
            set(self.df["r_fighter"].dropna().tolist())
            | set(self.df["b_fighter"].dropna().tolist())
        )
        print_metric("Total fights:", len(self.df))
        print_metric("Unique fighters:", len(self.all_fighters))
        print_metric("Weight classes:", len(self.weight_classes))
        print()

    # The raw CSV often contains cumulative career stats recorded at fight time, which
    # technically includes the current fight's outcome — a data-leakage source.
    # This method replays all fights in chronological order and rebuilds every stat
    # using only information available *before* each fight. It also computes:
    #   - ELO and Glicko-2 ratings snapshotted pre-fight
    #   - Rolling windows (last 3/5/10 fights)
    #   - Days since last fight, opponent ELO history, trajectory slopes
    # The result is written back into self.df as `r_pre_*` / `b_pre_*` columns.
    def fix_data_leakage(self):
        print_section("FIXING DATA LEAKAGE")
        self._log("Recalculating fighter stats chronologically...")
        self._wc_finish_counts = defaultdict(
            lambda: {"ko": 0, "sub": 0, "dec": 0})
        self._wc_finish_raw_events = defaultdict(list)

        stats_to_track = {
            "wins": 0,
            "losses": 0,
            "draws": 0,
            "ko_wins": 0,
            "sub_wins": 0,
            "dec_wins": 0,
            "ko_losses": 0,
            "sub_losses": 0,
            "dec_losses": 0,
            "unanimous_wins": 0,
            "majority_wins": 0,
            "split_wins": 0,
            "split_losses": 0,
            "total_fights": 0,
            "title_fights": 0,
            "title_wins": 0,
            "recent_wins": 0,
            "recent_losses": 0,
            "win_streak": 0,
            "loss_streak": 0,
            "finish_rate": 0.0,
            "recent_finish_rate": 0.0,
            "total_fight_time": 0.0,
            "avg_fight_time": 0.0,
            "total_sig_str": 0,
            "total_sig_str_att": 0,
            "total_str": 0,
            "total_str_att": 0,
            "total_td": 0,
            "total_td_att": 0,
            "total_sub_att": 0,
            "total_kd": 0,
            "total_head": 0,
            "total_body": 0,
            "total_leg": 0,
            "total_distance": 0,
            "total_clinch": 0,
            "total_ground": 0,
            "total_ctrl_sec": 0,
            "total_rev": 0,
            "rolling3_wins": 0,
            "rolling3_sig_str": 0.0,
            "rolling3_td": 0.0,
            "rolling3_kd": 0.0,
            "rolling3_sub_att": 0.0,
            "rolling3_ctrl": 0.0,
            "rolling5_wins": 0,
            "rolling5_sig_str": 0.0,
            "rolling5_td": 0.0,
            "rolling5_kd": 0.0,
            "pro_SLpM": 0.0,
            "pro_SApM": 0.0,
            "pro_sig_str_acc": 0.0,
            "pro_str_def": 0.0,
            "pro_td_avg": 0.0,
            "pro_td_acc": 0.0,
            "pro_td_def": 0.0,
            "pro_sub_avg": 0.0,
            "avg_opp_wins": 0.0,
            "avg_opp_losses": 0.0,
            "last_result": None,
            "_history": [],
            "kd_absorbed": 0,
            "fight_dates": [],
            "opp_elo_history": [],
            "distance_strikes": 0,
            "clinch_strikes": 0,
            "ground_strikes": 0,
            "head_strikes": 0,
            "body_strikes": 0,
            "leg_strikes": 0,
            "total_positional_strikes": 0,
            "str_def_attempts": 0,
            "str_def_success": 0,
            "td_def_attempts": 0,
            "td_def_success": 0,
            "early_finishes": 0,
            "late_finishes": 0,
            "first_round_kos": 0,
            "total_rounds_fought": 0,
            "five_round_fights": 0,
            "championship_rounds": 0,
            "rolling10_history": [],
            "career_elo_peak": 1500.0,
            "fights_since_peak": 0,
            "weight_class_history": [],
            "weight_class_tenure": 0,
            "last_fight_method": None,
            "last_fight_finish_round": 0,
            "last_fight_was_finish": False,
            "last_fight_was_win": False,
            "vs_elite_wins": 0,
            "vs_elite_fights": 0,
            "vs_striker_wins": 0,
            "vs_striker_fights": 0,
            "vs_grappler_wins": 0,
            "vs_grappler_fights": 0,
            # Tier 43: exact fight-history numerators (updated end of each fight row)
            "early_finish_losses": 0,
            "late_finish_losses": 0,
            "sub_wins_with_td": 0,
            "ko_wins_with_kd": 0,
            "finish_losses_after_lost_r1": 0,
            "fights_lost_r1": 0,
            "dec_wins_after_won_r1": 0,
            "fights_won_r1": 0,
            "sub_wins_after_lost_r1": 0,
            "ko_wins_after_lost_r1": 0,
            "finish_wins_phys_dom": 0,
            "t43_dom_wins": 0,
            "t43_finish_dom_wins": 0,
        }

        fighter_stats = defaultdict(lambda: copy.deepcopy(stats_to_track))
        fighter_fights_count = defaultdict(int)

        leakage_cols_r = [c for c in self.df.columns if c.startswith("r_")]
        leakage_cols_b = [c for c in self.df.columns if c.startswith("b_")]

        new_cols = [
            "r_pre_wins",
            "r_pre_losses",
            "r_pre_draws",
            "b_pre_wins",
            "b_pre_losses",
            "b_pre_draws",
            "r_pre_ko_wins",
            "r_pre_sub_wins",
            "r_pre_dec_wins",
            "b_pre_ko_wins",
            "b_pre_sub_wins",
            "b_pre_dec_wins",
            "r_pre_total_fights",
            "b_pre_total_fights",
            "r_pre_finish_rate",
            "b_pre_finish_rate",
            "r_pre_win_streak",
            "b_pre_win_streak",
            "r_pre_loss_streak",
            "b_pre_loss_streak",
            "r_pre_title_fights",
            "b_pre_title_fights",
            "r_pre_title_wins",
            "b_pre_title_wins",
            "r_pre_avg_fight_time",
            "b_pre_avg_fight_time",
            "r_pre_sig_str_acc",
            "b_pre_sig_str_acc",
            "r_pre_td_acc",
            "b_pre_td_acc",
            "r_pre_sub_att_rate",
            "b_pre_sub_att_rate",
            "r_pre_kd_rate",
            "b_pre_kd_rate",
            "r_pre_ctrl_avg",
            "b_pre_ctrl_avg",
            "r_pre_SLpM",
            "b_pre_SLpM",
            "r_pre_SApM",
            "b_pre_SApM",
            "r_pre_td_avg",
            "b_pre_td_avg",
            "r_rolling3_wins",
            "b_rolling3_wins",
            "r_rolling3_sig_str",
            "b_rolling3_sig_str",
            "r_rolling3_td",
            "b_rolling3_td",
            "r_rolling3_kd",
            "b_rolling3_kd",
            "r_rolling3_sub_att",
            "b_rolling3_sub_att",
            "r_rolling5_wins",
            "b_rolling5_wins",
            "r_rolling5_sig_str",
            "b_rolling5_sig_str",
            "r_rolling5_td",
            "b_rolling5_td",
            "r_rolling5_kd",
            "b_rolling5_kd",
            "r_glicko_pre_r",
            "r_glicko_pre_rd",
            "r_glicko_pre_vol",
            "b_glicko_pre_r",
            "b_glicko_pre_rd",
            "b_glicko_pre_vol",
            "r_days_since_last",
            "b_days_since_last",
            "r_pre_kd_absorbed",
            "b_pre_kd_absorbed",
            "r_avg_opp_elo_L5",
            "b_avg_opp_elo_L5",
            "r_trajectory_3",
            "b_trajectory_3",
            "r_pre_distance_pct",
            "b_pre_distance_pct",
            "r_pre_clinch_pct",
            "b_pre_clinch_pct",
            "r_pre_ground_pct",
            "b_pre_ground_pct",
            "r_pre_head_pct",
            "b_pre_head_pct",
            "r_pre_body_pct",
            "b_pre_body_pct",
            "r_pre_leg_pct",
            "b_pre_leg_pct",
            "r_pre_str_def",
            "b_pre_str_def",
            "r_pre_td_def",
            "b_pre_td_def",
            "r_pre_early_finish_rate",
            "b_pre_early_finish_rate",
            "r_pre_late_finish_rate",
            "b_pre_late_finish_rate",
            "r_pre_first_round_ko_rate",
            "b_pre_first_round_ko_rate",
            "r_pre_total_rounds_fought",
            "b_pre_total_rounds_fought",
            "r_pre_five_round_fights",
            "b_pre_five_round_fights",
            "r_rolling10_wins",
            "b_rolling10_wins",
            "r_rolling10_sig_str",
            "b_rolling10_sig_str",
            "r_rolling10_td",
            "b_rolling10_td",
            "r_rolling10_kd",
            "b_rolling10_kd",
            "r_rolling10_finishes",
            "b_rolling10_finishes",
            "r_career_elo_peak",
            "b_career_elo_peak",
            "r_fights_since_peak",
            "b_fights_since_peak",
            "r_vs_elite_win_rate",
            "b_vs_elite_win_rate",
            "r_vs_striker_win_rate",
            "b_vs_striker_win_rate",
            "r_vs_grappler_win_rate",
            "b_vs_grappler_win_rate",
            "r_last_fight_was_win",
            "b_last_fight_was_win",
            "r_last_fight_was_finish",
            "b_last_fight_was_finish",
            "r_pre_finish_rate_l5",
            "b_pre_finish_rate_l5",
            "r_pre_finish_rate_l10",
            "b_pre_finish_rate_l10",
            "r_pre_slpm_cv",
            "b_pre_slpm_cv",
            "r_pre_mileage_adj_age",
            "b_pre_mileage_adj_age",
            "r_pre_rolling3_wins",
            "b_pre_rolling3_wins",
            "r_pre_rolling5_slpm",
            "b_pre_rolling5_slpm",
            "r_pre_slpm_std_l10",
            "b_pre_slpm_std_l10",
            "r_pre_damage_ratio_std_l10",
            "b_pre_damage_ratio_std_l10",
            "r_pre_tactical_evolution",
            "b_pre_tactical_evolution",
            # Decision-type win/loss counts
            "r_pre_unanimous_wins",
            "b_pre_unanimous_wins",
            "r_pre_majority_wins",
            "b_pre_majority_wins",
            "r_pre_split_wins",
            "b_pre_split_wins",
            "r_pre_split_losses",
            "b_pre_split_losses",
            # Tier 43: method-path numerators → pre-fight rates (from fix_data_leakage counters)
            "r_pre_early_finish_loss_rate",
            "b_pre_early_finish_loss_rate",
            "r_pre_late_finish_loss_rate",
            "b_pre_late_finish_loss_rate",
            "r_pre_sub_win_after_td_rate",
            "b_pre_sub_win_after_td_rate",
            "r_pre_ko_win_after_kd_rate",
            "b_pre_ko_win_after_kd_rate",
            "r_pre_finish_loss_lost_r1_rate",
            "b_pre_finish_loss_lost_r1_rate",
            "r_pre_sub_win_lost_r1_rate",
            "b_pre_sub_win_lost_r1_rate",
            "r_pre_ko_win_lost_r1_rate",
            "b_pre_ko_win_lost_r1_rate",
            "r_pre_dec_win_won_r1_rate",
            "b_pre_dec_win_won_r1_rate",
            "r_pre_finish_win_phys_dom_rate",
            "b_pre_finish_win_phys_dom_rate",
            "r_pre_dom_finish_conv_exact",
            "b_pre_dom_finish_conv_exact",
            "r_pre_ex_sub_conv",
            "b_pre_ex_sub_conv",
            "r_pre_ex_ctrl_sub_conv",
            "b_pre_ex_ctrl_sub_conv",
            "r_pre_ex_ko_tko_conv",
            "b_pre_ex_ko_tko_conv",
            "r_pre_da_damage_margin",
            "b_pre_da_damage_margin",
            "r_pre_da_head_sig_abs_share",
            "b_pre_da_head_sig_abs_share",
            "r_pre_qa_finish_win",
            "b_pre_qa_finish_win",
            "r_pre_ra_finish_win",
            "b_pre_ra_finish_win",
            "wc_finish_route_prior",
            "_wc_chrono_seq",
            "cluster_pair_finish_rate",
            # Per-round derived stats (rolling L5 averages)
            "r_pre_rd1_slpm_avg",
            "b_pre_rd1_slpm_avg",
            "r_pre_rd1_kd_rate",
            "b_pre_rd1_kd_rate",
            "r_pre_rd1_td_rate",
            "b_pre_rd1_td_rate",
            "r_pre_rd1_head_pct",
            "b_pre_rd1_head_pct",
            "r_pre_cardio_index",
            "b_pre_cardio_index",
            "r_pre_late_td_rate",
            "b_pre_late_td_rate",
            # Output slope (pace change round-over-round)
            "r_pre_output_slope",
            "b_pre_output_slope",
            # Round-1 accuracy
            "r_pre_r1_acc",
            "b_pre_r1_acc",
            # Late-round accuracy (R3+)
            "r_pre_late_acc",
            "b_pre_late_acc",
            # Accuracy fade (R1 acc minus late acc)
            "r_pre_acc_fade",
            "b_pre_acc_fade",
            # Championship round (R4/R5) striking pace
            "r_pre_r45_slpm",
            "b_pre_r45_slpm",
            # Championship round pace relative to R1
            "r_pre_r45_vs_r1_ratio",
            "b_pre_r45_vs_r1_ratio",
            # Championship round TD rate
            "r_pre_r45_td_rate",
            "b_pre_r45_td_rate",
            # Body strike escalation (body hits rise in later rounds)
            "r_pre_body_escalation",
            "b_pre_body_escalation",
            # TD attempt slope (more TDs in later rounds = wrestler stamina)
            "r_pre_td_slope",
            "b_pre_td_slope",
            # Late-round KD rate
            "r_pre_late_kd_rate",
            "b_pre_late_kd_rate",
            # Weighted average round in which KDs land
            "r_pre_kd_round_avg",
            "b_pre_kd_round_avg",
            # Absorption slope (takes more strikes as fight progresses)
            "r_pre_absorption_slope",
            "b_pre_absorption_slope",
            # Late-round strikes absorbed per min
            "r_pre_late_absorbed",
            "b_pre_late_absorbed",
            # Ground-strike escalation (ground output rises later)
            "r_pre_ground_escalation",
            "b_pre_ground_escalation",
            # Late-round submission attempt rate
            "r_pre_late_sub_rate",
            "b_pre_late_sub_rate",
            # Control-time slope
            "r_pre_ctrl_slope",
            "b_pre_ctrl_slope",
            # ── Leg kick features ───────────────────────────────────────────────────
            "r_pre_r1_leg_rate",
            "b_pre_r1_leg_rate",  # R1 leg kicks per min
            "r_pre_leg_escalation",
            "b_pre_leg_escalation",  # R3 leg − R1 leg (count)
            "r_pre_leg_vs_head",
            "b_pre_leg_vs_head",  # R1 leg / max(R1 head, 1)
            "r_pre_leg_pct",
            "b_pre_leg_pct",  # leg / sig_str in R1
            # ── Clinch features ─────────────────────────────────────────────────────
            "r_pre_r1_clinch_rate",
            "b_pre_r1_clinch_rate",  # R1 clinch strikes per min
            "r_pre_clinch_escalation",
            "b_pre_clinch_escalation",  # R3 clinch − R1 clinch
            "r_pre_clinch_pct_r1",
            "b_pre_clinch_pct_r1",  # R1 clinch / R1 sig_str
            # ── Round 2 adjustment signals ──────────────────────────────────────────
            "r_pre_r1_to_r2_output",
            "b_pre_r1_to_r2_output",  # R2 sig_str − R1 sig_str
            "r_pre_r1_to_r2_td",
            "b_pre_r1_to_r2_td",  # R2 td − R1 td
            "r_pre_r2_body_rate",
            "b_pre_r2_body_rate",  # R2 body strikes avg
            # ── Volume vs power ratio ───────────────────────────────────────────────
            "r_pre_volume_ratio",
            "b_pre_volume_ratio",  # R1 total_str / sig_str
            "r_pre_vol_ratio_evo",
            "b_pre_vol_ratio_evo",  # R3 ratio − R1 ratio
            # ── Zone shift (distance → clinch/ground progression) ───────────────────
            "r_pre_zone_shift",
            "b_pre_zone_shift",  # R3 dist_pct − R1 dist_pct
            # ── TD accuracy evolution ───────────────────────────────────────────────
            "r_pre_r1_td_acc",
            "b_pre_r1_td_acc",  # R1 td / td_att
            "r_pre_late_td_acc",
            "b_pre_late_td_acc",  # R3 td / td_att
            "r_pre_td_acc_evo",
            "b_pre_td_acc_evo",  # late − r1 td acc
            # ── Championship round extras (R4/R5) ───────────────────────────────────
            "r_pre_r45_kd_rate",
            "b_pre_r45_kd_rate",  # avg KDs in R4+R5
            "r_pre_r45_body_rate",
            "b_pre_r45_body_rate",  # avg body in R4+R5
            "r_pre_r45_clinch_rate",
            "b_pre_r45_clinch_rate",  # avg clinch in R4+R5
            "r_pre_r45_ctrl",
            "b_pre_r45_ctrl",  # avg ctrl in R4+R5
            # ── Reversal rate ───────────────────────────────────────────────────────
            "r_pre_reversal_rate",
            "b_pre_reversal_rate",
            # ── Round momentum profile ──────────────────────────────────────────────
            "r_pre_r1_win_rate",
            "b_pre_r1_win_rate",  # % fights won R1 by output
            "r_pre_late_win_rate",
            "b_pre_late_win_rate",  # % fights won R3 by output
            # ── Per-round damage margins ─────────────────────────────────────────────
            "r_pre_r1_damage_margin",
            "b_pre_r1_damage_margin",  # R1 output − absorbed
            "r_pre_r2_damage_margin",
            "b_pre_r2_damage_margin",  # R2 output − absorbed
            "r_pre_r3_damage_margin",
            "b_pre_r3_damage_margin",  # R3 output − absorbed
            "r_pre_damage_margin_change",
            "b_pre_damage_margin_change",  # R3 margin − R1 margin
            # ── Head accuracy evolution ──────────────────────────────────────────────
            "r_pre_r1_head_acc",
            "b_pre_r1_head_acc",  # R1 head / head_att
            "r_pre_r3_head_acc",
            "b_pre_r3_head_acc",  # R3 head / head_att
            "r_pre_head_acc_fade",
            "b_pre_head_acc_fade",  # R1 − R3 head accuracy
            "r_pre_head_to_leg_late",
            "b_pre_head_to_leg_late",  # R3 leg_pct − R1 leg_pct
            # ── Grappling chain composites ───────────────────────────────────────────
            "r_pre_grapple_chain",
            "b_pre_grapple_chain",  # TD→ctrl→sub chain score
            "r_pre_gnp_score",
            "b_pre_gnp_score",  # ground-and-pound score
            "r_pre_sub_efficiency",
            "b_pre_sub_efficiency",  # sub_att per late TD
            # ── Style entropy ────────────────────────────────────────────────────────
            "r_pre_target_entropy",
            "b_pre_target_entropy",  # entropy of head/body/leg
            "r_pre_zone_entropy",
            "b_pre_zone_entropy",  # entropy of dist/clinch/grd
            # ── Finish timing profile ────────────────────────────────────────────────
            "r_pre_avg_finish_round",
            "b_pre_avg_finish_round",  # avg round of finishes
            "r_pre_r1_finish_rate",
            "b_pre_r1_finish_rate",  # % wins ended in R1
            "r_pre_early_finish_score",
            "b_pre_early_finish_score",  # rate × (1/avg_rnd)
            # ── Comeback and lead-hold rates ─────────────────────────────────────────
            "r_pre_comeback_rate",
            "b_pre_comeback_rate",  # won after losing R1
            "r_pre_lead_hold_rate",
            "b_pre_lead_hold_rate",  # won after winning R1
            # ── Power and chin signals ───────────────────────────────────────────────
            "r_pre_kd_efficiency",
            "b_pre_kd_efficiency",  # KDs per 100 sig_str
            "r_pre_r1_head_pressure",
            "b_pre_r1_head_pressure",  # opp R1 head strikes absorbed
            # ── Style composite indices ──────────────────────────────────────────────
            "r_pre_pressure_index",
            "b_pre_pressure_index",  # zone_shift + clinch_esc + td_slope
            "r_pre_counter_index",
            "b_pre_counter_index",  # R1 win rate × neg output slope
            # ── R2→R3 closing momentum ───────────────────────────────────────────────
            "r_pre_r2_to_r3_momentum",
            "b_pre_r2_to_r3_momentum",  # R3 sig_str − R2 sig_str
            # ── Output consistency ───────────────────────────────────────────────────
            "r_pre_output_cv",
            "b_pre_output_cv",  # coeff of variation R1/R2/R3
            "r_pre_output_consistency",
            "b_pre_output_consistency",  # 1/(1+cv)
            # ── Control-time ground-and-pound rate ───────────────────────────────────
            "r_pre_ctrl_gnp_rate",
            "b_pre_ctrl_gnp_rate",  # ground strikes per ctrl min
            # ── Strike quality evolution (sig/total ratio) ───────────────────────────
            "r_pre_r1_sig_pct",
            "b_pre_r1_sig_pct",  # R1 sig/total strike ratio
            "r_pre_r3_sig_pct",
            "b_pre_r3_sig_pct",  # R3 sig/total strike ratio
            "r_pre_sig_pct_evo",
            "b_pre_sig_pct_evo",  # R3 ratio − R1 ratio
            # ── Tactical divergence (R1 vs R3 zone shift) ───────────────────────────
            "r_pre_tactical_div",
            "b_pre_tactical_div",  # euclidean dist R1→R3 zones
            # ── Fatigue composite ────────────────────────────────────────────────────
            "r_pre_fatigue_composite",
            "b_pre_fatigue_composite",  # abs_slope × (1−cardio)
            # ── R1 absorption rate ───────────────────────────────────────────────────
            "r_pre_r1_abs_rate",
            "b_pre_r1_abs_rate",  # R1 absorbed per minute
            # ── Championship damage margin ───────────────────────────────────────────
            "r_pre_r45_dmg_margin",
            "b_pre_r45_dmg_margin",  # R4+R5 output − absorbed
            # ── Combination-punch rate ───────────────────────────────────────────────
            "r_pre_combo_rate",
            "b_pre_combo_rate",  # pct R1s with head+body > 0
            # ── Head strike momentum ─────────────────────────────────────────────────
            "r_pre_head_momentum",
            "b_pre_head_momentum",  # R3 head / R1 head ratio
            # ── Clinch vs ground ratio ───────────────────────────────────────────────
            "r_pre_clinch_grd_ratio",
            "b_pre_clinch_grd_ratio",  # dirty boxer vs wrestler
            # ── Conditional finish rate ──────────────────────────────────────────────
            "r_pre_finish_when_winning",
            "b_pre_finish_when_winning",  # finish rate when R1 won
            # ── Submission round timing ──────────────────────────────────────────────
            "r_pre_sub_round_avg",
            "b_pre_sub_round_avg",  # weighted avg round of sub_att
            # ── Integrated dominance ─────────────────────────────────────────────────
            "r_pre_integrated_dom",
            "b_pre_integrated_dom",  # margin × cardio × finish_rate
            # ── R2 KD threat ─────────────────────────────────────────────────────────
            "r_pre_r2_kd_threat",
            "b_pre_r2_kd_threat",  # avg R2 knockdowns
            # ── Round output consistency ─────────────────────────────────────────────
            "r_pre_r1_r3_consistency",
            "b_pre_r1_r3_consistency",  # 1−|R1−R3|/max(R1+R3)
            # ── R1 control-seeking rate ──────────────────────────────────────────────
            "r_pre_r1_ctrl_opening",
            "b_pre_r1_ctrl_opening",  # pct fights w/ R1 ctrl > 0
            # ── Quality-of-win damage ────────────────────────────────────────────────
            "r_pre_win_r3_margin",
            "b_pre_win_r3_margin",  # avg R3 margin in won fights
            # ── Tier 12l: Per-round defense, adjustment speed, grappling depth ──────
            "r_pre_r1_td_def_rate",
            "b_pre_r1_td_def_rate",  # R1 TD defense rate
            "r_pre_r1_td_pressure",
            "b_pre_r1_td_pressure",  # opp avg R1 TD attempts
            "r_pre_r2_head_acc",
            "b_pre_r2_head_acc",  # R2 head accuracy
            "r_pre_r1r2_head_adj",
            "b_pre_r1r2_head_adj",  # head acc R1→R2 adjustment
            "r_pre_chin_ratio",
            "b_pre_chin_ratio",  # R3 absorbed / R1 absorbed
            "r_pre_r2r3_kd_rate",
            "b_pre_r2r3_kd_rate",  # avg KDs in R2+R3
            "r_pre_sub_after_td",
            "b_pre_sub_after_td",  # R2 sub att when R1 td > 0
            "r_pre_r1_output_eff",
            "b_pre_r1_output_eff",  # R1 sig_str / absorbed
            "r_pre_r3_output_eff",
            "b_pre_r3_output_eff",  # R3 sig_str / absorbed
            "r_pre_output_eff_trend",
            "b_pre_output_eff_trend",  # R3 eff − R1 eff
            "r_pre_r45_head_rate",
            "b_pre_r45_head_rate",  # R4+R5 head strikes (5-rounders)
            "r_pre_r1r2_zone_adj",
            "b_pre_r1r2_zone_adj",  # zone shift R1→R2 (early adjustment)
            "r_pre_r23_sub_rate",
            "b_pre_r23_sub_rate",  # sub att in R2+R3 per fight
            "r_pre_r45_sub_rate_l",
            "b_pre_r45_sub_rate_l",  # R4+R5 sub attempts (5-rounders)
            "r_pre_r45_ground_rate",
            "b_pre_r45_ground_rate",  # R4+R5 ground strikes (5-rounders)
            "r_pre_r3_leg_rate",
            "b_pre_r3_leg_rate",  # R3 leg strikes avg
            # ── Tier 12m: Style entropy adaptation, grappling efficiency, resilience ─
            "r_pre_r2_target_ent",
            "b_pre_r2_target_ent",  # R2 head/body/leg entropy
            "r_pre_r2_zone_ent",
            "b_pre_r2_zone_ent",  # R2 dist/clinch/grd entropy
            "r_pre_r1r2_tgt_shift",
            "b_pre_r1r2_tgt_shift",  # R2 target ent − R1 target ent
            "r_pre_r1r2_zone_shift",
            "b_pre_r1r2_zone_shift",  # R2 zone ent − R1 zone ent
            "r_pre_ctrl_per_td",
            "b_pre_ctrl_per_td",  # R1 ctrl_sec / R1 td (dominance)
            "r_pre_finish_rnd_ent",
            "b_pre_finish_rnd_ent",  # entropy of finish rounds
            "r_pre_r3_body_pct",
            "b_pre_r3_body_pct",  # R3 body / R3 sig_str
            "r_pre_body_pct_evo",
            "b_pre_body_pct_evo",  # R3 body pct − R1 body pct
            "r_pre_r1_cg_link",
            "b_pre_r1_cg_link",  # R1 ground / R1 clinch
            "r_pre_output_in_loss",
            "b_pre_output_in_loss",  # avg sig_str in lost fights
            "r_pre_rev_per_ctrl",
            "b_pre_rev_per_ctrl",  # reversals per ctrl minute
            "r_pre_r2_r1_ratio",
            "b_pre_r2_r1_ratio",  # R2 / R1 output ratio
            "r_pre_r3_leg_pct",
            "b_pre_r3_leg_pct",  # R3 leg / R3 sig_str
            "r_pre_win_r3_output",
            "b_pre_win_r3_output",  # avg R3 sig_str in won fights
            # ── Tier 12n: Finish method diversity, mid-fight adjustment, decision quality ─
            "r_pre_method_entropy",
            "b_pre_method_entropy",  # KO/sub/dec win diversity
            "r_pre_r2_output_eff",
            "b_pre_r2_output_eff",  # R2 sig_str / absorbed
            "r_pre_r1_front_load",
            "b_pre_r1_front_load",  # R1 slpm / career avg slpm
            "r_pre_r2_td_acc",
            "b_pre_r2_td_acc",  # R2 td / td_att
            "r_pre_dec_win_margin",
            "b_pre_dec_win_margin",  # R3 margin in decision wins
            "r_pre_ctrl_esc_ratio",
            "b_pre_ctrl_esc_ratio",  # R3 ctrl / R1 ctrl ratio
            "r_pre_absorbed_in_loss",
            "b_pre_absorbed_in_loss",  # avg absorbed in lost fights
            "r_pre_grapple_usage",
            "b_pre_grapple_usage",  # total td_att+sub_att R1-R3
            "r_pre_kd_r1_conc",
            "b_pre_kd_r1_conc",  # % of KDs occurring in R1
            "r_pre_finish_when_behind",
            "b_pre_finish_when_behind",  # finish rate when R1 margin < 0
            "r_pre_r2_win_rate",
            "b_pre_r2_win_rate",  # % fights won R2 by output
            "r_pre_r3_ctrl_rate",
            "b_pre_r3_ctrl_rate",  # avg R3 control seconds
            "r_pre_r1_body_rate",
            "b_pre_r1_body_rate",  # avg R1 body strikes
            "r_pre_r3_dist_pct",
            "b_pre_r3_dist_pct",  # R3 dist / R3 sig_str
            # ── Tier 12o: Per-round reversals ────────────────────────────────────────────
            "r_pre_r1_rev_rate",
            "b_pre_r1_rev_rate",  # avg R1 reversals per fight
            "r_pre_late_rev_escalation",
            "b_pre_late_rev_escalation",  # R3 revs − R1 revs (scramble late)
            # ── Tier 12o: R4/R5 TD accuracy ──────────────────────────────────────────────
            "r_pre_r45_td_acc",
            "b_pre_r45_td_acc",  # R4+R5 TD accuracy (5-rounders)
            # ── Tier 12o: Full 4-stage grappling depth chain ─────────────────────────────
            "r_pre_grapple_depth_4",
            "b_pre_grapple_depth_4",  # TD×ctrl×GnP×sub multiplicative
            # ── Tier 12o: Career-normalised KD absorption rate ───────────────────────────
            "r_pre_kd_absorbed_rate",
            "b_pre_kd_absorbed_rate",  # KDs absorbed per fight
            # ── Tier 12o: R4/R5 leg escalation ───────────────────────────────────────────
            "r_pre_r45_leg_rate",
            "b_pre_r45_leg_rate",  # avg leg kicks per round in R4+R5
            "r_pre_r45_vs_r1_leg_ratio",
            "b_pre_r45_vs_r1_leg_ratio",  # R45 leg / R1 leg ratio
            # ── Tier 12o: R2 conditional adjustment scores ───────────────────────────────
            "r_pre_r2_adj_score_loss",
            "b_pre_r2_adj_score_loss",  # R2 output Δ when trailing after R1
            "r_pre_r2_adj_score_win",
            "b_pre_r2_adj_score_win",  # R2 output Δ when leading after R1
            # ── Tier 12o: Cardio endurance features ──────────────────────────────────────
            "r_pre_output_decay_rate",
            "b_pre_output_decay_rate",  # linear slope R1→R2→R3 per fight
            "r_pre_all_round_cardio",
            "b_pre_all_round_cardio",  # R5/R1 output in 5-rounders
            "r_pre_r2_retention",
            "b_pre_r2_retention",  # avg R2/R1 output ratio
            "r_pre_r3_retention",
            "b_pre_r3_retention",  # avg R3/R2 output ratio
            "r_pre_accuracy_retention",
            "b_pre_accuracy_retention",  # avg R3_acc / R1_acc ratio
            "r_pre_pace_variance_r123",
            "b_pre_pace_variance_r123",  # std of R1/R2/R3 output per fight
            "r_pre_endurance_composite",
            "b_pre_endurance_composite",  # cardio × r3_ret × acc_ret
            # ── Tier 12p: Round conditional momentum sequences ────────────────────────
            "r_pre_momentum_r12_win",
            "b_pre_momentum_r12_win",  # P(win R2 | win R1)
            "r_pre_momentum_r12_loss",
            "b_pre_momentum_r12_loss",  # P(win R2 | lose R1)
            "r_pre_momentum_r23_win",
            "b_pre_momentum_r23_win",  # P(win R3 | win R2)
            "r_pre_momentum_r23_loss",
            "b_pre_momentum_r23_loss",  # P(win R3 | lose R2)
            # ── Tier 12p: Quality-adjusted round output ────────────────────────────────
            "r_pre_r1_quality_output",
            "b_pre_r1_quality_output",  # R1 sig_str² / att (volume × accuracy)
            "r_pre_r3_quality_output",
            "b_pre_r3_quality_output",  # R3 sig_str² / att
            "r_pre_quality_output_ratio",
            "b_pre_quality_output_ratio",  # R3_quality / R1_quality
            # ── Tier 12p: Output under punishment ─────────────────────────────────────
            "r_pre_output_vs_punishment",
            "b_pre_output_vs_punishment",  # avg output in high-absorption rounds
            # ── Tier 12p: R5/R4 terminal surge ────────────────────────────────────────
            "r_pre_r5_r4_surge",
            "b_pre_r5_r4_surge",  # R5/R4 output ratio (final-round gear)
            # ── Tier 12p: Sub escalation ratio ────────────────────────────────────────
            "r_pre_sub_escalation_ratio",
            "b_pre_sub_escalation_ratio",  # R3/R1 sub attempt ratio
            # ── Tier 12p: Burst index ──────────────────────────────────────────────────
            "r_pre_burst_index",
            "b_pre_burst_index",  # R1 / (R1+R2+R3) output share
            # ── Tier 12p: Ctrl GnP efficiency slope ───────────────────────────────────
            "r_pre_ctrl_gnp_slope",
            "b_pre_ctrl_gnp_slope",  # (R3 grd/ctrl) − (R1 grd/ctrl)
            # ── Tier 12p: R3 striking under takedown threat ───────────────────────────
            "r_pre_r3_str_vs_td_pressure",
            "b_pre_r3_str_vs_td_pressure",  # R3 sig_str / max(R1 td_att, 1)
            # ── Tier 13: Exponentially Decayed Average Features ──────────────────────
            # da = decayed average (alpha=0.8, recent fights weighted ~2× vs 5 fights ago)
            # dapa = decayed average adjusted for opponent quality (× opp_elo / 1500)
            "r_pre_da_sig_str_acc",
            "b_pre_da_sig_str_acc",
            "r_pre_da_td_acc",
            "b_pre_da_td_acc",
            "r_pre_da_head_landed",
            "b_pre_da_head_landed",
            "r_pre_da_head_acc",
            "b_pre_da_head_acc",
            "r_pre_da_body_acc",
            "b_pre_da_body_acc",
            "r_pre_da_distance_acc",
            "b_pre_da_distance_acc",
            "r_pre_da_head_defense",
            "b_pre_da_head_defense",
            "r_pre_da_body_defense",
            "b_pre_da_body_defense",
            "r_pre_da_distance_defense",
            "b_pre_da_distance_defense",
            "r_pre_da_ground_defense",
            "b_pre_da_ground_defense",
            "r_pre_da_td_defense",
            "b_pre_da_td_defense",
            "r_pre_da_sub_att",
            "b_pre_da_sub_att",
            "r_pre_da_kd",
            "b_pre_da_kd",
            "r_pre_da_ko",
            "b_pre_da_ko",
            "r_pre_da_win_ratio",
            "b_pre_da_win_ratio",
            "r_pre_da_ctrl_r1",
            "b_pre_da_ctrl_r1",
            "r_pre_da_clinch_pm",
            "b_pre_da_clinch_pm",
            "r_pre_da_opp_leg_pm",
            "b_pre_da_opp_leg_pm",
            "r_pre_da_opp_ctrl_r1_pm",
            "b_pre_da_opp_ctrl_r1_pm",
            "r_pre_da_opp_sub_pm",
            "b_pre_da_opp_sub_pm",
            "r_pre_da_opp_rev_r1",
            "b_pre_da_opp_rev_r1",
            "r_pre_da_r1_strikes",
            "b_pre_da_r1_strikes",
            "r_pre_da_reversals",
            "b_pre_da_reversals",
            "r_pre_da_dist_landing_ratio",
            "b_pre_da_dist_landing_ratio",
            "r_pre_da_opp_kd",
            "b_pre_da_opp_kd",
            "r_pre_dapa_sig_str_acc",
            "b_pre_dapa_sig_str_acc",
            "r_pre_dapa_head_acc",
            "b_pre_dapa_head_acc",
            "r_pre_dapa_body_acc",
            "b_pre_dapa_body_acc",
            "r_pre_dapa_distance_acc",
            "b_pre_dapa_distance_acc",
            "r_pre_dapa_head_defense",
            "b_pre_dapa_head_defense",
            "r_pre_dapa_dist_defense",
            "b_pre_dapa_dist_defense",
            "r_pre_dapa_ground_defense",
            "b_pre_dapa_ground_defense",
            "r_pre_dapa_r1_strikes",
            "b_pre_dapa_r1_strikes",
            "r_pre_dapa_reversals",
            "b_pre_dapa_reversals",
            "r_pre_dapa_dist_landing_ratio",
            "b_pre_dapa_dist_landing_ratio",
            "r_pre_dapa_head_landing_ratio",
            "b_pre_dapa_head_landing_ratio",
            "r_pre_da_age",
            "b_pre_da_age",
            "r_pre_da_ufc_age",
            "b_pre_da_ufc_age",
            "r_pre_da_reach_ratio",
            "b_pre_da_reach_ratio",
            "r_pre_da_days_since",
            "b_pre_da_days_since",
        ]
        for col in new_cols:
            self.df[col] = 0.0

        if "r_elo_pre_fight" not in self.df.columns:
            self.df["r_elo_pre_fight"] = 1500.0
        if "b_elo_pre_fight" not in self.df.columns:
            self.df["b_elo_pre_fight"] = 1500.0

        for _wc_seq, (idx, row) in enumerate(self.df.iterrows()):
            r = str(row.get("r_fighter", "")).strip()
            b = str(row.get("b_fighter", "")).strip()
            winner = str(row.get("winner", "")).strip()
            method = str(row.get("method", "")).strip()
            is_title = bool(row.get("is_title_bout", False))

            rs = fighter_stats[r]
            bs = fighter_stats[b]
            self.df.at[idx, "_wc_chrono_seq"] = float(_wc_seq)

            def _safe(v, default=0.0):
                if v is None:
                    return default
                try:
                    f = float(v)
                    return default if math.isnan(f) else f
                except (TypeError, ValueError):
                    return default

            r_nf = max(rs["total_fights"], 1)
            b_nf = max(bs["total_fights"], 1)

            _wck = (
                str(row.get("weight_class", "")).strip(),
                str(row.get("gender", "")).lower().strip(),
                int(_safe(row.get("total_rounds", 3), 3)) or 3,
            )
            _wcc = self._wc_finish_counts[_wck]
            _wko, _wsu = _wcc["ko"], _wcc["sub"]
            _wfin = _wko + _wsu
            if _wfin > 0:
                self.df.at[idx, "wc_finish_route_prior"] = (_wsu / _wfin) - (
                    _wko / _wfin
                )
            else:
                self.df.at[idx, "wc_finish_route_prior"] = 0.0

            self.df.at[idx, "r_pre_wins"] = rs["wins"]
            self.df.at[idx, "r_pre_losses"] = rs["losses"]
            self.df.at[idx, "r_pre_draws"] = rs["draws"]
            self.df.at[idx, "b_pre_wins"] = bs["wins"]
            self.df.at[idx, "b_pre_losses"] = bs["losses"]
            self.df.at[idx, "b_pre_draws"] = bs["draws"]
            self.df.at[idx, "r_pre_ko_wins"] = rs["ko_wins"]
            self.df.at[idx, "r_pre_sub_wins"] = rs["sub_wins"]
            self.df.at[idx, "r_pre_dec_wins"] = rs["dec_wins"]
            self.df.at[idx, "b_pre_ko_wins"] = bs["ko_wins"]
            self.df.at[idx, "b_pre_sub_wins"] = bs["sub_wins"]
            self.df.at[idx, "b_pre_dec_wins"] = bs["dec_wins"]
            self.df.at[idx, "r_pre_ko_losses"] = rs["ko_losses"]
            self.df.at[idx, "r_pre_sub_losses"] = rs["sub_losses"]
            self.df.at[idx, "r_pre_dec_losses"] = rs["dec_losses"]
            self.df.at[idx, "b_pre_ko_losses"] = bs["ko_losses"]
            self.df.at[idx, "b_pre_sub_losses"] = bs["sub_losses"]
            self.df.at[idx, "b_pre_dec_losses"] = bs["dec_losses"]
            self.df.at[idx, "r_pre_total_fights"] = rs["total_fights"]
            self.df.at[idx, "b_pre_total_fights"] = bs["total_fights"]
            self.df.at[idx, "r_pre_finish_rate"] = rs["finish_rate"]
            self.df.at[idx, "b_pre_finish_rate"] = bs["finish_rate"]
            self.df.at[idx, "r_pre_win_streak"] = rs["win_streak"]
            self.df.at[idx, "b_pre_win_streak"] = bs["win_streak"]
            self.df.at[idx, "r_pre_loss_streak"] = rs["loss_streak"]
            self.df.at[idx, "b_pre_loss_streak"] = bs["loss_streak"]
            self.df.at[idx, "r_pre_title_fights"] = rs["title_fights"]
            self.df.at[idx, "b_pre_title_fights"] = bs["title_fights"]
            self.df.at[idx, "r_pre_title_wins"] = rs["title_wins"]
            self.df.at[idx, "b_pre_title_wins"] = bs["title_wins"]
            self.df.at[idx, "r_pre_avg_fight_time"] = rs["avg_fight_time"]
            self.df.at[idx, "b_pre_avg_fight_time"] = bs["avg_fight_time"]

            sig_str_att_r = max(rs["total_sig_str_att"], 1)
            sig_str_att_b = max(bs["total_sig_str_att"], 1)
            td_att_r = max(rs["total_td_att"], 1)
            td_att_b = max(bs["total_td_att"], 1)
            fight_time_r = max(rs["total_fight_time"], 1.0)
            fight_time_b = max(bs["total_fight_time"], 1.0)

            self.df.at[idx, "r_pre_sig_str_acc"] = rs["total_sig_str"] / \
                sig_str_att_r
            self.df.at[idx, "b_pre_sig_str_acc"] = bs["total_sig_str"] / \
                sig_str_att_b
            self.df.at[idx, "r_pre_td_acc"] = rs["total_td"] / td_att_r
            self.df.at[idx, "b_pre_td_acc"] = bs["total_td"] / td_att_b
            self.df.at[idx, "r_pre_sub_att_rate"] = rs["total_sub_att"] / r_nf
            self.df.at[idx, "b_pre_sub_att_rate"] = bs["total_sub_att"] / b_nf
            self.df.at[idx, "r_pre_kd_rate"] = rs["total_kd"] / r_nf
            self.df.at[idx, "b_pre_kd_rate"] = bs["total_kd"] / b_nf
            self.df.at[idx, "r_pre_ctrl_avg"] = rs["total_ctrl_sec"] / r_nf
            self.df.at[idx, "b_pre_ctrl_avg"] = bs["total_ctrl_sec"] / b_nf

            r_min = fight_time_r / 60.0
            b_min = fight_time_b / 60.0
            self.df.at[idx, "r_pre_SLpM"] = rs["total_sig_str"] / \
                max(r_min, 1.0)
            self.df.at[idx, "b_pre_SLpM"] = bs["total_sig_str"] / \
                max(b_min, 1.0)
            self.df.at[idx, "r_pre_SApM"] = rs["total_str"] / max(r_min, 1.0)
            self.df.at[idx, "b_pre_SApM"] = bs["total_str"] / max(b_min, 1.0)
            self.df.at[idx, "r_pre_td_avg"] = rs["total_td"] / max(r_min, 1.0)
            self.df.at[idx, "b_pre_td_avg"] = bs["total_td"] / max(b_min, 1.0)

            # Per-15-minute grappling rates (common time base for method features)
            self.df.at[idx, "r_pre_sub_att_per15"] = (
                15.0 * rs["total_sub_att"] / max(r_min, 0.25)
            )
            self.df.at[idx, "b_pre_sub_att_per15"] = (
                15.0 * bs["total_sub_att"] / max(b_min, 0.25)
            )
            self.df.at[idx, "r_pre_ctrl_min_per15"] = (
                15.0 * (rs["total_ctrl_sec"] / 60.0) / max(r_min, 0.25)
            )
            self.df.at[idx, "b_pre_ctrl_min_per15"] = (
                15.0 * (bs["total_ctrl_sec"] / 60.0) / max(b_min, 0.25)
            )
            self.df.at[idx, "r_pre_td_per15"] = 15.0 * \
                rs["total_td"] / max(r_min, 0.25)
            self.df.at[idx, "b_pre_td_per15"] = 15.0 * \
                bs["total_td"] / max(b_min, 0.25)

            hist_r3 = rs["_history"][-3:]
            hist_b3 = bs["_history"][-3:]
            self.df.at[idx, "r_rolling3_wins"] = sum(h["won"] for h in hist_r3)
            self.df.at[idx, "b_rolling3_wins"] = sum(h["won"] for h in hist_b3)
            self.df.at[idx, "r_rolling3_sig_str"] = (
                np.mean([h["sig_str"] for h in hist_r3]) if hist_r3 else 0.0
            )
            self.df.at[idx, "b_rolling3_sig_str"] = (
                np.mean([h["sig_str"] for h in hist_b3]) if hist_b3 else 0.0
            )
            self.df.at[idx, "r_rolling3_td"] = (
                np.mean([h["td"] for h in hist_r3]) if hist_r3 else 0.0
            )
            self.df.at[idx, "b_rolling3_td"] = (
                np.mean([h["td"] for h in hist_b3]) if hist_b3 else 0.0
            )
            self.df.at[idx, "r_rolling3_kd"] = (
                np.mean([h["kd"] for h in hist_r3]) if hist_r3 else 0.0
            )
            self.df.at[idx, "b_rolling3_kd"] = (
                np.mean([h["kd"] for h in hist_b3]) if hist_b3 else 0.0
            )
            self.df.at[idx, "r_rolling3_sub_att"] = (
                np.mean([h["sub_att"] for h in hist_r3]) if hist_r3 else 0.0
            )
            self.df.at[idx, "b_rolling3_sub_att"] = (
                np.mean([h["sub_att"] for h in hist_b3]) if hist_b3 else 0.0
            )

            hist_r5 = rs["_history"][-5:]
            hist_b5 = bs["_history"][-5:]
            self.df.at[idx, "r_rolling5_wins"] = sum(h["won"] for h in hist_r5)
            self.df.at[idx, "b_rolling5_wins"] = sum(h["won"] for h in hist_b5)
            self.df.at[idx, "r_rolling5_sig_str"] = (
                np.mean([h["sig_str"] for h in hist_r5]) if hist_r5 else 0.0
            )
            self.df.at[idx, "b_rolling5_sig_str"] = (
                np.mean([h["sig_str"] for h in hist_b5]) if hist_b5 else 0.0
            )
            self.df.at[idx, "r_rolling5_td"] = (
                np.mean([h["td"] for h in hist_r5]) if hist_r5 else 0.0
            )
            self.df.at[idx, "b_rolling5_td"] = (
                np.mean([h["td"] for h in hist_b5]) if hist_b5 else 0.0
            )
            self.df.at[idx, "r_rolling5_kd"] = (
                np.mean([h["kd"] for h in hist_r5]) if hist_r5 else 0.0
            )
            self.df.at[idx, "b_rolling5_kd"] = (
                np.mean([h["kd"] for h in hist_b5]) if hist_b5 else 0.0
            )

            r_glicko_snap = self.feature_engineer.glicko2_get(r)
            b_glicko_snap = self.feature_engineer.glicko2_get(b)
            self.df.at[idx, "r_glicko_pre_r"] = r_glicko_snap[0]
            self.df.at[idx, "r_glicko_pre_rd"] = r_glicko_snap[1]
            self.df.at[idx, "r_glicko_pre_vol"] = r_glicko_snap[2]
            self.df.at[idx, "b_glicko_pre_r"] = b_glicko_snap[0]
            self.df.at[idx, "b_glicko_pre_rd"] = b_glicko_snap[1]
            self.df.at[idx, "b_glicko_pre_vol"] = b_glicko_snap[2]

            current_date = row.get("event_date", None)
            for corner, fs in [("r", rs), ("b", bs)]:
                if fs["fight_dates"] and pd.notna(current_date):
                    last_date = fs["fight_dates"][-1]
                    try:
                        days_gap = (current_date - last_date).days
                    except Exception:
                        days_gap = 365
                else:
                    days_gap = 365
                self.df.at[idx, f"{corner}_days_since_last"] = days_gap

            self.df.at[idx, "r_pre_kd_absorbed"] = rs.get("kd_absorbed", 0)
            self.df.at[idx, "b_pre_kd_absorbed"] = bs.get("kd_absorbed", 0)

            # Decision-type win/loss counts (written pre-fight, before update_fighter)
            for _corner, _fs in [("r", rs), ("b", bs)]:
                self.df.at[idx,
                           f"{_corner}_pre_unanimous_wins"] = _fs["unanimous_wins"]
                self.df.at[idx,
                           f"{_corner}_pre_majority_wins"] = _fs["majority_wins"]
                self.df.at[idx,
                           f"{_corner}_pre_split_wins"] = _fs["split_wins"]
                self.df.at[idx,
                           f"{_corner}_pre_split_losses"] = _fs["split_losses"]
                _loss_d = max(_fs["losses"], 1)
                self.df.at[idx, f"{_corner}_pre_early_finish_loss_rate"] = (
                    _fs["early_finish_losses"] / _loss_d
                )
                self.df.at[idx, f"{_corner}_pre_late_finish_loss_rate"] = (
                    _fs["late_finish_losses"] / _loss_d
                )
                _sw_d = max(_fs["sub_wins"], 1)
                _kw_d = max(_fs["ko_wins"], 1)
                self.df.at[idx, f"{_corner}_pre_sub_win_after_td_rate"] = (
                    _fs["sub_wins_with_td"] / _sw_d
                )
                self.df.at[idx, f"{_corner}_pre_ko_win_after_kd_rate"] = (
                    _fs["ko_wins_with_kd"] / _kw_d
                )
                _fl1 = max(_fs["fights_lost_r1"], 1)
                self.df.at[idx, f"{_corner}_pre_finish_loss_lost_r1_rate"] = (
                    _fs["finish_losses_after_lost_r1"] / _fl1
                )
                self.df.at[idx, f"{_corner}_pre_sub_win_lost_r1_rate"] = (
                    _fs["sub_wins_after_lost_r1"] / _fl1
                )
                self.df.at[idx, f"{_corner}_pre_ko_win_lost_r1_rate"] = (
                    _fs["ko_wins_after_lost_r1"] / _fl1
                )
                _fw1 = max(_fs["fights_won_r1"], 1)
                self.df.at[idx, f"{_corner}_pre_dec_win_won_r1_rate"] = (
                    _fs["dec_wins_after_won_r1"] / _fw1
                )
                _dom_w = max(
                    _fs["unanimous_wins"]
                    + _fs["majority_wins"]
                    + _fs["ko_wins"]
                    + _fs["sub_wins"],
                    1,
                )
                self.df.at[idx, f"{_corner}_pre_finish_win_phys_dom_rate"] = (
                    _fs["finish_wins_phys_dom"] / _dom_w
                )
                self.df.at[idx, f"{_corner}_pre_dom_finish_conv_exact"] = _fs[
                    "t43_finish_dom_wins"
                ] / max(_fs["t43_dom_wins"], 1)
                _tex_td = max(_fs["total_td"] + _fs["total_sub_att"], 1)
                self.df.at[idx, f"{_corner}_pre_ex_sub_conv"] = (
                    _fs["sub_wins"] / _tex_td
                )
                _tex_cm = max(_fs["total_ctrl_sec"] / 60.0, 1e-6)
                self.df.at[idx, f"{_corner}_pre_ex_ctrl_sub_conv"] = (
                    _fs["sub_wins"] / _tex_cm
                )
                _tex_kh = max(_fs["total_kd"] + _fs["total_head"], 1)
                self.df.at[idx, f"{_corner}_pre_ex_ko_tko_conv"] = (
                    _fs["ko_wins"] / _tex_kh
                )

            r_opp_elos = rs["opp_elo_history"][-5:]
            b_opp_elos = bs["opp_elo_history"][-5:]
            self.df.at[idx, "r_avg_opp_elo_L5"] = (
                float(np.mean(r_opp_elos)) if r_opp_elos else 1500.0
            )
            self.df.at[idx, "b_avg_opp_elo_L5"] = (
                float(np.mean(b_opp_elos)) if b_opp_elos else 1500.0
            )

            def _trajectory_slope(history):
                if len(history) < 2:
                    return 0.0
                results = [h["won"] for h in history[-3:]]
                x = np.arange(len(results), dtype=float)
                try:
                    slope = np.polyfit(x, results, 1)[0]
                except Exception:
                    slope = 0.0
                return float(slope)

            self.df.at[idx, "r_trajectory_3"] = _trajectory_slope(
                rs["_history"])
            self.df.at[idx, "b_trajectory_3"] = _trajectory_slope(
                bs["_history"])

            for corner, fs in [("r", rs), ("b", bs)]:
                total_pos = max(
                    fs["total_distance"] + fs["total_clinch"] +
                    fs["total_ground"], 1
                )
                total_tgt = max(
                    fs["total_head"] + fs["total_body"] + fs["total_leg"], 1
                )
                self.df.at[idx, f"{corner}_pre_distance_pct"] = (
                    fs["total_distance"] / total_pos
                )
                self.df.at[idx, f"{corner}_pre_clinch_pct"] = (
                    fs["total_clinch"] / total_pos
                )
                self.df.at[idx, f"{corner}_pre_ground_pct"] = (
                    fs["total_ground"] / total_pos
                )
                self.df.at[idx,
                           f"{corner}_pre_head_pct"] = fs["total_head"] / total_tgt
                self.df.at[idx,
                           f"{corner}_pre_body_pct"] = fs["total_body"] / total_tgt
                self.df.at[idx,
                           f"{corner}_pre_leg_pct"] = fs["total_leg"] / total_tgt

                str_def_att = max(fs["str_def_attempts"], 1)
                td_def_att = max(fs["td_def_attempts"], 1)
                self.df.at[idx, f"{corner}_pre_str_def"] = (
                    fs["str_def_success"] / str_def_att
                )
                self.df.at[idx, f"{corner}_pre_td_def"] = (
                    fs["td_def_success"] / td_def_att
                )
                # Career-normalised KD absorption rate (chin durability signal)
                self.df.at[idx, f"{corner}_pre_kd_absorbed_rate"] = fs[
                    "kd_absorbed"
                ] / max(fs["total_fights"], 1)

                wins = max(fs["wins"], 1)
                self.df.at[idx, f"{corner}_pre_early_finish_rate"] = (
                    fs["early_finishes"] / wins
                )
                self.df.at[idx, f"{corner}_pre_late_finish_rate"] = (
                    fs["late_finishes"] / wins
                )
                self.df.at[idx, f"{corner}_pre_first_round_ko_rate"] = fs[
                    "first_round_kos"
                ] / max(fs["total_fights"], 1)
                self.df.at[idx, f"{corner}_pre_total_rounds_fought"] = fs[
                    "total_rounds_fought"
                ]
                self.df.at[idx, f"{corner}_pre_five_round_fights"] = fs[
                    "five_round_fights"
                ]

                hist10 = fs["rolling10_history"][-10:]
                self.df.at[idx, f"{corner}_rolling10_wins"] = sum(
                    h["won"] for h in hist10
                )
                self.df.at[idx, f"{corner}_rolling10_sig_str"] = (
                    float(np.mean([h["sig_str"]
                          for h in hist10])) if hist10 else 0.0
                )
                self.df.at[idx, f"{corner}_rolling10_td"] = (
                    float(np.mean([h["td"]
                          for h in hist10])) if hist10 else 0.0
                )
                self.df.at[idx, f"{corner}_rolling10_kd"] = (
                    float(np.mean([h["kd"]
                          for h in hist10])) if hist10 else 0.0
                )
                self.df.at[idx, f"{corner}_rolling10_finishes"] = sum(
                    h.get("finished", 0) for h in hist10
                )

                self.df.at[idx,
                           f"{corner}_career_elo_peak"] = fs["career_elo_peak"]
                self.df.at[idx,
                           f"{corner}_fights_since_peak"] = fs["fights_since_peak"]

                elite_fights = max(fs["vs_elite_fights"], 1)
                striker_fights = max(fs["vs_striker_fights"], 1)
                grappler_fights = max(fs["vs_grappler_fights"], 1)
                self.df.at[idx, f"{corner}_vs_elite_win_rate"] = (
                    fs["vs_elite_wins"] / elite_fights
                )
                self.df.at[idx, f"{corner}_vs_striker_win_rate"] = (
                    fs["vs_striker_wins"] / striker_fights
                )
                self.df.at[idx, f"{corner}_vs_grappler_win_rate"] = (
                    fs["vs_grappler_wins"] / grappler_fights
                )

                self.df.at[idx, f"{corner}_last_fight_was_win"] = float(
                    fs["last_fight_was_win"]
                )
                self.df.at[idx, f"{corner}_last_fight_was_finish"] = float(
                    fs["last_fight_was_finish"]
                )

                _h10_v = fs["rolling10_history"]
                _h5_v = _h10_v[-5:]
                _h10_vl = _h10_v[-10:]
                _fin_l5 = sum(h.get("finished", 0)
                              for h in _h5_v) / max(len(_h5_v), 1)
                _fin_l10 = sum(h.get("finished", 0) for h in _h10_vl) / max(
                    len(_h10_vl), 1
                )
                self.df.at[idx, f"{corner}_pre_finish_rate_l5"] = _fin_l5
                self.df.at[idx, f"{corner}_pre_finish_rate_l10"] = _fin_l10
                _sig_vals = [h.get("sig_str", 0) for h in _h10_vl]
                if len(_sig_vals) >= 3:
                    _mu_v = float(np.mean(_sig_vals))
                    _cv_v = float(np.std(_sig_vals)) / (_mu_v + 0.1)
                else:
                    _cv_v = 0.0
                self.df.at[idx, f"{corner}_pre_slpm_cv"] = _cv_v
                _corner_age = _safe(
                    row.get(f"{corner}_age_at_event", 28.0), 28.0)
                self.df.at[idx, f"{corner}_pre_mileage_adj_age"] = (
                    _corner_age * fs["total_rounds_fought"] / 100.0
                )

                self.df.at[idx, f"{corner}_pre_rolling3_wins"] = float(
                    sum(h.get("won", 0) for h in _h10_v[-3:])
                )

                _slpm_l5 = [h.get("slpm", 0.0) for h in _h10_v[-5:]]
                _slpm_l10 = [h.get("slpm", 0.0) for h in _h10_v[-10:]]
                self.df.at[idx, f"{corner}_pre_rolling5_slpm"] = (
                    float(np.mean(_slpm_l5)) if _slpm_l5 else 0.0
                )
                self.df.at[idx, f"{corner}_pre_slpm_std_l10"] = (
                    float(np.std(_slpm_l10)) if len(_slpm_l10) >= 2 else 0.0
                )

                _dr_l10 = [h.get("damage_ratio", 1.0) for h in _h10_v[-10:]]
                self.df.at[idx, f"{corner}_pre_damage_ratio_std_l10"] = (
                    float(np.std(_dr_l10)) if len(_dr_l10) >= 2 else 0.0
                )

                _dp_l5 = [h.get("distance_pct", 0.5) for h in _h10_v[-5:]]
                _dp_all = [h.get("distance_pct", 0.5) for h in _h10_v]
                _dp_l5m = float(np.mean(_dp_l5)) if _dp_l5 else 0.5
                _dp_allm = float(np.mean(_dp_all)) if _dp_all else 0.5
                self.df.at[idx,
                           f"{corner}_pre_tactical_evolution"] = _dp_l5m - _dp_allm

                # Per-round derived stats: rolling L5 averages from h10 history
                _h_rd = _h10_v[-5:]
                self.df.at[idx, f"{corner}_pre_rd1_slpm_avg"] = (
                    float(np.mean([h.get("rd1_slpm", 0.0) for h in _h_rd]))
                    if _h_rd
                    else 0.0
                )
                self.df.at[idx, f"{corner}_pre_rd1_kd_rate"] = (
                    float(np.mean([h.get("rd1_kd", 0.0) for h in _h_rd]))
                    if _h_rd
                    else 0.0
                )
                self.df.at[idx, f"{corner}_pre_rd1_td_rate"] = (
                    float(np.mean([h.get("rd1_td", 0.0) for h in _h_rd]))
                    if _h_rd
                    else 0.0
                )
                self.df.at[idx, f"{corner}_pre_late_td_rate"] = (
                    float(np.mean([h.get("rd3_td", 0.0) for h in _h_rd]))
                    if _h_rd
                    else 0.0
                )

                _rd1_head_pct_vals = [
                    h.get("rd1_head", 0.0) /
                    max(h.get("rd1_head_att", 1.0), 1.0)
                    for h in _h_rd
                    if h.get("rd1_head_att", 0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_rd1_head_pct"] = (
                    float(np.mean(_rd1_head_pct_vals)
                          ) if _rd1_head_pct_vals else 0.45
                )

                # Cardio index: avg R3 sig strikes relative to avg R1 sig strikes (only fights that reached R3).
                # Values below 1 mean output decays late; above 1 means the fighter grows stronger.
                _rd3_pairs = [
                    (h.get("rd1_sig_str", 0.0), h.get("rd3_sig_str", 0.0))
                    for h in _h_rd
                    if h.get("rd3_sig_str", 0.0) > 0
                ]
                if _rd3_pairs:
                    _avg_rd1_c = float(np.mean([p[0] for p in _rd3_pairs]))
                    _avg_rd3_c = float(np.mean([p[1] for p in _rd3_pairs]))
                    self.df.at[idx, f"{corner}_pre_cardio_index"] = _avg_rd3_c / max(
                        _avg_rd1_c, 1.0
                    )
                else:
                    self.df.at[idx, f"{corner}_pre_cardio_index"] = (
                        0.75  # neutral prior: slight expected decay
                    )

                # ── Output slope: pace change R1 → R3 (strikes per round, normalised) ──
                _slope_pairs = [
                    (h.get("rd1_sig_str", 0.0), h.get("rd3_sig_str", 0.0))
                    for h in _h_rd
                    if h.get("rd3_sig_str", 0.0) > 0
                ]
                if _slope_pairs:
                    _s_r1 = float(np.mean([p[0] for p in _slope_pairs]))
                    _s_r3 = float(np.mean([p[1] for p in _slope_pairs]))
                    self.df.at[idx, f"{corner}_pre_output_slope"] = (
                        _s_r3 - _s_r1
                    ) / 2.0
                else:
                    self.df.at[idx, f"{corner}_pre_output_slope"] = 0.0

                # ── Cardio endurance features ─────────────────────────────────────────
                # Output decay rate: linear slope over [R1, R2, R3] per fight (polyfit).
                # Negative = fades, positive = builds. Distinct from output_slope which
                # only compares R1 and R3 aggregate averages.
                _decay_slopes = []
                for _h in _h_rd:
                    _xs, _ys = [], []
                    if _h.get("rd1_sig_str", 0.0) > 0:
                        _xs.append(1)
                        _ys.append(_h["rd1_sig_str"])
                    if _h.get("rd2_sig_str", 0.0) > 0:
                        _xs.append(2)
                        _ys.append(_h["rd2_sig_str"])
                    if _h.get("rd3_sig_str", 0.0) > 0:
                        _xs.append(3)
                        _ys.append(_h["rd3_sig_str"])
                    if len(_xs) >= 2:
                        try:
                            _decay_slopes.append(
                                float(np.polyfit(_xs, _ys, 1)[0]))
                        except Exception:
                            pass
                self.df.at[idx, f"{corner}_pre_output_decay_rate"] = (
                    float(np.mean(_decay_slopes)) if _decay_slopes else 0.0
                )

                # R2 output retention: avg R2/R1 ratio (R2 fade specifically)
                _r2_ret_vals = [
                    _h.get("rd2_sig_str", 0.0) /
                    max(_h.get("rd1_sig_str", 1.0), 1.0)
                    for _h in _h_rd
                    if _h.get("rd2_sig_str", 0.0) > 0 and _h.get("rd1_sig_str", 0.0) > 0
                ]
                _r2_ret = float(np.mean(_r2_ret_vals)) if _r2_ret_vals else 1.0
                self.df.at[idx, f"{corner}_pre_r2_retention"] = _r2_ret

                # R3 output retention: avg R3/R2 ratio (late-round fade R2→R3)
                _r3_ret_vals = [
                    _h.get("rd3_sig_str", 0.0) /
                    max(_h.get("rd2_sig_str", 1.0), 1.0)
                    for _h in _h_rd
                    if _h.get("rd3_sig_str", 0.0) > 0 and _h.get("rd2_sig_str", 0.0) > 0
                ]
                _r3_ret = float(np.mean(_r3_ret_vals)) if _r3_ret_vals else 1.0
                self.df.at[idx, f"{corner}_pre_r3_retention"] = _r3_ret

                # Accuracy retention: avg R3_acc / R1_acc (does accuracy hold under fatigue?)
                _acc_ret_vals = []
                for _h in _h_rd:
                    _r1a = (
                        _h.get("rd1_sig_str", 0.0)
                        / max(_h.get("rd1_sig_str_att", 1.0), 1.0)
                        if _h.get("rd1_sig_str_att", 0.0) > 0
                        else None
                    )
                    _r3a = (
                        _h.get("rd3_sig_str", 0.0)
                        / max(_h.get("rd3_sig_str_att", 1.0), 1.0)
                        if _h.get("rd3_sig_str_att", 0.0) > 0
                        else None
                    )
                    if _r1a and _r3a and _r1a > 0:
                        _acc_ret_vals.append(_r3a / _r1a)
                _acc_ret = float(np.mean(_acc_ret_vals)
                                 ) if _acc_ret_vals else 1.0
                self.df.at[idx, f"{corner}_pre_accuracy_retention"] = _acc_ret

                # Pace variance: std of [R1, R2, R3] output per fight — high = erratic pacing
                _pv_vals = []
                for _h in _h_rd:
                    _rds = [
                        v
                        for v in [
                            _h.get("rd1_sig_str", 0.0),
                            _h.get("rd2_sig_str", 0.0),
                            _h.get("rd3_sig_str", 0.0),
                        ]
                        if v > 0
                    ]
                    if len(_rds) >= 2:
                        _pv_vals.append(float(np.std(_rds)))
                self.df.at[idx, f"{corner}_pre_pace_variance_r123"] = (
                    float(np.mean(_pv_vals)) if _pv_vals else 0.0
                )

                # ── Round-1 accuracy ──
                _r1_acc_vals = [
                    h.get("rd1_sig_str", 0.0) /
                    max(h.get("rd1_sig_str_att", 1.0), 1.0)
                    for h in _h_rd
                    if h.get("rd1_sig_str_att", 0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_r1_acc"] = (
                    float(np.mean(_r1_acc_vals)) if _r1_acc_vals else 0.45
                )

                # ── Late-round accuracy (R3) ──
                _late_acc_vals = [
                    h.get("rd3_sig_str", 0.0) /
                    max(h.get("rd3_sig_str_att", 1.0), 1.0)
                    for h in _h_rd
                    if h.get("rd3_sig_str_att", 0) > 0
                ]
                _late_acc = float(np.mean(_late_acc_vals)
                                  ) if _late_acc_vals else 0.45
                self.df.at[idx, f"{corner}_pre_late_acc"] = _late_acc

                # ── Accuracy fade: R1 acc − late acc (positive = fades under pressure) ──
                _r1_acc_val = self.df.at[idx, f"{corner}_pre_r1_acc"]
                self.df.at[idx,
                           f"{corner}_pre_acc_fade"] = _r1_acc_val - _late_acc

                # ── Championship round metrics (R4/R5, five-round bouts only) ──
                _r45_fights = [h for h in _h_rd if h.get(
                    "is_five_round", 0) == 1]
                if _r45_fights:
                    _r45_ss_avg = float(
                        np.mean(
                            [
                                (h.get("rd4_sig_str", 0.0) +
                                 h.get("rd5_sig_str", 0.0))
                                / 2.0
                                for h in _r45_fights
                            ]
                        )
                    )
                    _r1_ss_avg = float(
                        np.mean([h.get("rd1_sig_str", 0.0)
                                for h in _r45_fights])
                    )
                    _r45_td_avg = float(
                        np.mean(
                            [
                                h.get("rd4_td", 0.0) + h.get("rd5_td", 0.0)
                                for h in _r45_fights
                            ]
                        )
                    )
                    self.df.at[idx,
                               f"{corner}_pre_r45_slpm"] = _r45_ss_avg / 5.0
                    self.df.at[idx, f"{corner}_pre_r45_vs_r1_ratio"] = (
                        _r45_ss_avg / max(_r1_ss_avg, 1.0)
                    )
                    self.df.at[idx, f"{corner}_pre_r45_td_rate"] = _r45_td_avg
                    # R4/R5 TD accuracy (fights that reached championship rounds)
                    _r45_td_total = sum(
                        h.get("rd4_td", 0.0) + h.get("rd5_td", 0.0) for h in _r45_fights
                    )
                    _r45_td_att_total = sum(
                        h.get("rd4_td_att", 0.0) + h.get("rd5_td_att", 0.0)
                        for h in _r45_fights
                    )
                    self.df.at[idx, f"{corner}_pre_r45_td_acc"] = _r45_td_total / max(
                        _r45_td_att_total, 1.0
                    )
                    # R4/R5 leg escalation
                    _r45_leg_avg = float(
                        np.mean(
                            [
                                (h.get("rd4_leg", 0.0) +
                                 h.get("rd5_leg", 0.0)) / 2.0
                                for h in _r45_fights
                            ]
                        )
                    )
                    _r1_leg_in_r45 = float(
                        np.mean([h.get("rd1_leg", 0.0) for h in _r45_fights])
                    )
                    self.df.at[idx,
                               f"{corner}_pre_r45_leg_rate"] = _r45_leg_avg
                    self.df.at[idx, f"{corner}_pre_r45_vs_r1_leg_ratio"] = (
                        _r45_leg_avg / max(_r1_leg_in_r45, 1.0)
                    )
                    # All-round cardio: R5/R1 output ratio (ultimate 5-round endurance test)
                    _r5r1_vals = [
                        h.get("rd5_sig_str", 0.0) /
                        max(h.get("rd1_sig_str", 1.0), 1.0)
                        for h in _r45_fights
                        if h.get("rd1_sig_str", 0.0) > 0
                        and h.get("rd5_sig_str", 0.0) > 0
                    ]
                    self.df.at[idx, f"{corner}_pre_all_round_cardio"] = (
                        float(np.mean(_r5r1_vals)) if _r5r1_vals else 0.0
                    )
                    # R5/R4 terminal surge: did the fighter find another gear in the final round?
                    # Distinct from all_round_cardio (R5/R1) — this captures a late-round acceleration
                    # regardless of how the fight unfolded up to R4.
                    _r5r4_vals = [
                        h.get("rd5_sig_str", 0.0) /
                        max(h.get("rd4_sig_str", 1.0), 1.0)
                        for h in _r45_fights
                        if h.get("rd4_sig_str", 0.0) > 0
                        and h.get("rd5_sig_str", 0.0) > 0
                    ]
                    self.df.at[idx, f"{corner}_pre_r5_r4_surge"] = (
                        float(np.mean(_r5r4_vals)) if _r5r4_vals else 0.0
                    )
                else:
                    self.df.at[idx, f"{corner}_pre_r45_slpm"] = 0.0
                    self.df.at[idx, f"{corner}_pre_r45_vs_r1_ratio"] = 0.0
                    self.df.at[idx, f"{corner}_pre_r45_td_rate"] = 0.0
                    self.df.at[idx, f"{corner}_pre_r45_td_acc"] = 0.0
                    self.df.at[idx, f"{corner}_pre_r45_leg_rate"] = 0.0
                    self.df.at[idx, f"{corner}_pre_r45_vs_r1_leg_ratio"] = 0.0
                    self.df.at[idx, f"{corner}_pre_all_round_cardio"] = 0.0
                    self.df.at[idx, f"{corner}_pre_r5_r4_surge"] = 0.0

                # ── Body escalation: avg body shots R3 − R1 (positive = escalates body attack) ──
                _body_pairs = [
                    (h.get("rd1_body", 0.0), h.get("rd3_body", 0.0))
                    for h in _h_rd
                    if h.get("rd3_sig_str", 0.0) > 0
                ]
                if _body_pairs:
                    self.df.at[idx, f"{corner}_pre_body_escalation"] = float(
                        np.mean([p[1] - p[0] for p in _body_pairs])
                    )
                else:
                    self.df.at[idx, f"{corner}_pre_body_escalation"] = 0.0

                # ── TD slope: avg TDs R3 − R1 (positive = wrestler gasses in) ──
                _td_pairs = [
                    (h.get("rd1_td", 0.0), h.get("rd3_td", 0.0))
                    for h in _h_rd
                    if h.get("rd3_sig_str", 0.0) > 0
                ]
                if _td_pairs:
                    self.df.at[idx, f"{corner}_pre_td_slope"] = float(
                        np.mean([p[1] - p[0] for p in _td_pairs])
                    )
                else:
                    self.df.at[idx, f"{corner}_pre_td_slope"] = 0.0

                # ── Late KD rate: avg R3 knockdowns ──
                _late_kd_vals = [
                    h.get("rd3_kd", 0.0) for h in _h_rd if h.get("rd3_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_late_kd_rate"] = (
                    float(np.mean(_late_kd_vals)) if _late_kd_vals else 0.0
                )

                # ── KD round average: weighted avg round in which KDs land ──
                _kd_weighted_rounds = []
                for h in _h_rd:
                    _total_kd = (
                        h.get("rd1_kd", 0.0)
                        + h.get("rd2_kd", 0.0)
                        + h.get("rd3_kd", 0.0)
                    )
                    if _total_kd > 0:
                        _kd_avg_r = (
                            1 * h.get("rd1_kd", 0.0)
                            + 2 * h.get("rd2_kd", 0.0)
                            + 3 * h.get("rd3_kd", 0.0)
                        ) / _total_kd
                        _kd_weighted_rounds.append(_kd_avg_r)
                self.df.at[idx, f"{corner}_pre_kd_round_avg"] = (
                    float(np.mean(_kd_weighted_rounds)
                          ) if _kd_weighted_rounds else 2.0
                )

                # ── Absorption slope: strikes taken R3 − R1 (positive = tires defensively) ──
                _abs_pairs = [
                    (h.get("rd1_absorbed", 0.0), h.get("rd3_absorbed", 0.0))
                    for h in _h_rd
                    if h.get("rd3_absorbed", 0.0) > 0
                ]
                if _abs_pairs:
                    self.df.at[idx, f"{corner}_pre_absorption_slope"] = float(
                        np.mean([p[1] - p[0] for p in _abs_pairs])
                    )
                    self.df.at[idx, f"{corner}_pre_late_absorbed"] = float(
                        np.mean([p[1] / 5.0 for p in _abs_pairs])
                    )
                else:
                    self.df.at[idx, f"{corner}_pre_absorption_slope"] = 0.0
                    self.df.at[idx, f"{corner}_pre_late_absorbed"] = 0.0

                # ── Ground escalation: ground strikes R3 − R1 ──
                _grd_pairs = [
                    (h.get("rd1_ground", 0.0), h.get("rd3_ground", 0.0))
                    for h in _h_rd
                    if h.get("rd3_sig_str", 0.0) > 0
                ]
                if _grd_pairs:
                    self.df.at[idx, f"{corner}_pre_ground_escalation"] = float(
                        np.mean([p[1] - p[0] for p in _grd_pairs])
                    )
                else:
                    self.df.at[idx, f"{corner}_pre_ground_escalation"] = 0.0

                # ── Late sub rate: avg R3 submission attempts ──
                _late_sub_vals = [
                    h.get("rd3_sub_att", 0.0)
                    for h in _h_rd
                    if h.get("rd3_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_late_sub_rate"] = (
                    float(np.mean(_late_sub_vals)) if _late_sub_vals else 0.0
                )

                # ── Control-time slope: ctrl R3 − R1 (positive = grappler gets more dominant) ──
                _ctrl_pairs = [
                    (h.get("rd1_ctrl", 0.0), h.get("rd3_ctrl", 0.0))
                    for h in _h_rd
                    if h.get("rd3_sig_str", 0.0) > 0
                ]
                if _ctrl_pairs:
                    self.df.at[idx, f"{corner}_pre_ctrl_slope"] = float(
                        np.mean([p[1] - p[0] for p in _ctrl_pairs])
                    )
                else:
                    self.df.at[idx, f"{corner}_pre_ctrl_slope"] = 0.0

                # ── Leg kick features ──────────────────────────────────────────────
                _r1_leg_vals = [h.get("rd1_leg", 0.0) for h in _h_rd]
                self.df.at[idx, f"{corner}_pre_r1_leg_rate"] = (
                    float(np.mean(_r1_leg_vals)) / 5.0 if _r1_leg_vals else 0.0
                )
                _leg_esc_pairs = [
                    (h.get("rd1_leg", 0.0), h.get("rd3_leg", 0.0))
                    for h in _h_rd
                    if h.get("rd3_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_leg_escalation"] = (
                    float(np.mean([p[1] - p[0] for p in _leg_esc_pairs]))
                    if _leg_esc_pairs
                    else 0.0
                )
                _leg_head_vals = [
                    h.get("rd1_leg", 0.0) / max(h.get("rd1_head", 1.0), 1.0)
                    for h in _h_rd
                    if h.get("rd1_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_leg_vs_head"] = (
                    float(np.mean(_leg_head_vals)) if _leg_head_vals else 0.0
                )
                _leg_pct_vals = [
                    h.get("rd1_leg", 0.0) / max(h.get("rd1_sig_str", 1.0), 1.0)
                    for h in _h_rd
                    if h.get("rd1_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_leg_pct"] = (
                    float(np.mean(_leg_pct_vals)) if _leg_pct_vals else 0.0
                )

                # ── Clinch features ────────────────────────────────────────────────
                _r1_clinch_vals = [h.get("rd1_clinch", 0.0) for h in _h_rd]
                self.df.at[idx, f"{corner}_pre_r1_clinch_rate"] = (
                    float(np.mean(_r1_clinch_vals)) /
                    5.0 if _r1_clinch_vals else 0.0
                )
                _clinch_esc_pairs = [
                    (h.get("rd1_clinch", 0.0), h.get("rd3_clinch", 0.0))
                    for h in _h_rd
                    if h.get("rd3_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_clinch_escalation"] = (
                    float(np.mean([p[1] - p[0] for p in _clinch_esc_pairs]))
                    if _clinch_esc_pairs
                    else 0.0
                )
                _clinch_pct_vals = [
                    h.get("rd1_clinch", 0.0) /
                    max(h.get("rd1_sig_str", 1.0), 1.0)
                    for h in _h_rd
                    if h.get("rd1_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_clinch_pct_r1"] = (
                    float(np.mean(_clinch_pct_vals)
                          ) if _clinch_pct_vals else 0.0
                )

                # ── Round 2 adjustment signals ─────────────────────────────────────
                _r2_fights = [h for h in _h_rd if h.get(
                    "rd2_sig_str", 0.0) > 0]
                if _r2_fights:
                    self.df.at[idx, f"{corner}_pre_r1_to_r2_output"] = float(
                        np.mean(
                            [
                                h.get("rd2_sig_str", 0.0) -
                                h.get("rd1_sig_str", 0.0)
                                for h in _r2_fights
                            ]
                        )
                    )
                    self.df.at[idx, f"{corner}_pre_r1_to_r2_td"] = float(
                        np.mean(
                            [
                                h.get("rd2_td", 0.0) - h.get("rd1_td", 0.0)
                                for h in _r2_fights
                            ]
                        )
                    )
                    self.df.at[idx, f"{corner}_pre_r2_body_rate"] = float(
                        np.mean([h.get("rd2_body", 0.0) for h in _r2_fights])
                    )
                else:
                    self.df.at[idx, f"{corner}_pre_r1_to_r2_output"] = 0.0
                    self.df.at[idx, f"{corner}_pre_r1_to_r2_td"] = 0.0
                    self.df.at[idx, f"{corner}_pre_r2_body_rate"] = 0.0

                # ── Volume vs power ratio ──────────────────────────────────────────
                _vol_r1_vals = [
                    h.get("rd1_str", 0.0) / max(h.get("rd1_sig_str", 1.0), 1.0)
                    for h in _h_rd
                    if h.get("rd1_sig_str", 0.0) > 0
                ]
                _r1_vol_ratio = float(
                    np.mean(_vol_r1_vals)) if _vol_r1_vals else 1.0
                self.df.at[idx, f"{corner}_pre_volume_ratio"] = _r1_vol_ratio
                _vol_r3_vals = [
                    h.get("rd3_str", 0.0) / max(h.get("rd3_sig_str", 1.0), 1.0)
                    for h in _h_rd
                    if h.get("rd3_sig_str", 0.0) > 0
                ]
                _r3_vol_ratio = float(
                    np.mean(_vol_r3_vals)) if _vol_r3_vals else 1.0
                self.df.at[idx, f"{corner}_pre_vol_ratio_evo"] = (
                    _r3_vol_ratio - _r1_vol_ratio
                )

                # ── Zone shift: does fight move away from distance in later rounds? ─
                _zone_pairs = [
                    (
                        h.get("rd1_dist", 0.0) /
                        max(h.get("rd1_sig_str", 1.0), 1.0),
                        h.get("rd3_dist", 0.0) /
                        max(h.get("rd3_sig_str", 1.0), 1.0),
                    )
                    for h in _h_rd
                    if h.get("rd3_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_zone_shift"] = (
                    float(np.mean([p[1] - p[0] for p in _zone_pairs]))
                    if _zone_pairs
                    else 0.0
                )

                # ── TD accuracy evolution ──────────────────────────────────────────
                _r1_tdacc_vals = [
                    h.get("rd1_td", 0.0) / max(h.get("rd1_td_att", 1.0), 1.0)
                    for h in _h_rd
                    if h.get("rd1_td_att", 0.0) > 0
                ]
                _r1_td_acc = float(np.mean(_r1_tdacc_vals)
                                   ) if _r1_tdacc_vals else 0.0
                self.df.at[idx, f"{corner}_pre_r1_td_acc"] = _r1_td_acc
                _late_tdacc_vals = [
                    h.get("rd3_td", 0.0) / max(h.get("rd3_td_att", 1.0), 1.0)
                    for h in _h_rd
                    if h.get("rd3_td_att", 0.0) > 0
                ]
                _late_td_acc = (
                    float(np.mean(_late_tdacc_vals)
                          ) if _late_tdacc_vals else 0.0
                )
                self.df.at[idx, f"{corner}_pre_late_td_acc"] = _late_td_acc
                self.df.at[idx,
                           f"{corner}_pre_td_acc_evo"] = _late_td_acc - _r1_td_acc

                # ── Championship round extras (R4/R5) ─────────────────────────────
                if _r45_fights:  # reuse list computed above
                    self.df.at[idx, f"{corner}_pre_r45_kd_rate"] = float(
                        np.mean(
                            [
                                h.get("rd4_kd", 0.0) + h.get("rd5_kd", 0.0)
                                for h in _r45_fights
                            ]
                        )
                    )
                    self.df.at[idx, f"{corner}_pre_r45_body_rate"] = float(
                        np.mean(
                            [
                                (h.get("rd4_body", 0.0) +
                                 h.get("rd5_body", 0.0)) / 2.0
                                for h in _r45_fights
                            ]
                        )
                    )
                    self.df.at[idx, f"{corner}_pre_r45_clinch_rate"] = float(
                        np.mean(
                            [
                                (h.get("rd4_clinch", 0.0) +
                                 h.get("rd5_clinch", 0.0))
                                / 2.0
                                for h in _r45_fights
                            ]
                        )
                    )
                    self.df.at[idx, f"{corner}_pre_r45_ctrl"] = float(
                        np.mean(
                            [
                                (h.get("rd4_ctrl", 0.0) +
                                 h.get("rd5_ctrl", 0.0)) / 2.0
                                for h in _r45_fights
                            ]
                        )
                    )
                else:
                    self.df.at[idx, f"{corner}_pre_r45_kd_rate"] = 0.0
                    self.df.at[idx, f"{corner}_pre_r45_body_rate"] = 0.0
                    self.df.at[idx, f"{corner}_pre_r45_clinch_rate"] = 0.0
                    self.df.at[idx, f"{corner}_pre_r45_ctrl"] = 0.0

                # ── Reversal rate ──────────────────────────────────────────────────
                _rev_vals = [h.get("rev", 0.0) for h in _h_rd]
                self.df.at[idx, f"{corner}_pre_reversal_rate"] = (
                    float(np.mean(_rev_vals)) if _rev_vals else 0.0
                )

                # ── Per-round reversal escalation (scramble timing signal) ────────
                # R1 reversal rate: how often does the fighter reverse in the first round?
                _r1_rev_vals = [
                    h.get("rd1_rev", 0.0)
                    for h in _h_rd
                    if h.get("rd1_sig_str", 0.0) > 0
                ]
                _r1_rev_avg = float(np.mean(_r1_rev_vals)
                                    ) if _r1_rev_vals else 0.0
                self.df.at[idx, f"{corner}_pre_r1_rev_rate"] = _r1_rev_avg
                # Late reversal escalation: R3 revs − R1 revs (positive = more scrambles late)
                _r3_rev_vals = [
                    h.get("rd3_rev", 0.0)
                    for h in _h_rd
                    if h.get("rd3_sig_str", 0.0) > 0
                ]
                _r3_rev_avg = float(np.mean(_r3_rev_vals)
                                    ) if _r3_rev_vals else 0.0
                self.df.at[idx, f"{corner}_pre_late_rev_escalation"] = (
                    _r3_rev_avg - _r1_rev_avg
                )

                # ── Per-round damage margins ───────────────────────────────────────
                _r1_margins = [
                    h.get("rd1_sig_str", 0.0) - h.get("rd1_absorbed", 0.0)
                    for h in _h_rd
                    if (h.get("rd1_sig_str", 0.0) + h.get("rd1_absorbed", 0.0)) > 0
                ]
                _r2_margins = [
                    h.get("rd2_sig_str", 0.0) - h.get("rd2_absorbed", 0.0)
                    for h in _h_rd
                    if h.get("rd2_sig_str", 0.0) > 0
                ]
                _r3_margins = [
                    h.get("rd3_sig_str", 0.0) - h.get("rd3_absorbed", 0.0)
                    for h in _h_rd
                    if h.get("rd3_sig_str", 0.0) > 0
                ]
                _r1_mrg = float(np.mean(_r1_margins)) if _r1_margins else 0.0
                _r3_mrg = float(np.mean(_r3_margins)) if _r3_margins else 0.0
                self.df.at[idx, f"{corner}_pre_r1_damage_margin"] = _r1_mrg
                self.df.at[idx, f"{corner}_pre_r2_damage_margin"] = (
                    float(np.mean(_r2_margins)) if _r2_margins else 0.0
                )
                self.df.at[idx, f"{corner}_pre_r3_damage_margin"] = _r3_mrg
                self.df.at[idx, f"{corner}_pre_damage_margin_change"] = (
                    _r3_mrg - _r1_mrg
                )

                # ── R2 conditional adjustment scores ──────────────────────────────
                # When R1 margin < 0 (losing R1): does the fighter increase R2 output?
                # Positive = adjusts and comes back; negative = compounds the deficit.
                _r2_adj_loss = [
                    h.get("rd2_sig_str", 0.0) - h.get("rd1_sig_str", 0.0)
                    for h in _h_rd
                    if h.get("rd1_sig_str", 0.0) < h.get("rd1_absorbed", 0.0)
                    and h.get("rd2_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_r2_adj_score_loss"] = (
                    float(np.mean(_r2_adj_loss)) if _r2_adj_loss else 0.0
                )
                # When R1 margin > 0 (winning R1): does the fighter press or coast in R2?
                _r2_adj_win = [
                    h.get("rd2_sig_str", 0.0) - h.get("rd1_sig_str", 0.0)
                    for h in _h_rd
                    if h.get("rd1_sig_str", 0.0) > h.get("rd1_absorbed", 0.0)
                    and h.get("rd2_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_r2_adj_score_win"] = (
                    float(np.mean(_r2_adj_win)) if _r2_adj_win else 0.0
                )

                # ── Head accuracy evolution ────────────────────────────────────────
                _r1_hacc = [
                    h.get("rd1_head", 0.0) /
                    max(h.get("rd1_head_att", 1.0), 1.0)
                    for h in _h_rd
                    if h.get("rd1_head_att", 0.0) > 0
                ]
                _r3_hacc = [
                    h.get("rd3_head", 0.0) /
                    max(h.get("rd3_head_att", 1.0), 1.0)
                    for h in _h_rd
                    if h.get("rd3_head_att", 0.0) > 0
                ]
                _r1_ha = float(np.mean(_r1_hacc)) if _r1_hacc else 0.45
                _r3_ha = float(np.mean(_r3_hacc)) if _r3_hacc else 0.45
                self.df.at[idx, f"{corner}_pre_r1_head_acc"] = _r1_ha
                self.df.at[idx, f"{corner}_pre_r3_head_acc"] = _r3_ha
                self.df.at[idx,
                           f"{corner}_pre_head_acc_fade"] = _r1_ha - _r3_ha
                # Head-to-leg switch: positive = more leg pct in R3 than R1
                _htl_pairs = [
                    (
                        h.get("rd1_leg", 0.0) /
                        max(h.get("rd1_sig_str", 1.0), 1.0),
                        h.get("rd3_leg", 0.0) /
                        max(h.get("rd3_sig_str", 1.0), 1.0),
                    )
                    for h in _h_rd
                    if h.get("rd3_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_head_to_leg_late"] = (
                    float(np.mean([p[1] - p[0] for p in _htl_pairs]))
                    if _htl_pairs
                    else 0.0
                )

                # ── Grappling chain composites ─────────────────────────────────────
                # chain: avg td × avg ctrl growth × avg late sub attempts
                _td_avg_h = (
                    float(np.mean([h.get("td", 0.0)
                          for h in _h_rd])) if _h_rd else 0.0
                )
                _ctrl_pos = max(
                    self.df.at[idx, f"{corner}_pre_ctrl_slope"], 0.0)
                _late_sub = self.df.at[idx, f"{corner}_pre_late_sub_rate"]
                self.df.at[idx, f"{corner}_pre_grapple_chain"] = (
                    _td_avg_h * (1.0 + _ctrl_pos) * (1.0 + _late_sub)
                )
                # ground-and-pound: td frequency × ground strike escalation
                _grd_esc = self.df.at[idx, f"{corner}_pre_ground_escalation"]
                self.df.at[idx, f"{corner}_pre_gnp_score"] = _td_avg_h * max(
                    _grd_esc, 0.0
                )
                # sub efficiency: late sub attempts per average TD
                self.df.at[idx, f"{corner}_pre_sub_efficiency"] = _late_sub / max(
                    _td_avg_h, 0.01
                )

                # Full 4-stage grappling depth chain (multiplicative).
                # TD → control → GnP → submission: captures truly complete grapplers.
                _gnp_s = self.df.at[idx, f"{corner}_pre_gnp_score"]
                _sub_e = self.df.at[idx, f"{corner}_pre_sub_efficiency"]
                self.df.at[idx, f"{corner}_pre_grapple_depth_4"] = (
                    max(_td_avg_h, 0.0)
                    * (1.0 + max(_ctrl_pos, 0.0))
                    * max(_gnp_s, 0.01)
                    * (1.0 + max(_sub_e, 0.0))
                )

                # Endurance composite (computed here after all building blocks are ready):
                # cardio_index × r3_retention × accuracy_retention
                _cardio_idx = self.df.at[idx, f"{corner}_pre_cardio_index"]
                self.df.at[idx, f"{corner}_pre_endurance_composite"] = (
                    _cardio_idx * _r3_ret * _acc_ret
                )

                # ── Round conditional momentum sequences ──────────────────────────
                # P(win Rn+1 | win Rn) and P(win Rn+1 | lose Rn) — sequential
                # round patterns capture sustained pressure vs. reset/comeback
                # within fights. Distinct from r1_win_rate/r2_win_rate which are
                # unconditional.
                _r12_win_n, _r12_win_d = 0, 0  # P(win R2 | win R1)
                _r12_loss_n, _r12_loss_d = 0, 0  # P(win R2 | lose R1)
                _r23_win_n, _r23_win_d = 0, 0  # P(win R3 | win R2)
                _r23_loss_n, _r23_loss_d = 0, 0  # P(win R3 | lose R2)
                for _h in _h_rd:
                    _h_r1ss = _h.get("rd1_sig_str", 0.0)
                    _h_r2ss = _h.get("rd2_sig_str", 0.0)
                    _h_r3ss = _h.get("rd3_sig_str", 0.0)
                    _h_r1ab = _h.get("rd1_absorbed", 0.0)
                    _h_r2ab = _h.get("rd2_absorbed", 0.0)
                    _h_r3ab = _h.get("rd3_absorbed", 0.0)
                    if _h_r1ss > 0 and _h_r2ss > 0:
                        _r1m = _h_r1ss - _h_r1ab
                        _r2m = _h_r2ss - _h_r2ab
                        if _r1m > 0:
                            _r12_win_d += 1
                            if _r2m > 0:
                                _r12_win_n += 1
                        else:
                            _r12_loss_d += 1
                            if _r2m > 0:
                                _r12_loss_n += 1
                    if _h_r2ss > 0 and _h_r3ss > 0:
                        _r2m = _h_r2ss - _h_r2ab
                        _r3m = _h_r3ss - _h_r3ab
                        if _r2m > 0:
                            _r23_win_d += 1
                            if _r3m > 0:
                                _r23_win_n += 1
                        else:
                            _r23_loss_d += 1
                            if _r3m > 0:
                                _r23_loss_n += 1
                self.df.at[idx, f"{corner}_pre_momentum_r12_win"] = _r12_win_n / max(
                    _r12_win_d, 1
                )
                self.df.at[idx, f"{corner}_pre_momentum_r12_loss"] = _r12_loss_n / max(
                    _r12_loss_d, 1
                )
                self.df.at[idx, f"{corner}_pre_momentum_r23_win"] = _r23_win_n / max(
                    _r23_win_d, 1
                )
                self.df.at[idx, f"{corner}_pre_momentum_r23_loss"] = _r23_loss_n / max(
                    _r23_loss_d, 1
                )

                # ── Quality-adjusted round output ─────────────────────────────────
                # sig_str² / att = volume × accuracy product. Captures the joint
                # collapse of both dimensions under fatigue — not proxied by either
                # accuracy_retention or cardio_index alone.
                _r1_qual_vals = [
                    _h.get("rd1_sig_str", 0.0)
                    * (
                        _h.get("rd1_sig_str", 0.0)
                        / max(_h.get("rd1_sig_str_att", 1.0), 1.0)
                    )
                    for _h in _h_rd
                    if _h.get("rd1_sig_str_att", 0.0) > 0
                ]
                _r3_qual_vals = [
                    _h.get("rd3_sig_str", 0.0)
                    * (
                        _h.get("rd3_sig_str", 0.0)
                        / max(_h.get("rd3_sig_str_att", 1.0), 1.0)
                    )
                    for _h in _h_rd
                    if _h.get("rd3_sig_str_att", 0.0) > 0
                ]
                _r1_qual = float(np.mean(_r1_qual_vals)
                                 ) if _r1_qual_vals else 0.0
                _r3_qual = float(np.mean(_r3_qual_vals)
                                 ) if _r3_qual_vals else 0.0
                self.df.at[idx, f"{corner}_pre_r1_quality_output"] = _r1_qual
                self.df.at[idx, f"{corner}_pre_r3_quality_output"] = _r3_qual
                self.df.at[idx, f"{corner}_pre_quality_output_ratio"] = _r3_qual / max(
                    _r1_qual, 1.0
                )

                # ── Output under punishment ───────────────────────────────────────
                # In rounds where the fighter absorbed more than their rolling
                # average, how much did they still output? True heart/toughness
                # metric — no existing proxy captures fighting back while hurt.
                _all_abs_vals = []
                for _h in _h_rd:
                    for _rk in ["rd1_absorbed", "rd2_absorbed", "rd3_absorbed"]:
                        _av = _h.get(_rk, 0.0)
                        if _av > 0:
                            _all_abs_vals.append(_av)
                _mean_abs_p = float(np.mean(_all_abs_vals)
                                    ) if _all_abs_vals else 0.0
                _punish_outputs = []
                if _mean_abs_p > 0:
                    for _h in _h_rd:
                        for _rk, _sk in [
                            ("rd1_absorbed", "rd1_sig_str"),
                            ("rd2_absorbed", "rd2_sig_str"),
                            ("rd3_absorbed", "rd3_sig_str"),
                        ]:
                            if _h.get(_rk, 0.0) > _mean_abs_p:
                                _punish_outputs.append(_h.get(_sk, 0.0))
                self.df.at[idx, f"{corner}_pre_output_vs_punishment"] = (
                    float(np.mean(_punish_outputs)) if _punish_outputs else 0.0
                )

                # ── Sub escalation ratio ──────────────────────────────────────────
                # R3 sub attempts relative to R1 — grapplers who ratchet up
                # submission pressure as opponent tires. Distinct from late_sub_rate
                # (absolute R3 count) which doesn't capture escalation from R1.
                _sub_esc_vals = [
                    _h.get("rd3_sub_att", 0.0)
                    / max(_h.get("rd1_sub_att", 0.0) + 0.5, 0.5)
                    for _h in _h_rd
                    if _h.get("rd3_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_sub_escalation_ratio"] = (
                    float(np.mean(_sub_esc_vals)) if _sub_esc_vals else 1.0
                )

                # ── Burst index ───────────────────────────────────────────────────
                # Fraction of total R1-R3 output landed in R1. High = front-loader
                # who dumps volume early; low = patient/late builder.
                # Complements r1_front_load (R1 vs career avg slpm) but measures
                # output share within the specific fight rather than vs career norm.
                _burst_vals = [
                    _h.get("rd1_sig_str", 0.0)
                    / max(
                        _h.get("rd1_sig_str", 0.0)
                        + _h.get("rd2_sig_str", 0.0)
                        + _h.get("rd3_sig_str", 0.0),
                        1.0,
                    )
                    for _h in _h_rd
                    if _h.get("rd3_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_burst_index"] = (
                    float(np.mean(_burst_vals)) if _burst_vals else 1.0 / 3.0
                )

                # ── Ctrl GnP efficiency slope ─────────────────────────────────────
                # (R3 ground / R3 ctrl) − (R1 ground / R1 ctrl): does GnP output
                # per ctrl minute improve as the fight goes? Elite grapplers
                # tighten their GnP as opponent tires; ctrl_gnp_rate is career avg.
                _gnp_slope_vals = []
                for _h in _h_rd:
                    _r1g = _h.get("rd1_ground", 0.0)
                    _r1c = _h.get("rd1_ctrl", 0.0)
                    _r3g = _h.get("rd3_ground", 0.0)
                    _r3c = _h.get("rd3_ctrl", 0.0)
                    if _r1c > 0 or _r3c > 0:
                        _gnp_slope_vals.append(
                            (_r3g / max(_r3c, 1.0)) - (_r1g / max(_r1c, 1.0))
                        )
                self.df.at[idx, f"{corner}_pre_ctrl_gnp_slope"] = (
                    float(np.mean(_gnp_slope_vals)) if _gnp_slope_vals else 0.0
                )

                # ── R3 striking under takedown threat ─────────────────────────────
                # R3 sig_str / max(R1 td_att, 1): does the striker maintain output
                # in later rounds even when the opponent was shooting heavily in R1?
                # Measures cross-skill pressure resistance (striker vs wrestler).
                _str_td_vals = [
                    _h.get("rd3_sig_str", 0.0) /
                    max(_h.get("rd1_td_att", 1.0), 1.0)
                    for _h in _h_rd
                    if _h.get("rd3_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_r3_str_vs_td_pressure"] = (
                    float(np.mean(_str_td_vals)) if _str_td_vals else 0.0
                )

                # ── Tier 13: Exponentially Decayed Average Features ───────────────
                # alpha=0.8: each fight is weighted 80% of the next more-recent one.
                # "dapa" features also multiply by (opp_elo / 1500) to reward
                # achieving high stats against higher-quality opposition.
                def _da_fn(vals, alpha=0.8, default=0.0):
                    if not vals:
                        return default
                    n = len(vals)
                    weights = [alpha ** (n - 1 - i) for i in range(n)]
                    tw = sum(weights)
                    return sum(v * w for v, w in zip(vals, weights)) / tw

                def _dapa_fn(vals, opp_elos, alpha=0.8, base_elo=1500.0, default=0.0):
                    if not vals or not opp_elos:
                        return default
                    n = len(vals)
                    weights = [alpha ** (n - 1 - i) for i in range(n)]
                    adj = [v * (oe / base_elo)
                           for v, oe in zip(vals, opp_elos)]
                    tw = sum(weights)
                    return sum(a * w for a, w in zip(adj, weights)) / tw

                _h_da = _h10_v  # all available history for this fighter

                def _da_key(key, default=0.0, req_key=None):
                    vals = [
                        h.get(key, default)
                        for h in _h_da
                        if (req_key is None or h.get(req_key, 0) > 0)
                    ]
                    return _da_fn(vals, default=default) if vals else default

                def _dapa_key(key, default=0.0, req_key=None):
                    data = [
                        (h.get(key, default), h.get("opp_elo_pre", 1500.0))
                        for h in _h_da
                        if (req_key is None or h.get(req_key, 0) > 0)
                    ]
                    if not data:
                        return default
                    vals, elos = zip(*data)
                    return _dapa_fn(list(vals), list(elos), default=default)

                def _ra_fn_calendar(
                    vals, dates, ref_date, half_life=365.0, default=0.0
                ):
                    if not vals:
                        return default
                    try:
                        ref = pd.Timestamp(ref_date)
                        if pd.isna(ref):
                            return default
                    except Exception:
                        return default
                    tw, acc = 0.0, 0.0
                    lam = math.log(2) / max(half_life, 1.0)
                    for v, d in zip(vals, dates):
                        try:
                            td = pd.Timestamp(d)
                            if pd.isna(td):
                                continue
                        except Exception:
                            continue
                        days = max((ref - td).days, 0)
                        w = math.exp(-lam * days)
                        acc += float(v) * w
                        tw += w
                    return acc / tw if tw > 0 else default

                def _ra_key_finish(key, default=0.0):
                    ref_date = row.get("event_date", None)
                    vals = [h.get(key, default) for h in _h_da]
                    dates = [h.get("event_date") for h in _h_da]
                    return _ra_fn_calendar(vals, dates, ref_date, default=default)

                # ── Decayed averages (da) ─────────────────────────────────────────
                self.df.at[idx, f"{corner}_pre_da_sig_str_acc"] = _da_key(
                    "sig_str_acc_v", default=0.5, req_key="sig_str_att"
                )
                self.df.at[idx, f"{corner}_pre_da_td_acc"] = _da_key(
                    "td_acc_v", default=0.4
                )
                self.df.at[idx, f"{corner}_pre_da_head_landed"] = _da_key(
                    "total_head_landed", default=0.0
                )
                self.df.at[idx, f"{corner}_pre_da_head_acc"] = _da_key(
                    "head_acc_v", default=0.45, req_key="total_head_att"
                )
                self.df.at[idx, f"{corner}_pre_da_body_acc"] = _da_key(
                    "body_acc_v", default=0.4, req_key="total_body_att"
                )
                self.df.at[idx, f"{corner}_pre_da_distance_acc"] = _da_key(
                    "dist_acc_v", default=0.35, req_key="total_dist_att"
                )
                self.df.at[idx, f"{corner}_pre_da_head_defense"] = _da_key(
                    "head_def_v", default=0.55
                )
                self.df.at[idx, f"{corner}_pre_da_body_defense"] = _da_key(
                    "body_def_v", default=0.55
                )
                self.df.at[idx, f"{corner}_pre_da_distance_defense"] = _da_key(
                    "dist_def_v", default=0.55
                )
                self.df.at[idx, f"{corner}_pre_da_ground_defense"] = _da_key(
                    "grd_def_v", default=0.55
                )
                self.df.at[idx, f"{corner}_pre_da_td_defense"] = _da_key(
                    "td_def_v", default=0.6
                )
                self.df.at[idx, f"{corner}_pre_da_sub_att"] = _da_key(
                    "sub_att", default=0.0
                )
                self.df.at[idx, f"{corner}_pre_da_kd"] = _da_key(
                    "kd", default=0.0)
                self.df.at[idx, f"{corner}_pre_da_ko"] = _da_key(
                    "ko", default=0.0)
                self.df.at[idx, f"{corner}_pre_da_win_ratio"] = _da_key(
                    "won", default=0.5
                )
                self.df.at[idx, f"{corner}_pre_da_ctrl_r1"] = _da_key(
                    "rd1_ctrl", default=0.0
                )
                self.df.at[idx, f"{corner}_pre_da_clinch_pm"] = _da_key(
                    "clinch_pm_v", default=0.0
                )
                self.df.at[idx, f"{corner}_pre_da_opp_leg_pm"] = _da_key(
                    "opp_leg_pm_v", default=0.0
                )
                self.df.at[idx, f"{corner}_pre_da_opp_ctrl_r1_pm"] = _da_key(
                    "opp_ctrl_r1_pm_v", default=0.0
                )
                self.df.at[idx, f"{corner}_pre_da_opp_sub_pm"] = _da_key(
                    "opp_sub_pm_v", default=0.0
                )
                self.df.at[idx, f"{corner}_pre_da_opp_rev_r1"] = _da_key(
                    "opp_rd1_rev_v", default=0.0
                )
                self.df.at[idx, f"{corner}_pre_da_r1_strikes"] = _da_key(
                    "rd1_sig_str", default=0.0
                )
                self.df.at[idx, f"{corner}_pre_da_reversals"] = _da_key(
                    "rev", default=0.0
                )
                self.df.at[idx, f"{corner}_pre_da_dist_landing_ratio"] = _da_key(
                    "dist_landing_ratio_v", default=0.5
                )
                self.df.at[idx, f"{corner}_pre_da_opp_kd"] = _da_key(
                    "opp_kd_v", default=0.0
                )
                # ── Adjusted performance decayed averages (dapa) ──────────────────
                self.df.at[idx, f"{corner}_pre_dapa_sig_str_acc"] = _dapa_key(
                    "sig_str_acc_v", default=0.5, req_key="sig_str_att"
                )
                self.df.at[idx, f"{corner}_pre_dapa_head_acc"] = _dapa_key(
                    "head_acc_v", default=0.45, req_key="total_head_att"
                )
                self.df.at[idx, f"{corner}_pre_dapa_body_acc"] = _dapa_key(
                    "body_acc_v", default=0.4, req_key="total_body_att"
                )
                self.df.at[idx, f"{corner}_pre_dapa_distance_acc"] = _dapa_key(
                    "dist_acc_v", default=0.35, req_key="total_dist_att"
                )
                self.df.at[idx, f"{corner}_pre_dapa_head_defense"] = _dapa_key(
                    "head_def_v", default=0.55
                )
                self.df.at[idx, f"{corner}_pre_dapa_dist_defense"] = _dapa_key(
                    "dist_def_v", default=0.55
                )
                self.df.at[idx, f"{corner}_pre_dapa_ground_defense"] = _dapa_key(
                    "grd_def_v", default=0.55
                )
                self.df.at[idx, f"{corner}_pre_dapa_r1_strikes"] = _dapa_key(
                    "rd1_sig_str", default=0.0
                )
                self.df.at[idx, f"{corner}_pre_dapa_reversals"] = _dapa_key(
                    "rev", default=0.0
                )
                self.df.at[idx, f"{corner}_pre_dapa_dist_landing_ratio"] = _dapa_key(
                    "dist_landing_ratio_v", default=0.5
                )
                self.df.at[idx, f"{corner}_pre_dapa_head_landing_ratio"] = _dapa_key(
                    "head_landing_ratio_v", default=0.4
                )
                self.df.at[idx, f"{corner}_pre_da_age"] = _da_key(
                    "age_at_event", default=28.0
                )
                self.df.at[idx, f"{corner}_pre_da_ufc_age"] = _da_key(
                    "ufc_fight_num_v", default=1.0
                )
                self.df.at[idx, f"{corner}_pre_da_reach_ratio"] = _da_key(
                    "reach_ratio_v", default=1.0
                )
                self.df.at[idx, f"{corner}_pre_da_days_since"] = _da_key(
                    "days_since_v", default=365.0
                )
                self.df.at[idx, f"{corner}_pre_da_damage_margin"] = _da_key(
                    "damage_margin_r123", default=0.0
                )
                self.df.at[idx, f"{corner}_pre_da_head_sig_abs_share"] = _da_key(
                    "head_sig_abs_share_v", default=0.45
                )
                self.df.at[idx, f"{corner}_pre_qa_finish_win"] = _dapa_key(
                    "was_finish_win", default=0.0
                )
                self.df.at[idx, f"{corner}_pre_ra_finish_win"] = _ra_key_finish(
                    "was_finish_win", default=0.0
                )

                # ── Style entropy ──────────────────────────────────────────────────
                def _shannon(probs):
                    """Shannon entropy of a probability distribution (list of floats)."""
                    total = sum(probs)
                    if total <= 0:
                        return 0.0
                    ps = [p / total for p in probs if p > 0]
                    return -sum(p * math.log(p + 1e-12) for p in ps)

                _target_ents = []
                _zone_ents = []
                for h in _h_rd:
                    if h.get("rd1_sig_str", 0.0) > 0:
                        _target_ents.append(
                            _shannon(
                                [
                                    h.get("rd1_head", 0.0),
                                    h.get("rd1_body", 0.0),
                                    h.get("rd1_leg", 0.0),
                                ]
                            )
                        )
                        _zone_ents.append(
                            _shannon(
                                [
                                    h.get("rd1_dist", 0.0),
                                    h.get("rd1_clinch", 0.0),
                                    h.get("rd1_ground", 0.0),
                                ]
                            )
                        )
                self.df.at[idx, f"{corner}_pre_target_entropy"] = (
                    float(np.mean(_target_ents)) if _target_ents else 1.0
                )
                self.df.at[idx, f"{corner}_pre_zone_entropy"] = (
                    float(np.mean(_zone_ents)) if _zone_ents else 1.0
                )

                # ── Finish timing profile ──────────────────────────────────────────
                _finish_rounds = [
                    h.get("finish_round", 0.0)
                    for h in _h_rd
                    if h.get("finished", 0) == 1 and h.get("finish_round", 0.0) > 0
                ]
                if _finish_rounds:
                    _avg_fin_rnd = float(np.mean(_finish_rounds))
                    self.df.at[idx,
                               f"{corner}_pre_avg_finish_round"] = _avg_fin_rnd
                    _r1_fin_cnt = sum(1 for r in _finish_rounds if r <= 1)
                    _r1_fin_rate = _r1_fin_cnt / len(_finish_rounds)
                    self.df.at[idx,
                               f"{corner}_pre_r1_finish_rate"] = _r1_fin_rate
                    self.df.at[idx, f"{corner}_pre_early_finish_score"] = (
                        _r1_fin_rate / max(_avg_fin_rnd, 1.0)
                    )
                else:
                    _fr_all = (
                        self.df.at[idx, f"{corner}_pre_finish_rate"]
                        if f"{corner}_pre_finish_rate" in self.df.columns
                        else 0.0
                    )
                    self.df.at[idx, f"{corner}_pre_avg_finish_round"] = 2.5
                    self.df.at[idx, f"{corner}_pre_r1_finish_rate"] = 0.0
                    self.df.at[idx, f"{corner}_pre_early_finish_score"] = 0.0

                # ── Comeback and lead-hold rates ───────────────────────────────────
                _r1_results = [
                    (
                        h.get("rd1_sig_str", 0.0) > h.get("rd1_absorbed", 0.0),
                        h.get("won", 0),
                    )
                    for h in _h_rd
                    if (h.get("rd1_sig_str", 0.0) + h.get("rd1_absorbed", 0.0)) > 0
                ]
                _came_back = [
                    1 for (won_r1, won_f) in _r1_results if not won_r1 and won_f
                ]
                _held_lead = [1 for (won_r1, won_f)
                              in _r1_results if won_r1 and won_f]
                _lost_r1 = sum(1 for (wr1, _) in _r1_results if not wr1)
                _won_r1 = sum(1 for (wr1, _) in _r1_results if wr1)
                self.df.at[idx, f"{corner}_pre_comeback_rate"] = len(_came_back) / max(
                    _lost_r1, 1
                )
                self.df.at[idx, f"{corner}_pre_lead_hold_rate"] = len(_held_lead) / max(
                    _won_r1, 1
                )

                # ── Power and chin signals ─────────────────────────────────────────
                _total_ss_h = sum(h.get("sig_str", 0.0) for h in _h_rd)
                _total_kd_h = sum(h.get("kd", 0.0) for h in _h_rd)
                self.df.at[idx, f"{corner}_pre_kd_efficiency"] = (
                    _total_kd_h / max(_total_ss_h, 1.0)
                ) * 100.0
                # avg R1 head strikes the opponent threw against this fighter (chin pressure)
                _r1_hpres = [h.get("opp_rd1_head", 0.0) for h in _h_rd]
                self.df.at[idx, f"{corner}_pre_r1_head_pressure"] = (
                    float(np.mean(_r1_hpres)) if _r1_hpres else 0.0
                )

                # ── Style composite indices ────────────────────────────────────────
                _zone_sh = self.df.at[idx, f"{corner}_pre_zone_shift"]
                _clinch_e = self.df.at[idx, f"{corner}_pre_clinch_escalation"]
                _td_sl = self.df.at[idx, f"{corner}_pre_td_slope"]
                self.df.at[idx, f"{corner}_pre_pressure_index"] = (
                    max(-_zone_sh, 0.0) + max(_clinch_e, 0.0) + max(_td_sl, 0.0)
                )
                _r1_wr = self.df.at[idx, f"{corner}_pre_r1_win_rate"]
                _out_sl = self.df.at[idx, f"{corner}_pre_output_slope"]
                self.df.at[idx, f"{corner}_pre_counter_index"] = _r1_wr * max(
                    -_out_sl, 0.0
                )

                # ── Round momentum: % of fights winning R1/R3 by output margin ─────
                _r1_wins = [
                    1 if h.get("rd1_sig_str", 0.0) > h.get(
                        "rd1_absorbed", 0.0) else 0
                    for h in _h_rd
                    if (h.get("rd1_sig_str", 0.0) + h.get("rd1_absorbed", 0.0)) > 0
                ]
                self.df.at[idx, f"{corner}_pre_r1_win_rate"] = (
                    float(np.mean(_r1_wins)) if _r1_wins else 0.5
                )
                _late_wins = [
                    1 if h.get("rd3_sig_str", 0.0) > h.get(
                        "rd3_absorbed", 0.0) else 0
                    for h in _h_rd
                    if h.get("rd3_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_late_win_rate"] = (
                    float(np.mean(_late_wins)) if _late_wins else 0.5
                )

                # ── R2→R3 closing momentum ────────────────────────────────────────
                _r2r3_pairs = [
                    (h.get("rd2_sig_str", 0.0), h.get("rd3_sig_str", 0.0))
                    for h in _h_rd
                    if h.get("rd3_sig_str", 0.0) > 0 and h.get("rd2_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_r2_to_r3_momentum"] = (
                    float(np.mean([p[1] - p[0] for p in _r2r3_pairs]))
                    if _r2r3_pairs
                    else 0.0
                )

                # ── Output consistency (coefficient of variation) ─────────────────
                _cv_vals = []
                for h in _h_rd:
                    _rounds_out = [
                        v
                        for v in [
                            h.get("rd1_sig_str", 0.0),
                            h.get("rd2_sig_str", 0.0),
                            h.get("rd3_sig_str", 0.0),
                        ]
                        if v > 0
                    ]
                    if len(_rounds_out) >= 2:
                        _m = float(np.mean(_rounds_out))
                        _s = float(np.std(_rounds_out))
                        _cv_vals.append(_s / max(_m, 0.01))
                _cv = float(np.mean(_cv_vals)) if _cv_vals else 0.5
                self.df.at[idx, f"{corner}_pre_output_cv"] = _cv
                self.df.at[idx,
                           f"{corner}_pre_output_consistency"] = 1.0 / (1.0 + _cv)

                # ── Control-time GnP rate ─────────────────────────────────────────
                _gnp_vals = []
                for h in _h_rd:
                    _ctrl_total = (
                        h.get("rd1_ctrl", 0.0)
                        + h.get("rd2_ctrl", 0.0)
                        + h.get("rd3_ctrl", 0.0)
                    )
                    _grd_total = (
                        h.get("rd1_ground", 0.0)
                        + h.get("rd2_ground", 0.0)
                        + h.get("rd3_ground", 0.0)
                    )
                    if _ctrl_total > 0:
                        _gnp_vals.append(_grd_total / (_ctrl_total / 60.0))
                self.df.at[idx, f"{corner}_pre_ctrl_gnp_rate"] = (
                    float(np.mean(_gnp_vals)) if _gnp_vals else 0.0
                )

                # ── Strike quality evolution (sig/total ratio) ────────────────────
                _r1_sqvals = [
                    h.get("rd1_sig_str", 0.0) / max(h.get("rd1_str", 1.0), 1.0)
                    for h in _h_rd
                    if h.get("rd1_str", 0.0) > 0
                ]
                _r3_sqvals = [
                    h.get("rd3_sig_str", 0.0) / max(h.get("rd3_str", 1.0), 1.0)
                    for h in _h_rd
                    if h.get("rd3_str", 0.0) > 0
                ]
                _r1_sq = float(np.mean(_r1_sqvals)) if _r1_sqvals else 0.6
                _r3_sq = float(np.mean(_r3_sqvals)) if _r3_sqvals else 0.6
                self.df.at[idx, f"{corner}_pre_r1_sig_pct"] = _r1_sq
                self.df.at[idx, f"{corner}_pre_r3_sig_pct"] = _r3_sq
                self.df.at[idx, f"{corner}_pre_sig_pct_evo"] = _r3_sq - _r1_sq

                # ── Tactical divergence (zone distribution shift R1→R3) ────────────
                _tdiv_vals = []
                for h in _h_rd:
                    if h.get("rd1_sig_str", 0.0) > 0 and h.get("rd3_sig_str", 0.0) > 0:
                        _r1z = [
                            h.get("rd1_dist", 0.0)
                            / max(h.get("rd1_sig_str", 1.0), 1.0),
                            h.get("rd1_clinch", 0.0)
                            / max(h.get("rd1_sig_str", 1.0), 1.0),
                            h.get("rd1_ground", 0.0)
                            / max(h.get("rd1_sig_str", 1.0), 1.0),
                        ]
                        _r3z = [
                            h.get("rd3_dist", 0.0)
                            / max(h.get("rd3_sig_str", 1.0), 1.0),
                            h.get("rd3_clinch", 0.0)
                            / max(h.get("rd3_sig_str", 1.0), 1.0),
                            h.get("rd3_ground", 0.0)
                            / max(h.get("rd3_sig_str", 1.0), 1.0),
                        ]
                        _tdiv_vals.append(
                            float(
                                np.sqrt(sum((a - b) ** 2 for a,
                                        b in zip(_r1z, _r3z)))
                            )
                        )
                self.df.at[idx, f"{corner}_pre_tactical_div"] = (
                    float(np.mean(_tdiv_vals)) if _tdiv_vals else 0.0
                )

                # ── Fatigue composite ─────────────────────────────────────────────
                _abs_sl = self.df.at[idx, f"{corner}_pre_absorption_slope"]
                _cardio = self.df.at[idx, f"{corner}_pre_cardio_index"]
                self.df.at[idx, f"{corner}_pre_fatigue_composite"] = max(
                    _abs_sl, 0.0
                ) * max(1.0 - _cardio, 0.0)

                # ── R1 absorption rate ─────────────────────────────────────────────
                _r1_abs_vals = [
                    h.get("rd1_absorbed", 0.0) / 5.0 for h in _h_rd]
                self.df.at[idx, f"{corner}_pre_r1_abs_rate"] = (
                    float(np.mean(_r1_abs_vals)) if _r1_abs_vals else 0.0
                )

                # ── Championship damage margin (R4+R5) ────────────────────────────
                if _r45_fights:
                    _r45_margins = [
                        (
                            (h.get("rd4_sig_str", 0.0) + h.get("rd5_sig_str", 0.0))
                            - (h.get("rd4_absorbed", 0.0) +
                               h.get("rd5_absorbed", 0.0))
                        )
                        / 2.0
                        for h in _r45_fights
                    ]
                    self.df.at[idx, f"{corner}_pre_r45_dmg_margin"] = float(
                        np.mean(_r45_margins)
                    )
                else:
                    self.df.at[idx, f"{corner}_pre_r45_dmg_margin"] = 0.0

                # ── Combo rate (head + body both > 0 in R1) ───────────────────────
                _combo_vals = [
                    1
                    if h.get("rd1_head", 0.0) > 0 and h.get("rd1_body", 0.0) > 0
                    else 0
                    for h in _h_rd
                    if h.get("rd1_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_combo_rate"] = (
                    float(np.mean(_combo_vals)) if _combo_vals else 0.0
                )

                # ── Head strike momentum (R3 head / R1 head) ─────────────────────
                _hmom_vals = [
                    h.get("rd3_head", 0.0) / max(h.get("rd1_head", 1.0), 1.0)
                    for h in _h_rd
                    if h.get("rd3_sig_str", 0.0) > 0 and h.get("rd1_head", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_head_momentum"] = (
                    float(np.mean(_hmom_vals)) if _hmom_vals else 1.0
                )

                # ── Clinch vs ground ratio (dirty boxer vs pure wrestler) ──────────
                _cgr_vals = [
                    h.get("rd1_clinch", 0.0)
                    / max(h.get("rd1_clinch", 0.0) + h.get("rd1_ground", 0.0), 1.0)
                    for h in _h_rd
                    if (h.get("rd1_clinch", 0.0) + h.get("rd1_ground", 0.0)) > 0
                ]
                self.df.at[idx, f"{corner}_pre_clinch_grd_ratio"] = (
                    float(np.mean(_cgr_vals)) if _cgr_vals else 0.5
                )

                # ── Conditional finish rate (finish when winning R1) ──────────────
                _win_r1_fights = [
                    h
                    for h in _h_rd
                    if h.get("rd1_sig_str", 0.0) > h.get("rd1_absorbed", 0.0)
                    and (h.get("rd1_sig_str", 0.0) + h.get("rd1_absorbed", 0.0)) > 0
                ]
                if _win_r1_fights:
                    _fin_when_winning = [
                        1 if h.get("finished", 0) == 1 and h.get(
                            "won", 0) == 1 else 0
                        for h in _win_r1_fights
                    ]
                    self.df.at[idx, f"{corner}_pre_finish_when_winning"] = float(
                        np.mean(_fin_when_winning)
                    )
                else:
                    self.df.at[idx, f"{corner}_pre_finish_when_winning"] = 0.0

                # ── Sub attempt round timing ──────────────────────────────────────
                _sub_weighted = []
                for h in _h_rd:
                    _total_sub = (
                        h.get("rd1_sub_att", 0.0)
                        + h.get("rd2_sub_att", 0.0)
                        + h.get("rd3_sub_att", 0.0)
                    )
                    if _total_sub > 0:
                        _sub_rnd = (
                            1 * h.get("rd1_sub_att", 0.0)
                            + 2 * h.get("rd2_sub_att", 0.0)
                            + 3 * h.get("rd3_sub_att", 0.0)
                        ) / _total_sub
                        _sub_weighted.append(_sub_rnd)
                self.df.at[idx, f"{corner}_pre_sub_round_avg"] = (
                    float(np.mean(_sub_weighted)) if _sub_weighted else 2.0
                )

                # ── Integrated dominance ──────────────────────────────────────────
                _r1_dmg_v = self.df.at[idx, f"{corner}_pre_r1_damage_margin"]
                _cardio_v = self.df.at[idx, f"{corner}_pre_cardio_index"]
                _fin_rate = (
                    self.df.at[idx, f"{corner}_pre_finish_rate"]
                    if f"{corner}_pre_finish_rate" in self.df.columns
                    else 0.0
                )
                self.df.at[idx, f"{corner}_pre_integrated_dom"] = (
                    _r1_dmg_v * max(_cardio_v, 0.1) * (1.0 + _fin_rate)
                )

                # ── R2 KD threat ──────────────────────────────────────────────────
                _r2_kd_vals = [
                    h.get("rd2_kd", 0.0) for h in _h_rd if h.get("rd2_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_r2_kd_threat"] = (
                    float(np.mean(_r2_kd_vals)) if _r2_kd_vals else 0.0
                )

                # ── R1/R3 output consistency ──────────────────────────────────────
                _r1r3c_vals = []
                for h in _h_rd:
                    r1o = h.get("rd1_sig_str", 0.0)
                    r3o = h.get("rd3_sig_str", 0.0)
                    if (r1o + r3o) > 0:
                        _r1r3c_vals.append(1.0 - abs(r1o - r3o) / (r1o + r3o))
                self.df.at[idx, f"{corner}_pre_r1_r3_consistency"] = (
                    float(np.mean(_r1r3c_vals)) if _r1r3c_vals else 0.5
                )

                # ── R1 control-seeking rate ───────────────────────────────────────
                _r1_ctrl_fights = [
                    1 if h.get("rd1_ctrl", 0.0) > 0 else 0 for h in _h_rd
                ]
                self.df.at[idx, f"{corner}_pre_r1_ctrl_opening"] = (
                    float(np.mean(_r1_ctrl_fights)) if _r1_ctrl_fights else 0.0
                )

                # ── Quality-of-win R3 margin ──────────────────────────────────────
                _win_r3_margins = [
                    h.get("rd3_sig_str", 0.0) - h.get("rd3_absorbed", 0.0)
                    for h in _h_rd
                    if h.get("won", 0) == 1 and h.get("rd3_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_win_r3_margin"] = (
                    float(np.mean(_win_r3_margins)) if _win_r3_margins else 0.0
                )

                # ── Tier 12l: Per-round defense, adjustment speed, grappling depth ──

                # R1 TD defense rate: % of opponent R1 TD attempts that were stuffed
                _r1_td_def_vals = [
                    1.0
                    - (
                        h.get("rd1_td", 0.0)
                        / max(h.get("opp_rd1_td_att", 0.0) + h.get("rd1_td", 0.0), 1.0)
                    )
                    for h in _h_rd
                    if (h.get("opp_rd1_td_att", 0.0) + h.get("rd1_td", 0.0)) > 0
                ]
                self.df.at[idx, f"{corner}_pre_r1_td_def_rate"] = (
                    float(np.mean(_r1_td_def_vals)) if _r1_td_def_vals else 0.7
                )
                # Avg opponent R1 TD attempts (measures how much TD pressure the fighter faces)
                _r1_tdp_vals = [h.get("opp_rd1_td_att", 0.0) for h in _h_rd]
                self.df.at[idx, f"{corner}_pre_r1_td_pressure"] = (
                    float(np.mean(_r1_tdp_vals)) if _r1_tdp_vals else 0.0
                )

                # R2 head accuracy: R2 head / R2 sig_str_att (or R2 sig_str as proxy)
                _r2_hacc_vals = [
                    h.get("rd2_head", 0.0)
                    / max(h.get("rd2_sig_str_att", h.get("rd2_sig_str", 1.0)), 1.0)
                    for h in _h_rd
                    if h.get("rd2_sig_str", 0.0) > 0
                ]
                _r2_hacc = float(np.mean(_r2_hacc_vals)
                                 ) if _r2_hacc_vals else 0.3
                self.df.at[idx, f"{corner}_pre_r2_head_acc"] = _r2_hacc
                # R1→R2 head accuracy adjustment (positive = finding range in R2)
                _r1_hacc_l = self.df.at[idx, f"{corner}_pre_r1_head_acc"]
                self.df.at[idx,
                           f"{corner}_pre_r1r2_head_adj"] = _r2_hacc - _r1_hacc_l

                # Chin ratio: R3 absorbed / R1 absorbed (>1 means fading under volume)
                _chin_vals = [
                    h.get("rd3_absorbed", 0.0) /
                    max(h.get("rd1_absorbed", 1.0), 1.0)
                    for h in _h_rd
                    if h.get("rd3_absorbed", 0.0) > 0 and h.get("rd1_absorbed", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_chin_ratio"] = (
                    float(np.mean(_chin_vals)) if _chin_vals else 1.0
                )

                # R2+R3 KD rate: avg knockdowns across R2 and R3 (mid-to-late round finisher)
                _r2r3_kd_vals = [
                    h.get("rd2_kd", 0.0) + h.get("rd3_kd", 0.0)
                    for h in _h_rd
                    if h.get("rd2_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_r2r3_kd_rate"] = (
                    float(np.mean(_r2r3_kd_vals)) if _r2r3_kd_vals else 0.0
                )

                # Sub after TD: % fights where R1 td > 0 AND R2 sub_att > 0
                _sub_td_vals = [
                    1
                    if h.get("rd1_td", 0.0) > 0 and h.get("rd2_sub_att", 0.0) > 0
                    else 0
                    for h in _h_rd
                    if h.get("rd2_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_sub_after_td"] = (
                    float(np.mean(_sub_td_vals)) if _sub_td_vals else 0.0
                )

                # R1 output efficiency: R1 sig_str / max(R1 absorbed, 1)
                _r1eff_vals = [
                    h.get("rd1_sig_str", 0.0) /
                    max(h.get("rd1_absorbed", 1.0), 1.0)
                    for h in _h_rd
                    if h.get("rd1_sig_str", 0.0) > 0
                ]
                _r1_eff = float(np.mean(_r1eff_vals)) if _r1eff_vals else 1.0
                self.df.at[idx, f"{corner}_pre_r1_output_eff"] = _r1_eff

                # R3 output efficiency: R3 sig_str / max(R3 absorbed, 1)
                _r3eff_vals = [
                    h.get("rd3_sig_str", 0.0) /
                    max(h.get("rd3_absorbed", 1.0), 1.0)
                    for h in _h_rd
                    if h.get("rd3_sig_str", 0.0) > 0
                ]
                _r3_eff = float(np.mean(_r3eff_vals)) if _r3eff_vals else 1.0
                self.df.at[idx, f"{corner}_pre_r3_output_eff"] = _r3_eff
                # Trend: improving or worsening exchange quality over the fight
                self.df.at[idx,
                           f"{corner}_pre_output_eff_trend"] = _r3_eff - _r1_eff

                # R4+R5 head strikes (championship title-fight head work)
                if _r45_fights:
                    _r45_head_vals = [
                        h.get("rd4_head", 0.0) + h.get("rd5_head", 0.0)
                        for h in _r45_fights
                    ]
                    self.df.at[idx, f"{corner}_pre_r45_head_rate"] = float(
                        np.mean(_r45_head_vals)
                    )
                else:
                    self.df.at[idx, f"{corner}_pre_r45_head_rate"] = 0.0

                # R1→R2 zone adjustment: R2 dist_pct − R1 dist_pct
                # Positive = backing off in R2; Negative = pressing in harder after R1
                _r1r2z_vals = []
                for h in _h_rd:
                    if h.get("rd2_sig_str", 0.0) > 0 and h.get("rd1_sig_str", 0.0) > 0:
                        _r2_dp = h.get("rd2_dist", 0.0) / max(
                            h.get("rd2_sig_str", 1.0), 1.0
                        )
                        _r1_dp = h.get("rd1_dist", 0.0) / max(
                            h.get("rd1_sig_str", 1.0), 1.0
                        )
                        _r1r2z_vals.append(_r2_dp - _r1_dp)
                self.df.at[idx, f"{corner}_pre_r1r2_zone_adj"] = (
                    float(np.mean(_r1r2z_vals)) if _r1r2z_vals else 0.0
                )

                # R2+R3 sub attempts per fight (late-round submission hunter)
                _r23sub_vals = [
                    h.get("rd2_sub_att", 0.0) + h.get("rd3_sub_att", 0.0)
                    for h in _h_rd
                    if h.get("rd2_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_r23_sub_rate"] = (
                    float(np.mean(_r23sub_vals)) if _r23sub_vals else 0.0
                )

                # R4+R5 sub attempts (championship-round submission threat)
                if _r45_fights:
                    _r45sub_vals = [
                        h.get("rd4_sub_att", 0.0) + h.get("rd5_sub_att", 0.0)
                        for h in _r45_fights
                    ]
                    self.df.at[idx, f"{corner}_pre_r45_sub_rate_l"] = float(
                        np.mean(_r45sub_vals)
                    )
                else:
                    self.df.at[idx, f"{corner}_pre_r45_sub_rate_l"] = 0.0

                # R4+R5 ground strikes (championship GnP specialist)
                if _r45_fights:
                    _r45grd_vals = [
                        h.get("rd4_ground", 0.0) + h.get("rd5_ground", 0.0)
                        for h in _r45_fights
                    ]
                    self.df.at[idx, f"{corner}_pre_r45_ground_rate"] = float(
                        np.mean(_r45grd_vals)
                    )
                else:
                    self.df.at[idx, f"{corner}_pre_r45_ground_rate"] = 0.0

                # R3 leg rate: avg R3 leg strikes per fight
                _r3leg_vals = [
                    h.get("rd3_leg", 0.0)
                    for h in _h_rd
                    if h.get("rd3_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_r3_leg_rate"] = (
                    float(np.mean(_r3leg_vals)) if _r3leg_vals else 0.0
                )

                # ── Tier 12m: Style entropy adaptation, grappling efficiency, resilience ──

                # R2 target entropy (head/body/leg distribution in R2)
                _r2tent_vals = []
                _r2zent_vals = []
                for h in _h_rd:
                    if h.get("rd2_sig_str", 0.0) > 0:
                        _r2tent_vals.append(
                            _shannon(
                                [
                                    h.get("rd2_head", 0.0),
                                    h.get("rd2_body", 0.0),
                                    h.get("rd2_leg", 0.0),
                                ]
                            )
                        )
                        _r2zent_vals.append(
                            _shannon(
                                [
                                    h.get("rd2_dist", 0.0),
                                    h.get("rd2_clinch", 0.0),
                                    h.get("rd2_ground", 0.0),
                                ]
                            )
                        )
                _r2_tent = float(np.mean(_r2tent_vals)
                                 ) if _r2tent_vals else 1.0
                _r2_zent = float(np.mean(_r2zent_vals)
                                 ) if _r2zent_vals else 1.0
                self.df.at[idx, f"{corner}_pre_r2_target_ent"] = _r2_tent
                self.df.at[idx, f"{corner}_pre_r2_zone_ent"] = _r2_zent
                # Entropy shift R1→R2: positive = more diverse in R2 (adapting style)
                _r1_tent = self.df.at[idx, f"{corner}_pre_target_entropy"]
                _r1_zent = self.df.at[idx, f"{corner}_pre_zone_entropy"]
                self.df.at[idx,
                           f"{corner}_pre_r1r2_tgt_shift"] = _r2_tent - _r1_tent
                self.df.at[idx,
                           f"{corner}_pre_r1r2_zone_shift"] = _r2_zent - _r1_zent

                # Control per R1 TD (grappling dominance: ctrl seconds per successful takedown)
                _cptd_vals = [
                    h.get("rd1_ctrl", 0.0) / max(h.get("rd1_td", 0.5), 0.5)
                    for h in _h_rd
                    if h.get("rd1_td", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_ctrl_per_td"] = (
                    float(np.mean(_cptd_vals)) if _cptd_vals else 0.0
                )

                # Finish round entropy (versatile finisher vs. R1-only finisher)
                _fin_rnds = [
                    h.get("finish_round", 0.0)
                    for h in _h_rd
                    if h.get("finished", 0) == 1 and h.get("finish_round", 0.0) > 0
                ]
                if len(_fin_rnds) >= 2:
                    _rnd_counts = [
                        _fin_rnds.count(1),
                        _fin_rnds.count(2),
                        _fin_rnds.count(3),
                        _fin_rnds.count(4),
                        _fin_rnds.count(5),
                    ]
                    self.df.at[idx, f"{corner}_pre_finish_rnd_ent"] = _shannon(
                        _rnd_counts
                    )
                else:
                    self.df.at[idx, f"{corner}_pre_finish_rnd_ent"] = 0.0

                # R3 body pct and body pct evolution
                _r3bpct_vals = [
                    h.get("rd3_body", 0.0) /
                    max(h.get("rd3_sig_str", 1.0), 1.0)
                    for h in _h_rd
                    if h.get("rd3_sig_str", 0.0) > 0
                ]
                _r3_bpct = float(np.mean(_r3bpct_vals)
                                 ) if _r3bpct_vals else 0.0
                _r1_bpct_vals = [
                    h.get("rd1_body", 0.0) /
                    max(h.get("rd1_sig_str", 1.0), 1.0)
                    for h in _h_rd
                    if h.get("rd1_sig_str", 0.0) > 0
                ]
                _r1_bpct = float(np.mean(_r1_bpct_vals)
                                 ) if _r1_bpct_vals else 0.0
                self.df.at[idx, f"{corner}_pre_r3_body_pct"] = _r3_bpct
                self.df.at[idx,
                           f"{corner}_pre_body_pct_evo"] = _r3_bpct - _r1_bpct

                # R1 clinch→ground link (dirty boxer who transitions to wrestling)
                _cg_vals = [
                    h.get("rd1_ground", 0.0) /
                    max(h.get("rd1_clinch", 0.5), 0.5)
                    for h in _h_rd
                    if h.get("rd1_clinch", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_r1_cg_link"] = (
                    float(np.mean(_cg_vals)) if _cg_vals else 0.0
                )

                # Output in losses (competitive spirit: still throws punches when losing)
                _loss_output = [
                    h.get("sig_str", 0.0) for h in _h_rd if h.get("won", 0) == 0
                ]
                self.df.at[idx, f"{corner}_pre_output_in_loss"] = (
                    float(np.mean(_loss_output)) if _loss_output else 0.0
                )

                # Reversals per control minute (escape/scramble artist)
                _tot_rev = sum(h.get("rev", 0.0) for h in _h_rd)
                _tot_ctrl = sum(
                    h.get("rd1_ctrl", 0.0)
                    + h.get("rd2_ctrl", 0.0)
                    + h.get("rd3_ctrl", 0.0)
                    for h in _h_rd
                )
                self.df.at[idx, f"{corner}_pre_rev_per_ctrl"] = _tot_rev / max(
                    _tot_ctrl / 60.0, 1.0
                )

                # R2/R1 output ratio (pacing strategy: surges in R2 vs. conserves)
                _r2r1_vals = [
                    h.get("rd2_sig_str", 0.0) /
                    max(h.get("rd1_sig_str", 1.0), 1.0)
                    for h in _h_rd
                    if h.get("rd2_sig_str", 0.0) > 0 and h.get("rd1_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_r2_r1_ratio"] = (
                    float(np.mean(_r2r1_vals)) if _r2r1_vals else 1.0
                )

                # R3 leg pct (late-round leg targeting %)
                _r3lpct_vals = [
                    h.get("rd3_leg", 0.0) / max(h.get("rd3_sig_str", 1.0), 1.0)
                    for h in _h_rd
                    if h.get("rd3_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_r3_leg_pct"] = (
                    float(np.mean(_r3lpct_vals)) if _r3lpct_vals else 0.0
                )

                # R3 output in won fights (how dominant in the final scoring round of wins)
                _wr3out_vals = [
                    h.get("rd3_sig_str", 0.0)
                    for h in _h_rd
                    if h.get("won", 0) == 1 and h.get("rd3_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_win_r3_output"] = (
                    float(np.mean(_wr3out_vals)) if _wr3out_vals else 0.0
                )

                # ── Tier 12n: Finish method diversity, mid-fight adjustment, decision quality ──

                # Finish method entropy: Shannon diversity across KO / sub / decision wins
                _ko_wins = sum(h.get("ko", 0) for h in _h_rd)
                _sub_wins = sum(h.get("sub_win", 0) for h in _h_rd)
                _dec_wins = sum(
                    1
                    for h in _h_rd
                    if h.get("won", 0) == 1
                    and h.get("ko", 0) == 0
                    and h.get("sub_win", 0) == 0
                )
                self.df.at[idx, f"{corner}_pre_method_entropy"] = _shannon(
                    [_ko_wins, _sub_wins, _dec_wins]
                )

                # R2 output efficiency: R2 sig_str / max(R2 absorbed, 1)
                _r2eff_vals = [
                    h.get("rd2_sig_str", 0.0) /
                    max(h.get("rd2_absorbed", 1.0), 1.0)
                    for h in _h_rd
                    if h.get("rd2_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_r2_output_eff"] = (
                    float(np.mean(_r2eff_vals)) if _r2eff_vals else 1.0
                )

                # R1 front-load ratio: R1 slpm / career average slpm
                _overall_slpm = (
                    float(np.mean([h.get("slpm", 0.0) for h in _h_rd]))
                    if _h_rd
                    else 1.0
                )
                _r1_slpm_avg = self.df.at[idx, f"{corner}_pre_rd1_slpm_avg"]
                self.df.at[idx, f"{corner}_pre_r1_front_load"] = _r1_slpm_avg / max(
                    _overall_slpm, 0.5
                )

                # R2 TD accuracy: R2 td / R2 td_att
                _r2tacc_vals = [
                    h.get("rd2_td", 0.0) / max(h.get("rd2_td_att", 0.5), 0.5)
                    for h in _h_rd
                    if h.get("rd2_td_att", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_r2_td_acc"] = (
                    float(np.mean(_r2tacc_vals)) if _r2tacc_vals else 0.0
                )

                # Decision win margin: R3 sig_str − R3 absorbed in fights won by decision
                _dec_wins_h = [
                    h
                    for h in _h_rd
                    if h.get("won", 0) == 1
                    and h.get("ko", 0) == 0
                    and h.get("sub_win", 0) == 0
                    and h.get("rd3_sig_str", 0.0) > 0
                ]
                if _dec_wins_h:
                    _dwm_vals = [
                        h.get("rd3_sig_str", 0.0) - h.get("rd3_absorbed", 0.0)
                        for h in _dec_wins_h
                    ]
                    self.df.at[idx, f"{corner}_pre_dec_win_margin"] = float(
                        np.mean(_dwm_vals)
                    )
                else:
                    self.df.at[idx, f"{corner}_pre_dec_win_margin"] = 0.0

                # Control escalation ratio: R3 ctrl / R1 ctrl (wrestler builds control over time)
                _ctrl_esc_vals = [
                    h.get("rd3_ctrl", 0.0) / max(h.get("rd1_ctrl", 0.5), 0.5)
                    for h in _h_rd
                    if h.get("rd3_ctrl", 0.0) > 0 and h.get("rd1_ctrl", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_ctrl_esc_ratio"] = (
                    float(np.mean(_ctrl_esc_vals)) if _ctrl_esc_vals else 1.0
                )

                # Absorbed in losses: avg absorbed in lost fights (got badly outworked?)
                _abs_loss_vals = [
                    h.get("rd1_absorbed", 0.0)
                    + h.get("rd2_absorbed", 0.0)
                    + h.get("rd3_absorbed", 0.0)
                    for h in _h_rd
                    if h.get("won", 0) == 0
                ]
                self.df.at[idx, f"{corner}_pre_absorbed_in_loss"] = (
                    float(np.mean(_abs_loss_vals)) if _abs_loss_vals else 0.0
                )

                # Grappling usage index: total td_att + sub_att across R1-R3 per fight
                _grapple_usage_vals = [
                    (
                        h.get("rd1_td_att", 0.0)
                        + h.get("rd2_td_att", 0.0)
                        + h.get("rd3_td_att", 0.0)
                        + h.get("rd1_sub_att", 0.0)
                        + h.get("rd2_sub_att", 0.0)
                        + h.get("rd3_sub_att", 0.0)
                    )
                    for h in _h_rd
                ]
                self.df.at[idx, f"{corner}_pre_grapple_usage"] = (
                    float(np.mean(_grapple_usage_vals)
                          ) if _grapple_usage_vals else 0.0
                )

                # KD R1 concentration: fraction of all KDs that land in R1
                _all_kds = sum(
                    h.get("rd1_kd", 0.0) + h.get("rd2_kd", 0.0) +
                    h.get("rd3_kd", 0.0)
                    for h in _h_rd
                )
                _r1_kds = sum(h.get("rd1_kd", 0.0) for h in _h_rd)
                self.df.at[idx, f"{corner}_pre_kd_r1_conc"] = _r1_kds / max(
                    _all_kds, 1.0
                )

                # Finish when behind: finish rate in fights where R1 margin was negative
                _behind_fights = [
                    h
                    for h in _h_rd
                    if h.get("rd1_sig_str", 0.0) < h.get("rd1_absorbed", 0.0)
                    and (h.get("rd1_sig_str", 0.0) + h.get("rd1_absorbed", 0.0)) > 0
                ]
                if _behind_fights:
                    _fwb_vals = [
                        1
                        if (
                            h.get("won", 0) == 1
                            and (h.get("ko", 0) == 1 or h.get("sub_win", 0) == 1)
                        )
                        else 0
                        for h in _behind_fights
                    ]
                    self.df.at[idx, f"{corner}_pre_finish_when_behind"] = float(
                        np.mean(_fwb_vals)
                    )
                else:
                    self.df.at[idx, f"{corner}_pre_finish_when_behind"] = 0.0

                # R2 win rate: % fights where R2 output > R2 absorbed
                _r2wins_vals = [
                    1 if h.get("rd2_sig_str", 0.0) > h.get(
                        "rd2_absorbed", 0.0) else 0
                    for h in _h_rd
                    if (h.get("rd2_sig_str", 0.0) + h.get("rd2_absorbed", 0.0)) > 0
                ]
                self.df.at[idx, f"{corner}_pre_r2_win_rate"] = (
                    float(np.mean(_r2wins_vals)) if _r2wins_vals else 0.5
                )

                # R3 control rate: avg R3 control seconds (late-round wrestling time)
                _r3ctrl_vals = [
                    h.get("rd3_ctrl", 0.0)
                    for h in _h_rd
                    if h.get("rd3_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_r3_ctrl_rate"] = (
                    float(np.mean(_r3ctrl_vals)) if _r3ctrl_vals else 0.0
                )

                # R1 body rate: avg R1 body strikes (volume of early body work)
                _r1body_vals = [
                    h.get("rd1_body", 0.0)
                    for h in _h_rd
                    if h.get("rd1_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_r1_body_rate"] = (
                    float(np.mean(_r1body_vals)) if _r1body_vals else 0.0
                )

                # R3 distance pct: dist / sig_str in R3 (late-round range preference)
                _r3dpct_vals = [
                    h.get("rd3_dist", 0.0) /
                    max(h.get("rd3_sig_str", 1.0), 1.0)
                    for h in _h_rd
                    if h.get("rd3_sig_str", 0.0) > 0
                ]
                self.df.at[idx, f"{corner}_pre_r3_dist_pct"] = (
                    float(np.mean(_r3dpct_vals)) if _r3dpct_vals else 0.5
                )

            r_elo_pre = self.feature_engineer.elo_get(r)
            b_elo_pre = self.feature_engineer.elo_get(b)

            _finish_round = int(_safe(row.get("finish_round", 0), 0))
            _r_streak = rs["win_streak"]
            _b_streak = bs["win_streak"]
            _r_elo_pre = self.feature_engineer.elo_get(r)
            _b_elo_pre = self.feature_engineer.elo_get(b)
            self.feature_engineer.elo_update(
                r,
                b,
                winner,
                is_title,
                method,
                fighter_fights_count[r],
                fighter_fights_count[b],
                finish_round=_finish_round,
                winner_streak=_r_streak if winner == "Red" else _b_streak,
                opponent_elo=_b_elo_pre if winner == "Red" else _r_elo_pre,
                r_loss_streak=rs["loss_streak"],
                b_loss_streak=bs["loss_streak"],
            )
            self.df.at[idx, "r_elo_pre_fight"] = r_elo_pre
            self.df.at[idx, "b_elo_pre_fight"] = b_elo_pre

            r_glicko = self.feature_engineer.glicko2_get(r)
            b_glicko = self.feature_engineer.glicko2_get(b)
            r_score = 1.0 if winner == "Red" else (
                0.5 if winner == "Draw" else 0.0)
            b_score = 1.0 - r_score
            self.feature_engineer.glicko2_update(
                r, [(b_glicko[0], b_glicko[1], r_score)]
            )
            self.feature_engineer.glicko2_update(
                b, [(r_glicko[0], r_glicko[1], b_score)]
            )

            self.feature_engineer.update_common_opponents(r, b, winner)

            r_won = winner == "Red"
            b_won = winner == "Blue"
            ft = _safe(row.get("total_fight_time_sec", 0), 0.0)

            def update_fighter(fs, won, fighter_name):
                fs["total_fights"] += 1
                if won:
                    fs["wins"] += 1
                    fs["win_streak"] += 1
                    fs["loss_streak"] = 0
                    if method in ("KO/TKO",):
                        fs["ko_wins"] += 1
                    elif method in ("Submission",):
                        fs["sub_wins"] += 1
                    else:
                        fs["dec_wins"] += 1
                        if method == "Unanimous Decision":
                            fs["unanimous_wins"] += 1
                        elif method == "Majority Decision":
                            fs["majority_wins"] += 1
                        elif method == "Split Decision":
                            fs["split_wins"] += 1
                    if is_title:
                        fs["title_wins"] += 1
                elif winner == "Draw":
                    fs["draws"] += 1
                else:
                    fs["losses"] += 1
                    fs["loss_streak"] += 1
                    fs["win_streak"] = 0
                    if method in ("KO/TKO",):
                        fs["ko_losses"] += 1
                    elif method in ("Submission",):
                        fs["sub_losses"] += 1
                    else:
                        fs["dec_losses"] += 1
                        if method == "Split Decision":
                            fs["split_losses"] += 1
                if is_title:
                    fs["title_fights"] += 1
                if fs["wins"] > 0:
                    fs["finish_rate"] = (
                        fs["ko_wins"] + fs["sub_wins"]) / fs["wins"]
                fs["total_fight_time"] += ft
                fs["avg_fight_time"] = fs["total_fight_time"] / max(
                    fs["total_fights"], 1
                )

            def update_fight_stats(fs, prefix, row_data):
                ss = _safe(row_data.get(f"{prefix}_sig_str", 0))
                ss_att = _safe(row_data.get(f"{prefix}_sig_str_att", 0))
                st = _safe(row_data.get(f"{prefix}_str", 0))
                st_att = _safe(row_data.get(f"{prefix}_str_att", 0))
                td = _safe(row_data.get(f"{prefix}_td", 0))
                td_att = _safe(row_data.get(f"{prefix}_td_att", 0))
                sub = _safe(row_data.get(f"{prefix}_sub_att", 0))
                kd = _safe(row_data.get(f"{prefix}_kd", 0))
                ctrl = _safe(row_data.get(f"{prefix}_ctrl_sec", 0))
                rev = _safe(row_data.get(f"{prefix}_rev", 0))
                head = _safe(row_data.get(f"{prefix}_head", 0))
                body = _safe(row_data.get(f"{prefix}_body", 0))
                leg = _safe(row_data.get(f"{prefix}_leg", 0))
                dist = _safe(row_data.get(f"{prefix}_distance", 0))
                clinch = _safe(row_data.get(f"{prefix}_clinch", 0))
                ground = _safe(row_data.get(f"{prefix}_ground", 0))
                fs["total_sig_str"] += ss
                fs["total_sig_str_att"] += ss_att
                fs["total_str"] += st
                fs["total_str_att"] += st_att
                fs["total_td"] += td
                fs["total_td_att"] += td_att
                fs["total_sub_att"] += sub
                fs["total_kd"] += kd
                fs["total_ctrl_sec"] += ctrl
                fs["total_rev"] += rev
                fs["total_head"] += head
                fs["total_body"] += body
                fs["total_leg"] += leg
                fs["total_distance"] += dist
                fs["total_clinch"] += clinch
                fs["total_ground"] += ground
                return {
                    "won": 1
                    if (prefix == "r" and r_won) or (prefix == "b" and b_won)
                    else 0,
                    "sig_str": ss,
                    "td": td,
                    "kd": kd,
                    "sub_att": sub,
                    "ctrl": ctrl,
                }

            update_fighter(rs, r_won, r)
            update_fighter(bs, b_won, b)
            r_hist = update_fight_stats(rs, "r", row)
            b_hist = update_fight_stats(bs, "b", row)
            rs["_history"].append(r_hist)
            bs["_history"].append(b_hist)
            fighter_fights_count[r] += 1
            fighter_fights_count[b] += 1

            event_date = row.get("event_date", None)
            if pd.notna(event_date):
                rs["fight_dates"].append(event_date)
                bs["fight_dates"].append(event_date)

            rs["opp_elo_history"].append(b_elo_pre)
            bs["opp_elo_history"].append(r_elo_pre)

            r_kd_absorbed = int(_safe(row.get("b_kd", 0), 0))
            b_kd_absorbed = int(_safe(row.get("r_kd", 0), 0))
            rs["kd_absorbed"] += r_kd_absorbed
            bs["kd_absorbed"] += b_kd_absorbed

            _finish_rnd = int(_safe(row.get("finish_round", 0), 0))
            _total_rds = int(_safe(row.get("total_rounds", 3), 3))

            for corner, fs, opp_prefix, won in [
                ("r", rs, "b", r_won),
                ("b", bs, "r", b_won),
            ]:
                opp_ss_att = int(
                    _safe(row.get(f"{opp_prefix}_sig_str_att", 0), 0))
                opp_ss_lnd = int(_safe(row.get(f"{opp_prefix}_sig_str", 0), 0))
                opp_td_att = int(_safe(row.get(f"{opp_prefix}_td_att", 0), 0))
                opp_td_lnd = int(_safe(row.get(f"{opp_prefix}_td", 0), 0))
                fs["str_def_attempts"] += opp_ss_att
                fs["str_def_success"] += opp_ss_att - opp_ss_lnd
                fs["td_def_attempts"] += opp_td_att
                fs["td_def_success"] += opp_td_att - opp_td_lnd

                if won and method in ("KO/TKO", "Submission"):
                    if _finish_rnd <= 2:
                        fs["early_finishes"] += 1
                    else:
                        fs["late_finishes"] += 1
                    if _finish_rnd == 1 and method == "KO/TKO":
                        fs["first_round_kos"] += 1

                fs["total_rounds_fought"] += (
                    _finish_rnd if _finish_rnd > 0 else _total_rds
                )
                if _total_rds >= 5:
                    fs["five_round_fights"] += 1

                _h10_ss = _safe(row.get(f"{corner}_sig_str", 0))
                _h10_ft = max(ft / 60.0, 1.0 / 60.0)
                _h10_opp = max(opp_ss_lnd, 1)
                _h10_d = _safe(row.get(f"{corner}_distance", 0))
                _h10_c = _safe(row.get(f"{corner}_clinch", 0))
                _h10_g = _safe(row.get(f"{corner}_ground", 0))
                # Round 1 fields
                _rd1_ss = _safe(row.get(f"{corner}_rd1_sig_str", 0))
                _rd1_ss_att = _safe(row.get(f"{corner}_rd1_sig_str_att", 0))
                _rd1_kd = _safe(row.get(f"{corner}_rd1_kd", 0))
                _rd1_td = _safe(row.get(f"{corner}_rd1_td", 0))
                _rd1_td_att = _safe(row.get(f"{corner}_rd1_td_att", 0))
                _rd1_head = _safe(row.get(f"{corner}_rd1_head", 0))
                _rd1_hatt = _safe(row.get(f"{corner}_rd1_head_att", 0))
                _rd1_body = _safe(row.get(f"{corner}_rd1_body", 0))
                _rd1_leg = _safe(row.get(f"{corner}_rd1_leg", 0))
                _rd1_leg_att = _safe(row.get(f"{corner}_rd1_leg_att", 0))
                _rd1_clinch = _safe(row.get(f"{corner}_rd1_clinch", 0))
                _rd1_dist = _safe(row.get(f"{corner}_rd1_distance", 0))
                _rd1_ground = _safe(row.get(f"{corner}_rd1_ground", 0))
                _rd1_sub_att = _safe(row.get(f"{corner}_rd1_sub_att", 0))
                _rd1_ctrl = _safe(row.get(f"{corner}_rd1_ctrl_sec", 0))
                _rd1_str = _safe(row.get(f"{corner}_rd1_str", 0))
                _rd1_str_att = _safe(row.get(f"{corner}_rd1_str_att", 0))
                # Round 2 fields (full)
                _rd2_ss = _safe(row.get(f"{corner}_rd2_sig_str", 0))
                _rd2_kd = _safe(row.get(f"{corner}_rd2_kd", 0))
                _rd2_td = _safe(row.get(f"{corner}_rd2_td", 0))
                _rd2_td_att = _safe(row.get(f"{corner}_rd2_td_att", 0))
                _rd2_head = _safe(row.get(f"{corner}_rd2_head", 0))
                _rd2_body = _safe(row.get(f"{corner}_rd2_body", 0))
                _rd2_leg = _safe(row.get(f"{corner}_rd2_leg", 0))
                _rd2_clinch = _safe(row.get(f"{corner}_rd2_clinch", 0))
                _rd2_ground = _safe(row.get(f"{corner}_rd2_ground", 0))
                _rd2_sub_att = _safe(row.get(f"{corner}_rd2_sub_att", 0))
                _rd2_ctrl = _safe(row.get(f"{corner}_rd2_ctrl_sec", 0))
                _rd2_str = _safe(row.get(f"{corner}_rd2_str", 0))
                _rd2_str_att = _safe(row.get(f"{corner}_rd2_str_att", 0))
                # Round 3 fields
                _rd3_ss = _safe(row.get(f"{corner}_rd3_sig_str", 0))
                _rd3_ss_att = _safe(row.get(f"{corner}_rd3_sig_str_att", 0))
                _rd3_kd = _safe(row.get(f"{corner}_rd3_kd", 0))
                _rd3_td = _safe(row.get(f"{corner}_rd3_td", 0))
                _rd3_td_att = _safe(row.get(f"{corner}_rd3_td_att", 0))
                _rd3_body = _safe(row.get(f"{corner}_rd3_body", 0))
                _rd3_leg = _safe(row.get(f"{corner}_rd3_leg", 0))
                _rd3_clinch = _safe(row.get(f"{corner}_rd3_clinch", 0))
                _rd3_dist = _safe(row.get(f"{corner}_rd3_distance", 0))
                _rd3_ground = _safe(row.get(f"{corner}_rd3_ground", 0))
                _rd3_sub_att = _safe(row.get(f"{corner}_rd3_sub_att", 0))
                _rd3_ctrl = _safe(row.get(f"{corner}_rd3_ctrl_sec", 0))
                _rd3_str = _safe(row.get(f"{corner}_rd3_str", 0))
                _rd3_str_att = _safe(row.get(f"{corner}_rd3_str_att", 0))
                _rd3_head = _safe(row.get(f"{corner}_rd3_head", 0))
                _rd3_head_att = _safe(row.get(f"{corner}_rd3_head_att", 0))
                # Opponent per-round pressure (chin / TD-defense signals)
                _opp_rd1_head = _safe(row.get(f"{opp_prefix}_rd1_head", 0))
                _opp_rd1_td_att = _safe(row.get(f"{opp_prefix}_rd1_td_att", 0))
                # Opponent R4/R5 sig strikes (for championship damage margins)
                _rd4_abs = _safe(row.get(f"{opp_prefix}_rd4_sig_str", 0))
                _rd5_abs = _safe(row.get(f"{opp_prefix}_rd5_sig_str", 0))
                # Championship round fields (R4/R5) — full
                _rd4_ss = _safe(row.get(f"{corner}_rd4_sig_str", 0))
                _rd4_td = _safe(row.get(f"{corner}_rd4_td", 0))
                _rd4_kd = _safe(row.get(f"{corner}_rd4_kd", 0))
                _rd4_body = _safe(row.get(f"{corner}_rd4_body", 0))
                _rd4_leg = _safe(row.get(f"{corner}_rd4_leg", 0))
                _rd4_clinch = _safe(row.get(f"{corner}_rd4_clinch", 0))
                _rd4_ctrl = _safe(row.get(f"{corner}_rd4_ctrl_sec", 0))
                _rd5_ss = _safe(row.get(f"{corner}_rd5_sig_str", 0))
                _rd5_td = _safe(row.get(f"{corner}_rd5_td", 0))
                _rd5_kd = _safe(row.get(f"{corner}_rd5_kd", 0))
                _rd5_body = _safe(row.get(f"{corner}_rd5_body", 0))
                _rd5_leg = _safe(row.get(f"{corner}_rd5_leg", 0))
                _rd5_clinch = _safe(row.get(f"{corner}_rd5_clinch", 0))
                _rd5_ctrl = _safe(row.get(f"{corner}_rd5_ctrl_sec", 0))
                # Reversals (overall fight)
                _rev = _safe(row.get(f"{corner}_rev", 0))
                # Absorption: opponent's sig strikes against this fighter per round
                _rd1_abs = _safe(row.get(f"{opp_prefix}_rd1_sig_str", 0))
                _rd2_abs = _safe(row.get(f"{opp_prefix}_rd2_sig_str", 0))
                _rd3_abs = _safe(row.get(f"{opp_prefix}_rd3_sig_str", 0))
                # ── Tier 43: exact fight-history counters (this fight, after pre-fight snapshot) ──
                _t43_td = _safe(row.get(f"{corner}_td", 0))
                _t43_kd = _safe(row.get(f"{corner}_kd", 0))
                _t43_ctrl = _safe(row.get(f"{corner}_ctrl_sec", 0))
                _t43_r1_ex = (_rd1_ss + _rd1_abs) > 0
                _t43_won_r1 = _t43_r1_ex and _rd1_ss > _rd1_abs
                _t43_lost_r1 = _t43_r1_ex and _rd1_ss < _rd1_abs
                if not won and winner != "Draw" and method in ("KO/TKO", "Submission"):
                    if _finish_rnd > 0:
                        if _finish_rnd <= 2:
                            fs["early_finish_losses"] += 1
                        else:
                            fs["late_finish_losses"] += 1
                if won and method == "Submission" and _t43_td > 0:
                    fs["sub_wins_with_td"] += 1
                if won and method == "KO/TKO" and _t43_kd > 0:
                    fs["ko_wins_with_kd"] += 1
                if _t43_lost_r1:
                    fs["fights_lost_r1"] += 1
                    if (
                        not won
                        and winner != "Draw"
                        and method in ("KO/TKO", "Submission")
                    ):
                        fs["finish_losses_after_lost_r1"] += 1
                    if won and method == "Submission":
                        fs["sub_wins_after_lost_r1"] += 1
                    if won and method == "KO/TKO":
                        fs["ko_wins_after_lost_r1"] += 1
                if _t43_won_r1:
                    fs["fights_won_r1"] += 1
                    if won and method in (
                        "Unanimous Decision",
                        "Majority Decision",
                        "Split Decision",
                    ):
                        fs["dec_wins_after_won_r1"] += 1
                if won and method in ("KO/TKO", "Submission"):
                    if _t43_kd > 0 or _t43_ctrl >= 60.0 or _t43_td > 0:
                        fs["finish_wins_phys_dom"] += 1
                if _t43_kd > 0 or _t43_ctrl >= 60.0 or _t43_td > 0:
                    fs["t43_dom_wins"] += 1
                    if won and method in ("KO/TKO", "Submission"):
                        fs["t43_finish_dom_wins"] += 1
                # Tier 12l: opponent R1 TDs (for R1 TD defense), R2 zone, R4/R5 head/sub/ground
                _opp_rd1_td = _safe(row.get(f"{opp_prefix}_rd1_td", 0))
                _rd2_ss_att = _safe(row.get(f"{corner}_rd2_sig_str_att", 0))
                _rd2_dist = _safe(row.get(f"{corner}_rd2_distance", 0))
                _rd4_head = _safe(row.get(f"{corner}_rd4_head", 0))
                _rd5_head = _safe(row.get(f"{corner}_rd5_head", 0))
                _rd4_sub_att = _safe(row.get(f"{corner}_rd4_sub_att", 0))
                _rd5_sub_att = _safe(row.get(f"{corner}_rd5_sub_att", 0))
                _rd4_ground = _safe(row.get(f"{corner}_rd4_ground", 0))
                _rd5_ground = _safe(row.get(f"{corner}_rd5_ground", 0))
                # Per-round reversals (R1/R2/R3) — scramble escalation signal
                _rd1_rev = _safe(row.get(f"{corner}_rd1_rev", 0))
                _rd2_rev = _safe(row.get(f"{corner}_rd2_rev", 0))
                _rd3_rev = _safe(row.get(f"{corner}_rd3_rev", 0))
                # Championship round TD attempts — needed for R4/R5 TD accuracy
                _rd4_td_att = _safe(row.get(f"{corner}_rd4_td_att", 0))
                _rd5_td_att = _safe(row.get(f"{corner}_rd5_td_att", 0))
                # rounds_available: how many complete-ish rounds exist in this fight
                _rds_avail = _finish_rnd if _finish_rnd > 0 else _total_rds
                # ── Tier 13: Variables for decayed average features ──────────────────
                # Per-round head/body counts (counts, not percentages)
                _da_head_total = (
                    _rd1_head + _rd2_head + _rd3_head + _rd4_head + _rd5_head
                )
                _da_body_total = (
                    _rd1_body + _rd2_body + _rd3_body + _rd4_body + _rd5_body
                )
                # Total attempts from per-round data (actual counts)
                _da_sig_str_att = _safe(row.get(f"{corner}_sig_str_att", 0))
                _da_td_total = _safe(row.get(f"{corner}_td", 0))
                _da_td_att = _safe(row.get(f"{corner}_td_att", 0))
                _da_head_att = sum(
                    _safe(row.get(f"{corner}_rd{_r}_head_att", 0)) for _r in range(1, 6)
                )
                _da_body_att = sum(
                    _safe(row.get(f"{corner}_rd{_r}_body_att", 0)) for _r in range(1, 6)
                )
                _da_dist_att = sum(
                    _safe(row.get(f"{corner}_rd{_r}_distance_att", 0))
                    for _r in range(1, 6)
                )
                # Opponent totals for defense metrics (per-round counts)
                _da_opp_head = sum(
                    _safe(row.get(f"{opp_prefix}_rd{_r}_head", 0)) for _r in range(1, 6)
                )
                _da_opp_head_att = sum(
                    _safe(row.get(f"{opp_prefix}_rd{_r}_head_att", 0))
                    for _r in range(1, 6)
                )
                _da_opp_body = sum(
                    _safe(row.get(f"{opp_prefix}_rd{_r}_body", 0)) for _r in range(1, 6)
                )
                _da_opp_body_att = sum(
                    _safe(row.get(f"{opp_prefix}_rd{_r}_body_att", 0))
                    for _r in range(1, 6)
                )
                _da_opp_dist = sum(
                    _safe(row.get(f"{opp_prefix}_rd{_r}_distance", 0))
                    for _r in range(1, 6)
                )
                _da_opp_dist_att = sum(
                    _safe(row.get(f"{opp_prefix}_rd{_r}_distance_att", 0))
                    for _r in range(1, 6)
                )
                _da_opp_grd = sum(
                    _safe(row.get(f"{opp_prefix}_rd{_r}_ground", 0))
                    for _r in range(1, 6)
                )
                _da_opp_grd_att = sum(
                    _safe(row.get(f"{opp_prefix}_rd{_r}_ground_att", 0))
                    for _r in range(1, 6)
                )
                _da_opp_leg = _safe(row.get(f"{opp_prefix}_leg", 0)) * _safe(
                    row.get(f"{opp_prefix}_sig_str", 0)
                )
                _da_opp_sub = _safe(row.get(f"{opp_prefix}_sub_att", 0))
                # Fighter's own per-round distance landed (count)
                _da_dist_total = sum(
                    _safe(row.get(f"{corner}_rd{_r}_distance", 0)) for _r in range(1, 6)
                )
                _da_opp_td_att = _safe(row.get(f"{opp_prefix}_td_att", 0))
                _da_opp_td = _safe(row.get(f"{opp_prefix}_td", 0))
                _da_opp_kd = _safe(row.get(f"{opp_prefix}_kd", 0))
                _da_opp_rd1_ctrl = _safe(
                    row.get(f"{opp_prefix}_rd1_ctrl_sec", 0))
                _da_opp_rd1_rev = _safe(row.get(f"{opp_prefix}_rd1_rev", 0))
                _da_opp_elo = b_elo_pre if corner == "r" else r_elo_pre
                _da_age = _safe(row.get(f"{corner}_age_at_event", 28.0), 28.0)
                _da_fighter_reach = _safe(
                    row.get(f"{corner}_reach", 70.0), 70.0)
                _da_opp_reach = _safe(
                    row.get(f"{opp_prefix}_reach", 70.0), 70.0)
                _da_ufc_fight_num = float(len(fs["rolling10_history"]) + 1)
                _da_fight_dates = fs.get("fight_dates", [])
                if len(_da_fight_dates) >= 2:
                    try:
                        _da_days_since = float(
                            (_da_fight_dates[-1] - _da_fight_dates[-2]).days
                        )
                    except Exception:
                        _da_days_since = 365.0
                else:
                    _da_days_since = 365.0
                _t43_dmg_m123 = (
                    (_rd1_ss - _rd1_abs) +
                    (_rd2_ss - _rd2_abs) + (_rd3_ss - _rd3_abs)
                )
                _t43_fight_dom = _safe(row.get(f"{corner}_sig_str", 0)) - float(
                    _safe(row.get(f"{opp_prefix}_sig_str", 0), 0)
                )
                _t43_head_abs_share = float(
                    _da_opp_head) / max(float(opp_ss_lnd), 1.0)
                _was_finish_win = (
                    1 if (won and method in ("KO/TKO", "Submission")) else 0
                )
                h10 = {
                    "won": 1 if won else 0,
                    "sig_str": _h10_ss,
                    "td": _safe(row.get(f"{corner}_td", 0)),
                    "kd": _safe(row.get(f"{corner}_kd", 0)),
                    "sub_att": _safe(row.get(f"{corner}_sub_att", 0)),
                    "finished": 1
                    if (won and method in ("KO/TKO", "Submission"))
                    else 0,
                    "slpm": _h10_ss / _h10_ft,
                    "damage_ratio": _h10_ss / _h10_opp,
                    "distance_pct": _h10_d / max(_h10_d + _h10_c + _h10_g, 1.0),
                    # Metadata
                    "rounds_available": _rds_avail,
                    "is_five_round": 1 if _total_rds >= 5 else 0,
                    # Round 1
                    # Round 1
                    "rd1_sig_str": _rd1_ss,
                    "rd1_slpm": _rd1_ss / 5.0,
                    "rd1_sig_str_att": _rd1_ss_att,
                    "rd1_kd": _rd1_kd,
                    "rd1_td": _rd1_td,
                    "rd1_td_att": _rd1_td_att,
                    "rd1_head": _rd1_head,
                    "rd1_head_att": _rd1_hatt,
                    "rd1_body": _rd1_body,
                    "rd1_leg": _rd1_leg,
                    "rd1_leg_att": _rd1_leg_att,
                    "rd1_clinch": _rd1_clinch,
                    "rd1_dist": _rd1_dist,
                    "rd1_ground": _rd1_ground,
                    "rd1_sub_att": _rd1_sub_att,
                    "rd1_ctrl": _rd1_ctrl,
                    "rd1_str": _rd1_str,
                    "rd1_str_att": _rd1_str_att,
                    # Round 2 (full)
                    "rd2_sig_str": _rd2_ss,
                    "rd2_kd": _rd2_kd,
                    "rd2_td": _rd2_td,
                    "rd2_td_att": _rd2_td_att,
                    "rd2_head": _rd2_head,
                    "rd2_body": _rd2_body,
                    "rd2_leg": _rd2_leg,
                    "rd2_clinch": _rd2_clinch,
                    "rd2_ground": _rd2_ground,
                    "rd2_sub_att": _rd2_sub_att,
                    "rd2_ctrl": _rd2_ctrl,
                    "rd2_str": _rd2_str,
                    "rd2_str_att": _rd2_str_att,
                    # Round 3
                    "rd3_sig_str": _rd3_ss,
                    "rd3_sig_str_att": _rd3_ss_att,
                    "rd3_kd": _rd3_kd,
                    "rd3_td": _rd3_td,
                    "rd3_td_att": _rd3_td_att,
                    "rd3_body": _rd3_body,
                    "rd3_leg": _rd3_leg,
                    "rd3_clinch": _rd3_clinch,
                    "rd3_dist": _rd3_dist,
                    "rd3_ground": _rd3_ground,
                    "rd3_sub_att": _rd3_sub_att,
                    "rd3_ctrl": _rd3_ctrl,
                    "rd3_str": _rd3_str,
                    "rd3_str_att": _rd3_str_att,
                    "rd3_head": _rd3_head,
                    "rd3_head_att": _rd3_head_att,
                    # Championship rounds (full)
                    "rd4_sig_str": _rd4_ss,
                    "rd4_td": _rd4_td,
                    "rd4_kd": _rd4_kd,
                    "rd4_body": _rd4_body,
                    "rd4_leg": _rd4_leg,
                    "rd4_clinch": _rd4_clinch,
                    "rd4_ctrl": _rd4_ctrl,
                    "rd5_sig_str": _rd5_ss,
                    "rd5_td": _rd5_td,
                    "rd5_kd": _rd5_kd,
                    "rd5_body": _rd5_body,
                    "rd5_leg": _rd5_leg,
                    "rd5_clinch": _rd5_clinch,
                    "rd5_ctrl": _rd5_ctrl,
                    # Reversals (overall + per-round for scramble escalation)
                    "rev": _rev,
                    "rd1_rev": _rd1_rev,
                    "rd2_rev": _rd2_rev,
                    "rd3_rev": _rd3_rev,
                    # Finish round metadata
                    "finish_round": float(_finish_rnd),
                    # Absorption (opponent sig strikes per round)
                    "rd1_absorbed": _rd1_abs,
                    "rd2_absorbed": _rd2_abs,
                    "rd3_absorbed": _rd3_abs,
                    # Opponent per-round pressure
                    "opp_rd1_head": _opp_rd1_head,
                    "opp_rd1_td_att": _opp_rd1_td_att,
                    # Championship round absorption
                    "rd4_absorbed": _rd4_abs,
                    "rd5_absorbed": _rd5_abs,
                    # Tier 12l additions
                    "opp_rd1_td": _opp_rd1_td,
                    "rd2_sig_str_att": _rd2_ss_att,
                    "rd2_dist": _rd2_dist,
                    "rd4_head": _rd4_head,
                    "rd5_head": _rd5_head,
                    "rd4_sub_att": _rd4_sub_att,
                    "rd5_sub_att": _rd5_sub_att,
                    "rd4_ground": _rd4_ground,
                    "rd5_ground": _rd5_ground,
                    # Championship round TD attempts (for R4/R5 TD accuracy)
                    "rd4_td_att": _rd4_td_att,
                    "rd5_td_att": _rd5_td_att,
                    # Tier 12n: finish method flags (for method entropy)
                    "ko": 1 if (won and method == "KO/TKO") else 0,
                    "sub_win": 1 if (won and method == "Submission") else 0,
                    # ── Tier 13: pre-computed ratios for decayed average features ────
                    "sig_str_att": _da_sig_str_att,
                    "td_att": _da_td_att,
                    "total_head_landed": _da_head_total,
                    "total_body_landed": _da_body_total,
                    "total_head_att": _da_head_att,
                    "total_body_att": _da_body_att,
                    "total_dist_att": _da_dist_att,
                    "fight_time_min": _h10_ft,
                    "opp_elo_pre": _da_opp_elo,
                    "age_at_event": _da_age,
                    # Pre-computed accuracy / defense ratios (computed once, reused across fights)
                    "sig_str_acc_v": _h10_ss / max(_da_sig_str_att, 1),
                    "td_acc_v": _da_td_total / max(_da_td_att, 1),
                    "head_acc_v": _da_head_total / max(_da_head_att, 1),
                    "body_acc_v": _da_body_total / max(_da_body_att, 1),
                    "dist_acc_v": _da_dist_total / max(_da_dist_att, 1),
                    "dist_landing_ratio_v": _da_dist_total / max(_h10_ss, 1),
                    "clinch_pm_v": _h10_c * _h10_ss / max(_h10_ft, 0.0167),
                    "head_landing_ratio_v": _da_head_total / max(_h10_ss, 1),
                    "head_def_v": (_da_opp_head_att - _da_opp_head)
                    / max(_da_opp_head_att, 1),
                    "body_def_v": (_da_opp_body_att - _da_opp_body)
                    / max(_da_opp_body_att, 1),
                    "dist_def_v": (_da_opp_dist_att - _da_opp_dist)
                    / max(_da_opp_dist_att, 1),
                    "grd_def_v": (_da_opp_grd_att - _da_opp_grd)
                    / max(_da_opp_grd_att, 1),
                    "td_def_v": (_da_opp_td_att - _da_opp_td) / max(_da_opp_td_att, 1),
                    "opp_leg_pm_v": _da_opp_leg / max(_h10_ft, 0.0167),
                    "opp_ctrl_r1_pm_v": _da_opp_rd1_ctrl / 5.0,
                    "opp_sub_pm_v": _da_opp_sub / max(_h10_ft, 0.0167),
                    "opp_rd1_rev_v": _da_opp_rd1_rev,
                    "opp_kd_v": _da_opp_kd,
                    "reach_ratio_v": _da_fighter_reach / max(_da_opp_reach, 1.0),
                    "ufc_fight_num_v": _da_ufc_fight_num,
                    "days_since_v": _da_days_since,
                    # Tier 43 exact rolling signals (pre-fight da/qa/ra from history)
                    "damage_margin_r123": _t43_dmg_m123,
                    "fight_dom_raw": _t43_fight_dom,
                    "head_sig_abs_share_v": _t43_head_abs_share,
                    "was_finish_win": float(_was_finish_win),
                    "event_date": row.get("event_date"),
                }
                fs["rolling10_history"].append(h10)

                fs["last_fight_was_win"] = bool(won)
                fs["last_fight_was_finish"] = bool(
                    won and method in ("KO/TKO", "Submission")
                )
                fs["last_fight_method"] = method
                fs["last_fight_finish_round"] = _finish_rnd

                cur_elo = self.feature_engineer.elo_get(
                    r if corner == "r" else b)
                if cur_elo > fs["career_elo_peak"]:
                    fs["career_elo_peak"] = cur_elo
                    fs["fights_since_peak"] = 0
                else:
                    fs["fights_since_peak"] += 1

                opp_elo_pre = b_elo_pre if corner == "r" else r_elo_pre
                opp_slpm = _safe(row.get(f"{opp_prefix}_pre_SLpM", 0))
                opp_td_avg = _safe(row.get(f"{opp_prefix}_pre_td_avg", 0))
                if opp_elo_pre > 1600:
                    fs["vs_elite_fights"] += 1
                    if won:
                        fs["vs_elite_wins"] += 1
                if opp_slpm > 4.5:
                    fs["vs_striker_fights"] += 1
                    if won:
                        fs["vs_striker_wins"] += 1
                if opp_td_avg > 2.5:
                    fs["vs_grappler_fights"] += 1
                    if won:
                        fs["vs_grappler_wins"] += 1

            if winner in ("Red", "Blue"):
                _wck_post = (
                    str(row.get("weight_class", "")).strip(),
                    str(row.get("gender", "")).lower().strip(),
                    int(_safe(row.get("total_rounds", 3), 3)) or 3,
                )
                _wcp = self._wc_finish_counts[_wck_post]
                if method == "KO/TKO":
                    _wcp["ko"] += 1
                elif method == "Submission":
                    _wcp["sub"] += 1
                else:
                    _wcp["dec"] += 1
                _dk = 1 if method == "KO/TKO" else 0
                _ds = 1 if method == "Submission" else 0
                self._wc_finish_raw_events[_wck_post].append(
                    (_wc_seq, _dk, _ds))

        self._wc_chrono_total_fights = len(self.df)
        self._finalize_wc_finish_chrono_timelines()

        self._log(
            f"Data leakage fix complete. DataFrame shape: {self.df.shape}")

    @staticmethod
    def _apply_tier43_method_head_features(df):
        """Tier 43: Method-head vulnerability + conversion (Decision/Finish, KO/Sub)."""
        _eps = 1e-6

        def _g(c, d=0.0):
            return df.get(c, pd.Series(d, index=df.index)).fillna(d)

        r_fw = _g("r_pre_finish_rate", 0.0).clip(0.0, 1.0)
        b_fw = _g("b_pre_finish_rate", 0.0).clip(0.0, 1.0)
        r_kwr = _g("r_ko_win_rate", 0.0).clip(0.0, 1.0)
        b_kwr = _g("b_ko_win_rate", 0.0).clip(0.0, 1.0)
        r_swr = _g("r_sub_win_rate", 0.0).clip(0.0, 1.0)
        b_swr = _g("b_sub_win_rate", 0.0).clip(0.0, 1.0)
        r_dwr = _g("r_decision_win_rate", 0.0).clip(0.0, 1.0)
        b_dwr = _g("b_decision_win_rate", 0.0).clip(0.0, 1.0)
        r_klr = _g("r_ko_loss_rate", 0.0).clip(0.0, 1.0)
        b_klr = _g("b_ko_loss_rate", 0.0).clip(0.0, 1.0)
        r_slr = _g("r_sub_loss_rate", 0.0).clip(0.0, 1.0)
        b_slr = _g("b_sub_loss_rate", 0.0).clip(0.0, 1.0)
        r_flr = (r_klr + r_slr).clip(0.0, 1.0)
        b_flr = (b_klr + b_slr).clip(0.0, 1.0)
        r_surv = (1.0 - r_flr).clip(0.0, 1.0)
        b_surv = (1.0 - b_flr).clip(0.0, 1.0)
        r_dur = 1.0 / (1.0 + r_flr)
        b_dur = 1.0 / (1.0 + b_flr)
        r_efr = r_fw
        b_efr = b_fw
        r_early_w = _g("r_pre_early_finish_rate", 0.0).clip(0.0, 1.0)
        b_early_w = _g("b_pre_early_finish_rate", 0.0).clip(0.0, 1.0)
        r_late_w = _g("r_pre_late_finish_rate", 0.0).clip(0.0, 1.0)
        b_late_w = _g("b_pre_late_finish_rate", 0.0).clip(0.0, 1.0)
        r_early_lr = _g("r_pre_early_finish_loss_rate", 0.0).clip(0.0, 1.0)
        b_early_lr = _g("b_pre_early_finish_loss_rate", 0.0).clip(0.0, 1.0)
        r_late_lr = _g("r_pre_late_finish_loss_rate", 0.0).clip(0.0, 1.0)
        b_late_lr = _g("b_pre_late_finish_loss_rate", 0.0).clip(0.0, 1.0)
        r_kde = (_g("r_pre_kd_efficiency", 0.0) / 100.0).clip(0.0, 5.0)
        b_kde = (_g("b_pre_kd_efficiency", 0.0) / 100.0).clip(0.0, 5.0)
        r_vol_abs = _g("r_pre_SApM", 0.0).clip(0.0, 50.0) * _g(
            "r_pre_avg_fight_time", 1.0
        ).clip(0.1, 60.0)
        b_vol_abs = _g("b_pre_SApM", 0.0).clip(0.0, 50.0) * _g(
            "b_pre_avg_fight_time", 1.0
        ).clip(0.1, 60.0)
        r_kda = (_g("r_pre_kd_absorbed", 0.0) / np.maximum(r_vol_abs, _eps)).clip(
            0.0, 5.0
        )
        b_kda = (_g("b_pre_kd_absorbed", 0.0) / np.maximum(b_vol_abs, _eps)).clip(
            0.0, 5.0
        )
        r_subc = _g("r_pre_ex_sub_conv", 0.0).clip(0.0, 50.0)
        b_subc = _g("b_pre_ex_sub_conv", 0.0).clip(0.0, 50.0)
        r_ctrl_subc = _g("r_pre_ex_ctrl_sub_conv", 0.0).clip(0.0, 50.0)
        b_ctrl_subc = _g("b_pre_ex_ctrl_sub_conv", 0.0).clip(0.0, 50.0)
        r_kd_tko = _g("r_pre_ex_ko_tko_conv", 0.0).clip(0.0, 50.0)
        b_kd_tko = _g("b_pre_ex_ko_tko_conv", 0.0).clip(0.0, 50.0)
        r_dist = _g("r_pre_distance_pct", 0.33).clip(0.0, 1.0)
        b_dist = _g("b_pre_distance_pct", 0.33).clip(0.0, 1.0)
        r_clinch = _g("r_pre_clinch_pct", 0.33).clip(0.0, 1.0)
        b_clinch = _g("b_pre_clinch_pct", 0.33).clip(0.0, 1.0)
        r_grd = _g("r_pre_ground_pct", 0.33).clip(0.0, 1.0)
        b_grd = _g("b_pre_ground_pct", 0.33).clip(0.0, 1.0)
        r_ftm = _g("r_pre_avg_fight_time", 1.0).clip(0.1, 60.0) * 60.0
        b_ftm = _g("b_pre_avg_fight_time", 1.0).clip(0.1, 60.0) * 60.0
        r_grd_ctrl_sh = r_grd * \
            (_g("r_pre_ctrl_avg", 0.0) / np.maximum(r_ftm, _eps))
        b_grd_ctrl_sh = b_grd * \
            (_g("b_pre_ctrl_avg", 0.0) / np.maximum(b_ftm, _eps))
        r_dom_split = (r_dist + r_clinch) - r_grd_ctrl_sh
        b_dom_split = (b_dist + b_clinch) - b_grd_ctrl_sh
        r_grnd_str = (r_grd * _g("r_pre_SLpM", 0.0)).clip(0.0, 20.0)
        b_grnd_str = (b_grd * _g("b_pre_SLpM", 0.0)).clip(0.0, 20.0)
        r_sub_att_r = _g("r_pre_sub_att_rate", 0.0).clip(0.0, 50.0)
        b_sub_att_r = _g("b_pre_sub_att_rate", 0.0).clip(0.0, 50.0)
        r_clinch_td = (_g("r_pre_r1_clinch_rate", 0.0) * _g("r_pre_td_avg", 0.0)).clip(
            0.0, 50.0
        )
        b_clinch_td = (_g("b_pre_r1_clinch_rate", 0.0) * _g("b_pre_td_avg", 0.0)).clip(
            0.0, 50.0
        )
        r_head_share = _g("r_pre_head_pct", 0.33).clip(0.0, 1.0)
        b_head_share = _g("b_pre_head_pct", 0.33).clip(0.0, 1.0)
        r_head_frag = (
            _g("r_pre_da_head_sig_abs_share", 0.45).clip(0.0, 1.0) * r_kda
        ).clip(0.0, 5.0)
        b_head_frag = (
            _g("b_pre_da_head_sig_abs_share", 0.45).clip(0.0, 1.0) * b_kda
        ).clip(0.0, 5.0)
        r_dm_dec = _g("r_pre_da_damage_margin", 0.0)
        b_dm_dec = _g("b_pre_da_damage_margin", 0.0)
        r_ctrl_rate = _g("r_pre_ctrl_min_per15", 0.0) / 15.0
        b_ctrl_rate = _g("b_pre_ctrl_min_per15", 0.0) / 15.0
        r_dom_fc = _g("r_pre_dom_finish_conv_exact", 0.0).clip(0.0, 1.0)
        b_dom_fc = _g("b_pre_dom_finish_conv_exact", 0.0).clip(0.0, 1.0)
        r_surv_hurt = 1.0 - \
            _g("r_pre_finish_loss_lost_r1_rate", 0.0).clip(0.0, 1.0)
        b_surv_hurt = 1.0 - \
            _g("b_pre_finish_loss_lost_r1_rate", 0.0).clip(0.0, 1.0)
        r_lh_dec = _g("r_pre_dec_win_won_r1_rate", 0.0).clip(0.0, 1.0)
        b_lh_dec = _g("b_pre_dec_win_won_r1_rate", 0.0).clip(0.0, 1.0)
        r_fin = np.maximum(r_fw, _eps)
        b_fin = np.maximum(b_fw, _eps)
        r_pko = (r_kwr / r_fin).clip(0.0, 1.0)
        r_psub = (r_swr / r_fin).clip(0.0, 1.0)
        b_pko = (b_kwr / b_fin).clip(0.0, 1.0)
        b_psub = (b_swr / b_fin).clip(0.0, 1.0)
        r_purity = np.abs(r_pko - r_psub)
        b_purity = np.abs(b_pko - b_psub)
        r_ent = -r_pko * np.log(r_pko + _eps) - r_psub * np.log(r_psub + _eps)
        b_ent = -b_pko * np.log(b_pko + _eps) - b_psub * np.log(b_psub + _eps)
        r_entry = _g("r_pre_sub_win_after_td_rate", 0.0) - _g(
            "r_pre_ko_win_after_kd_rate", 0.0
        )
        b_entry = _g("b_pre_sub_win_after_td_rate", 0.0) - _g(
            "b_pre_ko_win_after_kd_rate", 0.0
        )
        r_neck = (
            _g("b_pre_ctrl_avg", 0.0) * r_slr /
            np.maximum(_g("r_pre_losses", 1.0), 1.0)
        )
        b_neck = (
            _g("r_pre_ctrl_avg", 0.0) * b_slr /
            np.maximum(_g("b_pre_losses", 1.0), 1.0)
        )
        r_reach_adv = (_g("r_reach", 72.0) - _g("b_reach", 72.0)) / 72.0
        b_reach_adv = -r_reach_adv
        r_sub_surv_def = r_slr - r_swr
        b_sub_surv_def = b_slr - b_swr
        r_ko_surv_def = r_klr - r_kwr
        b_ko_surv_def = b_klr - b_kwr
        r_late_sub = _g("r_pre_late_sub_rate", 0.0) * r_swr
        b_late_sub = _g("b_pre_late_sub_rate", 0.0) * b_swr
        r_early_ko = _g("r_pre_first_round_ko_rate", 0.0) * r_kwr
        b_early_ko = _g("b_pre_first_round_ko_rate", 0.0) * b_kwr
        r_cb_sub = _g("r_pre_sub_win_lost_r1_rate", 0.0).clip(0.0, 1.0)
        b_cb_sub = _g("b_pre_sub_win_lost_r1_rate", 0.0).clip(0.0, 1.0)
        r_cb_ko = _g("r_pre_ko_win_lost_r1_rate", 0.0).clip(0.0, 1.0)
        b_cb_ko = _g("b_pre_ko_win_lost_r1_rate", 0.0).clip(0.0, 1.0)
        r_tf = _g("r_pre_total_fights", 0.0)
        b_tf = _g("b_pre_total_fights", 0.0)
        r_mexp = np.where(r_tf >= b_tf, r_efr, b_efr)
        l_exp = np.where(r_tf >= b_tf, b_efr, r_efr)
        r_mlr = np.where(r_tf >= b_tf, r_flr, b_flr)
        l_lr = np.where(r_tf >= b_tf, b_flr, r_flr)

        df["mutual_finish_pressure"] = r_efr * b_flr + b_efr * r_flr
        df["mutual_durability"] = r_dur * b_dur
        df["finish_net_mismatch"] = (r_efr - b_surv) - (b_efr - r_surv)
        df["abs_finish_risk"] = np.maximum(r_efr * b_flr, b_efr * r_flr)
        df["decision_gravity"] = r_dwr * b_dwr
        df["decision_chaos"] = df["decision_gravity"] - \
            df["mutual_finish_pressure"]
        df["early_danger_cross"] = r_early_w * \
            b_early_lr + b_early_w * r_early_lr
        df["late_collapse_cross"] = r_late_w * b_late_lr + b_late_w * r_late_lr
        df["kd_finish_pressure"] = r_kde * b_flr + b_kde * r_flr
        df["damage_but_decision"] = r_dm_dec * r_dwr + b_dm_dec * b_dwr
        df["control_but_decision"] = r_ctrl_rate * r_dwr + b_ctrl_rate * b_dwr
        df["dom_finish_conv_gap"] = r_dom_fc - b_dom_fc
        df["survive_hurt_sum"] = r_surv_hurt + b_surv_hurt
        df["lead_hold_dec_total"] = r_lh_dec + b_lh_dec
        df["pace_sustain_clash"] = (
            _g("r_pre_cardio_index", 0.75)
            * _g("r_pre_output_consistency", 0.5)
            * _g("b_pre_cardio_index", 0.75)
            * _g("b_pre_output_consistency", 0.5)
        )
        df["frontload_violence"] = (
            _g("r_pre_r1_front_load", 1.0) * r_efr
            + _g("b_pre_r1_front_load", 1.0) * b_efr
        )
        df["fatigue_finish_cross"] = (
            _g("r_pre_fatigue_composite", 0.5) * r_flr
            + _g("b_pre_fatigue_composite", 0.5) * b_flr
        )
        _cpr = _g("cluster_pair_finish_rate", 0.5).clip(0.0, 1.0)
        df["style_finish_lookup"] = (_cpr - 0.5) * 2.0 * (r_efr + b_efr) * 0.5 + _g(
            "style_matchup_edge", 0.0
        ) * (r_efr + b_efr) * 0.25
        df["wc_route_prior"] = _g("wc_finish_route_prior", 0.0).clip(-1.0, 1.0)
        df["style_finish_delta"] = df["style_finish_lookup"] - np.tanh(
            df["wc_route_prior"] * 0.1
        )
        df["layoff_finish_var"] = np.abs(
            np.log1p(_g("r_days_since_last", 180.0).clip(0.0, 2000.0))
            - np.log1p(_g("b_days_since_last", 180.0).clip(0.0, 2000.0))
        ) * (r_flr + b_flr)
        df["exp_chaos"] = np.abs(r_tf - b_tf) * (r_mexp * l_lr + l_exp * r_mlr)

        r_ko_path = r_kwr * b_klr
        b_ko_path = b_kwr * r_klr
        r_sub_path = r_swr * b_slr
        b_sub_path = b_swr * r_slr
        df["r_ks_ko_path"] = r_ko_path
        df["b_ks_ko_path"] = b_ko_path
        df["r_ks_sub_path"] = r_sub_path
        df["b_ks_sub_path"] = b_sub_path
        df["ks_net_path"] = np.maximum(r_ko_path, b_ko_path) - np.maximum(
            r_sub_path, b_sub_path
        )
        df["ks_ko_danger"] = r_kde * b_head_frag + b_kde * r_head_frag
        df["ks_ground_capture"] = (
            _g("r_pre_td_per15", 0.0) *
            _g("r_pre_ctrl_per_td", 1.0) * r_subc * b_slr
            + _g("b_pre_td_per15", 0.0) *
            _g("b_pre_ctrl_per_td", 1.0) * b_subc * r_slr
        )
        df["ks_antiwrestle_break_pair"] = (
            1.0 - _g("r_pre_td_def", 0.5)
        ) * r_slr * b_swr + (1.0 - _g("b_pre_td_def", 0.5)) * b_slr * r_swr
        df["ks_scramble_sub_match"] = (
            _g("r_pre_reversal_rate", 0.0) * r_sub_att_r * b_slr
            + _g("b_pre_reversal_rate", 0.0) * b_sub_att_r * r_slr
        )
        df["r_ks_ctrl_sub_conv"] = r_ctrl_subc.clip(0.0, 50.0)
        df["b_ks_ctrl_sub_conv"] = b_ctrl_subc.clip(0.0, 50.0)
        df["r_ks_kd_tko_conv"] = r_kd_tko.clip(0.0, 50.0)
        df["b_ks_kd_tko_conv"] = b_kd_tko.clip(0.0, 50.0)
        df["r_ks_entry_split"] = r_entry
        df["b_ks_entry_split"] = b_entry
        df["r_ks_finish_purity"] = r_purity
        df["b_ks_finish_purity"] = b_purity
        df["r_ks_finish_entropy"] = r_ent
        df["b_ks_finish_entropy"] = b_ent
        df["r_ks_domain_split"] = r_dom_split
        df["b_ks_domain_split"] = b_dom_split
        df["r_ks_clinch_sub"] = r_clinch_td * r_sub_att_r
        df["b_ks_clinch_sub"] = b_clinch_td * b_sub_att_r
        df["ks_chin_power"] = r_kde * b_kda + b_kde * r_kda
        df["r_ks_neck_exposure"] = r_neck.clip(0.0, 50.0)
        df["b_ks_neck_exposure"] = b_neck.clip(0.0, 50.0)
        df["r_ks_headhunter"] = (
            r_head_share * r_kde *
            _g("r_pre_r1_head_pressure", 0.0).clip(0.0, 30.0)
        )
        df["b_ks_headhunter"] = (
            b_head_share * b_kde *
            _g("b_pre_r1_head_pressure", 0.0).clip(0.0, 30.0)
        )
        df["r_ks_gnp_vs_sub"] = r_grnd_str / np.maximum(r_sub_att_r, _eps)
        df["b_ks_gnp_vs_sub"] = b_grnd_str / np.maximum(b_sub_att_r, _eps)
        df["r_ks_sub_survival_deficit"] = r_sub_surv_def
        df["b_ks_sub_survival_deficit"] = b_sub_surv_def
        df["r_ks_ko_survival_deficit"] = r_ko_surv_def
        df["b_ks_ko_survival_deficit"] = b_ko_surv_def
        df["ks_route_diff"] = (r_kwr * b_klr + b_kwr * r_klr) - (
            r_swr * b_slr + b_swr * r_slr
        )
        df["r_ks_reach_power"] = r_reach_adv * r_kde * r_dist
        df["b_ks_reach_power"] = b_reach_adv * b_kde * b_dist
        df["r_ks_grapple_depth"] = _g(
            "r_pre_td_per15", 0.0) * r_ctrl_rate * r_sub_att_r
        df["b_ks_grapple_depth"] = _g(
            "b_pre_td_per15", 0.0) * b_ctrl_rate * b_sub_att_r
        df["r_ks_comeback_sub"] = r_cb_sub.clip(0.0, 5.0)
        df["b_ks_comeback_sub"] = b_cb_sub.clip(0.0, 5.0)
        df["r_ks_comeback_ko"] = r_cb_ko.clip(0.0, 5.0)
        df["b_ks_comeback_ko"] = b_cb_ko.clip(0.0, 5.0)
        df["r_ks_late_sub"] = r_late_sub
        df["b_ks_late_sub"] = b_late_sub
        df["r_ks_early_ko"] = r_early_ko
        df["b_ks_early_ko"] = b_early_ko
        df["ks_route_mismatch"] = np.abs(r_ko_path - r_sub_path) + np.abs(
            b_ko_path - b_sub_path
        )

        _q_r = _g("r_avg_opp_elo_L5", 1500.0) / 1500.0
        _q_b = _g("b_avg_opp_elo_L5", 1500.0) / 1500.0
        df["qa_ks_ko_path_r"] = r_kwr * _q_r * b_klr
        df["qa_ks_sub_path_r"] = r_swr * _q_r * b_slr
        df["qa_mutual_finish_pressure"] = (
            r_efr * _q_r * b_flr * _q_b + b_efr * _q_b * r_flr * _q_r
        ) + 0.15 * (
            _g("r_pre_qa_finish_win", 0.0).clip(0.0, 1.0)
            + _g("b_pre_qa_finish_win", 0.0).clip(0.0, 1.0)
        )
        r_sc = _g("r_style_cluster", -1.0)
        b_sc = _g("b_style_cluster", -1.0)
        _style_sim = (1.0 - np.minimum(np.abs(r_sc - b_sc),
                      7.0) / 8.0).clip(0.0, 1.0)
        df["sa_ks_ground_capture"] = df["ks_ground_capture"] * _style_sim
        df["sa_style_finish_lookup"] = df["style_finish_lookup"] * _style_sim
        _lam = 1.0
        r_age_d = np.log1p(_g("r_days_since_last", 180.0).clip(0.0, 2000.0))
        b_age_d = np.log1p(_g("b_days_since_last", 180.0).clip(0.0, 2000.0))
        _rec = np.exp(-_lam * (r_age_d + b_age_d) / 2.0)
        df["ra_early_danger"] = (
            df["early_danger_cross"]
            + 0.3
            * (
                _g("r_pre_ra_finish_win", 0.0).clip(0.0, 1.0)
                + _g("b_pre_ra_finish_win", 0.0).clip(0.0, 1.0)
            )
        ) * _rec
        df["ra_ks_late_sub"] = (r_late_sub + b_late_sub) * _rec
        df["ra_ks_early_ko"] = (r_early_ko + b_early_ko) * _rec

        df["ko_path_diff"] = r_ko_path - b_ko_path
        df["ko_path_sum"] = r_ko_path + b_ko_path
        df["ko_path_prod"] = r_ko_path * b_ko_path
        df["ko_path_max"] = np.maximum(r_ko_path, b_ko_path)
        df["ko_path_min"] = np.minimum(r_ko_path, b_ko_path)
        df["ko_path_asym"] = np.abs(r_ko_path - b_ko_path)
        df["sub_path_diff"] = r_sub_path - b_sub_path
        df["sub_path_sum"] = r_sub_path + b_sub_path
        df["sub_path_prod"] = r_sub_path * b_sub_path
        df["mutual_finish_pressure_prod"] = df["mutual_finish_pressure"] ** 2
        df["durability_asymmetry"] = np.abs(r_dur - b_dur)
        r_ff = _g("r_pre_fatigue_composite", 0.5) * r_flr
        b_ff = _g("b_pre_fatigue_composite", 0.5) * b_flr
        df["fatigue_finish_diff"] = r_ff - b_ff
        df["fatigue_finish_sum"] = r_ff + b_ff
        df["fatigue_finish_prod"] = r_ff * b_ff
        df["fatigue_finish_asym"] = np.abs(r_ff - b_ff)
        df["net_path_abs"] = np.abs(df["ks_net_path"])

        for _pfx, _opp in [("r", "b"), ("b", "r")]:
            df[f"{_pfx}_t43_dom_finish_conv"] = _g(
                f"{_pfx}_pre_dom_finish_conv_exact", 0.0
            ).clip(0.0, 1.0)
            df[f"{_pfx}_t43_survive_hurt"] = 1.0 - _g(
                f"{_pfx}_pre_finish_loss_lost_r1_rate", 0.0
            ).clip(0.0, 1.0)
            df[f"{_pfx}_t43_lead_hold_dec"] = _g(
                f"{_pfx}_pre_dec_win_won_r1_rate", 0.0
            ).clip(0.0, 1.0)
            df[f"{_pfx}_t43_dom_z"] = (
                _g(f"z_{_pfx}_pre_kd_rate", 0.0)
                + np.tanh(_g(f"{_pfx}_pre_ctrl_min_per15", 0.0) / 15.0 * 3.0)
                + np.tanh(_g(f"{_pfx}_pre_da_damage_margin", 0.0) * 0.05)
            )
        return df

    # Constructs 25+ tiers of model features from the leakage-cleaned DataFrame.
    # Each tier operates on the pre-fight stats written by fix_data_leakage():
    #   Tier 0  - Raw physical attribute diffs (height, reach, weight, age, ape index)
    #   Tier 1  - Pre-fight record/rate diffs (wins, streaks, finish rate, etc.)
    #   Tier 2  - Rolling 3/5 fight window diffs
    #   Tier 3  - ELO difference, ratio
    #   Tier 4  - Glicko-2 rating, RD, volatility
    #   Tier 5  - Weight-class Z-scores (compare fighter to peers in same division/year)
    #   Tier 6  - Common opponent features (mutual wins/losses vs shared opponents)
    #   Tier 7  - KMeans style-cluster assignment and cluster-vs-cluster win rates
    #   Tier 8  - Stance encoding (Orthodox/Southpaw/Switch/Open Stance)
    #   Tier 9-17 - Interaction features, polynomials, momentum, method-specific, SVD
    #   Tier 18-28 - Opponent-adjusted, career patterns, matchup exploitation, stat ratios
    #   Tier 43 - Method-head vulnerability + conversion (Decision/Finish, KO/Sub pathways)
    # All leaky in-fight columns (actual fight stats) are dropped before feature building.
    def build_all_features(self):
        print_section("BUILDING FEATURES (13+ TIERS)")
        df = self.df

        _LEAKY_COLS = [
            "r_sig_str",
            "b_sig_str",
            "sig_str_diff",
            "r_sig_str_att",
            "b_sig_str_att",
            "sig_str_att_diff",
            "r_sig_str_acc",
            "b_sig_str_acc",
            "sig_str_acc_diff",
            "r_str",
            "b_str",
            "str_diff",
            "r_str_att",
            "b_str_att",
            "str_att_diff",
            "r_str_acc",
            "b_str_acc",
            "str_acc_diff",
            "r_kd",
            "b_kd",
            "kd_diff",
            "r_head",
            "b_head",
            "head_diff",
            "r_body",
            "b_body",
            "body_diff",
            "r_leg",
            "b_leg",
            "leg_diff",
            "r_distance",
            "b_distance",
            "distance_diff",
            "r_clinch",
            "b_clinch",
            "clinch_diff",
            "r_ground",
            "b_ground",
            "ground_diff",
            "r_td",
            "b_td",
            "td_diff",
            "r_td_att",
            "b_td_att",
            "td_att_diff",
            "r_td_acc",
            "b_td_acc",
            "td_acc_diff",
            "r_sub_att",
            "b_sub_att",
            "sub_att_diff",
            "r_rev",
            "b_rev",
            "rev_diff",
            "r_ctrl_sec",
            "b_ctrl_sec",
            "ctrl_sec_diff",
            "total_fight_time_sec",
            "r_wins",
            "b_wins",
            "wins_diff",
            "r_losses",
            "b_losses",
            "losses_diff",
            "r_draws",
            "b_draws",
            "draws_diff",
            "r_win_loss_ratio",
            "b_win_loss_ratio",
            "win_loss_ratio_diff",
            "r_pro_SLpM",
            "b_pro_SLpM",
            "pro_SLpM_diff",
            "r_pro_sig_str_acc",
            "b_pro_sig_str_acc",
            "pro_sig_str_acc_diff",
            "r_pro_SApM",
            "b_pro_SApM",
            "pro_SApM_diff",
            "r_pro_str_def",
            "b_pro_str_def",
            "pro_str_def_diff",
            "r_pro_td_avg",
            "b_pro_td_avg",
            "pro_td_avg_diff",
            "r_pro_td_acc",
            "b_pro_td_acc",
            "pro_td_acc_diff",
            "r_pro_td_def",
            "b_pro_td_def",
            "pro_td_def_diff",
            "r_pro_sub_avg",
            "b_pro_sub_avg",
            "pro_sub_avg_diff",
        ]
        _leaky_present = [c for c in _LEAKY_COLS if c in df.columns]
        self.df = df = df.drop(columns=_leaky_present)
        print_step(f"Dropped {len(_leaky_present)} leaky in-fight columns.")

        # Drop all per-round columns (actual fight stats — leaky regardless of round number).
        _round_cols = [
            c for c in df.columns if any(f"_rd{n}_" in c for n in range(1, 6))
        ]
        if _round_cols:
            self.df = df = df.drop(columns=_round_cols)
            print_step(f"Dropped {len(_round_cols)} per-round leaky columns.")

        self._log("Tier 0: Raw column differences...")
        raw_pairs = [
            ("r_height", "b_height"),
            ("r_reach", "b_reach"),
            ("r_weight", "b_weight"),
            ("r_age_at_event", "b_age_at_event"),
            ("r_ape_index", "b_ape_index"),
        ]
        for rc, bc in raw_pairs:
            if rc in df.columns and bc in df.columns:
                df[f"diff_{rc[2:]}"] = pd.to_numeric(df[rc], errors="coerce").fillna(
                    0
                ) - pd.to_numeric(df[bc], errors="coerce").fillna(0)

        self._log("Tier 1: Pre-fight stat diffs...")
        pre_pairs = [
            ("r_pre_wins", "b_pre_wins"),
            ("r_pre_losses", "b_pre_losses"),
            ("r_pre_ko_wins", "b_pre_ko_wins"),
            ("r_pre_sub_wins", "b_pre_sub_wins"),
            ("r_pre_dec_wins", "b_pre_dec_wins"),
            ("r_pre_total_fights", "b_pre_total_fights"),
            ("r_pre_finish_rate", "b_pre_finish_rate"),
            ("r_pre_win_streak", "b_pre_win_streak"),
            ("r_pre_loss_streak", "b_pre_loss_streak"),
            ("r_pre_title_fights", "b_pre_title_fights"),
            ("r_pre_title_wins", "b_pre_title_wins"),
            ("r_pre_avg_fight_time", "b_pre_avg_fight_time"),
            ("r_pre_sig_str_acc", "b_pre_sig_str_acc"),
            ("r_pre_td_acc", "b_pre_td_acc"),
            ("r_pre_sub_att_rate", "b_pre_sub_att_rate"),
            ("r_pre_kd_rate", "b_pre_kd_rate"),
            ("r_pre_ctrl_avg", "b_pre_ctrl_avg"),
            ("r_pre_SLpM", "b_pre_SLpM"),
            ("r_pre_SApM", "b_pre_SApM"),
            ("r_pre_td_avg", "b_pre_td_avg"),
        ]
        for rc, bc in pre_pairs:
            if rc in df.columns and bc in df.columns:
                df[f"diff_{rc[6:]}"] = df[rc].fillna(0) - df[bc].fillna(0)

        self._log("Tier 2: Rolling window diffs...")
        rolling_pairs = [
            ("r_rolling3_wins", "b_rolling3_wins"),
            ("r_rolling3_sig_str", "b_rolling3_sig_str"),
            ("r_rolling3_td", "b_rolling3_td"),
            ("r_rolling3_kd", "b_rolling3_kd"),
            ("r_rolling3_sub_att", "b_rolling3_sub_att"),
            ("r_rolling5_wins", "b_rolling5_wins"),
            ("r_rolling5_sig_str", "b_rolling5_sig_str"),
            ("r_rolling5_td", "b_rolling5_td"),
            ("r_rolling5_kd", "b_rolling5_kd"),
        ]
        for rc, bc in rolling_pairs:
            if rc in df.columns and bc in df.columns:
                df[f"diff_{rc[2:]}"] = df[rc].fillna(0) - df[bc].fillna(0)

        self._log("Tier 3: ELO features...")
        if "r_elo_pre_fight" in df.columns and "b_elo_pre_fight" in df.columns:
            df["elo_diff"] = df["r_elo_pre_fight"] - df["b_elo_pre_fight"]
            df["elo_r"] = df["r_elo_pre_fight"]
            df["elo_b"] = df["b_elo_pre_fight"]
            df["elo_ratio"] = df["r_elo_pre_fight"] / \
                (df["b_elo_pre_fight"] + 1e-6)
        else:
            df["elo_diff"] = 0.0
            df["elo_r"] = 1500.0
            df["elo_b"] = 1500.0
            df["elo_ratio"] = 1.0

        self._log("Tier 4: Glicko-2 features (pre-fight snapshots)...")
        if "r_glicko_pre_r" in df.columns and "b_glicko_pre_r" in df.columns:
            df["r_glicko_r"] = df["r_glicko_pre_r"]
            df["r_glicko_rd"] = df["r_glicko_pre_rd"]
            df["r_glicko_vol"] = df["r_glicko_pre_vol"]
            df["b_glicko_r"] = df["b_glicko_pre_r"]
            df["b_glicko_rd"] = df["b_glicko_pre_rd"]
            df["b_glicko_vol"] = df["b_glicko_pre_vol"]
        else:
            r_glicko_r, r_glicko_rd, r_glicko_vol = [], [], []
            b_glicko_r, b_glicko_rd, b_glicko_vol = [], [], []
            for _, row in df.iterrows():
                rg = self.feature_engineer.glicko2_get(
                    str(row.get("r_fighter", "")))
                bg = self.feature_engineer.glicko2_get(
                    str(row.get("b_fighter", "")))
                r_glicko_r.append(rg[0])
                r_glicko_rd.append(rg[1])
                r_glicko_vol.append(rg[2])
                b_glicko_r.append(bg[0])
                b_glicko_rd.append(bg[1])
                b_glicko_vol.append(bg[2])
            df["r_glicko_r"] = r_glicko_r
            df["r_glicko_rd"] = r_glicko_rd
            df["r_glicko_vol"] = r_glicko_vol
            df["b_glicko_r"] = b_glicko_r
            df["b_glicko_rd"] = b_glicko_rd
            df["b_glicko_vol"] = b_glicko_vol
        df["glicko_diff"] = df["r_glicko_r"] - df["b_glicko_r"]
        df["glicko_rd_diff"] = df["r_glicko_rd"] - df["b_glicko_rd"]

        self._log("Tier 5: Weight-class Z-scores...")
        fe = self.feature_engineer
        z_feats = [
            "r_pre_SLpM",
            "r_pre_SApM",
            "r_pre_sig_str_acc",
            "r_pre_td_avg",
            "r_pre_sub_att_rate",
            "r_pre_kd_rate",
            "b_pre_SLpM",
            "b_pre_SApM",
            "b_pre_sig_str_acc",
            "b_pre_td_avg",
            "b_pre_sub_att_rate",
            "b_pre_kd_rate",
        ]

        fe.weight_class_stats = defaultdict(_defaultdict_list)
        for col in z_feats:
            df[f"z_{col}"] = 0.0
        for idx, row in df.sort_values("event_date", na_position="first").iterrows():
            wc = str(row.get("weight_class", ""))
            yr = row["event_date"].year if pd.notna(
                row.get("event_date")) else 2000
            stats_dict = {}
            for feat in z_feats:
                if feat in df.columns:
                    v = row.get(feat, 0)
                    try:
                        v = float(v)
                        if not math.isnan(v):
                            df.at[idx, f"z_{feat}"] = fe.get_z_score(
                                wc, yr, feat, v)
                            stats_dict[feat] = v
                    except (TypeError, ValueError):
                        pass
            fe.update_weight_class_stats(wc, yr, stats_dict)

        self._log("Tier 6: Common opponent features...")

        fe.fighter_opponents = defaultdict(set)
        fe.fight_outcomes = {}
        for col in [
            "n_common_opponents",
            "r_wins_vs_common",
            "b_wins_vs_common",
            "common_opp_edge",
        ]:
            df[col] = 0.0
        for idx, row in df.sort_values("event_date", na_position="first").iterrows():
            r = str(row.get("r_fighter", ""))
            b = str(row.get("b_fighter", ""))
            winner = str(row.get("winner", ""))
            feat = fe.get_common_opponent_features(r, b)
            df.at[idx, "n_common_opponents"] = feat["n_common_opponents"]
            df.at[idx, "r_wins_vs_common"] = feat["r_wins_vs_common"]
            df.at[idx, "b_wins_vs_common"] = feat["b_wins_vs_common"]
            df.at[idx, "common_opp_edge"] = feat["common_opp_edge"]
            fe.update_common_opponents(r, b, winner)

        self._log("Tier 7: Style cluster features...")

        df_sorted_dates = df.sort_values("event_date", na_position="first")
        _cluster_fit_cutoff = int(len(df_sorted_dates) * 0.80)
        df_cluster_fit = df_sorted_dates.iloc[:_cluster_fit_cutoff]
        fighter_style = {}
        for f in self.all_fighters:
            r_rows = df_cluster_fit[df_cluster_fit["r_fighter"] == f]
            b_rows = df_cluster_fit[df_cluster_fit["b_fighter"] == f]
            slpm = 0.0
            sapm = 0.0
            td = 0.0
            sub = 0.0
            finish = 0.0
            rd1_slpm = 0.0
            cardio = 0.0
            rd1_td = 0.0
            cnt = 0
            for _, row in r_rows.iterrows():
                slpm += float(row.get("r_pre_SLpM", 0) or 0)
                sapm += float(row.get("r_pre_SApM", 0) or 0)
                td += float(row.get("r_pre_td_avg", 0) or 0)
                sub += float(row.get("r_pre_sub_att_rate", 0) or 0)
                finish += float(row.get("r_pre_finish_rate", 0) or 0)
                rd1_slpm += float(row.get("r_pre_rd1_slpm_avg", 0) or 0)
                cardio += float(row.get("r_pre_cardio_index", 0.75) or 0.75)
                rd1_td += float(row.get("r_pre_rd1_td_rate", 0) or 0)
                cnt += 1
            for _, row in b_rows.iterrows():
                slpm += float(row.get("b_pre_SLpM", 0) or 0)
                sapm += float(row.get("b_pre_SApM", 0) or 0)
                td += float(row.get("b_pre_td_avg", 0) or 0)
                sub += float(row.get("b_pre_sub_att_rate", 0) or 0)
                finish += float(row.get("b_pre_finish_rate", 0) or 0)
                rd1_slpm += float(row.get("b_pre_rd1_slpm_avg", 0) or 0)
                cardio += float(row.get("b_pre_cardio_index", 0.75) or 0.75)
                rd1_td += float(row.get("b_pre_rd1_td_rate", 0) or 0)
                cnt += 1
            if cnt > 0:
                fighter_style[f] = {
                    "SLpM": slpm / cnt,
                    "rd1_slpm_avg": rd1_slpm / cnt,
                    "cardio_index": cardio / cnt,
                    "TD": td / cnt,
                    "rd1_td_rate": rd1_td / cnt,
                    "Sub": sub / cnt,
                    "Finish": finish / cnt,
                }
        fe.fit_clusters(fighter_style)

        fe.style_matchup_finish = defaultdict(_defaultdict_list)
        df["cluster_pair_finish_rate"] = 0.5

        _style_snap = {}
        # Order must match FeatureEngineer.style_features exactly.
        _r_cluster_cols = [
            "r_pre_SLpM",
            "r_pre_rd1_slpm_avg",
            "r_pre_cardio_index",
            "r_pre_td_avg",
            "r_pre_rd1_td_rate",
            "r_pre_sub_att_rate",
            "r_pre_finish_rate",
        ]
        _b_cluster_cols = [
            "b_pre_SLpM",
            "b_pre_rd1_slpm_avg",
            "b_pre_cardio_index",
            "b_pre_td_avg",
            "b_pre_rd1_td_rate",
            "b_pre_sub_att_rate",
            "b_pre_finish_rate",
        ]
        for idx, row in df.sort_values("event_date", na_position="first").iterrows():
            r = str(row.get("r_fighter", ""))
            b = str(row.get("b_fighter", ""))
            winner = str(row.get("winner", ""))
            method = str(row.get("method", ""))
            _fight_was_finish = method in ("KO/TKO", "Submission")
            rc = fe.predict_cluster(
                [float(row.get(c, 0) or 0) for c in _r_cluster_cols]
            )
            bc = fe.predict_cluster(
                [float(row.get(c, 0) or 0) for c in _b_cluster_cols]
            )
            if rc >= 0 and bc >= 0:
                df.at[idx, "cluster_pair_finish_rate"] = (
                    fe.get_style_matchup_finish_rate(rc, bc)
                )
                mf = fe.get_style_matchup_features(rc, bc)
                _style_snap[idx] = (
                    rc,
                    bc,
                    mf["r_style_win_vs_opp_cluster"],
                    mf["b_style_win_vs_opp_cluster"],
                    mf["style_matchup_edge"],
                )

                fe.update_style_performance(rc, bc, winner == "Red")
                fe.update_style_performance(bc, rc, winner == "Blue")
                fe.update_style_matchup_finish(rc, bc, _fight_was_finish)
                fe.update_style_matchup_finish(bc, rc, _fight_was_finish)
            else:
                _style_snap[idx] = (rc, bc, 0.5, 0.5, 0.0)

        r_cluster, b_cluster, style_edge = [], [], []
        r_style_win, b_style_win = [], []
        for idx, row in df.iterrows():
            snap = _style_snap.get(idx, (-1, -1, 0.5, 0.5, 0.0))
            rc, bc, rw, bw, edge = snap
            r_cluster.append(rc)
            b_cluster.append(bc)
            r_style_win.append(rw)
            b_style_win.append(bw)
            style_edge.append(edge)
        df["r_style_cluster"] = r_cluster
        df["b_style_cluster"] = b_cluster
        df["style_matchup_edge"] = style_edge
        df["r_style_win_vs_cluster"] = r_style_win
        df["b_style_win_vs_cluster"] = b_style_win

        self._log("Tier 8: Stance encoding...")
        stance_map = {"Orthodox": 0, "Southpaw": 1,
                      "Switch": 2, "Open Stance": 3}
        for col in ["r_stance", "b_stance"]:
            if col in df.columns:
                df[f"{col}_enc"] = df[col].map(stance_map).fillna(-1)
        if "r_stance_enc" in df.columns and "b_stance_enc" in df.columns:
            df["stance_matchup"] = (
                df["r_stance_enc"].astype(
                    str) + "_" + df["b_stance_enc"].astype(str)
            )
            df["same_stance"] = (df["r_stance_enc"] ==
                                 df["b_stance_enc"]).astype(int)

        self._log("Tier 9: Interaction features...")
        if "elo_diff" in df.columns and "diff_finish_rate" in df.columns:
            df["elo_x_finish_rate"] = df["elo_diff"] * df["diff_finish_rate"]
        if "diff_pre_win_streak" in df.columns and "diff_pre_finish_rate" in df.columns:
            df["streak_x_finish"] = (
                df["diff_pre_win_streak"] * df["diff_pre_finish_rate"]
            )
        if "diff_pre_SLpM" in df.columns and "diff_pre_SApM" in df.columns:
            df["striking_exchange"] = df["diff_pre_SLpM"] - df["diff_pre_SApM"]
        if "diff_pre_td_avg" in df.columns and "diff_pre_td_acc" in df.columns:
            df["td_efficiency"] = df["diff_pre_td_avg"] * df["diff_pre_td_acc"]
        if "diff_pre_sig_str_acc" in df.columns and "diff_pre_ctrl_avg" in df.columns:
            df["control_accuracy"] = (
                df["diff_pre_sig_str_acc"] * df["diff_pre_ctrl_avg"]
            )

        self._log("Tier 10: Polynomial features...")

        if "r_pre_wins" in df.columns and "r_pre_losses" in df.columns:
            df["diff_win_loss_ratio"] = df["r_pre_wins"] / (
                df["r_pre_losses"].clip(lower=0) + 1.0
            ) - df["b_pre_wins"] / (df["b_pre_losses"].clip(lower=0) + 1.0)
        poly_cols = ["elo_diff", "glicko_diff", "diff_win_loss_ratio"]
        for col in poly_cols:
            if col in df.columns:
                df[f"{col}_sq"] = df[col] ** 2
                df[f"{col}_abs"] = df[col].abs()

        self._log("Tier 11: Momentum indicators...")
        if "r_rolling3_wins" in df.columns and "b_rolling3_wins" in df.columns:
            df["momentum_diff_3"] = df["r_rolling3_wins"] - df["b_rolling3_wins"]
        if "r_rolling5_wins" in df.columns and "b_rolling5_wins" in df.columns:
            df["momentum_diff_5"] = df["r_rolling5_wins"] - df["b_rolling5_wins"]
        if "r_pre_win_streak" in df.columns and "b_pre_win_streak" in df.columns:
            df["streak_differential"] = df["r_pre_win_streak"] - \
                df["b_pre_win_streak"]

        self._log("Tier 12: Method-specific features...")
        if "r_pre_ko_wins" in df.columns and "b_pre_ko_wins" in df.columns:
            df["ko_threat_diff"] = df["r_pre_ko_wins"] - df["b_pre_ko_wins"]
        if "r_pre_sub_wins" in df.columns and "b_pre_sub_wins" in df.columns:
            df["sub_threat_diff"] = df["r_pre_sub_wins"] - df["b_pre_sub_wins"]
        if "r_pre_dec_wins" in df.columns and "b_pre_dec_wins" in df.columns:
            df["dec_tendency_diff"] = df["r_pre_dec_wins"] - df["b_pre_dec_wins"]
        if "r_pre_finish_rate" in df.columns and "b_pre_finish_rate" in df.columns:
            df["r_finishing_tendency"] = df["r_pre_finish_rate"]
            df["b_finishing_tendency"] = df["b_pre_finish_rate"]
            df["finishing_matchup"] = df["r_pre_finish_rate"] * \
                df["b_pre_finish_rate"]

        print_step("Career pattern features...")
        for prefix in ["r", "b"]:
            dec_col = f"{prefix}_pre_dec_wins"
            ko_col = f"{prefix}_pre_ko_wins"
            sub_col = f"{prefix}_pre_sub_wins"
            total_col = f"{prefix}_pre_total_fights"

            if all(c in df.columns for c in [dec_col, ko_col, sub_col, total_col]):
                denom = df[total_col].clip(lower=1)
                df[f"{prefix}_decision_win_rate"] = df[dec_col] / denom
                df[f"{prefix}_ko_win_rate"] = df[ko_col] / denom
                df[f"{prefix}_sub_win_rate"] = df[sub_col] / denom
                df[f"{prefix}_finish_rate"] = (
                    df[ko_col] + df[sub_col]) / denom

            dec_loss_col = f"{prefix}_pre_dec_losses"
            ko_loss_col = f"{prefix}_pre_ko_losses"
            sub_loss_col = f"{prefix}_pre_sub_losses"
            if all(c in df.columns for c in [dec_loss_col, ko_loss_col, sub_loss_col]):
                loss_denom = (
                    df[dec_loss_col] + df[ko_loss_col] + df[sub_loss_col]
                ).clip(lower=1)
                df[f"{prefix}_ko_loss_rate"] = df[ko_loss_col] / loss_denom
                df[f"{prefix}_sub_loss_rate"] = df[sub_loss_col] / loss_denom
                df[f"{prefix}_dec_loss_rate"] = df[dec_loss_col] / loss_denom

            if f"{prefix}_pre_title_fights" in df.columns:
                df[f"{prefix}_title_fight_exp"] = df[f"{prefix}_pre_title_fights"]

            if f"{prefix}_pre_five_round_fights" in df.columns:
                df[f"{prefix}_main_event_exp"] = df[f"{prefix}_pre_five_round_fights"]

        for feat in [
            "decision_win_rate",
            "ko_win_rate",
            "sub_win_rate",
            "finish_rate",
            "title_fight_exp",
            "main_event_exp",
        ]:
            r_col = f"r_{feat}"
            b_col = f"b_{feat}"
            if r_col in df.columns and b_col in df.columns:
                df[f"{feat}_diff"] = df[r_col] - df[b_col]
        # decision_win_rate covers all decision sub-types (unanimous + majority + split) intentionally.

        # Grappling pathway features (per-side, matchup, composites, diffs)
        # are all computed in _recompute_derived_features so they survive
        # corner-swap augmentation.

        self._log("Tier 12b: Advanced combat metrics...")

        # ring_rust discretization removed — raw r_days_since_last / b_days_since_last
        # already exist as features and trees learn their own optimal splits.
        if "r_days_since_last" in df.columns and "b_days_since_last" in df.columns:
            df["days_since_last_diff"] = (
                df["r_days_since_last"].fillna(365) - df["b_days_since_last"].fillna(365)
            )

        # weight_class_ko_factor static map removed — WC effects are captured
        # by z-scores, style clusters, and gender_enc. Trees learn WC splits directly.

        r_slpm = df.get(
            "r_pre_SLpM", df.get("r_pro_SLpM", pd.Series(0, index=df.index))
        ).fillna(0)
        b_slpm = df.get(
            "b_pre_SLpM", df.get("b_pro_SLpM", pd.Series(0, index=df.index))
        ).fillna(0)
        # Removed: style_clash_severity, upset_potential, power_vs_technique,
        # championship_pressure, finishing_pressure_diff, overactive, peak_score.
        # Their raw components (diff_pre_SLpM, diff_pre_td_avg, diff_pre_sig_str_acc,
        # ko_win_rate_diff, sub_win_rate_diff, days_since_last, age_at_event, etc.)
        # are already features — trees learn the optimal interactions directly.

        r_acc = df.get("r_pre_sig_str_acc", pd.Series(0.45, index=df.index)).fillna(
            0.45
        )
        b_acc = df.get("b_pre_sig_str_acc", pd.Series(0.45, index=df.index)).fillna(
            0.45
        )

        r_clinch_pct = df.get("r_pre_clinch_pct", pd.Series(0, index=df.index)).fillna(
            0
        )
        b_clinch_pct = df.get("b_pre_clinch_pct", pd.Series(0, index=df.index)).fillna(
            0
        )
        df["r_clinch_effectiveness"] = r_clinch_pct * r_slpm * r_acc
        df["b_clinch_effectiveness"] = b_clinch_pct * b_slpm * b_acc
        df["clinch_effectiveness_diff"] = (
            df["r_clinch_effectiveness"] - df["b_clinch_effectiveness"]
        )

        r_dec = df.get("r_decision_win_rate", pd.Series(
            0, index=df.index)).fillna(0)
        b_dec = df.get("b_decision_win_rate", pd.Series(
            0, index=df.index)).fillna(0)
        total_rounds_col = df.get("total_rounds", pd.Series(3, index=df.index)).fillna(
            3
        )
        df["five_round_cardio_advantage"] = (
            r_dec - b_dec) * (total_rounds_col / 3)

        r_fights = df.get("r_pre_total_fights", pd.Series(10, index=df.index)).fillna(
            10
        )
        b_fights = df.get("b_pre_total_fights", pd.Series(10, index=df.index)).fillna(
            10
        )
        r_kd_absorbed = df.get(
            "r_pre_kd_absorbed", pd.Series(0, index=df.index)
        ).fillna(0)
        b_kd_absorbed = df.get(
            "b_pre_kd_absorbed", pd.Series(0, index=df.index)
        ).fillna(0)
        r_fights_safe = r_fights.clip(lower=1)
        b_fights_safe = b_fights.clip(lower=1)
        df["r_chin_deterioration"] = r_kd_absorbed / r_fights_safe
        df["b_chin_deterioration"] = b_kd_absorbed / b_fights_safe
        df["chin_deterioration_diff"] = (
            df["r_chin_deterioration"] - df["b_chin_deterioration"]
        )

        self._log("Tier 12c: Opponent quality & trajectory slopes...")
        if "r_avg_opp_elo_L5" in df.columns and "b_avg_opp_elo_L5" in df.columns:
            df["opp_quality_diff"] = df["r_avg_opp_elo_L5"] - \
                df["b_avg_opp_elo_L5"]
        if "r_trajectory_3" in df.columns and "b_trajectory_3" in df.columns:
            df["trajectory_diff"] = df["r_trajectory_3"] - df["b_trajectory_3"]

        # Removed: uncertainty_score — its components (SLpM, td_avg, total_fights,
        # elo_diff) are all already features; the 0.25-weight composite adds noise.

        self._log("Tier 12f: Decision-type features (split vs. dominant wins)...")
        for prefix in ["r", "b"]:
            dec_col = f"{prefix}_pre_dec_wins"
            unani_col = f"{prefix}_pre_unanimous_wins"
            split_col = f"{prefix}_pre_split_wins"
            total_col = f"{prefix}_pre_total_fights"
            if all(c in df.columns for c in [unani_col, split_col, total_col]):
                denom = df[total_col].clip(lower=1)
                df[f"{prefix}_dominant_win_rate"] = df[unani_col] / denom
                df[f"{prefix}_split_win_rate"] = df[split_col] / denom
                split_loss_col = f"{prefix}_pre_split_losses"
                if split_loss_col in df.columns:
                    df[f"{prefix}_split_loss_rate"] = df[split_loss_col] / denom
        for feat in ["dominant_win_rate", "split_win_rate", "split_loss_rate"]:
            r_col, b_col = f"r_{feat}", f"b_{feat}"
            if r_col in df.columns and b_col in df.columns:
                df[f"{feat}_diff"] = df[r_col].fillna(0) - df[b_col].fillna(0)
        # A fighter whose wins are predominantly split decisions is more "beatable" —
        # they consistently win close fights, meaning opponents are competitive with them.
        if "r_split_win_rate" in df.columns and "b_split_win_rate" in df.columns:
            df["contested_fighter_diff"] = df["r_split_win_rate"].fillna(0) - df[
                "b_split_win_rate"
            ].fillna(0)
        # Dominant finisher vs. scraper matchup: positive when red is the decisive winner
        if "r_dominant_win_rate" in df.columns and "b_dominant_win_rate" in df.columns:
            df["dominance_style_diff"] = df["r_dominant_win_rate"].fillna(0) - df[
                "b_dominant_win_rate"
            ].fillna(0)

        self._log(
            "Tier 12g: Per-round stats features (R1 aggression, cardio, late grappling)..."
        )
        r_rd1_slpm = df.get(
            "r_pre_rd1_slpm_avg", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        b_rd1_slpm = df.get(
            "b_pre_rd1_slpm_avg", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        r_rd1_kd = df.get("r_pre_rd1_kd_rate", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        b_rd1_kd = df.get("b_pre_rd1_kd_rate", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        r_rd1_td = df.get("r_pre_rd1_td_rate", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        b_rd1_td = df.get("b_pre_rd1_td_rate", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        r_cardio = df.get("r_pre_cardio_index", pd.Series(0.75, index=df.index)).fillna(
            0.75
        )
        b_cardio = df.get("b_pre_cardio_index", pd.Series(0.75, index=df.index)).fillna(
            0.75
        )
        r_late_td = df.get("r_pre_late_td_rate", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        b_late_td = df.get("b_pre_late_td_rate", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        r_rd1_head = df.get(
            "r_pre_rd1_head_pct", pd.Series(0.45, index=df.index)
        ).fillna(0.45)
        b_rd1_head = df.get(
            "b_pre_rd1_head_pct", pd.Series(0.45, index=df.index)
        ).fillna(0.45)

        df["rd1_slpm_diff"] = r_rd1_slpm - b_rd1_slpm
        df["rd1_kd_diff"] = r_rd1_kd - b_rd1_kd
        df["rd1_td_diff"] = r_rd1_td - b_rd1_td
        df["cardio_diff"] = r_cardio - b_cardio
        df["late_td_diff"] = r_late_td - b_late_td
        df["rd1_head_pct_diff"] = r_rd1_head - b_rd1_head

        # Fast-starter vs. good-cardio matchup: red fighter starts explosively but blue holds up late.
        # Positive = red is the early blitzer facing a durable fader; negative = opposite.
        df["rd1_vs_late_style_clash"] = (r_rd1_slpm * b_cardio) - (
            b_rd1_slpm * r_cardio
        )

        # R1 KD threat vs. opponent's chin deterioration: explosive puncher vs. already-wobbled chin
        r_kd_abs = df.get("r_pre_kd_absorbed", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        b_kd_abs = df.get("b_pre_kd_absorbed", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        df["r_rd1_power_vs_chin"] = r_rd1_kd * (
            b_kd_abs
            / (
                df.get("b_pre_total_fights", pd.Series(1.0, index=df.index))
                .fillna(1.0)
                .clip(lower=1)
            )
        )
        df["b_rd1_power_vs_chin"] = b_rd1_kd * (
            r_kd_abs
            / (
                df.get("r_pre_total_fights", pd.Series(1.0, index=df.index))
                .fillna(1.0)
                .clip(lower=1)
            )
        )
        df["rd1_power_vs_chin_diff"] = (
            df["r_rd1_power_vs_chin"] - df["b_rd1_power_vs_chin"]
        )

        # Cardio-adjusted five-round advantage: fighters with good cardio indices gain more in long fights
        total_rds_12g = df.get("total_rounds_num", pd.Series(3, index=df.index)).fillna(
            3
        )
        df["cardio_x_five_rounds"] = df["cardio_diff"] * (total_rds_12g / 3.0)

        # Late-round grappling exploitation: wrestler who improves late vs. striker whose output decays
        df["r_late_grappler_advantage"] = r_late_td * (1.0 - b_cardio)
        df["b_late_grappler_advantage"] = b_late_td * (1.0 - r_cardio)
        df["late_grappler_advantage_diff"] = (
            df["r_late_grappler_advantage"] - df["b_late_grappler_advantage"]
        )

        # ── Tier 12h: Deep per-round evolution features ─────────────────────────────
        self._log(
            "Tier 12h: Deep per-round evolution features (slopes, fade, championship rounds)..."
        )

        def _get(col, default=0.0):
            return df.get(col, pd.Series(default, index=df.index)).fillna(default)

        r_output_slope = _get("r_pre_output_slope")
        b_output_slope = _get("b_pre_output_slope")
        r_acc_fade = _get("r_pre_acc_fade")
        b_acc_fade = _get("b_pre_acc_fade")
        r_r45_slpm = _get("r_pre_r45_slpm")
        b_r45_slpm = _get("b_pre_r45_slpm")
        r_r45_ratio = _get("r_pre_r45_vs_r1_ratio")
        b_r45_ratio = _get("b_pre_r45_vs_r1_ratio")
        r_body_esc = _get("r_pre_body_escalation")
        b_body_esc = _get("b_pre_body_escalation")
        r_td_slope = _get("r_pre_td_slope")
        b_td_slope = _get("b_pre_td_slope")
        r_late_kd = _get("r_pre_late_kd_rate")
        b_late_kd = _get("b_pre_late_kd_rate")
        r_kd_rnd_avg = _get("r_pre_kd_round_avg", 2.0)
        b_kd_rnd_avg = _get("b_pre_kd_round_avg", 2.0)
        r_abs_slope = _get("r_pre_absorption_slope")
        b_abs_slope = _get("b_pre_absorption_slope")
        r_late_abs = _get("r_pre_late_absorbed")
        b_late_abs = _get("b_pre_late_absorbed")
        r_grd_esc = _get("r_pre_ground_escalation")
        b_grd_esc = _get("b_pre_ground_escalation")
        r_late_sub = _get("r_pre_late_sub_rate")
        b_late_sub = _get("b_pre_late_sub_rate")
        r_ctrl_slope = _get("r_pre_ctrl_slope")
        b_ctrl_slope = _get("b_pre_ctrl_slope")
        r_r1_acc = _get("r_pre_r1_acc", 0.45)
        b_r1_acc = _get("b_pre_r1_acc", 0.45)
        r_late_acc = _get("r_pre_late_acc", 0.45)
        b_late_acc = _get("b_pre_late_acc", 0.45)

        # Simple diffs
        df["output_slope_diff"] = r_output_slope - b_output_slope
        df["acc_fade_diff"] = r_acc_fade - b_acc_fade
        df["r45_slpm_diff"] = r_r45_slpm - b_r45_slpm
        df["r45_ratio_diff"] = r_r45_ratio - b_r45_ratio
        df["body_escalation_diff"] = r_body_esc - b_body_esc
        df["td_slope_diff"] = r_td_slope - b_td_slope
        df["late_kd_diff"] = r_late_kd - b_late_kd
        df["kd_round_avg_diff"] = r_kd_rnd_avg - b_kd_rnd_avg
        df["absorption_slope_diff"] = r_abs_slope - b_abs_slope
        df["late_absorbed_diff"] = r_late_abs - b_late_abs
        df["ground_escalation_diff"] = r_grd_esc - b_grd_esc
        df["late_sub_diff"] = r_late_sub - b_late_sub
        df["ctrl_slope_diff"] = r_ctrl_slope - b_ctrl_slope

        # R1 vs late accuracy diff (measures mid-fight accuracy drift per fighter)
        df["r_acc_drift"] = r_r1_acc - r_late_acc
        df["b_acc_drift"] = b_r1_acc - b_late_acc
        df["acc_drift_diff"] = df["r_acc_drift"] - df["b_acc_drift"]

        # Interaction: output slope amplified in 5-round bouts
        total_rds_12h = df.get(
            "total_rounds", pd.Series(3, index=df.index)).fillna(3)
        df["output_slope_x_five_rounds"] = df["output_slope_diff"] * (
            total_rds_12h / 3.0
        )

        # Interaction: accuracy fade × ELO (better fighters fade less — catches cardio edge)
        r_elo_12h = _get("r_elo_pre_fight", 1500.0)
        b_elo_12h = _get("b_elo_pre_fight", 1500.0)
        elo_diff_12h = (r_elo_12h - b_elo_12h) / 400.0
        df["acc_fade_x_elo"] = df["acc_fade_diff"] * elo_diff_12h

        # Interaction: late grappler v2 (TD slope + ctrl slope vs opponent absorption slope)
        df["r_late_grappler_v2"] = (r_td_slope + r_ctrl_slope) * b_abs_slope.clip(
            lower=0
        )
        df["b_late_grappler_v2"] = (b_td_slope + b_ctrl_slope) * r_abs_slope.clip(
            lower=0
        )
        df["late_grappler_v2_diff"] = (
            df["r_late_grappler_v2"] - df["b_late_grappler_v2"]
        )

        # Interaction: body escalation × accuracy (body punching improves when acc holds up)
        df["body_x_acc_diff"] = (
            r_body_esc * r_late_acc) - (b_body_esc * b_late_acc)

        # Late KD threat: KDs land later in fight (punisher who waits for opening)
        df["late_kd_threat_diff"] = (r_late_kd * r_kd_rnd_avg) - (
            b_late_kd * b_kd_rnd_avg
        )

        # Championship cardio: R4/R5 pace advantage in long bouts
        df["r45_cardio_diff"] = r_r45_ratio - b_r45_ratio

        # Ground escalation vs absorption slope (grappler exploiting a tiring shell)
        df["grapple_exploitation_diff"] = (r_grd_esc + r_td_slope) * b_abs_slope.clip(
            lower=0
        ) - (b_grd_esc + b_td_slope) * r_abs_slope.clip(lower=0)

        # ── Tier 12i: Leg kicks, clinch, R2 adjustment, volume, zone, momentum ─────
        self._log(
            "Tier 12i: Leg, clinch, R2 adjustment, volume ratio, zone shift, momentum features..."
        )

        def _g(col, default=0.0):
            return df.get(col, pd.Series(default, index=df.index)).fillna(default)

        # Leg kick features
        r_r1_leg_rate = _g("r_pre_r1_leg_rate")
        b_r1_leg_rate = _g("b_pre_r1_leg_rate")
        r_leg_esc = _g("r_pre_leg_escalation")
        b_leg_esc = _g("b_pre_leg_escalation")
        r_leg_vs_head = _g("r_pre_leg_vs_head")
        b_leg_vs_head = _g("b_pre_leg_vs_head")
        r_leg_pct = _g("r_pre_leg_pct")
        b_leg_pct = _g("b_pre_leg_pct")

        df["r1_leg_rate_diff"] = r_r1_leg_rate - b_r1_leg_rate
        df["leg_escalation_diff"] = r_leg_esc - b_leg_esc
        df["leg_vs_head_diff"] = r_leg_vs_head - b_leg_vs_head
        df["leg_pct_diff"] = r_leg_pct - b_leg_pct

        # Kicker vs grappler clash: high leg pct attacker vs high TD opponent
        r_td_avg_12i = _g("r_pre_td_avg", 0.0)
        b_td_avg_12i = _g("b_pre_td_avg", 0.0)
        df["kicker_vs_wrestler"] = (r_leg_pct - b_leg_pct) * (
            b_td_avg_12i - r_td_avg_12i
        )

        # Leg escalation × 5-round bout (leg kickers gain more in longer fights)
        total_rds_12i = df.get(
            "total_rounds", pd.Series(3, index=df.index)).fillna(3)
        df["leg_esc_x_five_rounds"] = df["leg_escalation_diff"] * \
            (total_rds_12i / 3.0)

        # Clinch features
        r_r1_clinch = _g("r_pre_r1_clinch_rate")
        b_r1_clinch = _g("b_pre_r1_clinch_rate")
        r_clinch_esc = _g("r_pre_clinch_escalation")
        b_clinch_esc = _g("b_pre_clinch_escalation")
        r_clinch_pct = _g("r_pre_clinch_pct_r1")
        b_clinch_pct = _g("b_pre_clinch_pct_r1")

        df["r1_clinch_rate_diff"] = r_r1_clinch - b_r1_clinch
        df["clinch_escalation_diff"] = r_clinch_esc - b_clinch_esc
        df["clinch_pct_diff"] = r_clinch_pct - b_clinch_pct

        # Clinch escalation × opponent leg pct: clinch fighter neutralising kicker
        df["clinch_vs_kicker"] = df["clinch_escalation_diff"] * \
            (b_leg_pct - r_leg_pct)

        # Round 2 adjustment signals
        r_r1_to_r2 = _g("r_pre_r1_to_r2_output")
        b_r1_to_r2 = _g("b_pre_r1_to_r2_output")
        r_r1_to_r2_td = _g("r_pre_r1_to_r2_td")
        b_r1_to_r2_td = _g("b_pre_r1_to_r2_td")
        r_r2_body = _g("r_pre_r2_body_rate")
        b_r2_body = _g("b_pre_r2_body_rate")

        df["r1_to_r2_output_diff"] = r_r1_to_r2 - b_r1_to_r2
        df["r1_to_r2_td_diff"] = r_r1_to_r2_td - b_r1_to_r2_td
        df["r2_body_rate_diff"] = r_r2_body - b_r2_body

        # Adaptability: ramps output in R2 after losing R1
        r_r1_win = _g("r_pre_r1_win_rate", 0.5)
        b_r1_win = _g("b_pre_r1_win_rate", 0.5)
        df["r_adaptability"] = r_r1_to_r2 * \
            (1.0 - r_r1_win)  # escalates when behind
        df["b_adaptability"] = b_r1_to_r2 * (1.0 - b_r1_win)
        df["adaptability_diff"] = df["r_adaptability"] - df["b_adaptability"]

        # Volume ratio features
        r_vol_ratio = _g("r_pre_volume_ratio", 1.0)
        b_vol_ratio = _g("b_pre_volume_ratio", 1.0)
        r_vol_evo = _g("r_pre_vol_ratio_evo")
        b_vol_evo = _g("b_pre_vol_ratio_evo")

        df["volume_ratio_diff"] = r_vol_ratio - b_vol_ratio
        df["vol_ratio_evo_diff"] = r_vol_evo - b_vol_evo
        # High volume ratio = jab-heavy; interacts with KD threat (jabbers rarely KO)
        r_finish_rate_12i = _g("r_pre_finish_rate")
        b_finish_rate_12i = _g("b_pre_finish_rate")
        df["volume_vs_finisher"] = (b_vol_ratio - r_vol_ratio) * (
            r_finish_rate_12i - b_finish_rate_12i
        )

        # Zone shift
        r_zone_shift = _g("r_pre_zone_shift")
        b_zone_shift = _g("b_pre_zone_shift")
        df["zone_shift_diff"] = r_zone_shift - b_zone_shift
        # Negative zone shift (moving closer) × high TD rate = wrestler game plan
        df["zone_x_td"] = (-r_zone_shift).clip(lower=0) * r_td_avg_12i - (
            -b_zone_shift
        ).clip(lower=0) * b_td_avg_12i

        # TD accuracy evolution
        r_r1_td_acc = _g("r_pre_r1_td_acc")
        b_r1_td_acc = _g("b_pre_r1_td_acc")
        r_late_td_acc = _g("r_pre_late_td_acc")
        b_late_td_acc = _g("b_pre_late_td_acc")
        r_td_acc_evo = _g("r_pre_td_acc_evo")
        b_td_acc_evo = _g("b_pre_td_acc_evo")

        df["r1_td_acc_diff"] = r_r1_td_acc - b_r1_td_acc
        df["late_td_acc_diff"] = r_late_td_acc - b_late_td_acc
        df["td_acc_evo_diff"] = r_td_acc_evo - b_td_acc_evo

        # Championship round extras
        r_r45_kd = _g("r_pre_r45_kd_rate")
        b_r45_kd = _g("b_pre_r45_kd_rate")
        r_r45_body = _g("r_pre_r45_body_rate")
        b_r45_body = _g("b_pre_r45_body_rate")
        r_r45_clinch = _g("r_pre_r45_clinch_rate")
        b_r45_clinch = _g("b_pre_r45_clinch_rate")
        r_r45_ctrl = _g("r_pre_r45_ctrl")
        b_r45_ctrl = _g("b_pre_r45_ctrl")

        df["r45_kd_diff"] = r_r45_kd - b_r45_kd
        df["r45_body_diff"] = r_r45_body - b_r45_body
        df["r45_clinch_diff"] = r_r45_clinch - b_r45_clinch
        df["r45_ctrl_diff"] = r_r45_ctrl - b_r45_ctrl
        # Championship ctrl time × 5-round bout flag
        df["r45_ctrl_x_five"] = df["r45_ctrl_diff"] * (total_rds_12i / 3.0)

        # Reversal rate
        r_rev_rate = _g("r_pre_reversal_rate")
        b_rev_rate = _g("b_pre_reversal_rate")
        df["reversal_rate_diff"] = r_rev_rate - b_rev_rate
        # Reversal × ground escalation: escaping takedowns while threatening from bottom
        df["reversal_x_ground"] = (r_rev_rate * _g("r_pre_ground_escalation")) - (
            b_rev_rate * _g("b_pre_ground_escalation")
        )

        # Round momentum
        r_late_win = _g("r_pre_late_win_rate", 0.5)
        b_late_win = _g("b_pre_late_win_rate", 0.5)
        df["r1_win_rate_diff"] = r_r1_win - b_r1_win
        df["late_win_rate_diff"] = r_late_win - b_late_win
        # Composite momentum: wins late rounds AND escalates to clinch (grappling closer)
        df["late_momentum_composite"] = df["late_win_rate_diff"] * (
            1.0 - df["zone_shift_diff"].clip(upper=0)
        )
        # Early vs late rounder clash: one wins R1, the other wins R3
        df["early_vs_late_rounder"] = (
            r_r1_win - r_late_win) - (b_r1_win - b_late_win)

        # ── Tier 12j: Damage margins, head accuracy, grappling chains, entropy ─────
        self._log(
            "Tier 12j: Damage margins, head acc evolution, grappling chains, entropy, finish timing..."
        )

        def _h(col, default=0.0):
            return df.get(col, pd.Series(default, index=df.index)).fillna(default)

        # Per-round damage margins
        r_r1_dmg = _h("r_pre_r1_damage_margin")
        b_r1_dmg = _h("b_pre_r1_damage_margin")
        r_r2_dmg = _h("r_pre_r2_damage_margin")
        b_r2_dmg = _h("b_pre_r2_damage_margin")
        r_r3_dmg = _h("r_pre_r3_damage_margin")
        b_r3_dmg = _h("b_pre_r3_damage_margin")
        r_dmg_chg = _h("r_pre_damage_margin_change")
        b_dmg_chg = _h("b_pre_damage_margin_change")

        df["r1_damage_margin_diff"] = r_r1_dmg - b_r1_dmg
        df["r2_damage_margin_diff"] = r_r2_dmg - b_r2_dmg
        df["r3_damage_margin_diff"] = r_r3_dmg - b_r3_dmg
        df["damage_margin_chg_diff"] = r_dmg_chg - b_dmg_chg

        # Damage trajectory: fighter improves margin across rounds vs fighter who fades
        df["damage_trajectory_clash"] = (
            r_dmg_chg - b_dmg_chg
        )  # same as above but named semantically

        # Composite: who dominates both R1 AND R3 (consistent dominator)
        df["r_multi_round_dom"] = (r_r1_dmg > 0).astype(float) * (r_r3_dmg > 0).astype(
            float
        )
        df["b_multi_round_dom"] = (b_r1_dmg > 0).astype(float) * (b_r3_dmg > 0).astype(
            float
        )
        df["multi_round_dom_diff"] = df["r_multi_round_dom"] - \
            df["b_multi_round_dom"]

        # Head accuracy evolution
        r_r1_ha = _h("r_pre_r1_head_acc", 0.45)
        b_r1_ha = _h("b_pre_r1_head_acc", 0.45)
        r_r3_ha = _h("r_pre_r3_head_acc", 0.45)
        b_r3_ha = _h("b_pre_r3_head_acc", 0.45)
        r_ha_fade = _h("r_pre_head_acc_fade")
        b_ha_fade = _h("b_pre_head_acc_fade")
        r_htl = _h("r_pre_head_to_leg_late")
        b_htl = _h("b_pre_head_to_leg_late")

        df["r1_head_acc_diff"] = r_r1_ha - b_r1_ha
        df["r3_head_acc_diff"] = r_r3_ha - b_r3_ha
        df["head_acc_fade_diff"] = r_ha_fade - b_ha_fade
        df["head_to_leg_late_diff"] = r_htl - b_htl

        # Fighter whose head accuracy holds up late wins more exchanges — amplified in long fights
        total_rds_12j = df.get(
            "total_rounds", pd.Series(3, index=df.index)).fillna(3)
        df["head_acc_stamina_diff"] = (
            r_r3_ha - b_r3_ha) * (total_rds_12j / 3.0)

        # Head-to-leg switch vs opponent who closes distance: kick fighter being tackled
        r_pressure_j = _h("r_pre_pressure_index")
        b_pressure_j = _h("b_pre_pressure_index")
        df["kicker_vs_pressure"] = r_htl * b_pressure_j - b_htl * r_pressure_j

        # Grappling chain composites
        r_gc = _h("r_pre_grapple_chain")
        b_gc = _h("b_pre_grapple_chain")
        r_gnp = _h("r_pre_gnp_score")
        b_gnp = _h("b_pre_gnp_score")
        r_sbe = _h("r_pre_sub_efficiency")
        b_sbe = _h("b_pre_sub_efficiency")

        df["grapple_chain_diff"] = r_gc - b_gc
        df["gnp_score_diff"] = r_gnp - b_gnp
        df["sub_efficiency_diff"] = r_sbe - b_sbe

        # Full grappling system score: chain + gnp + sub combined
        df["r_full_grapple_system"] = r_gc + r_gnp * 0.5 + r_sbe * 0.5
        df["b_full_grapple_system"] = b_gc + b_gnp * 0.5 + b_sbe * 0.5
        df["full_grapple_system_diff"] = (
            df["r_full_grapple_system"] - df["b_full_grapple_system"]
        )

        # Style entropy — unpredictable fighters are harder to defend against
        r_tent = _h("r_pre_target_entropy", 1.0)
        b_tent = _h("b_pre_target_entropy", 1.0)
        r_zent = _h("r_pre_zone_entropy", 1.0)
        b_zent = _h("b_pre_zone_entropy", 1.0)

        df["target_entropy_diff"] = r_tent - b_tent
        df["zone_entropy_diff"] = r_zent - b_zent
        df["total_entropy_diff"] = (r_tent + r_zent) - (b_tent + b_zent)

        # High entropy vs low entropy: diverse attacker vs predictable opponent
        df["entropy_matchup"] = df["total_entropy_diff"] * (
            -df["r3_damage_margin_diff"]
        )

        # Finish timing
        r_afr = _h("r_pre_avg_finish_round", 2.5)
        b_afr = _h("b_pre_avg_finish_round", 2.5)
        r_r1fr = _h("r_pre_r1_finish_rate")
        b_r1fr = _h("b_pre_r1_finish_rate")
        r_efs = _h("r_pre_early_finish_score")
        b_efs = _h("b_pre_early_finish_score")

        df["avg_finish_round_diff"] = r_afr - b_afr
        df["r1_finish_rate_diff"] = r_r1fr - b_r1fr
        df["early_finish_score_diff"] = r_efs - b_efs

        # Early finisher meets late-fight threat: most dangerous clash
        df["early_vs_endurance_clash"] = r_efs - \
            b_efs * _h("b_pre_cardio_index", 0.75)

        # Comeback and lead-hold rates
        r_cbk = _h("r_pre_comeback_rate")
        b_cbk = _h("b_pre_comeback_rate")
        r_lhr = _h("r_pre_lead_hold_rate")
        b_lhr = _h("b_pre_lead_hold_rate")

        df["comeback_rate_diff"] = r_cbk - b_cbk
        df["lead_hold_rate_diff"] = r_lhr - b_lhr

        # Mental toughness composite: comes back AND holds leads
        df["r_mental_toughness"] = r_cbk * r_lhr
        df["b_mental_toughness"] = b_cbk * b_lhr
        df["mental_toughness_diff"] = (
            df["r_mental_toughness"] - df["b_mental_toughness"]
        )

        # Power and chin
        r_kd_eff = _h("r_pre_kd_efficiency")
        b_kd_eff = _h("b_pre_kd_efficiency")
        r_r1_hpres = _h("r_pre_r1_head_pressure")
        b_r1_hpres = _h("b_pre_r1_head_pressure")

        df["kd_efficiency_diff"] = r_kd_eff - b_kd_eff
        df["r1_head_pressure_diff"] = r_r1_hpres - b_r1_hpres

        # High KD efficiency vs opponent who absorbs heavy R1 head pressure
        df["ko_threat_vs_chin"] = r_kd_eff * b_r1_hpres - b_kd_eff * r_r1_hpres

        # Style composite indices
        df["pressure_index_diff"] = r_pressure_j - b_pressure_j
        r_counter = _h("r_pre_counter_index")
        b_counter = _h("b_pre_counter_index")
        df["counter_index_diff"] = r_counter - b_counter

        # Pressure fighter vs counter fighter: the canonical grappler-vs-counterpuncher clash
        df["pressure_vs_counter"] = r_pressure_j * \
            b_counter - b_pressure_j * r_counter

        # Aggressor identity: pressure fighter who also wins damage margins = dominant aggressor
        df["r_dominant_aggressor"] = r_pressure_j * (r_r1_dmg.clip(lower=0))
        df["b_dominant_aggressor"] = b_pressure_j * (b_r1_dmg.clip(lower=0))
        df["dominant_aggressor_diff"] = (
            df["r_dominant_aggressor"] - df["b_dominant_aggressor"]
        )

        # ── Tier 12k: Consistency, closing speed, quality, fatigue, tactical depth ─
        self._log(
            "Tier 12k: Output consistency, closing momentum, strike quality, fatigue, tactical depth..."
        )

        def _k(col, default=0.0):
            return df.get(col, pd.Series(default, index=df.index)).fillna(default)

        # R2→R3 closing momentum
        r_r2r3 = _k("r_pre_r2_to_r3_momentum")
        b_r2r3 = _k("b_pre_r2_to_r3_momentum")
        df["r2_to_r3_momentum_diff"] = r_r2r3 - b_r2r3
        # Closing momentum × fight length (matters more in 5-rounders where late rounds decide)
        total_rds_k = df.get("total_rounds", pd.Series(
            3, index=df.index)).fillna(3)
        df["closing_momentum_x_rounds"] = df["r2_to_r3_momentum_diff"] * (
            total_rds_k / 3.0
        )

        # Output consistency
        r_cv = _k("r_pre_output_cv", 0.5)
        b_cv = _k("b_pre_output_cv", 0.5)
        r_oc = _k("r_pre_output_consistency")
        b_oc = _k("b_pre_output_consistency")
        r_r1r3 = _k("r_pre_r1_r3_consistency", 0.5)
        b_r1r3 = _k("b_pre_r1_r3_consistency", 0.5)
        df["output_cv_diff"] = r_cv - b_cv
        df["output_consistency_diff"] = r_oc - b_oc
        df["r1_r3_consistency_diff"] = r_r1r3 - b_r1r3
        # Consistent fighter vs streaky fighter — consistent one wins decisions
        df["consistency_vs_streaky"] = r_oc * b_cv - b_oc * r_cv

        # Control-time GnP effectiveness
        r_cgnp = _k("r_pre_ctrl_gnp_rate")
        b_cgnp = _k("b_pre_ctrl_gnp_rate")
        df["ctrl_gnp_rate_diff"] = r_cgnp - b_cgnp
        # GnP rate × TD slope (wrestler who lands more often AND damages during control)
        df["gnp_x_td_slope"] = (r_cgnp * _k("r_pre_td_slope").clip(lower=0)) - (
            b_cgnp * _k("b_pre_td_slope").clip(lower=0)
        )

        # Strike quality evolution
        r_r1sq = _k("r_pre_r1_sig_pct", 0.6)
        b_r1sq = _k("b_pre_r1_sig_pct", 0.6)
        r_r3sq = _k("r_pre_r3_sig_pct", 0.6)
        b_r3sq = _k("b_pre_r3_sig_pct", 0.6)
        r_sqev = _k("r_pre_sig_pct_evo")
        b_sqev = _k("b_pre_sig_pct_evo")
        df["r1_sig_pct_diff"] = r_r1sq - b_r1sq
        df["r3_sig_pct_diff"] = r_r3sq - b_r3sq
        df["sig_pct_evo_diff"] = r_sqev - b_sqev
        # Maintains strike quality in R3 while opponent degrades = quality gap compounds
        df["quality_gap_r3"] = r_r3sq * \
            (1.0 - b_r3sq) - b_r3sq * (1.0 - r_r3sq)

        # Tactical divergence — adaptable fighter vs one-dimensional
        r_tdiv = _k("r_pre_tactical_div")
        b_tdiv = _k("b_pre_tactical_div")
        df["tactical_div_diff"] = r_tdiv - b_tdiv
        # High tactical divergence × high entropy = fully unpredictable game-planner
        df["r_adaptable_unpredictable"] = r_tdiv * \
            _k("r_pre_zone_entropy", 1.0)
        df["b_adaptable_unpredictable"] = b_tdiv * \
            _k("b_pre_zone_entropy", 1.0)
        df["adaptable_unpredictable_diff"] = (
            df["r_adaptable_unpredictable"] - df["b_adaptable_unpredictable"]
        )

        # Fatigue composite — both fighters' fatigue signals clash
        r_fat = _k("r_pre_fatigue_composite")
        b_fat = _k("b_pre_fatigue_composite")
        df["fatigue_composite_diff"] = r_fat - b_fat
        # If both fighters have high fatigue, the more durable one wins — interacts with cardio
        r_cardio_k = _k("r_pre_cardio_index", 0.75)
        b_cardio_k = _k("b_pre_cardio_index", 0.75)
        df["fatigue_vs_cardio"] = (b_fat - r_fat) * (r_cardio_k - b_cardio_k)

        # R1 absorption rate
        r_r1abs = _k("r_pre_r1_abs_rate")
        b_r1abs = _k("b_pre_r1_abs_rate")
        df["r1_abs_rate_diff"] = r_r1abs - b_r1abs
        # Striker whose R1 absorption is low AND deals high R1 damage = dominant opener
        r_r1dmg_k = _k("r_pre_r1_damage_margin")
        b_r1dmg_k = _k("b_pre_r1_damage_margin")
        df["r1_net_damage_quality"] = (
            r_r1dmg_k - r_r1abs) - (b_r1dmg_k - b_r1abs)

        # Championship damage margins
        r_r45dm = _k("r_pre_r45_dmg_margin")
        b_r45dm = _k("b_pre_r45_dmg_margin")
        df["r45_dmg_margin_diff"] = r_r45dm - b_r45dm
        df["r45_dmg_x_five_rounds"] = df["r45_dmg_margin_diff"] * \
            (total_rds_k / 3.0)

        # Combo rate
        r_combo = _k("r_pre_combo_rate")
        b_combo = _k("b_pre_combo_rate")
        df["combo_rate_diff"] = r_combo - b_combo
        # Combo puncher vs single-target striker — combos set up KOs
        df["combo_x_kd_eff"] = r_combo * _k("r_pre_kd_efficiency") - b_combo * _k(
            "b_pre_kd_efficiency"
        )

        # Head strike momentum
        r_hmom = _k("r_pre_head_momentum", 1.0)
        b_hmom = _k("b_pre_head_momentum", 1.0)
        df["head_momentum_diff"] = r_hmom - b_hmom
        # Fighter who escalates head work against opponent whose defense fades
        df["head_momentum_vs_acc_fade"] = (r_hmom - 1.0) * _k("b_pre_head_acc_fade") - (
            b_hmom - 1.0
        ) * _k("r_pre_head_acc_fade")

        # Clinch vs ground ratio
        r_cgr = _k("r_pre_clinch_grd_ratio", 0.5)
        b_cgr = _k("b_pre_clinch_grd_ratio", 0.5)
        df["clinch_grd_ratio_diff"] = r_cgr - b_cgr
        # Dirty boxer (high clinch) vs wrestler (high ground) — different threat profiles
        df["dirty_boxer_vs_wrestler"] = (r_cgr - b_cgr) * (
            b_cgr - r_cgr
        )  # symmetric clash

        # Conditional finish rate
        r_fww = _k("r_pre_finish_when_winning")
        b_fww = _k("b_pre_finish_when_winning")
        df["finish_when_winning_diff"] = r_fww - b_fww
        # Converts leads to finishes AND has high R1 win rate = lethal combination
        r_r1wr_k = _k("r_pre_r1_win_rate", 0.5)
        b_r1wr_k = _k("b_pre_r1_win_rate", 0.5)
        df["finish_conversion_index"] = (r_fww * r_r1wr_k) - (b_fww * b_r1wr_k)

        # Sub round timing
        r_srt = _k("r_pre_sub_round_avg", 2.0)
        b_srt = _k("b_pre_sub_round_avg", 2.0)
        df["sub_round_avg_diff"] = r_srt - b_srt
        # Early sub hunter (R1) vs late sub hunter (R3+): clash timing
        df["sub_timing_clash"] = (3.0 - r_srt) * _k("r_pre_late_sub_rate") - (
            3.0 - b_srt
        ) * _k("b_pre_late_sub_rate")

        # Integrated dominance — the composite "who outfights whom" score
        r_idom = _k("r_pre_integrated_dom")
        b_idom = _k("b_pre_integrated_dom")
        df["integrated_dom_diff"] = r_idom - b_idom

        # R2 KD threat — knockouts in the adjustment round
        r_r2kd = _k("r_pre_r2_kd_threat")
        b_r2kd = _k("b_pre_r2_kd_threat")
        df["r2_kd_threat_diff"] = r_r2kd - b_r2kd
        # R2 KD threat × comeback rate (lands KD when opponent pushes back in R2)
        df["r2_ko_vs_comeback"] = r_r2kd * _k("b_pre_comeback_rate") - b_r2kd * _k(
            "r_pre_comeback_rate"
        )

        # R1 control-seeking
        r_r1co = _k("r_pre_r1_ctrl_opening")
        b_r1co = _k("b_pre_r1_ctrl_opening")
        df["r1_ctrl_opening_diff"] = r_r1co - b_r1co
        # Wrestler who seeks early control vs striker who defends early TDs
        df["early_control_vs_td_def"] = r_r1co * (
            1.0 - _k("b_pre_r1_td_acc", 0.0)
        ) - b_r1co * (1.0 - _k("r_pre_r1_td_acc", 0.0))

        # Quality-of-win R3 margin
        r_wr3 = _k("r_pre_win_r3_margin")
        b_wr3 = _k("b_pre_win_r3_margin")
        df["win_r3_margin_diff"] = r_wr3 - b_wr3
        # Dominates the final round in wins = scores judges' scorecards convincingly
        df["closing_quality_diff"] = r_wr3 - b_wr3  # kept as named alias

        # Grand composite: integrated dominance × consistency × closing momentum
        df["r_grand_composite"] = r_idom * r_oc * (1.0 + r_r2r3.clip(lower=0))
        df["b_grand_composite"] = b_idom * b_oc * (1.0 + b_r2r3.clip(lower=0))
        df["grand_composite_diff"] = df["r_grand_composite"] - \
            df["b_grand_composite"]

        # ── Tier 12l: Per-round defense, adjustment speed, grappling depth ──────
        self._log(
            "Tier 12l: TD defense, head adjustment, chin fade, output efficiency, late grappling..."
        )

        def _l(col, default=0.0):
            return df.get(col, pd.Series(default, index=df.index)).fillna(default)

        # R1 TD defense
        r_r1tdd = _l("r_pre_r1_td_def_rate", 0.7)
        b_r1tdd = _l("b_pre_r1_td_def_rate", 0.7)
        r_r1tdp = _l("r_pre_r1_td_pressure")
        b_r1tdp = _l("b_pre_r1_td_pressure")
        df["r1_td_def_rate_diff"] = r_r1tdd - b_r1tdd
        df["r1_td_pressure_diff"] = r_r1tdp - b_r1tdp
        # High TD defense vs high TD pressure opponent: classic wrestler vs. stuffier clash
        df["td_def_vs_pressure"] = r_r1tdd * b_r1tdp - b_r1tdd * r_r1tdp
        # TD pressure absorbed × opponent KD efficiency = being taken down AND then hit hard
        df["td_pressure_x_kd"] = b_r1tdp * _l("r_pre_kd_efficiency") - r_r1tdp * _l(
            "b_pre_kd_efficiency"
        )

        # R2 head accuracy and range-finding
        r_r2hacc = _l("r_pre_r2_head_acc", 0.3)
        b_r2hacc = _l("b_pre_r2_head_acc", 0.3)
        r_r1r2adj = _l("r_pre_r1r2_head_adj")
        b_r1r2adj = _l("b_pre_r1r2_head_adj")
        df["r2_head_acc_diff"] = r_r2hacc - b_r2hacc
        df["r1r2_head_adj_diff"] = r_r1r2adj - b_r1r2adj
        # Fighter who finds head accuracy in R2 while opponent's head acc fades = decisive edge
        df["head_finding_range"] = r_r1r2adj * _l(
            "b_pre_head_acc_fade"
        ) - b_r1r2adj * _l("r_pre_head_acc_fade")
        # High R2 head accuracy × R2 KD threat = capitalization rate
        df["r2_head_ko_potential"] = r_r2hacc * _l(
            "r_pre_r2_kd_threat"
        ) - b_r2hacc * _l("b_pre_r2_kd_threat")

        # Chin durability across rounds
        r_chin = _l("r_pre_chin_ratio", 1.0)
        b_chin = _l("b_pre_chin_ratio", 1.0)
        df["chin_ratio_diff"] = r_chin - b_chin
        # Durable chin (ratio ≈ 1) vs power striker: high power + vulnerable chin = KO threat
        df["chin_vs_power"] = b_chin * _l("r_pre_kd_efficiency") - r_chin * _l(
            "b_pre_kd_efficiency"
        )
        # Chin ratio × absorption slope: double signal that a fighter is fading under volume
        df["chin_fade_signal"] = (
            b_chin * _l("b_pre_absorption_slope").clip(lower=0)
        ) - (r_chin * _l("r_pre_absorption_slope").clip(lower=0))

        # Mid-to-late round finishing (R2+R3 KDs)
        r_r2r3kd = _l("r_pre_r2r3_kd_rate")
        b_r2r3kd = _l("b_pre_r2r3_kd_rate")
        df["r2r3_kd_rate_diff"] = r_r2r3kd - b_r2r3kd
        # Mid-round KD threat × comeback rate: knocks down opponent when they push back
        df["mid_ko_vs_comeback"] = r_r2r3kd * _l("b_pre_comeback_rate") - b_r2r3kd * _l(
            "r_pre_comeback_rate"
        )

        # Sequential grappling (sub after TD)
        r_satd = _l("r_pre_sub_after_td")
        b_satd = _l("b_pre_sub_after_td")
        df["sub_after_td_diff"] = r_satd - b_satd
        # Sub-after-TD × grapple chain: dual grappling threat signals compound
        df["sequential_grapple_diff"] = (r_satd * _l("r_pre_grapple_chain")) - (
            b_satd * _l("b_pre_grapple_chain")
        )

        # Per-round output efficiency (net exchange quality)
        r_r1eff = _l("r_pre_r1_output_eff", 1.0)
        b_r1eff = _l("b_pre_r1_output_eff", 1.0)
        r_r3eff = _l("r_pre_r3_output_eff", 1.0)
        b_r3eff = _l("b_pre_r3_output_eff", 1.0)
        r_efft = _l("r_pre_output_eff_trend")
        b_efft = _l("b_pre_output_eff_trend")
        df["r1_output_eff_diff"] = r_r1eff - b_r1eff
        df["r3_output_eff_diff"] = r_r3eff - b_r3eff
        df["output_eff_trend_diff"] = r_efft - b_efft
        # R1 exchange winner AND stays efficient in R3 = dominant throughout
        df["sustained_efficiency"] = (r_r1eff * r_r3eff) - (b_r1eff * b_r3eff)
        # Improving efficiency (positive trend) vs deteriorating opponent = compounding advantage
        df["efficiency_momentum"] = (
            r_efft.clip(lower=0) * b_r1eff - b_efft.clip(lower=0) * r_r1eff
        )

        # Championship round head work (R4+R5) — 5-round fight dominance
        r_r45h = _l("r_pre_r45_head_rate")
        b_r45h = _l("b_pre_r45_head_rate")
        df["r45_head_rate_diff"] = r_r45h - b_r45h
        df["r45_head_x_five"] = df["r45_head_rate_diff"] * (total_rds_k / 3.0)
        # Championship head work × R3 output efficiency = dominant in all late rounds
        df["r45_head_x_r3_eff"] = (r_r45h * r_r3eff) - (b_r45h * b_r3eff)

        # R1→R2 zone adjustment (early tactical adaptation)
        r_r1r2z = _l("r_pre_r1r2_zone_adj")
        b_r1r2z = _l("b_pre_r1r2_zone_adj")
        df["r1r2_zone_adj_diff"] = r_r1r2z - b_r1r2z
        # Pressing in after R1 (negative adj = less distance) × pressure index = aggressor
        df["early_zone_initiative"] = (-r_r1r2z).clip(lower=0) * _l(
            "r_pre_pressure_index"
        ) - (-b_r1r2z).clip(lower=0) * _l("b_pre_pressure_index")

        # R2+R3 sub hunting
        r_r23s = _l("r_pre_r23_sub_rate")
        b_r23s = _l("b_pre_r23_sub_rate")
        df["r23_sub_rate_diff"] = r_r23s - b_r23s
        # R2+R3 sub hunting × R1 TD success = grappler with a full plan
        df["sub_hunt_vs_td_off"] = r_r23s * _l("r_pre_r1_td_acc", 0.5) - b_r23s * _l(
            "b_pre_r1_td_acc", 0.5
        )

        # Championship sub and GnP threats (R4+R5)
        r_r45s = _l("r_pre_r45_sub_rate_l")
        b_r45s = _l("b_pre_r45_sub_rate_l")
        r_r45gr = _l("r_pre_r45_ground_rate")
        b_r45gr = _l("b_pre_r45_ground_rate")
        df["r45_sub_rate_diff"] = r_r45s - b_r45s
        df["r45_ground_rate_diff"] = r_r45gr - b_r45gr
        # Championship GnP composite: sub threat + ground volume in R4+R5
        df["r45_grapple_composite"] = (r_r45s + r_r45gr) - (b_r45s + b_r45gr)
        df["r45_gc_x_five"] = df["r45_grapple_composite"] * (total_rds_k / 3.0)

        # R3 leg rate and late kicking game
        r_r3leg = _l("r_pre_r3_leg_rate")
        b_r3leg = _l("b_pre_r3_leg_rate")
        df["r3_leg_rate_diff"] = r_r3leg - b_r3leg
        # Late leg kicking vs kicker-stopper (wrestler with good R1 TD accuracy)
        df["late_kick_vs_wrestler"] = r_r3leg * (
            1.0 - _l("b_pre_r1_td_acc", 0.5)
        ) - b_r3leg * (1.0 - _l("r_pre_r1_td_acc", 0.5))
        # Late leg escalation: R3 leg > R1 leg (ramping kicks) × head acc fade (opponent fading)
        r_leg_esc = _l("r_pre_leg_escalation")
        b_leg_esc = _l("b_pre_leg_escalation")
        df["late_leg_ramp_diff"] = (
            r_r3leg + r_leg_esc) - (b_r3leg + b_leg_esc)

        # Full-round dominance composite: R1 efficiency × R3 efficiency × cardio × consistency
        df["r_full_round_dom"] = r_r1eff * r_r3eff * r_cardio_k * r_oc
        df["b_full_round_dom"] = b_r1eff * b_r3eff * b_cardio_k * b_oc
        df["full_round_dom_diff"] = df["r_full_round_dom"] - df["b_full_round_dom"]

        # ── Tier 12m: Style entropy adaptation, grappling efficiency, resilience ─
        self._log(
            "Tier 12m: R2 entropy, ctrl-per-TD, finish diversity, body evolution, resilience..."
        )

        def _m(col, default=0.0):
            return df.get(col, pd.Series(default, index=df.index)).fillna(default)

        # R2 target and zone entropy
        r_r2tent = _m("r_pre_r2_target_ent", 1.0)
        b_r2tent = _m("b_pre_r2_target_ent", 1.0)
        r_r2zent = _m("r_pre_r2_zone_ent", 1.0)
        b_r2zent = _m("b_pre_r2_zone_ent", 1.0)
        df["r2_target_ent_diff"] = r_r2tent - b_r2tent
        df["r2_zone_ent_diff"] = r_r2zent - b_r2zent

        # R1→R2 entropy shifts (positive = more diverse/adaptive in R2)
        r_tgt_sh = _m("r_pre_r1r2_tgt_shift")
        b_tgt_sh = _m("b_pre_r1r2_tgt_shift")
        r_zon_sh = _m("r_pre_r1r2_zone_shift")
        b_zon_sh = _m("b_pre_r1r2_zone_shift")
        df["r1r2_tgt_shift_diff"] = r_tgt_sh - b_tgt_sh
        df["r1r2_zone_shift_diff"] = r_zon_sh - b_zon_sh
        # Combined adaptation score: fighter who diversifies both targets AND zones in R2
        df["r_style_adaptation"] = r_tgt_sh.clip(
            lower=0) + r_zon_sh.clip(lower=0)
        df["b_style_adaptation"] = b_tgt_sh.clip(
            lower=0) + b_zon_sh.clip(lower=0)
        df["style_adaptation_diff"] = (
            df["r_style_adaptation"] - df["b_style_adaptation"]
        )
        # Adaptive fighter × high entropy (already diverse) = unpredictable game-planner
        r_zent = _m("r_pre_zone_entropy", 1.0)
        b_zent = _m("b_pre_zone_entropy", 1.0)
        df["adaptive_entropy_diff"] = (r_tgt_sh * r_zent) - (b_tgt_sh * b_zent)

        # Control per TD (grappling dominance: time spent in control per successful takedown)
        r_cptd = _m("r_pre_ctrl_per_td")
        b_cptd = _m("b_pre_ctrl_per_td")
        df["ctrl_per_td_diff"] = r_cptd - b_cptd
        # High ctrl per TD × TD slope (gets more TDs in later rounds AND dominates each one)
        r_tdsl = _m("r_pre_td_slope")
        b_tdsl = _m("b_pre_td_slope")
        df["grappling_dominance"] = (r_cptd * r_tdsl.clip(lower=0)) - (
            b_cptd * b_tdsl.clip(lower=0)
        )
        # Ctrl per TD × sub efficiency = wrestler who dominates AND hunts submissions
        df["ctrl_sub_chain_diff"] = (r_cptd * _m("r_pre_sub_efficiency")) - (
            b_cptd * _m("b_pre_sub_efficiency")
        )

        # Finish round entropy (versatile finisher: can finish in any round)
        r_frnd_ent = _m("r_pre_finish_rnd_ent")
        b_frnd_ent = _m("b_pre_finish_rnd_ent")
        df["finish_rnd_ent_diff"] = r_frnd_ent - b_frnd_ent
        # Versatile finisher × high finish rate = elite finisher who isn't round-dependent
        r_finr = (
            _m("r_pre_finish_rate")
            if "r_pre_finish_rate" in df.columns
            else _m("r_pre_r1_finish_rate")
        )
        b_finr = (
            _m("b_pre_finish_rate")
            if "b_pre_finish_rate" in df.columns
            else _m("b_pre_r1_finish_rate")
        )
        df["versatile_finisher_diff"] = (
            r_frnd_ent * r_finr) - (b_frnd_ent * b_finr)

        # R3 body pct and evolution
        r_r3bp = _m("r_pre_r3_body_pct")
        b_r3bp = _m("b_pre_r3_body_pct")
        r_bpev = _m("r_pre_body_pct_evo")
        b_bpev = _m("b_pre_body_pct_evo")
        df["r3_body_pct_diff"] = r_r3bp - b_r3bp
        df["body_pct_evo_diff"] = r_bpev - b_bpev
        # Escalating body work (positive evolution) AND high R3 body % = late body specialist
        df["body_escalation_quality"] = (r_bpev.clip(lower=0) * r_r3bp) - (
            b_bpev.clip(lower=0) * b_r3bp
        )
        # Body escalation vs opponent whose R3 efficiency is declining = exploit fading guard
        r_r3eff_m = _m("r_pre_r3_output_eff", 1.0)
        b_r3eff_m = _m("b_pre_r3_output_eff", 1.0)
        df["body_vs_fading_def"] = r_r3bp * (
            1.0 - b_r3eff_m.clip(upper=1.0)
        ) - b_r3bp * (1.0 - r_r3eff_m.clip(upper=1.0))

        # R1 clinch→ground link (dirty boxer who transitions to wrestling)
        r_cgl = _m("r_pre_r1_cg_link")
        b_cgl = _m("b_pre_r1_cg_link")
        df["r1_cg_link_diff"] = r_cgl - b_cgl
        # Clinch→ground link × clinch escalation: grappler who builds through the clinch
        r_cesc = _m("r_pre_clinch_escalation")
        b_cesc = _m("b_pre_clinch_escalation")
        df["clinch_wrestling_sys"] = (r_cgl * r_cesc.clip(lower=0)) - (
            b_cgl * b_cesc.clip(lower=0)
        )

        # Output in losses (competitive spirit: fighter who stays active when losing)
        r_oil = _m("r_pre_output_in_loss")
        b_oil = _m("b_pre_output_in_loss")
        df["output_in_loss_diff"] = r_oil - b_oil
        # High output in losses × high comeback rate = genuinely resilient competitor
        r_cbk_m = _m("r_pre_comeback_rate")
        b_cbk_m = _m("b_pre_comeback_rate")
        df["resilient_output"] = (r_oil * r_cbk_m) - (b_oil * b_cbk_m)

        # Reversals per ctrl minute (escape artist / scramble ability)
        r_rpc = _m("r_pre_rev_per_ctrl")
        b_rpc = _m("b_pre_rev_per_ctrl")
        df["rev_per_ctrl_diff"] = r_rpc - b_rpc
        # Escape artist vs wrestler: high rev/ctrl × opp TD pressure = stuffs/reverses attempts
        df["escape_vs_wrestler"] = r_rpc * _m("b_pre_r1_td_pressure") - b_rpc * _m(
            "r_pre_r1_td_pressure"
        )

        # R2/R1 output ratio (pacing strategy)
        r_r2r1 = _m("r_pre_r2_r1_ratio", 1.0)
        b_r2r1 = _m("b_pre_r2_r1_ratio", 1.0)
        df["r2_r1_ratio_diff"] = r_r2r1 - b_r2r1
        # R2 surge (ratio > 1) + R2 head accuracy improvement = finding range AND pressing hard
        r_r1r2adj_m = _m("r_pre_r1r2_head_adj")
        b_r1r2adj_m = _m("b_pre_r1r2_head_adj")
        df["r2_surge_quality"] = (r_r2r1.clip(lower=0) * r_r1r2adj_m.clip(lower=0)) - (
            b_r2r1.clip(lower=0) * b_r1r2adj_m.clip(lower=0)
        )

        # R3 leg pct (late-round kicking as a percentage)
        r_r3lp = _m("r_pre_r3_leg_pct")
        b_r3lp = _m("b_pre_r3_leg_pct")
        df["r3_leg_pct_diff"] = r_r3lp - b_r3lp
        # Late leg targeting × leg escalation = fighter ramping up the leg attack plan
        r_lesc = _m("r_pre_leg_escalation")
        b_lesc = _m("b_pre_leg_escalation")
        df["late_leg_plan_diff"] = (r_r3lp * r_lesc.clip(lower=0)) - (
            b_r3lp * b_lesc.clip(lower=0)
        )

        # R3 output in won fights (late-round dominance when ahead on the scorecards)
        r_wr3o = _m("r_pre_win_r3_output")
        b_wr3o = _m("b_pre_win_r3_output")
        df["win_r3_output_diff"] = r_wr3o - b_wr3o
        # Win R3 output × output consistency = dominates closing rounds AND stays consistent
        df["late_dominance_comp"] = (r_wr3o * r_oc) - (b_wr3o * b_oc)
        # Win R3 output × R3 efficiency (how much they outwork opponent in winning R3s)
        df["r3_win_eff_diff"] = (r_wr3o * r_r3eff_m) - (b_wr3o * b_r3eff_m)

        # ── Tier 12n: Finish method diversity, mid-fight adjustment, decision quality ─
        self._log(
            "Tier 12n: Method entropy, R2 efficiency, front-loading, decision margin, grapple usage..."
        )

        def _n(col, default=0.0):
            return df.get(col, pd.Series(default, index=df.index)).fillna(default)

        # Finish method entropy (KO / sub / decision diversity)
        r_ment = _n("r_pre_method_entropy")
        b_ment = _n("b_pre_method_entropy")
        df["method_entropy_diff"] = r_ment - b_ment
        # Versatile finisher (high entropy) × high finish rate = truly dangerous in any position
        r_fr_n = (
            _n("r_pre_finish_rate")
            if "r_pre_finish_rate" in df.columns
            else _n("r_pre_r1_finish_rate")
        )
        b_fr_n = (
            _n("b_pre_finish_rate")
            if "b_pre_finish_rate" in df.columns
            else _n("b_pre_r1_finish_rate")
        )
        df["method_diversity_threat"] = (r_ment * r_fr_n) - (b_ment * b_fr_n)

        # R2 output efficiency — mid-fight exchange quality
        r_r2eff = _n("r_pre_r2_output_eff", 1.0)
        b_r2eff = _n("b_pre_r2_output_eff", 1.0)
        df["r2_output_eff_diff"] = r_r2eff - b_r2eff
        # Full arc: R1 → R2 → R3 efficiency (consistent exchanger vs. fader vs. builder)
        r_r1eff_n = _n("r_pre_r1_output_eff", 1.0)
        b_r1eff_n = _n("b_pre_r1_output_eff", 1.0)
        r_r3eff_n = _n("r_pre_r3_output_eff", 1.0)
        b_r3eff_n = _n("b_pre_r3_output_eff", 1.0)
        df["r_eff_arc"] = r_r1eff_n * r_r2eff * r_r3eff_n
        df["b_eff_arc"] = b_r1eff_n * b_r2eff * b_r3eff_n
        df["eff_arc_diff"] = df["r_eff_arc"] - df["b_eff_arc"]

        # R1 front-load ratio (pacing strategy: front-loader vs. builder)
        r_fl = _n("r_pre_r1_front_load", 1.0)
        b_fl = _n("b_pre_r1_front_load", 1.0)
        df["r1_front_load_diff"] = r_fl - b_fl
        # Front-loader vs. slow starter: front-loader targets slow starter's R1 vulnerability
        df["front_load_vs_builder"] = r_fl * (1.0 / b_fl.clip(lower=0.5)) - b_fl * (
            1.0 / r_fl.clip(lower=0.5)
        )
        # Front-loader × R1 KD efficiency = crashes in immediately with KO power
        df["front_load_x_kd_eff"] = (r_fl * _n("r_pre_kd_efficiency")) - (
            b_fl * _n("b_pre_kd_efficiency")
        )

        # R2 TD accuracy
        r_r2tda = _n("r_pre_r2_td_acc")
        b_r2tda = _n("b_pre_r2_td_acc")
        df["r2_td_acc_diff"] = r_r2tda - b_r2tda
        # R2 TD accuracy × sub_after_td: takes down AND transitions in R2
        df["r2_td_sub_chain"] = (r_r2tda * _n("r_pre_sub_after_td")) - (
            b_r2tda * _n("b_pre_sub_after_td")
        )

        # Decision win margin (how convincingly they win judged fights in R3)
        r_dwm = _n("r_pre_dec_win_margin")
        b_dwm = _n("b_pre_dec_win_margin")
        df["dec_win_margin_diff"] = r_dwm - b_dwm
        # Decision margin × win R3 output = dominates late rounds in won decisions
        df["late_scorecard_edge"] = (r_dwm * _n("r_pre_win_r3_output")) - (
            b_dwm * _n("b_pre_win_r3_output")
        )

        # Control escalation ratio (wrestler builds dominance on the mat)
        r_cer = _n("r_pre_ctrl_esc_ratio", 1.0)
        b_cer = _n("b_pre_ctrl_esc_ratio", 1.0)
        df["ctrl_esc_ratio_diff"] = r_cer - b_cer
        # Control escalation × ctrl_gnp_rate = builds control AND lands damage during it
        df["ctrl_build_gnp"] = (r_cer * _n("r_pre_ctrl_gnp_rate")) - (
            b_cer * _n("b_pre_ctrl_gnp_rate")
        )

        # Absorbed in losses vs absorbed in wins (fight IQ under pressure)
        r_ail = _n("r_pre_absorbed_in_loss")
        b_ail = _n("b_pre_absorbed_in_loss")
        df["absorbed_in_loss_diff"] = r_ail - b_ail
        # High absorption in losses × low output in losses = fighter who just gets outworked badly
        r_oil_n = _n("r_pre_output_in_loss")
        b_oil_n = _n("b_pre_output_in_loss")
        df["outworked_index"] = b_ail / \
            (b_oil_n + 0.01) - r_ail / (r_oil_n + 0.01)

        # Grappling usage index (total wrestling activity per fight)
        r_gu = _n("r_pre_grapple_usage")
        b_gu = _n("b_pre_grapple_usage")
        df["grapple_usage_diff"] = r_gu - b_gu
        # High usage × high TD accuracy = active AND efficient wrestler
        df["grapple_eff_diff"] = (r_gu * _n("r_pre_r1_td_acc", 0.5)) - (
            b_gu * _n("b_pre_r1_td_acc", 0.5)
        )

        # KD R1 concentration (early KO specialist vs. late finisher)
        r_kd1c = _n("r_pre_kd_r1_conc", 0.5)
        b_kd1c = _n("b_pre_kd_r1_conc", 0.5)
        df["kd_r1_conc_diff"] = r_kd1c - b_kd1c
        # R1 KO specialist vs. durable fighter: R1 concentrated KD × opp chin_ratio (fades)
        df["r1_ko_vs_durable"] = r_kd1c * _n("b_pre_chin_ratio", 1.0) - b_kd1c * _n(
            "r_pre_chin_ratio", 1.0
        )

        # Finish when behind (resilient comeback finisher)
        r_fwb = _n("r_pre_finish_when_behind")
        b_fwb = _n("b_pre_finish_when_behind")
        df["finish_when_behind_diff"] = r_fwb - b_fwb
        # Finishes when behind AND high output in losses = never-say-die competitor
        df["never_say_die_diff"] = (r_fwb * r_oil_n) - (b_fwb * b_oil_n)

        # R2 win rate (mid-fight dominance)
        r_r2wr = _n("r_pre_r2_win_rate", 0.5)
        b_r2wr = _n("b_pre_r2_win_rate", 0.5)
        df["r2_win_rate_diff"] = r_r2wr - b_r2wr
        # Win R1 AND R2 = leads every round going into R3 = clear favorite for scorecards
        r_r1wr_n = _n("r_pre_r1_win_rate", 0.5)
        b_r1wr_n = _n("b_pre_r1_win_rate", 0.5)
        df["r1_r2_dominance"] = (r_r1wr_n * r_r2wr) - (b_r1wr_n * b_r2wr)

        # R3 control time (late-round wrestling dominance)
        r_r3ct = _n("r_pre_r3_ctrl_rate")
        b_r3ct = _n("b_pre_r3_ctrl_rate")
        df["r3_ctrl_rate_diff"] = r_r3ct - b_r3ct
        # R3 control × ctrl escalation ratio = builds and holds throughout the fight
        df["r3_ctrl_compound"] = (r_r3ct * r_cer) - (b_r3ct * b_cer)

        # R1 body rate and R3 distance pct
        r_r1br = _n("r_pre_r1_body_rate")
        b_r1br = _n("b_pre_r1_body_rate")
        r_r3dp = _n("r_pre_r3_dist_pct", 0.5)
        b_r3dp = _n("b_pre_r3_dist_pct", 0.5)
        df["r1_body_rate_diff"] = r_r1br - b_r1br
        df["r3_dist_pct_diff"] = r_r3dp - b_r3dp
        # High R1 body early AND moves to distance in R3 = wearing ribs then circling out
        df["body_to_distance_gameplan"] = (r_r1br * (1.0 - r_r3dp)) - (
            b_r1br * (1.0 - b_r3dp)
        )
        # Opponent who presses to distance in R3 (r3_dist_pct high) vs. body worker = stylistic clash
        df["r3_dist_vs_body"] = (b_r3dp * r_r1br) - (r_r3dp * b_r1br)

        # ── Tier 12o: New composite features ────────────────────────────────────
        self._log(
            "Tier 12o: Striking efficiency, pressure-cardio clash, reach-distance, chin-power, reversal escalation, R45 leg/TD accuracy, R2 adjustment, cardio endurance..."
        )

        def _o(col, default=0.0):
            return df.get(col, pd.Series(default, index=df.index)).fillna(default)

        total_rds_o = df.get("total_rounds", pd.Series(
            3, index=df.index)).fillna(3)

        # ── Striking efficiency (defensive volume striker composite) ──────────
        r_slpm_o = _o("r_pre_SLpM")
        b_slpm_o = _o("b_pre_SLpM")
        r_strdef = _o("r_pre_str_def", 0.5)
        b_strdef = _o("b_pre_str_def", 0.5)
        df["r_striking_efficiency"] = r_slpm_o * r_strdef
        df["b_striking_efficiency"] = b_slpm_o * b_strdef
        df["striking_efficiency_diff"] = (
            df["r_striking_efficiency"] - df["b_striking_efficiency"]
        )
        # High-volume defender vs pure volume fighter: efficiency edge
        df["eff_vs_volume_diff"] = (r_slpm_o * r_strdef) - (
            b_slpm_o * (1.0 - b_strdef.clip(upper=1.0))
        )

        # ── Pressure × opponent cardio interaction ────────────────────────────
        r_pressure_o = _o("r_pre_pressure_index")
        b_pressure_o = _o("b_pre_pressure_index")
        r_cardio_o = _o("r_pre_cardio_index", 0.75)
        b_cardio_o = _o("b_pre_cardio_index", 0.75)
        df["r_pressure_vs_opp_cardio"] = r_pressure_o * (1.0 - b_cardio_o)
        df["b_pressure_vs_opp_cardio"] = b_pressure_o * (1.0 - r_cardio_o)
        df["pressure_cardio_clash_diff"] = (
            df["r_pressure_vs_opp_cardio"] - df["b_pressure_vs_opp_cardio"]
        )
        # Five-round amplification: pressure vs. fading cardio is more decisive over 5 rounds
        df["pressure_cardio_x_five"] = df["pressure_cardio_clash_diff"] * (
            total_rds_o / 3.0
        )

        # ── Reach × distance fighting preference ─────────────────────────────
        r_dist_pct_o = _o("r_pre_distance_pct", 0.5)
        b_dist_pct_o = _o("b_pre_distance_pct", 0.5)
        reach_diff_o = _o("reach_diff", 0.0)
        df["reach_x_dist_preference"] = reach_diff_o * \
            (r_dist_pct_o - b_dist_pct_o)
        # Reach advantage is only valuable when the reach-holder prefers long range
        df["r_reach_utility"] = _o("r_reach", 70.0) * r_dist_pct_o
        df["b_reach_utility"] = _o("b_reach", 70.0) * b_dist_pct_o
        df["reach_utility_diff"] = df["r_reach_utility"] - df["b_reach_utility"]

        # ── Full 4-stage grappling depth chain ────────────────────────────────
        r_gd4 = _o("r_pre_grapple_depth_4")
        b_gd4 = _o("b_pre_grapple_depth_4")
        df["grapple_depth_4_diff"] = r_gd4 - b_gd4
        # 4-stage chain advantage amplified against opponents with weak grappling defense
        r_td_def_o = _o("r_pre_td_def", 0.7)
        b_td_def_o = _o("b_pre_td_def", 0.7)
        df["r_gd4_vs_td_def"] = r_gd4 * (1.0 - b_td_def_o)
        df["b_gd4_vs_td_def"] = b_gd4 * (1.0 - r_td_def_o)
        df["gd4_vs_td_def_diff"] = df["r_gd4_vs_td_def"] - df["b_gd4_vs_td_def"]

        # ── Chin vs. power matchup (career-rate normalised) ───────────────────
        r_kd_eff_o = _o("r_pre_kd_efficiency")
        b_kd_eff_o = _o("b_pre_kd_efficiency")
        r_kd_abs_rate = _o("r_pre_kd_absorbed_rate")
        b_kd_abs_rate = _o("b_pre_kd_absorbed_rate")
        df["kd_absorbed_rate_diff"] = r_kd_abs_rate - b_kd_abs_rate
        # Blue fighter's chin vulnerability × Red's KD efficiency = career-stable KO threat
        df["r_chin_vs_power_norm"] = b_kd_eff_o * r_kd_abs_rate
        df["b_chin_vs_power_norm"] = r_kd_eff_o * b_kd_abs_rate
        df["chin_vs_power_norm_diff"] = (
            df["b_chin_vs_power_norm"] - df["r_chin_vs_power_norm"]
        )
        # Combined: puncher's efficiency vs opponent whose chin degrades per fight
        df["ko_threat_career_norm"] = (r_kd_eff_o * b_kd_abs_rate) - (
            b_kd_eff_o * r_kd_abs_rate
        )

        # ── R4/R5 leg escalation (previously stored in h10 but unused) ────────
        r_r45_leg = _o("r_pre_r45_leg_rate")
        b_r45_leg = _o("b_pre_r45_leg_rate")
        r_r45_lrat = _o("r_pre_r45_vs_r1_leg_ratio", 1.0)
        b_r45_lrat = _o("b_pre_r45_vs_r1_leg_ratio", 1.0)
        df["r45_leg_rate_diff"] = r_r45_leg - b_r45_leg
        df["r45_leg_ratio_diff"] = r_r45_lrat - b_r45_lrat
        df["r45_leg_x_five"] = df["r45_leg_rate_diff"] * (total_rds_o / 3.0)
        # High R45 leg rate + existing leg escalation = full-fight leg specialist
        r_leg_esc_o = _o("r_pre_leg_escalation")
        b_leg_esc_o = _o("b_pre_leg_escalation")
        df["r45_leg_volume_diff"] = (r_r45_leg + r_leg_esc_o) - (
            b_r45_leg + b_leg_esc_o
        )

        # ── R4/R5 TD accuracy ────────────────────────────────────────────────
        r_r45_tda = _o("r_pre_r45_td_acc")
        b_r45_tda = _o("b_pre_r45_td_acc")
        df["r45_td_acc_diff"] = r_r45_tda - b_r45_tda
        # Championship TD accuracy × championship ctrl time = late-round grappling dominance
        r_r45_ctrl_o = _o("r_pre_r45_ctrl")
        b_r45_ctrl_o = _o("b_pre_r45_ctrl")
        df["r45_td_ctrl_composite"] = (r_r45_tda * r_r45_ctrl_o) - (
            b_r45_tda * b_r45_ctrl_o
        )
        df["r45_td_acc_x_five"] = df["r45_td_acc_diff"] * (total_rds_o / 3.0)

        # ── Per-round reversal escalation ─────────────────────────────────────
        r_r1_rev = _o("r_pre_r1_rev_rate")
        b_r1_rev = _o("b_pre_r1_rev_rate")
        r_rev_esc = _o("r_pre_late_rev_escalation")
        b_rev_esc = _o("b_pre_late_rev_escalation")
        df["r1_rev_rate_diff"] = r_r1_rev - b_r1_rev
        df["late_rev_escalation_diff"] = r_rev_esc - b_rev_esc
        # Scramble escalation × grapple chain: elite late-round scrambler who grapples better
        df["late_scramble_diff"] = (r_rev_esc * _o("r_pre_grapple_chain")) - (
            b_rev_esc * _o("b_pre_grapple_chain")
        )
        # R1 reversal rate vs R1 TD pressure: escape artist who stuffs early takedowns
        df["r1_rev_vs_td_pressure"] = r_r1_rev * _o(
            "b_pre_r1_td_pressure"
        ) - b_r1_rev * _o("r_pre_r1_td_pressure")

        # ── R2 conditional adjustment scores ─────────────────────────────────
        r_r2_adj_loss = _o("r_pre_r2_adj_score_loss")
        b_r2_adj_loss = _o("b_pre_r2_adj_score_loss")
        r_r2_adj_win = _o("r_pre_r2_adj_score_win")
        b_r2_adj_win = _o("b_pre_r2_adj_score_win")
        df["r2_adj_loss_diff"] = r_r2_adj_loss - b_r2_adj_loss
        df["r2_adj_win_diff"] = r_r2_adj_win - b_r2_adj_win
        # Tactically complete fighter: surges when behind AND presses when ahead
        df["r_tactical_complete"] = r_r2_adj_loss.clip(lower=0) * r_r2_adj_win.clip(
            lower=0
        )
        df["b_tactical_complete"] = b_r2_adj_loss.clip(lower=0) * b_r2_adj_win.clip(
            lower=0
        )
        df["tactical_complete_diff"] = (
            df["r_tactical_complete"] - df["b_tactical_complete"]
        )
        # Adjustor when losing × comeback rate: true resilient fighter
        df["resilient_adjustor_diff"] = (
            r_r2_adj_loss.clip(lower=0) * _o("r_pre_comeback_rate")
        ) - (b_r2_adj_loss.clip(lower=0) * _o("b_pre_comeback_rate"))

        # ── Cardio endurance features ─────────────────────────────────────────
        r_decay = _o("r_pre_output_decay_rate")
        b_decay = _o("b_pre_output_decay_rate")
        r_r2_ret = _o("r_pre_r2_retention", 1.0)
        b_r2_ret = _o("b_pre_r2_retention", 1.0)
        r_r3_ret = _o("r_pre_r3_retention", 1.0)
        b_r3_ret = _o("b_pre_r3_retention", 1.0)
        r_acc_ret = _o("r_pre_accuracy_retention", 1.0)
        b_acc_ret = _o("b_pre_accuracy_retention", 1.0)
        r_pv = _o("r_pre_pace_variance_r123")
        b_pv = _o("b_pre_pace_variance_r123")
        r_allcard = _o("r_pre_all_round_cardio")
        b_allcard = _o("b_pre_all_round_cardio")
        r_endcomp = _o("r_pre_endurance_composite")
        b_endcomp = _o("b_pre_endurance_composite")

        df["output_decay_rate_diff"] = r_decay - b_decay
        df["r2_retention_diff"] = r_r2_ret - b_r2_ret
        df["r3_retention_diff"] = r_r3_ret - b_r3_ret
        df["accuracy_retention_diff"] = r_acc_ret - b_acc_ret
        df["pace_variance_r123_diff"] = r_pv - b_pv
        df["all_round_cardio_diff"] = r_allcard - b_allcard
        df["endurance_composite_diff"] = r_endcomp - b_endcomp

        # Full endurance edge: all three retention axes combined
        df["r_endurance_edge"] = r_r2_ret * r_r3_ret * r_acc_ret
        df["b_endurance_edge"] = b_r2_ret * b_r3_ret * b_acc_ret
        df["endurance_edge_diff"] = df["r_endurance_edge"] - df["b_endurance_edge"]

        # Decayer vs builder: negative-slope fighter facing a positive-slope opponent
        df["decay_vs_builder"] = (-r_decay).clip(lower=0) * b_decay.clip(lower=0) - (
            -b_decay
        ).clip(lower=0) * r_decay.clip(lower=0)

        # 5-round cardio amplification
        df["all_round_cardio_x_five"] = df["all_round_cardio_diff"] * (
            total_rds_o / 3.0
        )
        df["r3_retention_x_five"] = df["r3_retention_diff"] * \
            (total_rds_o / 3.0)

        # Quality endurance: retains accuracy AND output — a class above mere cardio
        df["quality_endurance_diff"] = (r_acc_ret * r_r3_ret * r_cardio_o) - (
            b_acc_ret * b_r3_ret * b_cardio_o
        )

        # Pace consistency edge: low variance = reliable output, harder to game-plan against
        df["pace_consistency_edge"] = (
            1.0 / (r_pv + 1.0)) - (1.0 / (b_pv + 1.0))

        # R2 fade vs R2 surge: opponent who drops output in R2 facing a fighter who surges
        df["r2_fade_vs_surge"] = (b_r2_ret - r_r2_ret) * \
            (b_r2_adj_loss - r_r2_adj_loss)

        # Endurance under pressure: retains accuracy while absorbing opponent's escalating volume
        r_abs_slope_o = _o("r_pre_absorption_slope")
        b_abs_slope_o = _o("b_pre_absorption_slope")
        df["r_endurance_under_pressure"] = (
            r_acc_ret
            * r_r3_ret
            * (1.0 - r_abs_slope_o.clip(lower=0) / 10.0).clip(lower=0)
        )
        df["b_endurance_under_pressure"] = (
            b_acc_ret
            * b_r3_ret
            * (1.0 - b_abs_slope_o.clip(lower=0) / 10.0).clip(lower=0)
        )
        df["endurance_under_pressure_diff"] = (
            df["r_endurance_under_pressure"] - df["b_endurance_under_pressure"]
        )

        # ── Tier 12p: Novel per-round composites + strong-signal combinations ────
        self._log(
            "Tier 12p: Round momentum sequences, quality output, punishment response, terminal surge, Tier-C composites..."
        )

        def _p(col, default=0.0):
            return df.get(col, pd.Series(default, index=df.index)).fillna(default)

        # ── Round conditional momentum sequences ──────────────────────────────────
        r_mom_r12w = _p("r_pre_momentum_r12_win", 0.5)
        b_mom_r12w = _p("b_pre_momentum_r12_win", 0.5)
        r_mom_r12l = _p("r_pre_momentum_r12_loss", 0.5)
        b_mom_r12l = _p("b_pre_momentum_r12_loss", 0.5)
        r_mom_r23w = _p("r_pre_momentum_r23_win", 0.5)
        b_mom_r23w = _p("b_pre_momentum_r23_win", 0.5)
        r_mom_r23l = _p("r_pre_momentum_r23_loss", 0.5)
        b_mom_r23l = _p("b_pre_momentum_r23_loss", 0.5)

        # Sustained-pressure fighter: wins R2 after winning R1
        df["momentum_r12_win_diff"] = r_mom_r12w - b_mom_r12w
        # Reset/comeback fighter: wins R2 even after losing R1
        df["momentum_r12_loss_diff"] = r_mom_r12l - b_mom_r12l
        df["momentum_r23_win_diff"] = r_mom_r23w - b_mom_r23w
        df["momentum_r23_loss_diff"] = r_mom_r23l - b_mom_r23l
        # Sustained pressure composite: carries momentum across both transitions
        df["r_sustained_momentum"] = r_mom_r12w * r_mom_r23w
        df["b_sustained_momentum"] = b_mom_r12w * b_mom_r23w
        df["sustained_momentum_diff"] = (
            df["r_sustained_momentum"] - df["b_sustained_momentum"]
        )
        # Resilient reset composite: bounces back across both transitions
        df["r_resilient_reset"] = r_mom_r12l * r_mom_r23l
        df["b_resilient_reset"] = b_mom_r12l * b_mom_r23l
        df["resilient_reset_diff"] = df["r_resilient_reset"] - \
            df["b_resilient_reset"]
        # Sustained vs. reset asymmetry: fighter who presses differs from one who resets
        df["momentum_asymmetry_diff"] = (r_mom_r12w - r_mom_r12l) - (
            b_mom_r12w - b_mom_r12l
        )
        # Five-round amplification — momentum patterns are even more decisive over 5 rounds
        total_rds_p = df.get("total_rounds", pd.Series(
            3, index=df.index)).fillna(3)
        df["sustained_momentum_x5"] = df["sustained_momentum_diff"] * (
            total_rds_p / 3.0
        )
        df["resilient_reset_x5"] = df["resilient_reset_diff"] * \
            (total_rds_p / 3.0)

        # ── Quality-adjusted round output ─────────────────────────────────────────
        r_r1qo = _p("r_pre_r1_quality_output")
        b_r1qo = _p("b_pre_r1_quality_output")
        r_r3qo = _p("r_pre_r3_quality_output")
        b_r3qo = _p("b_pre_r3_quality_output")
        r_qratio = _p("r_pre_quality_output_ratio", 1.0)
        b_qratio = _p("b_pre_quality_output_ratio", 1.0)
        df["r1_quality_output_diff"] = r_r1qo - b_r1qo
        df["r3_quality_output_diff"] = r_r3qo - b_r3qo
        df["quality_output_ratio_diff"] = r_qratio - b_qratio
        # Late quality advantage: fighter who maintains BOTH volume AND accuracy in R3
        df["r3_quality_edge"] = r_r3qo - b_r3qo
        # Quality retention vs cardio: captures the dimensions orthogonally
        r_ci_p = _p("r_pre_cardio_index", 0.75)
        b_ci_p = _p("b_pre_cardio_index", 0.75)
        df["quality_x_cardio_diff"] = (r_qratio * r_ci_p) - (b_qratio * b_ci_p)
        # Quality drop interaction: fighter whose R3 quality is high vs opponent who drops
        df["quality_cliff_edge"] = r_r3qo * (
            1.0 - b_qratio.clip(upper=1.5) / 1.5
        ) - b_r3qo * (1.0 - r_qratio.clip(upper=1.5) / 1.5)

        # ── Output under punishment ───────────────────────────────────────────────
        r_ovp = _p("r_pre_output_vs_punishment")
        b_ovp = _p("b_pre_output_vs_punishment")
        df["output_vs_punishment_diff"] = r_ovp - b_ovp
        # Heart × chin: fights back when hurt AND doesn't get stopped
        r_kd_abs = _p("r_pre_kd_absorbed_rate")
        b_kd_abs = _p("b_pre_kd_absorbed_rate")
        df["r_heart_vs_chin"] = r_ovp * (1.0 - r_kd_abs.clip(upper=1.0))
        df["b_heart_vs_chin"] = b_ovp * (1.0 - b_kd_abs.clip(upper=1.0))
        df["heart_vs_chin_diff"] = df["r_heart_vs_chin"] - df["b_heart_vs_chin"]
        # Punishment response vs endurance: fights back AND doesn't fade
        r_end_p = _p("r_pre_endurance_composite")
        b_end_p = _p("b_pre_endurance_composite")
        df["punishment_endurance_diff"] = (r_ovp * r_end_p) - (b_ovp * b_end_p)

        # ── R5/R4 terminal surge ──────────────────────────────────────────────────
        r_surge = _p("r_pre_r5_r4_surge")
        b_surge = _p("b_pre_r5_r4_surge")
        df["r5_r4_surge_diff"] = r_surge - b_surge
        # Terminal surge × championship rounds only
        df["r5_r4_surge_x5"] = df["r5_r4_surge_diff"] * (total_rds_p / 3.0)
        # Surge advantage stacked on top of all_round_cardio (endurance + final gear)
        r_arc = _p("r_pre_all_round_cardio")
        b_arc = _p("b_pre_all_round_cardio")
        df["r_championship_endurance"] = r_arc * r_surge.clip(lower=0.5)
        df["b_championship_endurance"] = b_arc * b_surge.clip(lower=0.5)
        df["championship_endurance_diff"] = (
            df["r_championship_endurance"] - df["b_championship_endurance"]
        )

        # ── Sub escalation ratio ──────────────────────────────────────────────────
        r_subesc = _p("r_pre_sub_escalation_ratio", 1.0)
        b_subesc = _p("b_pre_sub_escalation_ratio", 1.0)
        df["sub_escalation_ratio_diff"] = r_subesc - b_subesc
        # Sub escalation × grapple depth: ratchets up subs AND has the chain to finish
        r_gd4_p = _p("r_pre_grapple_depth_4")
        b_gd4_p = _p("b_pre_grapple_depth_4")
        df["sub_esc_x_grapple_depth"] = (
            r_subesc * r_gd4_p) - (b_subesc * b_gd4_p)
        # Sub escalation × championship rounds: grappler pressure builds into R4/R5
        df["sub_esc_x5"] = df["sub_escalation_ratio_diff"] * \
            (total_rds_p / 3.0)

        # ── Burst index ───────────────────────────────────────────────────────────
        r_burst = _p("r_pre_burst_index", 1.0 / 3.0)
        b_burst = _p("b_pre_burst_index", 1.0 / 3.0)
        df["burst_index_diff"] = r_burst - b_burst
        # Burst fighter vs. decaying opponent: front-loader against someone who fades
        r_decay_p = _p("r_pre_output_decay_rate")
        b_decay_p = _p("b_pre_output_decay_rate")
        df["burst_vs_fader"] = r_burst * (-b_decay_p).clip(lower=0) - b_burst * (
            -r_decay_p
        ).clip(lower=0)
        # Clash: burst fighter facing a back-loaded builder
        df["burst_vs_builder"] = (r_burst - b_burst) * (b_qratio - r_qratio)

        # ── Ctrl GnP efficiency slope ─────────────────────────────────────────────
        r_gnps = _p("r_pre_ctrl_gnp_slope")
        b_gnps = _p("b_pre_ctrl_gnp_slope")
        df["ctrl_gnp_slope_diff"] = r_gnps - b_gnps
        # GnP efficiency improvement × championship rounds = dominant late grappler
        df["ctrl_gnp_slope_x5"] = df["ctrl_gnp_slope_diff"] * \
            (total_rds_p / 3.0)
        # GnP slope × grapple depth: compound late-round grappling dominance
        df["gnp_slope_x_depth_diff"] = (r_gnps * r_gd4_p) - (b_gnps * b_gd4_p)

        # ── R3 striking under TD threat ───────────────────────────────────────────
        r_str_td = _p("r_pre_r3_str_vs_td_pressure")
        b_str_td = _p("b_pre_r3_str_vs_td_pressure")
        df["r3_str_vs_td_pressure_diff"] = r_str_td - b_str_td
        # Striker maintaining output even when the opponent was shooting: style clash
        r_strdef_p = _p("r_pre_str_def", 0.5)
        b_strdef_p = _p("b_pre_str_def", 0.5)
        df["str_td_resistance"] = (
            r_str_td * r_strdef_p) - (b_str_td * b_strdef_p)

        # ── Tier C: Composites of already-strong signals ──────────────────────────
        # physical_style_dominance × cardio_index: dominant physical fighter who
        # doesn't fade — top RF and CAT feature × strong LGB/CAT feature
        r_psd_p = _p("r_physical_style_dominance")
        b_psd_p = _p("b_physical_style_dominance")
        df["r_phys_dom_cardio"] = r_psd_p * r_ci_p
        df["b_phys_dom_cardio"] = b_psd_p * b_ci_p
        df["phys_dom_cardio_diff"] = df["r_phys_dom_cardio"] - \
            df["b_phys_dom_cardio"]

        # endurance_composite × grapple_depth_4: grappler who maintains multi-stage
        # chain pressure and doesn't tire out of position
        df["endurance_grapple_depth_diff"] = (
            r_end_p * r_gd4_p) - (b_end_p * b_gd4_p)

        # Body-KO pipeline: R2 body work × R3 KD rate = setup-then-finish archetype
        r_r2br = _p("r_pre_r2_body_rate")
        b_r2br = _p("b_pre_r2_body_rate")
        r_lkd = _p("r_pre_late_kd_rate")
        b_lkd = _p("b_pre_late_kd_rate")
        df["body_ko_pipeline_diff"] = (r_r2br * r_lkd) - (b_r2br * b_lkd)
        # Pipeline × output under punishment: body-KO fighter who also fights back
        df["body_ko_heart_diff"] = (
            r_r2br * r_lkd * r_ovp) - (b_r2br * b_lkd * b_ovp)

        # Glicko stability × endurance composite: reliable Glicko rating + doesn't
        # fade. Mirrors r_glicko_pre_vol_inv (top LGB) × endurance.
        r_gvol = _p("r_glicko_pre_vol", 0.06)
        b_gvol = _p("b_glicko_pre_vol", 0.06)
        r_gvol_inv = 1.0 / (r_gvol + 0.06)
        b_gvol_inv = 1.0 / (b_gvol + 0.06)
        df["r_glicko_endurance"] = r_gvol_inv * r_end_p
        df["b_glicko_endurance"] = b_gvol_inv * b_end_p
        df["glicko_endurance_diff"] = (
            df["r_glicko_endurance"] - df["b_glicko_endurance"]
        )
        # Glicko stability × quality output retention
        df["glicko_quality_output_diff"] = (r_gvol_inv * r_qratio) - (
            b_gvol_inv * b_qratio
        )

        # common_opp_edge × r2_win_rate: beats same opponents AND comes alive in R2
        coe_p = _p("common_opp_edge")
        r_r2wr_p = _p("r_pre_r2_win_rate", 0.5)
        b_r2wr_p = _p("b_pre_r2_win_rate", 0.5)
        df["coe_x_r2_win_diff"] = coe_p * (r_r2wr_p - b_r2wr_p)
        # common_opp_edge × sustained_momentum: consistent performer AND chains rounds
        df["coe_x_momentum_diff"] = coe_p * df["sustained_momentum_diff"]

        # R1 head pressure × output_decay_rate (signed): opponent who applies heavy
        # R1 head pressure but fades — the R1 threat becomes manageable late
        r_r1hp_p = _p("r_pre_r1_head_pressure")
        b_r1hp_p = _p("b_pre_r1_head_pressure")
        # Negative decay = fading fighter; heavy R1 head pressure × fader = early-round threat only
        df["r1_pressure_vs_decay_diff"] = (b_r1hp_p * (-b_decay_p).clip(lower=0)) - (
            r_r1hp_p * (-r_decay_p).clip(lower=0)
        )
        # physical_style_dominance × quality_output_ratio: physical dominance that
        # also maintains precision — the elite physical technician
        df["phys_dom_quality_diff"] = (
            r_psd_p * r_qratio) - (b_psd_p * b_qratio)

        self._log("Tier 13: SVD decomposition per feature bucket...")
        _numeric_cols = set(df.select_dtypes(include="number").columns)
        striking_cols = [
            c
            for c in df.columns
            if c in _numeric_cols
            and any(
                x in c
                for x in [
                    "SLpM",
                    "SApM",
                    "sig_str",
                    "str_def",
                    "kd",
                    "head",
                    "body",
                    "leg",
                ]
            )
        ]
        grappling_cols = [
            c
            for c in df.columns
            if c in _numeric_cols
            and any(x in c for x in ["td", "sub", "ctrl", "rev", "ground"])
        ]
        physical_cols = [
            c
            for c in df.columns
            if c in _numeric_cols
            and any(x in c for x in ["height", "reach", "weight", "age", "ape_index"])
        ]
        form_cols = [
            c
            for c in df.columns
            if c in _numeric_cols
            and any(
                x in c
                for x in [
                    "rolling",
                    "streak",
                    "momentum",
                    "finish_rate",
                    "elo",
                    "glicko",
                ]
            )
        ]

        # Fit SVD on training portion only (first 80% of non-Draw fights,
        # matching the chronological split in train()), then transform all rows
        # so downstream tiers can build features from SVD components.
        _non_draw_idx = df[df["winner"] != "Draw"].index
        _svd_train_end = int(len(_non_draw_idx) * 0.80)
        _svd_train_idx = _non_draw_idx[:_svd_train_end]

        def apply_svd(cols, svd_obj, prefix, df, col_store_attr, train_idx):
            if len(cols) < 2:
                return
            X_train = df.loc[train_idx, cols].fillna(0).values
            if X_train.dtype == object:
                raise TypeError(
                    f"SVD '{prefix}': input matrix has object dtype. "
                    f"Non-numeric columns: {[c for c in cols if df[c].dtype == object]}"
                )
            n_comp = min(
                svd_obj.n_components, X_train.shape[1] -
                1, X_train.shape[0] - 1
            )
            if n_comp < 1:
                return
            svd_obj.n_components = n_comp
            svd_obj.fit(X_train)
            X_all = df[cols].fillna(0).values
            X_svd = svd_obj.transform(X_all)
            setattr(self, col_store_attr, cols)
            for i in range(n_comp):
                df[f"{prefix}_svd_{i}"] = X_svd[:, i]

        apply_svd(
            striking_cols,
            self.svd_striking,
            "striking",
            df,
            "svd_striking_cols",
            _svd_train_idx,
        )
        apply_svd(
            grappling_cols,
            self.svd_grappling,
            "grappling",
            df,
            "svd_grappling_cols",
            _svd_train_idx,
        )
        apply_svd(
            physical_cols,
            self.svd_physical,
            "physical",
            df,
            "svd_physical_cols",
            _svd_train_idx,
        )
        apply_svd(form_cols, self.svd_form, "form",
                  df, "svd_form_cols", _svd_train_idx)
        self.svd_fitted = True

        if "is_title_bout" in df.columns:
            df["is_title_enc"] = df["is_title_bout"].astype(int)
        if "total_rounds" in df.columns:
            df["total_rounds_num"] = pd.to_numeric(
                df["total_rounds"], errors="coerce"
            ).fillna(3)

        if "gender" in df.columns:
            df["gender_enc"] = (df["gender"].fillna("").str.lower() == "women").astype(
                int
            )

        self._log("Tier 14: Positional & target differentials...")
        for feat in [
            "distance_pct",
            "clinch_pct",
            "ground_pct",
            "head_pct",
            "body_pct",
            "leg_pct",
        ]:
            r_col, b_col = f"r_pre_{feat}", f"b_pre_{feat}"
            if r_col in df.columns and b_col in df.columns:
                df[f"diff_{feat}"] = df[r_col].fillna(0) - df[b_col].fillna(0)
        df["positional_striking_advantage"] = (
            df.get("diff_distance_pct", pd.Series(0.0, index=df.index))
            .fillna(0.0)
            .abs()
            + df.get("diff_clinch_pct", pd.Series(0.0, index=df.index))
            .fillna(0.0)
            .abs()
            + df.get("diff_ground_pct", pd.Series(0.0, index=df.index))
            .fillna(0.0)
            .abs()
        )
        df["target_distribution_advantage"] = (
            df.get("diff_head_pct", pd.Series(
                0.0, index=df.index)).fillna(0.0).abs()
            + df.get("diff_body_pct", pd.Series(0.0,
                     index=df.index)).fillna(0.0).abs()
            + df.get("diff_leg_pct", pd.Series(0.0,
                     index=df.index)).fillna(0.0).abs()
        )

        self._log("Tier 15: Defense differentials...")
        df["diff_str_def"] = df.get(
            "r_pre_str_def", pd.Series(0, index=df.index)
        ).fillna(0) - df.get("b_pre_str_def", pd.Series(0, index=df.index)).fillna(0)
        df["diff_td_def"] = df.get("r_pre_td_def", pd.Series(0, index=df.index)).fillna(
            0
        ) - df.get("b_pre_td_def", pd.Series(0, index=df.index)).fillna(0)
        df["defensive_composite"] = df["diff_str_def"].fillna(0) + df[
            "diff_td_def"
        ].fillna(0)

        self._log("Tier 16: Deep interaction features...")
        elo_d = df.get("elo_diff", pd.Series(0, index=df.index)).fillna(0)
        form3 = df.get("diff_rolling3_wins", pd.Series(
            0, index=df.index)).fillna(0)
        wlr = df.get("diff_win_loss_ratio", pd.Series(
            0, index=df.index)).fillna(0)
        fin_r = df.get("diff_finish_rate", pd.Series(
            0, index=df.index)).fillna(0)
        kd_abs = df.get("diff_kd_absorbed", pd.Series(
            0, index=df.index)).fillna(0)

        df["elo_x_form"] = elo_d * form3
        df["elo_x_win_ratio"] = elo_d * wlr
        df["elo_x_finish"] = elo_d * fin_r
        df["elo_x_durability"] = elo_d * kd_abs.abs()

        reach_d = df.get("diff_reach", pd.Series(0, index=df.index)).fillna(0)
        height_d = df.get("diff_height", pd.Series(
            0, index=df.index)).fillna(0)
        slpm_d = df.get("diff_pre_SLpM", pd.Series(
            0, index=df.index)).fillna(0)
        td_d = df.get("diff_pre_td_avg", pd.Series(
            0, index=df.index)).fillna(0)
        acc_d = df.get("diff_pre_sig_str_acc",
                       pd.Series(0, index=df.index)).fillna(0)
        age_d = df.get("diff_age_at_event", pd.Series(
            0, index=df.index)).fillna(0)
        streak_d = df.get("diff_pre_win_streak",
                          pd.Series(0, index=df.index)).fillna(0)
        exp_gap = df.get("diff_pre_total_fights", pd.Series(0, index=df.index)).fillna(
            0
        )

        df["reach_x_striking"] = reach_d * slpm_d
        df["height_x_reach"] = height_d * reach_d
        df["physical_x_striking"] = (height_d + reach_d) * slpm_d

        df["age_x_striking"] = age_d * slpm_d
        df["age_x_grappling"] = age_d * td_d
        df["age_x_durability"] = age_d * kd_abs.abs()
        df["age_x_win_streak"] = age_d * streak_d
        df["experience_x_age"] = exp_gap * age_d

        str_def_d = df.get("diff_str_def", pd.Series(
            0, index=df.index)).fillna(0)
        td_def_d = df.get("diff_td_def", pd.Series(
            0, index=df.index)).fillna(0)
        sub_rate_d = df.get("sub_threat_diff", pd.Series(
            0, index=df.index)).fillna(0)
        df["td_x_defense"] = td_d * td_def_d
        df["submission_x_grappling"] = sub_rate_d * td_d

        df["striking_x_accuracy"] = slpm_d * acc_d
        df["striking_x_defense"] = slpm_d * str_def_d
        df["ko_power_x_striking"] = (
            df.get("ko_threat_diff", pd.Series(
                0, index=df.index)).fillna(0) * slpm_d
        )

        momentum = df.get("momentum_diff_3", pd.Series(
            0, index=df.index)).fillna(0)
        df["momentum_x_win_streak"] = momentum * streak_d
        df["form_x_experience"] = form3 * exp_gap
        df["finish_x_momentum"] = fin_r * momentum
        df["form_x_durability"] = form3 * kd_abs.abs()

        df["elite_finisher"] = elo_d * fin_r * form3
        df["unstoppable_streak"] = streak_d * momentum * form3
        df["veteran_advantage"] = wlr * exp_gap * (-age_d)

        self._log("Tier 17: Extended polynomial features...")
        poly_extended = [
            "elo_diff",
            "glicko_diff",
            "diff_win_loss_ratio",
            "diff_age_at_event",
            "diff_reach",
            "diff_height",
            "diff_pre_SLpM",
            "diff_pre_sig_str_acc",
            "diff_pre_td_avg",
            "diff_pre_win_streak",
            "diff_finish_rate",
            "diff_pre_loss_streak",
            "diff_str_def",
            "diff_td_def",
            "diff_pre_kd_rate",
            "diff_pre_ctrl_avg",
            "elo_x_form",
            "streak_x_finish",
            "striking_exchange",
            "diff_distance_pct",
            "diff_clinch_pct",
            "diff_ground_pct",
        ]
        for col in poly_extended:
            if col in df.columns:
                df[f"{col}_sq"] = df[col] ** 2
                df[f"{col}_abs"] = df[col].abs()
        if "diff_age_at_event" in df.columns:
            df["diff_age_cubed"] = df["diff_age_at_event"] ** 3

        self._log("Tier 18: Opponent-adjusted performance...")
        if "r_vs_elite_win_rate" in df.columns and "b_vs_elite_win_rate" in df.columns:
            df["diff_win_rate_vs_elite"] = (
                df["r_vs_elite_win_rate"] - df["b_vs_elite_win_rate"]
            )
            df["diff_win_rate_vs_strikers"] = df.get(
                "r_vs_striker_win_rate", pd.Series(0, index=df.index)
            ).fillna(0) - df.get(
                "b_vs_striker_win_rate", pd.Series(0, index=df.index)
            ).fillna(0)
            df["diff_win_rate_vs_grapplers"] = df.get(
                "r_vs_grappler_win_rate", pd.Series(0, index=df.index)
            ).fillna(0) - df.get(
                "b_vs_grappler_win_rate", pd.Series(0, index=df.index)
            ).fillna(0)
            df["championship_readiness"] = df["diff_win_rate_vs_elite"] * df.get(
                "elo_diff", pd.Series(0.0, index=df.index)
            ).fillna(0.0)

        self._log("Tier 19: Career pattern features...")
        if "r_pre_early_finish_rate" in df.columns:
            df["diff_early_finish_rate"] = df.get(
                "r_pre_early_finish_rate", pd.Series(0.0, index=df.index)
            ).fillna(0) - df.get(
                "b_pre_early_finish_rate", pd.Series(0.0, index=df.index)
            ).fillna(0)
            df["diff_late_finish_rate"] = df.get(
                "r_pre_late_finish_rate", pd.Series(0, index=df.index)
            ).fillna(0) - df.get(
                "b_pre_late_finish_rate", pd.Series(0, index=df.index)
            ).fillna(0)
            df["diff_first_round_ko_rate"] = df.get(
                "r_pre_first_round_ko_rate", pd.Series(0, index=df.index)
            ).fillna(0) - df.get(
                "b_pre_first_round_ko_rate", pd.Series(0, index=df.index)
            ).fillna(0)
        if "r_pre_five_round_fights" in df.columns:
            df["diff_five_round_fights"] = df.get(
                "r_pre_five_round_fights", pd.Series(0.0, index=df.index)
            ).fillna(0) - df.get(
                "b_pre_five_round_fights", pd.Series(0.0, index=df.index)
            ).fillna(0)

        # Removed: prime_score — hardcoded peak age 29.5; raw age_at_event is already a feature.

        if "r_fights_since_peak" in df.columns:
            df["diff_fights_since_peak"] = df.get(
                "r_fights_since_peak", pd.Series(0.0, index=df.index)
            ).fillna(0) - df.get(
                "b_fights_since_peak", pd.Series(0.0, index=df.index)
            ).fillna(0)
            df["declining_phase_diff"] = df["diff_fights_since_peak"]

        if "r_last_fight_was_win" in df.columns:
            df["r_last_fight_momentum"] = df.get(
                "r_last_fight_was_win", pd.Series(0.0, index=df.index)
            ).fillna(0).astype(float) + df.get(
                "r_last_fight_was_finish", pd.Series(0.0, index=df.index)
            ).fillna(0).astype(float)
            df["b_last_fight_momentum"] = df.get(
                "b_last_fight_was_win", pd.Series(0.0, index=df.index)
            ).fillna(0).astype(float) + df.get(
                "b_last_fight_was_finish", pd.Series(0.0, index=df.index)
            ).fillna(0).astype(float)
            df["last_fight_momentum_diff"] = (
                df["r_last_fight_momentum"] - df["b_last_fight_momentum"]
            )

        for feat in ["wins", "sig_str", "td", "kd", "finishes"]:
            r_col10 = f"r_rolling10_{feat}"
            b_col10 = f"b_rolling10_{feat}"
            if r_col10 in df.columns and b_col10 in df.columns:
                df[f"diff_rolling10_{feat}"] = df[r_col10].fillna(0) - df[
                    b_col10
                ].fillna(0)

        self._log("Tier 20: Rounds-based strategy features...")
        total_rds_t20 = df.get("total_rounds_num", pd.Series(3, index=df.index)).fillna(
            3
        )
        dec_rate_d_t20 = df.get(
            "dec_tendency_diff", pd.Series(0, index=df.index)
        ).fillna(0)
        if "diff_finish_rate" in df.columns:
            df["rounds_x_cardio"] = total_rds_t20 * dec_rate_d_t20
            df["rounds_x_finish_rate"] = (5 - total_rds_t20) * df[
                "diff_finish_rate"
            ].fillna(0)
        kd_abs2_t20 = df.get(
            "chin_deterioration_diff", pd.Series(0, index=df.index)
        ).fillna(0)
        df["rounds_x_durability"] = total_rds_t20 * kd_abs2_t20

        self._log("Tier 21: Matchup-specific features...")
        r_slpm_t21 = df.get("r_pre_SLpM", pd.Series(
            3.0, index=df.index)).fillna(3.0)
        b_slpm_t21 = df.get("b_pre_SLpM", pd.Series(
            3.0, index=df.index)).fillna(3.0)
        r_acc_t21 = df.get("r_pre_sig_str_acc", pd.Series(0.45, index=df.index)).fillna(
            0.45
        )
        b_acc_t21 = df.get("b_pre_sig_str_acc", pd.Series(0.45, index=df.index)).fillna(
            0.45
        )
        r_str_def_t21 = df.get("r_pre_str_def", pd.Series(0.55, index=df.index)).fillna(
            0.55
        )
        b_str_def_t21 = df.get("b_pre_str_def", pd.Series(0.55, index=df.index)).fillna(
            0.55
        )
        r_td_t21 = df.get("r_pre_td_avg", pd.Series(
            1.5, index=df.index)).fillna(1.5)
        b_td_t21 = df.get("b_pre_td_avg", pd.Series(
            1.5, index=df.index)).fillna(1.5)
        r_td_def_t21 = df.get("r_pre_td_def", pd.Series(0.65, index=df.index)).fillna(
            0.65
        )
        b_td_def_t21 = df.get("b_pre_td_def", pd.Series(0.65, index=df.index)).fillna(
            0.65
        )
        r_sub_avg_t21 = df.get(
            "r_pre_sub_att_rate", pd.Series(0.3, index=df.index)
        ).fillna(0.3)
        b_sub_avg_t21 = df.get(
            "b_pre_sub_att_rate", pd.Series(0.3, index=df.index)
        ).fillna(0.3)
        r_td_acc_t21 = df.get("r_pre_td_acc", pd.Series(0.4, index=df.index)).fillna(
            0.4
        )
        b_td_acc_t21 = df.get("b_pre_td_acc", pd.Series(0.4, index=df.index)).fillna(
            0.4
        )
        r_ctrl_t21 = df.get("r_pre_ctrl_avg", pd.Series(
            60, index=df.index)).fillna(60)
        b_ctrl_t21 = df.get("b_pre_ctrl_avg", pd.Series(
            60, index=df.index)).fillna(60)
        r_sub_rate_t21 = df.get(
            "r_pre_sub_att_rate", pd.Series(0.3, index=df.index)
        ).fillna(0.3)
        b_sub_rate_t21 = df.get(
            "b_pre_sub_att_rate", pd.Series(0.3, index=df.index)
        ).fillna(0.3)

        df["r_striking_vs_b_defense"] = r_slpm_t21 * (1.0 - b_str_def_t21)
        df["b_striking_vs_r_defense"] = b_slpm_t21 * (1.0 - r_str_def_t21)
        df["striking_exploitation_diff"] = (
            df["r_striking_vs_b_defense"] - df["b_striking_vs_r_defense"]
        )

        df["r_td_vs_b_td_defense"] = r_td_t21 * (1.0 - b_td_def_t21)
        df["b_td_vs_r_td_defense"] = b_td_t21 * (1.0 - r_td_def_t21)
        df["td_exploitation_diff"] = (
            df["r_td_vs_b_td_defense"] - df["b_td_vs_r_td_defense"]
        )

        df["r_sub_setup_efficiency"] = r_sub_rate_t21 * r_td_acc_t21
        df["b_sub_setup_efficiency"] = b_sub_rate_t21 * b_td_acc_t21
        df["sub_setup_diff"] = (
            df["r_sub_setup_efficiency"] - df["b_sub_setup_efficiency"]
        )
        df["r_sub_threat_vs_td_defense"] = r_sub_avg_t21 * (1.0 - b_td_def_t21)
        df["b_sub_threat_vs_td_defense"] = b_sub_avg_t21 * (1.0 - r_td_def_t21)
        df["sub_threat_vs_defense_diff"] = (
            df["r_sub_threat_vs_td_defense"] - df["b_sub_threat_vs_td_defense"]
        )

        df["r_striking_quality"] = r_slpm_t21 * r_acc_t21
        df["b_striking_quality"] = b_slpm_t21 * b_acc_t21
        df["striking_quality_diff"] = (
            df["r_striking_quality"] - df["b_striking_quality"]
        )
        df["r_accuracy_under_fire"] = r_acc_t21 / (b_slpm_t21 + 0.1)
        df["b_accuracy_under_fire"] = b_acc_t21 / (r_slpm_t21 + 0.1)
        df["accuracy_under_fire_diff"] = (
            df["r_accuracy_under_fire"] - df["b_accuracy_under_fire"]
        )

        self._log("Tier 22: Statistical ratio features...")
        r_sapm_t22 = df.get("r_pre_SApM", pd.Series(
            3.0, index=df.index)).fillna(3.0)
        b_sapm_t22 = df.get("b_pre_SApM", pd.Series(
            3.0, index=df.index)).fillna(3.0)

        df["r_damage_ratio"] = r_slpm_t21 / (r_sapm_t22 + 0.1)
        df["b_damage_ratio"] = b_slpm_t21 / (b_sapm_t22 + 0.1)
        df["damage_ratio_diff"] = df["r_damage_ratio"] - df["b_damage_ratio"]

        df["r_striking_output_quality"] = r_slpm_t21 * \
            r_acc_t21 / (r_sapm_t22 + 0.1)
        df["b_striking_output_quality"] = b_slpm_t21 * \
            b_acc_t21 / (b_sapm_t22 + 0.1)
        df["striking_output_quality_diff"] = (
            df["r_striking_output_quality"] - df["b_striking_output_quality"]
        )

        df["r_grappling_quality"] = r_td_t21 * \
            r_td_acc_t21 * (r_ctrl_t21 / 60.0)
        df["b_grappling_quality"] = b_td_t21 * \
            b_td_acc_t21 * (b_ctrl_t21 / 60.0)
        df["grappling_quality_diff"] = (
            df["r_grappling_quality"] - df["b_grappling_quality"]
        )

        df["r_total_defense_index"] = r_str_def_t21 * r_td_def_t21
        df["b_total_defense_index"] = b_str_def_t21 * b_td_def_t21
        df["total_defense_diff"] = (
            df["r_total_defense_index"] - df["b_total_defense_index"]
        )

        df["r_complete_fighter_index"] = (
            (r_slpm_t21 + r_td_t21 + r_sub_avg_t21) *
            r_str_def_t21 * r_td_def_t21
        )
        df["b_complete_fighter_index"] = (
            (b_slpm_t21 + b_td_t21 + b_sub_avg_t21) *
            b_str_def_t21 * b_td_def_t21
        )
        df["complete_fighter_diff"] = (
            df["r_complete_fighter_index"] - df["b_complete_fighter_index"]
        )

        df["r_pressure_index"] = r_slpm_t21 * r_td_t21 * (r_ctrl_t21 / 60.0)
        df["b_pressure_index"] = b_slpm_t21 * b_td_t21 * (b_ctrl_t21 / 60.0)
        df["pressure_index_diff"] = df["r_pressure_index"] - df["b_pressure_index"]

        self._log("Tier 23: Extended statistical ratio features...")
        r_wins_t23 = df.get("r_pre_wins", pd.Series(
            5.0, index=df.index)).fillna(5.0)
        b_wins_t23 = df.get("b_pre_wins", pd.Series(
            5.0, index=df.index)).fillna(5.0)
        r_losses_t23 = df.get("r_pre_losses", pd.Series(2.0, index=df.index)).fillna(
            2.0
        )
        b_losses_t23 = df.get("b_pre_losses", pd.Series(2.0, index=df.index)).fillna(
            2.0
        )
        r_streak_t23 = df.get(
            "r_pre_win_streak", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        b_streak_t23 = df.get(
            "b_pre_win_streak", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        r_fr_t23 = df.get("r_pre_finish_rate", pd.Series(0.4, index=df.index)).fillna(
            0.4
        )
        b_fr_t23 = df.get("b_pre_finish_rate", pd.Series(0.4, index=df.index)).fillna(
            0.4
        )
        r_age_t23 = df.get("r_age_at_event", pd.Series(28.0, index=df.index)).fillna(
            28.0
        )
        b_age_t23 = df.get("b_age_at_event", pd.Series(28.0, index=df.index)).fillna(
            28.0
        )
        r_reach_t23 = df.get("r_reach", pd.Series(
            71.0, index=df.index)).fillna(71.0)
        b_reach_t23 = df.get("b_reach", pd.Series(
            71.0, index=df.index)).fillna(71.0)
        r_weight_t23 = df.get("r_weight", pd.Series(155.0, index=df.index)).fillna(
            155.0
        )
        b_weight_t23 = df.get("b_weight", pd.Series(155.0, index=df.index)).fillna(
            155.0
        )
        r_tf_t23 = df.get("r_pre_total_fights", pd.Series(10.0, index=df.index)).fillna(
            10.0
        )
        b_tf_t23 = df.get("b_pre_total_fights", pd.Series(10.0, index=df.index)).fillna(
            10.0
        )

        df["r_defense_offense_balance"] = (
            r_str_def_t21 + 0.01) / (r_acc_t21 + 0.01)
        df["b_defense_offense_balance"] = (
            b_str_def_t21 + 0.01) / (b_acc_t21 + 0.01)
        df["defense_offense_balance_diff"] = (
            df["r_defense_offense_balance"] - df["b_defense_offense_balance"]
        )
        df["r_td_defense_offense_balance"] = (r_td_def_t21 + 0.01) / (
            r_td_acc_t21 + 0.01
        )
        df["b_td_defense_offense_balance"] = (b_td_def_t21 + 0.01) / (
            b_td_acc_t21 + 0.01
        )
        df["td_defense_offense_balance_diff"] = (
            df["r_td_defense_offense_balance"] -
            df["b_td_defense_offense_balance"]
        )
        df["finish_efficiency_diff"] = r_fr_t23 - b_fr_t23

        df["r_precision_striking"] = r_acc_t21 / (r_slpm_t21 + 0.1)
        df["b_precision_striking"] = b_acc_t21 / (b_slpm_t21 + 0.1)
        df["precision_striking_diff"] = (
            df["r_precision_striking"] - df["b_precision_striking"]
        )
        df["r_quality_grappling_23"] = r_td_acc_t21 * (r_td_t21**0.5)
        df["b_quality_grappling_23"] = b_td_acc_t21 * (b_td_t21**0.5)
        df["quality_grappling_diff"] = (
            df["r_quality_grappling_23"] - df["b_quality_grappling_23"]
        )
        df["r_submission_threat_ratio"] = (
            r_sub_avg_t21 + 0.01) / (r_td_t21 + 0.01)
        df["b_submission_threat_ratio"] = (
            b_sub_avg_t21 + 0.01) / (b_td_t21 + 0.01)
        df["submission_threat_ratio_diff"] = (
            df["r_submission_threat_ratio"] - df["b_submission_threat_ratio"]
        )

        df["r_damage_absorption_efficiency"] = r_sapm_t22 / \
            (r_str_def_t21 + 0.01)
        df["b_damage_absorption_efficiency"] = b_sapm_t22 / \
            (b_str_def_t21 + 0.01)
        df["damage_absorption_efficiency_diff"] = (
            df["r_damage_absorption_efficiency"] -
            df["b_damage_absorption_efficiency"]
        )
        df["r_defense_versatility"] = (r_str_def_t21 * r_td_def_t21) ** 0.5
        df["b_defense_versatility"] = (b_str_def_t21 * b_td_def_t21) ** 0.5
        df["defense_versatility_diff"] = (
            df["r_defense_versatility"] - df["b_defense_versatility"]
        )

        df["r_total_offense_index"] = r_slpm_t21 + (r_td_t21 * 1.5)
        df["b_total_offense_index"] = b_slpm_t21 + (b_td_t21 * 1.5)
        df["total_offense_index_diff"] = (
            df["r_total_offense_index"] - df["b_total_offense_index"]
        )
        df["r_offensive_versatility"] = (r_slpm_t21 * r_td_t21) ** 0.5
        df["b_offensive_versatility"] = (b_slpm_t21 * b_td_t21) ** 0.5
        df["offensive_versatility_diff"] = (
            df["r_offensive_versatility"] - df["b_offensive_versatility"]
        )
        df["r_striker_index"] = (r_slpm_t21 + 0.1) / (r_td_t21 + 0.1)
        df["b_striker_index"] = (b_slpm_t21 + 0.1) / (b_td_t21 + 0.1)
        df["striker_index_diff"] = df["r_striker_index"] - df["b_striker_index"]

        r_wlr_t23 = r_wins_t23 / (r_losses_t23 + 1.0)
        b_wlr_t23 = b_wins_t23 / (b_losses_t23 + 1.0)
        df["win_loss_ratio_squared_diff"] = (r_wlr_t23**2) - (b_wlr_t23**2)
        df["r_experience_quality"] = r_wins_t23 / \
            (r_wins_t23 + r_losses_t23 + 1.0)
        df["b_experience_quality"] = b_wins_t23 / \
            (b_wins_t23 + b_losses_t23 + 1.0)
        df["experience_quality_diff"] = (
            df["r_experience_quality"] - df["b_experience_quality"]
        )
        df["r_win_efficiency"] = r_wins_t23 / (r_age_t23 - 18.0 + 1.0)
        df["b_win_efficiency"] = b_wins_t23 / (b_age_t23 - 18.0 + 1.0)
        df["win_efficiency_diff"] = df["r_win_efficiency"] - df["b_win_efficiency"]
        df["r_momentum_quality"] = (r_streak_t23 + 1.0) / (r_wins_t23 + 1.0)
        df["b_momentum_quality"] = (b_streak_t23 + 1.0) / (b_wins_t23 + 1.0)
        df["momentum_quality_diff"] = (
            df["r_momentum_quality"] - df["b_momentum_quality"]
        )

        df["r_reach_efficiency"] = r_slpm_t21 / (r_reach_t23 + 1.0)
        df["b_reach_efficiency"] = b_slpm_t21 / (b_reach_t23 + 1.0)
        df["reach_efficiency_diff"] = (
            df["r_reach_efficiency"] - df["b_reach_efficiency"]
        )
        df["r_size_adjusted_striking"] = r_slpm_t21 / \
            ((r_weight_t23 / 100.0) + 0.01)
        df["b_size_adjusted_striking"] = b_slpm_t21 / \
            ((b_weight_t23 / 100.0) + 0.01)
        df["size_adjusted_striking_diff"] = (
            df["r_size_adjusted_striking"] - df["b_size_adjusted_striking"]
        )
        df["r_size_adjusted_grappling"] = r_td_t21 / \
            ((r_weight_t23 / 100.0) + 0.01)
        df["b_size_adjusted_grappling"] = b_td_t21 / \
            ((b_weight_t23 / 100.0) + 0.01)
        df["size_adjusted_grappling_diff"] = (
            df["r_size_adjusted_grappling"] - df["b_size_adjusted_grappling"]
        )

        df["r_counter_fighter_index"] = (
            r_str_def_t21 + 0.1) / (r_slpm_t21 + 1.0)
        df["b_counter_fighter_index"] = (
            b_str_def_t21 + 0.1) / (b_slpm_t21 + 1.0)
        df["counter_fighter_index_diff"] = (
            df["r_counter_fighter_index"] - df["b_counter_fighter_index"]
        )
        df["r_finishing_threat_composite"] = (
            r_fr_t23 + 0.1) * (r_sub_avg_t21 + 0.1)
        df["b_finishing_threat_composite"] = (
            b_fr_t23 + 0.1) * (b_sub_avg_t21 + 0.1)
        df["finishing_threat_composite_diff"] = (
            df["r_finishing_threat_composite"] -
            df["b_finishing_threat_composite"]
        )
        df["r_complete_geo"] = (
            (r_slpm_t21 + 1.0) * (r_str_def_t21 + 0.1) * (r_fr_t23 + 0.1)
        ) ** (1.0 / 3.0)
        df["b_complete_geo"] = (
            (b_slpm_t21 + 1.0) * (b_str_def_t21 + 0.1) * (b_fr_t23 + 0.1)
        ) ** (1.0 / 3.0)
        df["complete_geo_diff"] = df["r_complete_geo"] - df["b_complete_geo"]
        df["r_pressure_fighter_index"] = (
            r_slpm_t21 + r_td_t21) / (r_str_def_t21 + 0.3)
        df["b_pressure_fighter_index"] = (
            b_slpm_t21 + b_td_t21) / (b_str_def_t21 + 0.3)
        df["pressure_fighter_index_diff"] = (
            df["r_pressure_fighter_index"] - df["b_pressure_fighter_index"]
        )

        _r_roll3 = df.get("r_pre_rolling3_wins", pd.Series(1.5, index=df.index)).fillna(
            1.5
        )
        _b_roll3 = df.get("b_pre_rolling3_wins", pd.Series(1.5, index=df.index)).fillna(
            1.5
        )
        _r_cwr = r_wins_t23 / (r_tf_t23 + 1.0)
        _b_cwr = b_wins_t23 / (b_tf_t23 + 1.0)
        df["r_recent_form_ratio"] = (_r_roll3 / 3.0 + 0.01) / (_r_cwr + 0.01)
        df["b_recent_form_ratio"] = (_b_roll3 / 3.0 + 0.01) / (_b_cwr + 0.01)
        df["recent_form_ratio_diff"] = (
            df["r_recent_form_ratio"] - df["b_recent_form_ratio"]
        )

        _r_ko_d = df.get("r_pre_ko_wins", pd.Series(
            0.0, index=df.index)).fillna(0.0)
        _b_ko_d = df.get("b_pre_ko_wins", pd.Series(
            0.0, index=df.index)).fillna(0.0)
        _r_sub_d = df.get("r_pre_sub_wins", pd.Series(
            0.0, index=df.index)).fillna(0.0)
        _b_sub_d = df.get("b_pre_sub_wins", pd.Series(
            0.0, index=df.index)).fillna(0.0)
        _r_dec_d = df.get("r_pre_dec_wins", pd.Series(
            0.0, index=df.index)).fillna(0.0)
        _b_dec_d = df.get("b_pre_dec_wins", pd.Series(
            0.0, index=df.index)).fillna(0.0)
        df["r_finish_method_diversity"] = (
            (_r_ko_d > 0).astype(float)
            + (_r_sub_d > 0).astype(float)
            + (_r_dec_d > 0).astype(float)
        )
        df["b_finish_method_diversity"] = (
            (_b_ko_d > 0).astype(float)
            + (_b_sub_d > 0).astype(float)
            + (_b_dec_d > 0).astype(float)
        )
        df["finish_method_diversity_diff"] = (
            df["r_finish_method_diversity"] - df["b_finish_method_diversity"]
        )

        df["r_cross_domain_compensation"] = np.maximum(
            0.0, r_td_t21 - 1.5
        ) - np.maximum(0.0, 4.0 - r_slpm_t21)
        df["b_cross_domain_compensation"] = np.maximum(
            0.0, b_td_t21 - 1.5
        ) - np.maximum(0.0, 4.0 - b_slpm_t21)
        df["cross_domain_compensation_index_diff"] = (
            df["r_cross_domain_compensation"] -
            df["b_cross_domain_compensation"]
        )

        self._log("Tier 24: Additional matchup-specific features...")
        df["r_absorption_vuln"] = r_sapm_t22 / (b_slpm_t21 + 0.1)
        df["b_absorption_vuln"] = b_sapm_t22 / (r_slpm_t21 + 0.1)
        df["absorption_vulnerability_index_diff"] = (
            df["r_absorption_vuln"] - df["b_absorption_vuln"]
        )
        df["r_combined_def_hole"] = (
            1.0 - r_str_def_t21) * (1.0 - r_td_def_t21)
        df["b_combined_def_hole"] = (
            1.0 - b_str_def_t21) * (1.0 - b_td_def_t21)
        df["combined_defensive_hole_diff"] = (
            df["r_combined_def_hole"] - df["b_combined_def_hole"]
        )
        df["r_td_pressure_t24"] = (1.0 - r_td_def_t21) * b_td_t21
        df["b_td_pressure_t24"] = (1.0 - b_td_def_t21) * r_td_t21
        df["td_vulnerability_under_pressure_diff"] = (
            df["r_td_pressure_t24"] - df["b_td_pressure_t24"]
        )
        df["r_strike_pressure_t24"] = (1.0 - r_str_def_t21) * b_slpm_t21
        df["b_strike_pressure_t24"] = (1.0 - b_str_def_t21) * r_slpm_t21
        df["strike_defense_under_volume_diff"] = (
            df["r_strike_pressure_t24"] - df["b_strike_pressure_t24"]
        )
        df["r_ctrl_sub_ratio"] = (r_ctrl_t21 / 60.0) / (r_sub_avg_t21 + 0.1)
        df["b_ctrl_sub_ratio"] = (b_ctrl_t21 / 60.0) / (b_sub_avg_t21 + 0.1)
        df["grappling_control_vs_submission_ratio_diff"] = (
            df["r_ctrl_sub_ratio"] - df["b_ctrl_sub_ratio"]
        )
        df["r_sub_def_necessity"] = b_sub_avg_t21 / (r_td_def_t21 + 0.1)
        df["b_sub_def_necessity"] = r_sub_avg_t21 / (b_td_def_t21 + 0.1)
        df["submission_defense_necessity_diff"] = (
            df["r_sub_def_necessity"] - df["b_sub_def_necessity"]
        )
        df["r_strike_synergy"] = (r_slpm_t21 * (r_acc_t21 + 0.01)) ** 0.5
        df["b_strike_synergy"] = (b_slpm_t21 * (b_acc_t21 + 0.01)) ** 0.5
        df["striking_volume_accuracy_synergy_diff"] = (
            df["r_strike_synergy"] - df["b_strike_synergy"]
        )
        df["r_td_paradox"] = (r_td_acc_t21 + 0.01) / (r_td_t21 + 0.5)
        df["b_td_paradox"] = (b_td_acc_t21 + 0.01) / (b_td_t21 + 0.5)
        df["takedown_efficiency_paradox_diff"] = df["r_td_paradox"] - \
            df["b_td_paradox"]
        df["r_total_off_eff"] = (r_slpm_t21 * (r_acc_t21 + 0.01)) ** 0.5 + (
            r_td_t21 * (r_td_acc_t21 + 0.01)
        ) ** 0.5
        df["b_total_off_eff"] = (b_slpm_t21 * (b_acc_t21 + 0.01)) ** 0.5 + (
            b_td_t21 * (b_td_acc_t21 + 0.01)
        ) ** 0.5
        df["total_offensive_efficiency_index_diff"] = (
            df["r_total_off_eff"] - df["b_total_off_eff"]
        )
        df["r_sg_corr"] = (r_slpm_t21 * (r_acc_t21 + 0.01)) / (
            r_td_t21 * (r_td_acc_t21 + 0.01) + 0.1
        )
        df["b_sg_corr"] = (b_slpm_t21 * (b_acc_t21 + 0.01)) / (
            b_td_t21 * (b_td_acc_t21 + 0.01) + 0.1
        )
        df["striking_grappling_efficiency_correlation_diff"] = (
            df["r_sg_corr"] - df["b_sg_corr"]
        )
        df["r_def_allocation_balance"] = (r_str_def_t21 - r_td_def_t21).abs()
        df["b_def_allocation_balance"] = (b_str_def_t21 - b_td_def_t21).abs()
        df["defense_allocation_balance_diff"] = (
            df["r_def_allocation_balance"] - df["b_def_allocation_balance"]
        )
        _r_cbt = (
            (r_slpm_t21 / 10.0 + 0.01)
            * (r_acc_t21 + 0.01)
            * (10.0 / (r_sapm_t22 + 0.01))
            * (r_str_def_t21 + 0.01)
            * (r_td_t21 / 5.0 + 0.01)
            * (r_td_acc_t21 + 0.01)
            * (r_td_def_t21 + 0.01)
            * (r_sub_avg_t21 / 2.0 + 0.01)
        ) ** (1.0 / 8.0)
        _b_cbt = (
            (b_slpm_t21 / 10.0 + 0.01)
            * (b_acc_t21 + 0.01)
            * (10.0 / (b_sapm_t22 + 0.01))
            * (b_str_def_t21 + 0.01)
            * (b_td_t21 / 5.0 + 0.01)
            * (b_td_acc_t21 + 0.01)
            * (b_td_def_t21 + 0.01)
            * (b_sub_avg_t21 / 2.0 + 0.01)
        ) ** (1.0 / 8.0)
        df["r_combat_eff"] = _r_cbt
        df["b_combat_eff"] = _b_cbt
        df["total_combat_efficiency_index_diff"] = _r_cbt - _b_cbt

        self._log("Tier 25: Named composite features...")
        df["net_striking_advantage"] = (r_slpm_t21 - b_slpm_t21) - (
            r_sapm_t22 - b_sapm_t22
        )
        df["striker_advantage"] = (
            r_slpm_t21 * r_acc_t21) - (b_slpm_t21 * b_acc_t21)
        df["grappler_advantage"] = (
            r_td_t21 * r_td_acc_t21) - (b_td_t21 * b_td_acc_t21)
        df["experience_gap"] = r_tf_t23 - b_tf_t23
        r_ko_wins_t25 = df.get("r_pre_ko_wins", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        b_ko_wins_t25 = df.get("b_pre_ko_wins", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        r_sub_wins_t25 = df.get(
            "r_pre_sub_wins", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        b_sub_wins_t25 = df.get(
            "b_pre_sub_wins", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        r_ko_rate_t25 = r_ko_wins_t25 / (r_tf_t23 + 1.0)
        b_ko_rate_t25 = b_ko_wins_t25 / (b_tf_t23 + 1.0)
        r_sub_rate_t25 = r_sub_wins_t25 / (r_tf_t23 + 1.0)
        b_sub_rate_t25 = b_sub_wins_t25 / (b_tf_t23 + 1.0)
        df["ko_specialist_gap"] = r_ko_rate_t25 - b_ko_rate_t25
        df["submission_specialist_gap"] = r_sub_rate_t25 - b_sub_rate_t25
        r_elo_t25 = df.get("elo_r", pd.Series(
            1500.0, index=df.index)).fillna(1500.0)
        b_elo_t25 = df.get("elo_b", pd.Series(
            1500.0, index=df.index)).fillna(1500.0)
        r_traj_t25 = df.get("r_trajectory_3", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        b_traj_t25 = df.get("b_trajectory_3", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        df["skill_momentum"] = (r_elo_t25 - b_elo_t25) * \
            (r_traj_t25 - b_traj_t25)
        r_loss_streak_t25 = df.get(
            "r_pre_loss_streak", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        b_loss_streak_t25 = df.get(
            "b_pre_loss_streak", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        r_win_rate_t25 = r_wins_t23 / (r_tf_t23 + 1.0)
        b_win_rate_t25 = b_wins_t23 / (b_tf_t23 + 1.0)
        df["r_desperation"] = r_loss_streak_t25 * \
            (1.0 / (r_win_rate_t25 + 0.1))
        df["b_desperation"] = b_loss_streak_t25 * \
            (1.0 / (b_win_rate_t25 + 0.1))
        df["desperation_diff"] = df["r_desperation"] - df["b_desperation"]
        r_days_t25 = df.get(
            "r_days_since_last", pd.Series(180.0, index=df.index)
        ).fillna(180.0)
        b_days_t25 = df.get(
            "b_days_since_last", pd.Series(180.0, index=df.index)
        ).fillna(180.0)
        df["r_freshness"] = np.exp(-((r_days_t25 - 135.0)
                                   ** 2) / (2.0 * 90.0**2))
        df["b_freshness"] = np.exp(-((b_days_t25 - 135.0)
                                   ** 2) / (2.0 * 90.0**2))
        df["freshness_advantage"] = df["r_freshness"] - df["b_freshness"]

        self._log("Tier 26: Stance directional features...")
        _r_st26 = (
            df.get("r_stance", pd.Series("", index=df.index))
            .fillna("")
            .astype(str)
            .str.strip()
            .str.lower()
        )
        _b_st26 = (
            df.get("b_stance", pd.Series("", index=df.index))
            .fillna("")
            .astype(str)
            .str.strip()
            .str.lower()
        )
        df["orthodox_vs_southpaw_advantage"] = np.where(
            (_r_st26 == "orthodox") & (_b_st26 == "southpaw"),
            1.0,
            np.where((_r_st26 == "southpaw") & (
                _b_st26 == "orthodox"), -1.0, 0.0),
        ).astype(float)
        df["orthodox_vs_switch_advantage"] = np.where(
            (_r_st26 == "orthodox") & (_b_st26 == "switch"),
            1.0,
            np.where((_r_st26 == "switch") & (
                _b_st26 == "orthodox"), -1.0, 0.0),
        ).astype(float)
        df["southpaw_vs_switch_advantage"] = np.where(
            (_r_st26 == "southpaw") & (_b_st26 == "switch"),
            1.0,
            np.where((_r_st26 == "switch") & (
                _b_st26 == "southpaw"), -1.0, 0.0),
        ).astype(float)
        df["mirror_matchup"] = (_r_st26 == _b_st26).astype(float)

        self._log("Tier 27: Extended polynomial squared terms...")

        def _signed_sq_t27(s):
            return np.sign(s) * (s**2)

        for _feat_sq in [
            "net_striking_advantage",
            "striker_advantage",
            "grappler_advantage",
            "experience_gap",
            "ko_specialist_gap",
            "submission_specialist_gap",
            "skill_momentum",
            "desperation_diff",
            "freshness_advantage",
            "combined_defensive_hole_diff",
            "striking_volume_accuracy_synergy_diff",
            "total_offensive_efficiency_index_diff",
            "finish_efficiency_diff",
            "defense_versatility_diff",
            "offensive_versatility_diff",
        ]:
            if _feat_sq in df.columns:
                df[f"{_feat_sq}_sq"] = _signed_sq_t27(df[_feat_sq])

        self._log("Tier 28: Volatility and career arc features...")
        r_fr_l5_t28 = df.get(
            "r_pre_finish_rate_l5", pd.Series(0.4, index=df.index)
        ).fillna(0.4)
        b_fr_l5_t28 = df.get(
            "b_pre_finish_rate_l5", pd.Series(0.4, index=df.index)
        ).fillna(0.4)
        r_fr_l10_t28 = df.get(
            "r_pre_finish_rate_l10", pd.Series(0.4, index=df.index)
        ).fillna(0.4)
        b_fr_l10_t28 = df.get(
            "b_pre_finish_rate_l10", pd.Series(0.4, index=df.index)
        ).fillna(0.4)
        df["r_finish_rate_accel"] = r_fr_l5_t28 - r_fr_l10_t28
        df["b_finish_rate_accel"] = b_fr_l5_t28 - b_fr_l10_t28
        df["finish_rate_acceleration_diff"] = (
            df["r_finish_rate_accel"] - df["b_finish_rate_accel"]
        )
        r_slpm_cv_t28 = df.get("r_pre_slpm_cv", pd.Series(0.3, index=df.index)).fillna(
            0.3
        )
        b_slpm_cv_t28 = df.get("b_pre_slpm_cv", pd.Series(0.3, index=df.index)).fillna(
            0.3
        )
        df["slpm_coefficient_of_variation_diff"] = r_slpm_cv_t28 - b_slpm_cv_t28
        r_mil_t28 = df.get(
            "r_pre_mileage_adj_age", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        b_mil_t28 = df.get(
            "b_pre_mileage_adj_age", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        df["mileage_adjusted_age_diff"] = r_mil_t28 - b_mil_t28
        df["performance_decline_velocity_diff"] = (
            df.get("r_trajectory_3", pd.Series(
                0.0, index=df.index)).fillna(0.0)
            - df.get("b_trajectory_3", pd.Series(0.0,
                     index=df.index)).fillna(0.0)
        ) * (-1.0)
        r_cur_elo_t28 = df.get("elo_r", pd.Series(1500.0, index=df.index)).fillna(
            1500.0
        )
        b_cur_elo_t28 = df.get("elo_b", pd.Series(1500.0, index=df.index)).fillna(
            1500.0
        )
        r_peak_t28 = df.get(
            "r_career_elo_peak", pd.Series(1500.0, index=df.index)
        ).fillna(1500.0)
        b_peak_t28 = df.get(
            "b_career_elo_peak", pd.Series(1500.0, index=df.index)
        ).fillna(1500.0)
        df["r_distance_from_peak"] = r_peak_t28 - r_cur_elo_t28
        df["b_distance_from_peak"] = b_peak_t28 - b_cur_elo_t28
        df["distance_from_career_peak_diff"] = (
            df["r_distance_from_peak"] - df["b_distance_from_peak"]
        )
        r_fsp_t28 = df.get(
            "r_fights_since_peak", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        b_fsp_t28 = df.get(
            "b_fights_since_peak", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        df["r_career_inflection"] = r_fsp_t28 / (r_tf_t23 + 1.0)
        df["b_career_inflection"] = b_fsp_t28 / (b_tf_t23 + 1.0)
        df["career_inflection_point_diff"] = (
            df["r_career_inflection"] - df["b_career_inflection"]
        )
        df["r_prime_exit_risk"] = (r_age_t23 > 33).astype(float) * np.clip(
            -r_traj_t25, 0.0, 1.0
        )
        df["b_prime_exit_risk"] = (b_age_t23 > 33).astype(float) * np.clip(
            -b_traj_t25, 0.0, 1.0
        )
        df["prime_exit_risk_diff"] = df["r_prime_exit_risk"] - \
            df["b_prime_exit_risk"]
        df["r_aging_power_penalty"] = (
            r_ko_rate_t25 * r_age_t23 * (r_age_t23 > 35).astype(float)
        )
        df["b_aging_power_penalty"] = (
            b_ko_rate_t25 * b_age_t23 * (b_age_t23 > 35).astype(float)
        )
        df["aging_power_striker_penalty_diff"] = (
            df["r_aging_power_penalty"] - df["b_aging_power_penalty"]
        )
        df["r_bayesian_finish"] = (r_ko_wins_t25 + r_sub_wins_t25 + 2.0) / (
            r_tf_t23 + 4.0
        )
        df["b_bayesian_finish"] = (b_ko_wins_t25 + b_sub_wins_t25 + 2.0) / (
            b_tf_t23 + 4.0
        )
        df["bayesian_finish_rate_diff"] = (
            df["r_bayesian_finish"] - df["b_bayesian_finish"]
        )
        df["r_layoff_veteran"] = r_days_t25 * r_tf_t23
        df["b_layoff_veteran"] = b_days_t25 * b_tf_t23
        df["layoff_veteran_interaction_diff"] = (
            df["r_layoff_veteran"] - df["b_layoff_veteran"]
        )
        df["r_elo_momentum"] = r_cur_elo_t28 * r_traj_t25
        df["b_elo_momentum"] = b_cur_elo_t28 * b_traj_t25
        df["elo_momentum_vs_competition_diff"] = (
            df["r_elo_momentum"] - df["b_elo_momentum"]
        )
        r_avg_opp_elo_t28 = df.get(
            "r_avg_opp_elo_L5", pd.Series(1500.0, index=df.index)
        ).fillna(1500.0)
        b_avg_opp_elo_t28 = df.get(
            "b_avg_opp_elo_L5", pd.Series(1500.0, index=df.index)
        ).fillna(1500.0)
        df["r_title_proximity"] = (
            r_streak_t23 * r_avg_opp_elo_t28 * r_cur_elo_t28 / 1.0e6
        )
        df["b_title_proximity"] = (
            b_streak_t23 * b_avg_opp_elo_t28 * b_cur_elo_t28 / 1.0e6
        )
        df["title_shot_proximity_score_diff"] = (
            df["r_title_proximity"] - df["b_title_proximity"]
        )
        df["r_elo_volatility"] = r_cur_elo_t28 * r_slpm_cv_t28
        df["b_elo_volatility"] = b_cur_elo_t28 * b_slpm_cv_t28
        df["elo_volatility_interaction_diff"] = (
            df["r_elo_volatility"] - df["b_elo_volatility"]
        )
        r_fin_l10_t28 = df.get(
            "r_rolling10_finishes", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        b_fin_l10_t28 = df.get(
            "b_rolling10_finishes", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        df["elite_performance_frequency_l10_diff"] = (r_fin_l10_t28 / 10.0) - (
            b_fin_l10_t28 / 10.0
        )
        _r_dr = df.get("r_damage_ratio", pd.Series(
            1.0, index=df.index)).fillna(1.0)
        _b_dr = df.get("b_damage_ratio", pd.Series(
            1.0, index=df.index)).fillna(1.0)
        df["r_conf_damage_ratio"] = _r_dr * (1.0 - 1.0 / (r_tf_t23**0.5 + 1.0))
        df["b_conf_damage_ratio"] = _b_dr * (1.0 - 1.0 / (b_tf_t23**0.5 + 1.0))
        df["confidence_weighted_damage_ratio_diff"] = (
            df["r_conf_damage_ratio"] - df["b_conf_damage_ratio"]
        )

        _r_r5slpm = df.get(
            "r_pre_rolling5_slpm", pd.Series(3.0, index=df.index)
        ).fillna(3.0)
        _b_r5slpm = df.get(
            "b_pre_rolling5_slpm", pd.Series(3.0, index=df.index)
        ).fillna(3.0)
        df["r_recent_vs_career_striking"] = _r_r5slpm / (r_slpm_t21 + 0.1)
        df["b_recent_vs_career_striking"] = _b_r5slpm / (b_slpm_t21 + 0.1)
        df["recent_vs_career_striking_diff"] = (
            df["r_recent_vs_career_striking"] -
            df["b_recent_vs_career_striking"]
        )

        _r_slpmstd = df.get(
            "r_pre_slpm_std_l10", pd.Series(1.0, index=df.index)
        ).fillna(1.0)
        _b_slpmstd = df.get(
            "b_pre_slpm_std_l10", pd.Series(1.0, index=df.index)
        ).fillna(1.0)
        df["r_striking_consistency_ratio"] = 1.0 / (_r_slpmstd + 0.1)
        df["b_striking_consistency_ratio"] = 1.0 / (_b_slpmstd + 0.1)
        df["striking_consistency_ratio_diff"] = (
            df["r_striking_consistency_ratio"] -
            df["b_striking_consistency_ratio"]
        )

        _r_drstd = df.get(
            "r_pre_damage_ratio_std_l10", pd.Series(0.3, index=df.index)
        ).fillna(0.3)
        _b_drstd = df.get(
            "b_pre_damage_ratio_std_l10", pd.Series(0.3, index=df.index)
        ).fillna(0.3)
        df["performance_volatility_l10_diff"] = _r_drstd - _b_drstd

        _r_tact = df.get(
            "r_pre_tactical_evolution", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        _b_tact = df.get(
            "b_pre_tactical_evolution", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        df["tactical_evolution_score_diff"] = _r_tact - _b_tact

        self._log(
            "Tier 29: Daniel custom feature and ELO/Glicko compound features...")
        # Daniel Custom Feature: offensive striking efficiency * grappling threat composite
        r_daniel_t29 = (r_slpm_t21 * r_acc_t21) / (
            r_sapm_t22 * r_str_def_t21 + 0.01
        ) + (r_td_t21 * r_td_acc_t21 * r_td_def_t21 * r_sub_avg_t21)
        b_daniel_t29 = (b_slpm_t21 * b_acc_t21) / (
            b_sapm_t22 * b_str_def_t21 + 0.01
        ) + (b_td_t21 * b_td_acc_t21 * b_td_def_t21 * b_sub_avg_t21)
        df["r_daniel_custom"] = r_daniel_t29
        df["b_daniel_custom"] = b_daniel_t29
        df["daniel_custom_feature_diff"] = r_daniel_t29 - b_daniel_t29

        # ELO discounted by Glicko RD — penalises ELO signal for fighters with few fights (high RD = uncertain)
        r_glicko_rd_t29 = df.get(
            "r_glicko_rd", pd.Series(200.0, index=df.index)
        ).fillna(200.0)
        b_glicko_rd_t29 = df.get(
            "b_glicko_rd", pd.Series(200.0, index=df.index)
        ).fillna(200.0)
        df["r_elo_confidence_weighted"] = r_cur_elo_t28 * np.clip(
            1.0 - r_glicko_rd_t29 / 500.0, 0.2, 1.0
        )
        df["b_elo_confidence_weighted"] = b_cur_elo_t28 * np.clip(
            1.0 - b_glicko_rd_t29 / 500.0, 0.2, 1.0
        )
        df["elo_confidence_weighted_diff"] = (
            df["r_elo_confidence_weighted"] - df["b_elo_confidence_weighted"]
        )

        # Win streak quality: winning streak weighted by recent opponent strength
        df["r_streak_quality"] = r_streak_t23 * r_avg_opp_elo_t28 / 1000.0
        df["b_streak_quality"] = b_streak_t23 * b_avg_opp_elo_t28 / 1000.0
        df["streak_quality_diff"] = df["r_streak_quality"] - df["b_streak_quality"]

        # Reach × striking defense: range + shell combines into defensive reach advantage
        df["r_reach_defense"] = r_reach_t23 * r_str_def_t21
        df["b_reach_defense"] = b_reach_t23 * b_str_def_t21
        df["reach_defense_diff"] = df["r_reach_defense"] - df["b_reach_defense"]

        # Net striking efficiency: effective output minus absorbed striking cost (differs from damage_ratio by using subtraction not division)
        df["r_net_striking_efficiency"] = (r_slpm_t21 * r_acc_t21) - (
            r_sapm_t22 * (1.0 - r_str_def_t21)
        )
        df["b_net_striking_efficiency"] = (b_slpm_t21 * b_acc_t21) - (
            b_sapm_t22 * (1.0 - b_str_def_t21)
        )
        df["net_striking_efficiency_diff"] = (
            df["r_net_striking_efficiency"] - df["b_net_striking_efficiency"]
        )

        self._log(
            "Tier 30: Complex hybrid features combining top-importance signals..."
        )
        r_glicko_r_t30 = df.get("r_glicko_r", pd.Series(1500.0, index=df.index)).fillna(
            1500.0
        )
        b_glicko_r_t30 = df.get("b_glicko_r", pd.Series(1500.0, index=df.index)).fillna(
            1500.0
        )
        r_swvc_t30 = df.get(
            "r_style_win_vs_cluster", pd.Series(0.5, index=df.index)
        ).fillna(0.5)
        b_swvc_t30 = df.get(
            "b_style_win_vs_cluster", pd.Series(0.5, index=df.index)
        ).fillna(0.5)
        _common_opp_t30 = df.get(
            "common_opp_edge", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        _style_edge_t30 = df.get(
            "style_matchup_edge", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        r_ape_t30 = pd.to_numeric(
            df.get("r_ape_index", pd.Series(0.0, index=df.index)), errors="coerce"
        ).fillna(0.0)
        b_ape_t30 = pd.to_numeric(
            df.get("b_ape_index", pd.Series(0.0, index=df.index)), errors="coerce"
        ).fillna(0.0)
        r_peak_t30 = df.get("r_peak_score", pd.Series(
            0.5, index=df.index)).fillna(0.5)
        b_peak_t30 = df.get("b_peak_score", pd.Series(
            0.5, index=df.index)).fillna(0.5)
        _fresh_r_t30 = df.get("r_freshness", pd.Series(
            0.5, index=df.index)).fillna(0.5)
        _fresh_b_t30 = df.get("b_freshness", pd.Series(
            0.5, index=df.index)).fillna(0.5)
        _press_r_t30 = df.get(
            "r_pressure_index", pd.Series(5.0, index=df.index)
        ).fillna(5.0)
        _press_b_t30 = df.get(
            "b_pressure_index", pd.Series(5.0, index=df.index)
        ).fillna(5.0)

        # Glicko certainty and age-prime factors used across multiple hybrids
        r_certainty_t30 = np.clip(1.0 - r_glicko_rd_t29 / 400.0, 0.25, 1.0)
        b_certainty_t30 = np.clip(1.0 - b_glicko_rd_t29 / 400.0, 0.25, 1.0)
        r_prime_t30 = np.clip(1.0 - np.abs(r_age_t23 - 29.0) / 12.0, 0.1, 1.0)
        b_prime_t30 = np.clip(1.0 - np.abs(b_age_t23 - 29.0) / 12.0, 0.1, 1.0)

        # Style × Glicko × Prime: style win rate against opponent cluster × established high rating × age prime
        # Highest when: fighter beats their style matchup AND has a well-established, high Glicko AND is 25-33
        df["r_style_glicko_prime"] = (
            r_swvc_t30 * (r_glicko_r_t30 / 1500.0) *
            r_prime_t30 * r_certainty_t30
        )
        df["b_style_glicko_prime"] = (
            b_swvc_t30 * (b_glicko_r_t30 / 1500.0) *
            b_prime_t30 * b_certainty_t30
        )
        df["style_glicko_prime_diff"] = (
            df["r_style_glicko_prime"] - df["b_style_glicko_prime"]
        )

        # Common opponent × style consensus: product is positive when both independent signals agree on the same fighter
        # Captures corroborating evidence — when head-to-head logic and style matchup point the same direction
        df["common_style_consensus"] = _common_opp_t30 * _style_edge_t30

        # Age × ELO × Trajectory climb: young fighters actively rising vs aging fighters declining
        # Youth factor decays linearly from 1.0 at age 20 to 0.0 at age 35
        r_youth_t30 = np.clip((35.0 - r_age_t23) / 15.0, 0.0, 1.0)
        b_youth_t30 = np.clip((35.0 - b_age_t23) / 15.0, 0.0, 1.0)
        df["r_age_elo_climb"] = (
            (r_cur_elo_t28 / 1500.0) *
            np.clip(r_traj_t25, -1.0, 1.0) * r_youth_t30
        )
        df["b_age_elo_climb"] = (
            (b_cur_elo_t28 / 1500.0) *
            np.clip(b_traj_t25, -1.0, 1.0) * b_youth_t30
        )
        df["age_elo_climb_diff"] = df["r_age_elo_climb"] - df["b_age_elo_climb"]

        # Quality battle-tested: win rate × recent opponent quality × log(career depth)
        # Separates elite records from padded ones: a 20-0 vs tomato cans vs a 15-3 vs top contenders
        r_expq_t30 = r_wins_t23 / (r_wins_t23 + r_losses_t23 + 1.0)
        b_expq_t30 = b_wins_t23 / (b_wins_t23 + b_losses_t23 + 1.0)
        df["r_quality_battle_tested"] = (
            r_expq_t30 * (r_avg_opp_elo_t28 / 1500.0) * np.log1p(r_tf_t23)
        )
        df["b_quality_battle_tested"] = (
            b_expq_t30 * (b_avg_opp_elo_t28 / 1500.0) * np.log1p(b_tf_t23)
        )
        df["quality_battle_tested_diff"] = (
            df["r_quality_battle_tested"] - df["b_quality_battle_tested"]
        )

        # Physical × Style dominance: ape index (relative reach advantage) × reach × style win rate
        # A long-armed striker who also wins their style matchup has a multiplicative physical-tactical edge
        df["r_physical_style_dominance"] = r_ape_t30 * \
            r_swvc_t30 * (r_reach_t23 / 71.0)
        df["b_physical_style_dominance"] = b_ape_t30 * \
            b_swvc_t30 * (b_reach_t23 / 71.0)
        df["physical_style_dominance_diff"] = (
            df["r_physical_style_dominance"] - df["b_physical_style_dominance"]
        )

        # Peak × Freshness × Pressure: at career peak, physically fresh (ideal ring time), applying high pressure
        df["r_peak_fresh_pressure"] = r_peak_t30 * \
            _fresh_r_t30 * (_press_r_t30 / 10.0)
        df["b_peak_fresh_pressure"] = b_peak_t30 * \
            _fresh_b_t30 * (_press_b_t30 / 10.0)
        df["peak_fresh_pressure_diff"] = (
            df["r_peak_fresh_pressure"] - df["b_peak_fresh_pressure"]
        )

        # Tri-rating consensus: ELO diff × Glicko diff × style edge — positive when all three systems agree
        # Large positive = everyone agrees red is better; near zero = systems conflict (genuine toss-up)
        _elo_sig_t30 = (r_cur_elo_t28 - b_cur_elo_t28) / 200.0
        _glicko_sig_t30 = (r_glicko_r_t30 - b_glicko_r_t30) / 200.0
        df["tri_rating_consensus"] = _elo_sig_t30 * \
            _glicko_sig_t30 * _style_edge_t30

        self._log("Tier 31: Z-score × SVD × physical hybrid features...")
        z_r_kd_t31 = df.get("z_r_pre_kd_rate", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        z_b_kd_t31 = df.get("z_b_pre_kd_rate", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        z_r_acc_t31 = df.get(
            "z_r_pre_sig_str_acc", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        z_b_acc_t31 = df.get(
            "z_b_pre_sig_str_acc", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        z_r_slpm_t31 = df.get("z_r_pre_SLpM", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        z_b_slpm_t31 = df.get("z_b_pre_SLpM", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        z_r_sub_t31 = df.get(
            "z_r_pre_sub_att_rate", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        z_b_sub_t31 = df.get(
            "z_b_pre_sub_att_rate", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        form_svd3_t31 = df.get("form_svd_3", pd.Series(
            0.0, index=df.index)).fillna(0.0)
        str_svd3_t31 = df.get("striking_svd_3", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        grp_svd2_t31 = df.get("grappling_svd_2", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        r_cbt_t31 = df.get("r_combat_eff", pd.Series(
            0.5, index=df.index)).fillna(0.5)
        b_cbt_t31 = df.get("b_combat_eff", pd.Series(
            0.5, index=df.index)).fillna(0.5)
        r_glicko_vol_t31 = df.get(
            "r_glicko_vol", pd.Series(0.06, index=df.index)
        ).fillna(0.06)
        b_glicko_vol_t31 = df.get(
            "b_glicko_vol", pd.Series(0.06, index=df.index)
        ).fillna(0.06)

        # WC-elite KD power × style win rate × ape index: divisionally superior finisher + tactical + physical
        df["r_wc_ko_style_physical"] = z_r_kd_t31 * \
            r_swvc_t30 * r_ape_t30 / 5.0
        df["b_wc_ko_style_physical"] = z_b_kd_t31 * \
            b_swvc_t30 * b_ape_t30 / 5.0
        df["wc_ko_style_physical_diff"] = (
            df["r_wc_ko_style_physical"] - df["b_wc_ko_style_physical"]
        )

        # WC-elite accuracy × ape index × style: extends physical_style_dominance with z-score precision
        df["r_wc_acc_physical_style"] = z_r_acc_t31 * \
            r_ape_t30 / 5.0 * r_swvc_t30
        df["b_wc_acc_physical_style"] = z_b_acc_t31 * \
            b_ape_t30 / 5.0 * b_swvc_t30
        df["wc_acc_physical_style_diff"] = (
            df["r_wc_acc_physical_style"] - df["b_wc_acc_physical_style"]
        )

        # Double Z-score striking dominance: elite at BOTH volume AND accuracy within the weight class
        df["r_wc_striking_dominance"] = z_r_slpm_t31 * z_r_acc_t31
        df["b_wc_striking_dominance"] = z_b_slpm_t31 * z_b_acc_t31
        df["wc_striking_dominance_diff"] = (
            df["r_wc_striking_dominance"] - df["b_wc_striking_dominance"]
        )

        # WC-elite submission threat × TD defense × style: grappling-threat compound with defensive context
        df["r_wc_sub_threat_compound"] = z_r_sub_t31 * r_td_def_t21 * r_swvc_t30
        df["b_wc_sub_threat_compound"] = z_b_sub_t31 * b_td_def_t21 * b_swvc_t30
        df["wc_sub_threat_compound_diff"] = (
            df["r_wc_sub_threat_compound"] - df["b_wc_sub_threat_compound"]
        )

        # Form SVD × style edge: latent recent-form factor AND style matchup agree on same fighter
        # Product is symmetric (fight-level) — high = both corroborate the same outcome direction
        df["form_svd_style_synergy"] = form_svd3_t31 * _style_edge_t30

        # Striking SVD amplified by physical gap: latent striking advantage × relative ape index gap
        df["striking_svd_physical"] = str_svd3_t31 * \
            (r_ape_t30 - b_ape_t30) / 5.0

        # Grappling SVD × style edge: grappling latent quality AND style matchup point same direction
        df["grappling_svd_style_synergy"] = grp_svd2_t31 * _style_edge_t30

        # Combat efficiency × Glicko stability × prime: the reliable skilled in-prime fighter
        # Low Glicko vol = predictable performance; high = erratic (penalised)
        r_vol_pen_t31 = np.clip(1.0 - r_glicko_vol_t31 * 5.0, 0.1, 1.0)
        b_vol_pen_t31 = np.clip(1.0 - b_glicko_vol_t31 * 5.0, 0.1, 1.0)
        df["r_stable_prime_combat"] = r_cbt_t31 * r_vol_pen_t31 * r_prime_t30
        df["b_stable_prime_combat"] = b_cbt_t31 * b_vol_pen_t31 * b_prime_t30
        df["stable_prime_combat_diff"] = (
            df["r_stable_prime_combat"] - df["b_stable_prime_combat"]
        )

        # Physical × ELO × Youth: young fighter with elite ELO AND physical gifts — the rising physical phenom
        df["r_physical_elo_youth"] = (
            r_ape_t30 * (r_reach_t23 / 71.0) *
            (r_cur_elo_t28 / 1500.0) * r_youth_t30
        )
        df["b_physical_elo_youth"] = (
            b_ape_t30 * (b_reach_t23 / 71.0) *
            (b_cur_elo_t28 / 1500.0) * b_youth_t30
        )
        df["physical_elo_youth_diff"] = (
            df["r_physical_elo_youth"] - df["b_physical_elo_youth"]
        )

        self._log("Tier 32: Extended physical compound features...")
        z_r_td_t32 = df.get("z_r_pre_td_avg", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        z_b_td_t32 = df.get("z_b_pre_td_avg", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )

        # Physical × Glicko × Youth: Glicko-rated version of physical_elo_youth (more sophisticated rating)
        df["r_physical_glicko_youth"] = (
            r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
            * r_youth_t30
        )
        df["b_physical_glicko_youth"] = (
            b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
            * b_youth_t30
        )
        df["physical_glicko_youth_diff"] = (
            df["r_physical_glicko_youth"] - df["b_physical_glicko_youth"]
        )

        # Physical × Style × Glicko × Prime: 5-way compound — physical + tactical + rating + certainty + prime age
        df["r_physical_style_glicko_prime"] = (
            r_ape_t30
            * r_swvc_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
            * r_prime_t30
        )
        df["b_physical_style_glicko_prime"] = (
            b_ape_t30
            * b_swvc_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
            * b_prime_t30
        )
        df["physical_style_glicko_prime_diff"] = (
            df["r_physical_style_glicko_prime"] -
            df["b_physical_style_glicko_prime"]
        )

        # Physical × Combat Efficiency × Prime: physically dominant complete fighter at career peak
        df["r_physical_combat_prime"] = (
            r_ape_t30 * (r_reach_t23 / 71.0) * r_cbt_t31 * r_prime_t30
        )
        df["b_physical_combat_prime"] = (
            b_ape_t30 * (b_reach_t23 / 71.0) * b_cbt_t31 * b_prime_t30
        )
        df["physical_combat_prime_diff"] = (
            df["r_physical_combat_prime"] - df["b_physical_combat_prime"]
        )

        # Freshness × Physical × ELO: optimally-rested fighter with physical gifts AND high ELO
        df["r_fresh_physical_elo"] = (
            _fresh_r_t30 * r_ape_t30 *
            (r_reach_t23 / 71.0) * (r_cur_elo_t28 / 1500.0)
        )
        df["b_fresh_physical_elo"] = (
            _fresh_b_t30 * b_ape_t30 *
            (b_reach_t23 / 71.0) * (b_cur_elo_t28 / 1500.0)
        )
        df["fresh_physical_elo_diff"] = (
            df["r_fresh_physical_elo"] - df["b_fresh_physical_elo"]
        )

        # Grappling SVD × Physical: grappling latent advantage amplified by ape index gap (grappling_svd_physical)
        df["grappling_svd_physical"] = grp_svd2_t31 * \
            (r_ape_t30 - b_ape_t30) / 5.0

        # WC-elite TD volume × ape index × style: grappling-threat version of wc_ko_style_physical
        df["r_wc_td_physical_style"] = z_r_td_t32 * \
            r_ape_t30 / 5.0 * r_swvc_t30
        df["b_wc_td_physical_style"] = z_b_td_t32 * \
            b_ape_t30 / 5.0 * b_swvc_t30
        df["wc_td_physical_style_diff"] = (
            df["r_wc_td_physical_style"] - df["b_wc_td_physical_style"]
        )

        self._log("Tier 33: Freshness/layoff/absorption compound features...")
        z_r_sapm_t33 = df.get("z_r_pre_SApM", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        z_b_sapm_t33 = df.get("z_b_pre_SApM", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        form_svd1_t33 = df.get("form_svd_1", pd.Series(
            0.0, index=df.index)).fillna(0.0)
        form_svd0_t33 = df.get("form_svd_0", pd.Series(
            0.0, index=df.index)).fillna(0.0)
        str_svd0_t33 = df.get("striking_svd_0", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        grp_svd3_t33 = df.get("grappling_svd_3", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )

        # Physical × low absorption: long reach + ape advantage + elite WC ability to avoid damage
        # Negative z_SApM = absorbs LESS than weight class average (inverted: better = lower)
        df["r_physical_defensive"] = (
            r_ape_t30 * (r_reach_t23 / 71.0) * np.clip(-z_r_sapm_t33, 0.0, 3.0)
        )
        df["b_physical_defensive"] = (
            b_ape_t30 * (b_reach_t23 / 71.0) * np.clip(-z_b_sapm_t33, 0.0, 3.0)
        )
        df["physical_defensive_diff"] = (
            df["r_physical_defensive"] - df["b_physical_defensive"]
        )

        # Freshness × Style × Glicko: rested fighter who also has tactical AND rating edge
        df["r_fresh_style_glicko"] = (
            _fresh_r_t30 * r_swvc_t30 *
            (r_glicko_r_t30 / 1500.0) * r_certainty_t30
        )
        df["b_fresh_style_glicko"] = (
            _fresh_b_t30 * b_swvc_t30 *
            (b_glicko_r_t30 / 1500.0) * b_certainty_t30
        )
        df["fresh_style_glicko_diff"] = (
            df["r_fresh_style_glicko"] - df["b_fresh_style_glicko"]
        )

        # Physical × Glicko × Prime (no style): pure physical + rating + age compound
        df["r_physical_glicko_prime"] = (
            r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
            * r_prime_t30
        )
        df["b_physical_glicko_prime"] = (
            b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
            * b_prime_t30
        )
        df["physical_glicko_prime_diff"] = (
            df["r_physical_glicko_prime"] - df["b_physical_glicko_prime"]
        )

        # Form SVD × physical: recent form latent factor amplified by physical ape gap
        df["form_svd_physical"] = form_svd3_t31 * (r_ape_t30 - b_ape_t30) / 5.0

        # Form SVD1 × style × glicko: second form latent component corroborated by style and rating
        df["form_svd1_style_glicko"] = (
            form_svd1_t33 * _style_edge_t30 *
            (r_glicko_r_t30 - b_glicko_r_t30) / 200.0
        )

        # Striking SVD0 × physical: first (largest) striking SVD component × physical ape gap
        df["striking_svd0_physical"] = str_svd0_t33 * \
            (r_ape_t30 - b_ape_t30) / 5.0

        # Form SVD0 × physical: largest form latent component × physical ape gap
        df["form_svd0_physical"] = form_svd0_t33 * \
            (r_ape_t30 - b_ape_t30) / 5.0

        # Grappling SVD3 × style edge: grappling latent factor 3 corroborated by style matchup
        df["grappling_svd3_style"] = grp_svd3_t33 * _style_edge_t30

        self._log(
            "Tier 34: Win rate, finish rate, and opponent quality hybrid features..."
        )
        r_tp_t34 = df.get("r_title_proximity", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        b_tp_t34 = df.get("b_title_proximity", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        r_ko_rate_t34 = _r_ko_d / (r_tf_t23 + 1.0)
        b_ko_rate_t34 = _b_ko_d / (b_tf_t23 + 1.0)
        r_sub_rate_t34 = _r_sub_d / (r_tf_t23 + 1.0)
        b_sub_rate_t34 = _b_sub_d / (b_tf_t23 + 1.0)

        # Win rate × Physical × Glicko certainty: proven winner with physical gifts and established elite rating
        df["r_winrate_physical_glicko"] = (
            r_expq_t30
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
        )
        df["b_winrate_physical_glicko"] = (
            b_expq_t30
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
        )
        df["winrate_physical_glicko_diff"] = (
            df["r_winrate_physical_glicko"] - df["b_winrate_physical_glicko"]
        )

        # Finish rate × Physical × Youth: young, physically gifted finisher — rising threat
        df["r_finish_physical_youth"] = (
            r_fr_t23 * r_ape_t30 * (r_reach_t23 / 71.0) * r_youth_t30
        )
        df["b_finish_physical_youth"] = (
            b_fr_t23 * b_ape_t30 * (b_reach_t23 / 71.0) * b_youth_t30
        )
        df["finish_physical_youth_diff"] = (
            df["r_finish_physical_youth"] - df["b_finish_physical_youth"]
        )

        # Opponent-quality-adjusted win rate × Physical: non-padded elite record combined with physical superiority
        df["r_elite_winrate_physical"] = (
            r_expq_t30 * (r_avg_opp_elo_t28 / 1500.0) *
            r_ape_t30 * (r_reach_t23 / 71.0)
        )
        df["b_elite_winrate_physical"] = (
            b_expq_t30 * (b_avg_opp_elo_t28 / 1500.0) *
            b_ape_t30 * (b_reach_t23 / 71.0)
        )
        df["elite_winrate_physical_diff"] = (
            df["r_elite_winrate_physical"] - df["b_elite_winrate_physical"]
        )

        # KO rate × Physical × Prime: physical knockout artist at career peak — the most dangerous archetype
        df["r_ko_rate_physical_prime"] = (
            r_ko_rate_t34 * r_ape_t30 * (r_reach_t23 / 71.0) * r_prime_t30
        )
        df["b_ko_rate_physical_prime"] = (
            b_ko_rate_t34 * b_ape_t30 * (b_reach_t23 / 71.0) * b_prime_t30
        )
        df["ko_rate_physical_prime_diff"] = (
            df["r_ko_rate_physical_prime"] - df["b_ko_rate_physical_prime"]
        )

        # Win rate × Style × Glicko × Prime: proven winner who also has style, rating, and age-prime edge
        df["r_winrate_style_glicko_prime"] = (
            r_expq_t30
            * r_swvc_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
            * r_prime_t30
        )
        df["b_winrate_style_glicko_prime"] = (
            b_expq_t30
            * b_swvc_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
            * b_prime_t30
        )
        df["winrate_style_glicko_prime_diff"] = (
            df["r_winrate_style_glicko_prime"] -
            df["b_winrate_style_glicko_prime"]
        )

        # Active streak × Physical × Glicko: fighter on a hot streak who is physically elite and highly rated
        df["r_streak_physical_glicko"] = (
            np.clip(r_streak_t23 / 5.0, 0.0, 1.0)
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
        )
        df["b_streak_physical_glicko"] = (
            np.clip(b_streak_t23 / 5.0, 0.0, 1.0)
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
        )
        df["streak_physical_glicko_diff"] = (
            df["r_streak_physical_glicko"] - df["b_streak_physical_glicko"]
        )

        # Title proximity × Physical × Glicko: near-belt contender with physical gifts and proven elite rating
        df["r_title_physical_glicko"] = r_tp_t34 * \
            r_ape_t30 * (r_glicko_r_t30 / 1500.0)
        df["b_title_physical_glicko"] = b_tp_t34 * \
            b_ape_t30 * (b_glicko_r_t30 / 1500.0)
        df["title_physical_glicko_diff"] = (
            df["r_title_physical_glicko"] - df["b_title_physical_glicko"]
        )

        # Recent form × Physical × Youth: hot young fighter with physical edge — peak upside scenario
        df["r_recent_winrate_physical_youth"] = (
            (_r_roll3 / 3.0) * r_ape_t30 * (r_reach_t23 / 71.0) * r_youth_t30
        )
        df["b_recent_winrate_physical_youth"] = (
            (_b_roll3 / 3.0) * b_ape_t30 * (b_reach_t23 / 71.0) * b_youth_t30
        )
        df["recent_winrate_physical_youth_diff"] = (
            df["r_recent_winrate_physical_youth"]
            - df["b_recent_winrate_physical_youth"]
        )

        # Submission rate × Physical × Youth: young physical sub threat — high ceiling grappler
        df["r_sub_rate_physical_youth"] = r_sub_rate_t34 * r_ape_t30 * r_youth_t30
        df["b_sub_rate_physical_youth"] = b_sub_rate_t34 * b_ape_t30 * b_youth_t30
        df["sub_rate_physical_youth_diff"] = (
            df["r_sub_rate_physical_youth"] - df["b_sub_rate_physical_youth"]
        )

        # 6-way ultimate compound: win rate × physical × Glicko × youth — every top dimension combined
        df["r_winrate_physical_glicko_youth"] = (
            r_expq_t30
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
            * r_youth_t30
        )
        df["b_winrate_physical_glicko_youth"] = (
            b_expq_t30
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
            * b_youth_t30
        )
        df["winrate_physical_glicko_youth_diff"] = (
            df["r_winrate_physical_glicko_youth"]
            - df["b_winrate_physical_glicko_youth"]
        )

        self._log(
            "Tier 35: Compound extensions of top-performing Tier 34 patterns...")
        r_mom_t35 = df.get("r_momentum_quality", pd.Series(0.2, index=df.index)).fillna(
            0.2
        )
        b_mom_t35 = df.get("b_momentum_quality", pd.Series(0.2, index=df.index)).fillna(
            0.2
        )

        # 7-way compound: win rate × ape × reach × style × Glicko × certainty × prime
        # Extends physical_style_glicko_prime (#1 prior tier) by adding the proven win-rate dimension
        df["r_winrate_physical_style_glicko_prime"] = (
            r_expq_t30
            * r_ape_t30
            * r_swvc_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
            * r_prime_t30
        )
        df["b_winrate_physical_style_glicko_prime"] = (
            b_expq_t30
            * b_ape_t30
            * b_swvc_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
            * b_prime_t30
        )
        df["winrate_physical_style_glicko_prime_diff"] = (
            df["r_winrate_physical_style_glicko_prime"]
            - df["b_winrate_physical_style_glicko_prime"]
        )

        # Recent form × ape × reach × Glicko × youth: rolling-3 version of the #1 feature winrate_physical_glicko_youth
        df["r_recent_physical_glicko_youth"] = (
            (_r_roll3 / 3.0)
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
            * r_youth_t30
        )
        df["b_recent_physical_glicko_youth"] = (
            (_b_roll3 / 3.0)
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
            * b_youth_t30
        )
        df["recent_physical_glicko_youth_diff"] = (
            df["r_recent_physical_glicko_youth"] -
            df["b_recent_physical_glicko_youth"]
        )

        # Opponent quality × Physical × Glicko: dominant against strong opposition AND physically superior AND highly rated
        df["r_opp_quality_physical_glicko"] = (
            (r_avg_opp_elo_t28 / 1500.0)
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
        )
        df["b_opp_quality_physical_glicko"] = (
            (b_avg_opp_elo_t28 / 1500.0)
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
        )
        df["opp_quality_physical_glicko_diff"] = (
            df["r_opp_quality_physical_glicko"] -
            df["b_opp_quality_physical_glicko"]
        )

        # Momentum quality × Physical × Glicko: currently building momentum (important LGB) + physical + elite rating
        df["r_momentum_physical_glicko"] = (
            r_mom_t35 * r_ape_t30 * (r_glicko_r_t30 / 1500.0) * r_certainty_t30
        )
        df["b_momentum_physical_glicko"] = (
            b_mom_t35 * b_ape_t30 * (b_glicko_r_t30 / 1500.0) * b_certainty_t30
        )
        df["momentum_physical_glicko_diff"] = (
            df["r_momentum_physical_glicko"] - df["b_momentum_physical_glicko"]
        )

        # Finish rate × Physical × Glicko × Prime: extends finish_physical_youth — seasoned prime finisher over rising youth
        df["r_finish_physical_glicko_prime"] = (
            r_fr_t23
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_prime_t30
        )
        df["b_finish_physical_glicko_prime"] = (
            b_fr_t23
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_prime_t30
        )
        df["finish_physical_glicko_prime_diff"] = (
            df["r_finish_physical_glicko_prime"] -
            df["b_finish_physical_glicko_prime"]
        )

        # Win rate × Style × Physical × Youth: tactical + physical + win rate + young — rising complete fighter
        df["r_winrate_style_physical_youth"] = (
            r_expq_t30 * r_swvc_t30 * r_ape_t30 * r_youth_t30
        )
        df["b_winrate_style_physical_youth"] = (
            b_expq_t30 * b_swvc_t30 * b_ape_t30 * b_youth_t30
        )
        df["winrate_style_physical_youth_diff"] = (
            df["r_winrate_style_physical_youth"] -
            df["b_winrate_style_physical_youth"]
        )

        # Streak × Win rate × Physical × Glicko: hot streak corroborated by career record + physical + rating
        df["r_streak_winrate_physical_glicko"] = (
            np.clip(r_streak_t23 / 5.0, 0.0, 1.0)
            * r_expq_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
        )
        df["b_streak_winrate_physical_glicko"] = (
            np.clip(b_streak_t23 / 5.0, 0.0, 1.0)
            * b_expq_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
        )
        df["streak_winrate_physical_glicko_diff"] = (
            df["r_streak_winrate_physical_glicko"]
            - df["b_streak_winrate_physical_glicko"]
        )

        # Opponent quality × Win rate × Physical × Youth: elite resume + win rate + physical + age upside
        df["r_opp_quality_winrate_physical_youth"] = (
            (r_avg_opp_elo_t28 / 1500.0) * r_expq_t30 * r_ape_t30 * r_youth_t30
        )
        df["b_opp_quality_winrate_physical_youth"] = (
            (b_avg_opp_elo_t28 / 1500.0) * b_expq_t30 * b_ape_t30 * b_youth_t30
        )
        df["opp_quality_winrate_physical_youth_diff"] = (
            df["r_opp_quality_winrate_physical_youth"]
            - df["b_opp_quality_winrate_physical_youth"]
        )

        # Title proximity × Physical × Glicko × Youth: contender who is young + physical + established elite rating
        df["r_title_physical_glicko_youth"] = (
            r_tp_t34 * r_ape_t30 * (r_glicko_r_t30 / 1500.0) * r_youth_t30
        )
        df["b_title_physical_glicko_youth"] = (
            b_tp_t34 * b_ape_t30 * (b_glicko_r_t30 / 1500.0) * b_youth_t30
        )
        df["title_physical_glicko_youth_diff"] = (
            df["r_title_physical_glicko_youth"] -
            df["b_title_physical_glicko_youth"]
        )

        # KO rate × Style × Glicko × Prime: devastating KO specialist who has style matchup AND is in their prime
        df["r_ko_rate_style_glicko_prime"] = (
            r_ko_rate_t34
            * r_swvc_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
            * r_prime_t30
        )
        df["b_ko_rate_style_glicko_prime"] = (
            b_ko_rate_t34
            * b_swvc_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
            * b_prime_t30
        )
        df["ko_rate_style_glicko_prime_diff"] = (
            df["r_ko_rate_style_glicko_prime"] -
            df["b_ko_rate_style_glicko_prime"]
        )

        self._log(
            "Tier 36: Fresh/defensive/peak dimension extensions of top compounds..."
        )
        r_career_peak_t36 = df.get(
            "r_career_elo_peak", pd.Series(1500.0, index=df.index)
        ).fillna(1500.0)
        b_career_peak_t36 = df.get(
            "b_career_elo_peak", pd.Series(1500.0, index=df.index)
        ).fillna(1500.0)

        # 7-way: freshness × win rate × physical × Glicko × youth — rested proven physical phenom
        # Adds freshness to the current #1 feature (winrate_physical_glicko_youth)
        df["r_fresh_winrate_physical_glicko_youth"] = (
            _fresh_r_t30
            * r_expq_t30
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
            * r_youth_t30
        )
        df["b_fresh_winrate_physical_glicko_youth"] = (
            _fresh_b_t30
            * b_expq_t30
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
            * b_youth_t30
        )
        df["fresh_winrate_physical_glicko_youth_diff"] = (
            df["r_fresh_winrate_physical_glicko_youth"]
            - df["b_fresh_winrate_physical_glicko_youth"]
        )

        # Strike defense × Physical × Glicko × Youth: elusive young physical fighter — hard to hit AND physically gifted
        df["r_strike_def_physical_glicko_youth"] = (
            r_str_def_t21
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_strike_def_physical_glicko_youth"] = (
            b_str_def_t21
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["strike_def_physical_glicko_youth_diff"] = (
            df["r_strike_def_physical_glicko_youth"]
            - df["b_strike_def_physical_glicko_youth"]
        )

        # Peak score × Physical × Glicko × Youth: fighter whose career peak aligns with being young and physically elite
        df["r_peak_physical_glicko_youth"] = (
            r_peak_t30
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_peak_physical_glicko_youth"] = (
            b_peak_t30
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["peak_physical_glicko_youth_diff"] = (
            df["r_peak_physical_glicko_youth"] -
            df["b_peak_physical_glicko_youth"]
        )

        # Win rate × Physical × Glicko × Prime × Opp Quality: opponent-validated 7-way elite compound
        df["r_elite_winrate_physical_glicko_prime"] = (
            r_expq_t30
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
            * r_prime_t30
            * (r_avg_opp_elo_t28 / 1500.0)
        )
        df["b_elite_winrate_physical_glicko_prime"] = (
            b_expq_t30
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
            * b_prime_t30
            * (b_avg_opp_elo_t28 / 1500.0)
        )
        df["elite_winrate_physical_glicko_prime_diff"] = (
            df["r_elite_winrate_physical_glicko_prime"]
            - df["b_elite_winrate_physical_glicko_prime"]
        )

        # Recent form × Opp quality × Physical × Glicko: rolling-3 hot form against strong competition × physical + rating
        df["r_recent_opp_quality_physical_glicko"] = (
            (_r_roll3 / 3.0)
            * (r_avg_opp_elo_t28 / 1500.0)
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
        )
        df["b_recent_opp_quality_physical_glicko"] = (
            (_b_roll3 / 3.0)
            * (b_avg_opp_elo_t28 / 1500.0)
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
        )
        df["recent_opp_quality_physical_glicko_diff"] = (
            df["r_recent_opp_quality_physical_glicko"]
            - df["b_recent_opp_quality_physical_glicko"]
        )

        # Daniel custom score × Physical × Glicko × Youth: combined offense + grappling threat composite × physical + rating + age
        df["r_daniel_physical_glicko_youth"] = (
            r_daniel_t29
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
            * r_youth_t30
        )
        df["b_daniel_physical_glicko_youth"] = (
            b_daniel_t29
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
            * b_youth_t30
        )
        df["daniel_physical_glicko_youth_diff"] = (
            df["r_daniel_physical_glicko_youth"] -
            df["b_daniel_physical_glicko_youth"]
        )

        # Title proximity × Win rate × Physical × Glicko × Youth: 5-way — belt contender + proven record + physical + rating + age
        df["r_title_winrate_physical_glicko_youth"] = (
            r_tp_t34 * r_expq_t30 * r_ape_t30 *
            (r_glicko_r_t30 / 1500.0) * r_youth_t30
        )
        df["b_title_winrate_physical_glicko_youth"] = (
            b_tp_t34 * b_expq_t30 * b_ape_t30 *
            (b_glicko_r_t30 / 1500.0) * b_youth_t30
        )
        df["title_winrate_physical_glicko_youth_diff"] = (
            df["r_title_winrate_physical_glicko_youth"]
            - df["b_title_winrate_physical_glicko_youth"]
        )

        # Complete defense × Physical × Glicko × Prime: both defensive walls up + physical gifts + elite rating + peak age
        r_tot_def_t36 = r_str_def_t21 * r_td_def_t21
        b_tot_def_t36 = b_str_def_t21 * b_td_def_t21
        df["r_complete_def_physical_glicko"] = (
            r_tot_def_t36
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_prime_t30
        )
        df["b_complete_def_physical_glicko"] = (
            b_tot_def_t36
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_prime_t30
        )
        df["complete_def_physical_glicko_diff"] = (
            df["r_complete_def_physical_glicko"] -
            df["b_complete_def_physical_glicko"]
        )

        # Career ELO peak × Physical × Glicko certainty: historical ceiling × current physical gifts × reliable elite rating
        df["r_career_peak_physical_glicko"] = (
            (r_career_peak_t36 / 1500.0)
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
        )
        df["b_career_peak_physical_glicko"] = (
            (b_career_peak_t36 / 1500.0)
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
        )
        df["career_peak_physical_glicko_diff"] = (
            df["r_career_peak_physical_glicko"] -
            df["b_career_peak_physical_glicko"]
        )

        # Finish rate × Strike defense × Physical × Glicko: the complete physical fighter — finishes and doesn't get finished
        df["r_finish_def_physical_glicko"] = (
            r_fr_t23
            * r_str_def_t21
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
        )
        df["b_finish_def_physical_glicko"] = (
            b_fr_t23
            * b_str_def_t21
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
        )
        df["finish_def_physical_glicko_diff"] = (
            df["r_finish_def_physical_glicko"] -
            df["b_finish_def_physical_glicko"]
        )

        self._log("Tier 37: Strike-defense and peak-score dimension extensions...")

        # Both defensive walls × Physical × Glicko × Youth: complete defensive physical phenom
        df["r_complete_def_physical_glicko_youth"] = (
            r_str_def_t21
            * r_td_def_t21
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_complete_def_physical_glicko_youth"] = (
            b_str_def_t21
            * b_td_def_t21
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["complete_def_physical_glicko_youth_diff"] = (
            df["r_complete_def_physical_glicko_youth"]
            - df["b_complete_def_physical_glicko_youth"]
        )

        # TD defense × Physical × Glicko × Youth: takedown-defensive version of strike_def_physical_glicko_youth
        df["r_td_def_physical_glicko_youth"] = (
            r_td_def_t21
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_td_def_physical_glicko_youth"] = (
            b_td_def_t21
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["td_def_physical_glicko_youth_diff"] = (
            df["r_td_def_physical_glicko_youth"] -
            df["b_td_def_physical_glicko_youth"]
        )

        # Peak × Win rate × Physical × Glicko × Youth: career-peak proven winner who is young and physically gifted
        df["r_peak_winrate_physical_glicko_youth"] = (
            r_peak_t30
            * r_expq_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_peak_winrate_physical_glicko_youth"] = (
            b_peak_t30
            * b_expq_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["peak_winrate_physical_glicko_youth_diff"] = (
            df["r_peak_winrate_physical_glicko_youth"]
            - df["b_peak_winrate_physical_glicko_youth"]
        )

        # Strike defense × Win rate × Physical × Glicko: proven winner who doesn't get hit + physical + elite rating
        df["r_strike_def_winrate_physical_glicko"] = (
            r_str_def_t21
            * r_expq_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
        )
        df["b_strike_def_winrate_physical_glicko"] = (
            b_str_def_t21
            * b_expq_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
        )
        df["strike_def_winrate_physical_glicko_diff"] = (
            df["r_strike_def_winrate_physical_glicko"]
            - df["b_strike_def_winrate_physical_glicko"]
        )

        # Peak × Style × Physical × Glicko × Youth: at career peak + style matchup + physical + rating + youth
        df["r_peak_style_physical_glicko_youth"] = (
            r_peak_t30
            * r_swvc_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_peak_style_physical_glicko_youth"] = (
            b_peak_t30
            * b_swvc_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["peak_style_physical_glicko_youth_diff"] = (
            df["r_peak_style_physical_glicko_youth"]
            - df["b_peak_style_physical_glicko_youth"]
        )

        # Career ELO peak × Win rate × Physical × Glicko: historical ceiling corroborated by career record + physical + rating
        df["r_career_peak_winrate_physical_glicko"] = (
            (r_career_peak_t36 / 1500.0)
            * r_expq_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
        )
        df["b_career_peak_winrate_physical_glicko"] = (
            (b_career_peak_t36 / 1500.0)
            * b_expq_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
        )
        df["career_peak_winrate_physical_glicko_diff"] = (
            df["r_career_peak_winrate_physical_glicko"]
            - df["b_career_peak_winrate_physical_glicko"]
        )

        # Strike defense × Style matchup × Physical × Glicko × Youth: tactical + defensive + physical + rating + youth
        df["r_strike_def_style_physical_glicko_youth"] = (
            r_str_def_t21
            * r_swvc_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_strike_def_style_physical_glicko_youth"] = (
            b_str_def_t21
            * b_swvc_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["strike_def_style_physical_glicko_youth_diff"] = (
            df["r_strike_def_style_physical_glicko_youth"]
            - df["b_strike_def_style_physical_glicko_youth"]
        )

        # Strike defense × Physical × Glicko × Prime: prime-age version (comparison against youth version above)
        df["r_strike_def_physical_glicko_prime"] = (
            r_str_def_t21
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_prime_t30
        )
        df["b_strike_def_physical_glicko_prime"] = (
            b_str_def_t21
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_prime_t30
        )
        df["strike_def_physical_glicko_prime_diff"] = (
            df["r_strike_def_physical_glicko_prime"]
            - df["b_strike_def_physical_glicko_prime"]
        )

        # Peak × Opp quality × Physical × Glicko: career peak + tough schedule + physical + reliable rating
        df["r_peak_opp_quality_physical_glicko"] = (
            r_peak_t30
            * (r_avg_opp_elo_t28 / 1500.0)
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
        )
        df["b_peak_opp_quality_physical_glicko"] = (
            b_peak_t30
            * (b_avg_opp_elo_t28 / 1500.0)
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
        )
        df["peak_opp_quality_physical_glicko_diff"] = (
            df["r_peak_opp_quality_physical_glicko"]
            - df["b_peak_opp_quality_physical_glicko"]
        )

        # Freshness × Strike defense × Physical × Glicko × Youth: rested, elusive, young, physically gifted + elite rating
        df["r_fresh_strike_def_physical_glicko_youth"] = (
            _fresh_r_t30
            * r_str_def_t21
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_fresh_strike_def_physical_glicko_youth"] = (
            _fresh_b_t30
            * b_str_def_t21
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["fresh_strike_def_physical_glicko_youth_diff"] = (
            df["r_fresh_strike_def_physical_glicko_youth"]
            - df["b_fresh_strike_def_physical_glicko_youth"]
        )

        self._log(
            "Tier 38: Combat efficiency, TD defense, and cross-dimension extensions..."
        )

        # TD defense × Win rate × Physical × Glicko × Youth: grappling-wall + proven winner + physical + young + rating
        df["r_td_def_winrate_physical_glicko_youth"] = (
            r_td_def_t21
            * r_expq_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_td_def_winrate_physical_glicko_youth"] = (
            b_td_def_t21
            * b_expq_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["td_def_winrate_physical_glicko_youth_diff"] = (
            df["r_td_def_winrate_physical_glicko_youth"]
            - df["b_td_def_winrate_physical_glicko_youth"]
        )

        # Combat efficiency × Physical × Glicko × Youth: damage-efficient + long reach + elite rating + youth
        df["r_combat_eff_physical_glicko_youth"] = (
            r_cbt_t31
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_combat_eff_physical_glicko_youth"] = (
            b_cbt_t31
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["combat_eff_physical_glicko_youth_diff"] = (
            df["r_combat_eff_physical_glicko_youth"]
            - df["b_combat_eff_physical_glicko_youth"]
        )

        # Finish rate × Strike defense × Physical × Glicko × Youth: can finish + hard to hit + physical + rating + youth
        df["r_finish_def_physical_glicko_youth"] = (
            r_fr_t23
            * r_str_def_t21
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_finish_def_physical_glicko_youth"] = (
            b_fr_t23
            * b_str_def_t21
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["finish_def_physical_glicko_youth_diff"] = (
            df["r_finish_def_physical_glicko_youth"]
            - df["b_finish_def_physical_glicko_youth"]
        )

        # Streak × Win rate × Physical × Glicko × Youth: hot streak + proven record + physical + rating + youth
        df["r_streak_winrate_physical_glicko_youth"] = (
            np.clip(r_streak_t23 / 5.0, 0.0, 1.0)
            * r_expq_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_streak_winrate_physical_glicko_youth"] = (
            np.clip(b_streak_t23 / 5.0, 0.0, 1.0)
            * b_expq_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["streak_winrate_physical_glicko_youth_diff"] = (
            df["r_streak_winrate_physical_glicko_youth"]
            - df["b_streak_winrate_physical_glicko_youth"]
        )

        # KO rate × Strike defense × Physical × Glicko × Youth: KO threat + evasive + physical + rating + youth
        df["r_ko_rate_def_physical_glicko_youth"] = (
            r_ko_rate_t34
            * r_str_def_t21
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_ko_rate_def_physical_glicko_youth"] = (
            b_ko_rate_t34
            * b_str_def_t21
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["ko_rate_def_physical_glicko_youth_diff"] = (
            df["r_ko_rate_def_physical_glicko_youth"]
            - df["b_ko_rate_def_physical_glicko_youth"]
        )

        # TD defense × Win rate × Physical × Glicko × Prime: prime-age grappling-defensive proven winner
        df["r_td_def_winrate_physical_glicko_prime"] = (
            r_td_def_t21
            * r_expq_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_prime_t30
        )
        df["b_td_def_winrate_physical_glicko_prime"] = (
            b_td_def_t21
            * b_expq_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_prime_t30
        )
        df["td_def_winrate_physical_glicko_prime_diff"] = (
            df["r_td_def_winrate_physical_glicko_prime"]
            - df["b_td_def_winrate_physical_glicko_prime"]
        )

        # Both defense walls × Win rate × Physical × Glicko: complete defensive wall + proven record + physical + rating
        df["r_complete_def_winrate_physical_glicko"] = (
            r_str_def_t21
            * r_td_def_t21
            * r_expq_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
        )
        df["b_complete_def_winrate_physical_glicko"] = (
            b_str_def_t21
            * b_td_def_t21
            * b_expq_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
        )
        df["complete_def_winrate_physical_glicko_diff"] = (
            df["r_complete_def_winrate_physical_glicko"]
            - df["b_complete_def_winrate_physical_glicko"]
        )

        # Career ELO peak × Strike defense × Physical × Glicko: historically elite + evasive + physical + current rating
        df["r_career_peak_strike_def_physical_glicko"] = (
            (r_career_peak_t36 / 1500.0)
            * r_str_def_t21
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
        )
        df["b_career_peak_strike_def_physical_glicko"] = (
            (b_career_peak_t36 / 1500.0)
            * b_str_def_t21
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
        )
        df["career_peak_strike_def_physical_glicko_diff"] = (
            df["r_career_peak_strike_def_physical_glicko"]
            - df["b_career_peak_strike_def_physical_glicko"]
        )

        # Peak × TD defense × Physical × Glicko × Youth: at career peak + grappling-solid + physical + rating + youth
        df["r_peak_td_def_physical_glicko_youth"] = (
            r_peak_t30
            * r_td_def_t21
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_peak_td_def_physical_glicko_youth"] = (
            b_peak_t30
            * b_td_def_t21
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["peak_td_def_physical_glicko_youth_diff"] = (
            df["r_peak_td_def_physical_glicko_youth"]
            - df["b_peak_td_def_physical_glicko_youth"]
        )

        # Combat efficiency × Win rate × Physical × Glicko: damage-efficient + proven record + physical + reliable rating
        df["r_combat_eff_winrate_physical_glicko"] = (
            r_cbt_t31
            * r_expq_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
        )
        df["b_combat_eff_winrate_physical_glicko"] = (
            b_cbt_t31
            * b_expq_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
        )
        df["combat_eff_winrate_physical_glicko_diff"] = (
            df["r_combat_eff_winrate_physical_glicko"]
            - df["b_combat_eff_winrate_physical_glicko"]
        )

        self._log(
            "Tier 39: Style dimension added to every top compound that lacks it..."
        )

        # Combat efficiency × Style × Physical × Glicko × Youth: cbt was #3 XGB without style; add it
        df["r_combat_eff_style_physical_glicko_youth"] = (
            r_cbt_t31 * r_swvc_t30 * r_ape_t30 *
            (r_glicko_r_t30 / 1500.0) * r_youth_t30
        )
        df["b_combat_eff_style_physical_glicko_youth"] = (
            b_cbt_t31 * b_swvc_t30 * b_ape_t30 *
            (b_glicko_r_t30 / 1500.0) * b_youth_t30
        )
        df["combat_eff_style_physical_glicko_youth_diff"] = (
            df["r_combat_eff_style_physical_glicko_youth"]
            - df["b_combat_eff_style_physical_glicko_youth"]
        )

        # Peak × Combat eff × Style × Physical × Glicko × Youth: merging #1 and #3 XGB features
        df["r_peak_cbt_style_physical_glicko_youth"] = (
            r_peak_t30
            * r_cbt_t31
            * r_swvc_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_peak_cbt_style_physical_glicko_youth"] = (
            b_peak_t30
            * b_cbt_t31
            * b_swvc_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["peak_cbt_style_physical_glicko_youth_diff"] = (
            df["r_peak_cbt_style_physical_glicko_youth"]
            - df["b_peak_cbt_style_physical_glicko_youth"]
        )

        # Freshness × Style × Physical × Glicko × Youth: rested + style advantage + physical + rating + youth
        df["r_fresh_style_physical_glicko_youth"] = (
            _fresh_r_t30
            * r_swvc_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_fresh_style_physical_glicko_youth"] = (
            _fresh_b_t30
            * b_swvc_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["fresh_style_physical_glicko_youth_diff"] = (
            df["r_fresh_style_physical_glicko_youth"]
            - df["b_fresh_style_physical_glicko_youth"]
        )

        # Recent form × Style × Physical × Glicko × Youth: hot streak wins + style + physical + rating + youth
        df["r_recent_style_physical_glicko_youth"] = (
            (_r_roll3 / 3.0)
            * r_swvc_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_recent_style_physical_glicko_youth"] = (
            (_b_roll3 / 3.0)
            * b_swvc_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["recent_style_physical_glicko_youth_diff"] = (
            df["r_recent_style_physical_glicko_youth"]
            - df["b_recent_style_physical_glicko_youth"]
        )

        # TD defense × Style × Physical × Glicko × Youth: grappling-defensive + style + physical + rating + youth
        df["r_td_def_style_physical_glicko_youth"] = (
            r_td_def_t21
            * r_swvc_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_td_def_style_physical_glicko_youth"] = (
            b_td_def_t21
            * b_swvc_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["td_def_style_physical_glicko_youth_diff"] = (
            df["r_td_def_style_physical_glicko_youth"]
            - df["b_td_def_style_physical_glicko_youth"]
        )

        # Finish rate × Style × Physical × Glicko × Youth: finisher + style matchup + physical + rating + youth
        df["r_finish_style_physical_glicko_youth"] = (
            r_fr_t23 * r_swvc_t30 * r_ape_t30 *
            (r_glicko_r_t30 / 1500.0) * r_youth_t30
        )
        df["b_finish_style_physical_glicko_youth"] = (
            b_fr_t23 * b_swvc_t30 * b_ape_t30 *
            (b_glicko_r_t30 / 1500.0) * b_youth_t30
        )
        df["finish_style_physical_glicko_youth_diff"] = (
            df["r_finish_style_physical_glicko_youth"]
            - df["b_finish_style_physical_glicko_youth"]
        )

        # Complete defense × Style × Physical × Glicko × Youth: both walls + style + physical + rating + youth
        df["r_complete_def_style_physical_glicko_youth"] = (
            r_str_def_t21
            * r_td_def_t21
            * r_swvc_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_complete_def_style_physical_glicko_youth"] = (
            b_str_def_t21
            * b_td_def_t21
            * b_swvc_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["complete_def_style_physical_glicko_youth_diff"] = (
            df["r_complete_def_style_physical_glicko_youth"]
            - df["b_complete_def_style_physical_glicko_youth"]
        )

        # Career ELO peak × Style × Physical × Glicko × Youth: historically great + style edge + physical + rating + youth
        df["r_career_peak_style_physical_glicko_youth"] = (
            (r_career_peak_t36 / 1500.0)
            * r_swvc_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_career_peak_style_physical_glicko_youth"] = (
            (b_career_peak_t36 / 1500.0)
            * b_swvc_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["career_peak_style_physical_glicko_youth_diff"] = (
            df["r_career_peak_style_physical_glicko_youth"]
            - df["b_career_peak_style_physical_glicko_youth"]
        )

        # Opp quality × Style × Physical × Glicko × Youth: tough schedule + style + physical + rating + youth
        df["r_opp_quality_style_physical_glicko_youth"] = (
            (r_avg_opp_elo_t28 / 1500.0)
            * r_swvc_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_opp_quality_style_physical_glicko_youth"] = (
            (b_avg_opp_elo_t28 / 1500.0)
            * b_swvc_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["opp_quality_style_physical_glicko_youth_diff"] = (
            df["r_opp_quality_style_physical_glicko_youth"]
            - df["b_opp_quality_style_physical_glicko_youth"]
        )

        # Streak × Style × Physical × Glicko × Youth: win momentum + style edge + physical + rating + youth
        df["r_streak_style_physical_glicko_youth"] = (
            np.clip(r_streak_t23 / 5.0, 0.0, 1.0)
            * r_swvc_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_streak_style_physical_glicko_youth"] = (
            np.clip(b_streak_t23 / 5.0, 0.0, 1.0)
            * b_swvc_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["streak_style_physical_glicko_youth_diff"] = (
            df["r_streak_style_physical_glicko_youth"]
            - df["b_streak_style_physical_glicko_youth"]
        )

        self._log(
            "Tier 40: Decayed average difference features (exponential decay α=0.8, opponent-quality adjusted)..."
        )

        # ── Tier 40: Decayed Average Difference Features ─────────────────────────
        def _da_diff(col):
            r_col = f"r_pre_{col}"
            b_col = f"b_pre_{col}"
            if r_col in df.columns and b_col in df.columns:
                df[f"{col}_diff"] = df[r_col].fillna(0) - df[b_col].fillna(0)

        _da_features = [
            "da_sig_str_acc",
            "da_td_acc",
            "da_head_landed",
            "da_head_acc",
            "da_body_acc",
            "da_distance_acc",
            "da_head_defense",
            "da_body_defense",
            "da_distance_defense",
            "da_ground_defense",
            "da_td_defense",
            "da_sub_att",
            "da_kd",
            "da_ko",
            "da_win_ratio",
            "da_ctrl_r1",
            "da_clinch_pm",
            "da_opp_leg_pm",
            "da_opp_ctrl_r1_pm",
            "da_opp_sub_pm",
            "da_opp_rev_r1",
            "da_r1_strikes",
            "da_reversals",
            "da_dist_landing_ratio",
            "da_opp_kd",
            "da_age",
            "da_ufc_age",
            "da_reach_ratio",
            "da_days_since",
            "dapa_sig_str_acc",
            "dapa_head_acc",
            "dapa_body_acc",
            "dapa_distance_acc",
            "dapa_head_defense",
            "dapa_dist_defense",
            "dapa_ground_defense",
            "dapa_r1_strikes",
            "dapa_reversals",
            "dapa_dist_landing_ratio",
            "dapa_head_landing_ratio",
        ]
        for _f in _da_features:
            _da_diff(_f)

        # ── Physical ratio features ───────────────────────────────────────────────
        # Age ratio: who is the relatively younger fighter (youth advantage)
        _r_age_t40 = df.get("r_age_at_event", pd.Series(28.0, index=df.index)).fillna(
            28.0
        )
        _b_age_t40 = df.get("b_age_at_event", pd.Series(28.0, index=df.index)).fillna(
            28.0
        )
        df["age_ratio_diff"] = (_r_age_t40 / (_b_age_t40 + 0.1)) - 1.0

        _r_reach_t40 = df.get("r_reach", pd.Series(
            70.0, index=df.index)).fillna(70.0)
        _b_reach_t40 = df.get("b_reach", pd.Series(
            70.0, index=df.index)).fillna(70.0)
        df["reach_ratio_diff"] = (_r_reach_t40 / (_b_reach_t40 + 0.1)) - 1.0

        # ── Composite decayed features ────────────────────────────────────────────
        # Striking precision composite: decayed sig_str_acc × head_acc (both decayed)
        _r_da_ss = df.get(
            "r_pre_da_sig_str_acc", pd.Series(0.5, index=df.index)
        ).fillna(0.5)
        _b_da_ss = df.get(
            "b_pre_da_sig_str_acc", pd.Series(0.5, index=df.index)
        ).fillna(0.5)
        _r_da_ha = df.get("r_pre_da_head_acc", pd.Series(0.45, index=df.index)).fillna(
            0.45
        )
        _b_da_ha = df.get("b_pre_da_head_acc", pd.Series(0.45, index=df.index)).fillna(
            0.45
        )
        df["da_striking_precision_diff"] = (
            _r_da_ss * _r_da_ha) - (_b_da_ss * _b_da_ha)

        # Defense composite: average of head, distance, TD defense (decayed)
        _r_da_hd = df.get(
            "r_pre_da_head_defense", pd.Series(0.55, index=df.index)
        ).fillna(0.55)
        _b_da_hd = df.get(
            "b_pre_da_head_defense", pd.Series(0.55, index=df.index)
        ).fillna(0.55)
        _r_da_dd = df.get(
            "r_pre_da_distance_defense", pd.Series(0.55, index=df.index)
        ).fillna(0.55)
        _b_da_dd = df.get(
            "b_pre_da_distance_defense", pd.Series(0.55, index=df.index)
        ).fillna(0.55)
        _r_da_td_def = df.get(
            "r_pre_da_td_defense", pd.Series(0.6, index=df.index)
        ).fillna(0.6)
        _b_da_td_def = df.get(
            "b_pre_da_td_defense", pd.Series(0.6, index=df.index)
        ).fillna(0.6)
        df["r_da_defense_composite"] = (
            _r_da_hd + _r_da_dd + _r_da_td_def) / 3.0
        df["b_da_defense_composite"] = (
            _b_da_hd + _b_da_dd + _b_da_td_def) / 3.0
        df["da_defense_composite_diff"] = (
            df["r_da_defense_composite"] - df["b_da_defense_composite"]
        )

        # Grappling threat: decayed td_acc × sub_att (offensive grappling)
        _r_da_tdacc = df.get("r_pre_da_td_acc", pd.Series(0.4, index=df.index)).fillna(
            0.4
        )
        _b_da_tdacc = df.get("b_pre_da_td_acc", pd.Series(0.4, index=df.index)).fillna(
            0.4
        )
        _r_da_sub = df.get("r_pre_da_sub_att", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        _b_da_sub = df.get("b_pre_da_sub_att", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        df["da_grapple_threat_diff"] = _r_da_tdacc * np.log1p(
            _r_da_sub
        ) - _b_da_tdacc * np.log1p(_b_da_sub)

        # Opponent pressure absorbed: opp_leg_pm + opp_sub_pm + opp_ctrl_r1_pm
        _r_da_olegpm = df.get(
            "r_pre_da_opp_leg_pm", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        _b_da_olegpm = df.get(
            "b_pre_da_opp_leg_pm", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        _r_da_osubpm = df.get(
            "r_pre_da_opp_sub_pm", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        _b_da_osubpm = df.get(
            "b_pre_da_opp_sub_pm", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        _r_da_ocr1 = df.get(
            "r_pre_da_opp_ctrl_r1_pm", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        _b_da_ocr1 = df.get(
            "b_pre_da_opp_ctrl_r1_pm", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        df["da_opp_pressure_diff"] = (_r_da_olegpm + _r_da_osubpm + _r_da_ocr1) - (
            _b_da_olegpm + _b_da_osubpm + _b_da_ocr1
        )

        # KO power × head accuracy (adjusted): finishing power composite
        _r_da_ko_t40 = df.get("r_pre_da_ko", pd.Series(
            0.0, index=df.index)).fillna(0.0)
        _b_da_ko_t40 = df.get("b_pre_da_ko", pd.Series(
            0.0, index=df.index)).fillna(0.0)
        _r_dapa_ha = df.get(
            "r_pre_dapa_head_acc", pd.Series(0.45, index=df.index)
        ).fillna(0.45)
        _b_dapa_ha = df.get(
            "b_pre_dapa_head_acc", pd.Series(0.45, index=df.index)
        ).fillna(0.45)
        df["da_ko_x_head_acc_diff"] = (_r_da_ko_t40 * _r_dapa_ha) - (
            _b_da_ko_t40 * _b_dapa_ha
        )

        # Win ratio × defense composite: consistent winners who also defend well
        _r_da_wr = df.get("r_pre_da_win_ratio", pd.Series(0.5, index=df.index)).fillna(
            0.5
        )
        _b_da_wr = df.get("b_pre_da_win_ratio", pd.Series(0.5, index=df.index)).fillna(
            0.5
        )
        df["da_win_x_defense_diff"] = (
            _r_da_wr * df["r_da_defense_composite"]
            - _b_da_wr * df["b_da_defense_composite"]
        )

        # ── Tier 41: Method-signal features ────────────────────────────────
        # Give the method model the same signals that the rule-based inference
        # blending uses, so it can learn optimal weighting during training
        # rather than relying on hand-tuned post-hoc blends.
        #
        # NOTE: Only r_/b_ prefixed features are created here — no _diff
        # columns.  The D+I decomposition automatically produces the
        # antisymmetric (differential) component from each r_/b_ pair, so
        # explicit diffs would be redundant and break under corner swap.
        # Symmetric features (cross_stance_ko_potential, etc.) become Inv
        # components, which is correct — they describe the fight, not a corner.
        self._log(
            "Tier 41: Method-signal features (susceptibility interactions, "
            "effective method rates, stance KO interactions)..."
        )

        # --- 41a: Susceptibility interaction features ---
        # "How I win" × "How opponent loses" — direct cross-fighter signal.
        _r_ko_wr = df.get("r_ko_win_rate", pd.Series(
            0, index=df.index)).fillna(0)
        _b_ko_wr = df.get("b_ko_win_rate", pd.Series(
            0, index=df.index)).fillna(0)
        _r_sub_wr = df.get("r_sub_win_rate", pd.Series(
            0, index=df.index)).fillna(0)
        _b_sub_wr = df.get("b_sub_win_rate", pd.Series(
            0, index=df.index)).fillna(0)
        _r_dec_wr = df.get("r_decision_win_rate", pd.Series(0, index=df.index)).fillna(
            0
        )
        _b_dec_wr = df.get("b_decision_win_rate", pd.Series(0, index=df.index)).fillna(
            0
        )

        _r_ko_lr = df.get("r_ko_loss_rate", pd.Series(
            0, index=df.index)).fillna(0)
        _b_ko_lr = df.get("b_ko_loss_rate", pd.Series(
            0, index=df.index)).fillna(0)
        _r_sub_lr = df.get("r_sub_loss_rate", pd.Series(
            0, index=df.index)).fillna(0)
        _b_sub_lr = df.get("b_sub_loss_rate", pd.Series(
            0, index=df.index)).fillna(0)
        _r_dec_lr = df.get("r_dec_loss_rate", pd.Series(
            0, index=df.index)).fillna(0)
        _b_dec_lr = df.get("b_dec_loss_rate", pd.Series(
            0, index=df.index)).fillna(0)

        # Red's KO power × Blue's KO vulnerability (and vice versa)
        df["r_ko_susc_interaction"] = _r_ko_wr * _b_ko_lr
        df["b_ko_susc_interaction"] = _b_ko_wr * _r_ko_lr

        df["r_sub_susc_interaction"] = _r_sub_wr * _b_sub_lr
        df["b_sub_susc_interaction"] = _b_sub_wr * _r_sub_lr

        df["r_dec_susc_interaction"] = _r_dec_wr * _b_dec_lr
        df["b_dec_susc_interaction"] = _b_dec_wr * _r_dec_lr

        # Finish threat: max(KO interaction, sub interaction) for each corner
        df["r_finish_threat"] = np.maximum(
            df["r_ko_susc_interaction"], df["r_sub_susc_interaction"]
        )
        df["b_finish_threat"] = np.maximum(
            df["b_ko_susc_interaction"], df["b_sub_susc_interaction"]
        )

        # --- 41b: Effective method rate features ---
        # Blend fighter's win rates with opponent's loss rates (susceptibility),
        # mirroring the inference rule-blend logic.
        _SUSC_DEFAULT = 0.35
        for _pfx, _opp in [("r", "b"), ("b", "r")]:
            _ko_w = df.get(f"{_pfx}_ko_win_rate", pd.Series(0, index=df.index)).fillna(
                0
            )
            _sub_w = df.get(
                f"{_pfx}_sub_win_rate", pd.Series(0, index=df.index)
            ).fillna(0)
            _dec_w = df.get(
                f"{_pfx}_decision_win_rate", pd.Series(0, index=df.index)
            ).fillna(0)
            _ko_l = df.get(f"{_opp}_ko_loss_rate", pd.Series(0, index=df.index)).fillna(
                0
            )
            _sub_l = df.get(
                f"{_opp}_sub_loss_rate", pd.Series(0, index=df.index)
            ).fillna(0)
            _dec_l = df.get(
                f"{_opp}_dec_loss_rate", pd.Series(0, index=df.index)
            ).fillna(0)

            eff_ko = (1.0 - _SUSC_DEFAULT) * _ko_w + _SUSC_DEFAULT * _ko_l
            eff_sub = (1.0 - _SUSC_DEFAULT) * _sub_w + _SUSC_DEFAULT * _sub_l
            eff_dec = (1.0 - _SUSC_DEFAULT) * _dec_w + _SUSC_DEFAULT * _dec_l
            eff_tot = (eff_ko + eff_sub + eff_dec).clip(lower=1e-9)
            df[f"{_pfx}_eff_ko_rate"] = eff_ko / eff_tot
            df[f"{_pfx}_eff_sub_rate"] = eff_sub / eff_tot
            df[f"{_pfx}_eff_dec_rate"] = eff_dec / eff_tot

        # Combined effective finish rate (KO + Sub) vs decision
        df["r_eff_finish_rate"] = df["r_eff_ko_rate"] + df["r_eff_sub_rate"]
        df["b_eff_finish_rate"] = df["b_eff_ko_rate"] + df["b_eff_sub_rate"]

        # --- 41c: Stance-specific KO features ---
        # Southpaw vs orthodox matchups historically produce more KOs due to
        # unfamiliar angles.  Encode this as an interaction with KO rates.
        _r_st41 = (
            df.get("r_stance", pd.Series("", index=df.index))
            .fillna("")
            .astype(str)
            .str.strip()
            .str.lower()
        )
        _b_st41 = (
            df.get("b_stance", pd.Series("", index=df.index))
            .fillna("")
            .astype(str)
            .str.strip()
            .str.lower()
        )

        # Is this a cross-stance matchup? (orthodox vs southpaw or vice versa)
        _cross_stance = (
            ((_r_st41 == "orthodox") & (_b_st41 == "southpaw"))
            | ((_r_st41 == "southpaw") & (_b_st41 == "orthodox"))
        ).astype(float)

        # Is this a mirror matchup? (same stance)
        _mirror = (_r_st41 == _b_st41).astype(float)

        # Southpaw indicator per corner (southpaw or switch fighters have
        # southpaw-angle advantage against orthodox)
        _r_southpaw = _r_st41.isin(
            ["southpaw", "switch", "open stance"]).astype(float)
        _b_southpaw = _b_st41.isin(
            ["southpaw", "switch", "open stance"]).astype(float)

        # KO rate amplified by cross-stance (unfamiliar angles boost KO chance)
        df["r_ko_rate_cross_stance"] = _r_ko_wr * _cross_stance
        df["b_ko_rate_cross_stance"] = _b_ko_wr * _cross_stance

        # Southpaw angle advantage × KO rate (southpaw facing orthodox gets a
        # boost; orthodox facing southpaw does too but from the other column)
        df["r_southpaw_ko_advantage"] = _r_southpaw * \
            _r_ko_wr * (1.0 - _b_southpaw)
        df["b_southpaw_ko_advantage"] = _b_southpaw * \
            _b_ko_wr * (1.0 - _r_southpaw)

        # Switch stance × finish rate (switch fighters are harder to read)
        _r_is_switch = _r_st41.isin(["switch", "open stance"]).astype(float)
        _b_is_switch = _b_st41.isin(["switch", "open stance"]).astype(float)
        df["r_switch_finish_threat"] = _r_is_switch * df.get(
            "r_finish_rate", pd.Series(0, index=df.index)
        ).fillna(0)
        df["b_switch_finish_threat"] = _b_is_switch * df.get(
            "b_finish_rate", pd.Series(0, index=df.index)
        ).fillna(0)

        # Symmetric features (fight-level, not corner-specific):
        # Cross-stance combined KO potential (both fighters face unfamiliar angles)
        df["cross_stance_ko_potential"] = _cross_stance * \
            (_r_ko_wr + _b_ko_wr) / 2
        # Mirror matchup decision tendency (same stance = more technical, less finishes)
        df["mirror_stance_dec_tendency"] = _mirror * \
            (_r_dec_wr + _b_dec_wr) / 2

        # ── Tier 42: Method-path closure, uncertainty routing, SVD method-alignment ──
        self._log(
            "Tier 42: Method-path closure, uncertainty routing, and SVD method-alignment features..."
        )
        _eps = 1e-6
        _cols_before_t42 = set(df.columns)

        def _g42(col, default=0.0):
            return df.get(col, pd.Series(default, index=df.index)).fillna(default)

        # --- helper: second-highest of three series ---
        def _second_max(a, b, c):
            stacked = np.column_stack([a, b, c])
            stacked.sort(axis=1)
            return stacked[:, 1]

        for _pfx, _opp in [("r", "b"), ("b", "r")]:
            # -- 1) submission path closure --
            # TD->control->submission chain closing into real submission threat
            _spc = (
                (
                    _g42(f"{_pfx}_pre_sub_after_td")
                    * _g42(f"{_pfx}_pre_ctrl_per_td")
                    * _g42(f"{_pfx}_pre_sub_efficiency")
                    * (1.0 - _g42(f"{_opp}_pre_td_def", 0.5))
                )
                .clip(0)
                .fillna(0)
            )
            df[f"{_pfx}_sub_path_closure"] = _spc

            # -- 2) late submission conversion --
            # Submission path that opens later in the fight
            df[f"{_pfx}_late_sub_conversion"] = (
                (
                    (
                        _g42(f"{_pfx}_pre_r23_sub_rate")
                        + 0.5 * _g42(f"{_pfx}_pre_r45_sub_rate_l")
                    )
                    * _g42(f"{_pfx}_pre_late_td_acc")
                    * (1.0 - _g42(f"{_opp}_pre_r1_td_def_rate", 0.5))
                )
                .clip(0)
                .fillna(0)
            )

            # -- 3) scramble submission window --
            # Messy scramble-driven submission danger
            df[f"{_pfx}_scramble_sub_window"] = (
                (
                    (
                        _g42(f"{_pfx}_pre_rev_per_ctrl")
                        + _g42(f"{_pfx}_pre_reversal_rate")
                    )
                    * _g42(f"{_pfx}_pre_sub_efficiency")
                    * (1.0 + _g42(f"{_opp}_pre_output_cv", 0.5))
                )
                .clip(0)
                .fillna(0)
            )

            # -- 4) anti-sub escape integrity --
            # Defender-side resistance to submission finishes
            df[f"{_pfx}_anti_sub_escape_integrity"] = (
                (
                    _g42(f"{_pfx}_pre_td_def", 0.5)
                    * _g42(f"{_pfx}_pre_r1_td_def_rate", 0.5)
                    * _g42(f"{_pfx}_pre_output_consistency", 0.5)
                    / (
                        1.0
                        + _g42(f"{_opp}_pre_sub_after_td")
                        + _g42(f"{_opp}_pre_sub_efficiency")
                    )
                )
                .clip(0)
                .fillna(0)
            )

            # -- 5) decision bankability --
            # True "minutes winner" who banks rounds safely
            _db = (
                (
                    _g42(f"{_pfx}_eff_dec_rate")
                    * _g42(f"{_pfx}_pre_output_consistency", 0.5)
                    * np.log1p(np.maximum(_g42(f"{_pfx}_pre_dec_win_margin").values, 0))
                    * (1.0 + _g42(f"{_pfx}_pre_lead_hold_rate"))
                )
                .clip(0)
                .fillna(0)
            )
            df[f"{_pfx}_decision_bankability"] = _db

            # -- 6) finish conversion pressure --
            # Offensive forcing style plus opponent fragility
            df[f"{_pfx}_finish_conversion_pressure"] = (
                (
                    _g42(f"{_pfx}_eff_finish_rate")
                    * np.log1p(
                        _g42(f"{_pfx}_pre_pressure_index").values
                        + _g42(f"{_pfx}_pre_r1_td_pressure").values
                    )
                    * (1.0 + _g42(f"{_pfx}_pre_output_eff_trend"))
                    * (
                        1.0
                        + _g42(f"{_opp}_pre_r1_abs_rate")
                        + _g42(f"{_opp}_pre_chin_ratio", 1.0)
                    )
                )
                .clip(0)
                .fillna(0)
            )

            # -- 7) attritional break score --
            # Late-breaking fighters who wear opponents down over time
            df[f"{_pfx}_attritional_break_score"] = (
                (
                    _g42(f"{_pfx}_pre_cardio_index", 0.75)
                    * _g42(f"{_pfx}_pre_late_win_rate", 0.5)
                    * _g42(f"{_pfx}_pre_r2_to_r3_momentum")
                    * (1.0 + _g42(f"{_opp}_pre_fatigue_composite"))
                )
                .clip(0)
                .fillna(0)
            )

            # -- 8) early chaos break score --
            # Immediate volatility and fast KO/TKO chaos
            _ecbs = (
                (
                    _g42(f"{_pfx}_pre_r1_finish_rate")
                    * _g42(f"{_pfx}_pre_kd_efficiency")
                    * (
                        1.0
                        + _g42(f"{_pfx}_ko_rate_cross_stance")
                        + _g42(f"{_pfx}_southpaw_ko_advantage")
                    )
                    * (1.0 + _g42(f"{_opp}_pre_r1_abs_rate"))
                )
                .clip(0)
                .fillna(0)
            )
            df[f"{_pfx}_early_chaos_break_score"] = _ecbs

            # -- 9) volatility finish chaos --
            # Uncertainty and instability driving finish routing
            df[f"{_pfx}_volatility_finish_chaos"] = (
                (
                    _g42(f"{_pfx}_glicko_pre_vol", 0.06)
                    * _g42(f"{_pfx}_eff_finish_rate")
                    * (1.0 + _g42(f"{_opp}_glicko_pre_vol", 0.06))
                    * (
                        1.0
                        + np.abs(
                            _g42(f"{_pfx}_pre_output_cv", 0.5).values
                            - _g42(f"{_opp}_pre_output_cv", 0.5).values
                        )
                    )
                )
                .clip(0)
                .fillna(0)
            )

            # -- 10) layoff decay mismatch --
            # Layoff + decayed sharpness interacting with opponent pressure
            df[f"{_pfx}_layoff_decay_mismatch"] = (
                (
                    np.log1p(_g42(f"{_pfx}_days_since_last", 180).values)
                    * (1.0 - _g42(f"{_pfx}_pre_accuracy_retention", 1.0))
                    * (
                        1.0
                        + _g42(f"{_opp}_pre_pressure_index")
                        + _g42(f"{_opp}_eff_finish_rate")
                    )
                )
                .clip(0)
                .fillna(0)
            )

            # -- 11) style-method mismatch --
            # Style advantage translated into method-route tendency
            df[f"{_pfx}_style_method_mismatch"] = (
                _g42(f"{_pfx}_style_win_vs_cluster", 0.5)
                * (
                    _g42(f"{_pfx}_eff_sub_rate")
                    + 0.5 * _g42(f"{_pfx}_eff_ko_rate")
                    - _g42(f"{_pfx}_eff_dec_rate")
                )
                * _g42(f"{_pfx}_physical_style_dominance")
            ).fillna(0)

            # -- 12) method path separation --
            # Clean method identity vs muddy one (gap between top-2 method rates)
            _eko = _g42(f"{_pfx}_eff_ko_rate").values
            _esub = _g42(f"{_pfx}_eff_sub_rate").values
            _edec = _g42(f"{_pfx}_eff_dec_rate").values
            _top = np.maximum(np.maximum(_eko, _esub), _edec)
            _second = _second_max(_eko, _esub, _edec)
            df[f"{_pfx}_method_path_separation"] = (
                pd.Series(_top - _second, index=df.index).clip(0).fillna(0)
            )

            # -- 13) SVD submission alignment --
            # Latent grappling/form structure routed into submission completion
            df[f"{_pfx}_svd_sub_alignment"] = (
                (0.6 * _g42("grappling_svd_0") + 0.4 * _g42("form_svd_1"))
                * df[f"{_pfx}_sub_path_closure"]
            ).fillna(0)

            # -- 14) SVD decision-control alignment --
            # Latent control archetype routed into decision winning
            df[f"{_pfx}_svd_decision_control_alignment"] = (
                (0.5 * _g42("grappling_svd_1") + 0.5 * _g42("form_svd_0"))
                * df[f"{_pfx}_decision_bankability"]
            ).fillna(0)

            # -- 15) SVD striking-break alignment --
            # Striking/form/physical latent structure routed into KO/TKO break potential
            df[f"{_pfx}_svd_striking_break_alignment"] = (
                (
                    0.5 * _g42("striking_svd_0")
                    + 0.3 * _g42("form_svd_2")
                    + 0.2 * _g42("physical_svd_0")
                )
                * df[f"{_pfx}_early_chaos_break_score"]
            ).fillna(0)

            # -- 16) SVD archetype instability --
            # Internally unstable or hard-to-classify archetypes hurting method confidence
            df[f"{_pfx}_svd_archetype_instability"] = (
                np.abs(_g42("striking_svd_0").values -
                       _g42("grappling_svd_0").values)
                * _g42(f"{_pfx}_glicko_pre_vol", 0.06).values
                * _g42(f"{_pfx}_pre_output_cv", 0.5).values
            )
            df[f"{_pfx}_svd_archetype_instability"] = df[
                f"{_pfx}_svd_archetype_instability"
            ].fillna(0)

        # --- diff columns for all 16 concepts ---
        _t42_concepts = [
            "sub_path_closure",
            "late_sub_conversion",
            "scramble_sub_window",
            "anti_sub_escape_integrity",
            "decision_bankability",
            "finish_conversion_pressure",
            "attritional_break_score",
            "early_chaos_break_score",
            "volatility_finish_chaos",
            "layoff_decay_mismatch",
            "style_method_mismatch",
            "method_path_separation",
            "svd_sub_alignment",
            "svd_decision_control_alignment",
            "svd_striking_break_alignment",
            "svd_archetype_instability",
        ]
        for _concept in _t42_concepts:
            df[f"{_concept}_diff"] = df[f"r_{_concept}"] - df[f"b_{_concept}"]

        _new_t42 = [c for c in df.columns if c not in _cols_before_t42]

        self._log(
            "Tier 43: Method-head vulnerability + conversion (Decision/Finish, KO/Sub)..."
        )
        df = self._apply_tier43_method_head_features(df)

        self.df = df
        print_metric("Feature columns added:", len(df.columns))

    def _get_feature_cols(self):
        exclude = {
            "event_date",
            "event_name",
            "event_location",
            "r_fighter",
            "b_fighter",
            "weight_class",
            "gender",
            "winner",
            "method",
            "referee",
            "r_stance",
            "b_stance",
            "r_date_of_birth",
            "b_date_of_birth",
            "finish_round",
            "time_sec",
            "r_name",
            "b_name",
            "stance_matchup",
            "_history",
        }
        cols = []
        for col in self.df.columns:
            if col in exclude:
                continue
            if str(col).startswith("_wc_"):
                continue
            if self.df[col].dtype in [object, "object"]:
                continue
            cols.append(col)
        return cols

    def _build_X_y(self, df=None):
        if df is None:
            df = self.df
        feat_cols = self._get_feature_cols()

        feat_cols = [c for c in feat_cols if c in df.columns]
        X = df[feat_cols].fillna(0).replace([np.inf, -np.inf], 0).values
        y_raw = df["winner"].values
        y = np.array([1 if w == "Red" else 0 for w in y_raw])
        return X, y, feat_cols

    # Antisymmetric feature decomposition: splits features into a corner-invariant
    # representation. Given X (original) and X_swap (red/blue corners swapped):
    #   D   = 0.5 * (X_orig - X_swap)  -> antisymmetric / "difference" component
    #   Inv = 0.5 * (X_orig + X_swap)  -> symmetric / "invariant" component
    # The D block captures genuine fighter advantage; the Inv block captures fight-level
    # context that doesn't depend on which corner a fighter is assigned.
    # At inference, flipping the sign of D columns is how we run corner-swap predictions.
    def _decompose_features(self, X_orig, X_swap):
        if isinstance(X_orig, pd.DataFrame):
            cols = X_orig.columns.tolist()
            D = 0.5 * (X_orig.values - X_swap.values)
            Inv = 0.5 * (X_orig.values + X_swap.values)
            D_df = pd.DataFrame(D, columns=cols, index=X_orig.index)
            I_df = pd.DataFrame(
                Inv, columns=[c + "_inv" for c in cols], index=X_orig.index
            )
            return pd.concat([D_df, I_df], axis=1)
        else:
            D = 0.5 * (X_orig - X_swap)
            Inv = 0.5 * (X_orig + X_swap)
            return np.concatenate([D, Inv], axis=1)

    # Full training pipeline:
    #   1. Chronological 90/10 train-test split (no shuffle — respects time order)
    #   2. Corner-swap augmentation doubles the training set and forces corner-invariance
    #   3. Antisymmetric D+I decomposition applied to all feature matrices
    #   4. Recency sample weights: fights from 2 years ago get ~50% weight of recent ones
    #   5. StandardScaler fitted on training data only
    #   6. Stability Selection (50 runs, LightGBM, 50% subsampling) selects stable feature subset
    #   7. Optuna hyperparameter search for XGBoost (25 trials) and LightGBM (15 trials)
    #   8. All base models fitted on full training set
    #   9. _ManualStackingEnsemble (OOF PurgedTimeSeriesSplit-3) with LR meta-learner
    #  10. Separate method classifier (Decision / KO / Submission) trained on same data
    def train(self, data_hash=None):
        print_section("TRAINING ENSEMBLE MODEL")
        self._log("Detecting GPU...")
        self.gpu_info = detect_gpu()
        print_metric("XGB GPU:", self.gpu_info["xgb"])
        print_metric("LGB GPU:", self.gpu_info["lgb"])
        print_metric("CAT GPU:", self.gpu_info["cat"])
        print()

        # ── Stage 2: preprocessing ──────────────────────────────────────
        _preproc_cached = (
            _load_stage_cache(
                "preprocessing", data_hash) if data_hash else None
        )
        if _preproc_cached is not None:
            X_tr_s = _preproc_cached["X_tr_s"]
            X_val_s = _preproc_cached["X_val_s"]
            X_test_s = _preproc_cached["X_test_s"]
            X_tr_raw = _preproc_cached["X_tr_raw"]
            X_val_raw = _preproc_cached["X_val_raw"]
            X_test_raw = _preproc_cached["X_test_raw"]
            y_tr = _preproc_cached["y_tr"]
            y_val = _preproc_cached["y_val"]
            y_test = _preproc_cached["y_test"]
            w_tr = _preproc_cached["w_tr"]
            _aug_groups = _preproc_cached["_aug_groups"]
            feat_col_list = _preproc_cached["feat_col_list"]
            df_tr = _preproc_cached["df_tr"]
            df_val = _preproc_cached["df_val"]
            df_test = _preproc_cached["df_test"]
            df_tr_aug = _preproc_cached["df_tr_aug"]
            df_train = _preproc_cached["df_train"]
            train_end = _preproc_cached["train_end"]
            val_end = _preproc_cached["val_end"]
            self.feature_cols = _preproc_cached["feature_cols"]
            self._decomposed_cols = _preproc_cached["_decomposed_cols"]
            self.scaler = _preproc_cached["scaler"]
            self._d_indices = _preproc_cached["_d_indices"]
            _HALF_LIFE_DAYS = _preproc_cached["_HALF_LIFE_DAYS"]
            _best_floor = _preproc_cached["_best_floor"]
            print_metric("Train samples (augmented):", len(X_tr_s))
            print_metric("Val samples (held out):", len(X_val_s))
            print_metric("Test samples (holdout):", len(df_test))
            print_metric("Features after decomposition:",
                         len(self._decomposed_cols))
            print()
        else:
            df_train = self.df[self.df["winner"] != "Draw"].copy()

            n = len(df_train)
            train_end = int(n * 0.80)
            val_end = int(n * 0.90)
            df_tr = df_train.iloc[:train_end]
            df_val = df_train.iloc[train_end:val_end]
            df_test = df_train.iloc[val_end:]

            df_aug = self._corner_swap(df_tr)
            df_tr_aug = pd.concat([df_tr, df_aug], ignore_index=True)

            X_tr, y_tr, feat_cols = self._build_X_y(df_tr_aug)
            X_val, y_val, _ = self._build_X_y(df_val)
            self.feature_cols = feat_cols

            print_metric("Train samples (augmented):", len(X_tr))
            print_metric("Val samples (held out):", len(X_val))
            print_metric("Test samples (holdout):", len(df_test))
            print_metric("Features:", len(feat_cols))
            print()

            print_step(
                "Applying antisymmetric feature decomposition (D + I)...")
            df_train_feat = df_tr_aug.copy()
            df_train_swap = self._corner_swap(df_train_feat)
            feat_col_list = self._get_feature_cols()
            feat_col_list = [
                c for c in feat_col_list if c in df_train_feat.columns]
            X_orig_df = df_train_feat[feat_col_list].fillna(0)
            X_swap_df = df_train_swap[feat_col_list].fillna(0)
            X_decomposed = self._decompose_features(X_orig_df, X_swap_df)
            self._decomposed_cols = X_decomposed.columns.tolist()
            print_metric("  Features after decomposition:",
                         X_decomposed.shape[1])
            print()

            df_val_feat = df_val.copy()
            df_val_swap = self._corner_swap(df_val_feat)
            X_val_orig_df = df_val_feat[feat_col_list].fillna(0)
            X_val_swap_df = df_val_swap[feat_col_list].fillna(0)
            X_val_decomposed = self._decompose_features(
                X_val_orig_df, X_val_swap_df)

            df_test_feat = df_test.copy()
            df_test_swap = self._corner_swap(df_test_feat)
            X_test_orig_df = df_test_feat[feat_col_list].fillna(0)
            X_test_swap_df = df_test_swap[feat_col_list].fillna(0)
            X_test_decomposed = self._decompose_features(
                X_test_orig_df, X_test_swap_df)

            X_tr_raw = X_decomposed.values
            X_val_raw = X_val_decomposed.values
            X_test_raw = X_test_decomposed.values
            y_tr = np.array(
                [1 if w == "Red" else 0 for w in df_train_feat["winner"].values]
            )
            y_val = np.array(
                [1 if w == "Red" else 0 for w in df_val_feat["winner"].values]
            )
            y_test = np.array(
                [1 if w == "Red" else 0 for w in df_test_feat["winner"].values]
            )

            _n_orig = len(df_tr)
            _aug_groups = np.tile(np.arange(_n_orig), 2)[: len(y_tr)]

            print_step(
                "Computing recency sample weights (2.5-year half-life, floor=0.08)..."
            )
            _dates_tr = pd.to_datetime(
                df_train_feat["event_date"], errors="coerce")
            _most_recent_date = _dates_tr.max()
            _days_ago = (
                (_most_recent_date - _dates_tr)
                .dt.days.fillna(0)
                .clip(lower=0)
                .values.astype(float)
            )

            X_tr_s = self.scaler.fit_transform(X_tr_raw)
            X_val_s = self.scaler.transform(X_val_raw)
            X_test_s = self.scaler.transform(X_test_raw)

            _HALF_LIFE_DAYS = 365.0 * 2.5
            _best_floor = 0.08
            w_tr = np.exp(-np.log(2) / _HALF_LIFE_DAYS * _days_ago)
            w_tr = np.maximum(w_tr, _best_floor)
            w_tr = w_tr / w_tr.mean()
            print_metric(
                "  Weight range:",
                f"{w_tr.min():.3f} – {w_tr.max():.3f}  (mean = 1.000)",
            )

            self._d_indices = [
                i for i, c in enumerate(self._decomposed_cols) if not c.endswith("_inv")
            ]
            print()

            if data_hash:
                _save_stage_cache(
                    "preprocessing",
                    data_hash,
                    {
                        "X_tr_s": X_tr_s,
                        "X_val_s": X_val_s,
                        "X_test_s": X_test_s,
                        "X_tr_raw": X_tr_raw,
                        "X_val_raw": X_val_raw,
                        "X_test_raw": X_test_raw,
                        "y_tr": y_tr,
                        "y_val": y_val,
                        "y_test": y_test,
                        "w_tr": w_tr,
                        "_aug_groups": _aug_groups,
                        "feat_col_list": feat_col_list,
                        "df_tr": df_tr,
                        "df_val": df_val,
                        "df_test": df_test,
                        "df_tr_aug": df_tr_aug,
                        "df_train": df_train,
                        "train_end": train_end,
                        "val_end": val_end,
                        "feature_cols": self.feature_cols,
                        "_decomposed_cols": self._decomposed_cols,
                        "scaler": self.scaler,
                        "_d_indices": self._d_indices,
                        "_HALF_LIFE_DAYS": _HALF_LIFE_DAYS,
                        "_best_floor": _best_floor,
                    },
                )

        # ── Stage 3: winner_model (cache check) ────────────────────────
        # If the winner model is cached, restore its state and skip directly
        # to the method model. This avoids re-running stability selection,
        # ensemble training, calibration, and retrain when only the method
        # model code has changed.
        _wm_cached = _load_stage_cache(
            "winner_model", data_hash) if data_hash else None
        if _wm_cached is not None:
            self.stacking_clf = _wm_cached["stacking_clf"]
            self._global_selector = _wm_cached["_global_selector"]
            self._selected_decomposed_cols = _wm_cached["_selected_decomposed_cols"]
            self._d_indices = _wm_cached["_d_indices"]
            self._passthrough_indices = _wm_cached["_passthrough_indices"]
            self._blend_alpha = _wm_cached["_blend_alpha"]
            self._opt_threshold = _wm_cached["_opt_threshold"]
            self._temp_scale = _wm_cached.get("_temp_scale", None)
            self.scaler = _wm_cached["scaler"]
            _wm_feat_col_list = _wm_cached["feat_col_list"]
            _wm_df_tr_val_aug = _wm_cached["df_tr_val_aug"]
            _wm_df_test = _wm_cached["df_test"]
            _wm_X_tv_s = _wm_cached["X_tv_s"]
            _wm_X_test_s = _wm_cached["X_test_s"]
            print_metric("  Features used:", len(
                self._selected_decomposed_cols))
            print_metric("Stacking ensemble:", "restored from cache")

            # Recompute test arrays for metrics + method model
            _wm_X_test_sel = self._global_selector.transform(_wm_X_test_s)
            _wm_y_test = np.array(
                [1 if w == "Red" else 0 for w in _wm_df_test["winner"].values]
            )
            self._print_winner_test_metrics(
                _wm_X_test_sel, _wm_y_test, _wm_df_test)

            self.is_trained = True
            self._log("Training complete!")
            return  # early exit — winner model was cached

        # Stability Selection (Bodinier et al., 2023 / Meinshausen & Bühlmann, 2010):
        # repeated subsamples + base learner; features kept when selected often enough.
        _df_rfecv = df_tr_aug.copy()
        _df_rfecv_swap = self._corner_swap(_df_rfecv.copy())
        _feat_list_rfecv = [c for c in feat_col_list if c in _df_rfecv.columns]
        _X_rfecv_orig = _df_rfecv[_feat_list_rfecv].fillna(0)
        _X_rfecv_swap = _df_rfecv_swap[_feat_list_rfecv].fillna(0)
        _X_rfecv_raw = self._decompose_features(
            _X_rfecv_orig, _X_rfecv_swap).values
        _X_rfecv_s = self.scaler.transform(_X_rfecv_raw)
        _y_rfecv = np.array(
            [1 if w == "Red" else 0 for w in _df_rfecv["winner"].values]
        )

        _N_RUNS = 50
        _SUBSAMPLE_FRAC = 0.50
        _n_feats = _X_rfecv_s.shape[1]

        _SEL_THRESHOLD = 0.70

        # Build pool of model factories for multi-model stability selection.
        # Rotates across model types so feature importance reflects what the
        # actual ensemble (ET, HGB, XGB, CAT, LGB) finds useful, not just LGB.
        from sklearn.ensemble import ExtraTreesClassifier as _ET_Sel

        _model_pool = []
        _model_pool.append(("ET", lambda rs: _ET_Sel(
            n_estimators=100, max_features="sqrt",
            random_state=rs, n_jobs=SAFE_N_JOBS)))
        _model_pool.append(("RF", lambda rs: RandomForestClassifier(
            n_estimators=100, max_features="sqrt",
            random_state=rs, n_jobs=SAFE_N_JOBS)))
        if HAS_LGB:
            _model_pool.append(("LGB", lambda rs: lgb.LGBMClassifier(
                n_estimators=100, num_leaves=31, learning_rate=0.1,
                min_child_samples=10, random_state=rs,
                n_jobs=SAFE_N_JOBS, verbose=-1)))
        if HAS_XGB:
            _model_pool.append(("XGB", lambda rs: xgb.XGBClassifier(
                n_estimators=100, max_depth=6, learning_rate=0.1,
                random_state=rs, n_jobs=SAFE_N_JOBS, verbosity=0)))
        if HAS_CAT:
            _model_pool.append(("CAT", lambda rs: cb.CatBoostClassifier(
                iterations=100, depth=6, learning_rate=0.1,
                random_state=rs, verbose=0, allow_writing_files=False)))

        _n_models = len(_model_pool)
        _model_names = [m[0] for m in _model_pool]
        print_step(
            "Running Stability Selection feature selection "
            f"({_N_RUNS} runs, {int(_SUBSAMPLE_FRAC * 100)}% subsampling, "
            f"multi-model union [{'/'.join(_model_names)}], "
            f"threshold={int(_SEL_THRESHOLD * 100)}% per model)..."
        )
        _n_sub = max(10, int(_SUBSAMPLE_FRAC * len(_X_rfecv_s)))
        # Track selection counts per model type separately
        _per_model_counts = {name: np.zeros(_n_feats) for name, _ in _model_pool}
        _per_model_runs = {name: 0 for name, _ in _model_pool}
        _selection_counts = np.zeros(_n_feats)  # global counts for sel_freq
        t0_rfecv = time.time()
        for _run in range(_N_RUNS):
            _rng = np.random.RandomState(_run)
            _idx = _rng.choice(len(_X_rfecv_s), size=_n_sub, replace=False)
            _X_sub = _X_rfecv_s[_idx]
            _y_sub = _y_rfecv[_idx]
            _w_sub = w_tr[_idx]

            _model_name, _model_factory = _model_pool[_run % _n_models]
            _sel_est = _model_factory(_run)
            try:
                _sel_est.fit(_X_sub, _y_sub, sample_weight=_w_sub)
            except TypeError:
                _sel_est.fit(_X_sub, _y_sub)

            _imp = _sel_est.feature_importances_
            _selected = _imp >= _imp.mean()
            _selection_counts[_selected] += 1
            _per_model_counts[_model_name][_selected] += 1
            _per_model_runs[_model_name] += 1

            _pct = (_run + 1) / _N_RUNS
            _fill = int(_pct * 40)
            _bar = "\u2588" * _fill + "\u2591" * (40 - _fill)
            print(
                f"\r  [{_bar}] {int(_pct * 100):3d}%  (run {_run + 1}/{_N_RUNS})",
                end="",
                flush=True,
            )

        print()

        # Union strategy: a feature is kept if it passes the threshold
        # in ANY individual model type's runs.
        _support = np.zeros(_n_feats, dtype=bool)
        for _mname, _ in _model_pool:
            _m_runs = _per_model_runs[_mname]
            if _m_runs > 0:
                _m_freq = _per_model_counts[_mname] / _m_runs
                _m_selected = _m_freq >= _SEL_THRESHOLD
                _support |= _m_selected
                print_metric(f"    {_mname} features >= {int(_SEL_THRESHOLD*100)}%:",
                             int(_m_selected.sum()))

        # Overall selection frequency (across all runs) for reporting
        _sel_freq = _selection_counts / _N_RUNS

        # Ensure at least 10 features are always selected
        if _support.sum() < 10:
            _top10 = np.argsort(_sel_freq)[-10:]
            _support = np.zeros(_n_feats, dtype=bool)
            _support[_top10] = True

        print_metric(
            "  Stability Selection time:",
            f"{time.time() - t0_rfecv:.1f}s",
        )
        print_metric("  Features selected (union):", int(_support.sum()))
        print_metric("  Features removed:", int((~_support).sum()))

        # ── Correlation-based deduplication ─────────────────────────────
        # Within the selected features, drop near-duplicates (|corr| > 0.85).
        # For each correlated cluster, keep only the feature with the highest
        # stability-selection frequency to maximise information per feature.
        _dedup_threshold = 0.85
        _sel_indices = np.where(_support)[0]
        _X_sel_dedup = _X_rfecv_s[:, _sel_indices]
        _corr_matrix = np.corrcoef(_X_sel_dedup, rowvar=False)
        _corr_matrix = np.nan_to_num(_corr_matrix)
        _n_sel = len(_sel_indices)
        _to_drop_dedup = set()
        for _ci in range(_n_sel):
            if _ci in _to_drop_dedup:
                continue
            for _cj in range(_ci + 1, _n_sel):
                if _cj in _to_drop_dedup:
                    continue
                if abs(_corr_matrix[_ci, _cj]) > _dedup_threshold:
                    # Drop the feature with lower selection frequency
                    _fi = _sel_indices[_ci]
                    _fj = _sel_indices[_cj]
                    if _sel_freq[_fi] >= _sel_freq[_fj]:
                        _to_drop_dedup.add(_cj)
                    else:
                        _to_drop_dedup.add(_ci)
                        break  # _ci is dropped, move on
        _keep_dedup = sorted(set(range(_n_sel)) - _to_drop_dedup)
        _n_before_dedup = int(_support.sum())
        for _di in _to_drop_dedup:
            _support[_sel_indices[_di]] = False
        _n_after_dedup = int(_support.sum())
        print_metric("  Corr dedup removed (|r|>0.85):",
                     _n_before_dedup - _n_after_dedup)
        print_metric("  Features after dedup:", _n_after_dedup)

        rfecv = _StabilitySelector(_support, _sel_freq)

        self._global_selector = rfecv
        X_tr_sel = rfecv.transform(X_tr_s)
        X_val_sel = rfecv.transform(X_val_s)
        X_test_sel = rfecv.transform(X_test_s)
        self._selected_decomposed_cols = [
            c for c, keep in zip(self._decomposed_cols, rfecv.support_) if keep
        ]

        self._d_indices = [
            i
            for i, c in enumerate(self._selected_decomposed_cols)
            if not c.endswith("_inv")
        ]

        _importances = rfecv.estimator_.feature_importances_
        _feat_imp = sorted(
            zip(self._selected_decomposed_cols, _importances),
            key=lambda x: x[1],
            reverse=True,
        )
        _perfect_feats = [
            (_fname, _fimp) for _fname, _fimp in _feat_imp if _fimp >= 1.0
        ]
        if _perfect_feats:
            print_divider()
            print(
                f"  Features with 100% selection frequency ({len(_perfect_feats)}):")
            print_divider()
            for _fname, _fimp in _perfect_feats:
                print(f"    {_fname}")
            print_divider()

        print_divider()
        print(f"  Selected features with importances ({len(_feat_imp)}):")
        print_divider()
        _name_w = max((len(_fname) for _fname, _ in _feat_imp), default=0)
        for _rank, (_fname, _fimp) in enumerate(_feat_imp, 1):
            print(f"    {_rank:>3}. {_fname:<{_name_w}}  {_fimp:.4f}")
        print_divider()

        print_metric("  Features used:", X_tr_sel.shape[1])
        print()

        # Identify top D-component features for passthrough, diversified by signal family.
        # One feature per family ensures the meta-learner gets distinct information dimensions.
        _d_feat_imp = [
            (i, fname, fimp)
            for i, (fname, fimp) in enumerate(
                zip(self._selected_decomposed_cols, _importances)
            )
            if not fname.endswith("_inv")
        ]
        _d_feat_imp.sort(key=lambda x: x[2], reverse=True)

        _PT_FAMILIES = {
            "physical": ["reach", "height", "weight", "ape_index"],
            "rating": ["elo", "glicko"],
            "striking": [
                "SLpM",
                "SApM",
                "sig_str",
                "kd_rate",
                "head_acc",
                "damage_ratio",
                "damage_margin",
                "striking_quality",
                "combat_eff",
            ],
            "grappling": [
                "td_acc",
                "td_avg",
                "td_def",
                "sub_att",
                "ctrl",
                "grapple",
                "ground",
            ],
            "form": [
                "rolling",
                "streak",
                "momentum",
                "trajectory",
                "win_ratio",
                "da_win",
            ],
            "defense": ["str_def", "td_def", "defense", "head_defense", "absorbed"],
            "cardio": ["cardio", "endurance", "retention", "fatigue", "output_slope"],
            "finish": ["finish", "ko_win", "sub_win", "early_finish"],
            "style": ["style", "cluster", "entropy", "zone"],
            "experience": [
                "total_fights",
                "title",
                "five_round",
                "age",
                "peak_score",
                "prime",
            ],
        }

        def _classify_family(fname):
            for family, keywords in _PT_FAMILIES.items():
                if any(kw in fname for kw in keywords):
                    return family
            return "other"

        _pt_selected = []
        _pt_families_used = set()
        for idx, fname, fimp in _d_feat_imp:
            fam = _classify_family(fname)
            if fam not in _pt_families_used:
                _pt_selected.append((idx, fname, fimp, fam))
                _pt_families_used.add(fam)
            if len(_pt_selected) >= 10:
                break

        # If fewer than 10 families, fill remaining slots with next-best unused features
        if len(_pt_selected) < 10:
            _used_idx = {idx for idx, _, _, _ in _pt_selected}
            for idx, fname, fimp in _d_feat_imp:
                if idx not in _used_idx:
                    fam = _classify_family(fname)
                    _pt_selected.append((idx, fname, fimp, fam))
                    _used_idx.add(idx)
                if len(_pt_selected) >= 10:
                    break

        self._passthrough_indices = [idx for idx, _, _, _ in _pt_selected]
        print_metric("  Passthrough D-features:",
                     len(self._passthrough_indices))
        for _pi_idx, _pi_name, _pi_imp, _pi_fam in _pt_selected:
            print(
                f"    [{_pi_idx:3d}] {_pi_name}  (freq={_pi_imp:.2f}, {_pi_fam})")
        print()

        t0_cv = time.time()
        print_step(
            "Running grouped PurgedTimeSeriesSplit cross-validation (5 folds)..."
        )
        tscv = PurgedTimeSeriesSplit(n_splits=5, purge_days=30)
        fold_scores = []
        for fold, (tr_idx, val_idx) in enumerate(
            tscv.split(X_tr_sel, groups=_aug_groups), 1
        ):
            fold_rf = RandomForestClassifier(
                n_estimators=100, n_jobs=SAFE_N_JOBS, random_state=42
            )
            fold_rf.fit(X_tr_sel[tr_idx], y_tr[tr_idx],
                        sample_weight=w_tr[tr_idx])
            fold_pred = fold_rf.predict_proba(X_tr_sel[val_idx])[:, 1]
            fold_acc = (fold_rf.predict(
                X_tr_sel[val_idx]) == y_tr[val_idx]).mean()
            fold_ll = log_loss(y_tr[val_idx], fold_pred)
            fold_scores.append(fold_acc)
            print_metric(
                f"  Fold {fold} Accuracy:",
                f"{fold_acc:.4f}  |  Log-Loss: {fold_ll:.4f}",
            )
        _mean_cv = f"{np.mean(fold_scores):.4f} \u00b1 {np.std(fold_scores):.4f}"
        print(f"  {'  Mean CV Accuracy:':<42}{_mean_cv}")
        _cv_time = f"{time.time() - t0_cv:.1f}s"
        print(f"  {'  CV time:':<42}{_cv_time}")
        print()
        # --- OPTUNA BELOW ---
        # --- Optuna LGB tuning (warm start: enqueue_trial uses prior best params) ---
        # self._optuna_best_lgb_params = {
        #     "n_estimators": 620,
        #     "num_leaves": 61,
        #     "max_depth": 7,
        #     "learning_rate": 0.011567185795357808,
        #     "min_child_samples": 35,
        #     "reg_alpha": 1.2770572166659555,
        #     "reg_lambda": 1.7644097708281317,
        #     "colsample_bytree": 0.5388530793226052,
        #     "subsample": 0.9615358361326504,
        # }
        if HAS_OPTUNA and HAS_LGB:
            LGB_TRIALS = 30
            print_step(
                f"Running Optuna hyperparameter search for LightGBM ({LGB_TRIALS} trials)..."
            )
            lgb_optuna_start = time.time()

            def lgb_optuna_objective(trial):
                params = {
                    "n_estimators": trial.suggest_int("n_estimators", 200, 1000),
                    "num_leaves": trial.suggest_int("num_leaves", 15, 127),
                    "max_depth": trial.suggest_int("max_depth", 3, 8),
                    "learning_rate": trial.suggest_float(
                        "learning_rate", 0.01, 0.1, log=True
                    ),
                    "min_child_samples": trial.suggest_int("min_child_samples", 15, 50),
                    "reg_alpha": trial.suggest_float("reg_alpha", 0.0, 5.0),
                    "reg_lambda": trial.suggest_float("reg_lambda", 0.0, 5.0),
                    "colsample_bytree": trial.suggest_float(
                        "colsample_bytree", 0.5, 1.0
                    ),
                    "subsample": trial.suggest_float("subsample", 0.6, 1.0),
                }
                lgb_trial = lgb.LGBMClassifier(
                    **params,
                    random_state=42,
                    verbose=-1,
                    n_jobs=SAFE_N_JOBS,
                    class_weight="balanced",
                )
                tscv3 = PurgedTimeSeriesSplit(n_splits=3, purge_days=30)
                scores = []
                for tr_i, val_i in tscv3.split(X_tr_sel, groups=_aug_groups):
                    lgb_trial.fit(
                        X_tr_sel[tr_i],
                        y_tr[tr_i],
                        sample_weight=w_tr[tr_i],
                        eval_set=[(X_tr_sel[val_i], y_tr[val_i])],
                        callbacks=[
                            lgb.early_stopping(20, verbose=False),
                            lgb.log_evaluation(-1),
                        ],
                    )
                    p = lgb_trial.predict_proba(X_tr_sel[val_i])[:, 1]
                    scores.append(-log_loss(y_tr[val_i], p))
                return np.mean(scores)

            lgb_study = optuna.create_study(
                direction="maximize",
                sampler=optuna.samplers.TPESampler(seed=RANDOM_SEED),
            )
            lgb_study.enqueue_trial(
                {
                    "n_estimators": 620,
                    "num_leaves": 61,
                    "max_depth": 7,
                    "learning_rate": 0.011567185795357808,
                    "min_child_samples": 35,
                    "reg_alpha": 1.2770572166659555,
                    "reg_lambda": 1.7644097708281317,
                    "colsample_bytree": 0.5388530793226052,
                    "subsample": 0.9615358361326504,
                }
            )
            _lgb_trials_done = [0]

            def _lgb_callback(study, trial):
                _lgb_trials_done[0] += 1
                best_val = (
                    -study.best_value if study.best_value is not None else float("nan")
                )
                bar_fill = int(_lgb_trials_done[0] / LGB_TRIALS * 30)
                bar = "█" * bar_fill + "░" * (30 - bar_fill)
                pct = int(_lgb_trials_done[0] / LGB_TRIALS * 100)
                print(
                    f"\r  [{bar}] {pct:3d}%  trial {_lgb_trials_done[0]:2d}/{LGB_TRIALS}"
                    f"  best log-loss: {best_val:.4f}",
                    end="",
                    flush=True,
                )

            lgb_study.optimize(
                lgb_optuna_objective,
                n_trials=LGB_TRIALS,
                show_progress_bar=False,
                callbacks=[_lgb_callback],
            )
            print()
            lgb_best_params = lgb_study.best_params
            lgb_elapsed = time.time() - lgb_optuna_start
            print_metric("LGB Optuna best log-loss:",
                         f"{-lgb_study.best_value:.4f}")
            print_metric("Time elapsed:", f"{lgb_elapsed:.1f}s")
            print_step("LGB best params:")
            for _k, _v in lgb_best_params.items():
                print(f"    {_k}: {_v}")
            self._optuna_best_lgb_params = lgb_best_params
        else:
            self._optuna_best_lgb_params = {}

        # --- Optuna XGB tuning (warm start: enqueue_trial uses prior best params) ---
        # self._optuna_best_xgb_params = {
        #     "n_estimators": 350,
        #     "max_depth": 6,
        #     "learning_rate": 0.01698375168217574,
        #     "min_child_weight": 3,
        #     "reg_alpha": 0.049757303739185965,
        #     "reg_lambda": 1.192127507782831,
        #     "colsample_bytree": 0.5025354479197125,
        #     "subsample": 0.9259308962023541,
        #     "gamma": 0.6564910729998177,
        # }
        if HAS_OPTUNA and HAS_XGB:
            XGB_TRIALS = 30
            print_step(
                f"Running Optuna hyperparameter search for XGBoost ({XGB_TRIALS} trials)..."
            )
            xgb_optuna_start = time.time()

            def xgb_optuna_objective(trial):
                params = {
                    "n_estimators": trial.suggest_int("n_estimators", 200, 1000),
                    "max_depth": trial.suggest_int("max_depth", 3, 8),
                    "learning_rate": trial.suggest_float(
                        "learning_rate", 0.01, 0.1, log=True
                    ),
                    "min_child_weight": trial.suggest_int("min_child_weight", 3, 30),
                    "reg_alpha": trial.suggest_float("reg_alpha", 0.0, 5.0),
                    "reg_lambda": trial.suggest_float("reg_lambda", 0.0, 5.0),
                    "colsample_bytree": trial.suggest_float(
                        "colsample_bytree", 0.5, 1.0
                    ),
                    "subsample": trial.suggest_float("subsample", 0.6, 1.0),
                    "gamma": trial.suggest_float("gamma", 0.0, 5.0),
                }
                xgb_trial_params = {
                    **params,
                    "random_state": RANDOM_SEED,
                    "verbosity": 0,
                    "eval_metric": "logloss",
                    "n_jobs": SAFE_N_JOBS,
                    "scale_pos_weight": 1.0,
                    "early_stopping_rounds": 20,
                }
                if self.gpu_info.get("xgb"):
                    xgb_trial_params["device"] = "cuda"
                else:
                    xgb_trial_params["tree_method"] = "hist"
                xgb_trial = xgb.XGBClassifier(**xgb_trial_params)
                tscv3 = PurgedTimeSeriesSplit(n_splits=3, purge_days=30)
                scores = []
                for tr_i, val_i in tscv3.split(X_tr_sel, groups=_aug_groups):
                    xgb_trial.fit(
                        X_tr_sel[tr_i],
                        y_tr[tr_i],
                        sample_weight=w_tr[tr_i],
                        eval_set=[(X_tr_sel[val_i], y_tr[val_i])],
                        verbose=False,
                    )
                    p = xgb_trial.predict_proba(X_tr_sel[val_i])[:, 1]
                    scores.append(-log_loss(y_tr[val_i], p))
                return np.mean(scores)

            xgb_study = optuna.create_study(
                direction="maximize",
                sampler=optuna.samplers.TPESampler(seed=RANDOM_SEED),
            )
            xgb_study.enqueue_trial(
                {
                    "n_estimators": 350,
                    "max_depth": 6,
                    "learning_rate": 0.01698375168217574,
                    "min_child_weight": 3,
                    "reg_alpha": 0.049757303739185965,
                    "reg_lambda": 1.192127507782831,
                    "colsample_bytree": 0.5025354479197125,
                    "subsample": 0.9259308962023541,
                    "gamma": 0.6564910729998177,
                }
            )
            _xgb_trials_done = [0]

            def _xgb_callback(study, trial):
                _xgb_trials_done[0] += 1
                best_val = (
                    -study.best_value if study.best_value is not None else float("nan")
                )
                bar_fill = int(_xgb_trials_done[0] / XGB_TRIALS * 30)
                bar = "█" * bar_fill + "░" * (30 - bar_fill)
                pct = int(_xgb_trials_done[0] / XGB_TRIALS * 100)
                print(
                    f"\r  [{bar}] {pct:3d}%  trial {_xgb_trials_done[0]:2d}/{XGB_TRIALS}"
                    f"  best log-loss: {best_val:.4f}",
                    end="",
                    flush=True,
                )

            xgb_study.optimize(
                xgb_optuna_objective,
                n_trials=XGB_TRIALS,
                show_progress_bar=False,
                callbacks=[_xgb_callback],
            )
            print()
            xgb_elapsed = time.time() - xgb_optuna_start
            print_metric("XGB Optuna best log-loss:",
                         f"{-xgb_study.best_value:.4f}")
            print_metric("Time elapsed:", f"{xgb_elapsed:.1f}s")
            print_step("XGB best params:")
            for _k, _v in xgb_study.best_params.items():
                print(f"    {_k}: {_v}")
            self._optuna_best_xgb_params = xgb_study.best_params
        else:
            self._optuna_best_xgb_params = {}

        # --- Optuna CatBoost tuning (warm start: enqueue_trial uses prior best params) ---
        # self._optuna_best_cat_params = {
        #     "iterations": 495,
        #     "depth": 7,
        #     "learning_rate": 0.010423998224768954,
        #     "l2_leaf_reg": 5.07448427186833,
        #     "bagging_temperature": 0.7529860264069669,
        #     "random_strength": 0.6169855140521886,
        # }
        if HAS_OPTUNA and HAS_CAT:
            CAT_TRIALS = 30
            print_step(
                f"Running Optuna hyperparameter search for CatBoost ({CAT_TRIALS} trials)..."
            )
            cat_optuna_start = time.time()

            def cat_optuna_objective(trial):
                params = {
                    "iterations": trial.suggest_int("iterations", 200, 600),
                    "depth": trial.suggest_int("depth", 4, 8),
                    "learning_rate": trial.suggest_float(
                        "learning_rate", 0.01, 0.1, log=True
                    ),
                    "l2_leaf_reg": trial.suggest_float("l2_leaf_reg", 1.0, 15.0),
                    "bagging_temperature": trial.suggest_float(
                        "bagging_temperature", 0.0, 1.0
                    ),
                    "random_strength": trial.suggest_float("random_strength", 0.5, 3.0),
                }
                if self.gpu_info.get("cat"):
                    params["task_type"] = "GPU"
                    params["border_count"] = 128
                cat_trial = cb.CatBoostClassifier(
                    **params,
                    random_seed=RANDOM_SEED,
                    verbose=0,
                    eval_metric="Logloss",
                    auto_class_weights="Balanced",
                    allow_writing_files=False,
                    early_stopping_rounds=20,
                )
                tscv3 = PurgedTimeSeriesSplit(n_splits=3, purge_days=30)
                scores = []
                for tr_i, val_i in tscv3.split(X_tr_sel, groups=_aug_groups):
                    cat_trial.fit(
                        X_tr_sel[tr_i],
                        y_tr[tr_i],
                        sample_weight=w_tr[tr_i],
                        eval_set=(X_tr_sel[val_i], y_tr[val_i]),
                    )
                    p = cat_trial.predict_proba(X_tr_sel[val_i])[:, 1]
                    scores.append(-log_loss(y_tr[val_i], p))
                return np.mean(scores)

            cat_study = optuna.create_study(
                direction="maximize",
                sampler=optuna.samplers.TPESampler(seed=RANDOM_SEED),
            )
            cat_study.enqueue_trial(
                {
                    "iterations": 495,
                    "depth": 7,
                    "learning_rate": 0.010423998224768954,
                    "l2_leaf_reg": 5.07448427186833,
                    "bagging_temperature": 0.7529860264069669,
                    "random_strength": 0.6169855140521886,
                }
            )
            _cat_trials_done = [0]

            def _cat_callback(study, trial):
                _cat_trials_done[0] += 1
                best_val = (
                    -study.best_value if study.best_value is not None else float("nan")
                )
                bar_fill = int(_cat_trials_done[0] / CAT_TRIALS * 30)
                bar = "█" * bar_fill + "░" * (30 - bar_fill)
                pct = int(_cat_trials_done[0] / CAT_TRIALS * 100)
                print(
                    f"\r  [{bar}] {pct:3d}%  trial {_cat_trials_done[0]:2d}/{CAT_TRIALS}"
                    f"  best log-loss: {best_val:.4f}",
                    end="",
                    flush=True,
                )

            cat_study.optimize(
                cat_optuna_objective,
                n_trials=CAT_TRIALS,
                show_progress_bar=False,
                callbacks=[_cat_callback],
            )
            print()
            cat_elapsed = time.time() - cat_optuna_start
            print_metric("CAT Optuna best log-loss:",
                         f"{-cat_study.best_value:.4f}")
            print_metric("Time elapsed:", f"{cat_elapsed:.1f}s")
            print_step("CAT best params:")
            for _k, _v in cat_study.best_params.items():
                print(f"    {_k}: {_v}")
            self._optuna_best_cat_params = cat_study.best_params
        else:
            self._optuna_best_cat_params = {}

        t0_estimators = time.time()
        estimators = self._build_estimators(
            X_tr_sel,
            y_tr,
            sample_weight=w_tr,
            X_val=X_val_sel,
            y_val=y_val,
        )
        print()
        print_metric("  Base estimators time:",
                     f"{time.time() - t0_estimators:.1f}s")
        print()

        print_step(
            "Building stacking ensemble (manual PurgedTimeSeriesSplit-5 OOF meta-learning)..."
        )
        t0_stack = time.time()
        _stk = _ManualStackingEnsemble(
            estimators=estimators,
            meta_C=0.20,
            n_splits=5,
            random_state=RANDOM_SEED,
            passthrough_indices=getattr(self, "_passthrough_indices", None),
            verbose=False,
        )

        _stk.fit(X_tr_sel, y_tr, groups=_aug_groups, sample_weight=w_tr)

        self._base_ensemble = _stk
        self.stacking_clf = _stk
        print()
        print_metric("  Stacking ensemble time:",
                     f"{time.time() - t0_stack:.1f}s")

        print_section("POST-TRAINING CALIBRATION")
        self._corner_bias_diagnostic(X_val_sel, y_val)
        print()
        print_step(
            "Calibrating blend weight and decision threshold on validation set..."
        )
        try:
            val_proba = self.stacking_clf.predict_proba(X_val_sel)
            _classes_v = list(self.stacking_clf.classes_)
            _r_idx_v = _classes_v.index(1) if 1 in _classes_v else 1

            _d_idx_v = getattr(self, "_d_indices", [])
            if len(_d_idx_v) > 0:
                X_val_swap_sel = X_val_sel.copy()
                X_val_swap_sel[:, _d_idx_v] *= -1
                val_proba_swap = self.stacking_clf.predict_proba(
                    X_val_swap_sel)

                # Sweep alpha to find optimal blend weight (minimize log-loss)
                best_alpha, best_alpha_ll = 0.5, float("inf")
                for alpha in np.arange(0.30, 1.005, 0.01):
                    _vpp = alpha * val_proba[:, _r_idx_v] + (1 - alpha) * (
                        1.0 - val_proba_swap[:, _r_idx_v]
                    )
                    _vpp = np.clip(_vpp, 1e-7, 1 - 1e-7)
                    _ll = log_loss(y_val, np.column_stack([1.0 - _vpp, _vpp]))
                    if _ll < best_alpha_ll:
                        best_alpha_ll = _ll
                        best_alpha = alpha
                self._blend_alpha = float(best_alpha)

                val_proba_pos = self._blend_alpha * val_proba[:, _r_idx_v] + (
                    1 - self._blend_alpha
                ) * (1.0 - val_proba_swap[:, _r_idx_v])
            else:
                self._blend_alpha = 0.5
                val_proba_pos = val_proba[:, _r_idx_v]

            # ── Temperature scaling (post-ensemble calibration) ─────
            # Single-parameter calibration: divides logits by temperature T
            # before sigmoid.  T < 1 sharpens (stretches probabilities away
            # from 0.5), T > 1 softens.  Corrects uniform under/over-
            # confidence without overfitting sparse tail bins.
            from scipy.optimize import minimize_scalar

            _logits = np.log(np.clip(val_proba_pos, 1e-9, 1 - 1e-9) /
                             np.clip(1 - val_proba_pos, 1e-9, 1 - 1e-9))
            _y_f = y_val.astype(float)

            def _temp_nll(T):
                _scaled = _logits / T
                _p = 1.0 / (1.0 + np.exp(-_scaled))
                _p = np.clip(_p, 1e-9, 1 - 1e-9)
                return -np.mean(_y_f * np.log(_p) +
                                (1 - _y_f) * np.log(1 - _p))

            _res = minimize_scalar(_temp_nll, bounds=(0.1, 5.0),
                                   method="bounded")
            self._temp_scale = float(_res.x)
            # Apply to val probs for threshold tuning
            _scaled_logits = _logits / self._temp_scale
            val_proba_pos = 1.0 / (1.0 + np.exp(-_scaled_logits))

            # Sweep threshold to find optimal decision boundary (maximize accuracy)
            best_thr, best_thr_acc = 0.5, 0.0
            for thr in np.arange(0.40, 0.605, 0.005):
                _acc = accuracy_score(
                    y_val, (val_proba_pos >= thr).astype(int))
                if _acc > best_thr_acc:
                    best_thr_acc = _acc
                    best_thr = thr
            self._opt_threshold = float(best_thr)
        except Exception:
            self._blend_alpha = 0.5
            self._opt_threshold = 0.5
            self._temp_scale = None

        # ── Validation set evaluation ──────────────────────────────────────────
        print_section("WINNER MODEL — VALIDATION SET RESULTS")
        try:
            _thr_v = getattr(self, "_opt_threshold", 0.5)
            _alpha_v = getattr(self, "_blend_alpha", 0.5)

            # Recompute val_proba_pos in case the calibration block was skipped
            _val_proba = self.stacking_clf.predict_proba(X_val_sel)
            _classes_val = list(self.stacking_clf.classes_)
            _r_idx_val = _classes_val.index(1) if 1 in _classes_val else 1
            _d_idx_val = getattr(self, "_d_indices", [])
            if len(_d_idx_val) > 0:
                _X_val_swap = X_val_sel.copy()
                _X_val_swap[:, _d_idx_val] *= -1
                _val_proba_swap = self.stacking_clf.predict_proba(_X_val_swap)
                _val_proba_pos = _alpha_v * _val_proba[:, _r_idx_val] + (
                    1 - _alpha_v
                ) * (1.0 - _val_proba_swap[:, _r_idx_val])
            else:
                _val_proba_pos = _val_proba[:, _r_idx_val]

            # Apply temperature scaling if fitted
            _T = getattr(self, "_temp_scale", None)
            if _T is not None:
                _logits_v = np.log(np.clip(_val_proba_pos, 1e-9, 1 - 1e-9) /
                                   np.clip(1 - _val_proba_pos, 1e-9, 1 - 1e-9))
                _val_proba_pos = 1.0 / (1.0 + np.exp(-_logits_v / _T))

            _val_pred = (_val_proba_pos >= _thr_v).astype(int)
            _val_acc = accuracy_score(y_val, _val_pred)
            _val_ll = log_loss(
                y_val, np.column_stack([1.0 - _val_proba_pos, _val_proba_pos])
            )
            _val_prec = precision_score(y_val, _val_pred, zero_division=0)
            _val_rec = recall_score(y_val, _val_pred, zero_division=0)
            _val_f1 = f1_score(y_val, _val_pred, zero_division=0)
            _val_brier = brier_score_loss(y_val, _val_proba_pos)
            _val_auc = roc_auc_score(y_val, _val_proba_pos)
            _val_majority = max(np.mean(y_val == 1), np.mean(y_val == 0))

            print_metric("Val Accuracy:", f"{_val_acc:.4f}")
            print_metric("Majority-class baseline:", f"{_val_majority:.4f}")
            print_metric("Lift over baseline:",
                         f"{_val_acc - _val_majority:+.4f}")
            print_metric("Val samples:", len(y_val))
            print_divider()
            print_metric("Precision:", f"{_val_prec:.4f}")
            print_metric("Recall:", f"{_val_rec:.4f}")
            print_metric("F1 Score:", f"{_val_f1:.4f}")
            print_divider()
            print_metric("Log-Loss:", f"{_val_ll:.4f}")
            print_metric("Brier Score:", f"{_val_brier:.4f}")
            print_metric("ROC AUC:", f"{_val_auc:.4f}")
            print_divider()
            print_metric("Blend alpha (val-tuned):", f"{_alpha_v:.3f}")
            print_metric("Decision threshold (val-tuned):", f"{_thr_v:.3f}")
        except Exception as _e:
            print(f"  (val metrics unavailable: {_e})")

        # ── Retrain full stacking ensemble on train + val ──────────────────────
        print_section("RETRAINING ON TRAIN + VAL")
        print_step("Folding validation set into training data...")

        df_tr_val = df_train.iloc[:val_end]
        df_tr_val_aug = pd.concat(
            [df_tr_val, self._corner_swap(df_tr_val)], ignore_index=True
        )
        df_tv_feat = df_tr_val_aug.copy()
        df_tv_swap = self._corner_swap(df_tv_feat)
        X_tv_orig = df_tv_feat[feat_col_list].fillna(0)
        X_tv_swap = df_tv_swap[feat_col_list].fillna(0)
        X_tv_decomposed = self._decompose_features(X_tv_orig, X_tv_swap)
        X_tv_raw = X_tv_decomposed.values
        y_tv = np.array(
            [1 if w == "Red" else 0 for w in df_tv_feat["winner"].values])

        _dates_tv = pd.to_datetime(df_tv_feat["event_date"], errors="coerce")
        _most_recent_tv = _dates_tv.max()
        _days_ago_tv = (
            (_most_recent_tv - _dates_tv)
            .dt.days.fillna(0)
            .clip(lower=0)
            .values.astype(float)
        )
        w_tv = np.exp(-np.log(2) / _HALF_LIFE_DAYS * _days_ago_tv)
        w_tv = np.maximum(w_tv, _best_floor)
        w_tv = w_tv / w_tv.mean()

        X_tv_s = self.scaler.fit_transform(X_tv_raw)
        X_test_s = self.scaler.transform(X_test_raw)
        X_tv_sel = self._global_selector.transform(X_tv_s)
        X_test_sel = self._global_selector.transform(X_test_s)

        _n_orig_tv = len(df_tr_val)
        _aug_groups_tv = np.tile(np.arange(_n_orig_tv), 2)[: len(y_tv)]

        print_metric("Train+Val samples (augmented):", len(X_tv_sel))

        # Collect early-stop iterations from original training to use as fixed budgets
        _es_iters = {}
        for _name, _est in self.stacking_clf.estimators:
            if _name == "xgb" and hasattr(_est, "best_iteration"):
                _es_iters["xgb"] = _est.best_iteration + 1
            elif _name == "lgb" and hasattr(_est, "best_iteration_"):
                _es_iters["lgb"] = _est.best_iteration_
            elif _name == "cat" and hasattr(_est, "best_iteration_"):
                _es_iters["cat"] = _est.best_iteration_

        # Step 1: Rebuild base estimators with fixed iteration counts and NO early stopping
        _tv_estimators = []
        for name, est in self.stacking_clf.estimators:
            fixed_iters = _es_iters.get(name)
            new_est = _clean_estimator_for_retrain(
                name, est, fixed_iters=fixed_iters)
            _tv_estimators.append((name, new_est))

        # Step 2: Build new stacking ensemble — generates OOF on train+val,
        #         re-searches meta-learner C, refits meta-learner on enlarged OOF
        t0_retrain = time.time()
        _stk_tv = _ManualStackingEnsemble(
            estimators=_tv_estimators,
            meta_C=0.05,  # not used by simplex combiner
            n_splits=5,
            random_state=RANDOM_SEED,
            passthrough_indices=getattr(self, "_passthrough_indices", None),
        )
        _stk_tv.fit(X_tv_sel, y_tv, groups=_aug_groups_tv, sample_weight=w_tv)
        print_step(
            "OOF meta-features rebuilt and meta-learner refit on train+val")

        # Step 3: Refit each base estimator on full train+val (no OOF holdout)
        retrained = []
        for name, est in _stk_tv.estimators:
            fixed_iters = _es_iters.get(name)
            new_est = _clean_estimator_for_retrain(
                name, est, fixed_iters=fixed_iters)
            fit_kw = {}
            X_fit_tv, y_fit_tv = X_tv_sel, y_tv
            if name in ("mlp", "mlp2"):
                _rng_sw = np.random.RandomState(RANDOM_SEED)
                _probs = w_tv / w_tv.sum()
                _resample_idx = _rng_sw.choice(
                    len(X_tv_sel), size=len(X_tv_sel), replace=True, p=_probs
                )
                X_fit_tv = X_tv_sel[_resample_idx]
                y_fit_tv = y_tv[_resample_idx]
            elif name == "xgb":
                fit_kw = {"sample_weight": w_tv, "verbose": False}
            elif name == "lgb":
                fit_kw = {"sample_weight": w_tv}
            elif name == "cat":
                fit_kw = {"sample_weight": w_tv}
            elif name in ("rf", "et", "ridge"):
                fit_kw = {"sample_weight": w_tv}
            try:
                new_est.fit(X_fit_tv, y_fit_tv, **fit_kw)
            except TypeError:
                new_est.fit(X_fit_tv, y_fit_tv)
            retrained.append((name, new_est))
            print(f"    {name} retrained")

        _stk_tv.estimators = retrained
        self.stacking_clf = _stk_tv
        print_metric("Retrain time:", f"{time.time() - t0_retrain:.1f}s")
        print()

        # ── Refit temperature scaling on retrained model's OOF probs ──
        # The temperature fitted earlier used the pre-retrain model's val probs.
        # Now the ensemble has changed (retrained on train+val), so refit
        # using OOF predictions from the new stacking ensemble.
        _oof_pos = getattr(_stk_tv, "_oof_pos_proba", None)
        if _oof_pos is not None:
            try:
                from scipy.optimize import minimize_scalar as _ms_refit
                _oof_clipped = np.clip(_oof_pos, 1e-9, 1 - 1e-9)
                _oof_logits = np.log(_oof_clipped / (1 - _oof_clipped))
                _y_tv_f = y_tv.astype(float)

                def _temp_nll_refit(T):
                    _s = _oof_logits / T
                    _p = np.clip(1.0 / (1.0 + np.exp(-_s)), 1e-9, 1 - 1e-9)
                    return -np.mean(_y_tv_f * np.log(_p) +
                                    (1 - _y_tv_f) * np.log(1 - _p))

                _res_refit = _ms_refit(_temp_nll_refit, bounds=(0.1, 5.0),
                                       method="bounded")
                self._temp_scale = float(_res_refit.x)
                print_step(f"Temperature scaling refit on retrained OOF "
                           f"(T={self._temp_scale:.4f})")

                # Re-tune decision threshold on retrained OOF probs
                _T_refit = self._temp_scale
                _cal_oof = 1.0 / (1.0 + np.exp(-_oof_logits / _T_refit))
                _best_thr_r, _best_acc_r = 0.5, 0.0
                for _thr_r in np.arange(0.40, 0.605, 0.005):
                    _acc_r = accuracy_score(
                        y_tv, (_cal_oof >= _thr_r).astype(int))
                    if _acc_r > _best_acc_r:
                        _best_acc_r = _acc_r
                        _best_thr_r = _thr_r
                self._opt_threshold = float(_best_thr_r)
                print_step(f"Decision threshold refit on retrained OOF "
                           f"(thr={self._opt_threshold:.3f})")
            except Exception as _e:
                print(f"  (temperature/threshold refit failed: {_e})")

        self._print_winner_test_metrics(X_test_sel, y_test, df_test)

        # ── Save Stage 3: winner_model ─────────────────────────────────
        if data_hash:
            _save_stage_cache(
                "winner_model",
                data_hash,
                {
                    "stacking_clf": self.stacking_clf,
                    "_global_selector": self._global_selector,
                    "_selected_decomposed_cols": self._selected_decomposed_cols,
                    "_d_indices": self._d_indices,
                    "_passthrough_indices": self._passthrough_indices,
                    "_blend_alpha": getattr(self, "_blend_alpha", 0.5),
                    "_opt_threshold": getattr(self, "_opt_threshold", 0.5),
                    "_temp_scale": getattr(self, "_temp_scale", None),
                    "scaler": self.scaler,
                    "feat_col_list": feat_col_list,
                    "df_tr_val_aug": df_tr_val_aug,
                    "df_test": df_test,
                    "X_tv_s": X_tv_s,
                    "X_test_s": X_test_s,
                },
            )

        self.is_trained = True
        self._log("Training complete!")

    # Measures how much the model is biased toward the red corner by checking:
    #   parity error: |P(red wins | X) + P(red wins | X_flipped)| should equal 1.0
    #   red bias: mean(P(red wins)) - 0.5 (0% = perfectly unbiased)
    # A 4-8% red bias is expected (real UFC red-corner win rate ~54-58%).
    # Over 8% suggests the model has over-learned corner assignment as a signal.
    def _print_winner_test_metrics(self, X_test_sel, y_test, df_test):
        """Print full holdout test set evaluation for the winner model."""
        print_section("WINNER MODEL — HOLDOUT TEST SET RESULTS")
        try:
            test_proba = self.stacking_clf.predict_proba(X_test_sel)
            classes_t = list(self.stacking_clf.classes_)
            r_idx_t = classes_t.index(1) if 1 in classes_t else 1

            d_idx = getattr(self, "_d_indices", [])
            _alpha = getattr(self, "_blend_alpha", 0.5)
            if len(d_idx) > 0:
                X_test_swap_sel = X_test_sel.copy()
                X_test_swap_sel[:, d_idx] *= -1
                test_proba_swap = self.stacking_clf.predict_proba(
                    X_test_swap_sel)
                test_proba_pos = _alpha * test_proba[:, r_idx_t] + (1 - _alpha) * (
                    1.0 - test_proba_swap[:, r_idx_t]
                )
            else:
                test_proba_pos = test_proba[:, r_idx_t]

            # Apply temperature scaling if fitted
            _T = getattr(self, "_temp_scale", None)
            if _T is not None:
                _logits_t = np.log(np.clip(test_proba_pos, 1e-9, 1 - 1e-9) /
                                   np.clip(1 - test_proba_pos, 1e-9, 1 - 1e-9))
                test_proba_pos = 1.0 / (1.0 + np.exp(-_logits_t / _T))

            _thr = getattr(self, "_opt_threshold", 0.5)
            test_pred = (test_proba_pos >= _thr).astype(int)
            test_acc = accuracy_score(y_test, test_pred)
            test_ll = log_loss(
                y_test, np.column_stack([1.0 - test_proba_pos, test_proba_pos])
            )
            test_prec = precision_score(y_test, test_pred, zero_division=0)
            test_rec = recall_score(y_test, test_pred, zero_division=0)
            test_f1 = f1_score(y_test, test_pred, zero_division=0)
            test_brier = brier_score_loss(y_test, test_proba_pos)
            majority_acc_t = max(np.mean(y_test == 1), np.mean(y_test == 0))

            # Primary metrics
            print_metric("Test Accuracy:", f"{test_acc:.4f}")
            print_metric("Majority-class baseline:", f"{majority_acc_t:.4f}")
            print_metric("Lift over baseline:",
                         f"{test_acc - majority_acc_t:+.4f}")
            print_metric("Test samples:", len(y_test))
            print_divider()

            # Classification metrics
            print_metric("Precision:", f"{test_prec:.4f}")
            print_metric("Recall:", f"{test_rec:.4f}")
            print_metric("F1 Score:", f"{test_f1:.4f}")
            print_divider()

            # Probabilistic metrics
            test_auc = roc_auc_score(y_test, test_proba_pos)
            print_metric("Log-Loss:", f"{test_ll:.4f}")
            print_metric("Brier Score:", f"{test_brier:.4f}")
            print_metric("ROC AUC:", f"{test_auc:.4f}")
            print_divider()

            # Calibration parameters
            print_metric("Blend alpha (val-tuned):", f"{_alpha:.3f}")
            print_metric("Decision threshold (val-tuned):", f"{_thr:.3f}")
            print_divider()

            # Confusion matrix
            _cm = confusion_matrix(y_test, test_pred)
            _cm_labels = ["Blue", "Red"]
            _w = 8
            print()
            print(f"    {'':>{_w}s}" +
                  "".join(f"  {lb:>{_w}s}" for lb in _cm_labels))
            print(
                f"    {'':>{_w}s}" +
                "".join(f"  {'-' * _w:>{_w}s}" for _ in _cm_labels)
            )
            for _ri, _rl in enumerate(_cm_labels):
                _vals = "".join(
                    f"  {_cm[_ri, j]:>{_w}d}" for j in range(len(_cm_labels))
                )
                print(f"    {_rl:>{_w}s}{_vals}")

            # ── Calibration curve ──────────────────────────────────────────────
            print()
            print_step("Calibration curve (10 bins):")
            _n_bins = 10
            _bin_edges = np.linspace(0.0, 1.0, _n_bins + 1)
            _BAR_WIDTH = 30
            print(
                f"    {'Bin':<14s} {'Mean Pred':>9s}  {'Actual':>6s}  {'Count':>6s}  {'Predicted vs Actual'}"
            )
            for _bi in range(_n_bins):
                _lo, _hi = _bin_edges[_bi], _bin_edges[_bi + 1]
                _mask = (test_proba_pos >= _lo) & (test_proba_pos < _hi)
                if _bi == _n_bins - 1:
                    _mask = (test_proba_pos >= _lo) & (test_proba_pos <= _hi)
                _cnt = int(_mask.sum())
                if _cnt == 0:
                    continue
                _mean_pred = float(test_proba_pos[_mask].mean())
                _actual = float(y_test[_mask].mean())
                _n_pred = int(_mean_pred * _BAR_WIDTH)
                _n_act = int(_actual * _BAR_WIDTH)
                _n_max = max(_n_pred, _n_act)
                _bar = ""
                for _ci in range(_n_max):
                    if _ci < _n_pred and _ci < _n_act:
                        _bar += "\u2588"
                    elif _ci < _n_pred:
                        _bar += "\u2593"
                    else:
                        _bar += "\u2591"
                _diff = (_mean_pred - _actual) * 100
                _arrow = f" ({_diff:+.1f}%)" if abs(_diff) > 0.5 else ""
                _bin_lbl = f"[{_lo * 100:.0f}%\u2013{_hi * 100:.0f}%)"
                print(
                    f"    {_bin_lbl:<14s} {_mean_pred * 100:8.1f}%  {_actual * 100:6.1f}%  {_cnt:6d}  {_bar}{_arrow}"
                )
            print(
                f"    {'':14s} \u2588 = overlap  \u2593 = pred > actual  \u2591 = actual > pred"
            )

            # ── Performance by weight class ────────────────────────────────────
            print()
            print_step("Accuracy by weight class:")
            _wc_test = df_test["weight_class"].values
            _wc_unique = sorted(set(_wc_test))
            print(f"    {'Weight Class':<28s}  {'Acc':>6s}  {'N':>5s}")
            print(f"    {'\u2500' * 28}  {'\u2500' * 6}  {'\u2500' * 5}")
            for _wc in _wc_unique:
                _wc_mask = np.array([w == _wc for w in _wc_test])
                _wc_n = int(_wc_mask.sum())
                if _wc_n == 0:
                    continue
                _wc_acc = float(
                    (test_pred[_wc_mask] == y_test[_wc_mask]).mean())
                print(f"    {_wc:<28s}  {_wc_acc:6.1%}  {_wc_n:5d}")

            # ── Performance by gender ──────────────────────────────────────────
            if "gender" in df_test.columns:
                print()
                print_step("Accuracy by gender:")
                _gen_test = df_test["gender"].fillna("Unknown").values
                _gen_unique = sorted(set(_gen_test))
                if len(_gen_unique) > 1:
                    print(f"    {'Gender':<12s}  {'Acc':>6s}  {'N':>5s}")
                    print(f"    {'\u2500' * 12}  {'\u2500' * 6}  {'\u2500' * 5}")
                    for _gen in _gen_unique:
                        _g_mask = np.array([g == _gen for g in _gen_test])
                        _g_n = int(_g_mask.sum())
                        if _g_n == 0:
                            continue
                        _g_acc = float(
                            (test_pred[_g_mask] == y_test[_g_mask]).mean())
                        print(f"    {_gen:<12s}  {_g_acc:6.1%}  {_g_n:5d}")

            # ── Performance by fighter experience ──────────────────────────────
            print()
            print_step("Accuracy by min fighter experience (prior fights):")
            _r_fights = (
                df_test.get("r_pre_total_fights",
                            pd.Series(0, index=df_test.index))
                .fillna(0)
                .values
            )
            _b_fights = (
                df_test.get("b_pre_total_fights",
                            pd.Series(0, index=df_test.index))
                .fillna(0)
                .values
            )
            _min_fights = np.minimum(_r_fights, _b_fights)
            _exp_bins = [
                (0, 3, "0\u20133 (debut)"),
                (4, 8, "4\u20138"),
                (9, 15, "9\u201315"),
                (16, 999, "16+"),
            ]
            print(f"    {'Experience':<16s}  {'Acc':>6s}  {'N':>5s}")
            print(f"    {'\u2500' * 16}  {'\u2500' * 6}  {'\u2500' * 5}")
            for _lo, _hi, _lbl in _exp_bins:
                _e_mask = (_min_fights >= _lo) & (_min_fights <= _hi)
                _e_n = int(_e_mask.sum())
                if _e_n == 0:
                    continue
                _e_acc = float((test_pred[_e_mask] == y_test[_e_mask]).mean())
                print(f"    {_lbl:<16s}  {_e_acc:6.1%}  {_e_n:5d}")

        except Exception as _e:
            print(f"  (test metrics unavailable: {_e})")
            import traceback

            traceback.print_exc()

    def _corner_bias_diagnostic(self, X_val, _y_val):
        print_step("Corner bias diagnostic:")

        d_indices = getattr(self, "_d_indices", None)
        if d_indices is None or len(d_indices) == 0:
            print("    Skipping: D-feature indices not available.")
            return 0.0, 0.0, 0.0

        X_flipped = X_val.copy()
        X_flipped[:, d_indices] = -X_val[:, d_indices]

        p_orig = self.stacking_clf.predict_proba(X_val)[:, 1]
        p_flip = self.stacking_clf.predict_proba(X_flipped)[:, 1]

        parity = p_orig + p_flip
        parity_error = np.abs(parity - 1.0)

        mean_error = parity_error.mean()
        max_error = parity_error.max()

        red_bias = p_orig.mean() - 0.5

        print_metric("  Mean parity error:", f"{mean_error:.4f}")
        print_metric("  Max parity error:", f"{max_error:.4f}")
        _bias_num = f"{red_bias:+.4f}"
        _bias_pct = f"({red_bias * 100:+.2f}%)"
        print(f"  {'  Red-corner bias:':<42}{_bias_num:>24} {_bias_pct}")

        if abs(red_bias) < 0.04:
            print("    Bias < 4% -- may underestimate real red-corner advantage")
        elif abs(red_bias) <= 0.08:
            print("    Bias 4-8%: consistent with UFC red-corner win rate")
        else:
            print(
                "    WARNING: Bias > 8% -- model may have over-learned corner position"
            )

        self._red_corner_bias = red_bias

        return mean_error, max_error, red_bias

    # Creates a mirrored copy of the DataFrame with red and blue corners swapped.
    # Used for two purposes:
    #   1. Training augmentation: doubles training data and teaches the model that
    #      fighter quality is independent of which corner they were assigned.
    #   2. Antisymmetric decomposition: paired with the original to compute D and Inv.
    # Derived/computed features (diffs, interactions, SVD components) are dropped and
    # recomputed via _recompute_derived_features() after the rename to stay consistent.
    def _corner_swap(self, df):
        df_s = df.copy()

        r_cols = [c for c in df_s.columns if c.startswith("r_")]
        b_cols = [c for c in df_s.columns if c.startswith("b_")]

        rename_map = {}
        for c in r_cols:
            rename_map[c] = "b_" + c[2:]
        for c in b_cols:
            rename_map[c] = "r_" + c[2:]
        df_s = df_s.rename(columns=rename_map)

        def flip_winner(w):
            if w == "Red":
                return "Blue"
            if w == "Blue":
                return "Red"
            return w

        df_s["winner"] = df_s["winner"].apply(flip_winner)

        derived_patterns = (
            "diff_",
            "elo_diff",
            "elo_r",
            "elo_b",
            "elo_ratio",
            "glicko_diff",
            "glicko_rd_diff",
            "r_glicko_r",
            "b_glicko_r",
            "r_glicko_rd",
            "b_glicko_rd",
            "r_glicko_vol",
            "b_glicko_vol",
            "n_common_",
            "common_opp_",
            "r_wins_vs_",
            "b_wins_vs_",
            "r_style_cluster",
            "b_style_cluster",
            "style_matchup_edge",
            "r_style_win_vs_cluster",
            "b_style_win_vs_cluster",
            "same_stance",
            "stance_matchup",
            "momentum_diff_",
            "streak_differential",
            "ko_threat_diff",
            "sub_threat_diff",
            "dec_tendency_diff",
            "r_finishing_tendency",
            "b_finishing_tendency",
            "finishing_matchup",
            "days_since_last_diff",
            "r_clinch_effectiveness",
            "b_clinch_effectiveness",
            "clinch_effectiveness_diff",
            "five_round_cardio_advantage",
            "r_chin_deterioration",
            "b_chin_deterioration",
            "chin_deterioration_diff",
            "opp_quality_diff",
            "trajectory_diff",
            "striking_svd_",
            "grappling_svd_",
            "physical_svd_",
            "form_svd_",
            "z_r_",
            "z_b_",
            "is_title_enc",
            "total_rounds_num",
            "gender_enc",
            "positional_striking_advantage",
            "target_distribution_advantage",
            "defensive_composite",
            "elo_x_form",
            "elo_x_win_ratio",
            "elo_x_finish",
            "elo_x_durability",
            "reach_x_striking",
            "height_x_reach",
            "physical_x_striking",
            "age_x_striking",
            "age_x_grappling",
            "age_x_durability",
            "age_x_win_streak",
            "experience_x_age",
            "td_x_defense",
            "submission_x_grappling",
            "striking_x_accuracy",
            "striking_x_defense",
            "ko_power_x_striking",
            "momentum_x_win_streak",
            "form_x_experience",
            "finish_x_momentum",
            "form_x_durability",
            "elite_finisher",
            "unstoppable_streak",
            "veteran_advantage",
            "diff_age_cubed",
            "diff_win_rate_vs_elite",
            "diff_win_rate_vs_strikers",
            "diff_win_rate_vs_grapplers",
            "championship_readiness",
            "declining_phase_diff",
            "r_last_fight_momentum",
            "b_last_fight_momentum",
            "last_fight_momentum_diff",
            "rounds_x_cardio",
            "rounds_x_finish_rate",
            "rounds_x_durability",
            "r_striking_vs_b_defense",
            "b_striking_vs_r_defense",
            "striking_exploitation_diff",
            "r_td_vs_b_td_defense",
            "b_td_vs_r_td_defense",
            "td_exploitation_diff",
            "r_sub_setup_efficiency",
            "b_sub_setup_efficiency",
            "sub_setup_diff",
            "r_sub_threat_vs_td_defense",
            "b_sub_threat_vs_td_defense",
            "sub_threat_vs_defense_diff",
            "r_striking_quality",
            "b_striking_quality",
            "striking_quality_diff",
            "r_accuracy_under_fire",
            "b_accuracy_under_fire",
            "accuracy_under_fire_diff",
            "r_damage_ratio",
            "b_damage_ratio",
            "damage_ratio_diff",
            "r_striking_output_quality",
            "b_striking_output_quality",
            "striking_output_quality_diff",
            "r_grappling_quality",
            "b_grappling_quality",
            "grappling_quality_diff",
            "r_total_defense_index",
            "b_total_defense_index",
            "total_defense_diff",
            "r_complete_fighter_index",
            "b_complete_fighter_index",
            "complete_fighter_diff",
            "r_pressure_index",
            "b_pressure_index",
            "pressure_index_diff",
            "elo_x_finish_rate",
            "streak_x_finish",
            "striking_exchange",
            "td_efficiency",
            "control_accuracy",
            "r_stance_enc",
            "b_stance_enc",
            "r_decision_win_rate",
            "b_decision_win_rate",
            "r_ko_win_rate",
            "b_ko_win_rate",
            "r_sub_win_rate",
            "b_sub_win_rate",
            "r_finish_rate",
            "b_finish_rate",
            "r_title_fight_exp",
            "b_title_fight_exp",
            "r_main_event_exp",
            "b_main_event_exp",
            "decision_win_rate_diff",
            "ko_win_rate_diff",
            "sub_win_rate_diff",
            "finish_rate_diff",
            "title_fight_exp_diff",
            "main_event_exp_diff",
            "r_defense_offense_balance",
            "b_defense_offense_balance",
            "defense_offense_balance_diff",
            "r_td_defense_offense_balance",
            "b_td_defense_offense_balance",
            "td_defense_offense_balance_diff",
            "finish_efficiency_diff",
            "r_precision_striking",
            "b_precision_striking",
            "precision_striking_diff",
            "r_quality_grappling_23",
            "b_quality_grappling_23",
            "quality_grappling_diff",
            "r_submission_threat_ratio",
            "b_submission_threat_ratio",
            "submission_threat_ratio_diff",
            "r_damage_absorption_efficiency",
            "b_damage_absorption_efficiency",
            "damage_absorption_efficiency_diff",
            "r_defense_versatility",
            "b_defense_versatility",
            "defense_versatility_diff",
            "r_total_offense_index",
            "b_total_offense_index",
            "total_offense_index_diff",
            "r_offensive_versatility",
            "b_offensive_versatility",
            "offensive_versatility_diff",
            "r_striker_index",
            "b_striker_index",
            "striker_index_diff",
            "win_loss_ratio_squared_diff",
            "r_experience_quality",
            "b_experience_quality",
            "experience_quality_diff",
            "r_win_efficiency",
            "b_win_efficiency",
            "win_efficiency_diff",
            "r_momentum_quality",
            "b_momentum_quality",
            "momentum_quality_diff",
            "r_reach_efficiency",
            "b_reach_efficiency",
            "reach_efficiency_diff",
            "r_size_adjusted_striking",
            "b_size_adjusted_striking",
            "size_adjusted_striking_diff",
            "r_size_adjusted_grappling",
            "b_size_adjusted_grappling",
            "size_adjusted_grappling_diff",
            "r_counter_fighter_index",
            "b_counter_fighter_index",
            "counter_fighter_index_diff",
            "r_finishing_threat_composite",
            "b_finishing_threat_composite",
            "finishing_threat_composite_diff",
            "r_complete_geo",
            "b_complete_geo",
            "complete_geo_diff",
            "r_pressure_fighter_index",
            "b_pressure_fighter_index",
            "pressure_fighter_index_diff",
            "r_absorption_vuln",
            "b_absorption_vuln",
            "absorption_vulnerability_index_diff",
            "r_combined_def_hole",
            "b_combined_def_hole",
            "combined_defensive_hole_diff",
            "r_td_pressure_t24",
            "b_td_pressure_t24",
            "td_vulnerability_under_pressure_diff",
            "r_strike_pressure_t24",
            "b_strike_pressure_t24",
            "strike_defense_under_volume_diff",
            "r_ctrl_sub_ratio",
            "b_ctrl_sub_ratio",
            "grappling_control_vs_submission_ratio_diff",
            "r_sub_def_necessity",
            "b_sub_def_necessity",
            "submission_defense_necessity_diff",
            "r_strike_synergy",
            "b_strike_synergy",
            "striking_volume_accuracy_synergy_diff",
            "r_td_paradox",
            "b_td_paradox",
            "takedown_efficiency_paradox_diff",
            "r_total_off_eff",
            "b_total_off_eff",
            "total_offensive_efficiency_index_diff",
            "r_sg_corr",
            "b_sg_corr",
            "striking_grappling_efficiency_correlation_diff",
            "r_def_allocation_balance",
            "b_def_allocation_balance",
            "defense_allocation_balance_diff",
            "r_combat_eff",
            "b_combat_eff",
            "total_combat_efficiency_index_diff",
            "net_striking_advantage",
            "striker_advantage",
            "grappler_advantage",
            "experience_gap",
            "ko_specialist_gap",
            "submission_specialist_gap",
            "skill_momentum",
            "r_desperation",
            "b_desperation",
            "desperation_diff",
            "r_freshness",
            "b_freshness",
            "freshness_advantage",
            "orthodox_vs_southpaw_advantage",
            "orthodox_vs_switch_advantage",
            "southpaw_vs_switch_advantage",
            "mirror_matchup",
            "r_finish_rate_accel",
            "b_finish_rate_accel",
            "finish_rate_acceleration_diff",
            "slpm_coefficient_of_variation_diff",
            "mileage_adjusted_age_diff",
            "performance_decline_velocity_diff",
            "r_distance_from_peak",
            "b_distance_from_peak",
            "distance_from_career_peak_diff",
            "r_career_inflection",
            "b_career_inflection",
            "career_inflection_point_diff",
            "r_prime_exit_risk",
            "b_prime_exit_risk",
            "prime_exit_risk_diff",
            "r_aging_power_penalty",
            "b_aging_power_penalty",
            "aging_power_striker_penalty_diff",
            "r_bayesian_finish",
            "b_bayesian_finish",
            "bayesian_finish_rate_diff",
            "r_layoff_veteran",
            "b_layoff_veteran",
            "layoff_veteran_interaction_diff",
            "r_elo_momentum",
            "b_elo_momentum",
            "elo_momentum_vs_competition_diff",
            "r_title_proximity",
            "b_title_proximity",
            "title_shot_proximity_score_diff",
            "r_elo_volatility",
            "b_elo_volatility",
            "elo_volatility_interaction_diff",
            "elite_performance_frequency_l10_diff",
            "r_conf_damage_ratio",
            "b_conf_damage_ratio",
            "confidence_weighted_damage_ratio_diff",
            "r_recent_form_ratio",
            "b_recent_form_ratio",
            "recent_form_ratio_diff",
            "r_finish_method_diversity",
            "b_finish_method_diversity",
            "finish_method_diversity_diff",
            "r_cross_domain_compensation",
            "b_cross_domain_compensation",
            "cross_domain_compensation_index_diff",
            "r_recent_vs_career_striking",
            "b_recent_vs_career_striking",
            "recent_vs_career_striking_diff",
            "r_striking_consistency_ratio",
            "b_striking_consistency_ratio",
            "striking_consistency_ratio_diff",
            "performance_volatility_l10_diff",
            "tactical_evolution_score_diff",
            "r_daniel_custom",
            "b_daniel_custom",
            "daniel_custom_feature_diff",
            "r_elo_confidence_weighted",
            "b_elo_confidence_weighted",
            "elo_confidence_weighted_diff",
            "r_streak_quality",
            "b_streak_quality",
            "streak_quality_diff",
            "r_reach_defense",
            "b_reach_defense",
            "reach_defense_diff",
            "r_net_striking_efficiency",
            "b_net_striking_efficiency",
            "net_striking_efficiency_diff",
            "r_style_glicko_prime",
            "b_style_glicko_prime",
            "style_glicko_prime_diff",
            "common_style_consensus",
            "r_age_elo_climb",
            "b_age_elo_climb",
            "age_elo_climb_diff",
            "r_quality_battle_tested",
            "b_quality_battle_tested",
            "quality_battle_tested_diff",
            "r_physical_style_dominance",
            "b_physical_style_dominance",
            "physical_style_dominance_diff",
            "r_peak_fresh_pressure",
            "b_peak_fresh_pressure",
            "peak_fresh_pressure_diff",
            "tri_rating_consensus",
            "r_wc_ko_style_physical",
            "b_wc_ko_style_physical",
            "wc_ko_style_physical_diff",
            "r_wc_acc_physical_style",
            "b_wc_acc_physical_style",
            "wc_acc_physical_style_diff",
            "r_wc_striking_dominance",
            "b_wc_striking_dominance",
            "wc_striking_dominance_diff",
            "r_wc_sub_threat_compound",
            "b_wc_sub_threat_compound",
            "wc_sub_threat_compound_diff",
            "form_svd_style_synergy",
            "striking_svd_physical",
            "grappling_svd_style_synergy",
            "r_stable_prime_combat",
            "b_stable_prime_combat",
            "stable_prime_combat_diff",
            "r_physical_elo_youth",
            "b_physical_elo_youth",
            "physical_elo_youth_diff",
            "r_physical_glicko_youth",
            "b_physical_glicko_youth",
            "physical_glicko_youth_diff",
            "r_physical_style_glicko_prime",
            "b_physical_style_glicko_prime",
            "physical_style_glicko_prime_diff",
            "r_physical_combat_prime",
            "b_physical_combat_prime",
            "physical_combat_prime_diff",
            "r_fresh_physical_elo",
            "b_fresh_physical_elo",
            "fresh_physical_elo_diff",
            "grappling_svd_physical",
            "r_wc_td_physical_style",
            "b_wc_td_physical_style",
            "wc_td_physical_style_diff",
            "r_physical_defensive",
            "b_physical_defensive",
            "physical_defensive_diff",
            "r_fresh_style_glicko",
            "b_fresh_style_glicko",
            "fresh_style_glicko_diff",
            "r_physical_glicko_prime",
            "b_physical_glicko_prime",
            "physical_glicko_prime_diff",
            "form_svd_physical",
            "form_svd1_style_glicko",
            "striking_svd0_physical",
            "form_svd0_physical",
            "grappling_svd3_style",
            "r_winrate_physical_glicko",
            "b_winrate_physical_glicko",
            "winrate_physical_glicko_diff",
            "r_finish_physical_youth",
            "b_finish_physical_youth",
            "finish_physical_youth_diff",
            "r_elite_winrate_physical",
            "b_elite_winrate_physical",
            "elite_winrate_physical_diff",
            "r_ko_rate_physical_prime",
            "b_ko_rate_physical_prime",
            "ko_rate_physical_prime_diff",
            "r_winrate_style_glicko_prime",
            "b_winrate_style_glicko_prime",
            "winrate_style_glicko_prime_diff",
            "r_streak_physical_glicko",
            "b_streak_physical_glicko",
            "streak_physical_glicko_diff",
            "r_title_physical_glicko",
            "b_title_physical_glicko",
            "title_physical_glicko_diff",
            "r_recent_winrate_physical_youth",
            "b_recent_winrate_physical_youth",
            "recent_winrate_physical_youth_diff",
            "r_sub_rate_physical_youth",
            "b_sub_rate_physical_youth",
            "sub_rate_physical_youth_diff",
            "r_winrate_physical_glicko_youth",
            "b_winrate_physical_glicko_youth",
            "winrate_physical_glicko_youth_diff",
            "r_winrate_physical_style_glicko_prime",
            "b_winrate_physical_style_glicko_prime",
            "winrate_physical_style_glicko_prime_diff",
            "r_recent_physical_glicko_youth",
            "b_recent_physical_glicko_youth",
            "recent_physical_glicko_youth_diff",
            "r_opp_quality_physical_glicko",
            "b_opp_quality_physical_glicko",
            "opp_quality_physical_glicko_diff",
            "r_momentum_physical_glicko",
            "b_momentum_physical_glicko",
            "momentum_physical_glicko_diff",
            "r_finish_physical_glicko_prime",
            "b_finish_physical_glicko_prime",
            "finish_physical_glicko_prime_diff",
            "r_winrate_style_physical_youth",
            "b_winrate_style_physical_youth",
            "winrate_style_physical_youth_diff",
            "r_streak_winrate_physical_glicko",
            "b_streak_winrate_physical_glicko",
            "streak_winrate_physical_glicko_diff",
            "r_opp_quality_winrate_physical_youth",
            "b_opp_quality_winrate_physical_youth",
            "opp_quality_winrate_physical_youth_diff",
            "r_title_physical_glicko_youth",
            "b_title_physical_glicko_youth",
            "title_physical_glicko_youth_diff",
            "r_ko_rate_style_glicko_prime",
            "b_ko_rate_style_glicko_prime",
            "ko_rate_style_glicko_prime_diff",
            "r_fresh_winrate_physical_glicko_youth",
            "b_fresh_winrate_physical_glicko_youth",
            "fresh_winrate_physical_glicko_youth_diff",
            "r_strike_def_physical_glicko_youth",
            "b_strike_def_physical_glicko_youth",
            "strike_def_physical_glicko_youth_diff",
            "r_peak_physical_glicko_youth",
            "b_peak_physical_glicko_youth",
            "peak_physical_glicko_youth_diff",
            "r_elite_winrate_physical_glicko_prime",
            "b_elite_winrate_physical_glicko_prime",
            "elite_winrate_physical_glicko_prime_diff",
            "r_recent_opp_quality_physical_glicko",
            "b_recent_opp_quality_physical_glicko",
            "recent_opp_quality_physical_glicko_diff",
            "r_daniel_physical_glicko_youth",
            "b_daniel_physical_glicko_youth",
            "daniel_physical_glicko_youth_diff",
            "r_title_winrate_physical_glicko_youth",
            "b_title_winrate_physical_glicko_youth",
            "title_winrate_physical_glicko_youth_diff",
            "r_complete_def_physical_glicko",
            "b_complete_def_physical_glicko",
            "complete_def_physical_glicko_diff",
            "r_career_peak_physical_glicko",
            "b_career_peak_physical_glicko",
            "career_peak_physical_glicko_diff",
            "r_finish_def_physical_glicko",
            "b_finish_def_physical_glicko",
            "finish_def_physical_glicko_diff",
            "r_complete_def_physical_glicko_youth",
            "b_complete_def_physical_glicko_youth",
            "complete_def_physical_glicko_youth_diff",
            "r_td_def_physical_glicko_youth",
            "b_td_def_physical_glicko_youth",
            "td_def_physical_glicko_youth_diff",
            "r_peak_winrate_physical_glicko_youth",
            "b_peak_winrate_physical_glicko_youth",
            "peak_winrate_physical_glicko_youth_diff",
            "r_strike_def_winrate_physical_glicko",
            "b_strike_def_winrate_physical_glicko",
            "strike_def_winrate_physical_glicko_diff",
            "r_peak_style_physical_glicko_youth",
            "b_peak_style_physical_glicko_youth",
            "peak_style_physical_glicko_youth_diff",
            "r_career_peak_winrate_physical_glicko",
            "b_career_peak_winrate_physical_glicko",
            "career_peak_winrate_physical_glicko_diff",
            "r_strike_def_style_physical_glicko_youth",
            "b_strike_def_style_physical_glicko_youth",
            "strike_def_style_physical_glicko_youth_diff",
            "r_strike_def_physical_glicko_prime",
            "b_strike_def_physical_glicko_prime",
            "strike_def_physical_glicko_prime_diff",
            "r_peak_opp_quality_physical_glicko",
            "b_peak_opp_quality_physical_glicko",
            "peak_opp_quality_physical_glicko_diff",
            "r_fresh_strike_def_physical_glicko_youth",
            "b_fresh_strike_def_physical_glicko_youth",
            "fresh_strike_def_physical_glicko_youth_diff",
            # Tier 38
            "r_td_def_winrate_physical_glicko_youth",
            "b_td_def_winrate_physical_glicko_youth",
            "td_def_winrate_physical_glicko_youth_diff",
            "r_combat_eff_physical_glicko_youth",
            "b_combat_eff_physical_glicko_youth",
            "combat_eff_physical_glicko_youth_diff",
            "r_finish_def_physical_glicko_youth",
            "b_finish_def_physical_glicko_youth",
            "finish_def_physical_glicko_youth_diff",
            "r_streak_winrate_physical_glicko_youth",
            "b_streak_winrate_physical_glicko_youth",
            "streak_winrate_physical_glicko_youth_diff",
            "r_ko_rate_def_physical_glicko_youth",
            "b_ko_rate_def_physical_glicko_youth",
            "ko_rate_def_physical_glicko_youth_diff",
            "r_td_def_winrate_physical_glicko_prime",
            "b_td_def_winrate_physical_glicko_prime",
            "td_def_winrate_physical_glicko_prime_diff",
            "r_complete_def_winrate_physical_glicko",
            "b_complete_def_winrate_physical_glicko",
            "complete_def_winrate_physical_glicko_diff",
            "r_career_peak_strike_def_physical_glicko",
            "b_career_peak_strike_def_physical_glicko",
            "career_peak_strike_def_physical_glicko_diff",
            "r_peak_td_def_physical_glicko_youth",
            "b_peak_td_def_physical_glicko_youth",
            "peak_td_def_physical_glicko_youth_diff",
            "r_combat_eff_winrate_physical_glicko",
            "b_combat_eff_winrate_physical_glicko",
            "combat_eff_winrate_physical_glicko_diff",
            # Tier 39
            "r_combat_eff_style_physical_glicko_youth",
            "b_combat_eff_style_physical_glicko_youth",
            "combat_eff_style_physical_glicko_youth_diff",
            "r_peak_cbt_style_physical_glicko_youth",
            "b_peak_cbt_style_physical_glicko_youth",
            "peak_cbt_style_physical_glicko_youth_diff",
            "r_fresh_style_physical_glicko_youth",
            "b_fresh_style_physical_glicko_youth",
            "fresh_style_physical_glicko_youth_diff",
            "r_recent_style_physical_glicko_youth",
            "b_recent_style_physical_glicko_youth",
            "recent_style_physical_glicko_youth_diff",
            "r_td_def_style_physical_glicko_youth",
            "b_td_def_style_physical_glicko_youth",
            "td_def_style_physical_glicko_youth_diff",
            "r_finish_style_physical_glicko_youth",
            "b_finish_style_physical_glicko_youth",
            "finish_style_physical_glicko_youth_diff",
            "r_complete_def_style_physical_glicko_youth",
            "b_complete_def_style_physical_glicko_youth",
            "complete_def_style_physical_glicko_youth_diff",
            "r_career_peak_style_physical_glicko_youth",
            "b_career_peak_style_physical_glicko_youth",
            "career_peak_style_physical_glicko_youth_diff",
            "r_opp_quality_style_physical_glicko_youth",
            "b_opp_quality_style_physical_glicko_youth",
            "opp_quality_style_physical_glicko_youth_diff",
            "r_streak_style_physical_glicko_youth",
            "b_streak_style_physical_glicko_youth",
            "streak_style_physical_glicko_youth_diff",
        )

        cols_to_drop_set = set()
        for col in df_s.columns:
            for pat in derived_patterns:
                if pat.endswith("_"):
                    if col.startswith(pat):
                        cols_to_drop_set.add(col)
                        break
                else:
                    if col == pat or col.endswith("_sq") or col.endswith("_abs"):
                        cols_to_drop_set.add(col)
                        break

        for col in df_s.columns:
            for svd_prefix in (
                "striking_svd_",
                "grappling_svd_",
                "physical_svd_",
                "form_svd_",
            ):
                if col.startswith(svd_prefix):
                    cols_to_drop_set.add(col)
            for z_prefix in ("z_r_", "z_b_"):
                if col.startswith(z_prefix):
                    cols_to_drop_set.add(col)

        df_s = df_s.drop(
            columns=[c for c in cols_to_drop_set if c in df_s.columns], errors="ignore"
        )

        df_s = self._recompute_derived_features(df_s)

        return df_s

    def _recompute_derived_features(self, df):

        self._apply_wc_finish_route_prior_chrono_safe(df)

        raw_pairs = [
            ("r_height", "b_height"),
            ("r_reach", "b_reach"),
            ("r_weight", "b_weight"),
            ("r_age_at_event", "b_age_at_event"),
            ("r_ape_index", "b_ape_index"),
        ]
        for rc, bc in raw_pairs:
            if rc in df.columns and bc in df.columns:
                df[f"diff_{rc[2:]}"] = pd.to_numeric(df[rc], errors="coerce").fillna(
                    0
                ) - pd.to_numeric(df[bc], errors="coerce").fillna(0)

        pre_pairs = [
            ("r_pre_wins", "b_pre_wins"),
            ("r_pre_losses", "b_pre_losses"),
            ("r_pre_ko_wins", "b_pre_ko_wins"),
            ("r_pre_sub_wins", "b_pre_sub_wins"),
            ("r_pre_dec_wins", "b_pre_dec_wins"),
            ("r_pre_total_fights", "b_pre_total_fights"),
            ("r_pre_finish_rate", "b_pre_finish_rate"),
            ("r_pre_win_streak", "b_pre_win_streak"),
            ("r_pre_loss_streak", "b_pre_loss_streak"),
            ("r_pre_title_fights", "b_pre_title_fights"),
            ("r_pre_title_wins", "b_pre_title_wins"),
            ("r_pre_avg_fight_time", "b_pre_avg_fight_time"),
            ("r_pre_sig_str_acc", "b_pre_sig_str_acc"),
            ("r_pre_td_acc", "b_pre_td_acc"),
            ("r_pre_sub_att_rate", "b_pre_sub_att_rate"),
            ("r_pre_kd_rate", "b_pre_kd_rate"),
            ("r_pre_ctrl_avg", "b_pre_ctrl_avg"),
            ("r_pre_SLpM", "b_pre_SLpM"),
            ("r_pre_SApM", "b_pre_SApM"),
            ("r_pre_td_avg", "b_pre_td_avg"),
            ("r_pre_sub_att_per15", "b_pre_sub_att_per15"),
            ("r_pre_ctrl_min_per15", "b_pre_ctrl_min_per15"),
            ("r_pre_td_per15", "b_pre_td_per15"),
        ]
        for rc, bc in pre_pairs:
            if rc in df.columns and bc in df.columns:
                df[f"diff_{rc[6:]}"] = df[rc].fillna(0) - df[bc].fillna(0)

        rolling_pairs = [
            ("r_rolling3_wins", "b_rolling3_wins"),
            ("r_rolling3_sig_str", "b_rolling3_sig_str"),
            ("r_rolling3_td", "b_rolling3_td"),
            ("r_rolling3_kd", "b_rolling3_kd"),
            ("r_rolling3_sub_att", "b_rolling3_sub_att"),
            ("r_rolling5_wins", "b_rolling5_wins"),
            ("r_rolling5_sig_str", "b_rolling5_sig_str"),
            ("r_rolling5_td", "b_rolling5_td"),
            ("r_rolling5_kd", "b_rolling5_kd"),
        ]
        for rc, bc in rolling_pairs:
            if rc in df.columns and bc in df.columns:
                df[f"diff_{rc[2:]}"] = df[rc].fillna(0) - df[bc].fillna(0)

        if "r_elo_pre_fight" in df.columns and "b_elo_pre_fight" in df.columns:
            df["elo_diff"] = df["r_elo_pre_fight"] - df["b_elo_pre_fight"]
            df["elo_r"] = df["r_elo_pre_fight"]
            df["elo_b"] = df["b_elo_pre_fight"]
            df["elo_ratio"] = df["r_elo_pre_fight"] / \
                (df["b_elo_pre_fight"] + 1e-6)
        else:
            df["elo_diff"] = 0.0
            df["elo_r"] = 1500.0
            df["elo_b"] = 1500.0
            df["elo_ratio"] = 1.0

        if "r_glicko_pre_r" in df.columns and "b_glicko_pre_r" in df.columns:
            df["r_glicko_r"] = df["r_glicko_pre_r"]
            df["r_glicko_rd"] = df["r_glicko_pre_rd"]
            df["r_glicko_vol"] = df["r_glicko_pre_vol"]
            df["b_glicko_r"] = df["b_glicko_pre_r"]
            df["b_glicko_rd"] = df["b_glicko_pre_rd"]
            df["b_glicko_vol"] = df["b_glicko_pre_vol"]
        df["glicko_diff"] = df.get(
            "r_glicko_r", pd.Series(1500.0, index=df.index)
        ).fillna(1500.0) - df.get(
            "b_glicko_r", pd.Series(1500.0, index=df.index)
        ).fillna(1500.0)
        df["glicko_rd_diff"] = df.get(
            "r_glicko_rd", pd.Series(200.0, index=df.index)
        ).fillna(200.0) - df.get(
            "b_glicko_rd", pd.Series(200.0, index=df.index)
        ).fillna(200.0)

        fe = self.feature_engineer
        z_feats = [
            "r_pre_SLpM",
            "r_pre_SApM",
            "r_pre_sig_str_acc",
            "r_pre_td_avg",
            "r_pre_sub_att_rate",
            "r_pre_kd_rate",
            "b_pre_SLpM",
            "b_pre_SApM",
            "b_pre_sig_str_acc",
            "b_pre_td_avg",
            "b_pre_sub_att_rate",
            "b_pre_kd_rate",
        ]
        for col in z_feats:
            df[f"z_{col}"] = 0.0
        if "weight_class" in df.columns and "event_date" in df.columns:
            for idx, row in df.iterrows():
                wc = str(row.get("weight_class", ""))
                yr_val = row.get("event_date", None)
                try:
                    yr = pd.Timestamp(yr_val).year if pd.notna(
                        yr_val) else 2000
                except Exception:
                    yr = 2000
                for feat in z_feats:
                    if feat in df.columns:
                        v = row.get(feat, 0)
                        try:
                            v = float(v)
                            if not math.isnan(v):
                                df.at[idx, f"z_{feat}"] = fe.get_z_score(
                                    wc, yr, feat, v
                                )
                        except (TypeError, ValueError):
                            pass

        if "r_fighter" in df.columns and "b_fighter" in df.columns:
            n_common, r_wins_c, b_wins_c, co_edge = [], [], [], []
            for _, row in df.iterrows():
                r = str(row.get("r_fighter", ""))
                b = str(row.get("b_fighter", ""))
                feat = fe.get_common_opponent_features(r, b)
                n_common.append(feat["n_common_opponents"])
                r_wins_c.append(feat["r_wins_vs_common"])
                b_wins_c.append(feat["b_wins_vs_common"])
                co_edge.append(feat["common_opp_edge"])
            df["n_common_opponents"] = n_common
            df["r_wins_vs_common"] = r_wins_c
            df["b_wins_vs_common"] = b_wins_c
            df["common_opp_edge"] = co_edge

        if "r_fighter" in df.columns and "b_fighter" in df.columns:
            r_cluster, b_cluster, style_edge_list = [], [], []
            r_style_win, b_style_win = [], []
            cluster_pair_finish = []
            for _, row in df.iterrows():
                r = str(row.get("r_fighter", ""))
                b = str(row.get("b_fighter", ""))
                rc = fe.get_fighter_cluster(r)
                bc = fe.get_fighter_cluster(b)
                mf = fe.get_style_matchup_features(rc, bc)
                r_cluster.append(rc)
                b_cluster.append(bc)
                style_edge_list.append(mf["style_matchup_edge"])
                r_style_win.append(mf["r_style_win_vs_opp_cluster"])
                b_style_win.append(mf["b_style_win_vs_opp_cluster"])
                if rc >= 0 and bc >= 0:
                    cluster_pair_finish.append(
                        fe.get_style_matchup_finish_rate(rc, bc))
                else:
                    cluster_pair_finish.append(0.5)
            df["r_style_cluster"] = r_cluster
            df["b_style_cluster"] = b_cluster
            df["style_matchup_edge"] = style_edge_list
            df["r_style_win_vs_cluster"] = r_style_win
            df["b_style_win_vs_cluster"] = b_style_win
            df["cluster_pair_finish_rate"] = cluster_pair_finish

        stance_map = {"Orthodox": 0, "Southpaw": 1,
                      "Switch": 2, "Open Stance": 3}
        for col in ["r_stance", "b_stance"]:
            if col in df.columns:
                df[f"{col}_enc"] = df[col].map(stance_map).fillna(-1)
        if "r_stance_enc" in df.columns and "b_stance_enc" in df.columns:
            df["stance_matchup"] = (
                df["r_stance_enc"].astype(
                    str) + "_" + df["b_stance_enc"].astype(str)
            )
            df["same_stance"] = (df["r_stance_enc"] ==
                                 df["b_stance_enc"]).astype(int)

        if "elo_diff" in df.columns and "diff_finish_rate" in df.columns:
            df["elo_x_finish_rate"] = df["elo_diff"] * df["diff_finish_rate"]
        if "diff_pre_win_streak" in df.columns and "diff_pre_finish_rate" in df.columns:
            df["streak_x_finish"] = (
                df["diff_pre_win_streak"] * df["diff_pre_finish_rate"]
            )
        if "diff_pre_SLpM" in df.columns and "diff_pre_SApM" in df.columns:
            df["striking_exchange"] = df["diff_pre_SLpM"] - df["diff_pre_SApM"]
        if "diff_pre_td_avg" in df.columns and "diff_pre_td_acc" in df.columns:
            df["td_efficiency"] = df["diff_pre_td_avg"] * df["diff_pre_td_acc"]
        if "diff_pre_sig_str_acc" in df.columns and "diff_pre_ctrl_avg" in df.columns:
            df["control_accuracy"] = (
                df["diff_pre_sig_str_acc"] * df["diff_pre_ctrl_avg"]
            )

        if "r_pre_wins" in df.columns and "r_pre_losses" in df.columns:
            df["diff_win_loss_ratio"] = df["r_pre_wins"] / (
                df["r_pre_losses"].clip(lower=0) + 1.0
            ) - df["b_pre_wins"] / (df["b_pre_losses"].clip(lower=0) + 1.0)
        poly_cols = ["elo_diff", "glicko_diff", "diff_win_loss_ratio"]
        for col in poly_cols:
            if col in df.columns:
                df[f"{col}_sq"] = df[col] ** 2
                df[f"{col}_abs"] = df[col].abs()

        if "r_rolling3_wins" in df.columns and "b_rolling3_wins" in df.columns:
            df["momentum_diff_3"] = df["r_rolling3_wins"] - df["b_rolling3_wins"]
        if "r_rolling5_wins" in df.columns and "b_rolling5_wins" in df.columns:
            df["momentum_diff_5"] = df["r_rolling5_wins"] - df["b_rolling5_wins"]
        if "r_pre_win_streak" in df.columns:
            df["streak_differential"] = df.get(
                "r_pre_win_streak", pd.Series(0.0, index=df.index)
            ).fillna(0) - df.get(
                "b_pre_win_streak", pd.Series(0.0, index=df.index)
            ).fillna(0)

        if "r_pre_ko_wins" in df.columns:
            df["ko_threat_diff"] = df.get(
                "r_pre_ko_wins", pd.Series(0.0, index=df.index)
            ).fillna(0) - df.get(
                "b_pre_ko_wins", pd.Series(0.0, index=df.index)
            ).fillna(0)
        if "r_pre_sub_wins" in df.columns:
            df["sub_threat_diff"] = df.get(
                "r_pre_sub_wins", pd.Series(0.0, index=df.index)
            ).fillna(0) - df.get(
                "b_pre_sub_wins", pd.Series(0.0, index=df.index)
            ).fillna(0)
        if "r_pre_dec_wins" in df.columns:
            df["dec_tendency_diff"] = df.get(
                "r_pre_dec_wins", pd.Series(0.0, index=df.index)
            ).fillna(0) - df.get(
                "b_pre_dec_wins", pd.Series(0.0, index=df.index)
            ).fillna(0)
        if "r_pre_finish_rate" in df.columns:
            df["r_finishing_tendency"] = df.get(
                "r_pre_finish_rate", pd.Series(0.0, index=df.index)
            ).fillna(0)
            df["b_finishing_tendency"] = df.get(
                "b_pre_finish_rate", pd.Series(0.0, index=df.index)
            ).fillna(0)
            df["finishing_matchup"] = (
                df["r_finishing_tendency"] * df["b_finishing_tendency"]
            )

        for prefix in ["r", "b"]:
            dec_col = f"{prefix}_pre_dec_wins"
            ko_col = f"{prefix}_pre_ko_wins"
            sub_col = f"{prefix}_pre_sub_wins"
            total_col = f"{prefix}_pre_total_fights"
            if all(c in df.columns for c in [dec_col, ko_col, sub_col, total_col]):
                denom = df[total_col].clip(lower=1)
                df[f"{prefix}_decision_win_rate"] = df[dec_col] / denom
                df[f"{prefix}_ko_win_rate"] = df[ko_col] / denom
                df[f"{prefix}_sub_win_rate"] = df[sub_col] / denom
                df[f"{prefix}_finish_rate"] = (
                    df[ko_col] + df[sub_col]) / denom
            if f"{prefix}_pre_title_fights" in df.columns:
                df[f"{prefix}_title_fight_exp"] = df[f"{prefix}_pre_title_fights"]
            if f"{prefix}_pre_five_round_fights" in df.columns:
                df[f"{prefix}_main_event_exp"] = df[f"{prefix}_pre_five_round_fights"]
        for feat in [
            "decision_win_rate",
            "ko_win_rate",
            "sub_win_rate",
            "finish_rate",
            "title_fight_exp",
            "main_event_exp",
        ]:
            r_col = f"r_{feat}"
            b_col = f"b_{feat}"
            if r_col in df.columns and b_col in df.columns:
                df[f"{feat}_diff"] = df[r_col] - df[b_col]

        # ── Grappling pathway: per-side derived features ─────────────────
        for pfx in ["r", "b"]:
            sa15 = f"{pfx}_pre_sub_att_per15"
            cm15 = f"{pfx}_pre_ctrl_min_per15"
            td15 = f"{pfx}_pre_td_per15"
            tda = f"{pfx}_pre_td_acc"
            wins_col = f"{pfx}_pre_wins"
            sub_w = f"{pfx}_pre_sub_wins"
            ko_w = f"{pfx}_pre_ko_wins"
            dec_w = f"{pfx}_pre_dec_wins"
            losses_col = f"{pfx}_pre_losses"
            sub_l = f"{pfx}_pre_sub_losses"
            ko_l = f"{pfx}_pre_ko_losses"
            dec_l = f"{pfx}_pre_dec_losses"

            if all(c in df.columns for c in [wins_col, sub_w, ko_w, dec_w]):
                _w = df[wins_col].clip(lower=1)
                df[f"{pfx}_sub_wins_share"] = df[sub_w].fillna(0) / _w
                df[f"{pfx}_dec_wins_share"] = df[dec_w].fillna(0) / _w
                df[f"{pfx}_finish_wins_share"] = (
                    df[ko_w].fillna(0) + df[sub_w].fillna(0)
                ) / _w

            if all(c in df.columns for c in [losses_col, sub_l, ko_l, dec_l]):
                _l = df[losses_col].clip(lower=1)
                df[f"{pfx}_sub_losses_share"] = df[sub_l].fillna(0) / _l
                df[f"{pfx}_dec_losses_share"] = df[dec_l].fillna(0) / _l
                _total_fin_l = df[ko_l].fillna(0) + df[sub_l].fillna(0)
                df[f"{pfx}_finish_loss_rate"] = _total_fin_l / df[
                    f"{pfx}_pre_total_fights"
                ].clip(lower=1)

            if all(c in df.columns for c in [sa15, cm15, td15, tda]):
                _sa = df[sa15].fillna(0)
                _cm = df[cm15].fillna(0)
                _td = df[td15].fillna(0)
                _ta = df[tda].fillna(0)
                df[f"{pfx}_sub_hunt_rate"] = _sa / (_cm + 0.25)
                df[f"{pfx}_ctrl_no_sub_rate"] = _cm / (_sa + 1.0)
                df[f"{pfx}_sub_att_per_td"] = _sa / (_td + 1.0)
                df[f"{pfx}_td_to_sub_chain"] = _td * _ta * _sa
                df[f"{pfx}_ctrl_to_sub_chain"] = _cm * (_sa / (_cm + 0.25))

            if all(c in df.columns for c in [sub_w, sa15, f"{pfx}_pre_total_fights"]):
                _nf = df[f"{pfx}_pre_total_fights"].clip(lower=1)
                _sa_total = df[sa15].fillna(0) * _nf / 15.0
                df[f"{pfx}_sub_win_conversion"] = df[sub_w].fillna(0) / (
                    _sa_total + 1.0
                )

        # ── Grappling pathway: matchup interaction features ──────────────
        for focal, opp in [("r", "b"), ("b", "r")]:
            fa = f"{focal}_pre_td_acc"
            od = f"{opp}_pre_td_def"
            if fa in df.columns and od in df.columns:
                df[f"{focal}_td_acc_vs_{opp}_td_def"] = df[fa].fillna(0) - df[
                    od
                ].fillna(0)

            fsws = f"{focal}_sub_wins_share"
            osls = f"{opp}_sub_losses_share"
            if fsws in df.columns and osls in df.columns:
                df[f"{focal}_sub_wr_vs_{opp}_sub_lr"] = df[fsws].fillna(0) - df[
                    osls
                ].fillna(0)

            fshr = f"{focal}_sub_hunt_rate"
            if fshr in df.columns and osls in df.columns:
                df[f"{focal}_sub_hunt_x_{opp}_sub_vuln"] = df[fshr].fillna(0) * df[
                    osls
                ].fillna(0)

            ftsc = f"{focal}_td_to_sub_chain"
            if ftsc in df.columns and od in df.columns:
                df[f"{focal}_td_sub_chain_x_{opp}_td_weak"] = df[ftsc].fillna(0) * (
                    1.0 - df[od].fillna(0)
                )

            fcns = f"{focal}_ctrl_no_sub_rate"
            odls = f"{opp}_dec_losses_share"
            if fcns in df.columns and odls in df.columns:
                df[f"{focal}_ctrl_dec_x_{opp}_dec_loss"] = df[fcns].fillna(0) * df[
                    odls
                ].fillna(0)

            ffws = f"{focal}_finish_wins_share"
            oflr = f"{opp}_finish_loss_rate"
            if ffws in df.columns and oflr in df.columns:
                df[f"{focal}_finish_x_{opp}_finish_vuln"] = df[ffws].fillna(0) * df[
                    oflr
                ].fillna(0)

        # ── Grappling pathway: composite route scores ────────────────────
        for pfx in ["r", "b"]:
            opp = "b" if pfx == "r" else "r"
            _cols_sub = [
                f"{pfx}_pre_td_per15",
                f"{pfx}_pre_td_acc",
                f"{pfx}_pre_ctrl_min_per15",
                f"{pfx}_sub_hunt_rate",
                f"{pfx}_sub_wins_share",
                f"{opp}_sub_losses_share",
                f"{opp}_pre_td_def",
            ]
            if all(c in df.columns for c in _cols_sub):
                _td15 = df[f"{pfx}_pre_td_per15"].fillna(0).clip(0, 10) / 10.0
                _tda = df[f"{pfx}_pre_td_acc"].fillna(0).clip(0, 1)
                _cm15 = df[f"{pfx}_pre_ctrl_min_per15"].fillna(
                    0).clip(0, 10) / 10.0
                _shr = df[f"{pfx}_sub_hunt_rate"].fillna(0).clip(0, 5) / 5.0
                _sws = df[f"{pfx}_sub_wins_share"].fillna(0).clip(0, 1)
                _osls = df[f"{opp}_sub_losses_share"].fillna(0).clip(0, 1)
                _otdw = 1.0 - df[f"{opp}_pre_td_def"].fillna(0).clip(0, 1)
                df[f"{pfx}_submission_path_score"] = (
                    _td15 + _tda + _cm15 + _shr + _sws + _osls + _otdw
                ) / 7.0

            _cols_dec = [
                f"{pfx}_pre_ctrl_min_per15",
                f"{pfx}_pre_td_per15",
                f"{pfx}_dec_wins_share",
                f"{pfx}_ctrl_no_sub_rate",
                f"{opp}_dec_losses_share",
            ]
            if all(c in df.columns for c in _cols_dec):
                _cm15 = df[f"{pfx}_pre_ctrl_min_per15"].fillna(
                    0).clip(0, 10) / 10.0
                _td15 = df[f"{pfx}_pre_td_per15"].fillna(0).clip(0, 10) / 10.0
                _dws = df[f"{pfx}_dec_wins_share"].fillna(0).clip(0, 1)
                _cns = df[f"{pfx}_ctrl_no_sub_rate"].fillna(
                    0).clip(0, 10) / 10.0
                _odls = df[f"{opp}_dec_losses_share"].fillna(0).clip(0, 1)
                df[f"{pfx}_decision_ctrl_score"] = (
                    _cm15 + _td15 + _dws + _cns + _odls
                ) / 5.0

        # Diff features for per-side pathway features
        for feat in [
            "sub_hunt_rate",
            "ctrl_no_sub_rate",
            "sub_att_per_td",
            "td_to_sub_chain",
            "ctrl_to_sub_chain",
            "sub_win_conversion",
            "sub_wins_share",
            "dec_wins_share",
            "finish_wins_share",
            "submission_path_score",
            "decision_ctrl_score",
        ]:
            rc = f"r_{feat}"
            bc = f"b_{feat}"
            if rc in df.columns and bc in df.columns:
                df[f"{feat}_diff"] = df[rc].fillna(0) - df[bc].fillna(0)

        # days_since_last diff (replaces discretized ring_rust)
        if "r_days_since_last" in df.columns and "b_days_since_last" in df.columns:
            df["days_since_last_diff"] = (
                df["r_days_since_last"].fillna(365) - df["b_days_since_last"].fillna(365)
            )

        r_slpm = df.get(
            "r_pre_SLpM", df.get("r_pro_SLpM", pd.Series(0.0, index=df.index))
        ).fillna(0)
        b_slpm = df.get(
            "b_pre_SLpM", df.get("b_pro_SLpM", pd.Series(0.0, index=df.index))
        ).fillna(0)

        r_acc = df.get("r_pre_sig_str_acc", pd.Series(0.45, index=df.index)).fillna(
            0.45
        )
        b_acc = df.get("b_pre_sig_str_acc", pd.Series(0.45, index=df.index)).fillna(
            0.45
        )

        r_clinch_pct = df.get(
            "r_pre_clinch_pct", pd.Series(0.0, index=df.index)
        ).fillna(0)
        b_clinch_pct = df.get(
            "b_pre_clinch_pct", pd.Series(0.0, index=df.index)
        ).fillna(0)
        df["r_clinch_effectiveness"] = r_clinch_pct * r_slpm * r_acc
        df["b_clinch_effectiveness"] = b_clinch_pct * b_slpm * b_acc
        df["clinch_effectiveness_diff"] = (
            df["r_clinch_effectiveness"] - df["b_clinch_effectiveness"]
        )

        r_dec = df.get("r_decision_win_rate", pd.Series(
            0.0, index=df.index)).fillna(0)
        b_dec = df.get("b_decision_win_rate", pd.Series(
            0.0, index=df.index)).fillna(0)
        total_rounds_col = df.get(
            "total_rounds", pd.Series(3.0, index=df.index)
        ).fillna(3)
        df["five_round_cardio_advantage"] = (
            r_dec - b_dec) * (total_rounds_col / 3)

        r_fights = df.get("r_pre_total_fights", pd.Series(10.0, index=df.index)).fillna(
            10
        )
        b_fights = df.get("b_pre_total_fights", pd.Series(10.0, index=df.index)).fillna(
            10
        )
        r_kd_absorbed = df.get(
            "r_pre_kd_absorbed", pd.Series(0.0, index=df.index)
        ).fillna(0)
        b_kd_absorbed = df.get(
            "b_pre_kd_absorbed", pd.Series(0.0, index=df.index)
        ).fillna(0)
        r_fights_safe = r_fights.clip(lower=1)
        b_fights_safe = b_fights.clip(lower=1)
        df["r_chin_deterioration"] = r_kd_absorbed / r_fights_safe
        df["b_chin_deterioration"] = b_kd_absorbed / b_fights_safe
        df["chin_deterioration_diff"] = (
            df["r_chin_deterioration"] - df["b_chin_deterioration"]
        )

        if "r_avg_opp_elo_L5" in df.columns and "b_avg_opp_elo_L5" in df.columns:
            df["opp_quality_diff"] = df["r_avg_opp_elo_L5"] - \
                df["b_avg_opp_elo_L5"]
        if "r_trajectory_3" in df.columns and "b_trajectory_3" in df.columns:
            df["trajectory_diff"] = df["r_trajectory_3"] - df["b_trajectory_3"]

        if self.svd_fitted:
            self._materialize_svd_columns(df)

        if "is_title_bout" in df.columns:
            df["is_title_enc"] = df["is_title_bout"].astype(int)
        if "total_rounds" in df.columns:
            df["total_rounds_num"] = pd.to_numeric(
                df["total_rounds"], errors="coerce"
            ).fillna(3)
        if "gender" in df.columns:
            df["gender_enc"] = (df["gender"].fillna("").str.lower() == "women").astype(
                int
            )

        for feat in [
            "distance_pct",
            "clinch_pct",
            "ground_pct",
            "head_pct",
            "body_pct",
            "leg_pct",
        ]:
            r_col, b_col = f"r_pre_{feat}", f"b_pre_{feat}"
            if r_col in df.columns and b_col in df.columns:
                df[f"diff_{feat}"] = df[r_col].fillna(0) - df[b_col].fillna(0)
        df["positional_striking_advantage"] = (
            df.get("diff_distance_pct", pd.Series(0.0, index=df.index))
            .fillna(0.0)
            .abs()
            + df.get("diff_clinch_pct", pd.Series(0.0, index=df.index))
            .fillna(0.0)
            .abs()
            + df.get("diff_ground_pct", pd.Series(0.0, index=df.index))
            .fillna(0.0)
            .abs()
        )
        df["target_distribution_advantage"] = (
            df.get("diff_head_pct", pd.Series(
                0.0, index=df.index)).fillna(0.0).abs()
            + df.get("diff_body_pct", pd.Series(0.0,
                     index=df.index)).fillna(0.0).abs()
            + df.get("diff_leg_pct", pd.Series(0.0,
                     index=df.index)).fillna(0.0).abs()
        )

        df["diff_str_def"] = df.get(
            "r_pre_str_def", pd.Series(0.0, index=df.index)
        ).fillna(0) - df.get("b_pre_str_def", pd.Series(0.0, index=df.index)).fillna(0)
        df["diff_td_def"] = df.get(
            "r_pre_td_def", pd.Series(0.0, index=df.index)
        ).fillna(0) - df.get("b_pre_td_def", pd.Series(0.0, index=df.index)).fillna(0)
        df["defensive_composite"] = df["diff_str_def"].fillna(0) + df[
            "diff_td_def"
        ].fillna(0)

        elo_d = df.get("elo_diff", pd.Series(0.0, index=df.index)).fillna(0)
        form3 = df.get("diff_rolling3_wins", pd.Series(
            0.0, index=df.index)).fillna(0)
        wlr = df.get("diff_win_loss_ratio", pd.Series(
            0.0, index=df.index)).fillna(0)
        fin_r = df.get("diff_finish_rate", pd.Series(
            0.0, index=df.index)).fillna(0)
        kd_abs = df.get("diff_kd_absorbed", pd.Series(
            0.0, index=df.index)).fillna(0)

        df["elo_x_form"] = elo_d * form3
        df["elo_x_win_ratio"] = elo_d * wlr
        df["elo_x_finish"] = elo_d * fin_r
        df["elo_x_durability"] = elo_d * kd_abs.abs()

        reach_d = df.get("diff_reach", pd.Series(
            0.0, index=df.index)).fillna(0)
        height_d = df.get("diff_height", pd.Series(
            0.0, index=df.index)).fillna(0)
        slpm_d = df.get("diff_pre_SLpM", pd.Series(
            0.0, index=df.index)).fillna(0)
        td_d = df.get("diff_pre_td_avg", pd.Series(
            0.0, index=df.index)).fillna(0)
        acc_d = df.get("diff_pre_sig_str_acc", pd.Series(
            0.0, index=df.index)).fillna(0)
        age_d = df.get("diff_age_at_event", pd.Series(
            0.0, index=df.index)).fillna(0)
        streak_d = df.get("diff_pre_win_streak", pd.Series(0.0, index=df.index)).fillna(
            0
        )
        exp_gap = df.get(
            "diff_pre_total_fights", pd.Series(0.0, index=df.index)
        ).fillna(0)

        df["reach_x_striking"] = reach_d * slpm_d
        df["height_x_reach"] = height_d * reach_d
        df["physical_x_striking"] = (height_d + reach_d) * slpm_d

        df["age_x_striking"] = age_d * slpm_d
        df["age_x_grappling"] = age_d * td_d
        df["age_x_durability"] = age_d * kd_abs.abs()
        df["age_x_win_streak"] = age_d * streak_d
        df["experience_x_age"] = exp_gap * age_d

        str_def_d = df.get("diff_str_def", pd.Series(
            0.0, index=df.index)).fillna(0)
        td_def_d = df.get("diff_td_def", pd.Series(
            0.0, index=df.index)).fillna(0)
        sub_rate_d = df.get("sub_threat_diff", pd.Series(
            0.0, index=df.index)).fillna(0)
        df["td_x_defense"] = td_d * td_def_d
        df["submission_x_grappling"] = sub_rate_d * td_d

        df["striking_x_accuracy"] = slpm_d * acc_d
        df["striking_x_defense"] = slpm_d * str_def_d
        df["ko_power_x_striking"] = (
            df.get("ko_threat_diff", pd.Series(
                0.0, index=df.index)).fillna(0) * slpm_d
        )

        momentum = df.get("momentum_diff_3", pd.Series(
            0.0, index=df.index)).fillna(0)
        df["momentum_x_win_streak"] = momentum * streak_d
        df["form_x_experience"] = form3 * exp_gap
        df["finish_x_momentum"] = fin_r * momentum
        df["form_x_durability"] = form3 * kd_abs.abs()

        df["elite_finisher"] = elo_d * fin_r * form3
        df["unstoppable_streak"] = streak_d * momentum * form3
        df["veteran_advantage"] = wlr * exp_gap * (-age_d)

        poly_extended = [
            "elo_diff",
            "glicko_diff",
            "diff_win_loss_ratio",
            "diff_age_at_event",
            "diff_reach",
            "diff_height",
            "diff_pre_SLpM",
            "diff_pre_sig_str_acc",
            "diff_pre_td_avg",
            "diff_pre_win_streak",
            "diff_finish_rate",
            "diff_pre_loss_streak",
            "diff_str_def",
            "diff_td_def",
            "diff_pre_kd_rate",
            "diff_pre_ctrl_avg",
            "elo_x_form",
            "streak_x_finish",
            "striking_exchange",
            "diff_distance_pct",
            "diff_clinch_pct",
            "diff_ground_pct",
        ]
        for col in poly_extended:
            if col in df.columns:
                df[f"{col}_sq"] = df[col] ** 2
                df[f"{col}_abs"] = df[col].abs()
        if "diff_age_at_event" in df.columns:
            df["diff_age_cubed"] = df["diff_age_at_event"] ** 3

        if "r_vs_elite_win_rate" in df.columns and "b_vs_elite_win_rate" in df.columns:
            df["diff_win_rate_vs_elite"] = (
                df["r_vs_elite_win_rate"] - df["b_vs_elite_win_rate"]
            )
            df["diff_win_rate_vs_strikers"] = df.get(
                "r_vs_striker_win_rate", pd.Series(0.0, index=df.index)
            ).fillna(0) - df.get(
                "b_vs_striker_win_rate", pd.Series(0.0, index=df.index)
            ).fillna(0)
            df["diff_win_rate_vs_grapplers"] = df.get(
                "r_vs_grappler_win_rate", pd.Series(0.0, index=df.index)
            ).fillna(0) - df.get(
                "b_vs_grappler_win_rate", pd.Series(0.0, index=df.index)
            ).fillna(0)
            df["championship_readiness"] = df["diff_win_rate_vs_elite"] * df.get(
                "elo_diff", pd.Series(0.0, index=df.index)
            ).fillna(0.0)

        if "r_pre_early_finish_rate" in df.columns:
            df["diff_early_finish_rate"] = df.get(
                "r_pre_early_finish_rate", pd.Series(0.0, index=df.index)
            ).fillna(0) - df.get(
                "b_pre_early_finish_rate", pd.Series(0.0, index=df.index)
            ).fillna(0)
            df["diff_late_finish_rate"] = df.get(
                "r_pre_late_finish_rate", pd.Series(0.0, index=df.index)
            ).fillna(0) - df.get(
                "b_pre_late_finish_rate", pd.Series(0.0, index=df.index)
            ).fillna(0)
            df["diff_first_round_ko_rate"] = df.get(
                "r_pre_first_round_ko_rate", pd.Series(0.0, index=df.index)
            ).fillna(0) - df.get(
                "b_pre_first_round_ko_rate", pd.Series(0.0, index=df.index)
            ).fillna(0)
        if "r_pre_five_round_fights" in df.columns:
            df["diff_five_round_fights"] = df.get(
                "r_pre_five_round_fights", pd.Series(0.0, index=df.index)
            ).fillna(0) - df.get(
                "b_pre_five_round_fights", pd.Series(0.0, index=df.index)
            ).fillna(0)

        if "r_fights_since_peak" in df.columns:
            df["diff_fights_since_peak"] = df.get(
                "r_fights_since_peak", pd.Series(0.0, index=df.index)
            ).fillna(0) - df.get(
                "b_fights_since_peak", pd.Series(0.0, index=df.index)
            ).fillna(0)
            df["declining_phase_diff"] = df["diff_fights_since_peak"]

        if "r_last_fight_was_win" in df.columns:
            df["r_last_fight_momentum"] = df.get(
                "r_last_fight_was_win", pd.Series(0.0, index=df.index)
            ).fillna(0).astype(float) + df.get(
                "r_last_fight_was_finish", pd.Series(0.0, index=df.index)
            ).fillna(0).astype(float)
            df["b_last_fight_momentum"] = df.get(
                "b_last_fight_was_win", pd.Series(0.0, index=df.index)
            ).fillna(0).astype(float) + df.get(
                "b_last_fight_was_finish", pd.Series(0.0, index=df.index)
            ).fillna(0).astype(float)
            df["last_fight_momentum_diff"] = (
                df["r_last_fight_momentum"] - df["b_last_fight_momentum"]
            )

        for feat in ["wins", "sig_str", "td", "kd", "finishes"]:
            r_col10 = f"r_rolling10_{feat}"
            b_col10 = f"b_rolling10_{feat}"
            if r_col10 in df.columns and b_col10 in df.columns:
                df[f"diff_rolling10_{feat}"] = df[r_col10].fillna(0) - df[
                    b_col10
                ].fillna(0)

        total_rds_t20 = df.get(
            "total_rounds_num", pd.Series(3.0, index=df.index)
        ).fillna(3)
        dec_rate_d_t20 = df.get(
            "dec_tendency_diff", pd.Series(0.0, index=df.index)
        ).fillna(0)
        if "diff_finish_rate" in df.columns:
            df["rounds_x_cardio"] = total_rds_t20 * dec_rate_d_t20
            df["rounds_x_finish_rate"] = (5 - total_rds_t20) * df[
                "diff_finish_rate"
            ].fillna(0)
        kd_abs2_t20 = df.get(
            "chin_deterioration_diff", pd.Series(0.0, index=df.index)
        ).fillna(0)
        df["rounds_x_durability"] = total_rds_t20 * kd_abs2_t20

        r_slpm_t21 = df.get("r_pre_SLpM", pd.Series(
            3.0, index=df.index)).fillna(3.0)
        b_slpm_t21 = df.get("b_pre_SLpM", pd.Series(
            3.0, index=df.index)).fillna(3.0)
        r_acc_t21 = df.get("r_pre_sig_str_acc", pd.Series(0.45, index=df.index)).fillna(
            0.45
        )
        b_acc_t21 = df.get("b_pre_sig_str_acc", pd.Series(0.45, index=df.index)).fillna(
            0.45
        )
        r_str_def_t21 = df.get("r_pre_str_def", pd.Series(0.55, index=df.index)).fillna(
            0.55
        )
        b_str_def_t21 = df.get("b_pre_str_def", pd.Series(0.55, index=df.index)).fillna(
            0.55
        )
        r_td_t21 = df.get("r_pre_td_avg", pd.Series(
            1.5, index=df.index)).fillna(1.5)
        b_td_t21 = df.get("b_pre_td_avg", pd.Series(
            1.5, index=df.index)).fillna(1.5)
        r_td_def_t21 = df.get("r_pre_td_def", pd.Series(0.65, index=df.index)).fillna(
            0.65
        )
        b_td_def_t21 = df.get("b_pre_td_def", pd.Series(0.65, index=df.index)).fillna(
            0.65
        )
        r_sub_avg_t21 = df.get(
            "r_pre_sub_att_rate", pd.Series(0.3, index=df.index)
        ).fillna(0.3)
        b_sub_avg_t21 = df.get(
            "b_pre_sub_att_rate", pd.Series(0.3, index=df.index)
        ).fillna(0.3)
        r_td_acc_t21 = df.get("r_pre_td_acc", pd.Series(0.4, index=df.index)).fillna(
            0.4
        )
        b_td_acc_t21 = df.get("b_pre_td_acc", pd.Series(0.4, index=df.index)).fillna(
            0.4
        )
        r_ctrl_t21 = df.get("r_pre_ctrl_avg", pd.Series(60.0, index=df.index)).fillna(
            60
        )
        b_ctrl_t21 = df.get("b_pre_ctrl_avg", pd.Series(60.0, index=df.index)).fillna(
            60
        )
        r_sub_rate_t21 = df.get(
            "r_pre_sub_att_rate", pd.Series(0.3, index=df.index)
        ).fillna(0.3)
        b_sub_rate_t21 = df.get(
            "b_pre_sub_att_rate", pd.Series(0.3, index=df.index)
        ).fillna(0.3)

        df["r_striking_vs_b_defense"] = r_slpm_t21 * (1.0 - b_str_def_t21)
        df["b_striking_vs_r_defense"] = b_slpm_t21 * (1.0 - r_str_def_t21)
        df["striking_exploitation_diff"] = (
            df["r_striking_vs_b_defense"] - df["b_striking_vs_r_defense"]
        )

        df["r_td_vs_b_td_defense"] = r_td_t21 * (1.0 - b_td_def_t21)
        df["b_td_vs_r_td_defense"] = b_td_t21 * (1.0 - r_td_def_t21)
        df["td_exploitation_diff"] = (
            df["r_td_vs_b_td_defense"] - df["b_td_vs_r_td_defense"]
        )

        df["r_sub_setup_efficiency"] = r_sub_rate_t21 * r_td_acc_t21
        df["b_sub_setup_efficiency"] = b_sub_rate_t21 * b_td_acc_t21
        df["sub_setup_diff"] = (
            df["r_sub_setup_efficiency"] - df["b_sub_setup_efficiency"]
        )
        df["r_sub_threat_vs_td_defense"] = r_sub_avg_t21 * (1.0 - b_td_def_t21)
        df["b_sub_threat_vs_td_defense"] = b_sub_avg_t21 * (1.0 - r_td_def_t21)
        df["sub_threat_vs_defense_diff"] = (
            df["r_sub_threat_vs_td_defense"] - df["b_sub_threat_vs_td_defense"]
        )

        df["r_striking_quality"] = r_slpm_t21 * r_acc_t21
        df["b_striking_quality"] = b_slpm_t21 * b_acc_t21
        df["striking_quality_diff"] = (
            df["r_striking_quality"] - df["b_striking_quality"]
        )
        df["r_accuracy_under_fire"] = r_acc_t21 / (b_slpm_t21 + 0.1)
        df["b_accuracy_under_fire"] = b_acc_t21 / (r_slpm_t21 + 0.1)
        df["accuracy_under_fire_diff"] = (
            df["r_accuracy_under_fire"] - df["b_accuracy_under_fire"]
        )

        r_sapm_t22 = df.get("r_pre_SApM", pd.Series(
            3.0, index=df.index)).fillna(3.0)
        b_sapm_t22 = df.get("b_pre_SApM", pd.Series(
            3.0, index=df.index)).fillna(3.0)

        df["r_damage_ratio"] = r_slpm_t21 / (r_sapm_t22 + 0.1)
        df["b_damage_ratio"] = b_slpm_t21 / (b_sapm_t22 + 0.1)
        df["damage_ratio_diff"] = df["r_damage_ratio"] - df["b_damage_ratio"]

        df["r_striking_output_quality"] = r_slpm_t21 * \
            r_acc_t21 / (r_sapm_t22 + 0.1)
        df["b_striking_output_quality"] = b_slpm_t21 * \
            b_acc_t21 / (b_sapm_t22 + 0.1)
        df["striking_output_quality_diff"] = (
            df["r_striking_output_quality"] - df["b_striking_output_quality"]
        )

        df["r_grappling_quality"] = r_td_t21 * \
            r_td_acc_t21 * (r_ctrl_t21 / 60.0)
        df["b_grappling_quality"] = b_td_t21 * \
            b_td_acc_t21 * (b_ctrl_t21 / 60.0)
        df["grappling_quality_diff"] = (
            df["r_grappling_quality"] - df["b_grappling_quality"]
        )

        df["r_total_defense_index"] = r_str_def_t21 * r_td_def_t21
        df["b_total_defense_index"] = b_str_def_t21 * b_td_def_t21
        df["total_defense_diff"] = (
            df["r_total_defense_index"] - df["b_total_defense_index"]
        )

        df["r_complete_fighter_index"] = (
            (r_slpm_t21 + r_td_t21 + r_sub_avg_t21) *
            r_str_def_t21 * r_td_def_t21
        )
        df["b_complete_fighter_index"] = (
            (b_slpm_t21 + b_td_t21 + b_sub_avg_t21) *
            b_str_def_t21 * b_td_def_t21
        )
        df["complete_fighter_diff"] = (
            df["r_complete_fighter_index"] - df["b_complete_fighter_index"]
        )

        df["r_pressure_index"] = r_slpm_t21 * r_td_t21 * (r_ctrl_t21 / 60.0)
        df["b_pressure_index"] = b_slpm_t21 * b_td_t21 * (b_ctrl_t21 / 60.0)
        df["pressure_index_diff"] = df["r_pressure_index"] - df["b_pressure_index"]

        r_wins_t23 = df.get("r_pre_wins", pd.Series(
            5.0, index=df.index)).fillna(5.0)
        b_wins_t23 = df.get("b_pre_wins", pd.Series(
            5.0, index=df.index)).fillna(5.0)
        r_losses_t23 = df.get("r_pre_losses", pd.Series(2.0, index=df.index)).fillna(
            2.0
        )
        b_losses_t23 = df.get("b_pre_losses", pd.Series(2.0, index=df.index)).fillna(
            2.0
        )
        r_streak_t23 = df.get(
            "r_pre_win_streak", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        b_streak_t23 = df.get(
            "b_pre_win_streak", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        r_fr_t23 = df.get("r_pre_finish_rate", pd.Series(0.4, index=df.index)).fillna(
            0.4
        )
        b_fr_t23 = df.get("b_pre_finish_rate", pd.Series(0.4, index=df.index)).fillna(
            0.4
        )
        r_age_t23 = df.get("r_age_at_event", pd.Series(28.0, index=df.index)).fillna(
            28.0
        )
        b_age_t23 = df.get("b_age_at_event", pd.Series(28.0, index=df.index)).fillna(
            28.0
        )
        r_reach_t23 = df.get("r_reach", pd.Series(
            71.0, index=df.index)).fillna(71.0)
        b_reach_t23 = df.get("b_reach", pd.Series(
            71.0, index=df.index)).fillna(71.0)
        r_weight_t23 = df.get("r_weight", pd.Series(155.0, index=df.index)).fillna(
            155.0
        )
        b_weight_t23 = df.get("b_weight", pd.Series(155.0, index=df.index)).fillna(
            155.0
        )
        r_tf_t23 = df.get("r_pre_total_fights", pd.Series(10.0, index=df.index)).fillna(
            10.0
        )
        b_tf_t23 = df.get("b_pre_total_fights", pd.Series(10.0, index=df.index)).fillna(
            10.0
        )
        df["r_defense_offense_balance"] = (
            r_str_def_t21 + 0.01) / (r_acc_t21 + 0.01)
        df["b_defense_offense_balance"] = (
            b_str_def_t21 + 0.01) / (b_acc_t21 + 0.01)
        df["defense_offense_balance_diff"] = (
            df["r_defense_offense_balance"] - df["b_defense_offense_balance"]
        )
        df["r_td_defense_offense_balance"] = (r_td_def_t21 + 0.01) / (
            r_td_acc_t21 + 0.01
        )
        df["b_td_defense_offense_balance"] = (b_td_def_t21 + 0.01) / (
            b_td_acc_t21 + 0.01
        )
        df["td_defense_offense_balance_diff"] = (
            df["r_td_defense_offense_balance"] -
            df["b_td_defense_offense_balance"]
        )
        df["finish_efficiency_diff"] = r_fr_t23 - b_fr_t23
        df["r_precision_striking"] = r_acc_t21 / (r_slpm_t21 + 0.1)
        df["b_precision_striking"] = b_acc_t21 / (b_slpm_t21 + 0.1)
        df["precision_striking_diff"] = (
            df["r_precision_striking"] - df["b_precision_striking"]
        )
        df["r_quality_grappling_23"] = r_td_acc_t21 * (r_td_t21**0.5)
        df["b_quality_grappling_23"] = b_td_acc_t21 * (b_td_t21**0.5)
        df["quality_grappling_diff"] = (
            df["r_quality_grappling_23"] - df["b_quality_grappling_23"]
        )
        df["r_submission_threat_ratio"] = (
            r_sub_avg_t21 + 0.01) / (r_td_t21 + 0.01)
        df["b_submission_threat_ratio"] = (
            b_sub_avg_t21 + 0.01) / (b_td_t21 + 0.01)
        df["submission_threat_ratio_diff"] = (
            df["r_submission_threat_ratio"] - df["b_submission_threat_ratio"]
        )
        df["r_damage_absorption_efficiency"] = r_sapm_t22 / \
            (r_str_def_t21 + 0.01)
        df["b_damage_absorption_efficiency"] = b_sapm_t22 / \
            (b_str_def_t21 + 0.01)
        df["damage_absorption_efficiency_diff"] = (
            df["r_damage_absorption_efficiency"] -
            df["b_damage_absorption_efficiency"]
        )
        df["r_defense_versatility"] = (r_str_def_t21 * r_td_def_t21) ** 0.5
        df["b_defense_versatility"] = (b_str_def_t21 * b_td_def_t21) ** 0.5
        df["defense_versatility_diff"] = (
            df["r_defense_versatility"] - df["b_defense_versatility"]
        )
        df["r_total_offense_index"] = r_slpm_t21 + (r_td_t21 * 1.5)
        df["b_total_offense_index"] = b_slpm_t21 + (b_td_t21 * 1.5)
        df["total_offense_index_diff"] = (
            df["r_total_offense_index"] - df["b_total_offense_index"]
        )
        df["r_offensive_versatility"] = (r_slpm_t21 * r_td_t21) ** 0.5
        df["b_offensive_versatility"] = (b_slpm_t21 * b_td_t21) ** 0.5
        df["offensive_versatility_diff"] = (
            df["r_offensive_versatility"] - df["b_offensive_versatility"]
        )
        df["r_striker_index"] = (r_slpm_t21 + 0.1) / (r_td_t21 + 0.1)
        df["b_striker_index"] = (b_slpm_t21 + 0.1) / (b_td_t21 + 0.1)
        df["striker_index_diff"] = df["r_striker_index"] - df["b_striker_index"]
        r_wlr_t23 = r_wins_t23 / (r_losses_t23 + 1.0)
        b_wlr_t23 = b_wins_t23 / (b_losses_t23 + 1.0)
        df["win_loss_ratio_squared_diff"] = (r_wlr_t23**2) - (b_wlr_t23**2)
        df["r_experience_quality"] = r_wins_t23 / \
            (r_wins_t23 + r_losses_t23 + 1.0)
        df["b_experience_quality"] = b_wins_t23 / \
            (b_wins_t23 + b_losses_t23 + 1.0)
        df["experience_quality_diff"] = (
            df["r_experience_quality"] - df["b_experience_quality"]
        )
        df["r_win_efficiency"] = r_wins_t23 / (r_age_t23 - 18.0 + 1.0)
        df["b_win_efficiency"] = b_wins_t23 / (b_age_t23 - 18.0 + 1.0)
        df["win_efficiency_diff"] = df["r_win_efficiency"] - df["b_win_efficiency"]
        df["r_momentum_quality"] = (r_streak_t23 + 1.0) / (r_wins_t23 + 1.0)
        df["b_momentum_quality"] = (b_streak_t23 + 1.0) / (b_wins_t23 + 1.0)
        df["momentum_quality_diff"] = (
            df["r_momentum_quality"] - df["b_momentum_quality"]
        )
        df["r_reach_efficiency"] = r_slpm_t21 / (r_reach_t23 + 1.0)
        df["b_reach_efficiency"] = b_slpm_t21 / (b_reach_t23 + 1.0)
        df["reach_efficiency_diff"] = (
            df["r_reach_efficiency"] - df["b_reach_efficiency"]
        )
        df["r_size_adjusted_striking"] = r_slpm_t21 / \
            ((r_weight_t23 / 100.0) + 0.01)
        df["b_size_adjusted_striking"] = b_slpm_t21 / \
            ((b_weight_t23 / 100.0) + 0.01)
        df["size_adjusted_striking_diff"] = (
            df["r_size_adjusted_striking"] - df["b_size_adjusted_striking"]
        )
        df["r_size_adjusted_grappling"] = r_td_t21 / \
            ((r_weight_t23 / 100.0) + 0.01)
        df["b_size_adjusted_grappling"] = b_td_t21 / \
            ((b_weight_t23 / 100.0) + 0.01)
        df["size_adjusted_grappling_diff"] = (
            df["r_size_adjusted_grappling"] - df["b_size_adjusted_grappling"]
        )
        df["r_counter_fighter_index"] = (
            r_str_def_t21 + 0.1) / (r_slpm_t21 + 1.0)
        df["b_counter_fighter_index"] = (
            b_str_def_t21 + 0.1) / (b_slpm_t21 + 1.0)
        df["counter_fighter_index_diff"] = (
            df["r_counter_fighter_index"] - df["b_counter_fighter_index"]
        )
        df["r_finishing_threat_composite"] = (
            r_fr_t23 + 0.1) * (r_sub_avg_t21 + 0.1)
        df["b_finishing_threat_composite"] = (
            b_fr_t23 + 0.1) * (b_sub_avg_t21 + 0.1)
        df["finishing_threat_composite_diff"] = (
            df["r_finishing_threat_composite"] -
            df["b_finishing_threat_composite"]
        )
        df["r_complete_geo"] = (
            (r_slpm_t21 + 1.0) * (r_str_def_t21 + 0.1) * (r_fr_t23 + 0.1)
        ) ** (1.0 / 3.0)
        df["b_complete_geo"] = (
            (b_slpm_t21 + 1.0) * (b_str_def_t21 + 0.1) * (b_fr_t23 + 0.1)
        ) ** (1.0 / 3.0)
        df["complete_geo_diff"] = df["r_complete_geo"] - df["b_complete_geo"]
        df["r_pressure_fighter_index"] = (
            r_slpm_t21 + r_td_t21) / (r_str_def_t21 + 0.3)
        df["b_pressure_fighter_index"] = (
            b_slpm_t21 + b_td_t21) / (b_str_def_t21 + 0.3)
        df["pressure_fighter_index_diff"] = (
            df["r_pressure_fighter_index"] - df["b_pressure_fighter_index"]
        )

        _r_roll3 = df.get("r_pre_rolling3_wins", pd.Series(1.5, index=df.index)).fillna(
            1.5
        )
        _b_roll3 = df.get("b_pre_rolling3_wins", pd.Series(1.5, index=df.index)).fillna(
            1.5
        )
        _r_cwr = r_wins_t23 / (r_tf_t23 + 1.0)
        _b_cwr = b_wins_t23 / (b_tf_t23 + 1.0)
        df["r_recent_form_ratio"] = (_r_roll3 / 3.0 + 0.01) / (_r_cwr + 0.01)
        df["b_recent_form_ratio"] = (_b_roll3 / 3.0 + 0.01) / (_b_cwr + 0.01)
        df["recent_form_ratio_diff"] = (
            df["r_recent_form_ratio"] - df["b_recent_form_ratio"]
        )

        _r_ko_d = df.get("r_pre_ko_wins", pd.Series(
            0.0, index=df.index)).fillna(0.0)
        _b_ko_d = df.get("b_pre_ko_wins", pd.Series(
            0.0, index=df.index)).fillna(0.0)
        _r_sub_d = df.get("r_pre_sub_wins", pd.Series(
            0.0, index=df.index)).fillna(0.0)
        _b_sub_d = df.get("b_pre_sub_wins", pd.Series(
            0.0, index=df.index)).fillna(0.0)
        _r_dec_d = df.get("r_pre_dec_wins", pd.Series(
            0.0, index=df.index)).fillna(0.0)
        _b_dec_d = df.get("b_pre_dec_wins", pd.Series(
            0.0, index=df.index)).fillna(0.0)
        df["r_finish_method_diversity"] = (
            (_r_ko_d > 0).astype(float)
            + (_r_sub_d > 0).astype(float)
            + (_r_dec_d > 0).astype(float)
        )
        df["b_finish_method_diversity"] = (
            (_b_ko_d > 0).astype(float)
            + (_b_sub_d > 0).astype(float)
            + (_b_dec_d > 0).astype(float)
        )
        df["finish_method_diversity_diff"] = (
            df["r_finish_method_diversity"] - df["b_finish_method_diversity"]
        )

        df["r_cross_domain_compensation"] = np.maximum(
            0.0, r_td_t21 - 1.5
        ) - np.maximum(0.0, 4.0 - r_slpm_t21)
        df["b_cross_domain_compensation"] = np.maximum(
            0.0, b_td_t21 - 1.5
        ) - np.maximum(0.0, 4.0 - b_slpm_t21)
        df["cross_domain_compensation_index_diff"] = (
            df["r_cross_domain_compensation"] -
            df["b_cross_domain_compensation"]
        )

        df["r_absorption_vuln"] = r_sapm_t22 / (b_slpm_t21 + 0.1)
        df["b_absorption_vuln"] = b_sapm_t22 / (r_slpm_t21 + 0.1)
        df["absorption_vulnerability_index_diff"] = (
            df["r_absorption_vuln"] - df["b_absorption_vuln"]
        )
        df["r_combined_def_hole"] = (
            1.0 - r_str_def_t21) * (1.0 - r_td_def_t21)
        df["b_combined_def_hole"] = (
            1.0 - b_str_def_t21) * (1.0 - b_td_def_t21)
        df["combined_defensive_hole_diff"] = (
            df["r_combined_def_hole"] - df["b_combined_def_hole"]
        )
        df["r_td_pressure_t24"] = (1.0 - r_td_def_t21) * b_td_t21
        df["b_td_pressure_t24"] = (1.0 - b_td_def_t21) * r_td_t21
        df["td_vulnerability_under_pressure_diff"] = (
            df["r_td_pressure_t24"] - df["b_td_pressure_t24"]
        )
        df["r_strike_pressure_t24"] = (1.0 - r_str_def_t21) * b_slpm_t21
        df["b_strike_pressure_t24"] = (1.0 - b_str_def_t21) * r_slpm_t21
        df["strike_defense_under_volume_diff"] = (
            df["r_strike_pressure_t24"] - df["b_strike_pressure_t24"]
        )
        df["r_ctrl_sub_ratio"] = (r_ctrl_t21 / 60.0) / (r_sub_avg_t21 + 0.1)
        df["b_ctrl_sub_ratio"] = (b_ctrl_t21 / 60.0) / (b_sub_avg_t21 + 0.1)
        df["grappling_control_vs_submission_ratio_diff"] = (
            df["r_ctrl_sub_ratio"] - df["b_ctrl_sub_ratio"]
        )
        df["r_sub_def_necessity"] = b_sub_avg_t21 / (r_td_def_t21 + 0.1)
        df["b_sub_def_necessity"] = r_sub_avg_t21 / (b_td_def_t21 + 0.1)
        df["submission_defense_necessity_diff"] = (
            df["r_sub_def_necessity"] - df["b_sub_def_necessity"]
        )
        df["r_strike_synergy"] = (r_slpm_t21 * (r_acc_t21 + 0.01)) ** 0.5
        df["b_strike_synergy"] = (b_slpm_t21 * (b_acc_t21 + 0.01)) ** 0.5
        df["striking_volume_accuracy_synergy_diff"] = (
            df["r_strike_synergy"] - df["b_strike_synergy"]
        )
        df["r_td_paradox"] = (r_td_acc_t21 + 0.01) / (r_td_t21 + 0.5)
        df["b_td_paradox"] = (b_td_acc_t21 + 0.01) / (b_td_t21 + 0.5)
        df["takedown_efficiency_paradox_diff"] = df["r_td_paradox"] - \
            df["b_td_paradox"]
        df["r_total_off_eff"] = (r_slpm_t21 * (r_acc_t21 + 0.01)) ** 0.5 + (
            r_td_t21 * (r_td_acc_t21 + 0.01)
        ) ** 0.5
        df["b_total_off_eff"] = (b_slpm_t21 * (b_acc_t21 + 0.01)) ** 0.5 + (
            b_td_t21 * (b_td_acc_t21 + 0.01)
        ) ** 0.5
        df["total_offensive_efficiency_index_diff"] = (
            df["r_total_off_eff"] - df["b_total_off_eff"]
        )
        df["r_sg_corr"] = (r_slpm_t21 * (r_acc_t21 + 0.01)) / (
            r_td_t21 * (r_td_acc_t21 + 0.01) + 0.1
        )
        df["b_sg_corr"] = (b_slpm_t21 * (b_acc_t21 + 0.01)) / (
            b_td_t21 * (b_td_acc_t21 + 0.01) + 0.1
        )
        df["striking_grappling_efficiency_correlation_diff"] = (
            df["r_sg_corr"] - df["b_sg_corr"]
        )
        df["r_def_allocation_balance"] = (r_str_def_t21 - r_td_def_t21).abs()
        df["b_def_allocation_balance"] = (b_str_def_t21 - b_td_def_t21).abs()
        df["defense_allocation_balance_diff"] = (
            df["r_def_allocation_balance"] - df["b_def_allocation_balance"]
        )
        _r_cbt = (
            (r_slpm_t21 / 10.0 + 0.01)
            * (r_acc_t21 + 0.01)
            * (10.0 / (r_sapm_t22 + 0.01))
            * (r_str_def_t21 + 0.01)
            * (r_td_t21 / 5.0 + 0.01)
            * (r_td_acc_t21 + 0.01)
            * (r_td_def_t21 + 0.01)
            * (r_sub_avg_t21 / 2.0 + 0.01)
        ) ** (1.0 / 8.0)
        _b_cbt = (
            (b_slpm_t21 / 10.0 + 0.01)
            * (b_acc_t21 + 0.01)
            * (10.0 / (b_sapm_t22 + 0.01))
            * (b_str_def_t21 + 0.01)
            * (b_td_t21 / 5.0 + 0.01)
            * (b_td_acc_t21 + 0.01)
            * (b_td_def_t21 + 0.01)
            * (b_sub_avg_t21 / 2.0 + 0.01)
        ) ** (1.0 / 8.0)
        df["r_combat_eff"] = _r_cbt
        df["b_combat_eff"] = _b_cbt
        df["total_combat_efficiency_index_diff"] = _r_cbt - _b_cbt

        df["net_striking_advantage"] = (r_slpm_t21 - b_slpm_t21) - (
            r_sapm_t22 - b_sapm_t22
        )
        df["striker_advantage"] = (
            r_slpm_t21 * r_acc_t21) - (b_slpm_t21 * b_acc_t21)
        df["grappler_advantage"] = (
            r_td_t21 * r_td_acc_t21) - (b_td_t21 * b_td_acc_t21)
        df["experience_gap"] = r_tf_t23 - b_tf_t23
        r_ko_wins_t25 = df.get("r_pre_ko_wins", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        b_ko_wins_t25 = df.get("b_pre_ko_wins", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        r_sub_wins_t25 = df.get(
            "r_pre_sub_wins", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        b_sub_wins_t25 = df.get(
            "b_pre_sub_wins", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        r_ko_rate_t25 = r_ko_wins_t25 / (r_tf_t23 + 1.0)
        b_ko_rate_t25 = b_ko_wins_t25 / (b_tf_t23 + 1.0)
        r_sub_rate_t25 = r_sub_wins_t25 / (r_tf_t23 + 1.0)
        b_sub_rate_t25 = b_sub_wins_t25 / (b_tf_t23 + 1.0)
        df["ko_specialist_gap"] = r_ko_rate_t25 - b_ko_rate_t25
        df["submission_specialist_gap"] = r_sub_rate_t25 - b_sub_rate_t25
        r_elo_t25 = df.get("elo_r", pd.Series(
            1500.0, index=df.index)).fillna(1500.0)
        b_elo_t25 = df.get("elo_b", pd.Series(
            1500.0, index=df.index)).fillna(1500.0)
        r_traj_t25 = df.get("r_trajectory_3", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        b_traj_t25 = df.get("b_trajectory_3", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        df["skill_momentum"] = (r_elo_t25 - b_elo_t25) * \
            (r_traj_t25 - b_traj_t25)
        r_loss_streak_t25 = df.get(
            "r_pre_loss_streak", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        b_loss_streak_t25 = df.get(
            "b_pre_loss_streak", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        r_win_rate_t25 = r_wins_t23 / (r_tf_t23 + 1.0)
        b_win_rate_t25 = b_wins_t23 / (b_tf_t23 + 1.0)
        df["r_desperation"] = r_loss_streak_t25 * \
            (1.0 / (r_win_rate_t25 + 0.1))
        df["b_desperation"] = b_loss_streak_t25 * \
            (1.0 / (b_win_rate_t25 + 0.1))
        df["desperation_diff"] = df["r_desperation"] - df["b_desperation"]
        r_days_t25 = df.get(
            "r_days_since_last", pd.Series(180.0, index=df.index)
        ).fillna(180.0)
        b_days_t25 = df.get(
            "b_days_since_last", pd.Series(180.0, index=df.index)
        ).fillna(180.0)
        df["r_freshness"] = np.exp(-((r_days_t25 - 135.0)
                                   ** 2) / (2.0 * 90.0**2))
        df["b_freshness"] = np.exp(-((b_days_t25 - 135.0)
                                   ** 2) / (2.0 * 90.0**2))
        df["freshness_advantage"] = df["r_freshness"] - df["b_freshness"]

        _r_st26 = (
            df.get("r_stance", pd.Series("", index=df.index))
            .fillna("")
            .astype(str)
            .str.strip()
            .str.lower()
        )
        _b_st26 = (
            df.get("b_stance", pd.Series("", index=df.index))
            .fillna("")
            .astype(str)
            .str.strip()
            .str.lower()
        )
        df["orthodox_vs_southpaw_advantage"] = np.where(
            (_r_st26 == "orthodox") & (_b_st26 == "southpaw"),
            1.0,
            np.where((_r_st26 == "southpaw") & (
                _b_st26 == "orthodox"), -1.0, 0.0),
        ).astype(float)
        df["orthodox_vs_switch_advantage"] = np.where(
            (_r_st26 == "orthodox") & (_b_st26 == "switch"),
            1.0,
            np.where((_r_st26 == "switch") & (
                _b_st26 == "orthodox"), -1.0, 0.0),
        ).astype(float)
        df["southpaw_vs_switch_advantage"] = np.where(
            (_r_st26 == "southpaw") & (_b_st26 == "switch"),
            1.0,
            np.where((_r_st26 == "switch") & (
                _b_st26 == "southpaw"), -1.0, 0.0),
        ).astype(float)
        df["mirror_matchup"] = (_r_st26 == _b_st26).astype(float)

        def _signed_sq_t27(s):
            return np.sign(s) * (s**2)

        for _feat_sq in [
            "net_striking_advantage",
            "striker_advantage",
            "grappler_advantage",
            "experience_gap",
            "ko_specialist_gap",
            "submission_specialist_gap",
            "skill_momentum",
            "desperation_diff",
            "freshness_advantage",
            "combined_defensive_hole_diff",
            "striking_volume_accuracy_synergy_diff",
            "total_offensive_efficiency_index_diff",
            "finish_efficiency_diff",
            "defense_versatility_diff",
            "offensive_versatility_diff",
        ]:
            if _feat_sq in df.columns:
                df[f"{_feat_sq}_sq"] = _signed_sq_t27(df[_feat_sq])

        r_fr_l5_t28 = df.get(
            "r_pre_finish_rate_l5", pd.Series(0.4, index=df.index)
        ).fillna(0.4)
        b_fr_l5_t28 = df.get(
            "b_pre_finish_rate_l5", pd.Series(0.4, index=df.index)
        ).fillna(0.4)
        r_fr_l10_t28 = df.get(
            "r_pre_finish_rate_l10", pd.Series(0.4, index=df.index)
        ).fillna(0.4)
        b_fr_l10_t28 = df.get(
            "b_pre_finish_rate_l10", pd.Series(0.4, index=df.index)
        ).fillna(0.4)
        df["r_finish_rate_accel"] = r_fr_l5_t28 - r_fr_l10_t28
        df["b_finish_rate_accel"] = b_fr_l5_t28 - b_fr_l10_t28
        df["finish_rate_acceleration_diff"] = (
            df["r_finish_rate_accel"] - df["b_finish_rate_accel"]
        )
        r_slpm_cv_t28 = df.get("r_pre_slpm_cv", pd.Series(0.3, index=df.index)).fillna(
            0.3
        )
        b_slpm_cv_t28 = df.get("b_pre_slpm_cv", pd.Series(0.3, index=df.index)).fillna(
            0.3
        )
        df["slpm_coefficient_of_variation_diff"] = r_slpm_cv_t28 - b_slpm_cv_t28
        r_mil_t28 = df.get(
            "r_pre_mileage_adj_age", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        b_mil_t28 = df.get(
            "b_pre_mileage_adj_age", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        df["mileage_adjusted_age_diff"] = r_mil_t28 - b_mil_t28
        df["performance_decline_velocity_diff"] = (
            df.get("r_trajectory_3", pd.Series(
                0.0, index=df.index)).fillna(0.0)
            - df.get("b_trajectory_3", pd.Series(0.0,
                     index=df.index)).fillna(0.0)
        ) * (-1.0)
        r_cur_elo_t28 = df.get("elo_r", pd.Series(1500.0, index=df.index)).fillna(
            1500.0
        )
        b_cur_elo_t28 = df.get("elo_b", pd.Series(1500.0, index=df.index)).fillna(
            1500.0
        )
        r_peak_t28 = df.get(
            "r_career_elo_peak", pd.Series(1500.0, index=df.index)
        ).fillna(1500.0)
        b_peak_t28 = df.get(
            "b_career_elo_peak", pd.Series(1500.0, index=df.index)
        ).fillna(1500.0)
        df["r_distance_from_peak"] = r_peak_t28 - r_cur_elo_t28
        df["b_distance_from_peak"] = b_peak_t28 - b_cur_elo_t28
        df["distance_from_career_peak_diff"] = (
            df["r_distance_from_peak"] - df["b_distance_from_peak"]
        )
        r_fsp_t28 = df.get(
            "r_fights_since_peak", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        b_fsp_t28 = df.get(
            "b_fights_since_peak", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        df["r_career_inflection"] = r_fsp_t28 / (r_tf_t23 + 1.0)
        df["b_career_inflection"] = b_fsp_t28 / (b_tf_t23 + 1.0)
        df["career_inflection_point_diff"] = (
            df["r_career_inflection"] - df["b_career_inflection"]
        )
        df["r_prime_exit_risk"] = (r_age_t23 > 33).astype(float) * np.clip(
            -r_traj_t25, 0.0, 1.0
        )
        df["b_prime_exit_risk"] = (b_age_t23 > 33).astype(float) * np.clip(
            -b_traj_t25, 0.0, 1.0
        )
        df["prime_exit_risk_diff"] = df["r_prime_exit_risk"] - \
            df["b_prime_exit_risk"]
        df["r_aging_power_penalty"] = (
            r_ko_rate_t25 * r_age_t23 * (r_age_t23 > 35).astype(float)
        )
        df["b_aging_power_penalty"] = (
            b_ko_rate_t25 * b_age_t23 * (b_age_t23 > 35).astype(float)
        )
        df["aging_power_striker_penalty_diff"] = (
            df["r_aging_power_penalty"] - df["b_aging_power_penalty"]
        )
        df["r_bayesian_finish"] = (r_ko_wins_t25 + r_sub_wins_t25 + 2.0) / (
            r_tf_t23 + 4.0
        )
        df["b_bayesian_finish"] = (b_ko_wins_t25 + b_sub_wins_t25 + 2.0) / (
            b_tf_t23 + 4.0
        )
        df["bayesian_finish_rate_diff"] = (
            df["r_bayesian_finish"] - df["b_bayesian_finish"]
        )
        df["r_layoff_veteran"] = r_days_t25 * r_tf_t23
        df["b_layoff_veteran"] = b_days_t25 * b_tf_t23
        df["layoff_veteran_interaction_diff"] = (
            df["r_layoff_veteran"] - df["b_layoff_veteran"]
        )
        df["r_elo_momentum"] = r_cur_elo_t28 * r_traj_t25
        df["b_elo_momentum"] = b_cur_elo_t28 * b_traj_t25
        df["elo_momentum_vs_competition_diff"] = (
            df["r_elo_momentum"] - df["b_elo_momentum"]
        )
        r_avg_opp_elo_t28 = df.get(
            "r_avg_opp_elo_L5", pd.Series(1500.0, index=df.index)
        ).fillna(1500.0)
        b_avg_opp_elo_t28 = df.get(
            "b_avg_opp_elo_L5", pd.Series(1500.0, index=df.index)
        ).fillna(1500.0)
        df["r_title_proximity"] = (
            r_streak_t23 * r_avg_opp_elo_t28 * r_cur_elo_t28 / 1.0e6
        )
        df["b_title_proximity"] = (
            b_streak_t23 * b_avg_opp_elo_t28 * b_cur_elo_t28 / 1.0e6
        )
        df["title_shot_proximity_score_diff"] = (
            df["r_title_proximity"] - df["b_title_proximity"]
        )
        df["r_elo_volatility"] = r_cur_elo_t28 * r_slpm_cv_t28
        df["b_elo_volatility"] = b_cur_elo_t28 * b_slpm_cv_t28
        df["elo_volatility_interaction_diff"] = (
            df["r_elo_volatility"] - df["b_elo_volatility"]
        )
        r_fin_l10_t28 = df.get(
            "r_rolling10_finishes", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        b_fin_l10_t28 = df.get(
            "b_rolling10_finishes", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        df["elite_performance_frequency_l10_diff"] = (r_fin_l10_t28 / 10.0) - (
            b_fin_l10_t28 / 10.0
        )
        _r_dr28 = df.get("r_damage_ratio", pd.Series(
            1.0, index=df.index)).fillna(1.0)
        _b_dr28 = df.get("b_damage_ratio", pd.Series(
            1.0, index=df.index)).fillna(1.0)
        df["r_conf_damage_ratio"] = _r_dr28 * \
            (1.0 - 1.0 / (r_tf_t23**0.5 + 1.0))
        df["b_conf_damage_ratio"] = _b_dr28 * \
            (1.0 - 1.0 / (b_tf_t23**0.5 + 1.0))
        df["confidence_weighted_damage_ratio_diff"] = (
            df["r_conf_damage_ratio"] - df["b_conf_damage_ratio"]
        )

        _r_r5slpm = df.get(
            "r_pre_rolling5_slpm", pd.Series(3.0, index=df.index)
        ).fillna(3.0)
        _b_r5slpm = df.get(
            "b_pre_rolling5_slpm", pd.Series(3.0, index=df.index)
        ).fillna(3.0)
        df["r_recent_vs_career_striking"] = _r_r5slpm / (r_slpm_t21 + 0.1)
        df["b_recent_vs_career_striking"] = _b_r5slpm / (b_slpm_t21 + 0.1)
        df["recent_vs_career_striking_diff"] = (
            df["r_recent_vs_career_striking"] -
            df["b_recent_vs_career_striking"]
        )

        _r_slpmstd = df.get(
            "r_pre_slpm_std_l10", pd.Series(1.0, index=df.index)
        ).fillna(1.0)
        _b_slpmstd = df.get(
            "b_pre_slpm_std_l10", pd.Series(1.0, index=df.index)
        ).fillna(1.0)
        df["r_striking_consistency_ratio"] = 1.0 / (_r_slpmstd + 0.1)
        df["b_striking_consistency_ratio"] = 1.0 / (_b_slpmstd + 0.1)
        df["striking_consistency_ratio_diff"] = (
            df["r_striking_consistency_ratio"] -
            df["b_striking_consistency_ratio"]
        )

        _r_drstd = df.get(
            "r_pre_damage_ratio_std_l10", pd.Series(0.3, index=df.index)
        ).fillna(0.3)
        _b_drstd = df.get(
            "b_pre_damage_ratio_std_l10", pd.Series(0.3, index=df.index)
        ).fillna(0.3)
        df["performance_volatility_l10_diff"] = _r_drstd - _b_drstd

        _r_tact = df.get(
            "r_pre_tactical_evolution", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        _b_tact = df.get(
            "b_pre_tactical_evolution", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        df["tactical_evolution_score_diff"] = _r_tact - _b_tact

        r_daniel_t29 = (r_slpm_t21 * r_acc_t21) / (
            r_sapm_t22 * r_str_def_t21 + 0.01
        ) + (r_td_t21 * r_td_acc_t21 * r_td_def_t21 * r_sub_avg_t21)
        b_daniel_t29 = (b_slpm_t21 * b_acc_t21) / (
            b_sapm_t22 * b_str_def_t21 + 0.01
        ) + (b_td_t21 * b_td_acc_t21 * b_td_def_t21 * b_sub_avg_t21)
        df["r_daniel_custom"] = r_daniel_t29
        df["b_daniel_custom"] = b_daniel_t29
        df["daniel_custom_feature_diff"] = r_daniel_t29 - b_daniel_t29

        r_glicko_rd_t29 = df.get(
            "r_glicko_rd", pd.Series(200.0, index=df.index)
        ).fillna(200.0)
        b_glicko_rd_t29 = df.get(
            "b_glicko_rd", pd.Series(200.0, index=df.index)
        ).fillna(200.0)
        df["r_elo_confidence_weighted"] = r_cur_elo_t28 * np.clip(
            1.0 - r_glicko_rd_t29 / 500.0, 0.2, 1.0
        )
        df["b_elo_confidence_weighted"] = b_cur_elo_t28 * np.clip(
            1.0 - b_glicko_rd_t29 / 500.0, 0.2, 1.0
        )
        df["elo_confidence_weighted_diff"] = (
            df["r_elo_confidence_weighted"] - df["b_elo_confidence_weighted"]
        )

        df["r_streak_quality"] = r_streak_t23 * r_avg_opp_elo_t28 / 1000.0
        df["b_streak_quality"] = b_streak_t23 * b_avg_opp_elo_t28 / 1000.0
        df["streak_quality_diff"] = df["r_streak_quality"] - df["b_streak_quality"]

        df["r_reach_defense"] = r_reach_t23 * r_str_def_t21
        df["b_reach_defense"] = b_reach_t23 * b_str_def_t21
        df["reach_defense_diff"] = df["r_reach_defense"] - df["b_reach_defense"]

        df["r_net_striking_efficiency"] = (r_slpm_t21 * r_acc_t21) - (
            r_sapm_t22 * (1.0 - r_str_def_t21)
        )
        df["b_net_striking_efficiency"] = (b_slpm_t21 * b_acc_t21) - (
            b_sapm_t22 * (1.0 - b_str_def_t21)
        )
        df["net_striking_efficiency_diff"] = (
            df["r_net_striking_efficiency"] - df["b_net_striking_efficiency"]
        )

        r_glicko_r_t30 = df.get("r_glicko_r", pd.Series(1500.0, index=df.index)).fillna(
            1500.0
        )
        b_glicko_r_t30 = df.get("b_glicko_r", pd.Series(1500.0, index=df.index)).fillna(
            1500.0
        )
        r_swvc_t30 = df.get(
            "r_style_win_vs_cluster", pd.Series(0.5, index=df.index)
        ).fillna(0.5)
        b_swvc_t30 = df.get(
            "b_style_win_vs_cluster", pd.Series(0.5, index=df.index)
        ).fillna(0.5)
        _common_opp_t30 = df.get(
            "common_opp_edge", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        _style_edge_t30 = df.get(
            "style_matchup_edge", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        r_ape_t30 = pd.to_numeric(
            df.get("r_ape_index", pd.Series(0.0, index=df.index)), errors="coerce"
        ).fillna(0.0)
        b_ape_t30 = pd.to_numeric(
            df.get("b_ape_index", pd.Series(0.0, index=df.index)), errors="coerce"
        ).fillna(0.0)
        r_peak_t30 = df.get("r_peak_score", pd.Series(
            0.5, index=df.index)).fillna(0.5)
        b_peak_t30 = df.get("b_peak_score", pd.Series(
            0.5, index=df.index)).fillna(0.5)
        _fresh_r_t30 = df.get("r_freshness", pd.Series(
            0.5, index=df.index)).fillna(0.5)
        _fresh_b_t30 = df.get("b_freshness", pd.Series(
            0.5, index=df.index)).fillna(0.5)
        _press_r_t30 = df.get(
            "r_pressure_index", pd.Series(5.0, index=df.index)
        ).fillna(5.0)
        _press_b_t30 = df.get(
            "b_pressure_index", pd.Series(5.0, index=df.index)
        ).fillna(5.0)

        r_certainty_t30 = np.clip(1.0 - r_glicko_rd_t29 / 400.0, 0.25, 1.0)
        b_certainty_t30 = np.clip(1.0 - b_glicko_rd_t29 / 400.0, 0.25, 1.0)
        r_prime_t30 = np.clip(1.0 - np.abs(r_age_t23 - 29.0) / 12.0, 0.1, 1.0)
        b_prime_t30 = np.clip(1.0 - np.abs(b_age_t23 - 29.0) / 12.0, 0.1, 1.0)

        df["r_style_glicko_prime"] = (
            r_swvc_t30 * (r_glicko_r_t30 / 1500.0) *
            r_prime_t30 * r_certainty_t30
        )
        df["b_style_glicko_prime"] = (
            b_swvc_t30 * (b_glicko_r_t30 / 1500.0) *
            b_prime_t30 * b_certainty_t30
        )
        df["style_glicko_prime_diff"] = (
            df["r_style_glicko_prime"] - df["b_style_glicko_prime"]
        )

        df["common_style_consensus"] = _common_opp_t30 * _style_edge_t30

        r_youth_t30 = np.clip((35.0 - r_age_t23) / 15.0, 0.0, 1.0)
        b_youth_t30 = np.clip((35.0 - b_age_t23) / 15.0, 0.0, 1.0)
        df["r_age_elo_climb"] = (
            (r_cur_elo_t28 / 1500.0) *
            np.clip(r_traj_t25, -1.0, 1.0) * r_youth_t30
        )
        df["b_age_elo_climb"] = (
            (b_cur_elo_t28 / 1500.0) *
            np.clip(b_traj_t25, -1.0, 1.0) * b_youth_t30
        )
        df["age_elo_climb_diff"] = df["r_age_elo_climb"] - df["b_age_elo_climb"]

        r_expq_t30 = r_wins_t23 / (r_wins_t23 + r_losses_t23 + 1.0)
        b_expq_t30 = b_wins_t23 / (b_wins_t23 + b_losses_t23 + 1.0)
        df["r_quality_battle_tested"] = (
            r_expq_t30 * (r_avg_opp_elo_t28 / 1500.0) * np.log1p(r_tf_t23)
        )
        df["b_quality_battle_tested"] = (
            b_expq_t30 * (b_avg_opp_elo_t28 / 1500.0) * np.log1p(b_tf_t23)
        )
        df["quality_battle_tested_diff"] = (
            df["r_quality_battle_tested"] - df["b_quality_battle_tested"]
        )

        df["r_physical_style_dominance"] = r_ape_t30 * \
            r_swvc_t30 * (r_reach_t23 / 71.0)
        df["b_physical_style_dominance"] = b_ape_t30 * \
            b_swvc_t30 * (b_reach_t23 / 71.0)
        df["physical_style_dominance_diff"] = (
            df["r_physical_style_dominance"] - df["b_physical_style_dominance"]
        )

        df["r_peak_fresh_pressure"] = r_peak_t30 * \
            _fresh_r_t30 * (_press_r_t30 / 10.0)
        df["b_peak_fresh_pressure"] = b_peak_t30 * \
            _fresh_b_t30 * (_press_b_t30 / 10.0)
        df["peak_fresh_pressure_diff"] = (
            df["r_peak_fresh_pressure"] - df["b_peak_fresh_pressure"]
        )

        _elo_sig_t30 = (r_cur_elo_t28 - b_cur_elo_t28) / 200.0
        _glicko_sig_t30 = (r_glicko_r_t30 - b_glicko_r_t30) / 200.0
        df["tri_rating_consensus"] = _elo_sig_t30 * \
            _glicko_sig_t30 * _style_edge_t30

        z_r_kd_t31 = df.get("z_r_pre_kd_rate", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        z_b_kd_t31 = df.get("z_b_pre_kd_rate", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        z_r_acc_t31 = df.get(
            "z_r_pre_sig_str_acc", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        z_b_acc_t31 = df.get(
            "z_b_pre_sig_str_acc", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        z_r_slpm_t31 = df.get("z_r_pre_SLpM", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        z_b_slpm_t31 = df.get("z_b_pre_SLpM", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        z_r_sub_t31 = df.get(
            "z_r_pre_sub_att_rate", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        z_b_sub_t31 = df.get(
            "z_b_pre_sub_att_rate", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        form_svd3_t31 = df.get("form_svd_3", pd.Series(
            0.0, index=df.index)).fillna(0.0)
        str_svd3_t31 = df.get("striking_svd_3", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        grp_svd2_t31 = df.get("grappling_svd_2", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        r_cbt_t31 = df.get("r_combat_eff", pd.Series(
            0.5, index=df.index)).fillna(0.5)
        b_cbt_t31 = df.get("b_combat_eff", pd.Series(
            0.5, index=df.index)).fillna(0.5)
        r_glicko_vol_t31 = df.get(
            "r_glicko_vol", pd.Series(0.06, index=df.index)
        ).fillna(0.06)
        b_glicko_vol_t31 = df.get(
            "b_glicko_vol", pd.Series(0.06, index=df.index)
        ).fillna(0.06)

        df["r_wc_ko_style_physical"] = z_r_kd_t31 * \
            r_swvc_t30 * r_ape_t30 / 5.0
        df["b_wc_ko_style_physical"] = z_b_kd_t31 * \
            b_swvc_t30 * b_ape_t30 / 5.0
        df["wc_ko_style_physical_diff"] = (
            df["r_wc_ko_style_physical"] - df["b_wc_ko_style_physical"]
        )

        df["r_wc_acc_physical_style"] = z_r_acc_t31 * \
            r_ape_t30 / 5.0 * r_swvc_t30
        df["b_wc_acc_physical_style"] = z_b_acc_t31 * \
            b_ape_t30 / 5.0 * b_swvc_t30
        df["wc_acc_physical_style_diff"] = (
            df["r_wc_acc_physical_style"] - df["b_wc_acc_physical_style"]
        )

        df["r_wc_striking_dominance"] = z_r_slpm_t31 * z_r_acc_t31
        df["b_wc_striking_dominance"] = z_b_slpm_t31 * z_b_acc_t31
        df["wc_striking_dominance_diff"] = (
            df["r_wc_striking_dominance"] - df["b_wc_striking_dominance"]
        )

        df["r_wc_sub_threat_compound"] = z_r_sub_t31 * r_td_def_t21 * r_swvc_t30
        df["b_wc_sub_threat_compound"] = z_b_sub_t31 * b_td_def_t21 * b_swvc_t30
        df["wc_sub_threat_compound_diff"] = (
            df["r_wc_sub_threat_compound"] - df["b_wc_sub_threat_compound"]
        )

        df["form_svd_style_synergy"] = form_svd3_t31 * _style_edge_t30
        df["striking_svd_physical"] = str_svd3_t31 * \
            (r_ape_t30 - b_ape_t30) / 5.0
        df["grappling_svd_style_synergy"] = grp_svd2_t31 * _style_edge_t30

        r_vol_pen_t31 = np.clip(1.0 - r_glicko_vol_t31 * 5.0, 0.1, 1.0)
        b_vol_pen_t31 = np.clip(1.0 - b_glicko_vol_t31 * 5.0, 0.1, 1.0)
        df["r_stable_prime_combat"] = r_cbt_t31 * r_vol_pen_t31 * r_prime_t30
        df["b_stable_prime_combat"] = b_cbt_t31 * b_vol_pen_t31 * b_prime_t30
        df["stable_prime_combat_diff"] = (
            df["r_stable_prime_combat"] - df["b_stable_prime_combat"]
        )

        df["r_physical_elo_youth"] = (
            r_ape_t30 * (r_reach_t23 / 71.0) *
            (r_cur_elo_t28 / 1500.0) * r_youth_t30
        )
        df["b_physical_elo_youth"] = (
            b_ape_t30 * (b_reach_t23 / 71.0) *
            (b_cur_elo_t28 / 1500.0) * b_youth_t30
        )
        df["physical_elo_youth_diff"] = (
            df["r_physical_elo_youth"] - df["b_physical_elo_youth"]
        )

        z_r_td_t32 = df.get("z_r_pre_td_avg", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        z_b_td_t32 = df.get("z_b_pre_td_avg", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )

        df["r_physical_glicko_youth"] = (
            r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
            * r_youth_t30
        )
        df["b_physical_glicko_youth"] = (
            b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
            * b_youth_t30
        )
        df["physical_glicko_youth_diff"] = (
            df["r_physical_glicko_youth"] - df["b_physical_glicko_youth"]
        )

        df["r_physical_style_glicko_prime"] = (
            r_ape_t30
            * r_swvc_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
            * r_prime_t30
        )
        df["b_physical_style_glicko_prime"] = (
            b_ape_t30
            * b_swvc_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
            * b_prime_t30
        )
        df["physical_style_glicko_prime_diff"] = (
            df["r_physical_style_glicko_prime"] -
            df["b_physical_style_glicko_prime"]
        )

        df["r_physical_combat_prime"] = (
            r_ape_t30 * (r_reach_t23 / 71.0) * r_cbt_t31 * r_prime_t30
        )
        df["b_physical_combat_prime"] = (
            b_ape_t30 * (b_reach_t23 / 71.0) * b_cbt_t31 * b_prime_t30
        )
        df["physical_combat_prime_diff"] = (
            df["r_physical_combat_prime"] - df["b_physical_combat_prime"]
        )

        df["r_fresh_physical_elo"] = (
            _fresh_r_t30 * r_ape_t30 *
            (r_reach_t23 / 71.0) * (r_cur_elo_t28 / 1500.0)
        )
        df["b_fresh_physical_elo"] = (
            _fresh_b_t30 * b_ape_t30 *
            (b_reach_t23 / 71.0) * (b_cur_elo_t28 / 1500.0)
        )
        df["fresh_physical_elo_diff"] = (
            df["r_fresh_physical_elo"] - df["b_fresh_physical_elo"]
        )

        df["grappling_svd_physical"] = grp_svd2_t31 * \
            (r_ape_t30 - b_ape_t30) / 5.0

        df["r_wc_td_physical_style"] = z_r_td_t32 * \
            r_ape_t30 / 5.0 * r_swvc_t30
        df["b_wc_td_physical_style"] = z_b_td_t32 * \
            b_ape_t30 / 5.0 * b_swvc_t30
        df["wc_td_physical_style_diff"] = (
            df["r_wc_td_physical_style"] - df["b_wc_td_physical_style"]
        )

        z_r_sapm_t33 = df.get("z_r_pre_SApM", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        z_b_sapm_t33 = df.get("z_b_pre_SApM", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        form_svd1_t33 = df.get("form_svd_1", pd.Series(
            0.0, index=df.index)).fillna(0.0)
        form_svd0_t33 = df.get("form_svd_0", pd.Series(
            0.0, index=df.index)).fillna(0.0)
        str_svd0_t33 = df.get("striking_svd_0", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        grp_svd3_t33 = df.get("grappling_svd_3", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )

        df["r_physical_defensive"] = (
            r_ape_t30 * (r_reach_t23 / 71.0) * np.clip(-z_r_sapm_t33, 0.0, 3.0)
        )
        df["b_physical_defensive"] = (
            b_ape_t30 * (b_reach_t23 / 71.0) * np.clip(-z_b_sapm_t33, 0.0, 3.0)
        )
        df["physical_defensive_diff"] = (
            df["r_physical_defensive"] - df["b_physical_defensive"]
        )

        df["r_fresh_style_glicko"] = (
            _fresh_r_t30 * r_swvc_t30 *
            (r_glicko_r_t30 / 1500.0) * r_certainty_t30
        )
        df["b_fresh_style_glicko"] = (
            _fresh_b_t30 * b_swvc_t30 *
            (b_glicko_r_t30 / 1500.0) * b_certainty_t30
        )
        df["fresh_style_glicko_diff"] = (
            df["r_fresh_style_glicko"] - df["b_fresh_style_glicko"]
        )

        df["r_physical_glicko_prime"] = (
            r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
            * r_prime_t30
        )
        df["b_physical_glicko_prime"] = (
            b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
            * b_prime_t30
        )
        df["physical_glicko_prime_diff"] = (
            df["r_physical_glicko_prime"] - df["b_physical_glicko_prime"]
        )

        df["form_svd_physical"] = form_svd3_t31 * (r_ape_t30 - b_ape_t30) / 5.0
        df["form_svd1_style_glicko"] = (
            form_svd1_t33 * _style_edge_t30 *
            (r_glicko_r_t30 - b_glicko_r_t30) / 200.0
        )
        df["striking_svd0_physical"] = str_svd0_t33 * \
            (r_ape_t30 - b_ape_t30) / 5.0
        df["form_svd0_physical"] = form_svd0_t33 * \
            (r_ape_t30 - b_ape_t30) / 5.0
        df["grappling_svd3_style"] = grp_svd3_t33 * _style_edge_t30

        r_tp_t34 = df.get("r_title_proximity", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        b_tp_t34 = df.get("b_title_proximity", pd.Series(0.0, index=df.index)).fillna(
            0.0
        )
        r_ko_rate_t34 = _r_ko_d / (r_tf_t23 + 1.0)
        b_ko_rate_t34 = _b_ko_d / (b_tf_t23 + 1.0)
        r_sub_rate_t34 = _r_sub_d / (r_tf_t23 + 1.0)
        b_sub_rate_t34 = _b_sub_d / (b_tf_t23 + 1.0)

        df["r_winrate_physical_glicko"] = (
            r_expq_t30
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
        )
        df["b_winrate_physical_glicko"] = (
            b_expq_t30
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
        )
        df["winrate_physical_glicko_diff"] = (
            df["r_winrate_physical_glicko"] - df["b_winrate_physical_glicko"]
        )

        df["r_finish_physical_youth"] = (
            r_fr_t23 * r_ape_t30 * (r_reach_t23 / 71.0) * r_youth_t30
        )
        df["b_finish_physical_youth"] = (
            b_fr_t23 * b_ape_t30 * (b_reach_t23 / 71.0) * b_youth_t30
        )
        df["finish_physical_youth_diff"] = (
            df["r_finish_physical_youth"] - df["b_finish_physical_youth"]
        )

        df["r_elite_winrate_physical"] = (
            r_expq_t30 * (r_avg_opp_elo_t28 / 1500.0) *
            r_ape_t30 * (r_reach_t23 / 71.0)
        )
        df["b_elite_winrate_physical"] = (
            b_expq_t30 * (b_avg_opp_elo_t28 / 1500.0) *
            b_ape_t30 * (b_reach_t23 / 71.0)
        )
        df["elite_winrate_physical_diff"] = (
            df["r_elite_winrate_physical"] - df["b_elite_winrate_physical"]
        )

        df["r_ko_rate_physical_prime"] = (
            r_ko_rate_t34 * r_ape_t30 * (r_reach_t23 / 71.0) * r_prime_t30
        )
        df["b_ko_rate_physical_prime"] = (
            b_ko_rate_t34 * b_ape_t30 * (b_reach_t23 / 71.0) * b_prime_t30
        )
        df["ko_rate_physical_prime_diff"] = (
            df["r_ko_rate_physical_prime"] - df["b_ko_rate_physical_prime"]
        )

        df["r_winrate_style_glicko_prime"] = (
            r_expq_t30
            * r_swvc_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
            * r_prime_t30
        )
        df["b_winrate_style_glicko_prime"] = (
            b_expq_t30
            * b_swvc_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
            * b_prime_t30
        )
        df["winrate_style_glicko_prime_diff"] = (
            df["r_winrate_style_glicko_prime"] -
            df["b_winrate_style_glicko_prime"]
        )

        df["r_streak_physical_glicko"] = (
            np.clip(r_streak_t23 / 5.0, 0.0, 1.0)
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
        )
        df["b_streak_physical_glicko"] = (
            np.clip(b_streak_t23 / 5.0, 0.0, 1.0)
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
        )
        df["streak_physical_glicko_diff"] = (
            df["r_streak_physical_glicko"] - df["b_streak_physical_glicko"]
        )

        df["r_title_physical_glicko"] = r_tp_t34 * \
            r_ape_t30 * (r_glicko_r_t30 / 1500.0)
        df["b_title_physical_glicko"] = b_tp_t34 * \
            b_ape_t30 * (b_glicko_r_t30 / 1500.0)
        df["title_physical_glicko_diff"] = (
            df["r_title_physical_glicko"] - df["b_title_physical_glicko"]
        )

        df["r_recent_winrate_physical_youth"] = (
            (_r_roll3 / 3.0) * r_ape_t30 * (r_reach_t23 / 71.0) * r_youth_t30
        )
        df["b_recent_winrate_physical_youth"] = (
            (_b_roll3 / 3.0) * b_ape_t30 * (b_reach_t23 / 71.0) * b_youth_t30
        )
        df["recent_winrate_physical_youth_diff"] = (
            df["r_recent_winrate_physical_youth"]
            - df["b_recent_winrate_physical_youth"]
        )

        df["r_sub_rate_physical_youth"] = r_sub_rate_t34 * r_ape_t30 * r_youth_t30
        df["b_sub_rate_physical_youth"] = b_sub_rate_t34 * b_ape_t30 * b_youth_t30
        df["sub_rate_physical_youth_diff"] = (
            df["r_sub_rate_physical_youth"] - df["b_sub_rate_physical_youth"]
        )

        df["r_winrate_physical_glicko_youth"] = (
            r_expq_t30
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
            * r_youth_t30
        )
        df["b_winrate_physical_glicko_youth"] = (
            b_expq_t30
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
            * b_youth_t30
        )
        df["winrate_physical_glicko_youth_diff"] = (
            df["r_winrate_physical_glicko_youth"]
            - df["b_winrate_physical_glicko_youth"]
        )

        r_mom_t35 = df.get("r_momentum_quality", pd.Series(0.2, index=df.index)).fillna(
            0.2
        )
        b_mom_t35 = df.get("b_momentum_quality", pd.Series(0.2, index=df.index)).fillna(
            0.2
        )

        df["r_winrate_physical_style_glicko_prime"] = (
            r_expq_t30
            * r_ape_t30
            * r_swvc_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
            * r_prime_t30
        )
        df["b_winrate_physical_style_glicko_prime"] = (
            b_expq_t30
            * b_ape_t30
            * b_swvc_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
            * b_prime_t30
        )
        df["winrate_physical_style_glicko_prime_diff"] = (
            df["r_winrate_physical_style_glicko_prime"]
            - df["b_winrate_physical_style_glicko_prime"]
        )

        df["r_recent_physical_glicko_youth"] = (
            (_r_roll3 / 3.0)
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
            * r_youth_t30
        )
        df["b_recent_physical_glicko_youth"] = (
            (_b_roll3 / 3.0)
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
            * b_youth_t30
        )
        df["recent_physical_glicko_youth_diff"] = (
            df["r_recent_physical_glicko_youth"] -
            df["b_recent_physical_glicko_youth"]
        )

        df["r_opp_quality_physical_glicko"] = (
            (r_avg_opp_elo_t28 / 1500.0)
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
        )
        df["b_opp_quality_physical_glicko"] = (
            (b_avg_opp_elo_t28 / 1500.0)
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
        )
        df["opp_quality_physical_glicko_diff"] = (
            df["r_opp_quality_physical_glicko"] -
            df["b_opp_quality_physical_glicko"]
        )

        df["r_momentum_physical_glicko"] = (
            r_mom_t35 * r_ape_t30 * (r_glicko_r_t30 / 1500.0) * r_certainty_t30
        )
        df["b_momentum_physical_glicko"] = (
            b_mom_t35 * b_ape_t30 * (b_glicko_r_t30 / 1500.0) * b_certainty_t30
        )
        df["momentum_physical_glicko_diff"] = (
            df["r_momentum_physical_glicko"] - df["b_momentum_physical_glicko"]
        )

        df["r_finish_physical_glicko_prime"] = (
            r_fr_t23
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_prime_t30
        )
        df["b_finish_physical_glicko_prime"] = (
            b_fr_t23
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_prime_t30
        )
        df["finish_physical_glicko_prime_diff"] = (
            df["r_finish_physical_glicko_prime"] -
            df["b_finish_physical_glicko_prime"]
        )

        df["r_winrate_style_physical_youth"] = (
            r_expq_t30 * r_swvc_t30 * r_ape_t30 * r_youth_t30
        )
        df["b_winrate_style_physical_youth"] = (
            b_expq_t30 * b_swvc_t30 * b_ape_t30 * b_youth_t30
        )
        df["winrate_style_physical_youth_diff"] = (
            df["r_winrate_style_physical_youth"] -
            df["b_winrate_style_physical_youth"]
        )

        df["r_streak_winrate_physical_glicko"] = (
            np.clip(r_streak_t23 / 5.0, 0.0, 1.0)
            * r_expq_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
        )
        df["b_streak_winrate_physical_glicko"] = (
            np.clip(b_streak_t23 / 5.0, 0.0, 1.0)
            * b_expq_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
        )
        df["streak_winrate_physical_glicko_diff"] = (
            df["r_streak_winrate_physical_glicko"]
            - df["b_streak_winrate_physical_glicko"]
        )

        df["r_opp_quality_winrate_physical_youth"] = (
            (r_avg_opp_elo_t28 / 1500.0) * r_expq_t30 * r_ape_t30 * r_youth_t30
        )
        df["b_opp_quality_winrate_physical_youth"] = (
            (b_avg_opp_elo_t28 / 1500.0) * b_expq_t30 * b_ape_t30 * b_youth_t30
        )
        df["opp_quality_winrate_physical_youth_diff"] = (
            df["r_opp_quality_winrate_physical_youth"]
            - df["b_opp_quality_winrate_physical_youth"]
        )

        df["r_title_physical_glicko_youth"] = (
            r_tp_t34 * r_ape_t30 * (r_glicko_r_t30 / 1500.0) * r_youth_t30
        )
        df["b_title_physical_glicko_youth"] = (
            b_tp_t34 * b_ape_t30 * (b_glicko_r_t30 / 1500.0) * b_youth_t30
        )
        df["title_physical_glicko_youth_diff"] = (
            df["r_title_physical_glicko_youth"] -
            df["b_title_physical_glicko_youth"]
        )

        df["r_ko_rate_style_glicko_prime"] = (
            r_ko_rate_t34
            * r_swvc_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
            * r_prime_t30
        )
        df["b_ko_rate_style_glicko_prime"] = (
            b_ko_rate_t34
            * b_swvc_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
            * b_prime_t30
        )
        df["ko_rate_style_glicko_prime_diff"] = (
            df["r_ko_rate_style_glicko_prime"] -
            df["b_ko_rate_style_glicko_prime"]
        )

        r_career_peak_t36 = df.get(
            "r_career_elo_peak", pd.Series(1500.0, index=df.index)
        ).fillna(1500.0)
        b_career_peak_t36 = df.get(
            "b_career_elo_peak", pd.Series(1500.0, index=df.index)
        ).fillna(1500.0)

        df["r_fresh_winrate_physical_glicko_youth"] = (
            _fresh_r_t30
            * r_expq_t30
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
            * r_youth_t30
        )
        df["b_fresh_winrate_physical_glicko_youth"] = (
            _fresh_b_t30
            * b_expq_t30
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
            * b_youth_t30
        )
        df["fresh_winrate_physical_glicko_youth_diff"] = (
            df["r_fresh_winrate_physical_glicko_youth"]
            - df["b_fresh_winrate_physical_glicko_youth"]
        )

        df["r_strike_def_physical_glicko_youth"] = (
            r_str_def_t21
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_strike_def_physical_glicko_youth"] = (
            b_str_def_t21
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["strike_def_physical_glicko_youth_diff"] = (
            df["r_strike_def_physical_glicko_youth"]
            - df["b_strike_def_physical_glicko_youth"]
        )

        df["r_peak_physical_glicko_youth"] = (
            r_peak_t30
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_peak_physical_glicko_youth"] = (
            b_peak_t30
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["peak_physical_glicko_youth_diff"] = (
            df["r_peak_physical_glicko_youth"] -
            df["b_peak_physical_glicko_youth"]
        )

        df["r_elite_winrate_physical_glicko_prime"] = (
            r_expq_t30
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
            * r_prime_t30
            * (r_avg_opp_elo_t28 / 1500.0)
        )
        df["b_elite_winrate_physical_glicko_prime"] = (
            b_expq_t30
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
            * b_prime_t30
            * (b_avg_opp_elo_t28 / 1500.0)
        )
        df["elite_winrate_physical_glicko_prime_diff"] = (
            df["r_elite_winrate_physical_glicko_prime"]
            - df["b_elite_winrate_physical_glicko_prime"]
        )

        df["r_recent_opp_quality_physical_glicko"] = (
            (_r_roll3 / 3.0)
            * (r_avg_opp_elo_t28 / 1500.0)
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
        )
        df["b_recent_opp_quality_physical_glicko"] = (
            (_b_roll3 / 3.0)
            * (b_avg_opp_elo_t28 / 1500.0)
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
        )
        df["recent_opp_quality_physical_glicko_diff"] = (
            df["r_recent_opp_quality_physical_glicko"]
            - df["b_recent_opp_quality_physical_glicko"]
        )

        df["r_daniel_physical_glicko_youth"] = (
            r_daniel_t29
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
            * r_youth_t30
        )
        df["b_daniel_physical_glicko_youth"] = (
            b_daniel_t29
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
            * b_youth_t30
        )
        df["daniel_physical_glicko_youth_diff"] = (
            df["r_daniel_physical_glicko_youth"] -
            df["b_daniel_physical_glicko_youth"]
        )

        df["r_title_winrate_physical_glicko_youth"] = (
            r_tp_t34 * r_expq_t30 * r_ape_t30 *
            (r_glicko_r_t30 / 1500.0) * r_youth_t30
        )
        df["b_title_winrate_physical_glicko_youth"] = (
            b_tp_t34 * b_expq_t30 * b_ape_t30 *
            (b_glicko_r_t30 / 1500.0) * b_youth_t30
        )
        df["title_winrate_physical_glicko_youth_diff"] = (
            df["r_title_winrate_physical_glicko_youth"]
            - df["b_title_winrate_physical_glicko_youth"]
        )

        r_tot_def_t36 = r_str_def_t21 * r_td_def_t21
        b_tot_def_t36 = b_str_def_t21 * b_td_def_t21
        df["r_complete_def_physical_glicko"] = (
            r_tot_def_t36
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_prime_t30
        )
        df["b_complete_def_physical_glicko"] = (
            b_tot_def_t36
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_prime_t30
        )
        df["complete_def_physical_glicko_diff"] = (
            df["r_complete_def_physical_glicko"] -
            df["b_complete_def_physical_glicko"]
        )

        df["r_career_peak_physical_glicko"] = (
            (r_career_peak_t36 / 1500.0)
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
        )
        df["b_career_peak_physical_glicko"] = (
            (b_career_peak_t36 / 1500.0)
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
        )
        df["career_peak_physical_glicko_diff"] = (
            df["r_career_peak_physical_glicko"] -
            df["b_career_peak_physical_glicko"]
        )

        df["r_finish_def_physical_glicko"] = (
            r_fr_t23
            * r_str_def_t21
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
        )
        df["b_finish_def_physical_glicko"] = (
            b_fr_t23
            * b_str_def_t21
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
        )
        df["finish_def_physical_glicko_diff"] = (
            df["r_finish_def_physical_glicko"] -
            df["b_finish_def_physical_glicko"]
        )

        df["r_complete_def_physical_glicko_youth"] = (
            r_str_def_t21
            * r_td_def_t21
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_complete_def_physical_glicko_youth"] = (
            b_str_def_t21
            * b_td_def_t21
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["complete_def_physical_glicko_youth_diff"] = (
            df["r_complete_def_physical_glicko_youth"]
            - df["b_complete_def_physical_glicko_youth"]
        )

        df["r_td_def_physical_glicko_youth"] = (
            r_td_def_t21
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_td_def_physical_glicko_youth"] = (
            b_td_def_t21
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["td_def_physical_glicko_youth_diff"] = (
            df["r_td_def_physical_glicko_youth"] -
            df["b_td_def_physical_glicko_youth"]
        )

        df["r_peak_winrate_physical_glicko_youth"] = (
            r_peak_t30
            * r_expq_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_peak_winrate_physical_glicko_youth"] = (
            b_peak_t30
            * b_expq_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["peak_winrate_physical_glicko_youth_diff"] = (
            df["r_peak_winrate_physical_glicko_youth"]
            - df["b_peak_winrate_physical_glicko_youth"]
        )

        df["r_strike_def_winrate_physical_glicko"] = (
            r_str_def_t21
            * r_expq_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
        )
        df["b_strike_def_winrate_physical_glicko"] = (
            b_str_def_t21
            * b_expq_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
        )
        df["strike_def_winrate_physical_glicko_diff"] = (
            df["r_strike_def_winrate_physical_glicko"]
            - df["b_strike_def_winrate_physical_glicko"]
        )

        df["r_peak_style_physical_glicko_youth"] = (
            r_peak_t30
            * r_swvc_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_peak_style_physical_glicko_youth"] = (
            b_peak_t30
            * b_swvc_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["peak_style_physical_glicko_youth_diff"] = (
            df["r_peak_style_physical_glicko_youth"]
            - df["b_peak_style_physical_glicko_youth"]
        )

        df["r_career_peak_winrate_physical_glicko"] = (
            (r_career_peak_t36 / 1500.0)
            * r_expq_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
        )
        df["b_career_peak_winrate_physical_glicko"] = (
            (b_career_peak_t36 / 1500.0)
            * b_expq_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
        )
        df["career_peak_winrate_physical_glicko_diff"] = (
            df["r_career_peak_winrate_physical_glicko"]
            - df["b_career_peak_winrate_physical_glicko"]
        )

        df["r_strike_def_style_physical_glicko_youth"] = (
            r_str_def_t21
            * r_swvc_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_strike_def_style_physical_glicko_youth"] = (
            b_str_def_t21
            * b_swvc_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["strike_def_style_physical_glicko_youth_diff"] = (
            df["r_strike_def_style_physical_glicko_youth"]
            - df["b_strike_def_style_physical_glicko_youth"]
        )

        df["r_strike_def_physical_glicko_prime"] = (
            r_str_def_t21
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_prime_t30
        )
        df["b_strike_def_physical_glicko_prime"] = (
            b_str_def_t21
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_prime_t30
        )
        df["strike_def_physical_glicko_prime_diff"] = (
            df["r_strike_def_physical_glicko_prime"]
            - df["b_strike_def_physical_glicko_prime"]
        )

        df["r_peak_opp_quality_physical_glicko"] = (
            r_peak_t30
            * (r_avg_opp_elo_t28 / 1500.0)
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
        )
        df["b_peak_opp_quality_physical_glicko"] = (
            b_peak_t30
            * (b_avg_opp_elo_t28 / 1500.0)
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
        )
        df["peak_opp_quality_physical_glicko_diff"] = (
            df["r_peak_opp_quality_physical_glicko"]
            - df["b_peak_opp_quality_physical_glicko"]
        )

        df["r_fresh_strike_def_physical_glicko_youth"] = (
            _fresh_r_t30
            * r_str_def_t21
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_fresh_strike_def_physical_glicko_youth"] = (
            _fresh_b_t30
            * b_str_def_t21
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["fresh_strike_def_physical_glicko_youth_diff"] = (
            df["r_fresh_strike_def_physical_glicko_youth"]
            - df["b_fresh_strike_def_physical_glicko_youth"]
        )

        # ── Tier 38 ──────────────────────────────────────────────────────────────
        # TD defense × Win rate × Physical × Glicko × Youth
        df["r_td_def_winrate_physical_glicko_youth"] = (
            r_td_def_t21
            * r_expq_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_td_def_winrate_physical_glicko_youth"] = (
            b_td_def_t21
            * b_expq_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["td_def_winrate_physical_glicko_youth_diff"] = (
            df["r_td_def_winrate_physical_glicko_youth"]
            - df["b_td_def_winrate_physical_glicko_youth"]
        )

        # Combat efficiency × Physical × Glicko × Youth
        df["r_combat_eff_physical_glicko_youth"] = (
            r_cbt_t31
            * r_ape_t30
            * (r_reach_t23 / 71.0)
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_combat_eff_physical_glicko_youth"] = (
            b_cbt_t31
            * b_ape_t30
            * (b_reach_t23 / 71.0)
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["combat_eff_physical_glicko_youth_diff"] = (
            df["r_combat_eff_physical_glicko_youth"]
            - df["b_combat_eff_physical_glicko_youth"]
        )

        # Finish rate × Strike defense × Physical × Glicko × Youth
        df["r_finish_def_physical_glicko_youth"] = (
            r_fr_t23
            * r_str_def_t21
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_finish_def_physical_glicko_youth"] = (
            b_fr_t23
            * b_str_def_t21
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["finish_def_physical_glicko_youth_diff"] = (
            df["r_finish_def_physical_glicko_youth"]
            - df["b_finish_def_physical_glicko_youth"]
        )

        # Streak × Win rate × Physical × Glicko × Youth
        df["r_streak_winrate_physical_glicko_youth"] = (
            np.clip(r_streak_t23 / 5.0, 0.0, 1.0)
            * r_expq_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_streak_winrate_physical_glicko_youth"] = (
            np.clip(b_streak_t23 / 5.0, 0.0, 1.0)
            * b_expq_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["streak_winrate_physical_glicko_youth_diff"] = (
            df["r_streak_winrate_physical_glicko_youth"]
            - df["b_streak_winrate_physical_glicko_youth"]
        )

        # KO rate × Strike defense × Physical × Glicko × Youth
        df["r_ko_rate_def_physical_glicko_youth"] = (
            r_ko_rate_t34
            * r_str_def_t21
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_ko_rate_def_physical_glicko_youth"] = (
            b_ko_rate_t34
            * b_str_def_t21
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["ko_rate_def_physical_glicko_youth_diff"] = (
            df["r_ko_rate_def_physical_glicko_youth"]
            - df["b_ko_rate_def_physical_glicko_youth"]
        )

        # TD defense × Win rate × Physical × Glicko × Prime
        df["r_td_def_winrate_physical_glicko_prime"] = (
            r_td_def_t21
            * r_expq_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_prime_t30
        )
        df["b_td_def_winrate_physical_glicko_prime"] = (
            b_td_def_t21
            * b_expq_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_prime_t30
        )
        df["td_def_winrate_physical_glicko_prime_diff"] = (
            df["r_td_def_winrate_physical_glicko_prime"]
            - df["b_td_def_winrate_physical_glicko_prime"]
        )

        # Both defense walls × Win rate × Physical × Glicko
        df["r_complete_def_winrate_physical_glicko"] = (
            r_str_def_t21
            * r_td_def_t21
            * r_expq_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
        )
        df["b_complete_def_winrate_physical_glicko"] = (
            b_str_def_t21
            * b_td_def_t21
            * b_expq_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
        )
        df["complete_def_winrate_physical_glicko_diff"] = (
            df["r_complete_def_winrate_physical_glicko"]
            - df["b_complete_def_winrate_physical_glicko"]
        )

        # Career ELO peak × Strike defense × Physical × Glicko
        df["r_career_peak_strike_def_physical_glicko"] = (
            (r_career_peak_t36 / 1500.0)
            * r_str_def_t21
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
        )
        df["b_career_peak_strike_def_physical_glicko"] = (
            (b_career_peak_t36 / 1500.0)
            * b_str_def_t21
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
        )
        df["career_peak_strike_def_physical_glicko_diff"] = (
            df["r_career_peak_strike_def_physical_glicko"]
            - df["b_career_peak_strike_def_physical_glicko"]
        )

        # Peak × TD defense × Physical × Glicko × Youth
        df["r_peak_td_def_physical_glicko_youth"] = (
            r_peak_t30
            * r_td_def_t21
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_peak_td_def_physical_glicko_youth"] = (
            b_peak_t30
            * b_td_def_t21
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["peak_td_def_physical_glicko_youth_diff"] = (
            df["r_peak_td_def_physical_glicko_youth"]
            - df["b_peak_td_def_physical_glicko_youth"]
        )

        # Combat efficiency × Win rate × Physical × Glicko
        df["r_combat_eff_winrate_physical_glicko"] = (
            r_cbt_t31
            * r_expq_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_certainty_t30
        )
        df["b_combat_eff_winrate_physical_glicko"] = (
            b_cbt_t31
            * b_expq_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_certainty_t30
        )
        df["combat_eff_winrate_physical_glicko_diff"] = (
            df["r_combat_eff_winrate_physical_glicko"]
            - df["b_combat_eff_winrate_physical_glicko"]
        )

        # ── Tier 39 ──────────────────────────────────────────────────────────────
        # Combat efficiency × Style × Physical × Glicko × Youth
        df["r_combat_eff_style_physical_glicko_youth"] = (
            r_cbt_t31 * r_swvc_t30 * r_ape_t30 *
            (r_glicko_r_t30 / 1500.0) * r_youth_t30
        )
        df["b_combat_eff_style_physical_glicko_youth"] = (
            b_cbt_t31 * b_swvc_t30 * b_ape_t30 *
            (b_glicko_r_t30 / 1500.0) * b_youth_t30
        )
        df["combat_eff_style_physical_glicko_youth_diff"] = (
            df["r_combat_eff_style_physical_glicko_youth"]
            - df["b_combat_eff_style_physical_glicko_youth"]
        )

        # Peak × Combat eff × Style × Physical × Glicko × Youth
        df["r_peak_cbt_style_physical_glicko_youth"] = (
            r_peak_t30
            * r_cbt_t31
            * r_swvc_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_peak_cbt_style_physical_glicko_youth"] = (
            b_peak_t30
            * b_cbt_t31
            * b_swvc_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["peak_cbt_style_physical_glicko_youth_diff"] = (
            df["r_peak_cbt_style_physical_glicko_youth"]
            - df["b_peak_cbt_style_physical_glicko_youth"]
        )

        # Freshness × Style × Physical × Glicko × Youth
        df["r_fresh_style_physical_glicko_youth"] = (
            _fresh_r_t30
            * r_swvc_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_fresh_style_physical_glicko_youth"] = (
            _fresh_b_t30
            * b_swvc_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["fresh_style_physical_glicko_youth_diff"] = (
            df["r_fresh_style_physical_glicko_youth"]
            - df["b_fresh_style_physical_glicko_youth"]
        )

        # Recent form × Style × Physical × Glicko × Youth
        df["r_recent_style_physical_glicko_youth"] = (
            (_r_roll3 / 3.0)
            * r_swvc_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_recent_style_physical_glicko_youth"] = (
            (_b_roll3 / 3.0)
            * b_swvc_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["recent_style_physical_glicko_youth_diff"] = (
            df["r_recent_style_physical_glicko_youth"]
            - df["b_recent_style_physical_glicko_youth"]
        )

        # TD defense × Style × Physical × Glicko × Youth
        df["r_td_def_style_physical_glicko_youth"] = (
            r_td_def_t21
            * r_swvc_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_td_def_style_physical_glicko_youth"] = (
            b_td_def_t21
            * b_swvc_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["td_def_style_physical_glicko_youth_diff"] = (
            df["r_td_def_style_physical_glicko_youth"]
            - df["b_td_def_style_physical_glicko_youth"]
        )

        # Finish rate × Style × Physical × Glicko × Youth
        df["r_finish_style_physical_glicko_youth"] = (
            r_fr_t23 * r_swvc_t30 * r_ape_t30 *
            (r_glicko_r_t30 / 1500.0) * r_youth_t30
        )
        df["b_finish_style_physical_glicko_youth"] = (
            b_fr_t23 * b_swvc_t30 * b_ape_t30 *
            (b_glicko_r_t30 / 1500.0) * b_youth_t30
        )
        df["finish_style_physical_glicko_youth_diff"] = (
            df["r_finish_style_physical_glicko_youth"]
            - df["b_finish_style_physical_glicko_youth"]
        )

        # Complete defense × Style × Physical × Glicko × Youth
        df["r_complete_def_style_physical_glicko_youth"] = (
            r_str_def_t21
            * r_td_def_t21
            * r_swvc_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_complete_def_style_physical_glicko_youth"] = (
            b_str_def_t21
            * b_td_def_t21
            * b_swvc_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["complete_def_style_physical_glicko_youth_diff"] = (
            df["r_complete_def_style_physical_glicko_youth"]
            - df["b_complete_def_style_physical_glicko_youth"]
        )

        # Career ELO peak × Style × Physical × Glicko × Youth
        df["r_career_peak_style_physical_glicko_youth"] = (
            (r_career_peak_t36 / 1500.0)
            * r_swvc_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_career_peak_style_physical_glicko_youth"] = (
            (b_career_peak_t36 / 1500.0)
            * b_swvc_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["career_peak_style_physical_glicko_youth_diff"] = (
            df["r_career_peak_style_physical_glicko_youth"]
            - df["b_career_peak_style_physical_glicko_youth"]
        )

        # Opp quality × Style × Physical × Glicko × Youth
        df["r_opp_quality_style_physical_glicko_youth"] = (
            (r_avg_opp_elo_t28 / 1500.0)
            * r_swvc_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_opp_quality_style_physical_glicko_youth"] = (
            (b_avg_opp_elo_t28 / 1500.0)
            * b_swvc_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["opp_quality_style_physical_glicko_youth_diff"] = (
            df["r_opp_quality_style_physical_glicko_youth"]
            - df["b_opp_quality_style_physical_glicko_youth"]
        )

        # Streak × Style × Physical × Glicko × Youth
        df["r_streak_style_physical_glicko_youth"] = (
            np.clip(r_streak_t23 / 5.0, 0.0, 1.0)
            * r_swvc_t30
            * r_ape_t30
            * (r_glicko_r_t30 / 1500.0)
            * r_youth_t30
        )
        df["b_streak_style_physical_glicko_youth"] = (
            np.clip(b_streak_t23 / 5.0, 0.0, 1.0)
            * b_swvc_t30
            * b_ape_t30
            * (b_glicko_r_t30 / 1500.0)
            * b_youth_t30
        )
        df["streak_style_physical_glicko_youth_diff"] = (
            df["r_streak_style_physical_glicko_youth"]
            - df["b_streak_style_physical_glicko_youth"]
        )

        # ── Tier 40: Decayed Average Difference Features (inference) ─────────────
        _da_cols_inf = [
            "da_sig_str_acc",
            "da_td_acc",
            "da_head_landed",
            "da_head_acc",
            "da_body_acc",
            "da_distance_acc",
            "da_head_defense",
            "da_body_defense",
            "da_distance_defense",
            "da_ground_defense",
            "da_td_defense",
            "da_sub_att",
            "da_kd",
            "da_ko",
            "da_win_ratio",
            "da_ctrl_r1",
            "da_clinch_pm",
            "da_opp_leg_pm",
            "da_opp_ctrl_r1_pm",
            "da_opp_sub_pm",
            "da_opp_rev_r1",
            "da_r1_strikes",
            "da_reversals",
            "da_dist_landing_ratio",
            "da_opp_kd",
            "da_age",
            "da_ufc_age",
            "da_reach_ratio",
            "da_days_since",
            "dapa_sig_str_acc",
            "dapa_head_acc",
            "dapa_body_acc",
            "dapa_distance_acc",
            "dapa_head_defense",
            "dapa_dist_defense",
            "dapa_ground_defense",
            "dapa_r1_strikes",
            "dapa_reversals",
            "dapa_dist_landing_ratio",
            "dapa_head_landing_ratio",
        ]
        for _col in _da_cols_inf:
            _rc = f"r_pre_{_col}"
            _bc = f"b_pre_{_col}"
            if _rc in df.columns and _bc in df.columns:
                df[f"{_col}_diff"] = df[_rc].fillna(0) - df[_bc].fillna(0)

        # Physical ratio features
        _r_age_inf = df.get("r_age_at_event", pd.Series(28.0, index=df.index)).fillna(
            28.0
        )
        _b_age_inf = df.get("b_age_at_event", pd.Series(28.0, index=df.index)).fillna(
            28.0
        )
        df["age_ratio_diff"] = (_r_age_inf / (_b_age_inf + 0.1)) - 1.0

        _r_reach_inf = df.get("r_reach", pd.Series(
            70.0, index=df.index)).fillna(70.0)
        _b_reach_inf = df.get("b_reach", pd.Series(
            70.0, index=df.index)).fillna(70.0)
        df["reach_ratio_diff"] = (_r_reach_inf / (_b_reach_inf + 0.1)) - 1.0

        # Composite decayed features
        _r_da_ss_inf = df.get(
            "r_pre_da_sig_str_acc", pd.Series(0.5, index=df.index)
        ).fillna(0.5)
        _b_da_ss_inf = df.get(
            "b_pre_da_sig_str_acc", pd.Series(0.5, index=df.index)
        ).fillna(0.5)
        _r_da_ha_inf = df.get(
            "r_pre_da_head_acc", pd.Series(0.45, index=df.index)
        ).fillna(0.45)
        _b_da_ha_inf = df.get(
            "b_pre_da_head_acc", pd.Series(0.45, index=df.index)
        ).fillna(0.45)
        df["da_striking_precision_diff"] = (_r_da_ss_inf * _r_da_ha_inf) - (
            _b_da_ss_inf * _b_da_ha_inf
        )

        _r_da_hd_inf = df.get(
            "r_pre_da_head_defense", pd.Series(0.55, index=df.index)
        ).fillna(0.55)
        _b_da_hd_inf = df.get(
            "b_pre_da_head_defense", pd.Series(0.55, index=df.index)
        ).fillna(0.55)
        _r_da_dd_inf = df.get(
            "r_pre_da_distance_defense", pd.Series(0.55, index=df.index)
        ).fillna(0.55)
        _b_da_dd_inf = df.get(
            "b_pre_da_distance_defense", pd.Series(0.55, index=df.index)
        ).fillna(0.55)
        _r_da_td_def_inf = df.get(
            "r_pre_da_td_defense", pd.Series(0.6, index=df.index)
        ).fillna(0.6)
        _b_da_td_def_inf = df.get(
            "b_pre_da_td_defense", pd.Series(0.6, index=df.index)
        ).fillna(0.6)
        df["r_da_defense_composite"] = (
            _r_da_hd_inf + _r_da_dd_inf + _r_da_td_def_inf
        ) / 3.0
        df["b_da_defense_composite"] = (
            _b_da_hd_inf + _b_da_dd_inf + _b_da_td_def_inf
        ) / 3.0
        df["da_defense_composite_diff"] = (
            df["r_da_defense_composite"] - df["b_da_defense_composite"]
        )

        _r_da_tdacc_inf = df.get(
            "r_pre_da_td_acc", pd.Series(0.4, index=df.index)
        ).fillna(0.4)
        _b_da_tdacc_inf = df.get(
            "b_pre_da_td_acc", pd.Series(0.4, index=df.index)
        ).fillna(0.4)
        _r_da_sub_inf = df.get(
            "r_pre_da_sub_att", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        _b_da_sub_inf = df.get(
            "b_pre_da_sub_att", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        df["da_grapple_threat_diff"] = _r_da_tdacc_inf * np.log1p(
            _r_da_sub_inf
        ) - _b_da_tdacc_inf * np.log1p(_b_da_sub_inf)

        _r_da_olegpm_inf = df.get(
            "r_pre_da_opp_leg_pm", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        _b_da_olegpm_inf = df.get(
            "b_pre_da_opp_leg_pm", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        _r_da_osubpm_inf = df.get(
            "r_pre_da_opp_sub_pm", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        _b_da_osubpm_inf = df.get(
            "b_pre_da_opp_sub_pm", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        _r_da_ocr1_inf = df.get(
            "r_pre_da_opp_ctrl_r1_pm", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        _b_da_ocr1_inf = df.get(
            "b_pre_da_opp_ctrl_r1_pm", pd.Series(0.0, index=df.index)
        ).fillna(0.0)
        df["da_opp_pressure_diff"] = (
            _r_da_olegpm_inf + _r_da_osubpm_inf + _r_da_ocr1_inf
        ) - (_b_da_olegpm_inf + _b_da_osubpm_inf + _b_da_ocr1_inf)

        _r_da_ko_inf = df.get("r_pre_da_ko", pd.Series(
            0.0, index=df.index)).fillna(0.0)
        _b_da_ko_inf = df.get("b_pre_da_ko", pd.Series(
            0.0, index=df.index)).fillna(0.0)
        _r_dapa_ha_inf = df.get(
            "r_pre_dapa_head_acc", pd.Series(0.45, index=df.index)
        ).fillna(0.45)
        _b_dapa_ha_inf = df.get(
            "b_pre_dapa_head_acc", pd.Series(0.45, index=df.index)
        ).fillna(0.45)
        df["da_ko_x_head_acc_diff"] = (_r_da_ko_inf * _r_dapa_ha_inf) - (
            _b_da_ko_inf * _b_dapa_ha_inf
        )

        _r_da_wr_inf = df.get(
            "r_pre_da_win_ratio", pd.Series(0.5, index=df.index)
        ).fillna(0.5)
        _b_da_wr_inf = df.get(
            "b_pre_da_win_ratio", pd.Series(0.5, index=df.index)
        ).fillna(0.5)
        df["da_win_x_defense_diff"] = (
            _r_da_wr_inf * df["r_da_defense_composite"]
            - _b_da_wr_inf * df["b_da_defense_composite"]
        )

        # ── Tier 41: Method-signal features (inference rebuild) ───────────
        _r_ko_wr_41 = df.get("r_ko_win_rate", pd.Series(
            0, index=df.index)).fillna(0)
        _b_ko_wr_41 = df.get("b_ko_win_rate", pd.Series(
            0, index=df.index)).fillna(0)
        _r_sub_wr_41 = df.get("r_sub_win_rate", pd.Series(
            0, index=df.index)).fillna(0)
        _b_sub_wr_41 = df.get("b_sub_win_rate", pd.Series(
            0, index=df.index)).fillna(0)
        _r_dec_wr_41 = df.get(
            "r_decision_win_rate", pd.Series(0, index=df.index)
        ).fillna(0)
        _b_dec_wr_41 = df.get(
            "b_decision_win_rate", pd.Series(0, index=df.index)
        ).fillna(0)
        _r_ko_lr_41 = df.get("r_ko_loss_rate", pd.Series(
            0, index=df.index)).fillna(0)
        _b_ko_lr_41 = df.get("b_ko_loss_rate", pd.Series(
            0, index=df.index)).fillna(0)
        _r_sub_lr_41 = df.get("r_sub_loss_rate", pd.Series(
            0, index=df.index)).fillna(0)
        _b_sub_lr_41 = df.get("b_sub_loss_rate", pd.Series(
            0, index=df.index)).fillna(0)
        _r_dec_lr_41 = df.get("r_dec_loss_rate", pd.Series(
            0, index=df.index)).fillna(0)
        _b_dec_lr_41 = df.get("b_dec_loss_rate", pd.Series(
            0, index=df.index)).fillna(0)

        # 41a: Susceptibility interactions
        df["r_ko_susc_interaction"] = _r_ko_wr_41 * _b_ko_lr_41
        df["b_ko_susc_interaction"] = _b_ko_wr_41 * _r_ko_lr_41
        df["r_sub_susc_interaction"] = _r_sub_wr_41 * _b_sub_lr_41
        df["b_sub_susc_interaction"] = _b_sub_wr_41 * _r_sub_lr_41
        df["r_dec_susc_interaction"] = _r_dec_wr_41 * _b_dec_lr_41
        df["b_dec_susc_interaction"] = _b_dec_wr_41 * _r_dec_lr_41
        df["r_finish_threat"] = np.maximum(
            df["r_ko_susc_interaction"], df["r_sub_susc_interaction"]
        )
        df["b_finish_threat"] = np.maximum(
            df["b_ko_susc_interaction"], df["b_sub_susc_interaction"]
        )

        # 41b: Effective method rates
        _SUSC_41 = 0.35
        for _pfx_41, _opp_41 in [("r", "b"), ("b", "r")]:
            _ek = (1.0 - _SUSC_41) * df.get(
                f"{_pfx_41}_ko_win_rate", pd.Series(0, index=df.index)
            ).fillna(0) + _SUSC_41 * df.get(
                f"{_opp_41}_ko_loss_rate", pd.Series(0, index=df.index)
            ).fillna(0)
            _es = (1.0 - _SUSC_41) * df.get(
                f"{_pfx_41}_sub_win_rate", pd.Series(0, index=df.index)
            ).fillna(0) + _SUSC_41 * df.get(
                f"{_opp_41}_sub_loss_rate", pd.Series(0, index=df.index)
            ).fillna(0)
            _ed = (1.0 - _SUSC_41) * df.get(
                f"{_pfx_41}_decision_win_rate", pd.Series(0, index=df.index)
            ).fillna(0) + _SUSC_41 * df.get(
                f"{_opp_41}_dec_loss_rate", pd.Series(0, index=df.index)
            ).fillna(0)
            _et = (_ek + _es + _ed).clip(lower=1e-9)
            df[f"{_pfx_41}_eff_ko_rate"] = _ek / _et
            df[f"{_pfx_41}_eff_sub_rate"] = _es / _et
            df[f"{_pfx_41}_eff_dec_rate"] = _ed / _et
        df["r_eff_finish_rate"] = df["r_eff_ko_rate"] + df["r_eff_sub_rate"]
        df["b_eff_finish_rate"] = df["b_eff_ko_rate"] + df["b_eff_sub_rate"]

        # 41c: Stance KO features
        _r_st41_r = (
            df.get("r_stance", pd.Series("", index=df.index))
            .fillna("")
            .astype(str)
            .str.strip()
            .str.lower()
        )
        _b_st41_r = (
            df.get("b_stance", pd.Series("", index=df.index))
            .fillna("")
            .astype(str)
            .str.strip()
            .str.lower()
        )
        _cross_41 = (
            ((_r_st41_r == "orthodox") & (_b_st41_r == "southpaw"))
            | ((_r_st41_r == "southpaw") & (_b_st41_r == "orthodox"))
        ).astype(float)
        _mirror_41 = (_r_st41_r == _b_st41_r).astype(float)
        _r_sp_41 = _r_st41_r.isin(
            ["southpaw", "switch", "open stance"]).astype(float)
        _b_sp_41 = _b_st41_r.isin(
            ["southpaw", "switch", "open stance"]).astype(float)

        df["r_ko_rate_cross_stance"] = _r_ko_wr_41 * _cross_41
        df["b_ko_rate_cross_stance"] = _b_ko_wr_41 * _cross_41
        df["r_southpaw_ko_advantage"] = _r_sp_41 * \
            _r_ko_wr_41 * (1.0 - _b_sp_41)
        df["b_southpaw_ko_advantage"] = _b_sp_41 * \
            _b_ko_wr_41 * (1.0 - _r_sp_41)
        _r_sw_41 = _r_st41_r.isin(["switch", "open stance"]).astype(float)
        _b_sw_41 = _b_st41_r.isin(["switch", "open stance"]).astype(float)
        df["r_switch_finish_threat"] = _r_sw_41 * df.get(
            "r_finish_rate", pd.Series(0, index=df.index)
        ).fillna(0)
        df["b_switch_finish_threat"] = _b_sw_41 * df.get(
            "b_finish_rate", pd.Series(0, index=df.index)
        ).fillna(0)
        df["cross_stance_ko_potential"] = _cross_41 * \
            (_r_ko_wr_41 + _b_ko_wr_41) / 2
        df["mirror_stance_dec_tendency"] = (
            _mirror_41 * (_r_dec_wr_41 + _b_dec_wr_41) / 2
        )

        # ── Tier 42: Method-path closure, uncertainty routing, SVD method-alignment (inference) ──
        _eps_42 = 1e-6

        def _gi42(col, default=0.0):
            return df.get(col, pd.Series(default, index=df.index)).fillna(default)

        def _second_max_42(a, b, c):
            stacked = np.column_stack([a, b, c])
            stacked.sort(axis=1)
            return stacked[:, 1]

        for _pfx_42, _opp_42 in [("r", "b"), ("b", "r")]:
            _spc_42 = (
                (
                    _gi42(f"{_pfx_42}_pre_sub_after_td")
                    * _gi42(f"{_pfx_42}_pre_ctrl_per_td")
                    * _gi42(f"{_pfx_42}_pre_sub_efficiency")
                    * (1.0 - _gi42(f"{_opp_42}_pre_td_def", 0.5))
                )
                .clip(0)
                .fillna(0)
            )
            df[f"{_pfx_42}_sub_path_closure"] = _spc_42

            df[f"{_pfx_42}_late_sub_conversion"] = (
                (
                    (
                        _gi42(f"{_pfx_42}_pre_r23_sub_rate")
                        + 0.5 * _gi42(f"{_pfx_42}_pre_r45_sub_rate_l")
                    )
                    * _gi42(f"{_pfx_42}_pre_late_td_acc")
                    * (1.0 - _gi42(f"{_opp_42}_pre_r1_td_def_rate", 0.5))
                )
                .clip(0)
                .fillna(0)
            )

            df[f"{_pfx_42}_scramble_sub_window"] = (
                (
                    (
                        _gi42(f"{_pfx_42}_pre_rev_per_ctrl")
                        + _gi42(f"{_pfx_42}_pre_reversal_rate")
                    )
                    * _gi42(f"{_pfx_42}_pre_sub_efficiency")
                    * (1.0 + _gi42(f"{_opp_42}_pre_output_cv", 0.5))
                )
                .clip(0)
                .fillna(0)
            )

            df[f"{_pfx_42}_anti_sub_escape_integrity"] = (
                (
                    _gi42(f"{_pfx_42}_pre_td_def", 0.5)
                    * _gi42(f"{_pfx_42}_pre_r1_td_def_rate", 0.5)
                    * _gi42(f"{_pfx_42}_pre_output_consistency", 0.5)
                    / (
                        1.0
                        + _gi42(f"{_opp_42}_pre_sub_after_td")
                        + _gi42(f"{_opp_42}_pre_sub_efficiency")
                    )
                )
                .clip(0)
                .fillna(0)
            )

            _db_42 = (
                (
                    _gi42(f"{_pfx_42}_eff_dec_rate")
                    * _gi42(f"{_pfx_42}_pre_output_consistency", 0.5)
                    * np.log1p(
                        np.maximum(
                            _gi42(f"{_pfx_42}_pre_dec_win_margin").values, 0)
                    )
                    * (1.0 + _gi42(f"{_pfx_42}_pre_lead_hold_rate"))
                )
                .clip(0)
                .fillna(0)
            )
            df[f"{_pfx_42}_decision_bankability"] = _db_42

            df[f"{_pfx_42}_finish_conversion_pressure"] = (
                (
                    _gi42(f"{_pfx_42}_eff_finish_rate")
                    * np.log1p(
                        _gi42(f"{_pfx_42}_pre_pressure_index").values
                        + _gi42(f"{_pfx_42}_pre_r1_td_pressure").values
                    )
                    * (1.0 + _gi42(f"{_pfx_42}_pre_output_eff_trend"))
                    * (
                        1.0
                        + _gi42(f"{_opp_42}_pre_r1_abs_rate")
                        + _gi42(f"{_opp_42}_pre_chin_ratio", 1.0)
                    )
                )
                .clip(0)
                .fillna(0)
            )

            df[f"{_pfx_42}_attritional_break_score"] = (
                (
                    _gi42(f"{_pfx_42}_pre_cardio_index", 0.75)
                    * _gi42(f"{_pfx_42}_pre_late_win_rate", 0.5)
                    * _gi42(f"{_pfx_42}_pre_r2_to_r3_momentum")
                    * (1.0 + _gi42(f"{_opp_42}_pre_fatigue_composite"))
                )
                .clip(0)
                .fillna(0)
            )

            _ecbs_42 = (
                (
                    _gi42(f"{_pfx_42}_pre_r1_finish_rate")
                    * _gi42(f"{_pfx_42}_pre_kd_efficiency")
                    * (
                        1.0
                        + _gi42(f"{_pfx_42}_ko_rate_cross_stance")
                        + _gi42(f"{_pfx_42}_southpaw_ko_advantage")
                    )
                    * (1.0 + _gi42(f"{_opp_42}_pre_r1_abs_rate"))
                )
                .clip(0)
                .fillna(0)
            )
            df[f"{_pfx_42}_early_chaos_break_score"] = _ecbs_42

            df[f"{_pfx_42}_volatility_finish_chaos"] = (
                (
                    _gi42(f"{_pfx_42}_glicko_pre_vol", 0.06)
                    * _gi42(f"{_pfx_42}_eff_finish_rate")
                    * (1.0 + _gi42(f"{_opp_42}_glicko_pre_vol", 0.06))
                    * (
                        1.0
                        + np.abs(
                            _gi42(f"{_pfx_42}_pre_output_cv", 0.5).values
                            - _gi42(f"{_opp_42}_pre_output_cv", 0.5).values
                        )
                    )
                )
                .clip(0)
                .fillna(0)
            )

            df[f"{_pfx_42}_layoff_decay_mismatch"] = (
                (
                    np.log1p(_gi42(f"{_pfx_42}_days_since_last", 180).values)
                    * (1.0 - _gi42(f"{_pfx_42}_pre_accuracy_retention", 1.0))
                    * (
                        1.0
                        + _gi42(f"{_opp_42}_pre_pressure_index")
                        + _gi42(f"{_opp_42}_eff_finish_rate")
                    )
                )
                .clip(0)
                .fillna(0)
            )

            df[f"{_pfx_42}_style_method_mismatch"] = (
                _gi42(f"{_pfx_42}_style_win_vs_cluster", 0.5)
                * (
                    _gi42(f"{_pfx_42}_eff_sub_rate")
                    + 0.5 * _gi42(f"{_pfx_42}_eff_ko_rate")
                    - _gi42(f"{_pfx_42}_eff_dec_rate")
                )
                * _gi42(f"{_pfx_42}_physical_style_dominance")
            ).fillna(0)

            _eko_42 = _gi42(f"{_pfx_42}_eff_ko_rate").values
            _esub_42 = _gi42(f"{_pfx_42}_eff_sub_rate").values
            _edec_42 = _gi42(f"{_pfx_42}_eff_dec_rate").values
            _top_42 = np.maximum(np.maximum(_eko_42, _esub_42), _edec_42)
            _sec_42 = _second_max_42(_eko_42, _esub_42, _edec_42)
            df[f"{_pfx_42}_method_path_separation"] = (
                pd.Series(_top_42 - _sec_42, index=df.index).clip(0).fillna(0)
            )

            df[f"{_pfx_42}_svd_sub_alignment"] = (
                (0.6 * _gi42("grappling_svd_0") + 0.4 * _gi42("form_svd_1"))
                * df[f"{_pfx_42}_sub_path_closure"]
            ).fillna(0)

            df[f"{_pfx_42}_svd_decision_control_alignment"] = (
                (0.5 * _gi42("grappling_svd_1") + 0.5 * _gi42("form_svd_0"))
                * df[f"{_pfx_42}_decision_bankability"]
            ).fillna(0)

            df[f"{_pfx_42}_svd_striking_break_alignment"] = (
                (
                    0.5 * _gi42("striking_svd_0")
                    + 0.3 * _gi42("form_svd_2")
                    + 0.2 * _gi42("physical_svd_0")
                )
                * df[f"{_pfx_42}_early_chaos_break_score"]
            ).fillna(0)

            df[f"{_pfx_42}_svd_archetype_instability"] = pd.Series(
                np.abs(_gi42("striking_svd_0").values -
                       _gi42("grappling_svd_0").values)
                * _gi42(f"{_pfx_42}_glicko_pre_vol", 0.06).values
                * _gi42(f"{_pfx_42}_pre_output_cv", 0.5).values,
                index=df.index,
            ).fillna(0)

        _t42_inf_concepts = [
            "sub_path_closure",
            "late_sub_conversion",
            "scramble_sub_window",
            "anti_sub_escape_integrity",
            "decision_bankability",
            "finish_conversion_pressure",
            "attritional_break_score",
            "early_chaos_break_score",
            "volatility_finish_chaos",
            "layoff_decay_mismatch",
            "style_method_mismatch",
            "method_path_separation",
            "svd_sub_alignment",
            "svd_decision_control_alignment",
            "svd_striking_break_alignment",
            "svd_archetype_instability",
        ]
        for _c42 in _t42_inf_concepts:
            df[f"{_c42}_diff"] = df[f"r_{_c42}"] - df[f"b_{_c42}"]

        df = self._apply_tier43_method_head_features(df)

        return df

    def _build_estimators(self, X_tr, y_tr, sample_weight=None, X_val=None, y_val=None):
        from sklearn.ensemble import ExtraTreesClassifier

        estimators = []

        n_pos = max(int(np.sum(y_tr == 1)), 1)
        n_neg = max(int(np.sum(y_tr == 0)), 1)
        spw = n_neg / n_pos
        print_metric("  Class ratio (neg/pos):", f"{spw:.3f}")
        _has_val = X_val is not None and y_val is not None
        print()
        if _has_val:
            print_step(
                "Early stopping enabled for boosting models (patience=50)")

        if HAS_XGB:
            xgb_params = {
                "n_estimators": 500,
                "max_depth": 6,
                "learning_rate": 0.05,
                "subsample": 0.8,
                "colsample_bytree": 0.8,
                "min_child_weight": 10,
                "reg_alpha": 0.5,
                "reg_lambda": 1.0,
                "gamma": 0.1,
                "random_state": RANDOM_SEED,
                "verbosity": 0,
                "eval_metric": "logloss",
                "n_jobs": SAFE_N_JOBS,
                "scale_pos_weight": 1.0,
            }
            if self.gpu_info.get("xgb"):
                xgb_params["device"] = "cuda"
            else:
                xgb_params["tree_method"] = "hist"
            if (
                hasattr(self, "_optuna_best_xgb_params")
                and self._optuna_best_xgb_params
            ):
                xgb_params.update(self._optuna_best_xgb_params)
            clf = xgb.XGBClassifier(**xgb_params)
            xgb_fit_kw = {"sample_weight": sample_weight, "verbose": False}
            if _has_val:
                xgb_fit_kw["eval_set"] = [(X_val, y_val)]
                xgb_params["early_stopping_rounds"] = 50
                clf = xgb.XGBClassifier(**xgb_params)
            clf.fit(X_tr, y_tr, **xgb_fit_kw)
            if _has_val and hasattr(clf, "best_iteration"):
                print_metric("  XGB early stop iteration:", clf.best_iteration)
            estimators.append(("xgb", clf))

        if HAS_LGB:
            lgb_params = {
                "n_estimators": 500,
                "num_leaves": 31,
                "max_depth": -1,
                "learning_rate": 0.05,
                "subsample": 0.8,
                "colsample_bytree": 0.8,
                "min_child_samples": 30,
                "reg_alpha": 0.5,
                "reg_lambda": 1.0,
                "random_state": RANDOM_SEED,
                "verbose": -1,
                "n_jobs": SAFE_N_JOBS,
                "class_weight": "balanced",
            }
            if (
                hasattr(self, "_optuna_best_lgb_params")
                and self._optuna_best_lgb_params
            ):
                lgb_params.update(self._optuna_best_lgb_params)
            clf = lgb.LGBMClassifier(**lgb_params)
            lgb_fit_kw = {"sample_weight": sample_weight}
            if _has_val:
                lgb_fit_kw["eval_set"] = [(X_val, y_val)]
                lgb_fit_kw["callbacks"] = [
                    lgb.early_stopping(50, verbose=False),
                    lgb.log_evaluation(-1),
                ]
            clf.fit(X_tr, y_tr, **lgb_fit_kw)
            if _has_val and hasattr(clf, "best_iteration_"):
                print_metric("  LGB early stop iteration:",
                             clf.best_iteration_)
            estimators.append(("lgb", clf))

        rf = RandomForestClassifier(
            n_estimators=300,
            max_depth=12,
            min_samples_split=6,
            min_samples_leaf=3,
            max_features="sqrt",
            random_state=RANDOM_SEED,
            n_jobs=SAFE_N_JOBS,
            class_weight="balanced",
        )
        rf.fit(X_tr, y_tr, sample_weight=sample_weight)
        estimators.append(("rf", rf))

        et = ExtraTreesClassifier(
            n_estimators=300,
            max_depth=None,
            min_samples_split=8,
            min_samples_leaf=4,
            max_features="sqrt",
            random_state=RANDOM_SEED + 1,
            n_jobs=SAFE_N_JOBS,
            class_weight="balanced",
        )
        et.fit(X_tr, y_tr, sample_weight=sample_weight)
        estimators.append(("et", et))

        from sklearn.neural_network import MLPClassifier

        # Resample training data proportional to recency weights for MLPs
        # (MLPClassifier doesn't accept sample_weight).
        if sample_weight is not None:
            _rng_mlp = np.random.RandomState(RANDOM_SEED)
            _probs_mlp = sample_weight / sample_weight.sum()
            _mlp_idx = _rng_mlp.choice(
                len(X_tr), size=len(X_tr), replace=True, p=_probs_mlp
            )
            X_tr_mlp = X_tr[_mlp_idx]
            y_tr_mlp = y_tr[_mlp_idx]
        else:
            X_tr_mlp = X_tr
            y_tr_mlp = y_tr

        mlp = MLPClassifier(
            hidden_layer_sizes=(128, 64),
            max_iter=500,
            early_stopping=True,
            validation_fraction=0.1,
            n_iter_no_change=20,
            learning_rate="adaptive",
            random_state=RANDOM_SEED,
        )
        mlp.fit(X_tr_mlp, y_tr_mlp)
        estimators.append(("mlp", mlp))

        mlp2 = MLPClassifier(
            hidden_layer_sizes=(256, 128, 64),
            activation="tanh",
            alpha=0.01,
            max_iter=500,
            early_stopping=True,
            validation_fraction=0.1,
            n_iter_no_change=20,
            learning_rate="adaptive",
            random_state=RANDOM_SEED + 2,
        )
        mlp2.fit(X_tr_mlp, y_tr_mlp)
        estimators.append(("mlp2", mlp2))

        if HAS_CAT:
            cat_params = {
                "iterations": 300,
                "depth": 5,
                "learning_rate": 0.05,
                "l2_leaf_reg": 8,
                "bagging_temperature": 0.8,
                "random_strength": 1.5,
                "random_seed": RANDOM_SEED,
                "verbose": 0,
                "eval_metric": "Logloss",
                "auto_class_weights": "Balanced",
            }
            if (
                hasattr(self, "_optuna_best_cat_params")
                and self._optuna_best_cat_params
            ):
                cat_params.update(self._optuna_best_cat_params)
            if self.gpu_info.get("cat"):
                cat_params["task_type"] = "GPU"
                cat_params["border_count"] = 128
            if _has_val:
                cat_params["early_stopping_rounds"] = 50
            clf = cb.CatBoostClassifier(
                **cat_params, allow_writing_files=False)
            cat_fit_kw = {"sample_weight": sample_weight}
            if _has_val:
                cat_fit_kw["eval_set"] = (X_val, y_val)
            clf.fit(X_tr, y_tr, **cat_fit_kw)
            if _has_val and hasattr(clf, "best_iteration_"):
                print_metric("  CAT early stop iteration:",
                             clf.best_iteration_)
            estimators.append(("cat", clf))

        from sklearn.ensemble import HistGradientBoostingClassifier

        hgb = HistGradientBoostingClassifier(
            max_iter=500,
            max_depth=6,
            learning_rate=0.05,
            max_leaf_nodes=31,
            min_samples_leaf=20,
            l2_regularization=1.0,
            early_stopping=False,
            random_state=RANDOM_SEED,
            class_weight="balanced",
        )
        hgb.fit(X_tr, y_tr, sample_weight=sample_weight)
        estimators.append(("hgb", hgb))

        from sklearn.linear_model import SGDClassifier

        ridge_lr = SGDClassifier(
            loss="log_loss",
            penalty="elasticnet",
            alpha=1e-4,
            l1_ratio=0.15,
            max_iter=2000,
            tol=1e-4,
            random_state=RANDOM_SEED,
            class_weight="balanced",
            n_jobs=SAFE_N_JOBS,
        )
        if sample_weight is not None:
            ridge_lr.fit(X_tr, y_tr, sample_weight=sample_weight)
        else:
            ridge_lr.fit(X_tr, y_tr)
        estimators.append(("ridge", ridge_lr))

        return estimators

    def predict_upcoming_fights(self, fights):
        print_section("PREDICTING UPCOMING FIGHTS")
        self.predictions = []

        for fight in fights:
            r_name = fight.get("r_fighter", "")
            b_name = fight.get("b_fighter", "")
            weight_class = fight.get("weight_class", "")
            gender = fight.get("gender", "Men")
            total_rounds = int(fight.get("total_rounds", 3))

            if r_name not in self.all_fighters:
                self._log(
                    f"  Skipping {r_name} vs {b_name} — '{r_name}' not in records (debut?)"
                )
                continue
            if b_name not in self.all_fighters:
                self._log(
                    f"  Skipping {r_name} vs {b_name} — '{b_name}' not in records (debut?)"
                )
                continue

            r_matched = r_name
            b_matched = b_name

            result = self._build_fight_feature_vector(
                r_matched, b_matched, weight_class, gender, total_rounds
            )
            if result is None:
                self._log(
                    f"  Skipping {r_name} vs {b_name} — could not build feature vector"
                )
                continue

            feat_vec, _ = result

            def _scale_select(fv):
                Xs = self.scaler.transform(np.array([fv]))
                if (
                    hasattr(self, "_global_selector")
                    and self._global_selector is not None
                ):
                    Xs = self._global_selector.transform(Xs)
                return Xs

            X_sel = _scale_select(feat_vec)

            win_proba = self.stacking_clf.predict_proba(X_sel)[0]
            classes = self.stacking_clf.classes_

            r_idx = list(classes).index(1) if 1 in classes else 1

            # Apply the same blend-alpha and corner-swap averaging used in
            # holdout evaluation so live predictions match reported metrics.
            _alpha = getattr(self, "_blend_alpha", 0.5)
            _d_idx = getattr(self, "_d_indices", [])
            if len(_d_idx) > 0 and _alpha < 1.0:
                X_sel_swap = X_sel.copy()
                X_sel_swap[:, _d_idx] *= -1
                win_proba_swap = self.stacking_clf.predict_proba(X_sel_swap)[0]
                r_win_prob = float(
                    _alpha * win_proba[r_idx]
                    + (1.0 - _alpha) * (1.0 - win_proba_swap[r_idx])
                )
            else:
                r_win_prob = float(win_proba[r_idx])

            # Apply temperature scaling if fitted
            _T = getattr(self, "_temp_scale", None)
            if _T is not None:
                _p = np.clip(r_win_prob, 1e-9, 1 - 1e-9)
                _logit = np.log(_p / (1 - _p))
                r_win_prob = float(1.0 / (1.0 + np.exp(-_logit / _T)))

            b_win_prob = 1.0 - r_win_prob

            _thr = getattr(self, "_opt_threshold", 0.5)
            winner_is_red = r_win_prob >= _thr
            winner_name = r_name if winner_is_red else b_name
            winner_conf = max(r_win_prob, b_win_prob)

            confidence = abs(r_win_prob - 0.5) * 2

            pred = {
                "r_fighter": r_name,
                "b_fighter": b_name,
                "weight_class": weight_class,
                "gender": gender,
                "total_rounds": total_rounds,
                "winner": winner_name,
                "winner_conf": winner_conf,
                "r_win_prob": r_win_prob,
                "b_win_prob": b_win_prob,
                "confidence": confidence,
            }
            self.predictions.append(pred)

        return self.predictions

    def _compute_row_features(self, base):

        df1 = pd.DataFrame([base])

        if "event_date" not in df1.columns:
            df1["event_date"] = pd.Timestamp.today()

        df1 = self._recompute_derived_features(df1)

        result = {}
        for col in df1.columns:
            if df1[col].dtype == object:
                continue
            try:
                val = df1[col].iloc[0]
                fval = float(val)
                result[col] = 0.0 if (math.isnan(
                    fval) or math.isinf(fval)) else fval
            except (TypeError, ValueError):
                pass

        return result

    def _build_fight_feature_vector(
        self, r_name, b_name, weight_class, gender, total_rounds
    ):
        r_last = self._get_fighter_last_stats(r_name, "r")
        b_last = self._get_fighter_last_stats(b_name, "b")

        if r_last is None and b_last is None:
            return None

        if r_last is None:
            r_last = {}
        if b_last is None:
            b_last = {}

        base = {}
        base.update(r_last)
        base.update(b_last)
        base["weight_class"] = weight_class
        base["gender"] = gender
        base["total_rounds"] = total_rounds
        base["total_rounds_num"] = float(total_rounds)
        base["is_title_enc"] = 0.0
        base["gender_enc"] = 1.0 if str(gender).lower() == "women" else 0.0

        r_elo = self.feature_engineer.elo_get(r_name)
        b_elo = self.feature_engineer.elo_get(b_name)
        rg = self.feature_engineer.glicko2_get(r_name)
        bg = self.feature_engineer.glicko2_get(b_name)
        base["r_elo_pre_fight"] = r_elo
        base["b_elo_pre_fight"] = b_elo
        base["r_glicko_pre_r"] = rg[0]
        base["r_glicko_pre_rd"] = rg[1]
        base["r_glicko_pre_vol"] = rg[2]
        base["b_glicko_pre_r"] = bg[0]
        base["b_glicko_pre_rd"] = bg[1]
        base["b_glicko_pre_vol"] = bg[2]

        common = self.feature_engineer.get_common_opponent_features(
            r_name, b_name)
        rc = self.feature_engineer.get_fighter_cluster(r_name)
        bc_cluster = self.feature_engineer.get_fighter_cluster(b_name)
        mf = self.feature_engineer.get_style_matchup_features(rc, bc_cluster)
        base.update(common)
        base.update(mf)

        orig_feats = self._compute_row_features(base)

        for k, v in base.items():
            if k not in orig_feats:
                try:
                    orig_feats[k] = float(v) if v is not None else 0.0
                except (TypeError, ValueError):
                    orig_feats[k] = 0.0

        swap_base = {}
        for k, v in base.items():
            if k.startswith("r_"):
                swap_base["b_" + k[2:]] = v
            elif k.startswith("b_"):
                swap_base["r_" + k[2:]] = v
            else:
                swap_base[k] = v

        swap_feats = self._compute_row_features(swap_base)
        for k, v in swap_base.items():
            if k not in swap_feats:
                try:
                    swap_feats[k] = float(v) if v is not None else 0.0
                except (TypeError, ValueError):
                    swap_feats[k] = 0.0

        ordered_cols = self.feature_cols
        orig_arr = np.array(
            [orig_feats.get(col, 0.0) for col in ordered_cols], dtype=float
        )
        swap_arr = np.array(
            [swap_feats.get(col, 0.0) for col in ordered_cols], dtype=float
        )
        orig_arr = np.nan_to_num(orig_arr)
        swap_arr = np.nan_to_num(swap_arr)

        D = 0.5 * (orig_arr - swap_arr)
        Inv = 0.5 * (orig_arr + swap_arr)
        feat_decomposed = np.concatenate([D, Inv])
        return feat_decomposed, orig_arr

    def _get_fighter_last_stats(self, fighter_name, corner):
        r_rows = self.df[self.df["r_fighter"] == fighter_name]
        b_rows = self.df[self.df["b_fighter"] == fighter_name]

        stats = {}
        last_r = r_rows.iloc[-1] if len(r_rows) > 0 else None
        last_b = b_rows.iloc[-1] if len(b_rows) > 0 else None

        if last_r is not None and last_b is not None:
            date_r = last_r.get("event_date", pd.NaT)
            date_b = last_b.get("event_date", pd.NaT)
            if pd.notna(date_r) and pd.notna(date_b):
                row = last_r if date_r >= date_b else last_b
                src_corner = "r" if date_r >= date_b else "b"
            else:
                row = last_r
                src_corner = "r"
        elif last_r is not None:
            row = last_r
            src_corner = "r"
        elif last_b is not None:
            row = last_b
            src_corner = "b"
        else:
            return None

        for col in self.df.columns:
            if col.startswith(f"{src_corner}_"):
                base = col[2:]
                target_col = f"{corner}_{base}"
                val = row.get(col, 0)
                try:
                    val = float(val)
                    if math.isnan(val):
                        val = 0.0
                except (TypeError, ValueError):
                    pass  # preserve string values (e.g. fighter names)
                stats[target_col] = val

        return stats

    def export_predictions_to_excel(self, output_path):
        print_section("EXPORTING TO EXCEL")
        self._log(f"Writing to: {output_path}")

        if not self.predictions:
            self._log("No predictions to export.")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "UFC Predictions"

        header_fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )
        header_font = Font(bold=True, size=11, color="000000")
        header_align = Alignment(horizontal="left", vertical="center")
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = [
            "Red Fighter",
            "Blue Fighter",
            "Weight Class",
            "Winner",
            "Win%",
        ]

        pct_col_indices = {5}

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = border

        for row_idx, pred in enumerate(self.predictions, 2):
            winner_name = pred.get("winner") or (
                pred["r_fighter"]
                if pred.get("r_win_prob", 0.5) > pred.get("b_win_prob", 0.5)
                else pred["b_fighter"]
            )
            winner_conf = pred.get(
                "winner_conf",
                max(pred.get("r_win_prob", 0.5), pred.get("b_win_prob", 0.5)),
            )

            row_data = [
                pred["r_fighter"],
                pred["b_fighter"],
                pred["weight_class"],
                winner_name,
                winner_conf,
            ]

            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(
                    horizontal="left", vertical="center")
                cell.border = border
                if col_idx in pct_col_indices:
                    cell.number_format = "0.00%"

        for col_idx in range(1, len(headers) + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        max_len = max(max_len, len(str(cell.value or "")))
                    except Exception:
                        pass
            ws.column_dimensions[col_letter].width = max_len + 2

        wb.save(output_path)
        self._log(f"Saved predictions to {output_path}")
        print_metric("Fights exported:", len(self.predictions))


class UFCPredictorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("UFC FIGHT PREDICTOR")
        self.root.geometry("900x700")
        self.root.resizable(True, True)
        self.root.configure(bg="#1a1a2e")

        self.predictor = None
        self.data_path_var = tk.StringVar(value=DEFAULT_DATA_PATH)
        self.output_path_var = tk.StringVar(
            value=os.path.join(SCRIPT_DIR, "UFC_predictions.xlsx")
        )
        self.is_running = False

        self._build_ui()

    def _build_ui(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TButton", font=("Helvetica", 10, "bold"), padding=6)
        style.configure(
            "TLabel", background="#1a1a2e", foreground="#e0e0e0", font=("Helvetica", 10)
        )
        style.configure("TEntry", font=("Helvetica", 10))
        style.configure("TFrame", background="#1a1a2e")

        title_frame = tk.Frame(self.root, bg="#16213e", pady=12)
        title_frame.pack(fill="x")
        tk.Label(
            title_frame,
            text="UFC FIGHT PREDICTOR",
            font=("Helvetica", 20, "bold"),
            fg="#e94560",
            bg="#16213e",
        ).pack()

        main = tk.Frame(self.root, bg="#1a1a2e", padx=16, pady=10)
        main.pack(fill="both", expand=True)

        df_frame = tk.Frame(main, bg="#1a1a2e")
        df_frame.pack(fill="x", pady=4)
        tk.Label(
            df_frame,
            text="Data File:",
            width=12,
            anchor="w",
            bg="#1a1a2e",
            fg="#e0e0e0",
            font=("Helvetica", 10),
        ).pack(side="left")
        tk.Entry(
            df_frame,
            textvariable=self.data_path_var,
            font=("Helvetica", 10),
            bg="#0f3460",
            fg="white",
            insertbackground="white",
            relief="flat",
            width=60,
        ).pack(side="left", padx=4, fill="x", expand=True)
        tk.Button(
            df_frame,
            text="Browse",
            command=self._browse_data,
            font=("Helvetica", 9, "bold"),
            bg="#e94560",
            fg="white",
            relief="flat",
            padx=10,
            cursor="hand2",
        ).pack(side="left", padx=4)

        out_frame = tk.Frame(main, bg="#1a1a2e")
        out_frame.pack(fill="x", pady=4)
        tk.Label(
            out_frame,
            text="Output File:",
            width=12,
            anchor="w",
            bg="#1a1a2e",
            fg="#e0e0e0",
            font=("Helvetica", 10),
        ).pack(side="left")
        tk.Entry(
            out_frame,
            textvariable=self.output_path_var,
            font=("Helvetica", 10),
            bg="#0f3460",
            fg="white",
            insertbackground="white",
            relief="flat",
            width=60,
        ).pack(side="left", padx=4, fill="x", expand=True)
        tk.Button(
            out_frame,
            text="Browse",
            command=self._browse_output,
            font=("Helvetica", 9, "bold"),
            bg="#e94560",
            fg="white",
            relief="flat",
            padx=10,
            cursor="hand2",
        ).pack(side="left", padx=4)

        lbl_frame = tk.Frame(main, bg="#1a1a2e")
        lbl_frame.pack(fill="x", pady=(8, 2))
        tk.Label(
            lbl_frame,
            text="Enter Fights  (one per line: Red Fighter, Blue Fighter, Weight Class, Gender, Rounds)",
            bg="#1a1a2e",
            fg="#a0c4ff",
            font=("Helvetica", 9, "italic"),
        ).pack(anchor="w")

        self.fight_input = tk.Text(
            main,
            height=11,
            font=("Courier New", 10),
            bg="#0f3460",
            fg="white",
            insertbackground="white",
            relief="flat",
            wrap="word",
        )
        self.fight_input.pack(fill="both", expand=True, pady=4)

        btn_frame = tk.Frame(main, bg="#1a1a2e")
        btn_frame.pack(fill="x", pady=6)

        tk.Button(
            btn_frame,
            text="Load Sample",
            command=self._load_sample,
            font=("Helvetica", 10, "bold"),
            bg="#533483",
            fg="white",
            relief="flat",
            padx=14,
            cursor="hand2",
        ).pack(side="left", padx=4)
        tk.Button(
            btn_frame,
            text="Clear",
            command=self._clear_input,
            font=("Helvetica", 10, "bold"),
            bg="#444466",
            fg="white",
            relief="flat",
            padx=14,
            cursor="hand2",
        ).pack(side="left", padx=4)

        self.run_btn = tk.Button(
            btn_frame,
            text="Generate Predictions",
            command=self._run_predictions,
            font=("Helvetica", 11, "bold"),
            bg="#e94560",
            fg="white",
            relief="flat",
            padx=20,
            cursor="hand2",
        )
        self.run_btn.pack(side="right", padx=4)

    def _browse_data(self):
        path = filedialog.askopenfilename(
            title="Select Fight Data CSV",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialdir=SCRIPT_DIR,
        )
        if path:
            self.data_path_var.set(path)

    def _browse_output(self):
        path = filedialog.asksaveasfilename(
            title="Save Predictions As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialdir=SCRIPT_DIR,
            initialfile="UFC_predictions.xlsx",
        )
        if path:
            self.output_path_var.set(path)

    def _load_sample(self):
        sample = (
            "Max Holloway,Dustin Poirier,Lightweight,Men,5\n"
            "Ilia Topuria,Charles Oliveira,Lightweight,Men,5\n"
            "Tatiana Suarez,Amanda Lemos,Strawweight,Women,3"
        )
        self.fight_input.delete("1.0", tk.END)
        self.fight_input.insert("1.0", sample)

    def _clear_input(self):
        self.fight_input.delete("1.0", tk.END)

    def _run_predictions(self):
        if self.is_running:
            return
        thread = threading.Thread(target=self._run_predictions_thread, daemon=True)
        thread.start()


    def _run_predictions_thread(self):
        self.is_running = True
        self.run_btn.config(state="disabled", text="Running...")
        try:
            data_path = self.data_path_var.get().strip()
            output_path = self.output_path_var.get().strip()

            if not os.path.isfile(data_path):
                messagebox.showerror("Error", f"Data file not found:\n{data_path}")
                return

            fights_text = self.fight_input.get("1.0", tk.END).strip()
            if not fights_text:
                messagebox.showerror("Error", "Please enter at least one fight.")
                return

            fights = self._parse_fights(fights_text)
            if not fights:
                return

            self.predictor = UFCPredictor(data_path=data_path)
            self.predictor.run_with_cache()

            self.predictor.predict_upcoming_fights(fights)
            self.predictor.export_predictions_to_excel(output_path)

            messagebox.showinfo(
                "Complete",
                f"Predictions generated successfully!\n\nSaved to:\n{output_path}",
            )

        except Exception as e:
            tb = traceback.format_exc()
            print(f"Error: {e}")
            messagebox.showerror("Error", f"An error occurred:\n\n{e}\n\n{tb[-500:]}")
        finally:
            self.is_running = False
            self.run_btn.config(state="normal", text="Generate Predictions")

    def _parse_fights(self, text):
        fights = []
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        for i, line in enumerate(lines, 1):
            parts = [p.strip() for p in line.split(",")]
            if len(parts) != 5:
                messagebox.showerror(
                    "Parse Error",
                    f"Line {i}: Expected 5 comma-separated fields.\n"
                    f"Format: Red Fighter, Blue Fighter, Weight Class, Gender, Rounds\n"
                    f"Got: {line}",
                )
                return []
            r_fighter, b_fighter, weight_class, gender, rounds_str = parts
            try:
                total_rounds = int(rounds_str)
            except ValueError:
                messagebox.showerror(
                    "Parse Error",
                    f"Line {i}: Rounds must be an integer, got '{rounds_str}'",
                )
                return []
            fights.append(
                {
                    "r_fighter": r_fighter,
                    "b_fighter": b_fighter,
                    "weight_class": weight_class,
                    "gender": gender,
                    "total_rounds": total_rounds,
                }
            )
        return fights


def main():
    mp.freeze_support()

    print_section("UFC FIGHT PREDICTOR")
    print_step(f"Python: {sys.version.split()[0]}")
    print_step(
        f"Libraries: XGB={HAS_XGB}, LGB={HAS_LGB}, CAT={HAS_CAT}, Optuna={HAS_OPTUNA}"
    )
    print_step(f"CPU cores: {mp.cpu_count()}  |  Safe n_jobs: {SAFE_N_JOBS}")
    print_step(f"Script dir: {SCRIPT_DIR}")
    print_step(f"Default data: {DEFAULT_DATA_PATH}")

    root = tk.Tk()
    app = UFCPredictorGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
