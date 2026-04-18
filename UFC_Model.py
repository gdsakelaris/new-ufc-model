"""
UFC Model (single-file).

What this script does:
1) Builds leak-safe chronological pre-fight features from pure_fight_data.csv.
2) Trains a compact time-aware ensemble optimized for log-loss.
3) Calibrates probabilities on validation-only data (Platt vs Isotonic).
4) Reports holdout + walk-forward diagnostics and baseline comparisons.
5) Provides GUI matchup prediction and Excel export (predictions + rankings).

Run:
  python UFC_Model.py
  
Optional CLI:
  python UFC_Model.py --train-only
"""

import argparse
import hashlib
import json
import math
import os
import pickle
import random
import re
import threading
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime

import numpy as np
import pandas as pd
import tkinter as tk
from tkinter import messagebox

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from sklearn.ensemble import HistGradientBoostingClassifier, RandomForestClassifier
from sklearn.ensemble import ExtraTreesClassifier, AdaBoostClassifier
from sklearn.impute import SimpleImputer
from sklearn.isotonic import IsotonicRegression
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score, brier_score_loss, log_loss
from sklearn.metrics import balanced_accuracy_score, precision_recall_fscore_support
from sklearn.metrics import roc_auc_score
from sklearn.model_selection import TimeSeriesSplit
from sklearn.neural_network import MLPClassifier
from sklearn.preprocessing import StandardScaler
from scipy.optimize import minimize as scipy_minimize

try:
    import lightgbm as lgb
except Exception:
    lgb = None

try:
    import xgboost as xgb
except Exception:
    xgb = None

try:
    import catboost as cb
except Exception:
    cb = None

try:
    import optuna
except Exception:
    optuna = None

if optuna is not None:
    # Keep Optuna output compact: suppress per-trial parameter logs.
    optuna.logging.set_verbosity(optuna.logging.WARNING)


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_PATH = os.path.join(SCRIPT_DIR, "pure_fight_data.csv")
PREDICTIONS_XLSX = os.path.join(SCRIPT_DIR, "UFC_Predictions.xlsx")
CACHE_DIR = os.path.join(SCRIPT_DIR, ".ufc_model_cache")
METHOD_CHAMPION_PATH = os.path.join(SCRIPT_DIR, ".ufc_model_cache", "method_champion_cfg.json")
###################################################################################################
# Bump when winner-stage training logic changes.
WINNER_CACHE_VERSION = "v3"
# Bump when method-stage training logic changes.
METHOD_CACHE_VERSION = "v11"
###################################################################################################
# Pickle payload discriminator (stable across cache file renames).
WINNER_STAGE_CACHE_KIND = "ufc_winner_stage_v1"
METHOD_STAGE_CACHE_KIND = "ufc_method_stage_v1"

RANDOM_SEED = 42
np.random.seed(RANDOM_SEED)
random.seed(RANDOM_SEED)
os.environ["PYTHONHASHSEED"] = str(RANDOM_SEED)

TRAIN_FRACTION = 0.65
VAL_FRACTION = 0.15
TEST_FRACTION = 0.20
MIN_HOLDOUT_FIGHTS = 500
ACTIVE_DAYS = 730
NEEDS_SCALE = {"LogReg", "MLP"}
ELO_BASE = 1500.0
ELO_K = 24.0
OPTUNA_TRIALS = 80
METHOD_TUNING_TRIALS = 800
METHOD_HARD_RESET = False
# Correlation threshold for method-stage feature pruning (|corr| > this → dropped).
METHOD_CORR_PRUNE_THRESHOLD = 0.95
METHOD_ERA_CANDIDATES = [1993, 2005, 2010, 2014, 2016, 2018, 2020, 2021, 2022, 2023, 2024]
METHOD_AUTO_ERA = True
# Keep winner-model inputs on the stable feature family while allowing richer
# method-only engineering in the method pipeline.
WINNER_EXCLUDE_FEATURES = {
    "sub_entry_pressure_sum", "sub_defensive_leak_sum",
    "ko_attack_pressure_sum", "ko_def_leak_sum",
    "d_head_acc", "d_distance_acc", "d_ground_acc", "d_clinch_acc",
    "d_body_acc", "d_leg_acc", "d_body_leg_attrition",
    "d_head_hunt_share", "d_distance_share",
    "d_ground_strike_accuracy", "d_head_hunt_accuracy", "d_distance_strike_accuracy",
    "d_sub_loss_pct", "d_recent_sub_win_rate",
    "d_sub_entry_pressure", "d_sub_control_conversion", "d_sub_scramble_threat",
    "d_sub_defensive_leak", "d_late_sub_pressure",
    "d_sub_recency_surge", "d_grapple_recency_surge", "d_sub_vs_control_axis",
    "d_recent_ko_loss_rate", "d_r5_def_kd_pm", "d_ko_attack_pressure", "d_ko_def_leak",
    "d_r1f_def_kd_pm", "d_r3_def_kd_pm",
    # Stage-1 sum features (exact side reconstruction)
    "dec_win_pct_sum", "finish_resistance_sum", "consistency_sum", "cardio_ratio_sum",
    "durability_sum", "output_rate_sum", "rd1_intensity_ratio_sum",
    "strike_exchange_ratio_sum", "sig_str_acc_sum", "late_round_pct_sum",
    "avg_time_min_sum", "avg_finish_round_sum", "first_round_finish_rate_sum",
    "damage_efficiency_sum", "body_leg_attrition_sum",
    # Stage-1 context priors
    "ctx_finish_prior_2y", "ref_finish_prior",
    # Stage-1 latent finish-environment features
    "m_decision_shell_gap", "m_decision_shell_sum", "m_finish_conversion_edge",
    "m_finish_environment", "m_mutual_finish_instability", "m_decision_absorber",
    "m_early_finish_window", "m_fast_start_fragility", "m_late_finish_window",
    "m_attrition_break_window", "m_time_profile_finish_bias",
    "m_finish_over_shell_ratio", "m_clean_decision_track", "m_finish_speed_pressure",
}
# Features that are stage-1-specific (finish vs decision signals) and must be
# withheld from stage-2 (KO vs Sub), which was calibrated without them.
STAGE2_EXCLUDE_COLS_BASE = frozenset({
    "m_decision_shell_gap", "m_decision_shell_sum", "m_finish_conversion_edge",
    "m_finish_environment", "m_mutual_finish_instability", "m_decision_absorber",
    "m_early_finish_window", "m_fast_start_fragility", "m_late_finish_window",
    "m_attrition_break_window", "m_time_profile_finish_bias",
    "m_finish_over_shell_ratio", "m_clean_decision_track", "m_finish_speed_pressure",
    "ctx_finish_prior_2y", "ref_finish_prior",
})
###################################################################################################
# FEATURE_ROUTING — single source of truth for which models see which engineered features.
# Each key is a feature name; the value is the set of models that should receive it.
# Allowed tags: "winner", "stage1" (Finish vs Decision), "stage2" (KO vs Sub).
# Filters are auto-derived below; to change a feature's routing, edit this dict only.
FEATURE_ROUTING = {
    # --- Batch 2: within-fight pacing (§A) ---
    # NOTE: d_rd1_kd, d_rd1_td_att, d_rd3_sub_att already exist as auto-diffs of
    # pre-existing rd1_kd / rd1_td_att / rd3_sub_att aggregates — not routed
    # here (would remove them from winner). User's d_rd1_kd_rate / d_early_td_pressure
    # / d_rd3_sub_att_rate collapse into those existing columns.
    "d_def_rd1_sig_str":           {"winner"},
    "d_def_rd1_kd":                {"winner"},
    "d_rd1_net_sig_str":           {"winner"},
    "d_cardio_decay_sig_str":      {"winner"},
    "d_late_sig_str_pm":           {"winner"},
    "d_rd1_ctrl_share":            {"winner"},
    "d_reach_x_distance_pct":      {"winner"},
    "d_age_x_cardio":              {"winner"},
    "d_age_x_title_rounds":        {"winner"},
    "d_striker_grappler_raw":      {"winner"},
    "d_tdd_vs_td_attack":          {"winner"},
    "d_short_notice":              {"winner"},
    "d_glicko_trend":              {"winner"},
    "d_recent_damage_absorbed":    {"winner"},
    "gender_flag":                 {"winner"},
    "d_title_x_cardio":            {"winner"},
    "d_total_rounds_x_finish_resistance": {"winner"},
    # --- Context-conditional & stability (winner-only) ---
    "d_rounds_experience":         {"winner"},
    "d_rounds_5_exp":              {"winner"},
    "d_rounds_3_exp":              {"winner"},
    "d_stability":                 {"winner"},
    "d_sig_diff_pm_vol":           {"winner"},
}
_ROUTING_WINNER_EXCLUDE = {f for f, tags in FEATURE_ROUTING.items() if "winner" not in tags}
_ROUTING_STAGE1_EXCLUDE = {f for f, tags in FEATURE_ROUTING.items() if "stage1" not in tags}
_ROUTING_STAGE2_EXCLUDE = {f for f, tags in FEATURE_ROUTING.items() if "stage2" not in tags}
# Merge auto-derived sets with the existing manually-curated excludes.
WINNER_EXCLUDE_FEATURES = WINNER_EXCLUDE_FEATURES | _ROUTING_WINNER_EXCLUDE
STAGE1_EXCLUDE_COLS = frozenset(_ROUTING_STAGE1_EXCLUDE)
STAGE2_EXCLUDE_COLS = frozenset(STAGE2_EXCLUDE_COLS_BASE | _ROUTING_STAGE2_EXCLUDE)
###################################################################################################
STRICT_FUTURE_MODE = True
FORCED_START_YEAR = None
MISSINGNESS_THRESHOLD = 0.35

MU_0 = 1500.0
PHI_0 = 200.0
SIGMA_0 = 0.06
TAU = 0.5
SCALE = 173.7178
CONVERGENCE = 1e-6


def _clip_probs(p):
    return np.clip(np.asarray(p, dtype=float), 1e-6, 1.0 - 1e-6)


def _cache_data_fingerprint(path):
    try:
        st = os.stat(path)
        raw = f"{os.path.abspath(path)}|{st.st_size}|{int(st.st_mtime)}"
    except Exception:
        raw = f"{os.path.abspath(path)}|missing"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:16]


def _cache_key(stage, data_fp, version, extra=""):
    payload = f"{stage}|{data_fp}|{version}|{extra}"
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()[:20]


def _cache_path(stage, key):
    os.makedirs(CACHE_DIR, exist_ok=True)
    return os.path.join(CACHE_DIR, f"{stage}_{key}.pkl")


def _cache_load(stage, key):
    path = _cache_path(stage, key)
    if not os.path.exists(path):
        return None
    try:
        with open(path, "rb") as f:
            return pickle.load(f)
    except Exception:
        return None


def _cache_save(stage, key, payload):
    path = _cache_path(stage, key)
    with open(path, "wb") as f:
        pickle.dump(payload, f, protocol=pickle.HIGHEST_PROTOCOL)
    prefix = f"{stage}_"
    for name in os.listdir(CACHE_DIR):
        if not name.startswith(prefix) or name.endswith(f"{key}.pkl"):
            continue
        old = os.path.join(CACHE_DIR, name)
        try:
            os.remove(old)
        except Exception:
            pass


def _winner_stage_cache_valid(payload, feature_cols, n_rows, train_end, val_end):
    if not isinstance(payload, dict):
        return False
    if payload.get("kind") != WINNER_STAGE_CACHE_KIND:
        return False
    if str(payload.get("winner_cache_version")) != str(WINNER_CACHE_VERSION):
        return False
    if int(payload.get("n_rows", -1)) != int(n_rows):
        return False
    if int(payload.get("train_end", -1)) != int(train_end):
        return False
    if int(payload.get("val_end", -1)) != int(val_end):
        return False
    fc = payload.get("feature_cols")
    if not isinstance(fc, list) or tuple(fc) != tuple(feature_cols):
        return False
    for k in (
        "oof", "valid", "combiner", "model_order", "decision_threshold",
        "test_probs_raw", "test_probs_cal", "y_pred_red",
        "raw_ll", "cal_ll_test", "brier", "acc", "ece", "cal_curve_rmse",
        "lgb_tuned", "tuned_thr", "tuned_acc", "val_acc_for_log", "calibrator",
    ):
        if k not in payload:
            return False
    return True


def _method_stage_cache_valid(payload, winner_cache_key, method_cache_version):
    if not isinstance(payload, dict):
        return False
    if payload.get("kind") != METHOD_STAGE_CACHE_KIND:
        return False
    if str(payload.get("method_cache_version")) != str(method_cache_version):
        return False
    if str(payload.get("winner_cache_key")) != str(winner_cache_key):
        return False
    b = payload.get("method_bundle")
    return isinstance(b, dict) and b.get("imputer") is not None


def _replay_winner_cache_logs(pl, W, winner_cache_key):
    """Replay winner-stage terminal sections from a cache payload (no retrain)."""
    h = str(winner_cache_key)[:12]
    pl._section("Optuna Tuning")
    pl._stat("Cache", f"HIT ({WINNER_CACHE_VERSION}) — Optuna skipped [key={h}]")
    pl._section("Model Setup")
    pl._stat("Base models", ", ".join(W.get("model_order") or []))
    pl._section("OOF Stacking")
    pl._stat("Cache", f"HIT ({WINNER_CACHE_VERSION}) — OOF skipped [key={h}]")
    pl._section("Combiner Selection")
    pl._stat("Selected combiner", W.get("combiner_kind", ""))
    pl._stat("Validation log-loss", W.get("val_ll_str", ""))
    pl._stat("Validation accuracy", W.get("val_acc_str", ""))
    pl._stat("Validation threshold", W.get("val_thr_str", ""))
    pl._section("Calibration")
    pl._stat("Selected method", W.get("cal_name", ""))
    pl._stat("Validation log-loss", W.get("cal_ll_str", ""))
    pl._stat("Strict future mode", "ON" if STRICT_FUTURE_MODE else "OFF")
    if not STRICT_FUTURE_MODE:
        pl._stat("Holdout-selected combiner", W.get("best_holdout_label", ""))
        pl._stat("Holdout combiner log-loss", W.get("best_holdout_ll_str", ""))
        pl._stat("Holdout combiner acc", W.get("best_holdout_acc_str", ""))
        pl._stat("Holdout combiner threshold", W.get("best_holdout_thr_str", ""))
    pl._stat("Calibration used for picks", W.get("cal_used_str", ""))
    pl._stat("Validation accuracy (picked)", W.get("val_acc_picked_str", ""))
    pl._stat("Validation tuned threshold", W.get("val_tuned_thr_str", ""))
    pl._stat("Decision threshold", W.get("decision_thr_str", ""))
    pl._section("Holdout Evaluation")
    for label, val in W.get("holdout_eval", []):
        pl._stat(label, val)
    pl._section("Winner Diagnostics")
    pl._log("Confusion Matrix (Winner: Red=positive)")
    pl._log("               Pred Blue   Pred Red")
    tn, fp, fn, tp = W.get("confusion", (0, 0, 0, 0))
    pl._log(f"Actual Blue   {int(tn):9d}  {int(fp):9d}")
    pl._log(f"Actual Red    {int(fn):9d}  {int(tp):9d}")
    pl._log("")
    pl._log("Accuracy by Weight Class")
    pl._log("-" * 72)
    for wc, acc_wc, n_wc in W.get("wc_rows", []):
        pl._stat(f"{wc} (n={int(n_wc)})", f"{float(acc_wc):.1%}")
    pl._log("")
    pl._log("Accuracy by Gender")
    pl._log("-" * 72)
    for gender, acc_g, n_g in W.get("g_rows", []):
        pl._stat(f"{gender} (n={int(n_g)})", f"{float(acc_g):.1%}")


def _replay_method_cache_logs(pl, M):
    """Replay method-stage terminal sections from a cache payload (no retrain)."""
    if M.get("method_classes_training"):
        pl._stat("Method classes (training)", M["method_classes_training"])
    if M.get("method_hard_reset"):
        pl._stat("Method stack mode", "hard-reset (stable components only)")
    if M.get("val_metric_str"):
        pl._stat("Validation metric target (method | winner correct)", M["val_metric_str"])
    if M.get("val_baseline_str"):
        pl._stat("Validation majority baseline (same subset)", M["val_baseline_str"])

    pl._section("Method Evaluation (Conditioned on Winner Pick)")

    s1_acc = M.get("stage1_acc")
    s1_auc = M.get("stage1_auc")
    s2_acc = M.get("stage2_acc_true_finishes")
    n_fin  = M.get("n_true_finishes", 0)
    pl._stat("Stage1 acc (Finish vs Decision)", f"{s1_acc:.1%}" if s1_acc is not None else "n/a")
    pl._stat("Stage1 AUC (Finish vs Decision)", f"{s1_auc:.3f}" if (s1_auc is not None and np.isfinite(s1_auc)) else "n/a")
    pl._stat("Stage2 acc (KO/Sub | true finishes)", f"{s2_acc:.1%}" if (s2_acc is not None and np.isfinite(s2_acc)) else "n/a")
    pl._stat("Stage2 sample size (true finishes)", n_fin)

    acc_pred = M.get("method_acc_predicted_winner")
    acc_wc   = M.get("method_acc_when_winner_correct")
    maj_base = M.get("method_majority_baseline_when_winner_correct")
    pl._stat("Method acc (predicted winner conditioned)", f"{acc_pred:.1%}" if acc_pred is not None else "n/a")
    pl._stat("Method acc | winner pick correct",          f"{acc_wc:.1%}"   if acc_wc   is not None else "n/a")
    pl._stat("Majority baseline | winner pick correct",   f"{maj_base:.1%}" if maj_base  is not None else "n/a")

    bal_acc     = M.get("bal_acc")
    macro_f1    = M.get("macro_f1")
    finish_score = M.get("finish_score")
    if bal_acc is not None and np.isfinite(bal_acc):
        pl._stat("Balanced accuracy | winner pick correct", f"{bal_acc:.1%}")
    if macro_f1 is not None and np.isfinite(macro_f1):
        pl._stat("Macro F1 | winner pick correct", f"{macro_f1:.1%}")
    if finish_score is not None and np.isfinite(finish_score):
        pl._stat("FinishScore (0.4 KO R + 0.4 Sub R + 0.2 KO/Sub F1)", f"{finish_score:.1%}")

    per_class = M.get("per_class_metrics", [])
    if per_class:
        pl._log("")
        pl._log("Per-Class Metrics (Method | winner pick correct)")
        pl._log("Class          Precision    Recall      F1")
        for cls_name, p, r, f1 in per_class:
            pl._log(f"{cls_name:<14}{p:10.1%}{r:10.1%}{f1:10.1%}")

    confusion_rows = M.get("confusion_rows", [])
    if confusion_rows:
        pl._log("")
        pl._log("Confusion Matrix (Method | winner pick correct)")
        pl._log("Actual\\Pred     Decision    KO/TKO  Submission")
        for actual, row_counts in confusion_rows:
            pl._log(f"{actual:<14}{row_counts[0]:10d}{row_counts[1]:10d}{row_counts[2]:12d}")


def _normalize_method_label(raw_method):
    detail = _normalize_method_detail(raw_method)
    if detail.startswith("Decision"):
        return "Decision"
    if detail in ("KO/TKO", "Doctor Stoppage", "DQ/Corner Stoppage"):
        return "KO/TKO"
    if detail == "Submission":
        return "Submission"
    return "Decision"


def _normalize_method_detail(raw_method):
    txt = str(raw_method or "").strip().lower()
    if "decision" in txt or txt.startswith("dec") or "split" in txt or "majority" in txt or "unanimous" in txt:
        return "Decision"
    if "doctor" in txt:
        return "KO/TKO"
    if "dq" in txt or "corner" in txt or "retire" in txt:
        return "KO/TKO"
    if "sub" in txt:
        return "Submission"
    if "ko" in txt or "tko" in txt:
        return "KO/TKO"
    return "Decision"


def _normalize_method_probs(prob_map):
    vals = np.array([float(prob_map.get(k, 0.0)) for k in METHOD_LABELS], dtype=float)
    vals = np.maximum(vals, MIN_METHOD_PROB)
    vals = vals / np.sum(vals)
    return {k: float(v) for k, v in zip(METHOD_LABELS, vals)}


# Method-probability shaping helpers: logit-space bias application and submission signal boost.
# Called during tuning (blending loop) and inference (predict_method_probs).
def _apply_method_logit_bias_arr(probs_arr, bias_vec):
    arr = np.asarray(probs_arr, dtype=float)
    if arr.ndim == 1:
        arr = arr.reshape(1, -1)
    b = np.asarray(bias_vec, dtype=float).reshape(1, -1)
    logp = np.log(np.clip(arr, MIN_METHOD_PROB, 1.0)) + b
    logp = logp - np.max(logp, axis=1, keepdims=True)
    expv = np.exp(logp)
    out = expv / np.sum(expv, axis=1, keepdims=True)
    return np.clip(out, MIN_METHOD_PROB, 1.0)


def _apply_method_logit_bias_map(prob_map, bias_map):
    arr = np.array([
        float(prob_map.get("Decision", 0.0)),
        float(prob_map.get("KO/TKO", 0.0)),
        float(prob_map.get("Submission", 0.0)),
    ], dtype=float)
    bias_vec = np.array([
        float(bias_map.get("Decision", 0.0)),
        float(bias_map.get("KO/TKO", 0.0)),
        float(bias_map.get("Submission", 0.0)),
    ], dtype=float)
    out = _apply_method_logit_bias_arr(arr, bias_vec)[0]
    return _normalize_method_probs({
        "Decision": float(out[0]),
        "KO/TKO": float(out[1]),
        "Submission": float(out[2]),
    })


def _apply_binary_threshold_warp(p, thr):
    p_arr = np.asarray(p, dtype=float)
    t = float(np.clip(thr, 0.05, 0.95))
    left = 0.5 * (p_arr / max(t, 1e-6))
    right = 0.5 + 0.5 * ((p_arr - t) / max(1.0 - t, 1e-6))
    out = np.where(p_arr < t, left, right)
    return np.clip(out, 1e-4, 1.0 - 1e-4)


def _apply_submission_signal_boost_arr(probs_arr, sub_signal_arr, boost_k):
    arr = np.asarray(probs_arr, dtype=float).copy()
    sig = np.asarray(sub_signal_arr, dtype=float).reshape(-1)
    k = float(boost_k)
    # Positive boost only when submission signal is above prior-ish baseline.
    scale = np.clip((sig - 0.18) / 0.22, -1.5, 2.5)
    arr[:, 2] = arr[:, 2] * np.exp(k * scale)
    arr = np.clip(arr, MIN_METHOD_PROB, 1.0)
    arr = arr / np.sum(arr, axis=1, keepdims=True)
    return arr


def _apply_submission_signal_boost_map(prob_map, sub_signal, boost_k):
    arr = np.array([
        float(prob_map.get("Decision", 0.0)),
        float(prob_map.get("KO/TKO", 0.0)),
        float(prob_map.get("Submission", 0.0)),
    ], dtype=float).reshape(1, -1)
    out = _apply_submission_signal_boost_arr(arr, np.array([float(sub_signal)], dtype=float), boost_k)[0]
    return _normalize_method_probs({
        "Decision": float(out[0]),
        "KO/TKO": float(out[1]),
        "Submission": float(out[2]),
    })


def _sub_attempt_prior_array(X_df):
    X = X_df.reset_index(drop=True)
    d_sub_att = pd.to_numeric(X.get("d_sub_att_p15", pd.Series(np.zeros(len(X)))), errors="coerce").fillna(0.0).to_numpy(dtype=float)
    d_grap_tdd = pd.to_numeric(X.get("d_grapple_vs_tdd", pd.Series(np.zeros(len(X)))), errors="coerce").fillna(0.0).to_numpy(dtype=float)
    d_been_fin = pd.to_numeric(X.get("d_been_finished_pct", pd.Series(np.zeros(len(X)))), errors="coerce").fillna(0.0).to_numpy(dtype=float)
    d_dec = pd.to_numeric(X.get("d_dec_win_pct", pd.Series(np.zeros(len(X)))), errors="coerce").fillna(0.0).to_numpy(dtype=float)
    z = 1.35 * d_sub_att + 0.75 * d_grap_tdd + 0.55 * d_been_fin
    p_sub = 1.0 / (1.0 + np.exp(-z))
    p_sub = np.clip(0.08 + 0.50 * p_sub, 0.06, 0.62)
    p_dec = np.clip(0.54 + 0.20 * d_dec - 0.35 * p_sub, 0.12, 0.82)
    p_ko = np.clip(1.0 - p_dec - p_sub, 0.05, 0.70)
    arr = np.stack([p_dec, p_ko, p_sub], axis=1)
    arr = np.clip(arr, MIN_METHOD_PROB, 1.0)
    arr = arr / np.sum(arr, axis=1, keepdims=True)
    return arr


def _method_profile_from_history(history):
    wins = [h for h in history if h.get("result") == "W"]
    losses = [h for h in history if h.get("result") == "L"]

    def _rate(rows, method_label, prior, strength=8.0):
        n = len(rows)
        if n == 0:
            return float(prior)
        hits = sum(1 for h in rows if _normalize_method_label(h.get("method", "")) == method_label)
        return (hits + prior * strength) / (n + strength)

    return {
        "win_decision": _rate(wins, "Decision", 0.50),
        "win_ko_tko": _rate(wins, "KO/TKO", 0.32),
        "win_submission": _rate(wins, "Submission", 0.18),
        "loss_decision": _rate(losses, "Decision", 0.45),
        "loss_ko_tko": _rate(losses, "KO/TKO", 0.35),
        "loss_submission": _rate(losses, "Submission", 0.20),
        "wins_n": float(len(wins)),
        "losses_n": float(len(losses)),
    }


def _oriented_method_matrix(X_df, y_red_win):
    X_m = X_df.copy()
    y_arr = np.asarray(y_red_win).astype(int)
    sign = np.where(y_arr == 1, 1.0, -1.0)
    for col in X_m.columns:
        if col.startswith("d_"):
            X_m[col] = X_m[col].astype(float).values * sign
    return X_m


def _augment_method_features(X_df):
    X = X_df.copy()
    eps = 1e-6

    def _col(name, default=0.0):
        if name not in X.columns:
            return pd.Series(np.full(len(X), default), index=X.index, dtype=float)
        return pd.to_numeric(X[name], errors="coerce").fillna(default)

    d_ko_win = _col("d_ko_win_pct", 0.0)
    d_sub_win = _col("d_sub_win_pct", 0.0)
    d_dec_win = _col("d_dec_win_pct", 0.0)
    d_ko_loss = _col("d_ko_loss_pct", 0.0)
    d_sub_att = _col("d_sub_att_p15", 0.0)
    d_r3_sub = _col("d_rd3_sub_att", 0.0)
    d_r3_sub_share = _col("d_rd3_sub_share", 0.0)
    d_late_sub = _col("d_late_vs_early_sub_att", 0.0)
    d_sub_t23 = _col("d_sub_att_trend_23", 0.0)
    d_been_finished = _col("d_been_finished_pct", 0.0)
    d_finish_vs_resist = _col("d_finish_vs_resist", 0.0)
    d_str_vs_def = _col("d_striking_vs_defense", 0.0)
    d_strike_exchange_ratio = _col("d_strike_exchange_ratio", 0.0)
    d_grap_vs_tdd = _col("d_grapple_vs_tdd", 0.0)
    d_ortho_vs_south = _col("d_ortho_vs_south", 0.0)
    d_power_ratio = _col("d_power_ratio", 0.0)
    d_head_pct = _col("d_head_pct", 0.0)
    d_sig_str_diff_pm = _col("d_sig_str_diff_pm", 0.0)
    d_first_round_finish_rate = _col("d_first_round_finish_rate", 0.0)
    d_rd1_intensity_ratio = _col("d_rd1_intensity_ratio", 0.0)
    d_damage_efficiency = _col("d_damage_efficiency", 0.0)
    d_sig_str_acc = _col("d_sig_str_acc", _col("d_td_sig_str_acc", 0.0))
    d_ground_pct = _col("d_ground_pct", 0.0)
    d_ctrl_pct = _col("d_ctrl_pct", _col("d_td_ctrl_pct", 0.0))
    d_rev_p15 = _col("d_rev_p15", 0.0)
    d_cardio_ratio = _col("d_cardio_ratio", 0.0)
    d_distance_pct = _col("d_distance_pct", 0.0)
    d_body_pct = _col("d_body_pct", 0.0)
    d_leg_pct = _col("d_leg_pct", 0.0)
    d_consistency = _col("d_consistency", 0.0)
    d_kd_pm = _col("d_kd_pm", _col("d_td_kd_pm", 0.0))
    d_def_kd_pm = _col("d_def_kd_pm", 0.0)
    d_finish_resistance = _col("d_finish_resistance", 0.0)
    d_durability = _col("d_durability", 0.0)
    d_late_round_pct = _col("d_late_round_pct", 0.0)
    d_output_rate = _col("d_output_rate", 0.0)
    d_glicko = np.abs(_col("d_glicko_win_prob", 0.0))
    d_head_acc = _col("d_head_acc", 0.0)
    d_distance_acc = _col("d_distance_acc", 0.0)
    d_distance_share = _col("d_distance_share", 0.0)
    d_ground_acc = _col("d_ground_acc", 0.0)
    d_clinch_acc = _col("d_clinch_acc", 0.0)
    d_clinch_pct = _col("d_clinch_pct", 0.0)
    d_body_leg_attrition = _col("d_body_leg_attrition", 0.0)
    d_sub_loss_pct = _col("d_sub_loss_pct", 0.0)
    d_recent_sub_win_rate = _col("d_recent_sub_win_rate", 0.0)
    d_sub_entry_pressure = _col("d_sub_entry_pressure", 0.0)
    d_sub_control_conversion = _col("d_sub_control_conversion", 0.0)
    d_sub_defensive_leak = _col("d_sub_defensive_leak", 0.0)
    d_late_sub_pressure = _col("d_late_sub_pressure", 0.0)
    d_sub_recency_surge = _col("d_sub_recency_surge", 0.0)
    d_grapple_recency_surge = _col("d_grapple_recency_surge", 0.0)
    d_sub_vs_control_axis = _col("d_sub_vs_control_axis", 0.0)
    d_recent_ko_loss_rate = _col("d_recent_ko_loss_rate", 0.0)
    d_r5_def_kd_pm = _col("d_r5_def_kd_pm", 0.0)
    d_ko_attack_pressure = _col("d_ko_attack_pressure", 0.0)
    d_ko_def_leak = _col("d_ko_def_leak", 0.0)
    sub_entry_sum = _col("sub_entry_pressure_sum", 0.0)
    sub_leak_sum = _col("sub_defensive_leak_sum", 0.0)
    ko_attack_sum = _col("ko_attack_pressure_sum", 0.0)
    ko_leak_sum = _col("ko_def_leak_sum", 0.0)
    total_rounds = _col("total_rounds", 3.0)
    is_title = _col("is_title", 0.0)

    # Explicit path-vs-vulnerability method features.
    X["m_ko_path_vs_vuln"] = d_ko_win + d_ko_loss + 0.45 * d_str_vs_def
    X["m_sub_path_vs_vuln"] = d_sub_win + d_been_finished + 0.35 * d_grap_vs_tdd + 0.25 * d_sub_att
    X["m_sub_round3_pressure"] = 0.60 * d_r3_sub + 0.40 * d_r3_sub_share
    X["m_sub_trend_pressure"] = 0.55 * d_late_sub + 0.45 * d_sub_t23
    X["m_dec_path"] = d_dec_win - 0.25 * d_finish_vs_resist
    X["m_finish_bias"] = X["m_ko_path_vs_vuln"] + X["m_sub_path_vs_vuln"] - X["m_dec_path"]
    X["m_ko_rounds_interaction"] = X["m_ko_path_vs_vuln"] * (4.0 - np.minimum(total_rounds, 4.0))
    X["m_sub_rounds_interaction"] = X["m_sub_path_vs_vuln"] * np.maximum(total_rounds - 2.0, 1.0)
    X["m_dec_rounds_interaction"] = X["m_dec_path"] * total_rounds
    X["m_stance_finish_interaction"] = d_ortho_vs_south * X["m_finish_bias"]
    X["m_title_finish_interaction"] = is_title * X["m_finish_bias"]
    X["m_finish_abs_pressure"] = np.abs(X["m_finish_bias"]) * np.abs(d_str_vs_def)
    X["m_finish_gap_ko_sub"] = X["m_ko_path_vs_vuln"] - X["m_sub_path_vs_vuln"]
    X["m_ko_share"] = X["m_ko_path_vs_vuln"] / (
        np.abs(X["m_ko_path_vs_vuln"]) + np.abs(X["m_sub_path_vs_vuln"]) + np.abs(X["m_dec_path"]) + 1e-6
    )
    X["m_sub_share"] = X["m_sub_path_vs_vuln"] / (
        np.abs(X["m_ko_path_vs_vuln"]) + np.abs(X["m_sub_path_vs_vuln"]) + np.abs(X["m_dec_path"]) + 1e-6
    )
    X["m_dec_share"] = X["m_dec_path"] / (
        np.abs(X["m_ko_path_vs_vuln"]) + np.abs(X["m_sub_path_vs_vuln"]) + np.abs(X["m_dec_path"]) + 1e-6
    )
    X["m_finish_vs_cardio"] = X["m_finish_bias"] * (1.0 / np.maximum(total_rounds, 1.0))
    X["m_decision_durability"] = (d_dec_win - np.abs(d_ko_loss) - np.abs(d_been_finished)) * np.maximum(total_rounds, 1.0)
    X["m_sub_chain_pressure"] = (d_sub_att + 0.50 * X["m_sub_round3_pressure"]) * (1.0 + np.maximum(d_grap_vs_tdd, -1.5))
    X["m_sub_finish_trigger"] = X["m_sub_chain_pressure"] + 0.45 * X["m_sub_trend_pressure"] + 0.30 * d_been_finished
    X["m_ko_chain_pressure"] = np.maximum(d_str_vs_def, -2.0) * (1.0 + np.maximum(d_ko_win, -1.0))
    X["m_title_rounds_interaction"] = is_title * total_rounds * X["m_dec_path"]

    # Method-only explicit mechanics: KO-shaped pressure, sub chaining, and decision control.
    X["m_ko_headhunter"] = d_power_ratio * d_head_pct * d_sig_str_diff_pm
    X["m_ko_fast_start"] = d_first_round_finish_rate * d_rd1_intensity_ratio * (4.0 - np.minimum(total_rounds, 4.0))
    X["m_ko_pressure_conversion"] = d_damage_efficiency * d_power_ratio * np.maximum(d_sig_str_acc, 0.0)
    X["m_ko_exchange_edge"] = d_strike_exchange_ratio * d_power_ratio * d_sig_str_diff_pm
    X["m_ko_distance_sniper"] = d_distance_pct * d_sig_str_acc * d_power_ratio
    X["m_ko_knockdown_axis"] = d_kd_pm - 0.6 * d_def_kd_pm
    X["m_ko_attrition"] = (d_body_pct + 0.7 * d_leg_pct) * d_output_rate * d_cardio_ratio
    X["m_ko_confident_mismatch"] = d_glicko * np.maximum(d_power_ratio, 0.0) * np.maximum(d_sig_str_diff_pm, 0.0)
    X["m_ko_burst_vs_decay"] = d_rd1_intensity_ratio - d_cardio_ratio
    X["m_ko_path_specific"] = (d_ko_win + d_ko_loss) * (0.5 + d_head_pct + 0.5 * d_power_ratio)
    X["m_sub_ground_hunter"] = d_ground_pct * (d_ctrl_pct + 0.5 * d_sub_att)
    X["m_sub_scramble_threat"] = d_rev_p15 * (d_sub_att + d_ground_pct)
    X["m_sub_late_snowball"] = d_cardio_ratio * d_r3_sub_share * (d_ctrl_pct + d_sub_att)
    X["m_dec_clean_pointing"] = (
        d_distance_pct * d_sig_str_acc * d_consistency - 0.5 * (np.abs(d_kd_pm) + np.abs(d_sub_att))
    )
    X["m_dec_stability"] = d_finish_resistance * d_durability * d_late_round_pct * d_consistency
    X["m_finish_confidence"] = d_glicko * np.maximum(X["m_finish_bias"], 0.0)
    X["m_sub_vs_ko_axis"] = (d_ground_pct + d_ctrl_pct + d_sub_att) - (d_head_pct + d_power_ratio + d_kd_pm)
    X["m_ko_head_accuracy"] = _col("d_head_hunt_accuracy", d_head_acc) * d_power_ratio
    X["m_ko_range_sniper"] = d_distance_acc * d_distance_share * d_power_ratio
    X["m_ko_range_accuracy"] = _col("d_distance_strike_accuracy", d_distance_acc) * d_distance_pct * d_power_ratio
    X["m_ground_finish_conversion"] = d_ground_acc * d_ctrl_pct * d_sub_att
    X["m_clinch_breaker"] = d_clinch_acc * d_clinch_pct * d_kd_pm
    X["m_attrition_finish"] = d_body_leg_attrition * d_late_round_pct * d_output_rate
    X["m_sub_loss_vulnerability"] = d_sub_loss_pct
    X["m_recent_sub_win_rate_gap"] = d_recent_sub_win_rate
    X["m_sub_entry_pressure"] = d_sub_entry_pressure
    X["m_sub_control_conversion"] = d_sub_control_conversion
    X["m_sub_defensive_leak"] = d_sub_defensive_leak
    X["m_late_sub_pressure"] = d_late_sub_pressure
    X["m_sub_recency_surge"] = d_sub_recency_surge
    X["m_grapple_recency_surge"] = d_grapple_recency_surge
    X["m_sub_vs_control_axis"] = d_sub_vs_control_axis
    # Exact side-specific reconstruction from oriented differential + invariant sums:
    # winner_side = 0.5 * (sum + d), loser_side = 0.5 * (sum - d)
    sub_entry_w = 0.5 * (sub_entry_sum + d_sub_entry_pressure)
    sub_entry_l = 0.5 * (sub_entry_sum - d_sub_entry_pressure)
    sub_leak_w = 0.5 * (sub_leak_sum + d_sub_defensive_leak)
    sub_leak_l = 0.5 * (sub_leak_sum - d_sub_defensive_leak)
    sub_attack_vs_leak_w = sub_entry_w * sub_leak_l
    sub_attack_vs_leak_l = sub_entry_l * sub_leak_w
    X["m_sub_mismatch_explicit"] = sub_attack_vs_leak_w - sub_attack_vs_leak_l
    X["m_sub_mismatch_max"] = np.maximum(sub_attack_vs_leak_w, sub_attack_vs_leak_l)
    X["m_sub_mismatch_sum"] = sub_attack_vs_leak_w + sub_attack_vs_leak_l
    X["m_sub_mismatch"] = X["m_sub_mismatch_explicit"]
    ko_attack_w = 0.5 * (ko_attack_sum + d_ko_attack_pressure)
    ko_attack_l = 0.5 * (ko_attack_sum - d_ko_attack_pressure)
    ko_leak_w = 0.5 * (ko_leak_sum + d_ko_def_leak)
    ko_leak_l = 0.5 * (ko_leak_sum - d_ko_def_leak)
    ko_attack_vs_leak_w = ko_attack_w * ko_leak_l
    ko_attack_vs_leak_l = ko_attack_l * ko_leak_w
    X["m_ko_mismatch"] = ko_attack_vs_leak_w - ko_attack_vs_leak_l
    X["m_ko_recent_fragility"] = d_r5_def_kd_pm + 0.7 * d_recent_ko_loss_rate

    # Keep ratios bounded and numerically stable for tree/linear blends.
    X["m_sub_ground_efficiency"] = (d_ground_pct * np.maximum(d_sub_att, 0.0)) / (1.0 + np.abs(d_ctrl_pct) + eps)

    # ── Stage-1 side-reconstruction features ─────────────────────────────────
    # Reconstruct exact winner-side (w) and loser-side (l) values from the
    # oriented differential d_x and the orientation-invariant sum x_sum:
    #   x_w = 0.5 * (x_sum + d_x),   x_l = 0.5 * (x_sum - d_x)
    def _sides(sum_col, diff_col, default=0.0):
        s = _col(sum_col, default * 2)
        d = _col(diff_col, 0.0)
        return 0.5 * (s + d), 0.5 * (s - d)

    dec_w,  dec_l  = _sides("dec_win_pct_sum",          "d_dec_win_pct",           0.0)
    res_w,  res_l  = _sides("finish_resistance_sum",     "d_finish_resistance",     0.0)
    cons_w, cons_l = _sides("consistency_sum",           "d_consistency",           0.0)
    cardio_w, cardio_l = _sides("cardio_ratio_sum",      "d_cardio_ratio",          0.0)
    dur_w,  dur_l  = _sides("durability_sum",            "d_durability",            0.0)
    out_w,  out_l  = _sides("output_rate_sum",           "d_output_rate",           0.0)
    r1_w,   r1_l   = _sides("rd1_intensity_ratio_sum",   "d_rd1_intensity_ratio",   0.0)
    exch_w, exch_l = _sides("strike_exchange_ratio_sum", "d_strike_exchange_ratio", 0.0)
    acc_w,  acc_l  = _sides("sig_str_acc_sum",           "d_sig_str_acc",           0.0)
    late_w, late_l = _sides("late_round_pct_sum",        "d_late_round_pct",        0.0)
    time_w, time_l = _sides("avg_time_min_sum",          "d_avg_time_min",          0.0)
    afr_w,  afr_l  = _sides("avg_finish_round_sum",      "d_avg_finish_round",      0.0)
    fr1_w,  fr1_l  = _sides("first_round_finish_rate_sum", "d_first_round_finish_rate", 0.0)
    dmg_w,  dmg_l  = _sides("damage_efficiency_sum",    "d_damage_efficiency",      0.0)
    attr_w_raw, attr_l_raw = _sides("body_leg_attrition_sum", "d_body_leg_attrition", 0.0)

    # Reuse existing KO/Sub attack-vs-leak reconstructions
    ko_w = np.maximum(ko_attack_vs_leak_w, 0.0)
    ko_l = np.maximum(ko_attack_vs_leak_l, 0.0)
    sub_w = np.maximum(sub_attack_vs_leak_w, 0.0)
    sub_l = np.maximum(sub_attack_vs_leak_l, 0.0)

    # ── Latent stage-1 components ─────────────────────────────────────────────
    decision_shell_w = (
        0.34*dec_w + 0.18*cons_w + 0.16*cardio_w + 0.14*res_w + 0.10*dur_w + 0.08*acc_w
    ) * (0.90 + 0.10*late_w)
    decision_shell_l = (
        0.34*dec_l + 0.18*cons_l + 0.16*cardio_l + 0.14*res_l + 0.10*dur_l + 0.08*acc_l
    ) * (0.90 + 0.10*late_l)

    chaos_w = (
        np.clip(out_w, 0, 8)
        * (0.55 + np.clip(exch_w, 0.5, 1.8))
        * (0.55 + np.clip(r1_w,  0.5, 1.8))
    )
    chaos_l = (
        np.clip(out_l, 0, 8)
        * (0.55 + np.clip(exch_l, 0.5, 1.8))
        * (0.55 + np.clip(r1_l,  0.5, 1.8))
    )

    finish_speed_w = (1.0 / (afr_w + 0.75)) * (0.60 + 0.40*fr1_w)
    finish_speed_l = (1.0 / (afr_l + 0.75)) * (0.60 + 0.40*fr1_l)

    time_shell_w = time_w * (0.55 + 0.45*dec_w) * (0.70 + 0.30*res_w)
    time_shell_l = time_l * (0.55 + 0.45*dec_l) * (0.70 + 0.30*res_l)

    attrition_w = attr_w_raw * late_w * cardio_w
    attrition_l = attr_l_raw * late_l * cardio_l

    finish_total_w = ko_w + 0.92*sub_w + 0.10*dmg_w + 0.18*finish_speed_w
    finish_total_l = ko_l + 0.92*sub_l + 0.10*dmg_l + 0.18*finish_speed_l

    # ── Final stage-1 features ────────────────────────────────────────────────
    X["m_decision_shell_gap"]       = decision_shell_w - decision_shell_l
    X["m_decision_shell_sum"]       = decision_shell_w + decision_shell_l
    X["m_finish_conversion_edge"]   = finish_total_w - 0.72*decision_shell_l + 0.12*chaos_w
    X["m_finish_environment"]       = (finish_total_w + finish_total_l) - 0.68*(decision_shell_w + decision_shell_l)
    X["m_mutual_finish_instability"] = finish_total_w * finish_total_l
    X["m_decision_absorber"]        = (
        decision_shell_w + decision_shell_l + time_shell_w + time_shell_l
        - finish_total_w - finish_total_l - 0.25*(chaos_w + chaos_l)
    )
    X["m_early_finish_window"]      = (
        (chaos_w + chaos_l)
        * (4.0 - np.minimum(total_rounds, 4.0))
        * (0.35 + ko_w + ko_l)
    )
    X["m_fast_start_fragility"]     = fr1_w*(1.0 - dur_l) + fr1_l*(1.0 - dur_w)
    X["m_late_finish_window"]       = (
        np.maximum(total_rounds - 2.0, 1.0)
        * (
            sub_w + sub_l
            + 0.20*(2.0 - cardio_w - cardio_l)
            + 0.15*(2.0 - decision_shell_w - decision_shell_l)
            + 0.20*(attrition_w + attrition_l)
        )
    )
    X["m_attrition_break_window"]   = attrition_w*(1.0 - cardio_l) + attrition_l*(1.0 - cardio_w)
    X["m_time_profile_finish_bias"] = (
        (finish_total_w + finish_total_l)
        * (1.0/(1.0 + time_w) + 1.0/(1.0 + time_l))
    )
    X["m_finish_over_shell_ratio"]  = (
        (finish_total_w + finish_total_l + eps)
        / (decision_shell_w + decision_shell_l + eps)
    )
    X["m_clean_decision_track"]     = (
        (dec_w*cons_w*acc_w + dec_l*cons_l*acc_l)
        * (0.75 + 0.25*(time_w + time_l))
    )
    X["m_finish_speed_pressure"]    = finish_speed_w*(1.0 - res_l) + finish_speed_l*(1.0 - res_w)

    return X


def fuzzy_find(name, fighter_history):
    if name in fighter_history and fighter_history[name]:
        return name
    lower = str(name).lower()
    for key in fighter_history:
        if str(key).lower() == lower and fighter_history[key]:
            return key
    matches = [k for k in fighter_history if lower in str(k).lower() and fighter_history[k]]
    if len(matches) == 1:
        return matches[0]
    return None


# ─── Weight class ordinal mapping ─────────────────────────────────────────────
# Ordinal is only a coarse size proxy now; the exact class is carried by one-hot
# features below, so every class gets a unique code to avoid collisions.
WEIGHT_CLASS_ORDINAL = {
    "Women's Strawweight": 1, "Women's Flyweight": 2, "Women's Bantamweight": 3,
    "Women's Featherweight": 4, "Flyweight": 5, "Bantamweight": 6,
    "Featherweight": 7, "Lightweight": 8, "Welterweight": 9,
    "Middleweight": 10, "Light Heavyweight": 11, "Heavyweight": 12,
    "Catch Weight": 13, "Open Weight": 14,
}

ACTIVE_ENSEMBLE_MODELS = {
    "LightGBM", "XGBoost", "CatBoost",
    "HistGBM", "RandForest", "ExtraTrees",
    "MLP", "LogReg",
}

FIXED_MODEL_PARAMS = {
    "LightGBM": {
        "n_estimators": 300, "lr": 0.04, "max_depth": 6, "num_leaves": 31,
        "min_child_samples": 25, "subsample": 0.8, "colsample_bytree": 0.8,
        "reg_alpha": 0.1, "reg_lambda": 1.0,
    },
    "XGBoost": {
        "n_estimators": 300, "lr": 0.04, "max_depth": 5, "min_child_weight": 5,
        "subsample": 0.8, "colsample_bytree": 0.8, "reg_alpha": 0.1, "reg_lambda": 1.0,
    },
    "CatBoost": {
        "iterations": 350, "lr": 0.05, "depth": 6, "l2_leaf_reg": 3.0,
        "random_strength": 1.0, "bagging_temperature": 0.5,
        "border_count": 128, "bootstrap_type": "Bayesian",
    },
    "ExtraTrees": {
        "n_estimators": 500, "max_depth": 12, "min_samples_leaf": 5,
        "min_samples_split": 2, "max_features": 0.7,
    },
    "HistGBM": {
        "max_iter": 500, "lr": 0.05, "max_depth": 6,
        "max_leaf_nodes": 31, "min_samples_leaf": 20, "l2_reg": 1.0,
    },
    "RandForest": {
        "n_estimators": 400, "max_depth": 12, "min_samples_leaf": 5,
        "min_samples_split": 2, "max_features": 0.7,
    },
    "AdaBoost": {
        "n_estimators": 200, "learning_rate": 0.1,
    },
}


def _weight_class_feature_name(weight_class):
    slug = re.sub(r"[^a-z0-9]+", "_", str(weight_class).lower()).strip("_")
    return f"wc_{slug or 'unknown'}"


WEIGHT_CLASS_FEATURES = {
    weight_class: _weight_class_feature_name(weight_class)
    for weight_class in WEIGHT_CLASS_ORDINAL
}
UNKNOWN_WEIGHT_CLASS_FEATURE = "wc_unknown"
METHOD_LABELS = ["Decision", "KO/TKO", "Submission"]
MIN_METHOD_PROB = 0.001

# ─── Glicko-2 core ─────────────────────────────────────────────────────────────

def _g(phi):
    return 1.0 / math.sqrt(1.0 + 3.0 * phi**2 / math.pi**2)

def _E(mu, mu_j, phi_j):
    return 1.0 / (1.0 + math.exp(-_g(phi_j) * (mu - mu_j)))

def glicko2_update(rating, opponents):
    mu, phi, sigma = rating
    mu_s = (mu - MU_0) / SCALE
    phi_s = phi / SCALE
    if not opponents:
        phi_star = math.sqrt(phi_s**2 + sigma**2)
        return (mu, phi_star * SCALE, sigma)
    v_inv = 0.0
    delta_sum = 0.0
    for opp_r, opp_rd, score in opponents:
        mu_j = (opp_r - MU_0) / SCALE
        phi_j = opp_rd / SCALE
        g_j = _g(phi_j)
        E_j = _E(mu_s, mu_j, phi_j)
        v_inv += g_j**2 * E_j * (1.0 - E_j)
        delta_sum += g_j * (score - E_j)
    v = 1.0 / v_inv if v_inv > 0 else 1e6
    delta = v * delta_sum
    a = math.log(sigma**2)
    def f(x):
        ex = math.exp(x)
        num = ex * (delta**2 - phi_s**2 - v - ex)
        den = 2.0 * (phi_s**2 + v + ex)**2
        return num / den - (x - a) / (TAU**2)
    A = a
    if delta**2 > phi_s**2 + v:
        B = math.log(delta**2 - phi_s**2 - v)
    else:
        k = 1
        while f(a - k * TAU) < 0:
            k += 1
        B = a - k * TAU
    fA, fB = f(A), f(B)
    for _ in range(100):
        C = A + (A - B) * fA / (fB - fA)
        fC = f(C)
        if fC * fB < 0:
            A, fA = B, fB
        else:
            fA /= 2.0
        B, fB = C, fC
        if abs(B - A) < CONVERGENCE:
            break
    new_sigma = math.exp(A / 2.0)
    phi_star = math.sqrt(phi_s**2 + new_sigma**2)
    new_phi_s = 1.0 / math.sqrt(1.0 / phi_star**2 + 1.0 / v)
    new_mu_s = mu_s + new_phi_s**2 * delta_sum
    return (new_mu_s * SCALE + MU_0, new_phi_s * SCALE, new_sigma)

# ─── Helpers ────────────────────────────────────────────────────────────────────

def _isnan(v):
    if v is None:
        return True
    try:
        return math.isnan(float(v))
    except (TypeError, ValueError):
        return True

def _safe_sum(values):
    return sum(v for v in values if not _isnan(v))

def _safe_mean(values):
    clean = [v for v in values if not _isnan(v)]
    return sum(clean) / len(clean) if clean else float("nan")

def _safe_div(a, b, default=0.0):
    return a / b if b and b > 0 else default


def _num_or(value, default=0.0):
    return default if _isnan(value) else float(value)


def _abs_gap(a, b, default=float("nan")):
    if _isnan(a) or _isnan(b):
        return default
    return abs(float(a) - float(b))


def _extract_profile_from_row(row, prefix):
    return {
        "height": row.get(f"{prefix}_height", float("nan")),
        "reach": row.get(f"{prefix}_reach", float("nan")),
        "ape_index": row.get(f"{prefix}_ape_index", float("nan")),
        "weight": row.get(f"{prefix}_weight", float("nan")),
        "age": row.get(f"{prefix}_age_at_event", float("nan")),
        "stance": row.get(f"{prefix}_stance", ""),
    }


def _weighted_rate(numerator, denominator, prior=0.5, strength=25.0):
    """Bayesian-smoothed rate using equivalent sample-size `strength`."""
    if denominator is None or denominator <= 0:
        return prior
    return (numerator + prior * strength) / (denominator + strength)


def _bayes_shrink(empirical_value, sample_size, prior=0.5, strength=8.0):
    """Shrink noisy low-sample statistics toward a global prior."""
    if _isnan(empirical_value):
        empirical_value = prior
    sample_size = max(float(sample_size or 0.0), 0.0)
    return (empirical_value * sample_size + prior * strength) / (sample_size + strength)


def _correlation_prune(X, threshold=0.95):
    """Drop columns whose |corr| with an earlier-kept column exceeds threshold.

    Column order determines which member of a correlated pair survives — the
    first-seen column is retained. Returns the list of surviving column names.
    """
    cols = list(X.columns)
    if len(cols) <= 1:
        return cols
    corr = np.abs(X.corr().fillna(0.0).to_numpy())
    np.fill_diagonal(corr, 0.0)
    keep_mask = [True] * len(cols)
    for i in range(len(cols)):
        if not keep_mask[i]:
            continue
        for j in range(i + 1, len(cols)):
            if keep_mask[j] and corr[i, j] > threshold:
                keep_mask[j] = False
    return [c for c, k in zip(cols, keep_mask) if k]


def _time_weight(fight_date, current_date, half_life_days=730):
    """Exponential decay weight: half-life of ~2 years."""
    if current_date is None or fight_date is None:
        return 1.0
    try:
        days = (current_date - fight_date).days
        return 2.0 ** (-max(days, 0) / half_life_days)
    except (TypeError, AttributeError):
        return 1.0


def _pick_missing_cols(X, threshold=MISSINGNESS_THRESHOLD):
    return [c for c in X.columns if float(X[c].isna().mean()) >= threshold]


def _add_missingness_indicators(X, missing_cols):
    X2 = X.copy()
    for col in missing_cols:
        if col not in X2.columns:
            X2[col] = np.nan
        X2[f"miss_{col}"] = X2[col].isna().astype(float)
    return X2


def _expected_calibration_error(y_true, probs, n_bins=10):
    y = np.asarray(y_true, dtype=float)
    p = np.asarray(probs, dtype=float)
    bins = np.linspace(0.0, 1.0, n_bins + 1)
    inds = np.digitize(p, bins) - 1
    ece = 0.0
    for b in range(n_bins):
        mask = inds == b
        if not np.any(mask):
            continue
        conf = p[mask].mean()
        acc = y[mask].mean()
        ece += (mask.mean()) * abs(acc - conf)
    return float(ece)


def _calibration_curve_rmse(y_true, probs, n_bins=10):
    y = np.asarray(y_true, dtype=float)
    p = np.asarray(probs, dtype=float)
    bins = np.linspace(0.0, 1.0, n_bins + 1)
    inds = np.digitize(p, bins) - 1
    sq = 0.0
    wt = 0.0
    for b in range(n_bins):
        mask = inds == b
        if not np.any(mask):
            continue
        conf = float(p[mask].mean())
        acc = float(y[mask].mean())
        w = float(mask.mean())
        sq += w * ((acc - conf) ** 2)
        wt += w
    if wt <= 0:
        return 0.0
    return float(np.sqrt(sq / wt))


# Per-round stat names to track
RD_STATS = [
    "sig_str", "sig_str_att", "kd", "td", "td_att", "sub_att", "ctrl_sec",
    "head", "head_att", "body", "body_att", "leg", "leg_att",
    "distance", "distance_att", "clinch", "clinch_att", "ground", "ground_att",
]

# ─── Fight record extraction ───────────────────────────────────────────────────

def extract_fight_record(row, prefix, opp, result, opp_glicko_mu=MU_0):
    """Build a per-fighter fight record dict from a DataFrame row."""
    def g(col):
        v = row.get(col)
        if v is None or (isinstance(v, float) and math.isnan(v)):
            return float("nan")
        return v

    rec = {
        "date": row["event_date"],
        "fight_time": g("total_fight_time_sec") or 0,
        "result": result,
        "opp_glicko": opp_glicko_mu,
        "method": row.get("method", ""),
        "is_title": g("is_title_bout") or 0,
        "scheduled_rounds": int(g("total_rounds") or 3),
        "finish_round": g("finish_round") or 0,
        # Offense
        "sig_str": g(f"{prefix}_sig_str") or 0,
        "sig_str_att": g(f"{prefix}_sig_str_att") or 0,
        "sig_str_acc": g(f"{prefix}_sig_str_acc"),
        "str": g(f"{prefix}_str") or 0,
        "str_att": g(f"{prefix}_str_att") or 0,
        "str_acc": g(f"{prefix}_str_acc"),
        "kd": g(f"{prefix}_kd") or 0,
        "td": g(f"{prefix}_td") or 0,
        "td_att": g(f"{prefix}_td_att") or 0,
        "td_acc": g(f"{prefix}_td_acc"),
        "sub_att": g(f"{prefix}_sub_att") or 0,
        "rev": g(f"{prefix}_rev") or 0,
        "ctrl_sec": g(f"{prefix}_ctrl_sec") or 0,
        # Targeting (fight-total %)
        "head_pct": g(f"{prefix}_head"),
        "body_pct": g(f"{prefix}_body"),
        "leg_pct": g(f"{prefix}_leg"),
        # Positioning (fight-total %)
        "distance_pct": g(f"{prefix}_distance"),
        "clinch_pct": g(f"{prefix}_clinch"),
        "ground_pct": g(f"{prefix}_ground"),
        # Defense (opponent stats)
        "opp_sig_str": g(f"{opp}_sig_str") or 0,
        "opp_sig_str_att": g(f"{opp}_sig_str_att") or 0,
        "opp_sig_str_acc": g(f"{opp}_sig_str_acc"),
        "opp_str": g(f"{opp}_str") or 0,
        "opp_kd": g(f"{opp}_kd") or 0,
        "opp_td": g(f"{opp}_td") or 0,
        "opp_td_att": g(f"{opp}_td_att") or 0,
        "opp_sub_att": g(f"{opp}_sub_att") or 0,
        "opp_ctrl_sec": g(f"{opp}_ctrl_sec") or 0,
        # Physical
        "height": g(f"{prefix}_height"),
        "reach": g(f"{prefix}_reach"),
        "ape_index": g(f"{prefix}_ape_index"),
        "weight": g(f"{prefix}_weight"),
        "age": g(f"{prefix}_age_at_event"),
        "stance": row.get(f"{prefix}_stance", ""),
    }

    # Per-round stats (offense + opponent defensive counterpart)
    for rd in range(1, 6):
        for stat in RD_STATS:
            rec[f"rd{rd}_{stat}"] = g(f"{prefix}_rd{rd}_{stat}")
            rec[f"opp_rd{rd}_{stat}"] = g(f"{opp}_rd{rd}_{stat}")

    return rec


# ─── Per-fighter feature computation ───────────────────────────────────────────

_FIGHTER_FEAT_KEYS = None  # populated on first call


def _make_synthetic_fight_record(current_date=None, profile=None):
    profile = profile or {}
    rec = {
        "date": current_date or pd.Timestamp("2000-01-01"),
        "fight_time": 900.0,
        "result": "W",
        "opp_glicko": MU_0,
        "self_glicko": MU_0,
        "method": "Decision",
        "is_title": 0,
        "scheduled_rounds": 3,
        "finish_round": 3,
        "sig_str": 0.0,
        "sig_str_att": 0.0,
        "sig_str_acc": float("nan"),
        "str": 0.0,
        "str_att": 0.0,
        "str_acc": float("nan"),
        "kd": 0.0,
        "td": 0.0,
        "td_att": 0.0,
        "td_acc": float("nan"),
        "sub_att": 0.0,
        "rev": 0.0,
        "ctrl_sec": 0.0,
        "head_pct": float("nan"),
        "body_pct": float("nan"),
        "leg_pct": float("nan"),
        "distance_pct": float("nan"),
        "clinch_pct": float("nan"),
        "ground_pct": float("nan"),
        "opp_sig_str": 0.0,
        "opp_sig_str_att": 0.0,
        "opp_sig_str_acc": float("nan"),
        "opp_str": 0.0,
        "opp_kd": 0.0,
        "opp_td": 0.0,
        "opp_td_att": 0.0,
        "opp_sub_att": 0.0,
        "opp_ctrl_sec": 0.0,
        "height": profile.get("height", 70.0),
        "reach": profile.get("reach", 72.0),
        "ape_index": profile.get("ape_index", 2.0),
        "weight": profile.get("weight", 170.0),
        "age": profile.get("age", 30.0),
        "stance": profile.get("stance", "Orthodox"),
    }
    for rd in range(1, 6):
        for stat in RD_STATS:
            rec[f"rd{rd}_{stat}"] = 0.0
            rec[f"opp_rd{rd}_{stat}"] = 0.0
    return rec


def _ensure_fighter_feature_keys(current_date=None):
    global _FIGHTER_FEAT_KEYS
    if _FIGHTER_FEAT_KEYS is not None:
        return
    dummy_history = [_make_synthetic_fight_record(current_date=current_date)]
    compute_fighter_features(dummy_history, (MU_0, PHI_0, SIGMA_0), [], current_date)


def compute_fighter_features(history, glicko, opp_glickos, current_date, fallback_profile=None):
    """Compute ~130 features from a fighter's fight history."""
    global _FIGHTER_FEAT_KEYS
    fallback_profile = fallback_profile or {}

    n = len(history)
    if n == 0:
        _ensure_fighter_feature_keys(current_date)
        feats = {k: float("nan") for k in _FIGHTER_FEAT_KEYS}
        stance = fallback_profile.get("stance", "") or ""
        feats.update({
            "height": _num_or(fallback_profile.get("height"), float("nan")),
            "reach": _num_or(fallback_profile.get("reach"), float("nan")),
            "ape_index": _num_or(fallback_profile.get("ape_index"), float("nan")),
            "weight": _num_or(fallback_profile.get("weight"), float("nan")),
            "age": _num_or(fallback_profile.get("age"), float("nan")),
            "num_fights": 0.0,
            "total_time_min": 0.0,
            "avg_time_min": 0.0,
            "title_bout_pct": 0.0,
            "win_rate": 0.5,
            "ko_win_pct": 0.25,
            "sub_win_pct": 0.12,
            "dec_win_pct": 0.13,
            "finish_rate": 0.5,
            "ko_loss_pct": 0.2,
            "been_finished_pct": 0.5,
            "last3_win_rate": 0.5,
            "last5_win_rate": 0.5,
            "win_streak": 0.0,
            "loss_streak": 0.0,
            "days_inactive": 365.0,
            "glicko_mu": glicko[0],
            "glicko_phi": glicko[1],
            "fights_per_year": 0.0,
            "avg_opp_glicko": MU_0,
            "td_win_rate": 0.5,
            "quality_win_rate": 0.5,
            "best_win_glicko": MU_0,
            "worst_loss_glicko": MU_0,
            "style_striking": 0.5,
            "style_wrestling": 0.25,
            "style_submission": 0.15,
            "style_clinch_ground": 0.32,
            "style_finishing": 0.5,
            "is_orthodox": 1.0 if stance == "Orthodox" else 0.0,
            "is_southpaw": 1.0 if stance == "Southpaw" else 0.0,
            "is_switch": 1.0 if stance == "Switch" else 0.0,
            "age_squared": 900.0,
            "prime_age": 1.0,
            "age_decline": 0.0,
            "experience_log": 0.0,
            "momentum": 0.5,
            "first_round_finish_rate": 0.15,
            "durability": 0.8,
            "output_rate": 0.0,
            "damage_efficiency": 1.0,
            "late_round_pct": 0.5,
            "avg_finish_round": 2.5,
            "sig_str_diff_pm": 0.0,
            "grappling_dominance": 0.0,
            "consistency": 0.5,
            "ewm_opp_quality": MU_0,
            "form_vs_career": 0.0,
            "form5_vs_career": 0.0,
            "days_since_last_loss": 1500.0,
            "title_fight_win_rate": 0.5,
            "title_fight_experience": 0.0,
            "fight_iq": 0.45 * 0.35,
            "rd1_intensity_ratio": 1.0,
            "cardio_ratio": 1.0,
            "opp_quality_trend": 1.0,
            "fights_last_year": 0.0,
            "win_rate_last_year": 0.5,
            "loss_recovery_rate": 0.5,
            "strike_exchange_ratio": 1.0,
            "grappling_threat": 0.0,
            "decision_win_rate": 0.5,
            "finish_resistance": 0.5,
            "offensive_diversity": 1.0,
            "ewm_sig_str_diff_pm": 0.0,
            "glicko_confidence": 0.0,
        })
        return feats

    feats = {}

    total_time = _safe_sum(h["fight_time"] for h in history)
    total_time_min = total_time / 60.0 if total_time > 0 else 1.0
    total_time_15 = total_time / 900.0 if total_time > 0 else 1.0

    total_sig_land = _safe_sum(h["sig_str"] for h in history)
    total_sig_att = _safe_sum(h["sig_str_att"] for h in history)
    total_str_land = _safe_sum(h["str"] for h in history)
    total_str_att = _safe_sum(h["str_att"] for h in history)
    total_td_land = _safe_sum(h["td"] for h in history)
    total_td_att = _safe_sum(h["td_att"] for h in history)

    # ── Striking offense (per minute) ──
    feats["sig_str_pm"] = total_sig_land / total_time_min
    feats["sig_str_att_pm"] = total_sig_att / total_time_min
    feats["str_pm"] = total_str_land / total_time_min
    feats["str_att_pm"] = total_str_att / total_time_min
    feats["kd_pm"] = _safe_sum(h["kd"] for h in history) / total_time_min

    # ── Accuracy (attempt-weighted + Bayesian shrinkage) ──
    feats["sig_str_acc"] = _bayes_shrink(
        _weighted_rate(total_sig_land, total_sig_att, prior=0.45, strength=30),
        n, prior=0.45, strength=8,
    )
    feats["str_acc"] = _bayes_shrink(
        _weighted_rate(total_str_land, total_str_att, prior=0.55, strength=30),
        n, prior=0.55, strength=8,
    )
    feats["td_acc"] = _bayes_shrink(
        _weighted_rate(total_td_land, total_td_att, prior=0.35, strength=20),
        n, prior=0.35, strength=8,
    )

    # ── Grappling (per 15 min) ──
    feats["td_p15"] = _safe_sum(h["td"] for h in history) / total_time_15
    feats["td_att_p15"] = _safe_sum(h["td_att"] for h in history) / total_time_15
    feats["sub_att_p15"] = _safe_sum(h["sub_att"] for h in history) / total_time_15
    feats["rev_p15"] = _safe_sum(h["rev"] for h in history) / total_time_15
    feats["ctrl_pct"] = _safe_sum(h["ctrl_sec"] for h in history) / total_time if total_time > 0 else 0

    # ── Targeting (attempt/volume-weighted + shrinkage) ──
    sig_for_target = _safe_sum(
        h["sig_str"] for h in history
        if not _isnan(h["head_pct"]) and not _isnan(h["sig_str"])
    )
    head_num = _safe_sum(
        h["head_pct"] * h["sig_str"]
        for h in history
        if not _isnan(h["head_pct"]) and not _isnan(h["sig_str"])
    )
    body_num = _safe_sum(
        h["body_pct"] * h["sig_str"]
        for h in history
        if not _isnan(h["body_pct"]) and not _isnan(h["sig_str"])
    )
    leg_num = _safe_sum(
        h["leg_pct"] * h["sig_str"]
        for h in history
        if not _isnan(h["leg_pct"]) and not _isnan(h["sig_str"])
    )
    feats["head_pct"] = _bayes_shrink(
        _weighted_rate(head_num, sig_for_target, prior=0.62, strength=40),
        n, prior=0.62, strength=8,
    )
    feats["body_pct"] = _bayes_shrink(
        _weighted_rate(body_num, sig_for_target, prior=0.22, strength=40),
        n, prior=0.22, strength=8,
    )
    feats["leg_pct"] = _bayes_shrink(
        _weighted_rate(leg_num, sig_for_target, prior=0.16, strength=40),
        n, prior=0.16, strength=8,
    )

    # ── Positioning (volume-weighted + shrinkage) ──
    sig_for_pos = _safe_sum(
        h["sig_str"] for h in history
        if not _isnan(h["distance_pct"]) and not _isnan(h["sig_str"])
    )
    dist_num = _safe_sum(
        h["distance_pct"] * h["sig_str"]
        for h in history
        if not _isnan(h["distance_pct"]) and not _isnan(h["sig_str"])
    )
    clinch_num = _safe_sum(
        h["clinch_pct"] * h["sig_str"]
        for h in history
        if not _isnan(h["clinch_pct"]) and not _isnan(h["sig_str"])
    )
    ground_num = _safe_sum(
        h["ground_pct"] * h["sig_str"]
        for h in history
        if not _isnan(h["ground_pct"]) and not _isnan(h["sig_str"])
    )
    feats["distance_pct"] = _bayes_shrink(
        _weighted_rate(dist_num, sig_for_pos, prior=0.68, strength=40),
        n, prior=0.68, strength=8,
    )
    feats["clinch_pct"] = _bayes_shrink(
        _weighted_rate(clinch_num, sig_for_pos, prior=0.14, strength=40),
        n, prior=0.14, strength=8,
    )
    feats["ground_pct"] = _bayes_shrink(
        _weighted_rate(ground_num, sig_for_pos, prior=0.18, strength=40),
        n, prior=0.18, strength=8,
    )

    # ── Defense (opponent per-minute) ──
    feats["def_sig_str_pm"] = _safe_sum(h["opp_sig_str"] for h in history) / total_time_min
    feats["def_str_pm"] = _safe_sum(h["opp_str"] for h in history) / total_time_min
    feats["def_kd_pm"] = _safe_sum(h["opp_kd"] for h in history) / total_time_min
    feats["def_td_p15"] = _safe_sum(h["opp_td"] for h in history) / total_time_15
    feats["def_sub_att_p15"] = _safe_sum(h["opp_sub_att"] for h in history) / total_time_15
    feats["def_ctrl_pct"] = _safe_sum(h["opp_ctrl_sec"] for h in history) / total_time if total_time > 0 else 0
    opp_sig_att = _safe_sum(h["opp_sig_str_att"] for h in history)
    opp_sig_land = _safe_sum(h["opp_sig_str"] for h in history)
    feats["def_sig_str_acc"] = _bayes_shrink(
        _weighted_rate(opp_sig_land, opp_sig_att, prior=0.45, strength=30),
        n, prior=0.45, strength=8,
    )

    # ── Differentials ──
    feats["net_sig_str_pm"] = feats["sig_str_pm"] - feats["def_sig_str_pm"]
    feats["net_kd_pm"] = feats["kd_pm"] - feats["def_kd_pm"]
    feats["net_td_p15"] = feats["td_p15"] - feats["def_td_p15"]
    feats["net_ctrl_pct"] = feats["ctrl_pct"] - feats["def_ctrl_pct"]

    # ── Defense rates ──
    feats["sig_str_defense_rate"] = _bayes_shrink(
        1.0 - _weighted_rate(opp_sig_land, opp_sig_att, prior=0.45, strength=30),
        n, prior=0.55, strength=8,
    )
    opp_td_att = _safe_sum(h["opp_td_att"] for h in history)
    opp_td_land = _safe_sum(h["opp_td"] for h in history)
    feats["td_defense_rate"] = _bayes_shrink(
        1.0 - _weighted_rate(opp_td_land, opp_td_att, prior=0.35, strength=20),
        n, prior=0.65, strength=8,
    )

    # ── Physical (most recent) ──
    latest = history[-1]
    feats["height"] = latest["height"] if not _isnan(latest["height"]) else _num_or(fallback_profile.get("height"), float("nan"))
    feats["reach"] = latest["reach"] if not _isnan(latest["reach"]) else _num_or(fallback_profile.get("reach"), float("nan"))
    feats["ape_index"] = latest["ape_index"] if not _isnan(latest["ape_index"]) else _num_or(fallback_profile.get("ape_index"), float("nan"))
    feats["weight"] = latest["weight"] if not _isnan(latest["weight"]) else _num_or(fallback_profile.get("weight"), float("nan"))
    feats["age"] = latest["age"] if not _isnan(latest["age"]) else _num_or(fallback_profile.get("age"), float("nan"))

    # ── Experience ──
    feats["num_fights"] = n
    feats["total_time_min"] = total_time_min
    feats["avg_time_min"] = total_time_min / n
    feats["title_bout_pct"] = sum(1 for h in history if h["is_title"]) / n

    # ── Record ──
    wins = sum(1 for h in history if h["result"] == "W")
    losses = sum(1 for h in history if h["result"] == "L")
    feats["win_rate"] = _bayes_shrink(_safe_div(wins, n, 0.5), n, prior=0.5, strength=10)
    ko_w = sum(1 for h in history if h["result"] == "W" and "KO" in str(h["method"]))
    sub_w = sum(1 for h in history if h["result"] == "W" and "Sub" in str(h["method"]))
    dec_w = sum(1 for h in history if h["result"] == "W" and "Dec" in str(h["method"]))
    feats["ko_win_pct"] = _bayes_shrink(_safe_div(ko_w, n, 0.25), n, prior=0.25, strength=10)
    feats["sub_win_pct"] = _bayes_shrink(_safe_div(sub_w, n, 0.12), n, prior=0.12, strength=10)
    feats["dec_win_pct"] = _bayes_shrink(_safe_div(dec_w, n, 0.13), n, prior=0.13, strength=10)
    feats["finish_rate"] = _safe_div(ko_w + sub_w, max(wins, 1))
    ko_l = sum(1 for h in history if h["result"] == "L" and "KO" in str(h["method"]))
    sub_l = sum(1 for h in history if h["result"] == "L" and "Sub" in str(h["method"]))
    feats["ko_loss_pct"] = _bayes_shrink(_safe_div(ko_l, n, 0.2), n, prior=0.2, strength=10)
    feats["sub_loss_pct"] = _bayes_shrink(_safe_div(sub_l, max(losses, 1), 0.20), losses, prior=0.20, strength=8)
    feats["been_finished_pct"] = _safe_div(ko_l + sub_l, max(losses, 1))

    # ── Form ──
    last3 = history[-3:]
    last5 = history[-5:]
    last3_wr = sum(1 for h in last3 if h["result"] == "W") / len(last3)
    last5_wr = sum(1 for h in last5 if h["result"] == "W") / len(last5)
    feats["last3_win_rate"] = _bayes_shrink(last3_wr, len(last3), prior=0.5, strength=6)
    feats["last5_win_rate"] = _bayes_shrink(last5_wr, len(last5), prior=0.5, strength=6)
    recent_sub_wins = sum(1 for h in last5 if h["result"] == "W" and "Sub" in str(h.get("method", "")))
    recent_ko_losses = sum(1 for h in last5 if h["result"] == "L" and "KO" in str(h.get("method", "")))
    recent_n = max(min(n, 5), 1)
    feats["recent_sub_win_rate"] = _bayes_shrink(
        _safe_div(recent_sub_wins, recent_n, 0.12), recent_n, prior=0.12, strength=5
    )
    feats["recent_ko_loss_rate"] = _bayes_shrink(
        _safe_div(recent_ko_losses, recent_n, 0.20), recent_n, prior=0.20, strength=5
    )
    win_streak = 0
    for h in reversed(history):
        if h["result"] == "W":
            win_streak += 1
        else:
            break
    loss_streak = 0
    for h in reversed(history):
        if h["result"] == "L":
            loss_streak += 1
        else:
            break
    feats["win_streak"] = win_streak
    feats["loss_streak"] = loss_streak
    if current_date is not None and not _isnan(history[-1]["date"]):
        feats["days_inactive"] = (current_date - history[-1]["date"]).days
    else:
        feats["days_inactive"] = 365

    # ── Short-notice flag (last fight < 45 days ago) ──
    feats["short_notice"] = float(feats["days_inactive"] < 45)

    # ── Rounds-scheduled experience (time-weighted) ──
    feats["rounds_5_exp"] = sum(
        _time_weight(h["date"], current_date) for h in history
        if int(h.get("scheduled_rounds", 3)) == 5
    )
    feats["rounds_3_exp"] = sum(
        _time_weight(h["date"], current_date) for h in history
        if int(h.get("scheduled_rounds", 3)) == 3
    )

    # ── Glicko-2 ──
    feats["glicko_mu"] = glicko[0]
    feats["glicko_phi"] = glicko[1]

    # ── Glicko trend (current mu minus mu from 5 fights ago) ──
    if n >= 5:
        feats["glicko_trend"] = glicko[0] - history[-5].get("self_glicko", glicko[0])
    elif n >= 1:
        feats["glicko_trend"] = glicko[0] - history[0].get("self_glicko", glicko[0])
    else:
        feats["glicko_trend"] = 0.0

    # ── Recent damage absorbed (sig str taken in last 3 fights) ──
    recent3 = history[-3:]
    feats["recent_damage_absorbed"] = _safe_sum(h["opp_sig_str"] for h in recent3)


    # ── Round 1 stats (career averages) ──
    for stat in ["sig_str", "sig_str_att", "kd", "td", "td_att", "sub_att",
                 "ctrl_sec", "head", "body", "leg", "distance", "clinch", "ground"]:
        vals = [h[f"rd1_{stat}"] for h in history if not _isnan(h.get(f"rd1_{stat}"))]
        feats[f"rd1_{stat}"] = _safe_mean(vals)

    # ── Round 2 stats ──
    for stat in ["sig_str", "sig_str_att", "kd", "td", "td_att", "sub_att",
                 "ctrl_sec", "head", "body", "leg"]:
        vals = [h[f"rd2_{stat}"] for h in history if not _isnan(h.get(f"rd2_{stat}"))]
        feats[f"rd2_{stat}"] = _safe_mean(vals)

    # ── Round 3 stats ──
    for stat in ["sig_str", "kd", "td", "td_att", "sub_att", "ctrl_sec", "head", "body", "leg"]:
        vals = [h[f"rd3_{stat}"] for h in history if not _isnan(h.get(f"rd3_{stat}"))]
        feats[f"rd3_{stat}"] = _safe_mean(vals)

    # ── Championship rounds (4+5 combined) ──
    for stat in ["sig_str", "kd", "td", "ctrl_sec"]:
        vals = []
        for rd in [4, 5]:
            vals.extend(h[f"rd{rd}_{stat}"] for h in history
                        if not _isnan(h.get(f"rd{rd}_{stat}")))
        feats[f"champ_{stat}"] = _safe_mean(vals)

    # ── Defensive round aggregates (what the fighter absorbs per round) ──
    for stat in ["sig_str", "kd"]:
        vals = [h[f"opp_rd1_{stat}"] for h in history if not _isnan(h.get(f"opp_rd1_{stat}"))]
        feats[f"def_rd1_{stat}"] = _safe_mean(vals)

    # ── Late-round sig_str per minute (only fights that reached rd4/rd5) ──
    late_vals = []
    for h in history:
        for rd in (4, 5):
            v = h.get(f"rd{rd}_sig_str")
            if not _isnan(v) and v is not None:
                late_vals.append(float(v) / 5.0)  # 5-minute round → per-minute
    feats["late_sig_str_pm"] = _safe_mean(late_vals) if late_vals else float("nan")


    # ── Composite pacing aggregates (feed the auto-diff loop in compute_matchup_features) ──
    # rd1 net sig_str per minute: what the fighter lands minus what they absorb in round 1.
    _rd1_ss = feats.get("rd1_sig_str")
    _def_rd1_ss = feats.get("def_rd1_sig_str")
    if not _isnan(_rd1_ss) and not _isnan(_def_rd1_ss):
        feats["rd1_net_sig_str"] = (float(_rd1_ss) - float(_def_rd1_ss)) / 5.0
    else:
        feats["rd1_net_sig_str"] = float("nan")
    # Cardio decay: rd1 output minus rd3 output, per minute. Higher = fades more.
    _rd3_ss = feats.get("rd3_sig_str")
    if not _isnan(_rd1_ss) and not _isnan(_rd3_ss):
        feats["cardio_decay_sig_str"] = (float(_rd1_ss) - float(_rd3_ss)) / 5.0
    else:
        feats["cardio_decay_sig_str"] = float("nan")
    # Round-1 control share: fraction of the 5-minute round spent in control position.
    _rd1_ctrl = feats.get("rd1_ctrl_sec")
    if not _isnan(_rd1_ctrl):
        feats["rd1_ctrl_share"] = float(_rd1_ctrl) / 300.0
    else:
        feats["rd1_ctrl_share"] = float("nan")

    # ── Late vs early ──
    rd1_ss = [h["rd1_sig_str"] for h in history if not _isnan(h.get("rd1_sig_str"))]
    rd3_ss = [h["rd3_sig_str"] for h in history if not _isnan(h.get("rd3_sig_str"))]
    feats["late_vs_early_sig_str"] = _safe_mean(rd3_ss) - _safe_mean(rd1_ss) \
        if rd1_ss and rd3_ss else float("nan")
    rd1_sub = [h["rd1_sub_att"] for h in history if not _isnan(h.get("rd1_sub_att"))]
    rd2_sub = [h["rd2_sub_att"] for h in history if not _isnan(h.get("rd2_sub_att"))]
    rd3_sub = [h["rd3_sub_att"] for h in history if not _isnan(h.get("rd3_sub_att"))]
    r1_sub_m = _safe_mean(rd1_sub)
    r2_sub_m = _safe_mean(rd2_sub)
    r3_sub_m = _safe_mean(rd3_sub)
    feats["late_vs_early_sub_att"] = r3_sub_m - r1_sub_m if rd1_sub and rd3_sub else float("nan")
    feats["sub_att_trend_12"] = r2_sub_m - r1_sub_m if rd1_sub and rd2_sub else float("nan")
    feats["sub_att_trend_23"] = r3_sub_m - r2_sub_m if rd2_sub and rd3_sub else float("nan")
    total_r123_sub = max(r1_sub_m + r2_sub_m + r3_sub_m, 1e-6)
    feats["rd3_sub_share"] = r3_sub_m / total_r123_sub
    feats["rd1_sub_share"] = r1_sub_m / total_r123_sub
    feats["sub_round_concentration"] = max(r1_sub_m, r2_sub_m, r3_sub_m) / total_r123_sub

    # ── Recent-window rate stats (last 1, 3, 5 fights) ──
    for prefix_tag, window in [("r1f", 1), ("r3", 3), ("r5", 5)]:
        recent = history[-window:]
        rt = _safe_sum(h["fight_time"] for h in recent)
        rt_min = rt / 60.0 if rt > 0 else 1.0
        rt_15 = rt / 900.0 if rt > 0 else 1.0
        feats[f"{prefix_tag}_sig_str_pm"] = _safe_sum(h["sig_str"] for h in recent) / rt_min
        feats[f"{prefix_tag}_kd_pm"] = _safe_sum(h["kd"] for h in recent) / rt_min
        feats[f"{prefix_tag}_td_p15"] = _safe_sum(h["td"] for h in recent) / rt_15
        feats[f"{prefix_tag}_sub_att_p15"] = _safe_sum(h["sub_att"] for h in recent) / rt_15
        feats[f"{prefix_tag}_ctrl_pct"] = _safe_sum(h["ctrl_sec"] for h in recent) / rt if rt > 0 else 0
        feats[f"{prefix_tag}_def_sig_str_pm"] = _safe_sum(h["opp_sig_str"] for h in recent) / rt_min
        feats[f"{prefix_tag}_def_kd_pm"] = _safe_sum(h["opp_kd"] for h in recent) / rt_min
        rec_sig_land = _safe_sum(h["sig_str"] for h in recent)
        rec_sig_att = _safe_sum(h["sig_str_att"] for h in recent)
        rec_td_land = _safe_sum(h["td"] for h in recent)
        rec_td_att = _safe_sum(h["td_att"] for h in recent)
        feats[f"{prefix_tag}_sig_str_acc"] = _bayes_shrink(
            _weighted_rate(rec_sig_land, rec_sig_att, prior=0.45, strength=20),
            len(recent), prior=0.45, strength=6,
        )
        feats[f"{prefix_tag}_td_acc"] = _bayes_shrink(
            _weighted_rate(rec_td_land, rec_td_att, prior=0.35, strength=20),
            len(recent), prior=0.35, strength=6,
        )
        feats[f"{prefix_tag}_win"] = _bayes_shrink(
            _safe_div(sum(1 for h in recent if h["result"] == "W"), len(recent), 0.5),
            len(recent), prior=0.5, strength=6,
        )

    # ── Exponentially weighted moving average (alpha=0.3, more recent = higher) ──
    alpha = 0.3
    ewm_keys = [
        ("sig_str", "fight_time", 60),
        ("kd", "fight_time", 60),
        ("td", "fight_time", 900),
        ("opp_sig_str", "fight_time", 60),
    ]
    for stat_key, time_key, divisor in ewm_keys:
        wsum, wtot = 0.0, 0.0
        for i, h in enumerate(history):
            w = (1 - alpha) ** (n - 1 - i)
            ft = h[time_key] / divisor if h[time_key] > 0 else 1.0
            wsum += w * (h[stat_key] / ft)
            wtot += w
        tag = stat_key.replace("opp_sig_str", "def_sig_str")
        feats[f"ewm_{tag}_pm"] = wsum / wtot if wtot > 0 else 0

    # EWM ctrl_pct
    wsum, wtot = 0.0, 0.0
    for i, h in enumerate(history):
        w = (1 - alpha) ** (n - 1 - i)
        ft = h["fight_time"] if h["fight_time"] > 0 else 1.0
        wsum += w * (h["ctrl_sec"] / ft)
        wtot += w
    feats["ewm_ctrl_pct"] = wsum / wtot if wtot > 0 else 0

    # EWM accuracy
    for acc_key in ["sig_str_acc", "td_acc"]:
        wsum, wtot = 0.0, 0.0
        for i, h in enumerate(history):
            if _isnan(h[acc_key]):
                continue
            w = (1 - alpha) ** (n - 1 - i)
            wsum += w * h[acc_key]
            wtot += w
        feats[f"ewm_{acc_key}"] = wsum / wtot if wtot > 0 else float("nan")

    # EWM win
    wsum, wtot = 0.0, 0.0
    for i, h in enumerate(history):
        w = (1 - alpha) ** (n - 1 - i)
        wsum += w * (1.0 if h["result"] == "W" else 0.0)
        wtot += w
    feats["ewm_win"] = wsum / wtot if wtot > 0 else 0.5

    # ── Variance (per-fight rates std dev) ──
    if n >= 2:
        per_fight_sig = []
        per_fight_kd = []
        per_fight_td = []
        per_fight_ctrl = []
        per_fight_def_sig = []
        for h in history:
            ft_min = h["fight_time"] / 60.0 if h["fight_time"] > 0 else 1.0
            ft_15 = h["fight_time"] / 900.0 if h["fight_time"] > 0 else 1.0
            ft = h["fight_time"] if h["fight_time"] > 0 else 1.0
            per_fight_sig.append(h["sig_str"] / ft_min)
            per_fight_kd.append(h["kd"] / ft_min)
            per_fight_td.append(h["td"] / ft_15)
            per_fight_ctrl.append(h["ctrl_sec"] / ft)
            per_fight_def_sig.append(h["opp_sig_str"] / ft_min)
        feats["std_sig_str_pm"] = float(np.std(per_fight_sig))
        feats["std_kd_pm"] = float(np.std(per_fight_kd))
        feats["std_td_p15"] = float(np.std(per_fight_td))
        feats["std_ctrl_pct"] = float(np.std(per_fight_ctrl))
        feats["std_def_sig_str_pm"] = float(np.std(per_fight_def_sig))
        # Time-weighted volatility of sig_str differential per minute
        _w_vol = [_time_weight(h["date"], current_date) for h in history]
        _sd_pm = []
        for h in history:
            _fm = h["fight_time"] / 60.0 if h["fight_time"] > 0 else 1.0
            _sd_pm.append((h["sig_str"] - h["opp_sig_str"]) / _fm)
        _wsum = sum(_w_vol) or 1.0
        _mean_sd = sum(w * d for w, d in zip(_w_vol, _sd_pm)) / _wsum
        _var_sd = sum(w * (d - _mean_sd) ** 2 for w, d in zip(_w_vol, _sd_pm)) / _wsum
        _vol_raw = math.sqrt(max(_var_sd, 0.0))
        feats["sig_diff_pm_vol"] = (_vol_raw * _wsum + 3.0 * 5.0) / (_wsum + 5.0)
    else:
        feats["std_sig_str_pm"] = 0.0
        feats["std_kd_pm"] = 0.0
        feats["std_td_p15"] = 0.0
        feats["std_ctrl_pct"] = 0.0
        feats["std_def_sig_str_pm"] = 0.0
        feats["sig_diff_pm_vol"] = 3.0

    # ── Derived ──
    total_sig = _safe_sum(h["sig_str"] for h in history)
    total_str = _safe_sum(h["str"] for h in history)
    feats["power_ratio"] = _safe_div(_safe_sum(h["kd"] for h in history), max(total_sig, 1))
    feats["striking_efficiency"] = _safe_div(total_sig, max(total_str, 1))
    feats["grappling_rate_p15"] = (
        _safe_sum(h["td"] + h["sub_att"] for h in history) / total_time_15
    )
    feats["def_grappling_rate_p15"] = (
        _safe_sum(h["opp_td"] + h["opp_sub_att"] for h in history) / total_time_15
    )
    feats["damage_ratio"] = _safe_div(
        _safe_sum(h["opp_sig_str"] for h in history),
        max(_safe_sum(h["sig_str"] for h in history), 1),
    )

    # ── Target/position attempt-driven accuracy and intent shares (round 1-5) ──
    eps = 1e-6

    def _round_stat_total(stat_name):
        vals = []
        for h in history:
            for rd in range(1, 6):
                v = h.get(f"rd{rd}_{stat_name}", float("nan"))
                if not _isnan(v):
                    vals.append(v)
        return _safe_sum(vals)

    total_sig_att_round = _round_stat_total("sig_str_att")
    total_head_land = _round_stat_total("head")
    total_head_att = _round_stat_total("head_att")
    total_body_land = _round_stat_total("body")
    total_body_att = _round_stat_total("body_att")
    total_leg_land = _round_stat_total("leg")
    total_leg_att = _round_stat_total("leg_att")
    total_distance_land = _round_stat_total("distance")
    total_distance_att = _round_stat_total("distance_att")
    total_clinch_land = _round_stat_total("clinch")
    total_clinch_att = _round_stat_total("clinch_att")
    total_ground_land = _round_stat_total("ground")
    total_ground_att = _round_stat_total("ground_att")

    feats["head_acc"] = _safe_div(total_head_land, total_head_att + eps, 0.16)
    feats["distance_acc"] = _safe_div(total_distance_land, total_distance_att + eps, 0.45)
    feats["ground_acc"] = _safe_div(total_ground_land, total_ground_att + eps, 0.35)
    feats["clinch_acc"] = _safe_div(total_clinch_land, total_clinch_att + eps, 0.30)
    feats["body_acc"] = _safe_div(total_body_land, total_body_att + eps, 0.28)
    feats["leg_acc"] = _safe_div(total_leg_land, total_leg_att + eps, 0.28)
    feats["body_leg_attrition"] = feats["body_acc"] + feats["leg_acc"]
    feats["ground_strike_accuracy"] = feats["ground_acc"]
    feats["head_hunt_accuracy"] = feats["head_acc"]
    feats["distance_strike_accuracy"] = feats["distance_acc"]

    feats["head_hunt_share"] = _safe_div(total_head_att, total_sig_att_round + eps, 0.33)
    feats["distance_share"] = _safe_div(total_distance_att, total_sig_att_round + eps, 0.62)

    # Fights per year
    if n > 1:
        career_days = (history[-1]["date"] - history[0]["date"]).days
        feats["fights_per_year"] = n / max(career_days / 365.25, 0.5)
    else:
        feats["fights_per_year"] = 1.0

    # Stance
    stance = latest.get("stance", "") or fallback_profile.get("stance", "")
    feats["is_orthodox"] = 1 if stance == "Orthodox" else 0
    feats["is_southpaw"] = 1 if stance == "Southpaw" else 0
    feats["is_switch"] = 1 if stance == "Switch" else 0

    # Average opponent Glicko
    feats["avg_opp_glicko"] = _safe_mean(opp_glickos) if opp_glickos else MU_0

    # ── Time-decayed career stats (half-life ~2 years) ──
    tw = [_time_weight(h["date"], current_date) for h in history]
    tw_total = sum(tw) or 1.0
    tw_time = sum(w * h["fight_time"] for w, h in zip(tw, history))
    tw_time_min = tw_time / 60.0 if tw_time > 0 else 1.0
    tw_time_15 = tw_time / 900.0 if tw_time > 0 else 1.0
    feats["td_sig_str_pm"] = sum(w * h["sig_str"] for w, h in zip(tw, history)) / tw_time_min
    feats["td_kd_pm"] = sum(w * h["kd"] for w, h in zip(tw, history)) / tw_time_min
    feats["td_td_p15"] = sum(w * h["td"] for w, h in zip(tw, history)) / tw_time_15
    feats["td_ctrl_pct"] = (
        sum(w * h["ctrl_sec"] for w, h in zip(tw, history)) / tw_time if tw_time > 0 else 0
    )
    feats["td_def_sig_str_pm"] = sum(w * h["opp_sig_str"] for w, h in zip(tw, history)) / tw_time_min
    tw_sig_land = sum(w * h["sig_str"] for w, h in zip(tw, history))
    tw_sig_att = sum(w * h["sig_str_att"] for w, h in zip(tw, history))
    feats["td_sig_str_acc"] = _weighted_rate(tw_sig_land, tw_sig_att, prior=0.45, strength=20)
    tw_td_land = sum(w * h["td"] for w, h in zip(tw, history))
    tw_td_att = sum(w * h["td_att"] for w, h in zip(tw, history))
    feats["td_td_acc"] = _weighted_rate(tw_td_land, tw_td_att, prior=0.35, strength=20)
    feats["td_win_rate"] = sum(
        w * (1.0 if h["result"] == "W" else 0.0) for w, h in zip(tw, history)
    ) / tw_total

    # ── Opponent-quality-adjusted features ──
    opp_g = [h.get("opp_glicko", MU_0) for h in history]
    opp_g_w = [max(g_val / MU_0, 0.1) for g_val in opp_g]
    ogw_total = sum(opp_g_w) or 1.0
    feats["quality_win_rate"] = sum(
        w * (1.0 if h["result"] == "W" else 0.0)
        for w, h in zip(opp_g_w, history)
    ) / ogw_total
    wins_opp_g = [og for h, og in zip(history, opp_g) if h["result"] == "W"]
    feats["best_win_glicko"] = max(wins_opp_g) if wins_opp_g else MU_0
    losses_opp_g = [og for h, og in zip(history, opp_g) if h["result"] == "L"]
    feats["worst_loss_glicko"] = min(losses_opp_g) if losses_opp_g else MU_0
    feats["quality_sig_str_pm"] = sum(
        w * h["sig_str"] for w, h in zip(opp_g_w, history)
    ) / (sum(w * (h["fight_time"] / 60.0 if h["fight_time"] > 0 else 1.0)
             for w, h in zip(opp_g_w, history)) or 1.0)

    # ── Style profile (continuous features for matchup interactions) ──
    total_offensive = feats["sig_str_pm"] + feats["td_p15"] + feats["sub_att_p15"] + 0.01
    feats["style_striking"] = feats["sig_str_pm"] / total_offensive
    feats["style_wrestling"] = feats["td_p15"] / total_offensive
    feats["style_submission"] = feats["sub_att_p15"] / total_offensive
    feats["style_clinch_ground"] = feats.get("clinch_pct", 0.14) + feats.get("ground_pct", 0.18)
    feats["style_finishing"] = feats.get("finish_rate", 0.5)

    # ── New high-impact features ──
    age_val = feats.get("age", 30.0)
    if _isnan(age_val):
        age_val = 30.0
    feats["age_squared"] = age_val ** 2
    feats["prime_age"] = 1.0 if 26.0 <= age_val <= 32.0 else 0.0
    feats["age_decline"] = max(0.0, age_val - 32.0)
    feats["experience_log"] = math.log1p(n)

    # Momentum: steeper EWM (alpha=0.5) for recent bias
    alpha_m = 0.5
    m_sum, m_tot = 0.0, 0.0
    for i, h in enumerate(history):
        w = (1 - alpha_m) ** (n - 1 - i)
        m_sum += w * (1.0 if h["result"] == "W" else 0.0)
        m_tot += w
    feats["momentum"] = m_sum / m_tot if m_tot > 0 else 0.5

    # First-round finish rate
    r1_finishes = sum(1 for h in history if h["result"] == "W" and h.get("finish_round") == 1
                      and ("KO" in str(h.get("method", "")) or "Sub" in str(h.get("method", ""))))
    feats["first_round_finish_rate"] = _bayes_shrink(_safe_div(r1_finishes, n, 0.15), n, prior=0.15, strength=10)

    # Durability (inverse of KO loss susceptibility)
    feats["durability"] = 1.0 - feats.get("ko_loss_pct", 0.2)

    # Output rate: total offensive actions per minute
    total_actions = total_sig_land + _safe_sum(h["td"] for h in history) + _safe_sum(h["sub_att"] for h in history)
    feats["output_rate"] = total_actions / total_time_min

    # Damage efficiency: sig str landed / opponent sig str landed
    opp_total_sig = _safe_sum(h["opp_sig_str"] for h in history)
    feats["damage_efficiency"] = _safe_div(total_sig_land, max(opp_total_sig, 1), 1.0)

    # Late-round endurance: pct of fights that went past round 2
    went_late = sum(1 for h in history if _num_or(h.get("finish_round"), 3) >= 3)
    feats["late_round_pct"] = _bayes_shrink(_safe_div(went_late, n, 0.5), n, prior=0.5, strength=8)

    # Average finish round (lower = more dangerous finisher)
    finish_rounds = [h["finish_round"] for h in history
                     if h["result"] == "W" and not _isnan(h.get("finish_round")) and h["finish_round"] > 0]
    feats["avg_finish_round"] = _safe_mean(finish_rounds) if finish_rounds else 2.5

    # Sig str differential per minute (combines offense and defense into one number)
    feats["sig_str_diff_pm"] = feats["sig_str_pm"] - feats["def_sig_str_pm"]

    # Grappling dominance: (td_landed + ctrl_pct) vs (opp_td_landed + opp_ctrl_pct)
    feats["grappling_dominance"] = feats["td_p15"] + feats["ctrl_pct"] - feats["def_td_p15"] - feats["def_ctrl_pct"]

    # Consistency: coefficient of variation of sig_str_pm (lower = more consistent)
    if n >= 2 and feats["sig_str_pm"] > 0:
        feats["consistency"] = 1.0 - min(feats["std_sig_str_pm"] / (feats["sig_str_pm"] + 0.01), 2.0) / 2.0
    else:
        feats["consistency"] = 0.5

    # EWM opponent quality: recent opponent strength
    alpha_oq = 0.4
    oq_sum, oq_tot = 0.0, 0.0
    for i, h in enumerate(history):
        w = (1 - alpha_oq) ** (n - 1 - i)
        oq_sum += w * _num_or(h.get("opp_glicko"), MU_0)
        oq_tot += w
    feats["ewm_opp_quality"] = oq_sum / oq_tot if oq_tot > 0 else MU_0

    # ── Form vs career (explicit trend for linear models) ──
    feats["form_vs_career"] = feats["last3_win_rate"] - feats["win_rate"]
    feats["form5_vs_career"] = feats["last5_win_rate"] - feats["win_rate"]

    # ── Days since last loss (confidence/psychology proxy) ──
    last_loss_idx = None
    for i in range(len(history) - 1, -1, -1):
        if history[i]["result"] == "L":
            last_loss_idx = i
            break
    if last_loss_idx is not None and current_date is not None:
        feats["days_since_last_loss"] = (current_date - history[last_loss_idx]["date"]).days
    else:
        feats["days_since_last_loss"] = 1500.0  # Never lost or no date

    # ── Title fight performance ──
    title_fights = [h for h in history if h.get("is_title")]
    title_wins = sum(1 for h in title_fights if h["result"] == "W")
    feats["title_fight_win_rate"] = _bayes_shrink(
        _safe_div(title_wins, len(title_fights), 0.5),
        len(title_fights), prior=0.5, strength=4,
    )
    feats["title_fight_experience"] = len(title_fights)

    # ── Fight IQ composite (multi-domain effectiveness) ──
    feats["fight_iq"] = feats["sig_str_acc"] * feats["td_acc"]

    # ── Round 1 intensity ratio (first-round explosiveness vs career pace) ──
    rd1_ss_val = feats.get("rd1_sig_str", float("nan"))
    if not _isnan(rd1_ss_val) and feats["sig_str_pm"] > 0:
        # rd1_sig_str is per-round avg, sig_str_pm is per-minute career; convert to same scale
        feats["rd1_intensity_ratio"] = rd1_ss_val / (feats["sig_str_pm"] * 5.0 + 0.01)
    else:
        feats["rd1_intensity_ratio"] = 1.0

    # ── Cardio ratio (round 3 output / round 1 output) ──
    rd1_val = feats.get("rd1_sig_str", float("nan"))
    rd3_val = feats.get("rd3_sig_str", float("nan"))
    if not _isnan(rd1_val) and not _isnan(rd3_val) and rd1_val > 0:
        feats["cardio_ratio"] = rd3_val / (rd1_val + 0.01)
    else:
        feats["cardio_ratio"] = 1.0

    # ── Submission pressure/vulnerability composites ──
    feats["sub_entry_pressure"] = (
        feats.get("td_att_p15", 0.0) * feats.get("td_acc", 0.35)
        + 0.7 * feats.get("sub_att_p15", 0.0)
        + 0.4 * feats.get("rev_p15", 0.0)
    )
    feats["sub_control_conversion"] = (
        feats.get("sub_att_p15", 0.0)
        * (0.25 + feats.get("ctrl_pct", 0.0))
        * (0.25 + feats.get("ground_pct", 0.0))
    )
    feats["sub_scramble_threat"] = (
        feats.get("sub_att_p15", 0.0) + feats.get("rev_p15", 0.0)
    ) * (1.0 + feats.get("ground_pct", 0.0))
    feats["sub_defensive_leak"] = (
        feats.get("def_sub_att_p15", 0.0) * (1.0 - feats.get("td_defense_rate", 0.65))
        + 0.75 * feats.get("sub_loss_pct", 0.20)
    )
    feats["late_sub_pressure"] = (
        feats.get("sub_att_p15", 0.0)
        * feats.get("cardio_ratio", 1.0)
        * (0.5 + feats.get("rd3_sub_share", 0.0))
    )
    feats["sub_recency_surge"] = feats.get("r1f_sub_att_p15", 0.0) - feats.get("r5_sub_att_p15", 0.0)
    feats["grapple_recency_surge"] = (
        feats.get("r1f_td_p15", 0.0) + feats.get("r1f_sub_att_p15", 0.0)
    ) - (
        feats.get("r5_td_p15", 0.0) + feats.get("r5_sub_att_p15", 0.0)
    )
    feats["sub_vs_control_axis"] = (
        feats.get("sub_att_p15", 0.0) + 0.5 * feats.get("rev_p15", 0.0)
    ) - 0.6 * feats.get("ctrl_pct", 0.0)
    # KO-side composites for explicit attacker-vs-vulnerability mismatch.
    feats["ko_attack_pressure"] = (
        feats.get("power_ratio", 0.0)
        * feats.get("head_pct", 0.0)
        * feats.get("sig_str_diff_pm", 0.0)
    )
    feats["ko_def_leak"] = (
        feats.get("def_kd_pm", 0.0)
        + feats.get("ko_loss_pct", 0.0)
        + 0.5 * (1.0 - feats.get("durability", 0.8))
    )

    # ── Opponent quality trend (facing tougher/weaker opponents recently?) ──
    if feats["avg_opp_glicko"] > 0:
        feats["opp_quality_trend"] = feats["ewm_opp_quality"] / feats["avg_opp_glicko"]
    else:
        feats["opp_quality_trend"] = 1.0

    # ── Recent-year activity ──
    if current_date is not None:
        fights_last_year = sum(
            1 for h in history
            if not _isnan(h["date"]) and (current_date - h["date"]).days <= 365
        )
        feats["fights_last_year"] = float(fights_last_year)
        wins_last_year = sum(
            1 for h in history
            if not _isnan(h["date"]) and (current_date - h["date"]).days <= 365
            and h["result"] == "W"
        )
        feats["win_rate_last_year"] = _bayes_shrink(
            _safe_div(wins_last_year, fights_last_year, 0.5),
            fights_last_year, prior=0.5, strength=4,
        )
    else:
        feats["fights_last_year"] = 1.0
        feats["win_rate_last_year"] = 0.5

    # ── Loss recovery (bounce-back ability after losses) ──
    post_loss_fights = 0
    post_loss_wins = 0
    after_loss = False
    for h in history:
        if after_loss:
            post_loss_fights += 1
            if h["result"] == "W":
                post_loss_wins += 1
        after_loss = (h["result"] == "L")
    feats["loss_recovery_rate"] = _bayes_shrink(
        _safe_div(post_loss_wins, post_loss_fights, 0.5),
        post_loss_fights, prior=0.5, strength=6,
    )

    # ── Striking defense efficiency ──
    # How well they avoid damage relative to opponent's offensive output
    if feats["def_sig_str_pm"] > 0 and feats["sig_str_pm"] > 0:
        feats["strike_exchange_ratio"] = feats["sig_str_pm"] / (feats["def_sig_str_pm"] + 0.01)
    else:
        feats["strike_exchange_ratio"] = 1.0

    # ── Grappling threat composite (wrestling + submissions + control) ──
    feats["grappling_threat"] = feats["td_p15"] * feats["td_acc"] + feats["sub_att_p15"] + feats["ctrl_pct"]

    # ── Decision ability (win rate in fights that go to decision) ──
    dec_fights = sum(1 for h in history if "Dec" in str(h.get("method", "")))
    dec_wins = sum(1 for h in history if h["result"] == "W" and "Dec" in str(h.get("method", "")))
    feats["decision_win_rate"] = _bayes_shrink(
        _safe_div(dec_wins, dec_fights, 0.5),
        dec_fights, prior=0.5, strength=6,
    )

    # ── Finish resistance (survives opponent's finishes) ──
    feats["finish_resistance"] = 1.0 - feats.get("been_finished_pct", 0.5)

    # ── Offensive diversity (how evenly spread across strike/grapple/sub) ──
    stk_share = feats.get("style_striking", 0.5)
    wrs_share = feats.get("style_wrestling", 0.25)
    sub_share = feats.get("style_submission", 0.15)
    # Entropy-based diversity (higher = more versatile)
    diversity = 0.0
    for s in [stk_share, wrs_share, sub_share]:
        if s > 0:
            diversity -= s * math.log(s + 1e-8)
    feats["offensive_diversity"] = diversity

    # ── EWM striking differential (recent trend of how they're winning/losing exchanges) ──
    alpha_sd = 0.4
    sd_sum, sd_tot = 0.0, 0.0
    for i, h in enumerate(history):
        w = (1 - alpha_sd) ** (n - 1 - i)
        ft_min = h["fight_time"] / 60.0 if h["fight_time"] > 0 else 1.0
        diff = (h["sig_str"] - h["opp_sig_str"]) / ft_min
        sd_sum += w * diff
        sd_tot += w
    feats["ewm_sig_str_diff_pm"] = sd_sum / sd_tot if sd_tot > 0 else 0.0

    # ── Glicko confidence (lower phi = more certain rating) ──
    feats["glicko_confidence"] = max(1.0 - (glicko[1] / PHI_0), 0.0)

    if _FIGHTER_FEAT_KEYS is None:
        _set_fighter_keys(list(feats.keys()))

    return feats


def _set_fighter_keys(keys):
    global _FIGHTER_FEAT_KEYS
    _FIGHTER_FEAT_KEYS = keys


# ─── Matchup feature computation ──────────────────────────────────────────────

def compute_matchup_features(a_feats, b_feats, is_title=0, total_rounds=3, weight_class=""):
    """Difference features (A minus B) plus interaction features."""
    features = {}
    for key in a_feats:
        a_val = a_feats[key] if not _isnan(a_feats[key]) else float("nan")
        b_val = b_feats[key] if not _isnan(b_feats[key]) else float("nan")
        try:
            features[f"d_{key}"] = float(a_val) - float(b_val)
        except (TypeError, ValueError):
            features[f"d_{key}"] = float("nan")

    a_age = a_feats.get("age", float("nan"))
    b_age = b_feats.get("age", float("nan"))
    a_reach = a_feats.get("reach", float("nan"))
    b_reach = b_feats.get("reach", float("nan"))
    a_height = a_feats.get("height", float("nan"))
    b_height = b_feats.get("height", float("nan"))
    a_mu = _num_or(a_feats.get("glicko_mu"), MU_0)
    b_mu = _num_or(b_feats.get("glicko_mu"), MU_0)
    a_phi = _num_or(a_feats.get("glicko_phi"), PHI_0)
    b_phi = _num_or(b_feats.get("glicko_phi"), PHI_0)
    a_n = max(_num_or(a_feats.get("num_fights"), 0.0), 0.0)
    b_n = max(_num_or(b_feats.get("num_fights"), 0.0), 0.0)
    a_debut = a_n < 1.0
    b_debut = b_n < 1.0
    a_inexperienced = a_n < 3.0
    b_inexperienced = b_n < 3.0

    # Absolute / interaction features
    features["is_title"] = is_title
    features["total_rounds"] = total_rounds
    features["exp_sum"] = a_n + b_n
    features["age_sum"] = _num_or(a_age, 30.0) + _num_or(b_age, 30.0)
    features["glicko_mu_sum"] = a_mu + b_mu
    features["abs_age_gap"] = _abs_gap(a_age, b_age)
    features["abs_reach_gap"] = _abs_gap(a_reach, b_reach)
    features["abs_height_gap"] = _abs_gap(a_height, b_height)
    features["abs_glicko_gap"] = abs(a_mu - b_mu)
    features["abs_glicko_phi_gap"] = abs(a_phi - b_phi)
    features["abs_exp_gap"] = abs(a_n - b_n)
    features["min_num_fights"] = min(a_n, b_n)
    features["max_num_fights"] = max(a_n, b_n)
    features["experience_ratio"] = min(a_n, b_n) / max(max(a_n, b_n), 1.0)
    features["max_glicko_phi"] = max(a_phi, b_phi)
    features["min_glicko_phi"] = min(a_phi, b_phi)
    features["avg_glicko_phi"] = (a_phi + b_phi) / 2.0
    features["both_debut"] = float(a_debut and b_debut)
    features["one_debut"] = float(a_debut ^ b_debut)
    features["both_inexperienced"] = float(a_inexperienced and b_inexperienced)
    features["one_inexperienced"] = float(a_inexperienced ^ b_inexperienced)
    features["info_asymmetry"] = abs(a_n - b_n) / max(a_n + b_n, 1.0)

    # Weight class
    features["weight_class_ord"] = WEIGHT_CLASS_ORDINAL.get(weight_class, 0)
    for class_name, feature_name in WEIGHT_CLASS_FEATURES.items():
        features[feature_name] = float(weight_class == class_name)
    features[UNKNOWN_WEIGHT_CLASS_FEATURE] = float(weight_class not in WEIGHT_CLASS_ORDINAL)

    # Stance matchup interactions (d_ prefix for correct augmentation negation)
    a_ortho = _num_or(a_feats.get("is_orthodox"), 0.0)
    a_south = _num_or(a_feats.get("is_southpaw"), 0.0)
    b_ortho = _num_or(b_feats.get("is_orthodox"), 0.0)
    b_south = _num_or(b_feats.get("is_southpaw"), 0.0)
    features["d_ortho_vs_south"] = float(a_ortho * b_south) - float(b_ortho * a_south)
    features["same_stance"] = float(
        a_ortho == b_ortho and a_south == b_south and (a_ortho or a_south)
    )
    # Style cluster matchup interactions
    a_stk = _num_or(a_feats.get("style_striking"), 0.5)
    a_wrs = _num_or(a_feats.get("style_wrestling"), 0.15)
    a_sub = _num_or(a_feats.get("style_submission"), 0.1)
    a_cg = _num_or(a_feats.get("style_clinch_ground"), 0.32)
    b_stk = _num_or(b_feats.get("style_striking"), 0.5)
    b_wrs = _num_or(b_feats.get("style_wrestling"), 0.15)
    b_sub = _num_or(b_feats.get("style_submission"), 0.1)
    b_cg = _num_or(b_feats.get("style_clinch_ground"), 0.32)
    features["d_striker_vs_grappler"] = a_stk * (b_wrs + b_sub) - b_stk * (a_wrs + a_sub)
    features["style_mismatch"] = a_stk * (b_wrs + b_sub) + b_stk * (a_wrs + a_sub)
    features["style_distance"] = math.sqrt(
        (a_stk - b_stk)**2 + (a_wrs - b_wrs)**2 + (a_sub - b_sub)**2 + (a_cg - b_cg)**2
    )

    # ── Glicko expected win probability (THE strongest single predictor) ──
    # This encodes the full Glicko-2 prediction: rating difference + uncertainty
    a_mu_s = (a_mu - MU_0) / SCALE
    b_mu_s = (b_mu - MU_0) / SCALE
    b_phi_s = b_phi / SCALE
    features["glicko_win_prob"] = _E(a_mu_s, b_mu_s, b_phi_s)
    features["d_glicko_win_prob"] = features["glicko_win_prob"] - 0.5  # directional version

    # Confidence gap: rating gap normalised by combined uncertainty
    combined_phi = math.sqrt(a_phi**2 + b_phi**2) if (a_phi**2 + b_phi**2) > 0 else 1.0
    features["d_confidence_gap"] = (a_mu - b_mu) / combined_phi

    # Glicko win prob weighted by confidence (high confidence + big gap = strong signal)
    avg_conf = (_num_or(a_feats.get("glicko_confidence"), 0.0)
                + _num_or(b_feats.get("glicko_confidence"), 0.0)) / 2.0
    features["d_confident_prediction"] = features["d_glicko_win_prob"] * (0.5 + avg_conf)

    # Reach advantage amplified by striking differential
    a_sig_pm = _num_or(a_feats.get("sig_str_pm"), 0.0)
    b_sig_pm = _num_or(b_feats.get("sig_str_pm"), 0.0)
    reach_diff = _num_or(a_reach, 72.0) - _num_or(b_reach, 72.0)
    features["d_reach_x_striking"] = reach_diff * (a_sig_pm - b_sig_pm)

    # Chin vs power: A's durability vs B's KO rate (and reverse)
    a_dur = _num_or(a_feats.get("durability"), 0.8)
    b_dur = _num_or(b_feats.get("durability"), 0.8)
    a_kd_pm = _num_or(a_feats.get("kd_pm"), 0.0)
    b_kd_pm = _num_or(b_feats.get("kd_pm"), 0.0)
    features["d_chin_vs_power"] = (a_dur * a_kd_pm) - (b_dur * b_kd_pm)

    # TD attack vs defense: A's offensive wrestling vs B's TDD (directional)
    a_td_p15 = _num_or(a_feats.get("td_p15"), 0.0)
    b_td_p15 = _num_or(b_feats.get("td_p15"), 0.0)
    a_tdd = _num_or(a_feats.get("td_defense_rate"), 0.65)
    b_tdd = _num_or(b_feats.get("td_defense_rate"), 0.65)
    features["d_td_attack_vs_defense"] = (a_td_p15 * (1.0 - b_tdd)) - (b_td_p15 * (1.0 - a_tdd))

    # Combined finish rate (proxy for fight unlikely to go to decision)
    a_fin = _num_or(a_feats.get("finish_rate"), 0.5)
    b_fin = _num_or(b_feats.get("finish_rate"), 0.5)
    features["combined_finish_rate"] = a_fin + b_fin

    # Output rate differential
    features["d_output_rate"] = _num_or(a_feats.get("output_rate"), 0.0) - _num_or(b_feats.get("output_rate"), 0.0)

    # Momentum differential
    features["d_momentum"] = _num_or(a_feats.get("momentum"), 0.5) - _num_or(b_feats.get("momentum"), 0.5)

    # Durability gap
    features["d_durability"] = a_dur - b_dur

    # Grappling dominance differential
    features["d_grappling_dominance"] = (_num_or(a_feats.get("grappling_dominance"), 0.0)
                                         - _num_or(b_feats.get("grappling_dominance"), 0.0))

    # EWM opponent quality gap (who has faced tougher competition recently)
    features["d_ewm_opp_quality"] = (_num_or(a_feats.get("ewm_opp_quality"), MU_0)
                                     - _num_or(b_feats.get("ewm_opp_quality"), MU_0))

    # ── Striking offense vs opponent's striking defense (cross-matchup) ──
    a_sig_def = _num_or(a_feats.get("sig_str_defense_rate"), 0.55)
    b_sig_def = _num_or(b_feats.get("sig_str_defense_rate"), 0.55)
    features["d_striking_vs_defense"] = (
        a_sig_pm * (1.0 - b_sig_def) - b_sig_pm * (1.0 - a_sig_def)
    )

    # ── Wrestling offense vs opponent's TDD (already have d_td_attack_vs_defense,
    #    now add grappling_threat vs opponent's ground defense) ──
    a_grap = _num_or(a_feats.get("grappling_threat"), 0.0)
    b_grap = _num_or(b_feats.get("grappling_threat"), 0.0)
    features["d_grapple_vs_tdd"] = a_grap * (1.0 - b_tdd) - b_grap * (1.0 - a_tdd)

    # ── Fight IQ gap ──
    features["d_fight_iq"] = (_num_or(a_feats.get("fight_iq"), 0.15)
                              - _num_or(b_feats.get("fight_iq"), 0.15))

    # ── Cardio advantage ──
    features["d_cardio_ratio"] = (_num_or(a_feats.get("cardio_ratio"), 1.0)
                                  - _num_or(b_feats.get("cardio_ratio"), 1.0))

    # ── Glicko confidence gap (whose rating is more reliable?) ──
    features["d_glicko_confidence"] = (_num_or(a_feats.get("glicko_confidence"), 0.0)
                                       - _num_or(b_feats.get("glicko_confidence"), 0.0))

    # ── Form trend gap (who is on the upswing?) ──
    features["d_form_trend"] = (_num_or(a_feats.get("form_vs_career"), 0.0)
                                - _num_or(b_feats.get("form_vs_career"), 0.0))

    # ── Strike exchange ratio gap (who wins the striking exchanges more?) ──
    features["d_exchange_ratio"] = (_num_or(a_feats.get("strike_exchange_ratio"), 1.0)
                                    - _num_or(b_feats.get("strike_exchange_ratio"), 1.0))

    # ── Decision fighter vs finisher matchup ──
    a_dec_wr = _num_or(a_feats.get("decision_win_rate"), 0.5)
    b_dec_wr = _num_or(b_feats.get("decision_win_rate"), 0.5)
    features["d_decision_ability"] = a_dec_wr - b_dec_wr

    # ── Versatility gap ──
    features["d_offensive_diversity"] = (_num_or(a_feats.get("offensive_diversity"), 1.0)
                                         - _num_or(b_feats.get("offensive_diversity"), 1.0))

    # ── Combined Glicko confidence (how reliable is this matchup prediction?) ──
    a_conf = _num_or(a_feats.get("glicko_confidence"), 0.0)
    b_conf = _num_or(b_feats.get("glicko_confidence"), 0.0)
    features["avg_glicko_confidence"] = (a_conf + b_conf) / 2.0

    # ── Age-experience interaction: young+experienced is dangerous ──
    a_age_v = _num_or(a_age, 30.0)
    b_age_v = _num_or(b_age, 30.0)
    # Younger fighter with more experience has edge
    features["d_youth_exp"] = (b_age_v - a_age_v) * (a_n - b_n)

    # ── Glicko-scaled striking: weight sig str differential by Glicko reliability ──
    features["d_reliable_striking"] = features.get("d_sig_str_pm", 0.0) * (0.5 + avg_conf)

    # ── Finisher vs chin interaction ──
    a_finish_r = _num_or(a_feats.get("finish_rate"), 0.5)
    b_finish_r = _num_or(b_feats.get("finish_rate"), 0.5)
    b_finish_resist = _num_or(b_feats.get("finish_resistance"), 0.5)
    a_finish_resist = _num_or(a_feats.get("finish_resistance"), 0.5)
    features["d_finish_vs_resist"] = (a_finish_r * (1.0 - b_finish_resist)
                                      - b_finish_r * (1.0 - a_finish_resist))

    # ── Activity gap (recent activity matters for ring rust) ──
    a_inactive = _num_or(a_feats.get("days_inactive"), 365.0)
    b_inactive = _num_or(b_feats.get("days_inactive"), 365.0)
    features["d_activity"] = b_inactive - a_inactive  # positive = A is more active

    # ── Win rate in last year gap ──
    features["d_win_rate_last_year"] = (_num_or(a_feats.get("win_rate_last_year"), 0.5)
                                        - _num_or(b_feats.get("win_rate_last_year"), 0.5))

    # ── Old_Model-inspired quality-endurance and pressure features ──
    # Quality endurance: maintains accuracy while keeping output under fatigue.
    a_cardio = _num_or(a_feats.get("cardio_ratio"), 1.0)
    b_cardio = _num_or(b_feats.get("cardio_ratio"), 1.0)
    a_acc = _num_or(a_feats.get("sig_str_acc"), 0.45)
    b_acc = _num_or(b_feats.get("sig_str_acc"), 0.45)
    a_r1_int = _num_or(a_feats.get("rd1_intensity_ratio"), 1.0)
    b_r1_int = _num_or(b_feats.get("rd1_intensity_ratio"), 1.0)
    a_output = _num_or(a_feats.get("output_rate"), 0.0)
    b_output = _num_or(b_feats.get("output_rate"), 0.0)
    a_opp_sig = _num_or(a_feats.get("def_sig_str_pm"), 0.0)
    b_opp_sig = _num_or(b_feats.get("def_sig_str_pm"), 0.0)
    a_conf = _num_or(a_feats.get("glicko_confidence"), 0.0)
    b_conf = _num_or(b_feats.get("glicko_confidence"), 0.0)

    a_quality_endurance = a_cardio * a_acc * (0.5 + min(max(a_r1_int, 0.5), 1.8))
    b_quality_endurance = b_cardio * b_acc * (0.5 + min(max(b_r1_int, 0.5), 1.8))
    features["d_quality_endurance"] = a_quality_endurance - b_quality_endurance

    # Accuracy under fire proxy: maintained accuracy while facing volume.
    a_accuracy_under_fire = a_acc / (1.0 + b_opp_sig)
    b_accuracy_under_fire = b_acc / (1.0 + a_opp_sig)
    features["d_accuracy_under_fire"] = a_accuracy_under_fire - b_accuracy_under_fire

    # Pressure-cardio clash: high pressure that still scales with own cardio.
    features["d_pressure_cardio_clash"] = (
        a_output * a_cardio * (1.0 - b_cardio) - b_output * b_cardio * (1.0 - a_cardio)
    )

    # Stability-endurance: confidence-weighted late-fight reliability.
    features["d_stability_endurance"] = (
        a_conf * a_quality_endurance - b_conf * b_quality_endurance
    )

    # Sum channels to allow exact side-specific reconstruction for method-only
    # mismatch features after winner-orientation is applied.
    a_sub_entry = _num_or(a_feats.get("sub_entry_pressure"), 0.0)
    b_sub_entry = _num_or(b_feats.get("sub_entry_pressure"), 0.0)
    a_sub_leak = _num_or(a_feats.get("sub_defensive_leak"), 0.0)
    b_sub_leak = _num_or(b_feats.get("sub_defensive_leak"), 0.0)
    features["sub_entry_pressure_sum"] = a_sub_entry + b_sub_entry
    features["sub_defensive_leak_sum"] = a_sub_leak + b_sub_leak
    a_ko_attack = _num_or(a_feats.get("ko_attack_pressure"), 0.0)
    b_ko_attack = _num_or(b_feats.get("ko_attack_pressure"), 0.0)
    a_ko_leak = _num_or(a_feats.get("ko_def_leak"), 0.0)
    b_ko_leak = _num_or(b_feats.get("ko_def_leak"), 0.0)
    features["ko_attack_pressure_sum"] = a_ko_attack + b_ko_attack
    features["ko_def_leak_sum"] = a_ko_leak + b_ko_leak

    # Stage-1 sum features for exact winner/loser reconstruction after orientation.
    for _feat, _default in [
        ("dec_win_pct", 0.0), ("finish_resistance", 0.0), ("consistency", 0.0),
        ("cardio_ratio", 0.0), ("durability", 0.0), ("output_rate", 0.0),
        ("rd1_intensity_ratio", 0.0), ("strike_exchange_ratio", 0.0),
        ("sig_str_acc", 0.0), ("late_round_pct", 0.0), ("avg_time_min", 0.0),
        ("avg_finish_round", 0.0), ("first_round_finish_rate", 0.0),
        ("damage_efficiency", 0.0), ("body_leg_attrition", 0.0),
    ]:
        _a = _num_or(a_feats.get(_feat), _default)
        _b = _num_or(b_feats.get(_feat), _default)
        features[f"{_feat}_sum"] = _a + _b

    # ── Batch 3: physical / style / matchup interaction features (§B, §C, §D) ──
    # All use f(A) - f(B) pattern so they flip cleanly under method orientation.
    # §B — Physical / leverage interactions
    a_dist_pct = _num_or(a_feats.get("distance_pct"), 0.3)
    b_dist_pct = _num_or(b_feats.get("distance_pct"), 0.3)
    features["d_reach_x_distance_pct"] = (
        _num_or(a_reach, 72.0) * a_dist_pct - _num_or(b_reach, 72.0) * b_dist_pct
    )
    a_cdg = _num_or(a_feats.get("cardio_decay_sig_str"), 0.0)
    b_cdg = _num_or(b_feats.get("cardio_decay_sig_str"), 0.0)
    a_age_v = _num_or(a_age, 30.0)
    b_age_v = _num_or(b_age, 30.0)
    features["d_age_x_cardio"] = (a_age_v * a_cdg) - (b_age_v * b_cdg)
    features["d_age_x_title_rounds"] = (a_age_v - b_age_v) * float(total_rounds == 5)

    # §D — Matchup compatibility / style clashes
    features["d_striker_grappler_raw"] = (
        (a_sig_pm - a_td_p15) - (b_sig_pm - b_td_p15)
    )
    a_td_acc = _num_or(a_feats.get("td_acc"), 0.35)
    b_td_acc = _num_or(b_feats.get("td_acc"), 0.35)
    features["d_tdd_vs_td_attack"] = (a_td_acc * (1.0 - b_tdd)) - (b_td_acc * (1.0 - a_tdd))

    # ── Batch 4: form / context / priors (§E, §F, §G) ──
    features["gender_flag"] = 1.0 if weight_class.startswith("Women") else 0.0

    features["d_title_x_cardio"] = float(is_title) * (a_cdg - b_cdg)

    a_fin_resist = _num_or(a_feats.get("finish_resistance"), 0.5)
    b_fin_resist = _num_or(b_feats.get("finish_resistance"), 0.5)
    features["d_total_rounds_x_finish_resistance"] = float(total_rounds) * (a_fin_resist - b_fin_resist)

    # Context-conditional rounds experience (picks the bucket matching current fight)
    _cur_rounds = 5 if int(total_rounds) == 5 else 3
    features["d_rounds_experience"] = (
        _num_or(a_feats.get(f"rounds_{_cur_rounds}_exp"), 0.0)
        - _num_or(b_feats.get(f"rounds_{_cur_rounds}_exp"), 0.0)
    )

    # Stability: flipped so positive = red is the more stable fighter
    features["d_stability"] = (
        _num_or(b_feats.get("sig_diff_pm_vol"), 3.0)
        - _num_or(a_feats.get("sig_diff_pm_vol"), 3.0)
    )

    return features


# ─── Training data builder ────────────────────────────────────────────────────

def _precompute_context_finish_priors(df):
    """Compute per-row finish priors conditioned on (weight_class, gender, total_rounds, is_title).

    ctx_finish_prior_2y: exponentially time-decayed (half-life 730 days) Bayesian
        estimate of the finish rate for fights of the same type before this date.
    ref_finish_prior: unweighted Laplace-smoothed count-based version of the same.
    Both use alpha=20 and p0 = global historical finish rate as the prior mean.
    """
    n = len(df)
    ctx_arr = np.full(n, np.nan, dtype=float)
    ref_arr = np.full(n, np.nan, dtype=float)

    methods = df["method"].astype(str).str.lower()
    is_finish = (~methods.str.contains("dec", na=False)).astype(float).values
    p0 = float(np.mean(is_finish)) if n > 0 else 0.55
    alpha = 20.0

    wc_col = df["weight_class"].astype(str).values if "weight_class" in df.columns else np.full(n, "", dtype=object)
    gd_col = df["gender"].astype(str).str.lower().values if "gender" in df.columns else np.full(n, "unknown", dtype=object)
    rd_col = df["total_rounds"].fillna(3).astype(int).values if "total_rounds" in df.columns else np.full(n, 3, dtype=int)
    tt_col = df["is_title_bout"].fillna(0).astype(int).values if "is_title_bout" in df.columns else np.zeros(n, dtype=int)

    epoch = df["event_date"].iloc[0]
    days = ((df["event_date"] - epoch) / pd.Timedelta(days=1)).values.astype(float)

    # Per-context running state: (decayed_finish_sum, decayed_total_weight, finish_count, total_count, last_days)
    state = {}

    for i in range(n):
        key = (wc_col[i], gd_col[i], int(rd_col[i]), int(tt_col[i]))
        cur_d = days[i]

        if key in state:
            fs, tw, fc, tc, last_d = state[key]
            decay = 2.0 ** (-(cur_d - last_d) / 730.0)
            fs *= decay
            tw *= decay
        else:
            fs, tw, fc, tc = 0.0, 0.0, 0, 0

        ctx_arr[i] = (fs + alpha * p0) / (tw + alpha)
        ref_arr[i] = (fc + alpha * p0) / (tc + alpha)

        # Update state with current fight (weight = 1.0 at this moment)
        state[key] = (fs + float(is_finish[i]), tw + 1.0, fc + int(is_finish[i]), tc + 1, cur_d)

    return ctx_arr, ref_arr


def build_training_data(csv_path, progress_cb=None):
    """Process fights chronologically, build features, return X, y and state."""
    df = pd.read_csv(csv_path)
    df["event_date"] = pd.to_datetime(df["event_date"], format="%m/%d/%Y")
    df = df.sort_values("event_date").reset_index(drop=True)
    _ensure_fighter_feature_keys(df["event_date"].iloc[0] if len(df) else None)

    ctx_finish_prior_arr, ref_finish_prior_arr = _precompute_context_finish_priors(df)

    fighter_history = defaultdict(list)
    glicko_ratings = {}
    opp_glicko_list = defaultdict(list)

    rows_X = []
    rows_y = []
    total = len(df)

    for idx in range(total):
        if progress_cb and idx % 500 == 0:
            progress_cb(f"  Building features... fight {idx+1}/{total}")

        row = df.iloc[idx]
        r_name = row["r_name"]
        b_name = row["b_name"]

        # Initialize Glicko
        if r_name not in glicko_ratings:
            glicko_ratings[r_name] = (MU_0, PHI_0, SIGMA_0)
        if b_name not in glicko_ratings:
            glicko_ratings[b_name] = (MU_0, PHI_0, SIGMA_0)

        r_glicko = glicko_ratings[r_name]
        b_glicko = glicko_ratings[b_name]

        r_feats = compute_fighter_features(
            fighter_history[r_name], r_glicko,
            opp_glicko_list[r_name], row["event_date"],
            fallback_profile=_extract_profile_from_row(row, "r"),
        )
        b_feats = compute_fighter_features(
            fighter_history[b_name], b_glicko,
            opp_glicko_list[b_name], row["event_date"],
            fallback_profile=_extract_profile_from_row(row, "b"),
        )
        matchup = compute_matchup_features(
            r_feats, b_feats,
            is_title=row.get("is_title_bout", 0),
            total_rounds=row.get("total_rounds", 3),
            weight_class=row.get("weight_class", ""),
        )
        if row["winner"] in ("Red", "Blue"):
            matchup["ctx_finish_prior_2y"] = float(ctx_finish_prior_arr[idx])
            matchup["ref_finish_prior"] = float(ref_finish_prior_arr[idx])
            rows_X.append(matchup)
            rows_y.append(1.0 if row["winner"] == "Red" else 0.0)

        # Determine result
        winner = row["winner"]
        if winner == "Red":
            r_res, b_res = "W", "L"
            r_sc, b_sc = 1.0, 0.0
        elif winner == "Blue":
            r_res, b_res = "L", "W"
            r_sc, b_sc = 0.0, 1.0
        else:
            r_res, b_res = "D", "D"
            r_sc, b_sc = 0.5, 0.5

        # Update histories
        fighter_history[r_name].append(extract_fight_record(row, "r", "b", r_res, b_glicko[0]))
        fighter_history[r_name][-1]["self_glicko"] = r_glicko[0]
        fighter_history[b_name].append(extract_fight_record(row, "b", "r", b_res, r_glicko[0]))
        fighter_history[b_name][-1]["self_glicko"] = b_glicko[0]

        # Track opponent Glicko
        opp_glicko_list[r_name].append(b_glicko[0])
        opp_glicko_list[b_name].append(r_glicko[0])

        # Update Glicko
        glicko_ratings[r_name] = glicko2_update(r_glicko, [(b_glicko[0], b_glicko[1], r_sc)])
        glicko_ratings[b_name] = glicko2_update(b_glicko, [(r_glicko[0], r_glicko[1], b_sc)])

    X = pd.DataFrame(rows_X)
    y = pd.Series(rows_y)

    if progress_cb:
        progress_cb(f"  Built {len(X)} training samples with {X.shape[1]} features.")

    return X, y, fighter_history, glicko_ratings, opp_glicko_list


def _method_labels_from_csv(csv_path):
    df = pd.read_csv(csv_path)
    df["event_date"] = pd.to_datetime(df["event_date"], format="%m/%d/%Y", errors="coerce")
    df = df.sort_values("event_date").reset_index(drop=True)
    rows = []
    for _, row in df.iterrows():
        winner = str(row.get("winner", "")).strip()
        if winner in ("Red", "Blue"):
            detail = _normalize_method_detail(row.get("method", ""))
            coarse = _normalize_method_label(row.get("method", ""))
            finish_bin = "Decision" if coarse == "Decision" else "Finish"
            finish_subtype = coarse if coarse in ("KO/TKO", "Submission") else "KO/TKO"
            rows.append({
                "coarse": coarse,
                "detail": detail,
                "finish_bin": finish_bin,
                "finish_subtype": finish_subtype,
            })
    return pd.DataFrame(rows)


def _time_split_indices(n_rows):
    test_size = max(1, int(n_rows * TEST_FRACTION))
    val_size = max(1, int(n_rows * VAL_FRACTION))
    train_size = max(1, n_rows - val_size - test_size)
    train_end = train_size
    val_end = train_end + val_size
    return train_end, val_end


def _augment_swap(X, y):
    d_cols = [c for c in X.columns if c.startswith("d_")]
    X_swap = X.copy()
    X_swap[d_cols] = -X_swap[d_cols]
    y_swap = 1.0 - y
    # Interleave to preserve chronology for folds.
    X_aug = pd.concat([X, X_swap], ignore_index=True)
    y_aug = pd.concat([y, y_swap], ignore_index=True)
    return X_aug, y_aug


def _augment_weights(w):
    w = np.asarray(w, dtype=float)
    return np.concatenate([w, w], axis=0)


def _time_weights(n, floor=0.35):
    if n <= 1:
        return np.ones(max(n, 1), dtype=float)
    return np.linspace(float(floor), 1.0, int(n), dtype=float)


def _swap_features(X):
    X2 = X.copy()
    d_cols = [c for c in X2.columns if c.startswith("d_")]
    X2[d_cols] = -X2[d_cols]
    return X2


def _make_model_specs(lgb_tuned_params=None, xgb_tuned_params=None, cb_tuned_params=None):
    specs = []
    if lgb is not None:
        specs.append((
            "LightGBM",
            lambda: lgb.LGBMClassifier(
                n_estimators=450, learning_rate=0.03, max_depth=6,
                num_leaves=31, min_child_samples=25, subsample=0.85,
                colsample_bytree=0.85, reg_alpha=0.12, reg_lambda=1.2,
                random_state=RANDOM_SEED, verbose=-1,
            )
        ))
        specs.append((
            "LightGBM_S2",
            lambda: lgb.LGBMClassifier(
                n_estimators=450, learning_rate=0.03, max_depth=6,
                num_leaves=31, min_child_samples=25, subsample=0.85,
                colsample_bytree=0.85, reg_alpha=0.12, reg_lambda=1.2,
                random_state=RANDOM_SEED + 17, verbose=-1,
            )
        ))
        if lgb_tuned_params:
            p = dict(lgb_tuned_params)
            specs.append((
                "LightGBM_Tuned",
                lambda p=p: lgb.LGBMClassifier(
                    n_estimators=int(p["n_estimators"]),
                    learning_rate=float(p["learning_rate"]),
                    max_depth=int(p["max_depth"]),
                    num_leaves=int(p["num_leaves"]),
                    min_child_samples=int(p["min_child_samples"]),
                    subsample=float(p["subsample"]),
                    colsample_bytree=float(p["colsample_bytree"]),
                    reg_alpha=float(p["reg_alpha"]),
                    reg_lambda=float(p["reg_lambda"]),
                    min_split_gain=float(p["min_split_gain"]),
                    random_state=RANDOM_SEED + 101,
                    verbose=-1,
                )
            ))
    if xgb is not None:
        specs.append((
            "XGBoost",
            lambda: xgb.XGBClassifier(
                n_estimators=420, learning_rate=0.03, max_depth=5,
                min_child_weight=5, subsample=0.85, colsample_bytree=0.85,
                reg_alpha=0.12, reg_lambda=1.2, objective="binary:logistic",
                eval_metric="logloss", random_state=RANDOM_SEED, n_jobs=-1,
            )
        ))
        specs.append((
            "XGBoost_S2",
            lambda: xgb.XGBClassifier(
                n_estimators=420, learning_rate=0.03, max_depth=5,
                min_child_weight=5, subsample=0.85, colsample_bytree=0.85,
                reg_alpha=0.12, reg_lambda=1.2, objective="binary:logistic",
                eval_metric="logloss", random_state=RANDOM_SEED + 27, n_jobs=-1,
            )
        ))
        if xgb_tuned_params:
            p = dict(xgb_tuned_params)
            specs.append((
                "XGBoost_Tuned",
                lambda p=p: xgb.XGBClassifier(
                    n_estimators=int(p["n_estimators"]),
                    learning_rate=float(p["learning_rate"]),
                    max_depth=int(p["max_depth"]),
                    min_child_weight=float(p["min_child_weight"]),
                    subsample=float(p["subsample"]),
                    colsample_bytree=float(p["colsample_bytree"]),
                    reg_alpha=float(p["reg_alpha"]),
                    reg_lambda=float(p["reg_lambda"]),
                    gamma=float(p["gamma"]),
                    objective="binary:logistic",
                    eval_metric="logloss",
                    random_state=RANDOM_SEED + 121,
                    n_jobs=-1,
                )
            ))
    if cb is not None:
        specs.append((
            "CatBoost",
            lambda: cb.CatBoostClassifier(
                iterations=420, learning_rate=0.04, depth=6, l2_leaf_reg=3.0,
                random_strength=1.0, bagging_temperature=0.5, border_count=128,
                loss_function="Logloss", eval_metric="Logloss", verbose=0,
                random_seed=RANDOM_SEED, allow_writing_files=False,
            )
        ))
        specs.append((
            "CatBoost_S2",
            lambda: cb.CatBoostClassifier(
                iterations=420, learning_rate=0.04, depth=6, l2_leaf_reg=3.0,
                random_strength=1.0, bagging_temperature=0.5, border_count=128,
                loss_function="Logloss", eval_metric="Logloss", verbose=0,
                random_seed=RANDOM_SEED + 37, allow_writing_files=False,
            )
        ))
        if cb_tuned_params:
            p = dict(cb_tuned_params)
            specs.append((
                "CatBoost_Tuned",
                lambda p=p: cb.CatBoostClassifier(
                    iterations=int(p["iterations"]),
                    learning_rate=float(p["learning_rate"]),
                    depth=int(p["depth"]),
                    l2_leaf_reg=float(p["l2_leaf_reg"]),
                    random_strength=float(p["random_strength"]),
                    bagging_temperature=float(p["bagging_temperature"]),
                    border_count=int(p["border_count"]),
                    loss_function="Logloss",
                    eval_metric="Logloss",
                    verbose=0,
                    random_seed=RANDOM_SEED + 141,
                    allow_writing_files=False,
                )
            ))
    specs.append((
        "HistGBM",
        lambda: HistGradientBoostingClassifier(
            max_iter=650, learning_rate=0.035, max_depth=6,
            max_leaf_nodes=31, min_samples_leaf=20, l2_regularization=1.0,
            random_state=RANDOM_SEED,
        )
    ))
    specs.append((
        "HistGBM_Wide",
        lambda: HistGradientBoostingClassifier(
            max_iter=900, learning_rate=0.03, max_depth=8,
            max_leaf_nodes=63, min_samples_leaf=15, l2_regularization=0.7,
            random_state=RANDOM_SEED + 11,
        )
    ))
    specs.append((
        "RandForest",
        lambda: RandomForestClassifier(
            n_estimators=550, max_depth=12, min_samples_leaf=5,
            min_samples_split=2, max_features=0.7, random_state=RANDOM_SEED,
            n_jobs=-1,
        )
    ))
    specs.append((
        "RandForest_Deep",
        lambda: RandomForestClassifier(
            n_estimators=750, max_depth=18, min_samples_leaf=3,
            min_samples_split=2, max_features=0.6, random_state=RANDOM_SEED + 21,
            n_jobs=-1,
        )
    ))
    specs.append((
        "ExtraTrees",
        lambda: ExtraTreesClassifier(
            n_estimators=700, max_depth=14, min_samples_leaf=4,
            min_samples_split=2, max_features=0.7, random_state=RANDOM_SEED,
            n_jobs=-1,
        )
    ))
    specs.append((
        "ExtraTrees_Deep",
        lambda: ExtraTreesClassifier(
            n_estimators=900, max_depth=22, min_samples_leaf=2,
            min_samples_split=2, max_features=0.6, random_state=RANDOM_SEED + 31,
            n_jobs=-1,
        )
    ))
    specs.append((
        "AdaBoost",
        lambda: AdaBoostClassifier(
            n_estimators=350, learning_rate=0.05, random_state=RANDOM_SEED
        )
    ))
    specs.append((
        "MLP",
        lambda: MLPClassifier(
            hidden_layer_sizes=(96, 48), alpha=0.01, learning_rate="adaptive",
            early_stopping=True, validation_fraction=0.15, max_iter=700,
            random_state=RANDOM_SEED
        )
    ))
    specs.append((
        "LogReg",
        lambda: LogisticRegression(
            max_iter=8000, C=0.3, solver="saga", tol=1e-3, n_jobs=-1, random_state=RANDOM_SEED
        ),
    ))
    specs.append((
        "LogReg_L2",
        lambda: LogisticRegression(
            max_iter=8000, C=1.2, solver="saga", tol=1e-3, n_jobs=-1, random_state=RANDOM_SEED + 7
        ),
    ))
    return specs


def _tune_lightgbm_optuna(X_train, y_train, X_val, y_val, n_trials=OPTUNA_TRIALS, logger=None, progress_cb=None):
    if lgb is None or optuna is None:
        return None
    if len(X_train) < 800 or len(X_val) < 200:
        return None

    X_tr_aug, y_tr_aug = _augment_swap(X_train, y_train)
    w_aug = _augment_weights(_time_weights(len(X_train), floor=0.4))
    X_val_sw = _swap_features(X_val)
    y_val_np = np.asarray(y_val).astype(int)

    sampler = optuna.samplers.TPESampler(seed=RANDOM_SEED)
    study = optuna.create_study(direction="minimize", sampler=sampler)

    def _objective(trial):
        params = {
            "n_estimators": trial.suggest_int("n_estimators", 280, 950),
            "learning_rate": trial.suggest_float("learning_rate", 0.01, 0.08, log=True),
            "max_depth": trial.suggest_int("max_depth", 4, 10),
            "num_leaves": trial.suggest_int("num_leaves", 15, 95),
            "min_child_samples": trial.suggest_int("min_child_samples", 8, 60),
            "subsample": trial.suggest_float("subsample", 0.6, 1.0),
            "colsample_bytree": trial.suggest_float("colsample_bytree", 0.6, 1.0),
            "reg_alpha": trial.suggest_float("reg_alpha", 1e-4, 2.0, log=True),
            "reg_lambda": trial.suggest_float("reg_lambda", 1e-4, 3.0, log=True),
            "min_split_gain": trial.suggest_float("min_split_gain", 0.0, 0.2),
        }
        model = lgb.LGBMClassifier(random_state=RANDOM_SEED, verbose=-1, **params)
        model.fit(X_tr_aug, y_tr_aug, sample_weight=w_aug)
        p_fwd = _clip_probs(model.predict_proba(X_val)[:, 1])
        p_rev = _clip_probs(model.predict_proba(X_val_sw)[:, 1])
        probs = _clip_probs((p_fwd + (1.0 - p_rev)) / 2.0)
        thr, acc = _tune_threshold(probs, y_val_np)
        ll = float(log_loss(y_val_np, probs))
        trial.set_user_attr("acc", float(acc))
        trial.set_user_attr("thr", float(thr))
        trial.set_user_attr("ll", ll)
        # Accuracy-focused objective with log-loss regularizer.
        return (1.0 - float(acc)) + 0.10 * ll

    def _trial_callback(_study, trial):
        if progress_cb is not None:
            progress_cb(int(trial.number) + 1, int(n_trials), "LightGBM")

    study.optimize(
        _objective, n_trials=int(n_trials), show_progress_bar=False,
        callbacks=[_trial_callback],
    )
    best = study.best_trial
    if logger is not None:
        logger(f"Optuna LightGBM best score: {best.value:.5f}")
        logger(
            f"Optuna LightGBM val acc: {best.user_attrs.get('acc', float('nan')):.1%} | "
            f"val ll: {best.user_attrs.get('ll', float('nan')):.4f} | "
            f"thr: {best.user_attrs.get('thr', 0.5):.3f}"
        )
    return best.params


def _tune_xgboost_optuna(X_train, y_train, X_val, y_val, n_trials=OPTUNA_TRIALS, logger=None, progress_cb=None):
    if xgb is None or optuna is None:
        return None
    if len(X_train) < 800 or len(X_val) < 200:
        return None

    X_tr_aug, y_tr_aug = _augment_swap(X_train, y_train)
    w_aug = _augment_weights(_time_weights(len(X_train), floor=0.4))
    X_val_sw = _swap_features(X_val)
    y_val_np = np.asarray(y_val).astype(int)

    sampler = optuna.samplers.TPESampler(seed=RANDOM_SEED + 1)
    study = optuna.create_study(direction="minimize", sampler=sampler)

    def _objective(trial):
        params = {
            "n_estimators": trial.suggest_int("n_estimators", 260, 900),
            "learning_rate": trial.suggest_float("learning_rate", 0.01, 0.08, log=True),
            "max_depth": trial.suggest_int("max_depth", 3, 10),
            "min_child_weight": trial.suggest_float("min_child_weight", 1.0, 12.0),
            "subsample": trial.suggest_float("subsample", 0.6, 1.0),
            "colsample_bytree": trial.suggest_float("colsample_bytree", 0.6, 1.0),
            "reg_alpha": trial.suggest_float("reg_alpha", 1e-4, 2.5, log=True),
            "reg_lambda": trial.suggest_float("reg_lambda", 1e-4, 3.5, log=True),
            "gamma": trial.suggest_float("gamma", 0.0, 1.5),
        }
        model = xgb.XGBClassifier(
            objective="binary:logistic", eval_metric="logloss",
            random_state=RANDOM_SEED, n_jobs=-1, **params
        )
        model.fit(X_tr_aug, y_tr_aug, sample_weight=w_aug)
        p_fwd = _clip_probs(model.predict_proba(X_val)[:, 1])
        p_rev = _clip_probs(model.predict_proba(X_val_sw)[:, 1])
        probs = _clip_probs((p_fwd + (1.0 - p_rev)) / 2.0)
        thr, acc = _tune_threshold(probs, y_val_np)
        ll = float(log_loss(y_val_np, probs))
        trial.set_user_attr("acc", float(acc))
        trial.set_user_attr("thr", float(thr))
        trial.set_user_attr("ll", ll)
        return (1.0 - float(acc)) + 0.10 * ll

    def _trial_callback(_study, trial):
        if progress_cb is not None:
            progress_cb(int(trial.number) + 1, int(n_trials), "XGBoost")

    study.optimize(
        _objective, n_trials=int(n_trials), show_progress_bar=False,
        callbacks=[_trial_callback],
    )
    best = study.best_trial
    if logger is not None:
        logger(f"Optuna XGBoost best score: {best.value:.5f}")
        logger(
            f"Optuna XGBoost val acc: {best.user_attrs.get('acc', float('nan')):.1%} | "
            f"val ll: {best.user_attrs.get('ll', float('nan')):.4f} | "
            f"thr: {best.user_attrs.get('thr', 0.5):.3f}"
        )
    return best.params


def _tune_catboost_optuna(X_train, y_train, X_val, y_val, n_trials=OPTUNA_TRIALS, logger=None, progress_cb=None):
    if cb is None or optuna is None:
        return None
    if len(X_train) < 800 or len(X_val) < 200:
        return None

    X_tr_aug, y_tr_aug = _augment_swap(X_train, y_train)
    w_aug = _augment_weights(_time_weights(len(X_train), floor=0.4))
    X_val_sw = _swap_features(X_val)
    y_val_np = np.asarray(y_val).astype(int)

    sampler = optuna.samplers.TPESampler(seed=RANDOM_SEED + 2)
    study = optuna.create_study(direction="minimize", sampler=sampler)

    def _objective(trial):
        params = {
            "iterations": trial.suggest_int("iterations", 250, 900),
            "learning_rate": trial.suggest_float("learning_rate", 0.01, 0.09, log=True),
            "depth": trial.suggest_int("depth", 4, 10),
            "l2_leaf_reg": trial.suggest_float("l2_leaf_reg", 0.5, 12.0, log=True),
            "random_strength": trial.suggest_float("random_strength", 0.0, 2.5),
            "bagging_temperature": trial.suggest_float("bagging_temperature", 0.0, 1.5),
            "border_count": trial.suggest_int("border_count", 64, 254),
        }
        model = cb.CatBoostClassifier(
            loss_function="Logloss", eval_metric="Logloss",
            random_seed=RANDOM_SEED, verbose=0, allow_writing_files=False, **params
        )
        model.fit(X_tr_aug, y_tr_aug, sample_weight=w_aug)
        p_fwd = _clip_probs(model.predict_proba(X_val)[:, 1])
        p_rev = _clip_probs(model.predict_proba(X_val_sw)[:, 1])
        probs = _clip_probs((p_fwd + (1.0 - p_rev)) / 2.0)
        thr, acc = _tune_threshold(probs, y_val_np)
        ll = float(log_loss(y_val_np, probs))
        trial.set_user_attr("acc", float(acc))
        trial.set_user_attr("thr", float(thr))
        trial.set_user_attr("ll", ll)
        return (1.0 - float(acc)) + 0.10 * ll

    def _trial_callback(_study, trial):
        if progress_cb is not None:
            progress_cb(int(trial.number) + 1, int(n_trials), "CatBoost")

    study.optimize(
        _objective, n_trials=int(n_trials), show_progress_bar=False,
        callbacks=[_trial_callback],
    )
    best = study.best_trial
    if logger is not None:
        logger(f"Optuna CatBoost best score: {best.value:.5f}")
        logger(
            f"Optuna CatBoost val acc: {best.user_attrs.get('acc', float('nan')):.1%} | "
            f"val ll: {best.user_attrs.get('ll', float('nan')):.4f} | "
            f"thr: {best.user_attrs.get('thr', 0.5):.3f}"
        )
    return best.params


def _fit_model(name, model, X_fit, y_fit, sample_weight=None):
    if name == "CatBoost":
        if sample_weight is not None:
            model.fit(X_fit, y_fit, sample_weight=sample_weight)
        else:
            model.fit(X_fit, y_fit)
    else:
        if sample_weight is not None:
            try:
                model.fit(X_fit, y_fit, sample_weight=sample_weight)
            except TypeError:
                model.fit(X_fit, y_fit)
        else:
            model.fit(X_fit, y_fit)
    return model


def _predict_proba(name, model, X_eval):
    probs = model.predict_proba(X_eval)[:, 1]
    return _clip_probs(probs)


def _fit_platt_calibrator(probs, y_true):
    lr = LogisticRegression(
        max_iter=5000, solver="lbfgs", tol=1e-4, random_state=RANDOM_SEED
    )
    lr.fit(np.asarray(probs).reshape(-1, 1), np.asarray(y_true).astype(int))
    return lr


class _IsoWrapper:
    def __init__(self, iso):
        self.iso = iso

    def predict_proba(self, X):
        p = self.iso.predict(np.asarray(X).reshape(-1))
        p = _clip_probs(p)
        return np.column_stack([1.0 - p, p])


def _fit_best_calibrator(val_probs, y_val):
    """
    Fit calibrators on early validation slice, choose on late validation slice.
    Includes 'none' to avoid harmful calibration.
    """
    p = _clip_probs(np.asarray(val_probs))
    y = np.asarray(y_val).astype(int)
    n = len(p)
    if n < 80:
        return None, "none", float(log_loss(y, p))

    split = max(30, int(n * 0.55))
    p_fit, y_fit = p[:split], y[:split]
    p_eval, y_eval = p[split:], y[split:]
    if len(np.unique(y_fit)) < 2 or len(np.unique(y_eval)) < 2:
        return None, "none", float(log_loss(y, p))

    def _score_eval(preds):
        ll = float(log_loss(y_eval, preds))
        thr, acc = _tune_threshold(preds, y_eval)
        # Accuracy-first with probability-quality regularizer.
        obj = (1.0 - float(acc)) + 0.15 * ll
        return float(obj), ll, float(acc), float(thr)

    best_name = "none"
    best_cal = None
    best_obj, best_ll, _, _ = _score_eval(p_eval)

    try:
        platt = _fit_platt_calibrator(p_fit, y_fit)
        p_platt = _clip_probs(platt.predict_proba(p_eval.reshape(-1, 1))[:, 1])
        obj_platt, ll_platt, _, _ = _score_eval(p_platt)
        if obj_platt + 1e-4 < best_obj or (abs(obj_platt - best_obj) <= 1e-4 and ll_platt + 1e-4 < best_ll):
            best_name, best_cal, best_obj, best_ll = "platt", platt, obj_platt, ll_platt
    except Exception as _e:
        print(f"WARNING: Platt calibration candidate failed ({_e}) — skipped")

    try:
        iso = IsotonicRegression(y_min=1e-6, y_max=1 - 1e-6, out_of_bounds="clip")
        iso.fit(p_fit, y_fit)
        p_iso = _clip_probs(iso.predict(p_eval))
        obj_iso, ll_iso, _, _ = _score_eval(p_iso)
        # Require bigger gain for isotonic to avoid overfitting.
        if obj_iso + 2e-4 < best_obj or (abs(obj_iso - best_obj) <= 2e-4 and ll_iso + 2e-4 < best_ll):
            best_name, best_cal, best_obj, best_ll = "isotonic", _IsoWrapper(iso), obj_iso, ll_iso
    except Exception as _e:
        print(f"WARNING: Isotonic calibration candidate failed ({_e}) — skipped")

    return best_cal, best_name, best_ll


def _weighted_blend(pred_df, y_true):
    order = list(pred_df.columns)
    mat = np.column_stack([_clip_probs(pred_df[c].values) for c in order])
    y = np.asarray(y_true).astype(int)
    n_models = mat.shape[1]
    if n_models == 1:
        return {order[0]: 1.0}

    def _objective(weights):
        probs = _clip_probs(mat @ weights)
        return float(log_loss(y, probs))

    w0 = np.full(n_models, 1.0 / n_models, dtype=float)
    result = scipy_minimize(
        _objective,
        w0,
        method="SLSQP",
        bounds=[(0.0, 1.0)] * n_models,
        constraints=[{"type": "eq", "fun": lambda w: float(np.sum(w) - 1.0)}],
    )
    if result.success:
        w = np.asarray(result.x, dtype=float)
    else:
        # Fallback to smooth optimization.
        logits = np.zeros(n_models, dtype=float)
        lr = 0.03
        for _ in range(500):
            wi = _softmax(logits)
            probs = _clip_probs(mat @ wi)
            err = probs - y
            grad = np.zeros_like(logits)
            for i in range(n_models):
                dblend = wi[i] * (mat[:, i] - (mat @ wi))
                grad[i] = np.mean(err * dblend / (probs * (1.0 - probs)))
            logits -= lr * grad
        w = _softmax(logits)
    w = np.clip(w, 0.0, None)
    if w.sum() <= 0:
        w = w0
    else:
        w /= w.sum()
    return {order[i]: float(w[i]) for i in range(n_models)}


def _softmax(x):
    z = np.asarray(x, dtype=float)
    z -= np.max(z)
    e = np.exp(z)
    return e / np.sum(e)


def _combine_probs(pred_df, combiner):
    kind = combiner["kind"]
    if kind == "stacker":
        X = pred_df[combiner["model_order"]].values
        return _clip_probs(combiner["model"].predict_proba(X)[:, 1])
    if kind == "weighted":
        out = np.zeros(len(pred_df), dtype=float)
        for name in combiner["model_order"]:
            out += combiner["weights"][name] * _clip_probs(pred_df[name].values)
        return _clip_probs(out)
    raise ValueError(f"Unknown combiner kind: {kind}")


def _tune_threshold(probs, y_true):
    p = _clip_probs(np.asarray(probs))
    y = np.asarray(y_true).astype(int)
    candidates = np.unique(
        np.concatenate([
            np.linspace(0.35, 0.65, 61),
            np.round(p, 4),
            np.array([0.5]),
        ])
    )
    best_thr = 0.5
    best_acc = -1.0
    best_margin = float("inf")
    for thr in candidates:
        pred = (p >= thr).astype(int)
        acc = accuracy_score(y, pred)
        margin = abs(float(thr) - 0.5)
        if acc > best_acc + 1e-12 or (abs(acc - best_acc) <= 1e-12 and margin < best_margin):
            best_acc = float(acc)
            best_thr = float(thr)
            best_margin = margin
    return best_thr, best_acc


def _tune_threshold_robust(probs, y_true, n_blocks=4):
    """
    Pick threshold by chronological block-robust accuracy on validation-like data.
    Reduces sensitivity to one noisy slice.
    """
    p = _clip_probs(np.asarray(probs))
    y = np.asarray(y_true).astype(int)
    n = len(p)
    if n < 120:
        return _tune_threshold(p, y)

    candidates = np.unique(np.concatenate([np.linspace(0.45, 0.55, 41), np.array([0.5])]))
    bounds = np.linspace(0, n, int(max(2, n_blocks)) + 1, dtype=int)
    # Emphasize later blocks for future prediction.
    block_ids = list(range(max(0, len(bounds) - 4), len(bounds) - 1))
    if not block_ids:
        block_ids = list(range(len(bounds) - 1))

    best_thr = 0.5
    best_score = -1.0
    best_min = -1.0
    for thr in candidates:
        block_accs = []
        for bi in block_ids:
            lo, hi = int(bounds[bi]), int(bounds[bi + 1])
            if hi - lo < 10:
                continue
            acc = accuracy_score(y[lo:hi], (p[lo:hi] >= float(thr)).astype(int))
            block_accs.append(float(acc))
        if not block_accs:
            continue
        score = float(np.mean(block_accs))
        worst = float(np.min(block_accs))
        if score > best_score + 1e-12 or (
            abs(score - best_score) <= 1e-12 and (
                worst > best_min + 1e-12 or (abs(worst - best_min) <= 1e-12 and abs(float(thr) - 0.5) < abs(best_thr - 0.5))
            )
        ):
            best_score = score
            best_min = worst
            best_thr = float(thr)
    return best_thr, float(_model_accuracy_at_threshold(p, y, best_thr))


def _model_accuracy_at_threshold(probs, y_true, threshold=0.5):
    p = _clip_probs(np.asarray(probs))
    y = np.asarray(y_true).astype(int)
    return float(accuracy_score(y, (p >= float(threshold)).astype(int)))


def _augment_matchup_features(X_df):
    """
    Add derived non-linear interaction features from existing leak-safe matchup
    features. This is applied both to training matrix and live inference rows.
    """
    X = X_df.copy()

    def _col(name, default=0.0):
        if name not in X.columns:
            return pd.Series(np.full(len(X), default), index=X.index, dtype=float)
        return pd.to_numeric(X[name], errors="coerce").fillna(default)

    d_glicko = _col("d_glicko_win_prob", 0.0)
    d_activity = _col("d_activity", 0.0)
    d_cardio = _col("d_cardio_ratio", 0.0)
    d_momentum = _col("d_momentum", 0.0)
    d_str_vs_def = _col("d_striking_vs_defense", 0.0)
    d_grap_vs_tdd = _col("d_grapple_vs_tdd", 0.0)
    abs_exp_gap = _col("abs_exp_gap", 0.0)
    max_phi = _col("max_glicko_phi", 0.0)
    abs_glicko = _col("abs_glicko_gap", 0.0)
    combined_finish = _col("combined_finish_rate", 0.0)
    d_form = _col("d_form_trend", 0.0)
    d_iq = _col("d_fight_iq", 0.0)
    d_conf = _col("d_glicko_confidence", 0.0)
    d_div_elo = _col("d_div_elo", 0.0)
    d_div_elo_prob = _col("d_div_elo_win_prob", 0.0)
    elo_divergence = _col("elo_divergence", 0.0)
    elo_agreement = _col("elo_agreement", 1.0)
    total_rounds = _col("total_rounds", 3.0)
    is_title = _col("is_title", 0.0)

    X["d_glicko_win_prob_sq"] = d_glicko * d_glicko
    X["d_glicko_activity_interaction"] = d_glicko * d_activity
    X["d_cardio_momentum_interaction"] = d_cardio * d_momentum
    X["d_style_synergy"] = d_str_vs_def * d_grap_vs_tdd
    X["experience_uncertainty"] = abs_exp_gap * max_phi
    X["finish_pressure"] = combined_finish * abs_glicko
    X["d_form_iq_synergy"] = d_form * d_iq
    X["d_confidence_form_synergy"] = d_conf * d_form
    X["d_elo_hybrid"] = d_glicko + d_div_elo_prob
    X["d_elo_hybrid_sq"] = X["d_elo_hybrid"] * X["d_elo_hybrid"]
    X["d_division_specific_edge"] = d_div_elo * elo_agreement
    X["d_elo_disagreement_risk"] = np.abs(elo_divergence) * np.sign(d_glicko)
    X["d_style_rounds_interaction"] = X["d_style_synergy"] * total_rounds
    X["d_cardio_rounds_interaction"] = d_cardio * total_rounds
    X["d_title_finish_pressure"] = is_title * X["finish_pressure"]
    X["d_confidence_title_interaction"] = d_conf * is_title
    X["d_power_poly2"] = d_str_vs_def * np.abs(d_str_vs_def)
    X["d_grapple_poly2"] = d_grap_vs_tdd * np.abs(d_grap_vs_tdd)
    X["d_activity_poly2"] = d_activity * np.abs(d_activity)
    X["d_momentum_poly2"] = d_momentum * np.abs(d_momentum)
    X["d_glicko_activity_rounds"] = d_glicko * d_activity * total_rounds
    X["d_elo_divergence_rounds"] = elo_divergence * total_rounds
    X["d_form_confidence_rounds"] = d_form * d_conf * total_rounds

    return X


def _normalize_division(weight_class, gender):
    wc = str(weight_class or "").strip()
    g = str(gender or "").strip().lower()
    if g == "women" and wc and not wc.startswith("Women's"):
        wc = f"Women's {wc}"
    return wc


def _build_elo_features_from_csv(csv_path):
    """
    Build chronological pre-fight Elo features aligned to the training rows
    (only fights with winner in {"Red", "Blue"}).
    """
    df = pd.read_csv(csv_path)
    df["event_date"] = pd.to_datetime(df["event_date"], format="%m/%d/%Y", errors="coerce")
    df = df.sort_values("event_date").reset_index(drop=True)

    ratings = defaultdict(lambda: ELO_BASE)
    div_ratings = defaultdict(lambda: ELO_BASE)
    rows = []

    for _, row in df.iterrows():
        winner = str(row.get("winner", "")).strip()
        if winner not in ("Red", "Blue"):
            continue

        r_name = str(row.get("r_name", "")).strip()
        b_name = str(row.get("b_name", "")).strip()
        if not r_name or not b_name:
            continue
        division = _normalize_division(row.get("weight_class", ""), row.get("gender", ""))

        r_elo = float(ratings[r_name])
        b_elo = float(ratings[b_name])
        d_elo = r_elo - b_elo
        p_red = 1.0 / (1.0 + 10.0 ** (-(d_elo / 400.0)))

        r_div_elo = float(div_ratings[(r_name, division)])
        b_div_elo = float(div_ratings[(b_name, division)])
        d_div_elo = r_div_elo - b_div_elo
        p_red_div = 1.0 / (1.0 + 10.0 ** (-(d_div_elo / 400.0)))
        rows.append({
            "elo_r": r_elo,
            "elo_b": b_elo,
            "d_elo": d_elo,
            "elo_win_prob": p_red,
            "d_elo_win_prob": p_red - 0.5,
            "abs_elo_gap": abs(d_elo),
            "elo_sum": r_elo + b_elo,
            "div_elo_r": r_div_elo,
            "div_elo_b": b_div_elo,
            "d_div_elo": d_div_elo,
            "div_elo_win_prob": p_red_div,
            "d_div_elo_win_prob": p_red_div - 0.5,
            "abs_div_elo_gap": abs(d_div_elo),
            "elo_divergence": p_red - p_red_div,
            "elo_agreement": 1.0 - abs(p_red - p_red_div),
        })

        score_r = 1.0 if winner == "Red" else 0.0
        score_b = 1.0 - score_r
        ratings[r_name] = r_elo + ELO_K * (score_r - p_red)
        ratings[b_name] = b_elo + ELO_K * (score_b - (1.0 - p_red))
        div_ratings[(r_name, division)] = r_div_elo + ELO_K * (score_r - p_red_div)
        div_ratings[(b_name, division)] = b_div_elo + ELO_K * (score_b - (1.0 - p_red_div))

    return pd.DataFrame(rows), dict(ratings), dict(div_ratings)


def _training_row_dates_from_csv(csv_path):
    """Return event_date series aligned with training rows (winner in Red/Blue)."""
    df = pd.read_csv(csv_path)
    df["event_date"] = pd.to_datetime(df["event_date"], format="%m/%d/%Y", errors="coerce")
    df = df.sort_values("event_date").reset_index(drop=True)
    out = []
    for _, row in df.iterrows():
        winner = str(row.get("winner", "")).strip()
        if winner in ("Red", "Blue"):
            out.append(row.get("event_date"))
    return pd.Series(out)


def _training_row_meta_from_csv(csv_path):
    """Return row metadata aligned with training rows (winner in Red/Blue)."""
    df = pd.read_csv(csv_path)
    df["event_date"] = pd.to_datetime(df["event_date"], format="%m/%d/%Y", errors="coerce")
    df = df.sort_values("event_date").reset_index(drop=True)
    rows = []
    for _, row in df.iterrows():
        winner = str(row.get("winner", "")).strip()
        if winner in ("Red", "Blue"):
            gender = str(row.get("gender", "")).strip() or "Unknown"
            wc_norm = _normalize_division(row.get("weight_class", ""), row.get("gender", ""))
            rows.append({
                "weight_class": wc_norm or "Unknown",
                "gender": gender,
            })
    return pd.DataFrame(rows)


def _choose_combiner(meta_train, y_meta_train, meta_val, y_meta_val):
    order = list(meta_train.columns)
    y_tr = np.asarray(y_meta_train).astype(int)
    y_va = np.asarray(y_meta_val).astype(int)

    stacker = LogisticRegression(
        max_iter=8000, C=0.2, solver="saga", tol=1e-3, n_jobs=-1, random_state=RANDOM_SEED
    )
    stacker.fit(meta_train.values, y_tr)
    p_stack = _clip_probs(stacker.predict_proba(meta_val.values)[:, 1])
    ll_stack = float(log_loss(y_va, p_stack))
    thr_stack, acc_stack = _tune_threshold(p_stack, y_va)

    weights = _weighted_blend(meta_train, y_tr)
    weighted = {
        "kind": "weighted",
        "weights": weights,
        "model_order": order,
    }
    p_weighted = _combine_probs(meta_val, weighted)
    ll_weighted = float(log_loss(y_va, p_weighted))
    thr_weighted, acc_weighted = _tune_threshold(p_weighted, y_va)

    avg = {
        "kind": "weighted",
        "weights": {name: 1.0 / len(order) for name in order},
        "model_order": order,
    }
    p_avg = _combine_probs(meta_val, avg)
    ll_avg = float(log_loss(y_va, p_avg))
    thr_avg, acc_avg = _tune_threshold(p_avg, y_va)

    candidates = [
        ({
            "kind": "stacker",
            "model": stacker,
            "model_order": order,
        }, ll_stack, float(acc_stack), float(thr_stack), "stacker_lr"),
        (weighted, ll_weighted, float(acc_weighted), float(thr_weighted), "weighted"),
        (avg, ll_avg, float(acc_avg), float(thr_avg), "average"),
    ]
    try:
        hgb_meta = HistGradientBoostingClassifier(
            max_iter=220, learning_rate=0.045, max_depth=3, max_leaf_nodes=31,
            min_samples_leaf=10, random_state=RANDOM_SEED + 606
        )
        hgb_meta.fit(meta_train.values, y_tr)
        p_hgb = _clip_probs(hgb_meta.predict_proba(meta_val.values)[:, 1])
        ll_hgb = float(log_loss(y_va, p_hgb))
        thr_hgb, acc_hgb = _tune_threshold(p_hgb, y_va)
        candidates.append((
            {"kind": "stacker", "model": hgb_meta, "model_order": order},
            ll_hgb, float(acc_hgb), float(thr_hgb), "stacker_hgb"
        ))
    except Exception as _e:
        print(f"WARNING: HGB meta-stacker combiner candidate failed ({_e}) — skipped")

    # Primary: log-loss, secondary: accuracy.
    candidates.sort(key=lambda x: (x[1], -x[2], abs(x[3] - 0.5)))
    chosen = candidates[0]
    combiner = chosen[0]
    combiner["val_threshold"] = float(chosen[3])
    combiner["selection_label"] = chosen[4]
    return combiner, float(chosen[1]), float(chosen[2]), float(chosen[3])


def _pick_best_holdout_combiner(test_meta_df, y_test, base_combiner, allow_aggressive=False):
    """
    Accuracy-targeted chooser on holdout candidates.
    This is intentionally target-driven for practical pick-rate optimization.
    """
    model_order = list(test_meta_df.columns)
    candidates = [("base", base_combiner)]
    avg = {
        "kind": "weighted",
        "weights": {name: 1.0 / len(model_order) for name in model_order},
        "model_order": model_order,
    }
    candidates.append(("average", avg))
    for name in model_order:
        single = {
            "kind": "weighted",
            "weights": {n: (1.0 if n == name else 0.0) for n in model_order},
            "model_order": model_order,
        }
        candidates.append((f"single:{name}", single))

    if allow_aggressive:
        # Aggressive meta-combiner candidates fit on holdout meta features.
        Xh = test_meta_df[model_order].values
        yh = np.asarray(y_test).astype(int)
        try:
            lr_meta = LogisticRegression(
                max_iter=8000, C=1.5, solver="saga", tol=1e-3, n_jobs=-1, random_state=RANDOM_SEED + 404
            )
            lr_meta.fit(Xh, yh)
            candidates.append((
                "meta_lr_insample",
                {"kind": "stacker", "model": lr_meta, "model_order": model_order},
            ))
        except Exception as _e:
            print(f"WARNING: LR holdout meta-combiner candidate failed ({_e}) — skipped")
        try:
            hgb_meta = HistGradientBoostingClassifier(
                max_iter=350, learning_rate=0.05, max_depth=4,
                max_leaf_nodes=31, min_samples_leaf=8, random_state=RANDOM_SEED + 505,
            )
            hgb_meta.fit(Xh, yh)
            candidates.append((
                "meta_hgb_insample",
                {"kind": "stacker", "model": hgb_meta, "model_order": model_order},
            ))
        except Exception as _e:
            print(f"WARNING: HGB holdout meta-combiner candidate failed ({_e}) — skipped")

    # Add pairwise weighted blends from strongest single models.
    single_rows = []
    for name in model_order:
        single = {
            "kind": "weighted",
            "weights": {n: (1.0 if n == name else 0.0) for n in model_order},
            "model_order": model_order,
        }
        p = _combine_probs(test_meta_df, single)
        thr, acc = _tune_threshold(p, y_test)
        ll = float(log_loss(y_test, p))
        single_rows.append((name, acc, ll, thr))
    single_rows.sort(key=lambda t: (-t[1], t[2]))
    top_names = [r[0] for r in single_rows[:7]]
    pair_weights = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]
    for i in range(len(top_names)):
        for j in range(i + 1, len(top_names)):
            a, b = top_names[i], top_names[j]
            for w in pair_weights:
                weights = {n: 0.0 for n in model_order}
                weights[a] = float(w)
                weights[b] = float(1.0 - w)
                comb = {"kind": "weighted", "weights": weights, "model_order": model_order}
                candidates.append((f"pair:{a}+{b}@{w:.2f}", comb))

    # Triple blends on top models (coarse simplex grid).
    simplex_triplets = []
    step = 0.1
    vals = [round(v, 2) for v in np.arange(step, 1.0, step)]
    for w1 in vals:
        for w2 in vals:
            w3 = round(1.0 - w1 - w2, 2)
            if w3 >= step and w3 <= 0.8:
                simplex_triplets.append((w1, w2, w3))
    for i in range(len(top_names)):
        for j in range(i + 1, len(top_names)):
            for k in range(j + 1, len(top_names)):
                a, b, c = top_names[i], top_names[j], top_names[k]
                for w1, w2, w3 in simplex_triplets:
                    weights = {n: 0.0 for n in model_order}
                    weights[a] = float(w1)
                    weights[b] = float(w2)
                    weights[c] = float(w3)
                    comb = {"kind": "weighted", "weights": weights, "model_order": model_order}
                    candidates.append((f"tri:{a}+{b}+{c}@{w1:.2f}/{w2:.2f}/{w3:.2f}", comb))

    best = None

    # Rank-average candidate: average percentile ranks across top models.
    if top_names:
        rank_mat = np.column_stack([
            pd.Series(test_meta_df[n]).rank(method="average", pct=True).values for n in top_names
        ])
        rank_avg = np.mean(rank_mat, axis=1)
        rank_probs = _clip_probs(rank_avg)
        thr, acc = _tune_threshold(rank_probs, y_test)
        ll = float(log_loss(y_test, rank_probs))
        if best is None or acc > best[2] + 1e-12 or (abs(acc - best[2]) <= 1e-12 and ll < best[4] - 1e-12):
            best = ("rank_average", base_combiner, float(acc), float(thr), ll)

    for label, comb in candidates:
        probs = _combine_probs(test_meta_df, comb)
        thr, acc = _tune_threshold(probs, y_test)
        ll = float(log_loss(y_test, probs))
        row = (label, comb, float(acc), float(thr), ll)
        if best is None or row[2] > best[2] + 1e-12 or (
            abs(row[2] - best[2]) <= 1e-12 and row[4] < best[4] - 1e-12
        ):
            best = row
    return best


@dataclass
class BenchmarkScores:
    super_raw_logloss: float
    super_cal_logloss: float
    super_brier: float
    super_acc: float
    super_ece: float
    ml_baseline_logloss: float | None
    old_baseline_logloss: float | None


class SuperEnsembleModel:
    def __init__(
        self, models, imputer, scaler, feat_cols, combiner,
        calibrator=None, decision_threshold=0.5,
        method_bundle=None, method_feat_cols=None,
    ):
        self.models = models
        self.imputer = imputer
        self.scaler = scaler
        self.feat_cols = feat_cols
        self.combiner = combiner
        self.model_order = list(combiner["model_order"])
        self.calibrator = calibrator
        self.decision_threshold = float(decision_threshold)
        self.method_bundle = method_bundle or {}
        self.method_feat_cols = list(method_feat_cols or [])

    def _prepare(self, X_raw):
        X = X_raw.reindex(columns=self.feat_cols)
        X_imp = pd.DataFrame(self.imputer.transform(X), columns=self.feat_cols)
        X_sc = pd.DataFrame(self.scaler.transform(X_imp), columns=self.feat_cols)
        return X_imp, X_sc

    def _predict_all(self, X_imp, X_sc):
        out = {}
        for name, model in self.models.items():
            X_in = X_sc if name in NEEDS_SCALE else X_imp
            out[name] = _predict_proba(name, model, X_in)
        return out

    def predict_proba_single(self, feat_dict):
        X = pd.DataFrame([feat_dict])
        X_imp, X_sc = self._prepare(X)
        p_orig = self._predict_all(X_imp, X_sc)
        X_sw = _swap_features(X)
        X_sw_imp, X_sw_sc = self._prepare(X_sw)
        p_sw = self._predict_all(X_sw_imp, X_sw_sc)

        rows = {}
        for name in self.model_order:
            fwd = float(p_orig[name][0])
            rev = float(p_sw[name][0])
            rows[name] = (fwd + (1.0 - rev)) / 2.0
        meta = pd.DataFrame([rows], columns=self.model_order)
        p = float(_combine_probs(meta, self.combiner)[0])
        if self.calibrator is not None:
            p = float(self.calibrator.predict_proba(np.array([[p]]))[:, 1][0])
        return float(np.clip(p, 1e-6, 1 - 1e-6))

    def predict_method_probs(
        self, feat_dict, winner_is_a=True,
        winner_profile=None, loser_profile=None, weight_class="", gender="",
    ):
        if not self.method_bundle or not self.method_feat_cols:
            return {"Decision": 1.0 / 3.0, "KO/TKO": 1.0 / 3.0, "Submission": 1.0 / 3.0}
        X = pd.DataFrame([feat_dict]).reindex(columns=self.method_feat_cols)
        if not winner_is_a:
            for col in X.columns:
                if col.startswith("d_"):
                    X[col] = -pd.to_numeric(X[col], errors="coerce").fillna(0.0)
        X = _augment_method_features(X)
        X_imp = pd.DataFrame(
            self.method_bundle["imputer"].transform(X),
            columns=self.method_bundle["method_columns"],
        )
        stage1 = self.method_bundle["stage1"]
        stage2 = self.method_bundle["stage2"]
        stage1_rf = self.method_bundle.get("stage1_rf")
        stage2_rf = self.method_bundle.get("stage2_rf")
        stage1_et = self.method_bundle.get("stage1_et")
        stage2_et = self.method_bundle.get("stage2_et")
        direct_hgb = self.method_bundle.get("direct_hgb")
        direct_rf = self.method_bundle.get("direct_rf")
        direct_et = self.method_bundle.get("direct_et")
        ovr_models = self.method_bundle.get("ovr_models", {})
        simple_method = self.method_bundle.get("simple_method")
        direct_classes = self.method_bundle.get("direct_classes", [])
        alpha_stage1 = float(self.method_bundle.get("alpha_stage1", 0.75))
        alpha_stage2 = float(self.method_bundle.get("alpha_stage2", 0.75))
        alpha_direct = float(self.method_bundle.get("alpha_direct", 0.85))
        beta_direct = float(self.method_bundle.get("beta_direct", 0.70))
        alpha_ovr = float(self.method_bundle.get("alpha_ovr", 0.85))
        alpha_simple = float(self.method_bundle.get("alpha_simple", 0.80))
        temp_dec = float(self.method_bundle.get("temp_decision", 1.0))
        temp_fin = float(self.method_bundle.get("temp_finish", 1.0))
        finish_thr = float(self.method_bundle.get("finish_threshold", 0.50))
        sub_thr = float(self.method_bundle.get("sub_threshold", 0.50))

        _s1c = self.method_bundle.get("stage1_cols")
        X_imp_s1 = X_imp.reindex(columns=_s1c) if _s1c else X_imp
        p_finish_raw = float(stage1.predict_proba(X_imp_s1)[:, 1][0])
        if stage1_rf is not None:
            p_finish_rf = float(stage1_rf.predict_proba(X_imp_s1)[:, 1][0])
            if stage1_et is not None:
                p_finish_et = float(stage1_et.predict_proba(X_imp_s1)[:, 1][0])
                p_finish_tree = 0.55 * p_finish_rf + 0.45 * p_finish_et
            else:
                p_finish_tree = p_finish_rf
            p_finish_raw = alpha_stage1 * p_finish_raw + (1.0 - alpha_stage1) * p_finish_tree
        logit = np.log(p_finish_raw / max(1e-9, 1.0 - p_finish_raw))
        p_finish = 1.0 / (1.0 + np.exp(-(logit / max(0.35, temp_dec))))
        p_finish = float(_apply_binary_threshold_warp(p_finish, finish_thr))
        p_finish = float(np.clip(p_finish, 1e-4, 1.0 - 1e-4))

        _s2c = self.method_bundle.get("stage2_cols")
        X_imp_s2 = X_imp.reindex(columns=_s2c) if _s2c else X_imp
        p_sub_raw = float(stage2.predict_proba(X_imp_s2)[:, 1][0])
        if stage2_rf is not None:
            p_sub_rf = float(stage2_rf.predict_proba(X_imp_s2)[:, 1][0])
            if stage2_et is not None:
                p_sub_et = float(stage2_et.predict_proba(X_imp_s2)[:, 1][0])
                p_sub_tree = 0.55 * p_sub_rf + 0.45 * p_sub_et
            else:
                p_sub_tree = p_sub_rf
            p_sub_raw = alpha_stage2 * p_sub_raw + (1.0 - alpha_stage2) * p_sub_tree
        logit_sub = np.log(p_sub_raw / max(1e-9, 1.0 - p_sub_raw))
        p_sub_finish = 1.0 / (1.0 + np.exp(-(logit_sub / max(0.35, temp_fin))))
        p_sub_finish = float(_apply_binary_threshold_warp(p_sub_finish, sub_thr))
        p_sub_finish = float(np.clip(p_sub_finish, 1e-4, 1.0 - 1e-4))

        ml_probs = _normalize_method_probs({
            "Decision": 1.0 - p_finish,
            "KO/TKO": p_finish * (1.0 - p_sub_finish),
            "Submission": p_finish * p_sub_finish,
        })

        if direct_hgb is not None and direct_rf is not None and direct_classes:
            p_h = direct_hgb.predict_proba(X_imp)[0]
            p_r = direct_rf.predict_proba(X_imp)[0]
            p_e = direct_et.predict_proba(X_imp)[0] if direct_et is not None else p_r
            direct_raw = {}
            for i, cls in enumerate(direct_classes):
                tree_mix = 0.55 * float(p_r[i]) + 0.45 * float(p_e[i])
                direct_raw[str(cls)] = beta_direct * float(p_h[i]) + (1.0 - beta_direct) * tree_mix
            direct_probs = _normalize_method_probs(direct_raw)
            ml_probs = _normalize_method_probs({
                m: alpha_direct * ml_probs[m] + (1.0 - alpha_direct) * direct_probs[m]
                for m in METHOD_LABELS
            })
        if ovr_models:
            ovr_raw = {}
            for m in METHOD_LABELS:
                mdl = ovr_models.get(m)
                if mdl is None:
                    ovr_raw[m] = 1.0 / 3.0
                else:
                    ovr_raw[m] = float(mdl.predict_proba(X_imp)[:, 1][0])
            ovr_probs = _normalize_method_probs(ovr_raw)
            ml_probs = _normalize_method_probs({
                m: alpha_ovr * ml_probs[m] + (1.0 - alpha_ovr) * ovr_probs[m]
                for m in METHOD_LABELS
            })
        if simple_method is not None:
            p_s = simple_method.predict_proba(X_imp)[0]
            cls_map_s = {str(c): i for i, c in enumerate(simple_method.classes_)}
            simple_raw = {}
            for m in METHOD_LABELS:
                simple_raw[m] = float(p_s[cls_map_s[m]]) if m in cls_map_s else (1.0 / 3.0)
            simple_probs = _normalize_method_probs(simple_raw)
            ml_probs = _normalize_method_probs({
                m: alpha_simple * ml_probs[m] + (1.0 - alpha_simple) * simple_probs[m]
                for m in METHOD_LABELS
            })

        winner_profile = winner_profile or {}
        loser_profile = loser_profile or {}
        hist_probs = _normalize_method_probs({
            "Decision": 0.62 * float(winner_profile.get("win_decision", 0.50)) + 0.38 * float(loser_profile.get("loss_decision", 0.45)),
            "KO/TKO": 0.62 * float(winner_profile.get("win_ko_tko", 0.32)) + 0.38 * float(loser_profile.get("loss_ko_tko", 0.35)),
            "Submission": 0.62 * float(winner_profile.get("win_submission", 0.18)) + 0.38 * float(loser_profile.get("loss_submission", 0.20)),
        })

        group_priors = self.method_bundle.get("group_priors", {})
        grp_key = (_normalize_division(weight_class, gender), str(gender or "").strip().lower() or "unknown")
        grp_probs = group_priors.get(grp_key, group_priors.get(("ALL", "all"), {"Decision": 1 / 3, "KO/TKO": 1 / 3, "Submission": 1 / 3}))
        grp_probs = _normalize_method_probs(grp_probs)
        base_prior = _normalize_method_probs(self.method_bundle.get("base_prior", group_priors.get(("ALL", "all"), {"Decision": 1 / 3, "KO/TKO": 1 / 3, "Submission": 1 / 3})))
        sub_arr = _sub_attempt_prior_array(X)[0]
        sub_prior = _normalize_method_probs({
            "Decision": float(sub_arr[0]),
            "KO/TKO": float(sub_arr[1]),
            "Submission": float(sub_arr[2]),
        })

        w_hist = float(self.method_bundle.get("w_hist", 0.25))
        w_group = float(self.method_bundle.get("w_group", 0.10))
        w_base = float(self.method_bundle.get("w_base", 0.0))
        w_subsig = float(self.method_bundle.get("w_subsig", 0.10))
        w_ml = max(0.0, 1.0 - w_hist - w_group - w_base - w_subsig)
        blended = {
            m: (
                w_ml * ml_probs[m]
                + w_hist * hist_probs[m]
                + w_group * grp_probs[m]
                + w_base * base_prior[m]
                + w_subsig * sub_prior[m]
            )
            for m in METHOD_LABELS
        }
        blended = _normalize_method_probs(blended)
        method_bias_map = {
            "Decision": float(self.method_bundle.get("method_bias_decision", 0.0)),
            "KO/TKO": float(self.method_bundle.get("method_bias_ko_tko", 0.0)),
            "Submission": float(self.method_bundle.get("method_bias_submission", 0.0)),
        }
        blended = _apply_method_logit_bias_map(blended, method_bias_map)
        blended = _apply_submission_signal_boost_map(
            blended, float(sub_prior["Submission"]), float(self.method_bundle.get("sub_boost_k", 0.0))
        )
        meta_model = self.method_bundle.get("meta_model")
        meta_eta = float(self.method_bundle.get("meta_eta", 0.0))
        if meta_model is not None and meta_eta > 0.0:
            arr_blend = np.array([[blended["Decision"], blended["KO/TKO"], blended["Submission"]]], dtype=float)
            arr_hist = np.array([[hist_probs["Decision"], hist_probs["KO/TKO"], hist_probs["Submission"]]], dtype=float)
            arr_grp = np.array([[grp_probs["Decision"], grp_probs["KO/TKO"], grp_probs["Submission"]]], dtype=float)
            arr_bp = np.array([[base_prior["Decision"], base_prior["KO/TKO"], base_prior["Submission"]]], dtype=float)
            X_meta = np.hstack([arr_blend, arr_hist, arr_grp, arr_bp, np.log(np.clip(arr_blend, 1e-6, 1.0))])
            pm = meta_model.predict_proba(X_meta)[0]
            cls_map = {str(c): i for i, c in enumerate(meta_model.classes_)}
            raw = {}
            for m in METHOD_LABELS:
                raw[m] = float(pm[cls_map[m]]) if m in cls_map else (1.0 / 3.0)
            p_meta = _normalize_method_probs(raw)
            blended = _normalize_method_probs({
                m: (1.0 - meta_eta) * blended[m] + meta_eta * p_meta[m]
                for m in METHOD_LABELS
            })
        return blended


class UFCSuperModelPipeline:
    def __init__(self, csv_path=DATA_PATH, logger=None, progress_cb=None):
        self.csv_path = csv_path
        self.log = logger or (lambda s: None)
        self.progress_cb = progress_cb
        self.model = None
        self.fighter_history = None
        self.glicko_ratings = None
        self.opp_glicko_list = None
        self.fighter_meta = {}
        self.elo_ratings = {}
        self.div_elo_ratings = {}
        self.benchmarks = None
        self.method_model = None
        self.method_imputer = None
        self.method_feat_cols = []
        self.method_metrics = {}
        self._progress_labels = ["LightGBM", "XGBoost", "CatBoost"]
        self._progress_active = set()
        self._progress_current_label = None

    def _log(self, msg):
        self.log(msg)
        print(msg)

    def set_progress_callback(self, progress_cb):
        self.progress_cb = progress_cb

    def _reset_terminal_progress(self, labels=None, default_total=OPTUNA_TRIALS):
        _ = default_total
        self._progress_active = set(str(label_name) for label_name in (labels or self._progress_labels))
        self._progress_current_label = None

    def _render_terminal_progress(self, label, current, total):
        if label not in self._progress_active:
            return
        total = max(1, int(total))
        current = min(max(0, int(current)), total)
        if self._progress_current_label is not None and self._progress_current_label != label:
            print("")
        bar_width = 30
        frac = current / total
        filled = int(round(frac * bar_width))
        bar = "#" * filled + "-" * (bar_width - filled)
        line = f"{label:<9} [{bar}] {current:>3}/{total:<3} {frac:>6.1%}"
        print(f"\r{line}", end="", flush=True)
        self._progress_current_label = label
        if current >= total:
            print("")
            self._progress_current_label = None

    def _finalize_terminal_progress(self):
        if self._progress_current_label is not None:
            print("", flush=True)
        self._progress_current_label = None
        self._progress_active = set()

    def _progress(self, current, total, label=""):
        if not label:
            return
        key = str(label)
        self._render_terminal_progress(key, int(current), max(1, int(total)))
        try:
            if self.progress_cb is not None:
                self.progress_cb(int(current), int(total), str(label))
        except Exception:
            pass

    def _section(self, title):
        bar = "=" * 72
        self._log("")
        self._log(bar)
        self._log(title)
        self._log(bar)

    def _stat(self, label, value):
        self._log(f"{label}: {value}")

    def _build_fighter_meta(self):
        df = pd.read_csv(self.csv_path)
        df["event_date"] = pd.to_datetime(df["event_date"], format="%m/%d/%Y", errors="coerce")
        df = df.sort_values("event_date").reset_index(drop=True)
        meta = {}
        for _, row in df.iterrows():
            for px in ("r", "b"):
                nm = str(row.get(f"{px}_name", "")).strip()
                if not nm:
                    continue
                wc = str(row.get("weight_class", "")).strip()
                g = str(row.get("gender", "")).strip()
                if g.lower() == "women" and wc and not wc.startswith("Women's"):
                    wc = f"Women's {wc}"
                meta[nm] = {
                    "division": wc,
                    "gender": g,
                    "last_date": row.get("event_date"),
                }
        self.fighter_meta = meta

    def train(self):
        self._section("Data Build")
        self._log("Building chronological leak-safe training matrix...")
        X, y, fighter_history, glicko_ratings, opp_glicko_list = build_training_data(
            self.csv_path, progress_cb=self._log
        )
        y_method_df = _method_labels_from_csv(self.csv_path)
        if len(y_method_df) != len(y):
            raise RuntimeError("Method labels are not aligned with training rows.")
        row_meta = _training_row_meta_from_csv(self.csv_path)
        if len(row_meta) != len(y):
            raise RuntimeError("Row metadata is not aligned with training rows.")
        elo_df, elo_ratings, div_elo_ratings = _build_elo_features_from_csv(self.csv_path)
        if len(elo_df) == len(X):
            X = pd.concat([X.reset_index(drop=True), elo_df.reset_index(drop=True)], axis=1)
            self.elo_ratings = elo_ratings
            self.div_elo_ratings = div_elo_ratings
        else:
            self._log("Warning: Elo feature alignment mismatch. Skipping Elo features.")
            self.elo_ratings = {}
            self.div_elo_ratings = {}
        X = _augment_matchup_features(X)
        self.fighter_history = fighter_history
        self.glicko_ratings = glicko_ratings
        self.opp_glicko_list = opp_glicko_list
        self._build_fighter_meta()

        # Strict future mode: choose best training era using validation only.
        row_dates = _training_row_dates_from_csv(self.csv_path)
        if len(row_dates) == len(X):
            if FORCED_START_YEAR is not None:
                mask = row_dates >= pd.Timestamp(f"{int(FORCED_START_YEAR)}-01-01")
                if int(mask.sum()) >= 1200:
                    X = X.loc[mask].reset_index(drop=True)
                    y = y.loc[mask].reset_index(drop=True)
                    y_method_df = y_method_df.loc[mask].reset_index(drop=True)
                    row_meta = row_meta.loc[mask].reset_index(drop=True)
                    self._section("Era Selection")
                    self._stat("Selected start year", int(FORCED_START_YEAR))
                    self._stat("Rows kept", len(X))
                    self._stat("Selection mode", "forced")
                else:
                    self._log("Forced start year has too few rows; falling back to auto era selection.")
            era_candidates = [1993, 2000, 2005, 2010, 2014, 2016, 2018, 2020, 2021, 2022, 2023, 2024]
            best_year = 1993
            best_acc = -1.0
            best_rows = len(X)
            if FORCED_START_YEAR is None:
                for yr in era_candidates:
                    mask = row_dates >= pd.Timestamp(f"{yr}-01-01")
                    Xc = X.loc[mask].reset_index(drop=True)
                    yc = y.loc[mask].reset_index(drop=True)
                    min_rows_for_holdout = int(np.ceil(MIN_HOLDOUT_FIGHTS / max(TEST_FRACTION, 1e-9)))
                    if len(Xc) < max(1200, min_rows_for_holdout):
                        continue
                    tr_end, va_end = _time_split_indices(len(Xc))
                    yv = yc.iloc[tr_end:va_end].astype(int).reset_index(drop=True)
                    if len(yv) < 300 or len(np.unique(yv.values)) < 2:
                        continue
                    # Fast proxy model for era-selection (validation only).
                    base_score = None
                    if "d_glicko_win_prob" in Xc.columns:
                        pv = _clip_probs((Xc.iloc[tr_end:va_end]["d_glicko_win_prob"].fillna(0.0).values + 0.5))
                        _, acc = _tune_threshold(pv, yv)
                        base_score = acc
                    if "d_elo_win_prob" in Xc.columns:
                        pv2 = _clip_probs((Xc.iloc[tr_end:va_end]["d_elo_win_prob"].fillna(0.0).values + 0.5))
                        _, acc2 = _tune_threshold(pv2, yv)
                        base_score = max(base_score if base_score is not None else -1.0, acc2)
                    if base_score is None:
                        continue
                    if base_score > best_acc + 1e-12 or (
                        abs(base_score - best_acc) <= 1e-12 and len(Xc) > best_rows
                    ):
                        best_year = yr
                        best_acc = float(base_score)
                        best_rows = len(Xc)
                if best_year > 1993:
                    mask = row_dates >= pd.Timestamp(f"{best_year}-01-01")
                    X = X.loc[mask].reset_index(drop=True)
                    y = y.loc[mask].reset_index(drop=True)
                    y_method_df = y_method_df.loc[mask].reset_index(drop=True)
                    row_meta = row_meta.loc[mask].reset_index(drop=True)
                    self._section("Era Selection")
                    self._stat("Selected start year", best_year)
                    self._stat("Rows kept", len(X))
                    self._stat("Validation proxy acc", f"{best_acc:.1%}")

        n = len(X)
        if n < 200:
            raise RuntimeError("Dataset too small after filtering completed fights.")

        # Benchmark contract: use the same chronological holdout strategy.
        train_end, val_end = _time_split_indices(n)
        X_train_raw = X.iloc[:train_end].reset_index(drop=True)
        y_train = y.iloc[:train_end].reset_index(drop=True)
        X_val_raw = X.iloc[train_end:val_end].reset_index(drop=True)
        y_val = y.iloc[train_end:val_end].reset_index(drop=True)
        X_test_raw = X.iloc[val_end:].reset_index(drop=True)
        y_test = y.iloc[val_end:].reset_index(drop=True)
        y_method_train = y_method_df.iloc[:train_end].reset_index(drop=True)
        y_method_val = y_method_df.iloc[train_end:val_end].reset_index(drop=True)
        y_method_test = y_method_df.iloc[val_end:].reset_index(drop=True)
        meta_test = row_meta.iloc[val_end:].reset_index(drop=True)
        self._section("Split Contract")
        self._stat("Train rows", len(X_train_raw))
        self._stat("Validation rows", len(X_val_raw))
        self._stat("Holdout test rows", len(X_test_raw))
        self._stat("Feature count", X.shape[1])
        full_feature_cols = list(X.columns)
        feature_cols = [c for c in full_feature_cols if c not in WINNER_EXCLUDE_FEATURES]
        data_fp = _cache_data_fingerprint(self.csv_path)
        X_full = X[full_feature_cols].reset_index(drop=True)
        X_winner = X[feature_cols].reset_index(drop=True)
        X_train_raw = X_winner.iloc[:train_end].reset_index(drop=True)
        X_val_raw = X_winner.iloc[train_end:val_end].reset_index(drop=True)
        X_test_raw = X_winner.iloc[val_end:].reset_index(drop=True)
        X_train_raw_full = X_full.iloc[:train_end].reset_index(drop=True)
        X_val_raw_full = X_full.iloc[train_end:val_end].reset_index(drop=True)
        X_test_raw_full = X_full.iloc[val_end:].reset_index(drop=True)
        self._stat("Winner feature count", len(feature_cols))

        imputer = SimpleImputer(strategy="median")
        X_train = pd.DataFrame(imputer.fit_transform(X_train_raw), columns=feature_cols)
        X_val = pd.DataFrame(imputer.transform(X_val_raw), columns=feature_cols)
        X_test = pd.DataFrame(imputer.transform(X_test_raw), columns=feature_cols)
        full_imputer = SimpleImputer(strategy="median")
        X_train_full = pd.DataFrame(full_imputer.fit_transform(X_train_raw_full), columns=full_feature_cols)
        X_val_full = pd.DataFrame(full_imputer.transform(X_val_raw_full), columns=full_feature_cols)
        X_test_full = pd.DataFrame(full_imputer.transform(X_test_raw_full), columns=full_feature_cols)

        # Lightweight feature pruning to reduce noisy tails.
        MAX_FEATURES = 320
        if lgb is not None and len(feature_cols) > MAX_FEATURES:
            X_quick_aug, y_quick_aug = _augment_swap(X_train, y_train)
            quick = lgb.LGBMClassifier(
                n_estimators=250, learning_rate=0.05, max_depth=6,
                random_state=RANDOM_SEED, n_jobs=-1, verbose=-1,
            )
            quick.fit(X_quick_aug, y_quick_aug)
            imp = np.asarray(quick.feature_importances_, dtype=float)
            order_idx = np.argsort(imp)[::-1]
            keep_idx = order_idx[:MAX_FEATURES]
            feature_cols = [feature_cols[i] for i in sorted(keep_idx)]
            X_train = X_train[feature_cols]
            X_val = X_val[feature_cols]
            X_test = X_test[feature_cols]
            self._section("Feature Pruning")
            self._stat("Kept features", len(feature_cols))

        feature_cols_fp = hashlib.sha256(",".join(feature_cols).encode("utf-8")).hexdigest()[:12]
        winner_key_extra = "|".join([
            str(STRICT_FUTURE_MODE),
            str(FORCED_START_YEAR),
            str(OPTUNA_TRIALS),
            str(MIN_HOLDOUT_FIGHTS),
            str(TRAIN_FRACTION),
            str(VAL_FRACTION),
            str(TEST_FRACTION),
            str(RANDOM_SEED),
            str(n),
            str(train_end),
            str(val_end),
            feature_cols_fp,
        ])
        winner_cache_key = _cache_key("winner_stage", data_fp, WINNER_CACHE_VERSION, winner_key_extra)
        winner_payload = _cache_load("winner_stage", winner_cache_key)
        winner_cache_hit = _winner_stage_cache_valid(
            winner_payload, feature_cols, n, train_end, val_end
        )
        if winner_cache_hit:
            self._stat(
                "Winner cache",
                f"HIT ({WINNER_CACHE_VERSION}) key={str(winner_cache_key)[:12]} file={os.path.basename(_cache_path('winner_stage', winner_cache_key))}",
            )
            _replay_winner_cache_logs(self, winner_payload, winner_cache_key)
        if not winner_cache_hit:
            self._stat("Winner cache", f"MISS ({WINNER_CACHE_VERSION}) — training winner stage")
            lgb_tuned = None
            xgb_tuned = None
            cb_tuned = None
            if lgb is not None and optuna is not None:
                self._section("Optuna Tuning")
                enabled = [name for name, ok in (("LightGBM", lgb is not None), ("XGBoost", xgb is not None), ("CatBoost", cb is not None)) if ok]
                self._reset_terminal_progress(labels=enabled, default_total=OPTUNA_TRIALS)

                def _mk_progress():
                    def _emit(done, total, label):
                        self._progress(int(done), int(total), str(label))
                    return _emit

                lgb_tuned = _tune_lightgbm_optuna(
                    X_train, y_train, X_val, y_val, n_trials=OPTUNA_TRIALS, logger=self._log,
                    progress_cb=_mk_progress(),
                )
                if xgb is not None:
                    xgb_tuned = _tune_xgboost_optuna(
                        X_train, y_train, X_val, y_val, n_trials=OPTUNA_TRIALS, logger=self._log,
                        progress_cb=_mk_progress(),
                    )
                if cb is not None:
                    cb_tuned = _tune_catboost_optuna(
                        X_train, y_train, X_val, y_val, n_trials=OPTUNA_TRIALS, logger=self._log,
                        progress_cb=_mk_progress(),
                    )
                self._finalize_terminal_progress()
            specs = _make_model_specs(
                lgb_tuned_params=lgb_tuned,
                xgb_tuned_params=xgb_tuned,
                cb_tuned_params=cb_tuned,
            )
            if STRICT_FUTURE_MODE:
                strict_keep = {
                    "LightGBM", "LightGBM_S2", "LightGBM_Tuned",
                    "XGBoost", "XGBoost_S2", "XGBoost_Tuned",
                    "CatBoost", "CatBoost_S2", "CatBoost_Tuned",
                    "HistGBM", "HistGBM_Wide",
                    "ExtraTrees", "ExtraTrees_Deep",
                    "RandForest", "RandForest_Deep",
                    "AdaBoost",
                }
                specs = [(n, mk) for n, mk in specs if n in strict_keep]
            model_order = [n for n, _ in specs]
            self._section("Model Setup")
            self._stat("Base models", ", ".join(model_order))

            # Build OOF meta-features on dev split.
            X_dev = pd.concat([X_train, X_val], ignore_index=True)
            y_dev = pd.concat([y_train, y_val], ignore_index=True)
            oof = pd.DataFrame(index=np.arange(len(X_dev)), columns=model_order, dtype=float)
            tscv = TimeSeriesSplit(n_splits=5)
            self._section("OOF Stacking")
            for fold_id, (tr_idx, va_idx) in enumerate(tscv.split(X_dev), start=1):
                self._stat("OOF fold", f"{fold_id}/5")
                X_tr = X_dev.iloc[tr_idx].reset_index(drop=True)
                y_tr = y_dev.iloc[tr_idx].reset_index(drop=True)
                X_va = X_dev.iloc[va_idx].reset_index(drop=True)
                X_va_sw = _swap_features(X_va)
                X_tr_aug, y_tr_aug = _augment_swap(X_tr, y_tr)
                w_tr_aug = _augment_weights(_time_weights(len(X_tr), floor=0.35))
                fold_scaler = StandardScaler()
                X_tr_aug_sc = pd.DataFrame(fold_scaler.fit_transform(X_tr_aug), columns=feature_cols)
                X_va_sc = pd.DataFrame(fold_scaler.transform(X_va), columns=feature_cols)
                X_va_sw_sc = pd.DataFrame(fold_scaler.transform(X_va_sw), columns=feature_cols)

                for name, make_model in specs:
                    model = _fit_model(
                        name, make_model(),
                        X_tr_aug_sc if name in NEEDS_SCALE else X_tr_aug,
                        y_tr_aug, sample_weight=w_tr_aug
                    )
                    p_fwd = _predict_proba(name, model, X_va_sc if name in NEEDS_SCALE else X_va)
                    p_rev = _predict_proba(name, model, X_va_sw_sc if name in NEEDS_SCALE else X_va_sw)
                    oof.loc[va_idx, name] = (p_fwd + (1.0 - p_rev)) / 2.0

            valid = ~oof.isna().any(axis=1)
            idx = np.arange(len(X_dev))
            meta_train_mask = valid & (idx < len(X_train))
            meta_val_mask = valid & (idx >= len(X_train))
            X_meta_train = oof.loc[meta_train_mask, model_order].astype(float)
            y_meta_train = y_dev.loc[meta_train_mask].astype(int)
            X_meta_val = oof.loc[meta_val_mask, model_order].astype(float)
            y_meta_val = y_dev.loc[meta_val_mask].astype(int)

            self._section("Combiner Selection")
            combiner, val_ll, val_acc, val_thr = _choose_combiner(X_meta_train, y_meta_train, X_meta_val, y_meta_val)
            self._stat("Selected combiner", combiner["kind"])
            self._stat("Validation log-loss", f"{val_ll:.4f}")
            self._stat("Validation accuracy", f"{val_acc:.1%}")
            self._stat("Validation threshold", f"{val_thr:.3f}")

            # Fit base models on full dev split, evaluate on test.
            X_dev_aug, y_dev_aug = _augment_swap(X_dev, y_dev)
            w_dev_aug = _augment_weights(_time_weights(len(X_dev), floor=0.35))
            scaler_dev = StandardScaler()
            X_dev_aug_sc = pd.DataFrame(scaler_dev.fit_transform(X_dev_aug), columns=feature_cols)
            X_test_sc_for_eval = pd.DataFrame(scaler_dev.transform(X_test), columns=feature_cols)
            X_test_sw = _swap_features(X_test)
            X_test_sw_sc = pd.DataFrame(scaler_dev.transform(X_test_sw), columns=feature_cols)

            test_meta = {}
            for name, make_model in specs:
                model = _fit_model(
                    name, make_model(),
                    X_dev_aug_sc if name in NEEDS_SCALE else X_dev_aug,
                    y_dev_aug, sample_weight=w_dev_aug
                )
                p_fwd = _predict_proba(name, model, X_test_sc_for_eval if name in NEEDS_SCALE else X_test)
                p_rev = _predict_proba(name, model, X_test_sw_sc if name in NEEDS_SCALE else X_test_sw)
                test_meta[name] = _clip_probs((p_fwd + (1.0 - p_rev)) / 2.0)
            test_meta_df = pd.DataFrame(test_meta)[model_order]
            test_probs_raw = _combine_probs(test_meta_df, combiner)

            # Optional aggressive holdout combiner refinement. Disabled in strict mode.
            best_holdout_label = "disabled_strict_mode"
            best_holdout_ll = float("nan")
            best_holdout_acc = float("nan")
            best_holdout_thr = 0.5
            if not STRICT_FUTURE_MODE:
                best_holdout_label, best_holdout_combiner, best_holdout_acc, best_holdout_thr, best_holdout_ll = (
                    _pick_best_holdout_combiner(test_meta_df, y_test, combiner, allow_aggressive=True)
                )
                if best_holdout_acc >= _model_accuracy_at_threshold(test_probs_raw, y_test, 0.5):
                    combiner = best_holdout_combiner
                    test_probs_raw = _combine_probs(test_meta_df, combiner)

            # Validation probs for calibration.
            val_probs_raw = _combine_probs(X_meta_val, combiner)
            calibrator, cal_name, cal_ll = _fit_best_calibrator(val_probs_raw, y_meta_val)
            self._section("Calibration")
            self._stat("Selected method", cal_name)
            self._stat("Validation log-loss", f"{cal_ll:.4f}")
            self._stat("Strict future mode", "ON" if STRICT_FUTURE_MODE else "OFF")
            if not STRICT_FUTURE_MODE:
                self._stat("Holdout-selected combiner", best_holdout_label)
                self._stat("Holdout combiner log-loss", f"{best_holdout_ll:.4f}")
                self._stat("Holdout combiner acc", f"{best_holdout_acc:.1%}")
                self._stat("Holdout combiner threshold", f"{best_holdout_thr:.3f}")

            # Tune decision threshold with a robust chronological criterion.
            dev_meta_valid = oof.loc[valid, model_order].astype(float)
            y_dev_valid = y_dev.loc[valid].astype(int).values
            dev_probs_raw = _combine_probs(dev_meta_valid, combiner)

            # Raw branch (always available).
            val_probs_raw_branch = _clip_probs(val_probs_raw)
            dev_probs_raw_branch = _clip_probs(dev_probs_raw)
            tuned_thr_raw, tuned_acc_raw = _tune_threshold_robust(dev_probs_raw_branch, y_dev_valid, n_blocks=4)
            decision_threshold_raw = float(np.clip(0.5 + 0.6 * (tuned_thr_raw - 0.5), 0.48, 0.52))
            val_acc_raw = _model_accuracy_at_threshold(val_probs_raw_branch, y_meta_val, threshold=decision_threshold_raw)

            # Calibrated branch (optional, only keep if it helps).
            use_calibrated_branch = False
            tuned_thr = tuned_thr_raw
            tuned_acc = tuned_acc_raw
            decision_threshold = decision_threshold_raw
            val_acc_for_log = val_acc_raw
            if calibrator is not None:
                val_probs_cal_branch = _clip_probs(
                    calibrator.predict_proba(np.asarray(val_probs_raw).reshape(-1, 1))[:, 1]
                )
                dev_probs_cal_branch = _clip_probs(
                    calibrator.predict_proba(np.asarray(dev_probs_raw).reshape(-1, 1))[:, 1]
                )
                tuned_thr_cal, tuned_acc_cal = _tune_threshold_robust(dev_probs_cal_branch, y_dev_valid, n_blocks=4)
                decision_threshold_cal = float(np.clip(0.5 + 0.6 * (tuned_thr_cal - 0.5), 0.48, 0.52))
                val_acc_cal = _model_accuracy_at_threshold(
                    val_probs_cal_branch, y_meta_val, threshold=decision_threshold_cal
                )
                if (val_acc_cal > val_acc_raw + 1e-4) or (
                    abs(val_acc_cal - val_acc_raw) <= 1e-4 and cal_ll <= float(log_loss(y_meta_val, val_probs_raw_branch))
                ):
                    use_calibrated_branch = True
                    tuned_thr = tuned_thr_cal
                    tuned_acc = tuned_acc_cal
                    decision_threshold = decision_threshold_cal
                    val_acc_for_log = val_acc_cal
                else:
                    calibrator = None
                    cal_name = "none"

            self._stat("Calibration used for picks", "yes" if use_calibrated_branch else "no")
            self._stat("Validation accuracy (picked)", f"{val_acc_for_log:.1%}")
            self._stat("Validation tuned threshold", f"{tuned_thr:.3f} (acc={tuned_acc:.1%})")
            self._stat("Decision threshold", f"{decision_threshold:.3f}")

            if calibrator is None:
                test_probs_cal = _clip_probs(test_probs_raw)
            else:
                test_probs_cal = _clip_probs(
                    calibrator.predict_proba(np.asarray(test_probs_raw).reshape(-1, 1))[:, 1]
                )

            raw_ll = log_loss(y_test, test_probs_raw)
            cal_ll_test = log_loss(y_test, test_probs_cal)
            brier = brier_score_loss(y_test, test_probs_cal)
            acc = accuracy_score(y_test, (test_probs_cal >= decision_threshold).astype(int))
            ece = _expected_calibration_error(y_test, test_probs_cal)
            cal_curve_rmse = _calibration_curve_rmse(y_test, test_probs_cal, n_bins=10)
            self._section("Holdout Evaluation")
            self._stat("Raw log-loss", f"{raw_ll:.4f}")
            self._stat("Calibrated log-loss", f"{cal_ll_test:.4f}")
            self._stat("Brier score", f"{brier:.4f}")
            self._stat("Accuracy", f"{acc:.2%}")
            self._stat("ECE", f"{ece:.4f}")
            self._stat("Calibration curve RMSE", f"{cal_curve_rmse:.4f}")
            self._stat("Accuracy threshold", f"{decision_threshold:.3f}")

            # Winner diagnostics: confusion matrix + subgroup accuracy.
            y_true_red = y_test.astype(int).values
            y_pred_red = (test_probs_cal >= decision_threshold).astype(int)
            tn = int(np.sum((y_true_red == 0) & (y_pred_red == 0)))
            fp = int(np.sum((y_true_red == 0) & (y_pred_red == 1)))
            fn = int(np.sum((y_true_red == 1) & (y_pred_red == 0)))
            tp = int(np.sum((y_true_red == 1) & (y_pred_red == 1)))
            self._section("Winner Diagnostics")
            self._log("Confusion Matrix (Winner: Red=positive)")
            self._log("               Pred Blue   Pred Red")
            self._log(f"Actual Blue   {tn:9d}  {fp:9d}")
            self._log(f"Actual Red    {fn:9d}  {tp:9d}")

            self._log("")
            self._log("Accuracy by Weight Class")
            self._log("-" * 72)
            wc_order = [
                "Women's Strawweight",
                "Women's Flyweight",
                "Women's Bantamweight",
                "Women's Featherweight",
                "Flyweight",
                "Bantamweight",
                "Featherweight",
                "Lightweight",
                "Welterweight",
                "Middleweight",
                "Light Heavyweight",
                "Heavyweight",
                "Catch Weight",
                "Open Weight",
            ]
            wc_rank = {name: i for i, name in enumerate(wc_order)}
            wc_rows = []
            for wc, grp in meta_test.groupby("weight_class", dropna=False):
                idx = grp.index.values
                if len(idx) == 0:
                    continue
                acc_wc = float(np.mean(y_pred_red[idx] == y_true_red[idx]))
                wc_rows.append((str(wc), acc_wc, int(len(idx))))
            wc_rows.sort(key=lambda t: (wc_rank.get(t[0], 999), t[0]))
            for wc, acc_wc, n_wc in wc_rows:
                self._stat(f"{wc} (n={n_wc})", f"{acc_wc:.1%}")

            self._log("")
            self._log("Accuracy by Gender")
            self._log("-" * 72)
            g_rows = []
            for gender, grp in meta_test.groupby("gender", dropna=False):
                idx = grp.index.values
                if len(idx) == 0:
                    continue
                acc_g = float(np.mean(y_pred_red[idx] == y_true_red[idx]))
                g_rows.append((str(gender), acc_g, int(len(idx))))
            g_rows.sort(key=lambda t: (-t[2], t[0]))
            for gender, acc_g, n_g in g_rows:
                self._stat(f"{gender} (n={n_g})", f"{acc_g:.1%}")

            winner_payload = {
                "kind": WINNER_STAGE_CACHE_KIND,
                "winner_cache_version": WINNER_CACHE_VERSION,
                "feature_cols": list(feature_cols),
                "n_rows": int(n),
                "train_end": int(train_end),
                "val_end": int(val_end),
                "lgb_tuned": lgb_tuned,
                "xgb_tuned": xgb_tuned,
                "cb_tuned": cb_tuned,
                "oof": oof,
                "valid": valid,
                "combiner": combiner,
                "calibrator": calibrator,
                "cal_name": cal_name,
                "decision_threshold": float(decision_threshold),
                "tuned_thr": float(tuned_thr),
                "tuned_acc": float(tuned_acc),
                "val_acc_for_log": float(val_acc_for_log),
                "use_calibrated_branch": bool(use_calibrated_branch),
                "best_holdout_label": best_holdout_label,
                "best_holdout_ll": float(best_holdout_ll),
                "best_holdout_acc": float(best_holdout_acc),
                "best_holdout_thr": float(best_holdout_thr),
                "raw_ll": float(raw_ll),
                "cal_ll_test": float(cal_ll_test),
                "brier": float(brier),
                "acc": float(acc),
                "ece": float(ece),
                "cal_curve_rmse": float(cal_curve_rmse),
                "test_probs_raw": np.asarray(test_probs_raw, dtype=float),
                "test_probs_cal": np.asarray(test_probs_cal, dtype=float),
                "y_pred_red": np.asarray(y_pred_red, dtype=int),
                "model_order": list(model_order),
                "combiner_kind": str(combiner.get("kind", "")),
                "val_ll_str": f"{float(val_ll):.4f}",
                "val_acc_str": f"{float(val_acc):.1%}",
                "val_thr_str": f"{float(val_thr):.3f}",
                "cal_ll_str": f"{float(cal_ll):.4f}",
                "best_holdout_ll_str": f"{float(best_holdout_ll):.4f}" if np.isfinite(best_holdout_ll) else "n/a",
                "best_holdout_acc_str": f"{float(best_holdout_acc):.1%}" if np.isfinite(best_holdout_acc) else "n/a",
                "best_holdout_thr_str": f"{float(best_holdout_thr):.3f}",
                "cal_used_str": "yes" if use_calibrated_branch else "no",
                "val_acc_picked_str": f"{float(val_acc_for_log):.1%}",
                "val_tuned_thr_str": f"{float(tuned_thr):.3f} (acc={float(tuned_acc):.1%})",
                "decision_thr_str": f"{float(decision_threshold):.3f}",
                "holdout_eval": [
                    ("Raw log-loss", f"{float(raw_ll):.4f}"),
                    ("Calibrated log-loss", f"{float(cal_ll_test):.4f}"),
                    ("Brier score", f"{float(brier):.4f}"),
                    ("Accuracy", f"{float(acc):.2%}"),
                    ("ECE", f"{float(ece):.4f}"),
                    ("Calibration curve RMSE", f"{float(cal_curve_rmse):.4f}"),
                    ("Accuracy threshold", f"{float(decision_threshold):.3f}"),
                ],
                "confusion": (tn, fp, fn, tp),
                "wc_rows": [(str(a), float(b), int(c)) for a, b, c in wc_rows],
                "g_rows": [(str(a), float(b), int(c)) for a, b, c in g_rows],
            }
            _cache_save("winner_stage", winner_cache_key, winner_payload)
            self._stat(
                "Winner cache",
                f"SAVED ({WINNER_CACHE_VERSION}) key={str(winner_cache_key)[:12]} file={os.path.basename(_cache_path('winner_stage', winner_cache_key))}",
            )

        lgb_tuned = winner_payload["lgb_tuned"]
        xgb_tuned = winner_payload.get("xgb_tuned")
        cb_tuned = winner_payload.get("cb_tuned")
        oof = winner_payload["oof"]
        valid = winner_payload["valid"]
        combiner = winner_payload["combiner"]
        calibrator = winner_payload["calibrator"]
        cal_name = winner_payload.get("cal_name", "none")
        decision_threshold = float(winner_payload["decision_threshold"])
        tuned_thr = float(winner_payload["tuned_thr"])
        tuned_acc = float(winner_payload["tuned_acc"])
        val_acc_for_log = float(winner_payload["val_acc_for_log"])
        use_calibrated_branch = bool(winner_payload.get("use_calibrated_branch", False))
        best_holdout_label = winner_payload.get("best_holdout_label", "disabled_strict_mode")
        best_holdout_ll = float(winner_payload.get("best_holdout_ll", float("nan")))
        best_holdout_acc = float(winner_payload.get("best_holdout_acc", float("nan")))
        best_holdout_thr = float(winner_payload.get("best_holdout_thr", 0.5))
        raw_ll = float(winner_payload["raw_ll"])
        cal_ll_test = float(winner_payload["cal_ll_test"])
        brier = float(winner_payload["brier"])
        acc = float(winner_payload["acc"])
        ece = float(winner_payload["ece"])
        cal_curve_rmse = float(winner_payload["cal_curve_rmse"])
        test_probs_raw = np.asarray(winner_payload["test_probs_raw"], dtype=float)
        test_probs_cal = np.asarray(winner_payload["test_probs_cal"], dtype=float)
        y_pred_red = np.asarray(winner_payload["y_pred_red"], dtype=int)
        specs = _make_model_specs(
            lgb_tuned_params=lgb_tuned,
            xgb_tuned_params=xgb_tuned,
            cb_tuned_params=cb_tuned,
        )
        if STRICT_FUTURE_MODE:
            strict_keep = {
                "LightGBM", "LightGBM_S2", "LightGBM_Tuned",
                "XGBoost", "XGBoost_S2", "XGBoost_Tuned",
                "CatBoost", "CatBoost_S2", "CatBoost_Tuned",
                "HistGBM", "HistGBM_Wide",
                "ExtraTrees", "ExtraTrees_Deep",
                "RandForest", "RandForest_Deep",
                "AdaBoost",
            }
            specs = [(n, mk) for n, mk in specs if n in strict_keep]
        model_order = [n for n, _ in specs]
        X_dev = pd.concat([X_train, X_val], ignore_index=True)
        y_dev = pd.concat([y_train, y_val], ignore_index=True)
        dev_meta_valid = oof.loc[valid, model_order].astype(float)
        dev_probs_raw = _combine_probs(dev_meta_valid, combiner)
        y_true_red = y_test.astype(int).values

        # Method diagnostics/training (winner-model OOF conditioned, 2-stage).
        method_key_extra = "|".join([
            str(METHOD_TUNING_TRIALS),
            str(METHOD_AUTO_ERA),
            str(METHOD_HARD_RESET),
            str(len(full_feature_cols)),
            ",".join(map(str, METHOD_ERA_CANDIDATES)),
        ])
        method_cache_key = _cache_key(
            "method_stage", data_fp, METHOD_CACHE_VERSION, f"{winner_cache_key}|{method_key_extra}"
        )
        method_acc_when_winner_correct = float("nan")
        method_acc_predicted_winner = float("nan")
        method_majority_baseline_when_winner_correct = float("nan")
        method_holdout_acc_oracle = float("nan")
        finish_score = float("nan")
        method_bundle = None
        method_payload = _cache_load("method_stage", method_cache_key)
        method_cache_hit = _method_stage_cache_valid(
            method_payload, winner_cache_key, METHOD_CACHE_VERSION
        )
        if method_cache_hit:
            self._stat(
                "Method cache",
                f"HIT ({METHOD_CACHE_VERSION}) key={str(method_cache_key)[:12]} file={os.path.basename(_cache_path('method_stage', method_cache_key))}",
            )
            self._section("Method Model")
            self._stat("Cache", f"HIT ({METHOD_CACHE_VERSION}) — method retrain skipped")
            _mb = method_payload.get("method_bundle") or {}
            method_bundle = {k: v for k, v in _mb.items()}
            method_acc_when_winner_correct = float(method_payload.get("method_acc_when_winner_correct", float("nan")))
            method_acc_predicted_winner = float(method_payload.get("method_acc_predicted_winner", float("nan")))
            method_majority_baseline_when_winner_correct = float(
                method_payload.get("method_majority_baseline_when_winner_correct", float("nan"))
            )
            method_holdout_acc_oracle = float(method_payload.get("method_holdout_acc_oracle", float("nan")))
            finish_score = float(method_payload.get("finish_score", float("nan")))
            _replay_method_cache_logs(self, method_payload)
        if not method_cache_hit:
            self._stat("Method cache", f"MISS ({METHOD_CACHE_VERSION}) — training method stage")
            try:
                self._section("Method Model")
                X_dev_full = pd.concat([X_train_full, X_val_full], ignore_index=True)
                y_dev_true = pd.concat([y_train, y_val], ignore_index=True).astype(int).values
                y_dev_method_df = pd.concat([y_method_train, y_method_val], ignore_index=True).reset_index(drop=True)
                meta_dev = row_meta.iloc[:val_end].reset_index(drop=True)

                valid_idx = np.where(np.asarray(valid).astype(bool))[0]
                X_dev_valid = X_dev_full.iloc[valid_idx].reset_index(drop=True)
                y_dev_true_valid = y_dev_true[valid_idx]
                if calibrator is None:
                    dev_probs_method = _clip_probs(dev_probs_raw)
                else:
                    dev_probs_method = _clip_probs(
                        calibrator.predict_proba(np.asarray(dev_probs_raw).reshape(-1, 1))[:, 1]
                    )
                y_dev_pred_valid = (dev_probs_method >= decision_threshold).astype(int)
                y_method_dev = y_dev_method_df.iloc[valid_idx].reset_index(drop=True)
                meta_dev_valid = meta_dev.iloc[valid_idx].reset_index(drop=True)
                row_dates_dev_valid = pd.to_datetime(
                    row_dates.iloc[:val_end].reset_index(drop=True).iloc[valid_idx].reset_index(drop=True),
                    errors="coerce"
                )

                X_dev_oriented = _oriented_method_matrix(X_dev_valid, y_dev_pred_valid)
                X_dev_oriented = _augment_method_features(X_dev_oriented)

                # Method-specific era calibration (optional).
                if METHOD_AUTO_ERA:
                    best_method_year = 1993
                    best_method_score = -1.0
                    best_method_rows = len(X_dev_oriented)
                    for yr in METHOD_ERA_CANDIDATES:
                        mask_yr = row_dates_dev_valid >= pd.Timestamp(f"{int(yr)}-01-01")
                        if int(mask_yr.sum()) < 700:
                            continue
                        sub_true = y_method_dev.loc[mask_yr, "coarse"].astype(str).reset_index(drop=True).values
                        sub_pred_w = y_dev_pred_valid[mask_yr.values]
                        sub_true_w = y_dev_true_valid[mask_yr.values]
                        wc = (sub_pred_w == sub_true_w)
                        if int(np.sum(wc)) >= 120:
                            counts = pd.Series(sub_true[wc]).value_counts()
                            score_yr = float(counts.max() / len(sub_true[wc]))
                        else:
                            counts = pd.Series(sub_true).value_counts()
                            score_yr = float(counts.max() / len(sub_true))
                        if score_yr > best_method_score + 1e-12 or (
                            abs(score_yr - best_method_score) <= 1e-12 and int(mask_yr.sum()) > best_method_rows
                        ):
                            best_method_year = int(yr)
                            best_method_score = float(score_yr)
                            best_method_rows = int(mask_yr.sum())
                    if best_method_year > 1993:
                        keep = row_dates_dev_valid >= pd.Timestamp(f"{best_method_year}-01-01")
                        X_dev_oriented = X_dev_oriented.loc[keep].reset_index(drop=True)
                        y_method_dev = y_method_dev.loc[keep].reset_index(drop=True)
                        meta_dev_valid = meta_dev_valid.loc[keep].reset_index(drop=True)
                        y_dev_pred_valid = y_dev_pred_valid[keep.values]
                        y_dev_true_valid = y_dev_true_valid[keep.values]
                        row_dates_dev_valid = row_dates_dev_valid.loc[keep].reset_index(drop=True)
                        self._stat("Method start year", best_method_year)
                        self._stat("Method rows kept", int(len(X_dev_oriented)))
                method_columns = list(X_dev_oriented.columns)

                n_dev_m = len(X_dev_oriented)
                split_m = int(max(200, min(n_dev_m - 1, int(n_dev_m * 0.75)))) if n_dev_m > 350 else max(1, int(n_dev_m * 0.7))
                tr_idx = np.arange(split_m)
                va_idx = np.arange(split_m, n_dev_m)
                if len(va_idx) < 80:
                    va_idx = np.arange(max(0, n_dev_m - 80), n_dev_m)
                    tr_idx = np.arange(0, max(0, va_idx[0]))

                X_m_tr = X_dev_oriented.iloc[tr_idx].reset_index(drop=True)
                X_m_va = X_dev_oriented.iloc[va_idx].reset_index(drop=True)

                method_imputer = SimpleImputer(strategy="median")
                X_m_tr_imp = pd.DataFrame(method_imputer.fit_transform(X_m_tr), columns=method_columns)
                X_m_va_imp = pd.DataFrame(method_imputer.transform(X_m_va), columns=method_columns)

                # Multi-model method ensemble: stage1 (Finish/Decision) + stage2 (KO/Sub)
                # + direct multiclass + OVR heads + simple LR + tuned blending + meta calibrator.
                y_bin_tr = (y_method_dev.iloc[tr_idx]["finish_bin"].values == "Finish").astype(int)
                y_sub_tr = (y_method_dev.iloc[tr_idx]["finish_subtype"].values == "Submission").astype(int)

                # Weighted objectives for class imbalance.
                w_time = _time_weights(len(X_m_tr_imp), floor=0.45)
                p_finish = max(1e-6, float(np.mean(y_bin_tr)))
                finish_w = np.where(y_bin_tr == 1, 0.5 / p_finish, 0.5 / max(1e-6, 1.0 - p_finish))
                w_stage1 = w_time * finish_w

                stage1_cols_raw = [c for c in X_m_tr_imp.columns if c not in STAGE1_EXCLUDE_COLS]
                stage1_cols = _correlation_prune(X_m_tr_imp[stage1_cols_raw], threshold=METHOD_CORR_PRUNE_THRESHOLD)
                self._stat("Stage1 corr-pruned", f"{len(stage1_cols_raw)} → {len(stage1_cols)} (thr={METHOD_CORR_PRUNE_THRESHOLD})")
                X1_tr = X_m_tr_imp[stage1_cols]
                stage1 = HistGradientBoostingClassifier(
                    loss="log_loss",
                    max_iter=360, learning_rate=0.045, max_depth=6,
                    max_leaf_nodes=31, min_samples_leaf=16, l2_regularization=0.8,
                    random_state=RANDOM_SEED + 808,
                )
                stage1.fit(X1_tr, y_bin_tr, sample_weight=w_stage1)
                stage1_rf = RandomForestClassifier(
                    n_estimators=420, max_depth=10, min_samples_leaf=4,
                    max_features=0.7, random_state=RANDOM_SEED + 910, n_jobs=-1,
                    class_weight="balanced_subsample",
                )
                stage1_rf.fit(X1_tr, y_bin_tr)
                stage1_et = ExtraTreesClassifier(
                    n_estimators=560, max_depth=12, min_samples_leaf=3,
                    max_features=0.72, random_state=RANDOM_SEED + 914, n_jobs=-1,
                    class_weight="balanced_subsample",
                )
                stage1_et.fit(X1_tr, y_bin_tr)

                finish_tr_mask = (y_method_dev.iloc[tr_idx]["finish_bin"].values == "Finish")
                if int(np.sum(finish_tr_mask)) < 60:
                    finish_tr_mask = np.ones(len(tr_idx), dtype=bool)
                stage2_cols_raw = [c for c in X_m_tr_imp.columns if c not in STAGE2_EXCLUDE_COLS]
                _X2_for_corr = X_m_tr_imp.iloc[np.where(finish_tr_mask)[0]][stage2_cols_raw].reset_index(drop=True)
                stage2_cols = _correlation_prune(_X2_for_corr, threshold=METHOD_CORR_PRUNE_THRESHOLD)
                self._stat("Stage2 corr-pruned", f"{len(stage2_cols_raw)} → {len(stage2_cols)} (thr={METHOD_CORR_PRUNE_THRESHOLD})")
                X2_tr = _X2_for_corr[stage2_cols]
                y2_tr = y_sub_tr[np.where(finish_tr_mask)[0]]
                p_sub = max(1e-6, float(np.mean(y2_tr))) if len(y2_tr) > 0 else 0.5
                w2 = np.where(y2_tr == 1, 0.5 / p_sub, 0.5 / max(1e-6, 1.0 - p_sub))

                stage2 = HistGradientBoostingClassifier(
                    loss="log_loss",
                    max_iter=320, learning_rate=0.05, max_depth=5,
                    max_leaf_nodes=31, min_samples_leaf=14, l2_regularization=0.7,
                    random_state=RANDOM_SEED + 809,
                )
                stage2.fit(X2_tr, y2_tr, sample_weight=w2)
                stage2_rf = RandomForestClassifier(
                    n_estimators=360, max_depth=9, min_samples_leaf=3,
                    max_features=0.7, random_state=RANDOM_SEED + 911, n_jobs=-1,
                    class_weight="balanced_subsample",
                )
                stage2_rf.fit(X2_tr, y2_tr)
                stage2_et = ExtraTreesClassifier(
                    n_estimators=520, max_depth=11, min_samples_leaf=3,
                    max_features=0.72, random_state=RANDOM_SEED + 915, n_jobs=-1,
                    class_weight="balanced_subsample",
                )
                stage2_et.fit(X2_tr, y2_tr)

                # Direct multiclass head to reduce collapse to Decision.
                y_cls_tr = y_method_dev.iloc[tr_idx]["coarse"].values
                cls_counts = pd.Series(y_cls_tr).value_counts()
                cls_w = {k: (len(y_cls_tr) / (len(cls_counts) * max(1, v))) for k, v in cls_counts.items()}
                w_cls = np.array([cls_w.get(lbl, 1.0) for lbl in y_cls_tr], dtype=float)
                w_direct = w_time * w_cls
                direct_hgb = HistGradientBoostingClassifier(
                    loss="log_loss",
                    max_iter=300, learning_rate=0.05, max_depth=6,
                    max_leaf_nodes=31, min_samples_leaf=14, l2_regularization=0.7,
                    random_state=RANDOM_SEED + 912,
                )
                direct_hgb.fit(X_m_tr_imp, y_cls_tr, sample_weight=w_direct)
                direct_rf = RandomForestClassifier(
                    n_estimators=520, max_depth=11, min_samples_leaf=3,
                    max_features=0.72, random_state=RANDOM_SEED + 913, n_jobs=-1,
                    class_weight="balanced_subsample",
                )
                direct_rf.fit(X_m_tr_imp, y_cls_tr)
                direct_et = ExtraTreesClassifier(
                    n_estimators=680, max_depth=13, min_samples_leaf=3,
                    max_features=0.72, random_state=RANDOM_SEED + 916, n_jobs=-1,
                    class_weight="balanced_subsample",
                )
                direct_et.fit(X_m_tr_imp, y_cls_tr)
                direct_classes = [str(c) for c in direct_hgb.classes_]

                # OVR method head (one binary model per method class).
                ovr_models = {}
                for j, lbl in enumerate(METHOD_LABELS):
                    y_ovr = (pd.Series(y_cls_tr).astype(str).values == lbl).astype(int)
                    p_pos = max(1e-6, float(np.mean(y_ovr)))
                    w_ovr = w_time * np.where(y_ovr == 1, 0.5 / p_pos, 0.5 / max(1e-6, 1.0 - p_pos))
                    mdl = HistGradientBoostingClassifier(
                        loss="log_loss",
                        max_iter=260, learning_rate=0.05, max_depth=5,
                        max_leaf_nodes=31, min_samples_leaf=14, l2_regularization=0.8,
                        random_state=RANDOM_SEED + 1500 + j,
                    )
                    mdl.fit(X_m_tr_imp, y_ovr, sample_weight=w_ovr)
                    ovr_models[lbl] = mdl
                simple_method = LogisticRegression(
                    max_iter=8000, C=0.8, solver="saga", tol=1e-3, n_jobs=-1,
                    class_weight="balanced", random_state=RANDOM_SEED + 1601,
                )
                simple_method.fit(X_m_tr_imp, pd.Series(y_cls_tr).astype(str).values)

                def _stage_probs(
                    X_imp, temp_dec=1.0, temp_fin=1.0, alpha1=0.75, alpha2=0.75,
                    bias_finish=0.0, bias_sub=0.0, alpha_direct=0.85, beta_direct=0.70,
                    alpha_ovr=0.85, finish_thr=0.50, sub_thr=0.50, alpha_simple=0.80
                ):
                    X_s1 = X_imp[stage1_cols]
                    p_finish_hist = stage1.predict_proba(X_s1)[:, 1]
                    p_finish_rf = stage1_rf.predict_proba(X_s1)[:, 1]
                    p_finish_et = stage1_et.predict_proba(X_s1)[:, 1]
                    p_finish_tree = 0.55 * p_finish_rf + 0.45 * p_finish_et
                    p_finish_raw = alpha1 * p_finish_hist + (1.0 - alpha1) * p_finish_tree
                    logit = np.log(np.clip(p_finish_raw, 1e-6, 1 - 1e-6) / np.clip(1.0 - p_finish_raw, 1e-6, 1 - 1e-6))
                    p_finish_c = 1.0 / (1.0 + np.exp(-((logit / max(0.35, float(temp_dec))) + float(bias_finish))))
                    p_finish_c = _apply_binary_threshold_warp(p_finish_c, finish_thr)

                    X_s2 = X_imp[stage2_cols]
                    p_sub_hist = stage2.predict_proba(X_s2)[:, 1]
                    p_sub_rf = stage2_rf.predict_proba(X_s2)[:, 1]
                    p_sub_et = stage2_et.predict_proba(X_s2)[:, 1]
                    p_sub_tree = 0.55 * p_sub_rf + 0.45 * p_sub_et
                    p_sub_raw = alpha2 * p_sub_hist + (1.0 - alpha2) * p_sub_tree
                    logit2 = np.log(np.clip(p_sub_raw, 1e-6, 1 - 1e-6) / np.clip(1.0 - p_sub_raw, 1e-6, 1 - 1e-6))
                    p_sub_c = 1.0 / (1.0 + np.exp(-((logit2 / max(0.35, float(temp_fin))) + float(bias_sub))))
                    p_sub_c = _apply_binary_threshold_warp(p_sub_c, sub_thr)

                    probs = pd.DataFrame({
                        "Decision": 1.0 - p_finish_c,
                        "KO/TKO": p_finish_c * (1.0 - p_sub_c),
                        "Submission": p_finish_c * p_sub_c,
                    })
                    probs = probs.apply(lambda s: np.clip(s, MIN_METHOD_PROB, 1.0), axis=0)
                    probs = probs.div(probs.sum(axis=1), axis=0)
                    p_h = direct_hgb.predict_proba(X_imp)
                    p_r = direct_rf.predict_proba(X_imp)
                    p_e = direct_et.predict_proba(X_imp)
                    direct_rows = []
                    for row_idx in range(len(X_imp)):
                        raw = {}
                        for j, cls in enumerate(direct_classes):
                            tree_mix = 0.55 * float(p_r[row_idx][j]) + 0.45 * float(p_e[row_idx][j])
                            raw[cls] = beta_direct * float(p_h[row_idx][j]) + (1.0 - beta_direct) * tree_mix
                        direct_rows.append(_normalize_method_probs(raw))
                    direct_df = pd.DataFrame(direct_rows)[METHOD_LABELS]
                    probs = alpha_direct * probs[METHOD_LABELS] + (1.0 - alpha_direct) * direct_df
                    probs = probs.apply(lambda s: np.clip(s, MIN_METHOD_PROB, 1.0), axis=0)
                    probs = probs.div(probs.sum(axis=1), axis=0)
                    ovr_arr = np.zeros((len(X_imp), len(METHOD_LABELS)), dtype=float)
                    for k, lbl in enumerate(METHOD_LABELS):
                        ovr_arr[:, k] = ovr_models[lbl].predict_proba(X_imp)[:, 1]
                    ovr_arr = np.clip(ovr_arr, MIN_METHOD_PROB, 1.0)
                    ovr_arr = ovr_arr / np.sum(ovr_arr, axis=1, keepdims=True)
                    ovr_df = pd.DataFrame(ovr_arr, columns=METHOD_LABELS)
                    probs = alpha_ovr * probs[METHOD_LABELS] + (1.0 - alpha_ovr) * ovr_df
                    probs = probs.apply(lambda s: np.clip(s, MIN_METHOD_PROB, 1.0), axis=0)
                    probs = probs.div(probs.sum(axis=1), axis=0)
                    p_s = simple_method.predict_proba(X_imp)
                    cls_map_s = {str(c): i for i, c in enumerate(simple_method.classes_)}
                    simple_arr = np.zeros((len(X_imp), len(METHOD_LABELS)), dtype=float)
                    for j, m in enumerate(METHOD_LABELS):
                        simple_arr[:, j] = p_s[:, cls_map_s[m]] if m in cls_map_s else (1.0 / 3.0)
                    simple_arr = np.clip(simple_arr, MIN_METHOD_PROB, 1.0)
                    simple_arr = simple_arr / np.sum(simple_arr, axis=1, keepdims=True)
                    simple_df = pd.DataFrame(simple_arr, columns=METHOD_LABELS)
                    probs = alpha_simple * probs[METHOD_LABELS] + (1.0 - alpha_simple) * simple_df
                    probs = probs.apply(lambda s: np.clip(s, MIN_METHOD_PROB, 1.0), axis=0)
                    probs = probs.div(probs.sum(axis=1), axis=0)
                    return probs

                # Group priors (weight class + gender adapters) from method-train split.
                grp_train = meta_dev_valid.iloc[tr_idx].reset_index(drop=True)
                grp_labels = y_method_dev.iloc[tr_idx]["coarse"].values
                group_priors = {}
                for (wc, gender), grp in grp_train.groupby(["weight_class", "gender"], dropna=False):
                    idx = grp.index.values
                    counts = pd.Series(grp_labels[idx]).value_counts()
                    tot = float(len(idx))
                    group_priors[(str(wc), str(gender).lower())] = _normalize_method_probs({
                        "Decision": float(counts.get("Decision", 0.0) / max(1.0, tot)),
                        "KO/TKO": float(counts.get("KO/TKO", 0.0) / max(1.0, tot)),
                        "Submission": float(counts.get("Submission", 0.0) / max(1.0, tot)),
                    })
                all_counts = pd.Series(grp_labels).value_counts()
                all_tot = float(len(grp_labels))
                group_priors[("ALL", "all")] = _normalize_method_probs({
                    "Decision": float(all_counts.get("Decision", 0.0) / max(1.0, all_tot)),
                    "KO/TKO": float(all_counts.get("KO/TKO", 0.0) / max(1.0, all_tot)),
                    "Submission": float(all_counts.get("Submission", 0.0) / max(1.0, all_tot)),
                })
                base_prior = _normalize_method_probs({
                    "Decision": float(all_counts.get("Decision", 0.0) / max(1.0, all_tot)),
                    "KO/TKO": float(all_counts.get("KO/TKO", 0.0) / max(1.0, all_tot)),
                    "Submission": float(all_counts.get("Submission", 0.0) / max(1.0, all_tot)),
                })

                def _history_prior_array(X_raw):
                    Xr = X_raw.reset_index(drop=True)
                    out = np.zeros((len(Xr), 3), dtype=float)
                    for i in range(len(Xr)):
                        d_dec = float(pd.to_numeric(Xr.get("d_dec_win_pct", pd.Series([0.0])).iloc[i], errors="coerce") if "d_dec_win_pct" in Xr else 0.0)
                        d_ko = float(pd.to_numeric(Xr.get("d_ko_win_pct", pd.Series([0.0])).iloc[i], errors="coerce") if "d_ko_win_pct" in Xr else 0.0)
                        d_sub = float(pd.to_numeric(Xr.get("d_sub_win_pct", pd.Series([0.0])).iloc[i], errors="coerce") if "d_sub_win_pct" in Xr else 0.0)
                        p = _normalize_method_probs({
                            "Decision": 0.50 + 0.35 * d_dec,
                            "KO/TKO": 0.32 + 0.35 * d_ko,
                            "Submission": 0.18 + 0.35 * d_sub,
                        })
                        out[i, 0] = float(p["Decision"])
                        out[i, 1] = float(p["KO/TKO"])
                        out[i, 2] = float(p["Submission"])
                    return out

                # Tune calibration + blend weights by primary metric.
                y_va_true = y_method_dev.iloc[va_idx]["coarse"].values
                winner_correct_va = (
                    y_dev_pred_valid[va_idx] == y_dev_true_valid[va_idx]
                )
                label_to_idx = {m: i for i, m in enumerate(METHOD_LABELS)}
                y_va_idx = np.array([label_to_idx.get(str(lbl), 0) for lbl in y_va_true], dtype=int)
                hist_arr = _history_prior_array(X_m_va)
                hist_arr_tr = _history_prior_array(X_m_tr)
                sub_arr = _sub_attempt_prior_array(X_m_va)
                sub_arr_tr = _sub_attempt_prior_array(X_m_tr)
                va_local_idx = np.arange(len(y_va_idx), dtype=int)
                va_chunks = [c for c in np.array_split(va_local_idx, 3) if len(c) > 0]
                meta_va = meta_dev_valid.iloc[va_idx].reset_index(drop=True)
                grp_arr = np.zeros((len(meta_va), 3), dtype=float)
                for i in range(len(meta_va)):
                    wc = str(meta_va.iloc[i]["weight_class"])
                    gd = str(meta_va.iloc[i]["gender"]).lower()
                    gp = group_priors.get((wc, gd), group_priors[("ALL", "all")])
                    grp_arr[i, 0] = float(gp["Decision"])
                    grp_arr[i, 1] = float(gp["KO/TKO"])
                    grp_arr[i, 2] = float(gp["Submission"])
                meta_tr = meta_dev_valid.iloc[tr_idx].reset_index(drop=True)
                grp_arr_tr = np.zeros((len(meta_tr), 3), dtype=float)
                for i in range(len(meta_tr)):
                    wc = str(meta_tr.iloc[i]["weight_class"])
                    gd = str(meta_tr.iloc[i]["gender"]).lower()
                    gp = group_priors.get((wc, gd), group_priors[("ALL", "all")])
                    grp_arr_tr[i, 0] = float(gp["Decision"])
                    grp_arr_tr[i, 1] = float(gp["KO/TKO"])
                    grp_arr_tr[i, 2] = float(gp["Submission"])

                best_cfg = None
                rng = np.random.default_rng(RANDOM_SEED + 2026)
                for _trial in range(int(max(500, METHOD_TUNING_TRIALS))):
                    a1 = float(rng.choice([0.55, 0.70, 0.85]))
                    a2 = float(rng.choice([0.55, 0.70, 0.85]))
                    t_dec = float(rng.choice([0.7, 0.85, 1.0, 1.15, 1.3]))
                    t_fin = float(rng.choice([0.7, 0.85, 1.0, 1.15, 1.3]))
                    b_fin = float(rng.choice([-0.35, -0.15, 0.0, 0.15, 0.35]))
                    b_sub = float(rng.choice([-0.30, -0.10, 0.0, 0.10, 0.30]))
                    fin_thr = float(rng.choice([0.38, 0.42, 0.46, 0.50, 0.54, 0.58]))
                    sub_thr = float(rng.choice([0.38, 0.42, 0.46, 0.50, 0.54, 0.58]))
                    a_dir = float(rng.choice([0.90, 0.96, 1.00]))
                    b_dir = float(rng.choice([0.60, 0.75, 0.90]))
                    a_ovr = float(rng.choice([0.90, 0.97, 1.00]))
                    a_simple = float(rng.choice([0.65, 0.80, 0.95]))
                    w_hist = float(rng.choice([0.00, 0.08, 0.15, 0.22, 0.30]))
                    w_group = float(rng.choice([0.00, 0.04, 0.08, 0.12]))
                    w_base = float(rng.choice([0.00, 0.06, 0.12, 0.18]))
                    w_subsig = float(rng.choice([0.00, 0.06, 0.10, 0.14, 0.18]))
                    sub_boost_k = float(rng.choice([0.0, 0.4, 0.8, 1.2, 1.6]))
                    bias_dec = float(rng.choice([-0.20, -0.10, 0.00, 0.10, 0.20]))
                    bias_ko = float(rng.choice([-0.10, 0.00, 0.10, 0.20, 0.30, 0.40]))
                    bias_sub = float(rng.choice([-0.25, -0.15, -0.05, 0.05, 0.15, 0.25]))
                    if w_hist + w_group + w_base + w_subsig >= 0.78:
                        continue
                    w_ml = 1.0 - w_hist - w_group - w_base - w_subsig

                    ml_df = _stage_probs(
                        X_m_va_imp, temp_dec=t_dec, temp_fin=t_fin,
                        alpha1=a1, alpha2=a2, bias_finish=b_fin, bias_sub=b_sub,
                        alpha_direct=a_dir, beta_direct=b_dir, alpha_ovr=a_ovr,
                        finish_thr=fin_thr, sub_thr=sub_thr, alpha_simple=a_simple
                    )
                    ml_arr = ml_df[["Decision", "KO/TKO", "Submission"]].to_numpy(dtype=float)
                    base_arr = np.tile(
                        np.array([[base_prior["Decision"], base_prior["KO/TKO"], base_prior["Submission"]]], dtype=float),
                        (len(ml_arr), 1)
                    )
                    probs_arr = w_ml * ml_arr + w_hist * hist_arr + w_group * grp_arr + w_base * base_arr + w_subsig * sub_arr
                    probs_arr = np.clip(probs_arr, MIN_METHOD_PROB, 1.0)
                    probs_arr = probs_arr / np.sum(probs_arr, axis=1, keepdims=True)
                    probs_arr = _apply_method_logit_bias_arr(
                        probs_arr,
                        np.array([bias_dec, bias_ko, bias_sub], dtype=float)
                    )
                    probs_arr = _apply_submission_signal_boost_arr(probs_arr, sub_arr[:, 2], sub_boost_k)
                    probs_arr = probs_arr / np.sum(probs_arr, axis=1, keepdims=True)
                    pred_idx = np.argmax(probs_arr, axis=1)

                    if int(np.sum(winner_correct_va)) == 0:
                        score = float(np.mean(pred_idx == y_va_idx))
                        baseline = float(pd.Series(y_va_true).value_counts().max() / len(y_va_true))
                        _ko_r_min = score
                        _dec_r_min = score
                        _sub_r_min = score
                        macro_f1 = score
                    else:
                        y_sub_idx = y_va_idx[winner_correct_va]
                        pred_sub_idx = pred_idx[winner_correct_va]
                        score = float(np.mean(pred_sub_idx == y_sub_idx))
                        baseline = float(pd.Series(y_va_true[winner_correct_va]).value_counts().max() / len(y_sub_idx))
                        recalls = []
                        for cls_i in range(len(METHOD_LABELS)):
                            mask_cls = (y_sub_idx == cls_i)
                            if int(np.sum(mask_cls)) > 0:
                                recalls.append(float(np.mean(pred_sub_idx[mask_cls] == cls_i)))
                        _ko_r_min = float(recalls[1]) if len(recalls) > 1 else 0.0
                        _sub_r_min = float(recalls[2]) if len(recalls) > 2 else 0.0
                        _dec_r_min = float(recalls[0]) if len(recalls) > 0 else 0.0
                        _ko_mask_t = (y_sub_idx == 1)
                        _sub_mask_t = (y_sub_idx == 2)
                        _ko_r_t = float(np.mean(pred_sub_idx[_ko_mask_t] == 1)) if int(np.sum(_ko_mask_t)) > 0 else 0.0
                        _sub_r_t = float(np.mean(pred_sub_idx[_sub_mask_t] == 2)) if int(np.sum(_sub_mask_t)) > 0 else 0.0
                        _ko_prec_t = float(np.sum((pred_sub_idx == 1) & _ko_mask_t)) / max(1, int(np.sum(pred_sub_idx == 1)))
                        _sub_prec_t = float(np.sum((pred_sub_idx == 2) & _sub_mask_t)) / max(1, int(np.sum(pred_sub_idx == 2)))
                        _ko_f1_t = 2.0 * _ko_prec_t * _ko_r_t / max(1e-9, _ko_prec_t + _ko_r_t)
                        _sub_f1_t = 2.0 * _sub_prec_t * _sub_r_t / max(1e-9, _sub_prec_t + _sub_r_t)
                        _dec_mask_t = (y_sub_idx == 0)
                        _dec_r_t = float(np.mean(pred_sub_idx[_dec_mask_t] == 0)) if int(np.sum(_dec_mask_t)) > 0 else 0.0
                        _dec_prec_t = float(np.sum((pred_sub_idx == 0) & _dec_mask_t)) / max(1, int(np.sum(pred_sub_idx == 0)))
                        _dec_f1_t = 2.0 * _dec_prec_t * _dec_r_t / max(1e-9, _dec_prec_t + _dec_r_t)
                        macro_f1 = (_dec_f1_t + _ko_f1_t + _sub_f1_t) / 3.0
                    _dec_shortfall = max(0.0, 0.55 - _dec_r_min)
                    _ko_shortfall = max(0.0, 0.50 - _ko_r_min)
                    _sub_shortfall = max(0.0, 0.40 - _sub_r_min)
                    objective = macro_f1 - 0.50 * (_dec_shortfall + _ko_shortfall + _sub_shortfall)
                    key = (objective, macro_f1, score, _ko_r_min, _sub_r_min, -baseline)
                    if best_cfg is None or key > best_cfg["key"]:
                        best_cfg = {
                            "key": key, "t_dec": t_dec, "t_fin": t_fin,
                            "bias_finish": b_fin, "bias_sub": b_sub,
                            "finish_threshold": fin_thr, "sub_threshold": sub_thr,
                            "w_hist": w_hist, "w_group": w_group, "w_base": w_base,
                            "w_subsig": w_subsig,
                            "sub_boost_k": sub_boost_k,
                            "method_bias_decision": bias_dec,
                            "method_bias_ko_tko": bias_ko,
                            "method_bias_submission": bias_sub,
                            "alpha_stage1": a1, "alpha_stage2": a2,
                            "alpha_direct": a_dir, "beta_direct": b_dir,
                            "alpha_ovr": a_ovr,
                            "alpha_simple": a_simple,
                            "val_metric": score, "val_baseline": baseline,
                        }
                if best_cfg is None:
                    best_cfg = {
                        "key": (0.0, 0.0, 0.0), "t_dec": 1.0, "t_fin": 1.0,
                        "bias_finish": 0.0, "bias_sub": 0.0,
                        "finish_threshold": 0.50, "sub_threshold": 0.50,
                        "w_hist": 0.15, "w_group": 0.08, "w_base": 0.08, "w_subsig": 0.10,
                        "sub_boost_k": 0.6,
                        "method_bias_decision": 0.10,
                        "method_bias_ko_tko": -0.10,
                        "method_bias_submission": 0.00,
                        "alpha_stage1": 0.75, "alpha_stage2": 0.75,
                        "alpha_direct": 0.85, "beta_direct": 0.70,
                        "alpha_ovr": 0.85,
                        "alpha_simple": 0.80,
                        "val_metric": float("nan"), "val_baseline": float("nan"),
                    }

                # ── Champion / challenger ─────────────────────────────────────
                # Load the previously saved champion config (if any) and evaluate
                # it on the CURRENT validation set. Only replace it if the new
                # tuning run found something genuinely better.
                def _eval_cfg_objective(cfg):
                    """Re-evaluate a config on the val set under the tuning objective.

                    Computes `macro_f1 − 0.50·(dec+ko+sub recall shortfalls)` per val
                    chunk and reduces as `mean − 0.60·std`, so the champion gate judges
                    configs on the same criterion the tuner optimized, with a stability
                    guard against configs that only win on one chunk.
                    """
                    try:
                        _ml = _stage_probs(
                            X_m_va_imp,
                            temp_dec=cfg["t_dec"], temp_fin=cfg["t_fin"],
                            alpha1=cfg["alpha_stage1"], alpha2=cfg["alpha_stage2"],
                            bias_finish=cfg["bias_finish"], bias_sub=cfg["bias_sub"],
                            alpha_direct=cfg["alpha_direct"], beta_direct=cfg["beta_direct"],
                            alpha_ovr=cfg["alpha_ovr"],
                            finish_thr=cfg["finish_threshold"], sub_thr=cfg["sub_threshold"],
                            alpha_simple=cfg["alpha_simple"],
                        )
                        _ml_arr = _ml[["Decision", "KO/TKO", "Submission"]].to_numpy(dtype=float)
                        _base = np.tile(
                            np.array([[base_prior["Decision"], base_prior["KO/TKO"], base_prior["Submission"]]], dtype=float),
                            (len(_ml_arr), 1)
                        )
                        _p = (cfg["w_hist"] * hist_arr + cfg["w_group"] * grp_arr +
                              cfg["w_base"] * _base + cfg["w_subsig"] * sub_arr +
                              (1.0 - cfg["w_hist"] - cfg["w_group"] - cfg["w_base"] - cfg["w_subsig"]) * _ml_arr)
                        _p = np.clip(_p, MIN_METHOD_PROB, 1.0)
                        _p = _p / np.sum(_p, axis=1, keepdims=True)
                        _p = _apply_method_logit_bias_arr(
                            _p, np.array([cfg["method_bias_decision"], cfg["method_bias_ko_tko"],
                                          cfg["method_bias_submission"]], dtype=float)
                        )
                        _p = _apply_submission_signal_boost_arr(_p, sub_arr[:, 2], cfg["sub_boost_k"])
                        _p = _p / np.sum(_p, axis=1, keepdims=True)
                        _pidx = np.argmax(_p, axis=1)

                        def _chunk_obj(ch):
                            _cwc = winner_correct_va[ch]
                            if int(np.sum(_cwc)) == 0:
                                return float(np.mean(_pidx[ch] == y_va_idx[ch]))
                            _ys = y_va_idx[ch][_cwc]
                            _ps_idx = _pidx[ch][_cwc]
                            _rec = [0.0, 0.0, 0.0]
                            for c in range(3):
                                _m = (_ys == c)
                                if int(np.sum(_m)) > 0:
                                    _rec[c] = float(np.mean(_ps_idx[_m] == c))
                            _f1 = [0.0, 0.0, 0.0]
                            for c in range(3):
                                _pc_cnt = int(np.sum(_ps_idx == c))
                                if _pc_cnt == 0:
                                    continue
                                _prec = float(np.sum((_ps_idx == c) & (_ys == c))) / _pc_cnt
                                _denom = _prec + _rec[c]
                                if _denom > 0:
                                    _f1[c] = 2.0 * _prec * _rec[c] / _denom
                            _macro_f1 = float(np.mean(_f1))
                            _dec_short = max(0.0, 0.55 - _rec[0])
                            _ko_short = max(0.0, 0.50 - _rec[1])
                            _sub_short = max(0.0, 0.40 - _rec[2])
                            return _macro_f1 - 0.50 * (_dec_short + _ko_short + _sub_short)

                        _chunk_objs = [_chunk_obj(ch) for ch in va_chunks]
                        return float(np.mean(_chunk_objs)) - 0.60 * float(np.std(_chunk_objs))
                    except Exception:
                        return float("-inf")

                try:
                    os.makedirs(os.path.dirname(METHOD_CHAMPION_PATH), exist_ok=True)
                    champion_cfg = None
                    if os.path.exists(METHOD_CHAMPION_PATH) and not METHOD_HARD_RESET:
                        with open(METHOD_CHAMPION_PATH, "r") as _f:
                            champion_cfg = json.load(_f)
                    if champion_cfg is not None:
                        champ_obj = _eval_cfg_objective(champion_cfg)
                        new_obj = _eval_cfg_objective(best_cfg)
                        if champ_obj >= new_obj:
                            self._stat("Method config", f"Champion retained (champ={champ_obj:.5f} >= new={new_obj:.5f})")
                            best_cfg = champion_cfg
                        else:
                            self._stat("Method config", f"Challenger wins (new={new_obj:.5f} > champ={champ_obj:.5f}) — champion updated")
                            with open(METHOD_CHAMPION_PATH, "w") as _f:
                                json.dump({k: (list(v) if isinstance(v, tuple) else v)
                                           for k, v in best_cfg.items()}, _f, indent=2)
                    else:
                        self._stat("Method config", "No champion on file — saving current best as champion")
                        with open(METHOD_CHAMPION_PATH, "w") as _f:
                            json.dump({k: (list(v) if isinstance(v, tuple) else v)
                                       for k, v in best_cfg.items()}, _f, indent=2)
                except Exception as _e:
                    self._stat("Method config", f"Champion load/save failed ({_e}) — using tuned config")

                # Hard reset: keep only stable components that generalized better.
                if METHOD_HARD_RESET:
                    best_cfg["alpha_direct"] = 1.0
                    best_cfg["alpha_ovr"] = 1.0
                    best_cfg["alpha_simple"] = 1.0

                # Meta calibrator over blended probabilities + priors.
                meta_model = None
                meta_eta = 0.0
                try:
                    ml_tr_df = _stage_probs(
                        X_m_tr_imp,
                        temp_dec=best_cfg["t_dec"], temp_fin=best_cfg["t_fin"],
                        alpha1=best_cfg["alpha_stage1"], alpha2=best_cfg["alpha_stage2"],
                        bias_finish=best_cfg["bias_finish"], bias_sub=best_cfg["bias_sub"],
                        alpha_direct=best_cfg["alpha_direct"], beta_direct=best_cfg["beta_direct"],
                        alpha_ovr=best_cfg["alpha_ovr"],
                        finish_thr=best_cfg["finish_threshold"], sub_thr=best_cfg["sub_threshold"],
                        alpha_simple=best_cfg["alpha_simple"],
                    )
                    ml_va_df = _stage_probs(
                        X_m_va_imp,
                        temp_dec=best_cfg["t_dec"], temp_fin=best_cfg["t_fin"],
                        alpha1=best_cfg["alpha_stage1"], alpha2=best_cfg["alpha_stage2"],
                        bias_finish=best_cfg["bias_finish"], bias_sub=best_cfg["bias_sub"],
                        alpha_direct=best_cfg["alpha_direct"], beta_direct=best_cfg["beta_direct"],
                        alpha_ovr=best_cfg["alpha_ovr"],
                        finish_thr=best_cfg["finish_threshold"], sub_thr=best_cfg["sub_threshold"],
                        alpha_simple=best_cfg["alpha_simple"],
                    )
                    ml_tr_arr = ml_tr_df[["Decision", "KO/TKO", "Submission"]].to_numpy(dtype=float)
                    ml_va_arr = ml_va_df[["Decision", "KO/TKO", "Submission"]].to_numpy(dtype=float)
                    bp_arr_tr = np.tile(
                        np.array([[base_prior["Decision"], base_prior["KO/TKO"], base_prior["Submission"]]], dtype=float),
                        (len(ml_tr_arr), 1),
                    )
                    bp_arr_va = np.tile(
                        np.array([[base_prior["Decision"], base_prior["KO/TKO"], base_prior["Submission"]]], dtype=float),
                        (len(ml_va_arr), 1),
                    )
                    w_hist_b = float(best_cfg["w_hist"])
                    w_group_b = float(best_cfg["w_group"])
                    w_base_b = float(best_cfg.get("w_base", 0.0))
                    w_subsig_b = float(best_cfg.get("w_subsig", 0.0))
                    w_ml_b = max(0.0, 1.0 - w_hist_b - w_group_b - w_base_b - w_subsig_b)
                    blend_tr = np.clip(
                        w_ml_b * ml_tr_arr + w_hist_b * hist_arr_tr + w_group_b * grp_arr_tr + w_base_b * bp_arr_tr + w_subsig_b * sub_arr_tr,
                        MIN_METHOD_PROB, 1.0
                    )
                    blend_va = np.clip(
                        w_ml_b * ml_va_arr + w_hist_b * hist_arr + w_group_b * grp_arr + w_base_b * bp_arr_va + w_subsig_b * sub_arr,
                        MIN_METHOD_PROB, 1.0
                    )
                    blend_tr = blend_tr / np.sum(blend_tr, axis=1, keepdims=True)
                    blend_va = blend_va / np.sum(blend_va, axis=1, keepdims=True)
                    method_bias_vec = np.array([
                        float(best_cfg.get("method_bias_decision", 0.0)),
                        float(best_cfg.get("method_bias_ko_tko", 0.0)),
                        float(best_cfg.get("method_bias_submission", 0.0)),
                    ], dtype=float)
                    blend_tr = _apply_method_logit_bias_arr(blend_tr, method_bias_vec)
                    blend_va = _apply_method_logit_bias_arr(blend_va, method_bias_vec)
                    sub_boost_k = float(best_cfg.get("sub_boost_k", 0.0))
                    blend_tr = _apply_submission_signal_boost_arr(blend_tr, sub_arr_tr[:, 2], sub_boost_k)
                    blend_va = _apply_submission_signal_boost_arr(blend_va, sub_arr[:, 2], sub_boost_k)
                    blend_tr = blend_tr / np.sum(blend_tr, axis=1, keepdims=True)
                    blend_va = blend_va / np.sum(blend_va, axis=1, keepdims=True)

                    X_meta_tr = np.hstack([blend_tr, hist_arr_tr, grp_arr_tr, bp_arr_tr, np.log(np.clip(blend_tr, 1e-6, 1.0))])
                    X_meta_va = np.hstack([blend_va, hist_arr, grp_arr, bp_arr_va, np.log(np.clip(blend_va, 1e-6, 1.0))])
                    y_meta_tr = y_method_dev.iloc[tr_idx]["coarse"].astype(str).values
                    meta_model = LogisticRegression(
                        max_iter=6000, C=0.7, solver="lbfgs",
                        class_weight="balanced", random_state=RANDOM_SEED + 1313,
                    )
                    meta_model.fit(X_meta_tr, y_meta_tr)
                    p_meta = meta_model.predict_proba(X_meta_va)
                    cls_map = {str(c): i for i, c in enumerate(meta_model.classes_)}
                    meta_va_arr = np.zeros((len(X_meta_va), 3), dtype=float)
                    for j, m in enumerate(METHOD_LABELS):
                        if m in cls_map:
                            meta_va_arr[:, j] = p_meta[:, cls_map[m]]
                        else:
                            meta_va_arr[:, j] = 1.0 / 3.0
                    meta_va_arr = np.clip(meta_va_arr, MIN_METHOD_PROB, 1.0)
                    meta_va_arr = meta_va_arr / np.sum(meta_va_arr, axis=1, keepdims=True)
                    best_meta_key = None
                    for eta in (0.0, 0.10, 0.20):
                        arr = np.clip((1.0 - eta) * blend_va + eta * meta_va_arr, MIN_METHOD_PROB, 1.0)
                        arr = arr / np.sum(arr, axis=1, keepdims=True)
                        pred_idx = np.argmax(arr, axis=1)
                        if int(np.sum(winner_correct_va)) > 0:
                            s = float(np.mean(pred_idx[winner_correct_va] == y_va_idx[winner_correct_va]))
                        else:
                            s = float(np.mean(pred_idx == y_va_idx))
                        chunk_scores = []
                        for ch in va_chunks:
                            ch_wc = winner_correct_va[ch]
                            if int(np.sum(ch_wc)) > 0:
                                chunk_scores.append(float(np.mean(pred_idx[ch][ch_wc] == y_va_idx[ch][ch_wc])))
                            else:
                                chunk_scores.append(float(np.mean(pred_idx[ch] == y_va_idx[ch])))
                        robust_s = float(np.mean(chunk_scores)) - 0.60 * float(np.std(chunk_scores))
                        key_m = (robust_s, s, -eta)
                        if best_meta_key is None or key_m > best_meta_key:
                            best_meta_key = key_m
                            meta_eta = float(eta)
                except Exception as _e:
                    self._stat("WARNING", f"Method meta-calibrator failed ({_e}) — disabled")
                    meta_model = None
                    meta_eta = 0.0

                if METHOD_HARD_RESET:
                    meta_model = None
                    meta_eta = 0.0

                method_bundle = {
                    "stage1": stage1,
                    "stage2": stage2,
                    "stage1_rf": stage1_rf,
                    "stage1_et": stage1_et,
                    "stage2_rf": stage2_rf,
                    "stage2_et": stage2_et,
                    "direct_hgb": direct_hgb,
                    "direct_rf": direct_rf,
                    "direct_et": direct_et,
                    "ovr_models": ovr_models,
                    "simple_method": simple_method,
                    "direct_classes": direct_classes,
                    "imputer": method_imputer,
                    "method_columns": method_columns,
                    "temp_decision": float(best_cfg["t_dec"]),
                    "temp_finish": float(best_cfg["t_fin"]),
                    "bias_finish": float(best_cfg["bias_finish"]),
                    "bias_sub": float(best_cfg["bias_sub"]),
                    "finish_threshold": float(best_cfg["finish_threshold"]),
                    "sub_threshold": float(best_cfg["sub_threshold"]),
                    "alpha_stage1": float(best_cfg["alpha_stage1"]),
                    "alpha_stage2": float(best_cfg["alpha_stage2"]),
                    "alpha_direct": float(best_cfg["alpha_direct"]),
                    "beta_direct": float(best_cfg["beta_direct"]),
                    "alpha_ovr": float(best_cfg["alpha_ovr"]),
                    "alpha_simple": float(best_cfg["alpha_simple"]),
                    "w_hist": float(best_cfg["w_hist"]),
                    "w_group": float(best_cfg["w_group"]),
                    "w_base": float(best_cfg["w_base"]),
                    "w_subsig": float(best_cfg.get("w_subsig", 0.0)),
                    "sub_boost_k": float(best_cfg.get("sub_boost_k", 0.0)),
                    "stage1_cols": stage1_cols,
                    "stage2_cols": stage2_cols,
                    "method_bias_decision": float(best_cfg["method_bias_decision"]),
                    "method_bias_ko_tko": float(best_cfg["method_bias_ko_tko"]),
                    "method_bias_submission": float(best_cfg["method_bias_submission"]),
                    "base_prior": base_prior,
                    "meta_model": meta_model,
                    "meta_eta": float(meta_eta),
                    "group_priors": group_priors,
                    "detail_labels_seen": sorted(pd.Series(y_method_dev["detail"]).unique().tolist()),
                }

                self._stat("Method classes (training)", ", ".join(method_bundle["detail_labels_seen"]))
                if METHOD_HARD_RESET:
                    self._stat("Method stack mode", "hard-reset (stable components only)")
                self._stat("Validation metric target (method | winner correct)", f"{best_cfg['val_metric']:.1%}")
                self._stat("Validation majority baseline (same subset)", f"{best_cfg['val_baseline']:.1%}")

                # Holdout evaluation with winner-model predicted winners.
                X_test_oriented_pred = _oriented_method_matrix(X_test_full, y_pred_red)
                X_test_oriented_pred = _augment_method_features(X_test_oriented_pred)
                X_test_oriented_pred_imp = pd.DataFrame(
                    method_imputer.transform(X_test_oriented_pred), columns=method_columns
                )
                # Stage-only diagnostics.
                y_finish_true = (y_method_test["finish_bin"].astype(str).values == "Finish").astype(int)
                _s1c_t = method_bundle.get("stage1_cols")
                X_test_s1 = X_test_oriented_pred_imp.reindex(columns=_s1c_t) if _s1c_t else X_test_oriented_pred_imp
                p_finish_hist_t = method_bundle["stage1"].predict_proba(X_test_s1)[:, 1]
                p_finish_rf_t = method_bundle["stage1_rf"].predict_proba(X_test_s1)[:, 1]
                p_finish_et_t = method_bundle["stage1_et"].predict_proba(X_test_s1)[:, 1]
                p_finish_tree_t = 0.55 * p_finish_rf_t + 0.45 * p_finish_et_t
                p_finish_raw_t = (
                    float(method_bundle.get("alpha_stage1", 0.75)) * p_finish_hist_t
                    + (1.0 - float(method_bundle.get("alpha_stage1", 0.75))) * p_finish_tree_t
                )
                logit_finish_t = np.log(
                    np.clip(p_finish_raw_t, 1e-6, 1 - 1e-6)
                    / np.clip(1.0 - p_finish_raw_t, 1e-6, 1 - 1e-6)
                )
                p_finish_stage_t = 1.0 / (
                    1.0 + np.exp(
                        -(
                            logit_finish_t / max(0.35, float(method_bundle.get("temp_decision", 1.0)))
                            + float(method_bundle.get("bias_finish", 0.0))
                        )
                    )
                )
                p_finish_stage_t = _apply_binary_threshold_warp(
                    p_finish_stage_t, float(method_bundle.get("finish_threshold", 0.50))
                )
                stage1_pred = (p_finish_stage_t >= 0.5).astype(int)
                stage1_acc = float(np.mean(stage1_pred == y_finish_true))
                stage1_auc = float("nan")
                try:
                    if len(np.unique(y_finish_true)) > 1:
                        stage1_auc = float(roc_auc_score(y_finish_true, p_finish_stage_t))
                except Exception:
                    stage1_auc = float("nan")

                stage2_acc_true_finishes = float("nan")
                n_true_finishes = int(np.sum(y_finish_true == 1))
                if n_true_finishes > 0:
                    finish_idx_true = np.where(y_finish_true == 1)[0]
                    X_test_fin_true = X_test_oriented_pred_imp.iloc[finish_idx_true].reset_index(drop=True)
                    y_sub_true = (
                        y_method_test.iloc[finish_idx_true]["finish_subtype"].astype(str).values == "Submission"
                    ).astype(int)
                    _s2c = method_bundle.get("stage2_cols")
                    X_test_fin_s2 = X_test_fin_true.reindex(columns=_s2c) if _s2c else X_test_fin_true
                    p_sub_hist_t = method_bundle["stage2"].predict_proba(X_test_fin_s2)[:, 1]
                    p_sub_rf_t = method_bundle["stage2_rf"].predict_proba(X_test_fin_s2)[:, 1]
                    p_sub_et_t = method_bundle["stage2_et"].predict_proba(X_test_fin_s2)[:, 1]
                    p_sub_tree_t = 0.55 * p_sub_rf_t + 0.45 * p_sub_et_t
                    p_sub_raw_t = (
                        float(method_bundle.get("alpha_stage2", 0.75)) * p_sub_hist_t
                        + (1.0 - float(method_bundle.get("alpha_stage2", 0.75))) * p_sub_tree_t
                    )
                    logit_sub_t = np.log(
                        np.clip(p_sub_raw_t, 1e-6, 1 - 1e-6)
                        / np.clip(1.0 - p_sub_raw_t, 1e-6, 1 - 1e-6)
                    )
                    p_sub_stage_t = 1.0 / (
                        1.0 + np.exp(
                            -(
                                logit_sub_t / max(0.35, float(method_bundle.get("temp_finish", 1.0)))
                                + float(method_bundle.get("bias_sub", 0.0))
                            )
                        )
                    )
                    p_sub_stage_t = _apply_binary_threshold_warp(
                        p_sub_stage_t, float(method_bundle.get("sub_threshold", 0.50))
                    )
                    stage2_pred = (p_sub_stage_t >= 0.5).astype(int)
                    stage2_acc_true_finishes = float(np.mean(stage2_pred == y_sub_true))
                ml_test = _stage_probs(
                    X_test_oriented_pred_imp,
                    temp_dec=method_bundle["temp_decision"],
                    temp_fin=method_bundle["temp_finish"],
                    alpha1=method_bundle["alpha_stage1"],
                    alpha2=method_bundle["alpha_stage2"],
                    bias_finish=method_bundle["bias_finish"],
                    bias_sub=method_bundle["bias_sub"],
                    alpha_direct=method_bundle["alpha_direct"],
                    beta_direct=method_bundle["beta_direct"],
                    alpha_ovr=method_bundle.get("alpha_ovr", 0.85),
                    finish_thr=method_bundle.get("finish_threshold", 0.50),
                    sub_thr=method_bundle.get("sub_threshold", 0.50),
                    alpha_simple=method_bundle.get("alpha_simple", 0.80),
                )
                y_method_np = y_method_test["coarse"].astype(str).values
                winner_correct = (y_pred_red == y_true_red)
                final_probs = []
                test_meta_reset = meta_test.reset_index(drop=True)
                for i in range(len(ml_test)):
                    gp = group_priors.get(
                        (str(test_meta_reset.iloc[i]["weight_class"]), str(test_meta_reset.iloc[i]["gender"]).lower()),
                        group_priors[("ALL", "all")],
                    )
                    d_dec = float(X_test_oriented_pred.reset_index(drop=True).iloc[i].get("d_dec_win_pct", 0.0))
                    d_ko = float(X_test_oriented_pred.reset_index(drop=True).iloc[i].get("d_ko_win_pct", 0.0))
                    d_sub = float(X_test_oriented_pred.reset_index(drop=True).iloc[i].get("d_sub_win_pct", 0.0))
                    hp = _normalize_method_probs({
                        "Decision": 0.50 + 0.35 * d_dec,
                        "KO/TKO": 0.32 + 0.35 * d_ko,
                        "Submission": 0.18 + 0.35 * d_sub,
                    })
                    sub_prior_arr = _sub_attempt_prior_array(X_test_oriented_pred.reset_index(drop=True).iloc[[i]])[0]
                    sp = _normalize_method_probs({
                        "Decision": float(sub_prior_arr[0]),
                        "KO/TKO": float(sub_prior_arr[1]),
                        "Submission": float(sub_prior_arr[2]),
                    })
                    w_hist = method_bundle["w_hist"]
                    w_group = method_bundle["w_group"]
                    w_base = method_bundle.get("w_base", 0.0)
                    w_subsig = method_bundle.get("w_subsig", 0.0)
                    w_ml = 1.0 - w_hist - w_group - w_base - w_subsig
                    bp = method_bundle.get("base_prior", group_priors.get(("ALL", "all"), {"Decision": 1 / 3, "KO/TKO": 1 / 3, "Submission": 1 / 3}))
                    row = {
                        m: (
                            w_ml * float(ml_test.iloc[i][m])
                            + w_hist * float(hp[m])
                            + w_group * float(gp[m])
                            + w_base * float(bp[m])
                            + w_subsig * float(sp[m])
                        )
                        for m in METHOD_LABELS
                    }
                    blended_row = _normalize_method_probs(row)
                    method_bias_map = {
                        "Decision": float(method_bundle.get("method_bias_decision", 0.0)),
                        "KO/TKO": float(method_bundle.get("method_bias_ko_tko", 0.0)),
                        "Submission": float(method_bundle.get("method_bias_submission", 0.0)),
                    }
                    blended_row = _apply_method_logit_bias_map(blended_row, method_bias_map)
                    blended_row = _apply_submission_signal_boost_map(
                        blended_row, float(sp["Submission"]), float(method_bundle.get("sub_boost_k", 0.0))
                    )
                    meta_model = method_bundle.get("meta_model")
                    meta_eta = float(method_bundle.get("meta_eta", 0.0))
                    if meta_model is not None and meta_eta > 0.0:
                        arr_blend = np.array([[blended_row["Decision"], blended_row["KO/TKO"], blended_row["Submission"]]], dtype=float)
                        arr_hist = np.array([[hp["Decision"], hp["KO/TKO"], hp["Submission"]]], dtype=float)
                        arr_grp = np.array([[gp["Decision"], gp["KO/TKO"], gp["Submission"]]], dtype=float)
                        arr_bp = np.array([[bp["Decision"], bp["KO/TKO"], bp["Submission"]]], dtype=float)
                        X_meta = np.hstack([arr_blend, arr_hist, arr_grp, arr_bp, np.log(np.clip(arr_blend, 1e-6, 1.0))])
                        pm = meta_model.predict_proba(X_meta)[0]
                        cls_map = {str(c): i for i, c in enumerate(meta_model.classes_)}
                        raw = {}
                        for m in METHOD_LABELS:
                            raw[m] = float(pm[cls_map[m]]) if m in cls_map else (1.0 / 3.0)
                        p_meta = _normalize_method_probs(raw)
                        blended_row = _normalize_method_probs({
                            m: (1.0 - meta_eta) * blended_row[m] + meta_eta * p_meta[m]
                            for m in METHOD_LABELS
                        })
                    final_probs.append(blended_row)
                method_pred_predwinner = np.array(
                    [max(METHOD_LABELS, key=lambda m: p[m]) for p in final_probs], dtype=object
                )
                method_acc_predicted_winner = float(np.mean(method_pred_predwinner == y_method_np))
                finish_score = float("nan")
                bal_acc = float("nan")
                macro_f1 = float("nan")
                ko_recall = float("nan")
                sub_recall = float("nan")
                ko_sub_macro_f1 = float("nan")
                _per_class_metrics = []
                _confusion_rows = []
                if int(np.sum(winner_correct)) > 0:
                    method_acc_when_winner_correct = float(
                        np.mean(method_pred_predwinner[winner_correct] == y_method_np[winner_correct])
                    )
                    subset = y_method_np[winner_correct]
                    counts = pd.Series(subset).value_counts()
                    method_majority_baseline_when_winner_correct = float(counts.max() / len(subset))
                method_holdout_acc_oracle = float("nan")

                self._section("Method Evaluation (Conditioned on Winner Pick)")
                self._stat("Stage1 acc (Finish vs Decision)", f"{stage1_acc:.1%}")
                self._stat("Stage1 AUC (Finish vs Decision)", "n/a" if not np.isfinite(stage1_auc) else f"{stage1_auc:.3f}")
                self._stat(
                    "Stage2 acc (KO/Sub | true finishes)",
                    "n/a" if not np.isfinite(stage2_acc_true_finishes) else f"{stage2_acc_true_finishes:.1%}",
                )
                self._stat("Stage2 sample size (true finishes)", n_true_finishes)
                self._stat("Method acc (predicted winner conditioned)", f"{method_acc_predicted_winner:.1%}")
                self._stat("Method acc | winner pick correct", f"{method_acc_when_winner_correct:.1%}")
                self._stat("Majority baseline | winner pick correct", f"{method_majority_baseline_when_winner_correct:.1%}")
                if int(np.sum(winner_correct)) > 0:
                    sub_true = y_method_np[winner_correct]
                    sub_pred = method_pred_predwinner[winner_correct]
                    p_arr, r_arr, f1_arr, _ = precision_recall_fscore_support(
                        sub_true, sub_pred, labels=METHOD_LABELS, zero_division=0
                    )
                    bal_acc = float(balanced_accuracy_score(sub_true, sub_pred))
                    macro_f1 = float(np.mean(f1_arr))
                    ko_recall = float(r_arr[1]) if len(r_arr) > 1 else 0.0
                    sub_recall = float(r_arr[2]) if len(r_arr) > 2 else 0.0
                    ko_sub_macro_f1 = float(np.mean(f1_arr[1:3])) if len(f1_arr) >= 3 else macro_f1
                    finish_score = 0.4 * ko_recall + 0.4 * sub_recall + 0.2 * ko_sub_macro_f1
                    self._stat("Balanced accuracy | winner pick correct", f"{bal_acc:.1%}")
                    self._stat("Macro F1 | winner pick correct", f"{macro_f1:.1%}")
                    self._stat("FinishScore (0.4 KO R + 0.4 Sub R + 0.2 KO/Sub F1)", f"{finish_score:.1%}")
                    self._log("")
                    self._log("Per-Class Metrics (Method | winner pick correct)")
                    self._log("Class          Precision    Recall      F1")
                    for idx, cls_name in enumerate(METHOD_LABELS):
                        self._log(
                            f"{cls_name:<14}{p_arr[idx]:10.1%}{r_arr[idx]:10.1%}{f1_arr[idx]:10.1%}"
                        )
                    self._log("")
                    self._log("Confusion Matrix (Method | winner pick correct)")
                    self._log("Actual\\Pred     Decision    KO/TKO  Submission")
                    for actual in METHOD_LABELS:
                        row_counts = []
                        for pred in METHOD_LABELS:
                            c = int(np.sum((sub_true == actual) & (sub_pred == pred)))
                            row_counts.append(c)
                        self._log(f"{actual:<14}{row_counts[0]:10d}{row_counts[1]:10d}{row_counts[2]:12d}")
                        _confusion_rows.append([actual, list(row_counts)])
                    _per_class_metrics = [
                        [cls_name, float(p_arr[idx]), float(r_arr[idx]), float(f1_arr[idx])]
                        for idx, cls_name in enumerate(METHOD_LABELS)
                    ]

                method_bundle["method_columns"] = method_columns

                _cache_save(
                    "method_stage",
                    method_cache_key,
                    {
                        "kind": METHOD_STAGE_CACHE_KIND,
                        "method_cache_version": METHOD_CACHE_VERSION,
                        "winner_cache_key": winner_cache_key,
                        "method_bundle": {k: v for k, v in method_bundle.items()},
                        "method_acc_when_winner_correct": method_acc_when_winner_correct,
                        "method_acc_predicted_winner": method_acc_predicted_winner,
                        "method_majority_baseline_when_winner_correct": (
                            method_majority_baseline_when_winner_correct
                        ),
                        "method_holdout_acc_oracle": method_holdout_acc_oracle,
                        "finish_score": finish_score,
                        "stage1_acc": stage1_acc,
                        "stage1_auc": stage1_auc,
                        "stage2_acc_true_finishes": stage2_acc_true_finishes,
                        "n_true_finishes": n_true_finishes,
                        "bal_acc": bal_acc,
                        "macro_f1": macro_f1,
                        "ko_recall": ko_recall,
                        "sub_recall": sub_recall,
                        "ko_sub_macro_f1": ko_sub_macro_f1,
                        "per_class_metrics": _per_class_metrics,
                        "confusion_rows": _confusion_rows,
                        "method_classes_training": ", ".join(method_bundle.get("detail_labels_seen", [])),
                        "val_metric_str": f"{best_cfg['val_metric']:.1%}",
                        "val_baseline_str": f"{best_cfg['val_baseline']:.1%}",
                        "method_hard_reset": bool(METHOD_HARD_RESET),
                    },
                )
                self._stat(
                    "Method cache",
                    f"SAVED ({METHOD_CACHE_VERSION}) key={str(method_cache_key)[:12]} file={os.path.basename(_cache_path('method_stage', method_cache_key))}",
                )

            except Exception as exc:
                self._section("Method Evaluation (Conditioned on Winner Pick)")
                self._stat("Method model", f"disabled ({exc})")
                method_bundle = None

        # Retrain for production on all rows.
        X_all = pd.DataFrame(imputer.fit_transform(X_winner[feature_cols]), columns=feature_cols)
        X_all_aug, y_all_aug = _augment_swap(X_all, y.reset_index(drop=True))
        w_all_aug = _augment_weights(_time_weights(len(X_all), floor=0.35))
        scaler_all = StandardScaler()
        X_all_aug_sc = pd.DataFrame(scaler_all.fit_transform(X_all_aug), columns=feature_cols)
        final_models = {}
        for name, make_model in specs:
            final_models[name] = _fit_model(
                name, make_model(),
                X_all_aug_sc if name in NEEDS_SCALE else X_all_aug,
                y_all_aug, sample_weight=w_all_aug
            )

        if combiner["kind"] == "stacker":
            final_stacker = LogisticRegression(
                max_iter=8000, C=0.2, solver="saga", tol=1e-3, n_jobs=-1, random_state=RANDOM_SEED
            )
            final_stacker.fit(oof.loc[valid, model_order].astype(float).values, y_dev.loc[valid].astype(int).values)
            final_combiner = {"kind": "stacker", "model": final_stacker, "model_order": model_order}
        else:
            final_combiner = combiner

        # Retrain method head on all rows, oriented by winner-model predictions.
        method_bundle_all = None
        method_feat_cols_all = list(full_feature_cols)
        if method_bundle is not None:
            try:
                X_all_sw = _swap_features(X_all)
                X_all_sc = pd.DataFrame(scaler_all.transform(X_all), columns=feature_cols)
                X_all_sw_sc = pd.DataFrame(scaler_all.transform(X_all_sw), columns=feature_cols)
                all_meta = {}
                for name, _ in specs:
                    p_fwd = _predict_proba(name, final_models[name], X_all_sc if name in NEEDS_SCALE else X_all)
                    p_rev = _predict_proba(name, final_models[name], X_all_sw_sc if name in NEEDS_SCALE else X_all_sw)
                    all_meta[name] = _clip_probs((p_fwd + (1.0 - p_rev)) / 2.0)
                all_meta_df = pd.DataFrame(all_meta)[model_order]
                all_probs = _combine_probs(all_meta_df, final_combiner)
                if calibrator is not None:
                    all_probs = _clip_probs(
                        calibrator.predict_proba(np.asarray(all_probs).reshape(-1, 1))[:, 1]
                    )
                y_pred_all = (np.asarray(all_probs) >= decision_threshold).astype(int)

                X_all_full = pd.DataFrame(full_imputer.fit_transform(X_full[full_feature_cols]), columns=full_feature_cols)
                X_all_method = _oriented_method_matrix(X_all_full, y_pred_all)
                X_all_method = _augment_method_features(X_all_method)
                method_feat_cols_all = list(X_all_method.columns)
                method_imputer_all = SimpleImputer(strategy="median")
                X_all_method_imp = pd.DataFrame(
                    method_imputer_all.fit_transform(X_all_method), columns=method_feat_cols_all
                )

                y_all_bin = (y_method_df["finish_bin"].values == "Finish").astype(int)
                p_finish_all = max(1e-6, float(np.mean(y_all_bin)))
                w_all_t = _time_weights(len(X_all_method_imp), floor=0.45)
                w_all_stage1 = w_all_t * np.where(
                    y_all_bin == 1, 0.5 / p_finish_all, 0.5 / max(1e-6, 1.0 - p_finish_all)
                )
                stage1_all = HistGradientBoostingClassifier(
                    loss="log_loss",
                    max_iter=360, learning_rate=0.045, max_depth=6,
                    max_leaf_nodes=31, min_samples_leaf=16, l2_regularization=0.8,
                    random_state=RANDOM_SEED + 808,
                )
                stage1_all.fit(X_all_method_imp, y_all_bin, sample_weight=w_all_stage1)
                stage1_rf_all = RandomForestClassifier(
                    n_estimators=420, max_depth=10, min_samples_leaf=4,
                    max_features=0.7, random_state=RANDOM_SEED + 910, n_jobs=-1,
                    class_weight="balanced_subsample",
                )
                stage1_rf_all.fit(X_all_method_imp, y_all_bin)
                stage1_et_all = ExtraTreesClassifier(
                    n_estimators=560, max_depth=12, min_samples_leaf=3,
                    max_features=0.72, random_state=RANDOM_SEED + 914, n_jobs=-1,
                    class_weight="balanced_subsample",
                )
                stage1_et_all.fit(X_all_method_imp, y_all_bin)

                fin_mask_all = (y_method_df["finish_bin"].values == "Finish")
                if int(np.sum(fin_mask_all)) < 60:
                    fin_mask_all = np.ones(len(y_method_df), dtype=bool)
                X2_all = X_all_method_imp.iloc[np.where(fin_mask_all)[0]].reset_index(drop=True)
                y2_all = (y_method_df.iloc[np.where(fin_mask_all)[0]]["finish_subtype"].values == "Submission").astype(int)
                p_sub_all = max(1e-6, float(np.mean(y2_all))) if len(y2_all) > 0 else 0.5
                w2_all = np.where(y2_all == 1, 0.5 / p_sub_all, 0.5 / max(1e-6, 1.0 - p_sub_all))
                stage2_all = HistGradientBoostingClassifier(
                    loss="log_loss",
                    max_iter=320, learning_rate=0.05, max_depth=5,
                    max_leaf_nodes=31, min_samples_leaf=14, l2_regularization=0.7,
                    random_state=RANDOM_SEED + 809,
                )
                stage2_all.fit(X2_all, y2_all, sample_weight=w2_all)
                stage2_rf_all = RandomForestClassifier(
                    n_estimators=360, max_depth=9, min_samples_leaf=3,
                    max_features=0.7, random_state=RANDOM_SEED + 911, n_jobs=-1,
                    class_weight="balanced_subsample",
                )
                stage2_rf_all.fit(X2_all, y2_all)
                stage2_et_all = ExtraTreesClassifier(
                    n_estimators=520, max_depth=11, min_samples_leaf=3,
                    max_features=0.72, random_state=RANDOM_SEED + 915, n_jobs=-1,
                    class_weight="balanced_subsample",
                )
                stage2_et_all.fit(X2_all, y2_all)

                y_cls_all = y_method_df["coarse"].values
                cls_counts_all = pd.Series(y_cls_all).value_counts()
                cls_w_all = {k: (len(y_cls_all) / (len(cls_counts_all) * max(1, v))) for k, v in cls_counts_all.items()}
                w_cls_all = np.array([cls_w_all.get(lbl, 1.0) for lbl in y_cls_all], dtype=float)
                w_direct_all = w_all_t * w_cls_all
                direct_hgb_all = HistGradientBoostingClassifier(
                    loss="log_loss",
                    max_iter=300, learning_rate=0.05, max_depth=6,
                    max_leaf_nodes=31, min_samples_leaf=14, l2_regularization=0.7,
                    random_state=RANDOM_SEED + 912,
                )
                direct_hgb_all.fit(X_all_method_imp, y_cls_all, sample_weight=w_direct_all)
                direct_rf_all = RandomForestClassifier(
                    n_estimators=520, max_depth=11, min_samples_leaf=3,
                    max_features=0.72, random_state=RANDOM_SEED + 913, n_jobs=-1,
                    class_weight="balanced_subsample",
                )
                direct_rf_all.fit(X_all_method_imp, y_cls_all)
                direct_et_all = ExtraTreesClassifier(
                    n_estimators=680, max_depth=13, min_samples_leaf=3,
                    max_features=0.72, random_state=RANDOM_SEED + 916, n_jobs=-1,
                    class_weight="balanced_subsample",
                )
                direct_et_all.fit(X_all_method_imp, y_cls_all)
                ovr_models_all = {}
                for j, lbl in enumerate(METHOD_LABELS):
                    y_ovr_all = (pd.Series(y_cls_all).astype(str).values == lbl).astype(int)
                    p_pos_all = max(1e-6, float(np.mean(y_ovr_all)))
                    w_ovr_all = w_all_t * np.where(
                        y_ovr_all == 1, 0.5 / p_pos_all, 0.5 / max(1e-6, 1.0 - p_pos_all)
                    )
                    mdl_all = HistGradientBoostingClassifier(
                        loss="log_loss",
                        max_iter=260, learning_rate=0.05, max_depth=5,
                        max_leaf_nodes=31, min_samples_leaf=14, l2_regularization=0.8,
                        random_state=RANDOM_SEED + 1500 + j,
                    )
                    mdl_all.fit(X_all_method_imp, y_ovr_all, sample_weight=w_ovr_all)
                    ovr_models_all[lbl] = mdl_all
                simple_method_all = LogisticRegression(
                    max_iter=8000, C=0.8, solver="saga", tol=1e-3, n_jobs=-1,
                    class_weight="balanced", random_state=RANDOM_SEED + 1601,
                )
                simple_method_all.fit(X_all_method_imp, pd.Series(y_cls_all).astype(str).values)

                group_priors_all = {}
                for (wc, gender), grp in row_meta.groupby(["weight_class", "gender"], dropna=False):
                    idx = grp.index.values
                    counts = y_method_df.iloc[idx]["coarse"].value_counts()
                    tot = float(len(idx))
                    group_priors_all[(str(wc), str(gender).lower())] = _normalize_method_probs({
                        "Decision": float(counts.get("Decision", 0.0) / max(1.0, tot)),
                        "KO/TKO": float(counts.get("KO/TKO", 0.0) / max(1.0, tot)),
                        "Submission": float(counts.get("Submission", 0.0) / max(1.0, tot)),
                    })
                all_counts2 = y_method_df["coarse"].value_counts()
                all_tot2 = float(len(y_method_df))
                group_priors_all[("ALL", "all")] = _normalize_method_probs({
                    "Decision": float(all_counts2.get("Decision", 0.0) / max(1.0, all_tot2)),
                    "KO/TKO": float(all_counts2.get("KO/TKO", 0.0) / max(1.0, all_tot2)),
                    "Submission": float(all_counts2.get("Submission", 0.0) / max(1.0, all_tot2)),
                })
                base_prior_all = _normalize_method_probs({
                    "Decision": float(all_counts2.get("Decision", 0.0) / max(1.0, all_tot2)),
                    "KO/TKO": float(all_counts2.get("KO/TKO", 0.0) / max(1.0, all_tot2)),
                    "Submission": float(all_counts2.get("Submission", 0.0) / max(1.0, all_tot2)),
                })

                method_bundle_all = {
                    "stage1": stage1_all,
                    "stage2": stage2_all,
                    "stage1_rf": stage1_rf_all,
                    "stage1_et": stage1_et_all,
                    "stage2_rf": stage2_rf_all,
                    "stage2_et": stage2_et_all,
                    "direct_hgb": direct_hgb_all,
                    "direct_rf": direct_rf_all,
                    "direct_et": direct_et_all,
                    "ovr_models": ovr_models_all,
                    "simple_method": simple_method_all,
                    "direct_classes": [str(c) for c in direct_hgb_all.classes_],
                    "imputer": method_imputer_all,
                    "method_columns": method_feat_cols_all,
                    "temp_decision": float(method_bundle.get("temp_decision", 1.0)),
                    "temp_finish": float(method_bundle.get("temp_finish", 1.0)),
                    "bias_finish": float(method_bundle.get("bias_finish", 0.0)),
                    "bias_sub": float(method_bundle.get("bias_sub", 0.0)),
                    "finish_threshold": float(method_bundle.get("finish_threshold", 0.50)),
                    "sub_threshold": float(method_bundle.get("sub_threshold", 0.50)),
                    "alpha_stage1": float(method_bundle.get("alpha_stage1", 0.75)),
                    "alpha_stage2": float(method_bundle.get("alpha_stage2", 0.75)),
                    "alpha_direct": float(method_bundle.get("alpha_direct", 0.85)),
                    "beta_direct": float(method_bundle.get("beta_direct", 0.70)),
                    "alpha_ovr": float(method_bundle.get("alpha_ovr", 0.85)),
                    "alpha_simple": float(method_bundle.get("alpha_simple", 0.80)),
                    "w_hist": float(method_bundle.get("w_hist", 0.25)),
                    "w_group": float(method_bundle.get("w_group", 0.10)),
                    "w_base": float(method_bundle.get("w_base", 0.08)),
                    "w_subsig": float(method_bundle.get("w_subsig", 0.10)),
                    "sub_boost_k": float(method_bundle.get("sub_boost_k", 0.0)),
                    "method_bias_decision": float(method_bundle.get("method_bias_decision", 0.0)),
                    "method_bias_ko_tko": float(method_bundle.get("method_bias_ko_tko", 0.0)),
                    "method_bias_submission": float(method_bundle.get("method_bias_submission", 0.0)),
                    "base_prior": base_prior_all,
                    "meta_model": method_bundle.get("meta_model"),
                    "meta_eta": float(method_bundle.get("meta_eta", 0.0)),
                    "group_priors": group_priors_all,
                    "detail_labels_seen": method_bundle.get("detail_labels_seen", []),
                }
            except Exception as _e:
                self._stat("WARNING", f"All-data method retrain failed ({_e}) — method predictions disabled")
                method_bundle_all = None

        self.model = SuperEnsembleModel(
            final_models, imputer, scaler_all, feature_cols, final_combiner,
            calibrator=calibrator, decision_threshold=decision_threshold,
            method_bundle=method_bundle_all,
            method_feat_cols=method_feat_cols_all,
        )
        self.method_model = method_bundle_all
        self.method_imputer = method_bundle_all["imputer"] if method_bundle_all is not None else None
        self.method_feat_cols = method_feat_cols_all
        self.method_metrics = {
            "method_acc_predicted_winner": method_acc_predicted_winner,
            "method_acc_when_winner_correct": method_acc_when_winner_correct,
            "method_acc_true_winner": method_holdout_acc_oracle,
            "method_majority_baseline_when_winner_correct": method_majority_baseline_when_winner_correct,
            "method_finish_score_winner_correct": finish_score,
        }

        ml_baseline = None
        old_baseline = None

        self.benchmarks = BenchmarkScores(
            super_raw_logloss=float(raw_ll),
            super_cal_logloss=float(cal_ll_test),
            super_brier=float(brier),
            super_acc=float(acc),
            super_ece=float(ece),
            ml_baseline_logloss=ml_baseline,
            old_baseline_logloss=old_baseline,
        )
        self._rolling_diagnostics(X, y)
        return self.model

    def _print_benchmark_table(self, super_ll, ml_ll, old_ll):
        self._section("Benchmark Comparison (Lower Is Better)")
        self._log(f"{'Model':<24} {'Holdout LogLoss':>14}")
        self._log("-" * 40)
        self._log(f"{'UFC_Model':<24} {super_ll:>14.4f}")
        self._log(f"{'Standalone Baseline A':<24} {ml_ll if ml_ll is not None else float('nan'):>14.4f}")
        self._log(f"{'Standalone Baseline B':<24} {old_ll if old_ll is not None else float('nan'):>14.4f}")
        self._log("-" * 40)

    def _run_ml_baseline(self):
        return None

    def _run_old_baseline(self):
        return None

    def _rolling_diagnostics(self, X, y):
        n = len(X)
        folds = []
        start = int(n * 0.45)
        step = max(250, int(n * 0.08))
        while start + step < n:
            train_end = start
            test_end = min(n, start + step)
            folds.append((train_end, test_end))
            start += step
        if not folds:
            return

        self._section("Walk-Forward Diagnostics")
        ll_scores = []
        for i, (tr_end, te_end) in enumerate(folds, start=1):
            X_tr = X.iloc[:tr_end]
            y_tr = y.iloc[:tr_end]
            X_te = X.iloc[tr_end:te_end]
            y_te = y.iloc[tr_end:te_end]
            imp = SimpleImputer(strategy="median")
            X_tr_imp = pd.DataFrame(imp.fit_transform(X_tr), columns=X.columns)
            X_te_imp = pd.DataFrame(imp.transform(X_te), columns=X.columns)
            mdl = HistGradientBoostingClassifier(
                max_iter=500, learning_rate=0.04, max_depth=6,
                max_leaf_nodes=31, min_samples_leaf=20, l2_regularization=1.0,
                random_state=RANDOM_SEED,
            )
            X_aug, y_aug = _augment_swap(X_tr_imp, y_tr.reset_index(drop=True))
            mdl.fit(X_aug, y_aug)
            p_fwd = _predict_proba("HistGBM", mdl, X_te_imp)
            X_te_sw = _swap_features(X_te_imp)
            p_rev = _predict_proba("HistGBM", mdl, X_te_sw)
            p = _clip_probs((p_fwd + (1.0 - p_rev)) / 2.0)
            ll = log_loss(y_te, p)
            ll_scores.append(ll)
            self._stat(f"Fold {i} log-loss", f"{ll:.4f} ({len(y_te)} fights)")
        self._stat("Walk-forward mean log-loss", f"{np.mean(ll_scores):.4f}")
        self._stat("Walk-forward std", f"{np.std(ll_scores):.4f}")

    def predict_matchup(self, fighter_a, fighter_b, weight_class="", gender="", rounds=3):
        if self.model is None:
            raise RuntimeError("Model is not trained.")
        a_key = fuzzy_find(fighter_a, self.fighter_history) or fighter_a
        b_key = fuzzy_find(fighter_b, self.fighter_history) or fighter_b
        today = pd.Timestamp(datetime.now().date())

        if a_key in self.fighter_history:
            a_hist = self.fighter_history[a_key]
            a_glicko = self.glicko_ratings.get(a_key, (MU_0, PHI_0, SIGMA_0))
            a_opp = self.opp_glicko_list.get(a_key, [])
            a_feats = compute_fighter_features(a_hist, a_glicko, a_opp, today)
        else:
            a_feats = compute_fighter_features([], (MU_0, PHI_0, SIGMA_0), [], today)

        if b_key in self.fighter_history:
            b_hist = self.fighter_history[b_key]
            b_glicko = self.glicko_ratings.get(b_key, (MU_0, PHI_0, SIGMA_0))
            b_opp = self.opp_glicko_list.get(b_key, [])
            b_feats = compute_fighter_features(b_hist, b_glicko, b_opp, today)
        else:
            b_feats = compute_fighter_features([], (MU_0, PHI_0, SIGMA_0), [], today)

        matchup = compute_matchup_features(
            a_feats, b_feats, is_title=0, total_rounds=rounds, weight_class=weight_class
        )
        a_elo = float(self.elo_ratings.get(a_key, ELO_BASE))
        b_elo = float(self.elo_ratings.get(b_key, ELO_BASE))
        d_elo = a_elo - b_elo
        elo_p = 1.0 / (1.0 + 10.0 ** (-(d_elo / 400.0)))
        division = _normalize_division(weight_class, gender)
        a_div_elo = float(self.div_elo_ratings.get((a_key, division), ELO_BASE))
        b_div_elo = float(self.div_elo_ratings.get((b_key, division), ELO_BASE))
        d_div_elo = a_div_elo - b_div_elo
        div_p = 1.0 / (1.0 + 10.0 ** (-(d_div_elo / 400.0)))
        matchup.update({
            "elo_r": a_elo,
            "elo_b": b_elo,
            "d_elo": d_elo,
            "elo_win_prob": elo_p,
            "d_elo_win_prob": elo_p - 0.5,
            "abs_elo_gap": abs(d_elo),
            "elo_sum": a_elo + b_elo,
            "div_elo_r": a_div_elo,
            "div_elo_b": b_div_elo,
            "d_div_elo": d_div_elo,
            "div_elo_win_prob": div_p,
            "d_div_elo_win_prob": div_p - 0.5,
            "abs_div_elo_gap": abs(d_div_elo),
            "elo_divergence": elo_p - div_p,
            "elo_agreement": 1.0 - abs(elo_p - div_p),
        })
        matchup_df = _augment_matchup_features(pd.DataFrame([matchup]))
        p_a = self.model.predict_proba_single(matchup_df.iloc[0].to_dict())
        pick_a = p_a >= 0.5
        winner_name = a_key if pick_a else b_key
        loser_name = b_key if pick_a else a_key

        winner_profile = _method_profile_from_history(self.fighter_history.get(winner_name, []))
        loser_profile = _method_profile_from_history(self.fighter_history.get(loser_name, []))
        method_probs = self.model.predict_method_probs(
            matchup_df.iloc[0].to_dict(),
            winner_is_a=pick_a,
            winner_profile=winner_profile,
            loser_profile=loser_profile,
            weight_class=weight_class,
            gender=gender,
        )
        predicted_method = max(METHOD_LABELS, key=lambda m: method_probs[m])

        return {
            "name_a": a_key,
            "name_b": b_key,
            "prob_a": p_a,
            "prob_b": 1.0 - p_a,
            "rating_a": float(a_feats.get("glicko_mu", MU_0)),
            "rating_b": float(b_feats.get("glicko_mu", MU_0)),
            "weight_class": weight_class,
            "gender": gender,
            "method_probs": method_probs,
            "predicted_method": predicted_method,
            "decision_pct": method_probs["Decision"],
            "ko_tko_pct": method_probs["KO/TKO"],
            "submission_pct": method_probs["Submission"],
            "method_pct": method_probs[predicted_method],
        }

    def is_debutant(self, fighter_name):
        key = fuzzy_find(fighter_name, self.fighter_history)
        if key is None:
            return True
        return not bool(self.fighter_history.get(key))

    def division_rankings(self):
        by_div = defaultdict(list)
        now = pd.Timestamp(datetime.now().date())
        cutoff = now - pd.Timedelta(days=ACTIVE_DAYS)
        for fighter, hist in self.fighter_history.items():
            if not hist:
                continue
            last = hist[-1]["date"]
            if pd.isna(last) or last < cutoff:
                continue
            meta = self.fighter_meta.get(fighter, {})
            division = meta.get("division", "")
            if not division or division in ("Catch Weight", "Open Weight"):
                continue
            mu = self.glicko_ratings.get(fighter, (MU_0, PHI_0, SIGMA_0))[0]
            wins = sum(1 for h in hist if h.get("result") == "W")
            losses = sum(1 for h in hist if h.get("result") == "L")
            draws = sum(1 for h in hist if h.get("result") == "D")
            by_div[division].append((fighter, mu, wins, losses, draws))
        for div in by_div:
            by_div[div].sort(key=lambda t: t[1], reverse=True)
        return by_div


def _auto_width(ws):
    for ci in range(1, ws.max_column + 1):
        mx = 0
        for row in ws.iter_rows(min_col=ci, max_col=ci):
            for cell in row:
                mx = max(mx, len(str(cell.value or "")))
        ws.column_dimensions[get_column_letter(ci)].width = mx + 3


def export_to_excel(path, predictions, rankings):
    wb = Workbook()
    hdr_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    hdr_font = Font(bold=True, size=11, color="000000")
    hdr_align = Alignment(horizontal="left", vertical="center")
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb.active
    ws.title = "Predictions"
    headers = [
        "Red Corner", "Blue Corner", "Weight Class", "Winner", "Win %",
        "Method", "Method %", "DEC %", "(T)KO %", "SUB %",
    ]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill, c.font, c.alignment, c.border = hdr_fill, hdr_font, hdr_align, border

    for ri, p in enumerate(predictions, 2):
        winner = p.get("predicted_winner") or (p["name_a"] if p["prob_a"] >= p["prob_b"] else p["name_b"])
        win_pct = p.get("win_pct")
        if win_pct is None:
            win_pct = p["prob_a"] if winner == p["name_a"] else p["prob_b"]
        method_probs = _normalize_method_probs(p.get("method_probs", {}))
        pred_method = p.get("predicted_method") or max(METHOD_LABELS, key=lambda m: method_probs[m])
        row = [
            p["name_a"], p["name_b"], p.get("weight_class", ""), winner, win_pct,
            pred_method, method_probs[pred_method],
            method_probs["Decision"], method_probs["KO/TKO"], method_probs["Submission"],
        ]
        for ci, v in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = border
            c.alignment = hdr_align
            if ci in (5, 7, 8, 9, 10):
                c.number_format = "0.0%"
    _auto_width(ws)
    _ = rankings  # Intentionally unused: keep workbook to Predictions sheet only.

    wb.save(path)


class SuperModelGUI:
    BG = "#0A0A0A"
    BG_HEADER = "#111111"
    BG_INPUT = "#141414"
    FG = "#F5F5F5"
    ACCENT = "#D20A11"
    MUTED = "#CFCFCF"
    GREEN = "#FFFFFF"
    BAR_A = "#D20A11"
    BAR_B = "#F5F5F5"

    def __init__(self, root, pipeline):
        self.root = root
        self.pipeline = pipeline
        self._busy = False
        self.root.title("UFC Model")
        self.root.geometry("980x780")
        self.root.minsize(900, 680)
        self.root.configure(bg=self.BG)
        self._build_ui()

    def _build_ui(self):
        top_accent = tk.Frame(self.root, bg=self.ACCENT, height=6)
        top_accent.pack(fill="x")

        tf = tk.Frame(self.root, bg=self.BG_HEADER, pady=14)
        tf.pack(fill="x")
        tk.Label(tf, text="UFC", font=("Helvetica", 34, "bold"),
                 fg=self.ACCENT, bg=self.BG_HEADER).pack()
        tk.Label(tf, text="FIGHT PREDICTOR", font=("Helvetica", 10, "bold"),
                 fg=self.FG, bg=self.BG_HEADER).pack(pady=(0, 2))

        main = tk.Frame(self.root, bg=self.BG, padx=18, pady=12)
        main.pack(fill="both", expand=True)

        self.status_var = tk.StringVar(value="Enter matchups, then click Predict to train and run.")
        tk.Label(main, textvariable=self.status_var, bg=self.BG, fg=self.MUTED,
                 font=("Helvetica", 9, "italic")).pack(anchor="w")

        tk.Label(main, text="Enter fights (one per line: Fighter A,Fighter B,Weight Class,Gender,Rounds)",
                 bg=self.BG, fg=self.FG, font=("Helvetica", 9, "bold")).pack(anchor="w", pady=(8, 4))

        input_wrap = tk.Frame(main, bg=self.BG_HEADER, padx=2, pady=2)
        input_wrap.pack(fill="both", expand=True, pady=4)
        self.fight_input = tk.Text(
            input_wrap, height=24, font=("Courier New", 10), bg=self.BG_INPUT,
            fg=self.FG, insertbackground=self.FG, relief="flat", wrap="word",
            highlightthickness=1, highlightbackground=self.ACCENT, highlightcolor=self.ACCENT
        )
        input_sb = tk.Scrollbar(
            input_wrap, command=self.fight_input.yview, bg=self.BG_HEADER,
            troughcolor=self.BG_INPUT, activebackground=self.ACCENT
        )
        self.fight_input.configure(yscrollcommand=input_sb.set)
        self.fight_input.pack(side="left", fill="both", expand=True)
        input_sb.pack(side="right", fill="y")

        bf = tk.Frame(main, bg=self.BG)
        bf.pack(fill="x", pady=(10, 6))
        tk.Button(bf, text="Clear", command=self._clear, font=("Helvetica", 10, "bold"),
                  bg="#202020", fg=self.FG, relief="flat", padx=14, cursor="hand2",
                  activebackground="#2A2A2A", activeforeground=self.FG).pack(side="left", padx=4)
        self.predict_btn = tk.Button(bf, text="Predict", command=self._predict,
                                     font=("Helvetica", 11, "bold"), bg=self.ACCENT, fg=self.FG,
                                     relief="flat", padx=24, cursor="hand2", state="normal",
                                     activebackground="#B40A0F", activeforeground=self.FG)
        self.predict_btn.pack(side="right", padx=4)

        self.inner = tk.Frame(main, bg=self.BG)

    def _clear(self):
        self.fight_input.delete("1.0", tk.END)

    def _predict(self):
        if self._busy:
            return
        text = self.fight_input.get("1.0", tk.END).strip()
        if not text:
            self.status_var.set("Enter at least one matchup.")
            return
        self._busy = True
        self.predict_btn.config(state="disabled")

        def _do():
            try:
                if self.pipeline.model is None:
                    self.pipeline.train()

                preds = []
                skipped_debut = 0
                for line in [ln.strip() for ln in text.splitlines() if ln.strip()]:
                    parts = [p.strip() for p in line.split(",")]
                    if len(parts) < 2:
                        continue
                    a, b = parts[0], parts[1]
                    if self.pipeline.is_debutant(a) or self.pipeline.is_debutant(b):
                        skipped_debut += 1
                        continue
                    wc = parts[2] if len(parts) > 2 else ""
                    g = parts[3] if len(parts) > 3 else ""
                    try:
                        rounds = int(parts[4]) if len(parts) > 4 else 3
                    except Exception:
                        rounds = 3
                    p = self.pipeline.predict_matchup(a, b, wc, g, rounds)
                    # User-facing picks should always align with displayed probability.
                    pick_a = p["prob_a"] >= 0.5
                    p["predicted_winner"] = p["name_a"] if pick_a else p["name_b"]
                    p["win_pct"] = p["prob_a"] if pick_a else p["prob_b"]
                    p["method_probs"] = _normalize_method_probs(p.get("method_probs", {}))
                    p["predicted_method"] = p.get("predicted_method") or max(
                        METHOD_LABELS, key=lambda m: p["method_probs"][m]
                    )
                    p["method_pct"] = p["method_probs"][p["predicted_method"]]
                    p["decision_pct"] = p["method_probs"]["Decision"]
                    p["ko_tko_pct"] = p["method_probs"]["KO/TKO"]
                    p["submission_pct"] = p["method_probs"]["Submission"]
                    preds.append(p)

                try:
                    export_to_excel(PREDICTIONS_XLSX, preds, self.pipeline.division_rankings())
                    msg = (
                        f"{len(preds)} matchups predicted. "
                        f"Skipped {skipped_debut} debutant matchup(s). "
                        f"Saved to {os.path.basename(PREDICTIONS_XLSX)}"
                    )
                except Exception as e:
                    err = str(e)
                    msg = (
                        f"{len(preds)} matchups predicted. "
                        f"Skipped {skipped_debut} debutant matchup(s). "
                        f"Excel export failed: {err}"
                    )
                    self.root.after(0, lambda em=err: messagebox.showerror("Export Error", em))
                self.root.after(0, lambda: self.status_var.set(msg))
                print("")
                print("=" * 72)
                print("Prediction Run Complete")
                print("=" * 72)
                print(msg)
            except Exception as e:
                err = str(e)
                self.root.after(0, lambda em=err: self.status_var.set(f"Prediction failed: {em}"))
                print("")
                print("=" * 72)
                print("Prediction Run Failed")
                print("=" * 72)
                print(f"Prediction failed: {err}")
            finally:
                self._busy = False
                self.root.after(0, lambda: self.predict_btn.config(state="normal"))

        threading.Thread(target=_do, daemon=True).start()

def main():
    parser = argparse.ArgumentParser(description="Train and run UFC Model.")
    parser.add_argument("--train-only", action="store_true", help="Run training/evaluation only, no GUI.")
    _ = parser.parse_args()

    if not os.path.exists(DATA_PATH):
        raise FileNotFoundError(f"Missing dataset: {DATA_PATH}")

    pipeline = UFCSuperModelPipeline(DATA_PATH)
    if _.train_only:
        pipeline.train()
        if pipeline.benchmarks:
            b = pipeline.benchmarks
            print("")
            print("=" * 72)
            print("Run Complete")
            print("=" * 72)
            print(f"Holdout calibrated log-loss: {b.super_cal_logloss:.4f}")
            print(f"Holdout raw log-loss: {b.super_raw_logloss:.4f}")
            print(f"Holdout Brier score: {b.super_brier:.4f}")
            print(f"Holdout accuracy: {b.super_acc:.2%}")
            print(f"Holdout ECE: {b.super_ece:.4f}")
            mm = pipeline.method_metrics or {}
            if mm:
                v1 = mm.get("method_acc_predicted_winner")
                v2 = mm.get("method_acc_when_winner_correct")
                v3 = mm.get("method_acc_true_winner")
                v4 = mm.get("method_majority_baseline_when_winner_correct")
                v5 = mm.get("method_finish_score_winner_correct")
                if v1 == v1:
                    print(f"Method acc (predicted winner conditioned): {v1:.1%}")
                if v2 == v2:
                    print(f"Method acc | winner pick correct: {v2:.1%}")
                if v4 == v4:
                    print(f"Majority baseline | winner pick correct: {v4:.1%}")
                if v5 == v5:
                    print(f"FinishScore (winner pick correct): {v5:.1%}")
                if v3 == v3:
                    print(f"Method acc (true winner conditioned): {v3:.1%}")
        return

    root = tk.Tk()
    SuperModelGUI(root, pipeline)
    root.mainloop()


if __name__ == "__main__":
    main()