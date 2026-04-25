"""
Monte Carlo simulator for UFC cards.

Reuses the trained pipeline from UFC_Model.py to get per-fight win and method
probabilities, then samples N independent realizations of the entire card to
estimate distributions you can't read off the point predictions: parlay
hit-rates, the spread of correct picks, expected finish counts, etc.

Usage: paste matchups into the GUI text box, one per line, in the format
    Fighter A,Fighter B,Weight Class,Gender,Rounds
"""

import os
import threading
import tkinter as tk
from tkinter import messagebox

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from UFC_Model import (
    DATA_PATH,
    METHOD_LABELS,
    UFCSuperModelPipeline,
    _normalize_method_probs,
)
from prop_stats import (
    build_fighter_stats,
    project_matchup_rates,
    round_of_finish,
    sample_durations,
)


# Stats we project per fight (subset of prop_stats.STAT_COLS — the ones
# sportsbooks actually book lines on). Tuples are (key, label, per-fighter, combined).
PROP_DISPLAY = [
    ("sig_str_landed",   "Sig strikes",    True, True),
    ("total_str_landed", "Total strikes",  True, True),
    ("td_landed",        "Takedowns",      True, True),
    ("kd_landed",        "Knockdowns",     True, False),
]

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_OUT_PATH = os.path.join(SCRIPT_DIR, "Monte_Carlo_Predictions.xlsx")


def simulate_card(card, n_trials, seed=42, prop_library=None, lookup_cache=None):
    """Run the Monte Carlo. `card` is a list of dicts with prob_a + method_probs.

    If `prop_library` is provided and each card item has `rates_a`/`rates_b`/`rounds`,
    also simulates prop totals (sig strikes, total strikes, takedowns, knockdowns)
    per trial via Poisson draws on (rate × sampled duration).
    """
    rng = np.random.default_rng(seed)
    n_fights = len(card)
    method_idx = {m: i for i, m in enumerate(METHOD_LABELS)}

    prob_a = np.array([f["prob_a"] for f in card])
    method_mat = np.array([
        [f["method_probs"][m] for m in METHOD_LABELS] for f in card
    ])
    method_mat = method_mat / method_mat.sum(axis=1, keepdims=True)

    winners = (rng.random((n_trials, n_fights)) >= prob_a).astype(np.int8)

    cum = np.cumsum(method_mat, axis=1)
    u = rng.random((n_trials, n_fights))
    methods = (u[:, :, None] >= cum[None, :, :]).sum(axis=2).astype(np.int8)

    favorite_idx = np.array([0 if f["prob_a"] >= 0.5 else 1 for f in card])
    correct_per_trial = (winners == favorite_idx).sum(axis=1)
    finishes_per_trial = (methods != method_idx["Decision"]).sum(axis=1)

    aggregates = {
        "n_trials": n_trials,
        "n_fights": n_fights,
        "mean_correct": float(correct_per_trial.mean()),
        "median_correct": float(np.median(correct_per_trial)),
        "p_all_favorites": float((correct_per_trial == n_fights).mean()),
        "p_at_least_half": float((correct_per_trial >= np.ceil(n_fights / 2)).mean()),
        "mean_finishes": float(finishes_per_trial.mean()),
        "p_at_least_3_finishes": float((finishes_per_trial >= 3).mean()),
        "p_no_finishes": float((finishes_per_trial == 0).mean()),
        "correct_distribution": np.bincount(correct_per_trial, minlength=n_fights + 1) / n_trials,
        "finishes_distribution": np.bincount(finishes_per_trial, minlength=n_fights + 1) / n_trials,
        "per_fight_a_rate": (winners == 0).mean(axis=0),
        "per_fight_finish_rate": (methods != method_idx["Decision"]).mean(axis=0),
    }

    fight_summary = None
    has_props = prop_library is not None and all(
        ("rates_a" in f and "rates_b" in f and "rounds" in f) for f in card
    )
    if has_props:
        rounds_per_fight = [f["rounds"] for f in card]
        durations = sample_durations(card, winners, methods, prop_library, rng, lookup_cache)
        finish_rounds = round_of_finish(durations, methods, rounds_per_fight)

        fight_summary = {"per_fight": [], "card_totals": {}}
        card_total = {key: np.zeros((n_trials,)) for key, *_ in PROP_DISPLAY}

        for fi, f in enumerate(card):
            rounds = f["rounds"]
            fight_methods = methods[:, fi]
            fight_winners = winners[:, fi]
            fight_rd_finish = finish_rounds[:, fi]

            # Winner & method probabilities per fight.
            p_a = float((fight_winners == 0).mean())
            method_p = {
                "Decision":   float((fight_methods == method_idx["Decision"]).mean()),
                "KO/TKO":     float((fight_methods == method_idx["KO/TKO"]).mean()),
                "Submission": float((fight_methods == method_idx["Submission"]).mean()),
            }
            # Round outcome distribution: Decision + R1..R{rounds} (unconditional).
            round_p = {"Decision": method_p["Decision"]}
            for r in range(1, rounds + 1):
                round_p[f"R{r}"] = float((fight_rd_finish == r).mean())
            # Conditional on a finish (excluding decisions). Useful because
            # the unconditional mode can be R1 even when E[duration] is in R2,
            # which confuses readers — this version always answers
            # "IF the fight ends early, when?"
            p_finish = max(1.0 - round_p["Decision"], 1e-9)
            round_p_if_finish = {
                f"R{r}": round_p[f"R{r}"] / p_finish for r in range(1, rounds + 1)
            }

            # Prop totals: Poisson(rate × sampled_duration).
            stats = {}
            for key, _label, _pf, combined in PROP_DISPLAY:
                lam_a = f["rates_a"].get(key, 0.0) * durations[:, fi]
                lam_b = f["rates_b"].get(key, 0.0) * durations[:, fi]
                sa = rng.poisson(lam_a)
                sb = rng.poisson(lam_b)
                stats[key] = {
                    "a_mean": float(sa.mean()),
                    "a_p10": float(np.percentile(sa, 10)),
                    "a_p90": float(np.percentile(sa, 90)),
                    "b_mean": float(sb.mean()),
                    "b_p10": float(np.percentile(sb, 10)),
                    "b_p90": float(np.percentile(sb, 90)),
                    "combined_mean": float((sa + sb).mean()),
                    "combined_p10": float(np.percentile(sa + sb, 10)),
                    "combined_p90": float(np.percentile(sa + sb, 90)),
                }
                if combined:
                    card_total[key] += sa + sb

            finish_mask = fight_methods != 0
            mean_dur_if_finish = (
                float(durations[finish_mask, fi].mean()) if finish_mask.any()
                else float(durations[:, fi].mean())
            )
            fight_summary["per_fight"].append({
                "name_a": f["name_a"],
                "name_b": f["name_b"],
                "weight_class": f.get("weight_class", ""),
                "rounds": rounds,
                "prob_a": p_a,
                "method_p": method_p,
                "round_p": round_p,
                "round_p_if_finish": round_p_if_finish,
                "expected_duration_min": float(durations[:, fi].mean()),
                "expected_duration_if_finish_min": mean_dur_if_finish,
                "stats": stats,
            })

        for key, label, _pf, combined in PROP_DISPLAY:
            if not combined:
                continue
            arr = card_total[key]
            fight_summary["card_totals"][key] = {
                "label": label,
                "mean": float(arr.mean()),
                "p10": float(np.percentile(arr, 10)),
                "p90": float(np.percentile(arr, 90)),
            }

    return aggregates, winners, methods, fight_summary


def _fmt_per_fight_block(idx, entry):
    """Build one fight's prediction block (winner / method / round / props)."""
    a, b = entry["name_a"], entry["name_b"]
    rounds = entry["rounds"]
    p_a = entry["prob_a"]
    winner = a if p_a >= 0.5 else b
    win_pct = max(p_a, 1.0 - p_a)
    loser = b if winner == a else a
    loser_pct = 1.0 - win_pct

    method_p = entry["method_p"]
    pred_method = max(method_p, key=method_p.get)
    round_if = entry["round_p_if_finish"]
    finish_keys = [f"R{r}" for r in range(1, rounds + 1)]
    pred_finish_key = max(finish_keys, key=lambda k: round_if.get(k, 0.0))
    pred_finish_label = f"Round {pred_finish_key[1:]}"

    lines = []
    wc = entry.get("weight_class", "")
    header = f"  {idx:2d}.  {a}  vs  {b}"
    header += f"   ({wc}, {rounds} rds)" if wc else f"   ({rounds} rds)"
    lines.append(header)
    lines.append(f"        Winner :  {winner:<28} {win_pct:.1%}   "
                 f"(vs {loser} {loser_pct:.1%})")
    lines.append(f"        Method :  {pred_method:<28} {method_p[pred_method]:.1%}   "
                 f"(DEC {method_p['Decision']:.0%} | KO {method_p['KO/TKO']:.0%} | "
                 f"SUB {method_p['Submission']:.0%})")
    finish_breakdown = " | ".join(f"R{r} {round_if[f'R{r}']:.0%}" for r in range(1, rounds + 1))
    lines.append(f"        If finish, in: {pred_finish_label:<22} {round_if[pred_finish_key]:.1%}   "
                 f"({finish_breakdown} of finishes)")
    mean_dur = entry["expected_duration_min"]
    mean_dur_finish = entry["expected_duration_if_finish_min"]
    lines.append(f"        Mean duration: {mean_dur:.1f} min overall   |   "
                 f"{mean_dur_finish:.1f} min if finish   "
                 f"(P(decision) = {method_p['Decision']:.0%})")

    lines.append("")
    lines.append(f"        Prop projections   (mean [P10-P90], counts):")
    a_disp = (a[:14] + "…") if len(a) > 15 else a
    b_disp = (b[:14] + "…") if len(b) > 15 else b
    for key, label, _per_fighter, combined in PROP_DISPLAY:
        s = entry["stats"][key]
        row = f"          {label:<14}"
        row += f"  {a_disp:<15} {s['a_mean']:5.1f} [{s['a_p10']:>3.0f}-{s['a_p90']:>3.0f}]"
        row += f"   {b_disp:<15} {s['b_mean']:5.1f} [{s['b_p10']:>3.0f}-{s['b_p90']:>3.0f}]"
        if combined:
            row += f"   Total {s['combined_mean']:5.1f} [{s['combined_p10']:>3.0f}-{s['combined_p90']:>3.0f}]"
        lines.append(row)
    return lines


def format_report(card, agg, fight_summary=None):
    n = agg["n_fights"]
    lines = []
    lines.append("=" * 78)
    lines.append(f"MONTE CARLO RESULTS  —  {agg['n_trials']:,} trials over {n} fights")
    lines.append("=" * 78)

    if fight_summary is not None:
        lines.append("")
        lines.append("PER-FIGHT PREDICTIONS")
        lines.append("=" * 78)
        for fi, entry in enumerate(fight_summary["per_fight"], 1):
            lines.extend(_fmt_per_fight_block(fi, entry))
            lines.append("")
    else:
        # Fallback: no library — show winner/method only.
        lines.append("")
        lines.append("Per-fight model probabilities:")
        lines.append("-" * 78)
        for i, f in enumerate(card):
            fav = f["name_a"] if f["prob_a"] >= 0.5 else f["name_b"]
            fav_pct = max(f["prob_a"], 1.0 - f["prob_a"])
            mp = f["method_probs"]
            lines.append(
                f"  {i+1:2d}. {f['name_a']} vs {f['name_b']}  ({f.get('weight_class','')})\n"
                f"      Favorite: {fav} {fav_pct:.1%}   "
                f"DEC {mp['Decision']:.0%} | KO {mp['KO/TKO']:.0%} | SUB {mp['Submission']:.0%}"
            )

    lines.append("=" * 78)
    lines.append("CARD AGGREGATES")
    lines.append("=" * 78)
    lines.append(f"  Expected correct picks (favorites):  {agg['mean_correct']:.2f} / {n}")
    lines.append(f"  Median correct picks:                {agg['median_correct']:.0f} / {n}")
    lines.append(f"  P(all favorites win — full parlay):  {agg['p_all_favorites']:.2%}")
    lines.append(f"  P(at least half correct):            {agg['p_at_least_half']:.2%}")
    lines.append(f"  Expected finishes on the card:       {agg['mean_finishes']:.2f} / {n}")
    lines.append(f"  P(>= 3 finishes):                    {agg['p_at_least_3_finishes']:.2%}")
    lines.append(f"  P(zero finishes — all decisions):    {agg['p_no_finishes']:.2%}")

    lines.append("")
    lines.append("Distribution of correct picks across trials:")
    lines.append("-" * 78)
    dist = agg["correct_distribution"]
    bar_max = 50
    peak = dist.max() or 1.0
    for k, p in enumerate(dist):
        bar = "#" * int(round((p / peak) * bar_max))
        lines.append(f"  {k:2d} correct: {p:6.2%}  {bar}")

    lines.append("")
    lines.append("Distribution of finishes across trials:")
    lines.append("-" * 78)
    dist = agg["finishes_distribution"]
    peak = dist.max() or 1.0
    for k, p in enumerate(dist):
        bar = "#" * int(round((p / peak) * bar_max))
        lines.append(f"  {k:2d} finishes: {p:6.2%}  {bar}")

    if fight_summary is not None and fight_summary["card_totals"]:
        lines.append("")
        lines.append("Card-wide prop totals  (sum across all fights):")
        lines.append("-" * 78)
        for _key, ct in fight_summary["card_totals"].items():
            lines.append(
                f"  {ct['label']:<16} mean {ct['mean']:6.1f}   "
                f"P10 {ct['p10']:.0f}   P90 {ct['p90']:.0f}"
            )

    return "\n".join(lines)


def _auto_width(ws, max_width=42):
    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        longest = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_letter].width = min(max(longest + 2, 8), max_width)


def export_to_excel(path, card, agg, fight_summary):
    """Write per-fight predictions, props, round distributions, and card aggregates
    to a multi-sheet workbook. Color-codes red/blue corner columns."""
    hdr_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    red_hdr_fill = PatternFill(start_color="E89090", end_color="E89090", fill_type="solid")
    blue_hdr_fill = PatternFill(start_color="90B0E5", end_color="90B0E5", fill_type="solid")
    red_cell_fill = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
    blue_cell_fill = PatternFill(start_color="E5EEFF", end_color="E5EEFF", fill_type="solid")
    hdr_font = Font(bold=True, size=11, color="000000")
    align_left = Alignment(horizontal="left", vertical="center")
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_header(ws, headers, red_cols=(), blue_cols=()):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font, c.alignment, c.border = hdr_font, align_left, border
            if ci in red_cols:
                c.fill = red_hdr_fill
            elif ci in blue_cols:
                c.fill = blue_hdr_fill
            else:
                c.fill = hdr_fill

    def write_row(ws, ri, values, pct_cols=(), red_cols=(), blue_cols=(), winner_color=None):
        for ci, v in enumerate(values, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = border
            c.alignment = align_left
            if ci in pct_cols:
                c.number_format = "0.0%"
            if winner_color is not None and ci == winner_color["col"]:
                c.fill = red_cell_fill if winner_color["side"] == "red" else blue_cell_fill
            elif ci in red_cols:
                c.fill = red_cell_fill
            elif ci in blue_cols:
                c.fill = blue_cell_fill

    wb = Workbook()

    # ───────── Sheet 1: Predictions (per-fight summary) ─────────
    ws = wb.active
    ws.title = "Predictions"
    pred_headers = [
        "#", "Red Corner", "Blue Corner", "Weight Class", "Rounds",
        "Predicted Winner", "Win %",
        "Predicted Method", "Method %",
        "DEC %", "KO/TKO %", "SUB %",
        "If Finish, Round", "P(Round | Finish)",
        "Mean Duration (min)", "Mean Duration if Finish (min)",
    ]
    pct_cols = (7, 9, 10, 11, 12, 14)
    red_cols = (2,)
    blue_cols = (3,)
    write_header(ws, pred_headers, red_cols=red_cols, blue_cols=blue_cols)
    if fight_summary is not None:
        for ri, entry in enumerate(fight_summary["per_fight"], 2):
            a, b = entry["name_a"], entry["name_b"]
            p_a = entry["prob_a"]
            winner = a if p_a >= 0.5 else b
            win_pct = max(p_a, 1.0 - p_a)
            mp = entry["method_p"]
            pred_method = max(mp, key=mp.get)
            rounds = entry["rounds"]
            rif = entry["round_p_if_finish"]
            finish_keys = [f"R{r}" for r in range(1, rounds + 1)]
            pred_finish_key = max(finish_keys, key=lambda k: rif.get(k, 0.0))
            pred_finish_label = f"Round {pred_finish_key[1:]}"
            write_row(ws, ri, [
                ri - 1, a, b, entry.get("weight_class", ""), rounds,
                winner, win_pct,
                pred_method, mp[pred_method],
                mp["Decision"], mp["KO/TKO"], mp["Submission"],
                pred_finish_label, rif[pred_finish_key],
                round(entry["expected_duration_min"], 2),
                round(entry["expected_duration_if_finish_min"], 2),
            ], pct_cols=pct_cols, red_cols=red_cols, blue_cols=blue_cols,
               winner_color={"col": 6, "side": "red" if winner == a else "blue"})
    else:
        for ri, f in enumerate(card, 2):
            p_a = f["prob_a"]
            winner = f["name_a"] if p_a >= 0.5 else f["name_b"]
            win_pct = max(p_a, 1.0 - p_a)
            mp = f["method_probs"]
            pred_method = max(mp, key=mp.get)
            write_row(ws, ri, [
                ri - 1, f["name_a"], f["name_b"], f.get("weight_class", ""),
                f.get("rounds", 3),
                winner, win_pct, pred_method, mp[pred_method],
                mp["Decision"], mp["KO/TKO"], mp["Submission"],
                "—", None, None, None,
            ], pct_cols=pct_cols, red_cols=red_cols, blue_cols=blue_cols,
               winner_color={"col": 6, "side": "red" if winner == f["name_a"] else "blue"})
    _auto_width(ws)

    # ───────── Sheet 2: Props (per-fight × per-stat) ─────────
    if fight_summary is not None:
        ws = wb.create_sheet("Props")
        prop_headers = [
            "Fight #", "Red Corner", "Blue Corner", "Stat",
            "Red Mean", "Red P10", "Red P90",
            "Blue Mean", "Blue P10", "Blue P90",
            "Total Mean", "Total P10", "Total P90",
        ]
        prop_red_cols = (2, 5, 6, 7)
        prop_blue_cols = (3, 8, 9, 10)
        write_header(ws, prop_headers, red_cols=prop_red_cols, blue_cols=prop_blue_cols)
        ri = 2
        for fi, entry in enumerate(fight_summary["per_fight"], 1):
            for key, label, _per_fighter, combined in PROP_DISPLAY:
                s = entry["stats"][key]
                row = [
                    fi, entry["name_a"], entry["name_b"], label,
                    round(s["a_mean"], 2), round(s["a_p10"], 1), round(s["a_p90"], 1),
                    round(s["b_mean"], 2), round(s["b_p10"], 1), round(s["b_p90"], 1),
                ]
                if combined:
                    row += [round(s["combined_mean"], 2),
                            round(s["combined_p10"], 1), round(s["combined_p90"], 1)]
                else:
                    row += [None, None, None]
                write_row(ws, ri, row, red_cols=prop_red_cols, blue_cols=prop_blue_cols)
                ri += 1
        _auto_width(ws)

    # ───────── Sheet 3: Round Distribution ─────────
    if fight_summary is not None:
        ws = wb.create_sheet("Round Distribution")
        rd_headers = ["#", "Red Corner", "Blue Corner", "Rounds Scheduled",
                      "DEC %", "R1 %", "R2 %", "R3 %", "R4 %", "R5 %"]
        rd_pct_cols = (5, 6, 7, 8, 9, 10)
        rd_red_cols = (2,)
        rd_blue_cols = (3,)
        write_header(ws, rd_headers, red_cols=rd_red_cols, blue_cols=rd_blue_cols)
        for ri, entry in enumerate(fight_summary["per_fight"], 2):
            rp = entry["round_p"]
            rounds = entry["rounds"]
            row = [
                ri - 1, entry["name_a"], entry["name_b"], rounds,
                rp.get("Decision", 0.0),
                rp.get("R1", 0.0), rp.get("R2", 0.0), rp.get("R3", 0.0),
                rp.get("R4") if rounds >= 4 else None,
                rp.get("R5") if rounds >= 5 else None,
            ]
            write_row(ws, ri, row, pct_cols=rd_pct_cols,
                      red_cols=rd_red_cols, blue_cols=rd_blue_cols)
        _auto_width(ws)

    # ───────── Sheet 4: Card Aggregates ─────────
    ws = wb.create_sheet("Card Aggregates")
    write_header(ws, ["Metric", "Value"])
    rows = [
        ("Trials", agg["n_trials"]),
        ("Fights", agg["n_fights"]),
        ("Mean correct picks (favorites)", round(agg["mean_correct"], 3)),
        ("Median correct picks", agg["median_correct"]),
        ("P(all favorites win)", agg["p_all_favorites"]),
        ("P(at least half correct)", agg["p_at_least_half"]),
        ("Mean finishes", round(agg["mean_finishes"], 3)),
        ("P(>= 3 finishes)", agg["p_at_least_3_finishes"]),
        ("P(zero finishes)", agg["p_no_finishes"]),
    ]
    pct_metrics = {"P(all favorites win)", "P(at least half correct)",
                   "P(>= 3 finishes)", "P(zero finishes)"}
    for ri, (k, v) in enumerate(rows, 2):
        write_row(ws, ri, [k, v], pct_cols=(2,) if k in pct_metrics else ())

    # Distributions (correct picks + finishes)
    ws.cell(row=len(rows) + 4, column=1, value="Correct picks distribution").font = hdr_font
    base = len(rows) + 5
    for k, p in enumerate(agg["correct_distribution"]):
        c1 = ws.cell(row=base + k, column=1, value=f"{k} correct")
        c2 = ws.cell(row=base + k, column=2, value=float(p))
        c2.number_format = "0.00%"
        c1.border = border; c2.border = border

    base2 = base + len(agg["correct_distribution"]) + 2
    ws.cell(row=base2 - 1, column=1, value="Finishes distribution").font = hdr_font
    for k, p in enumerate(agg["finishes_distribution"]):
        c1 = ws.cell(row=base2 + k, column=1, value=f"{k} finishes")
        c2 = ws.cell(row=base2 + k, column=2, value=float(p))
        c2.number_format = "0.00%"
        c1.border = border; c2.border = border

    if fight_summary is not None and fight_summary.get("card_totals"):
        base3 = base2 + len(agg["finishes_distribution"]) + 2
        ws.cell(row=base3 - 1, column=1, value="Card-wide prop totals").font = hdr_font
        ws.cell(row=base3, column=1, value="Stat").font = hdr_font
        ws.cell(row=base3, column=2, value="Mean").font = hdr_font
        ws.cell(row=base3, column=3, value="P10").font = hdr_font
        ws.cell(row=base3, column=4, value="P90").font = hdr_font
        for ri, (_key, ct) in enumerate(fight_summary["card_totals"].items(), base3 + 1):
            ws.cell(row=ri, column=1, value=ct["label"])
            ws.cell(row=ri, column=2, value=round(ct["mean"], 2))
            ws.cell(row=ri, column=3, value=round(ct["p10"], 1))
            ws.cell(row=ri, column=4, value=round(ct["p90"], 1))
    _auto_width(ws)

    wb.save(path)


class MonteCarloGUI:
    BG = "#0A0A0A"
    BG_HEADER = "#111111"
    BG_INPUT = "#141414"
    FG = "#F5F5F5"
    ACCENT = "#D20A11"
    MUTED = "#CFCFCF"

    def __init__(self, root, pipeline, prop_library=None):
        self.root = root
        self.pipeline = pipeline
        self.prop_library = prop_library
        self._prop_lookup_cache = {}
        self._busy = False
        self.root.title("UFC Monte Carlo")
        self.root.geometry("1100x900")
        self.root.minsize(960, 720)
        self.root.configure(bg=self.BG)
        self._build_ui()

    def _build_ui(self):
        tk.Frame(self.root, bg=self.ACCENT, height=6).pack(fill="x")

        tf = tk.Frame(self.root, bg=self.BG_HEADER, pady=14)
        tf.pack(fill="x")
        tk.Label(tf, text="UFC", font=("Helvetica", 34, "bold"),
                 fg=self.ACCENT, bg=self.BG_HEADER).pack()
        tk.Label(tf, text="MONTE CARLO SIMULATOR", font=("Helvetica", 10, "bold"),
                 fg=self.FG, bg=self.BG_HEADER).pack(pady=(0, 2))

        main = tk.Frame(self.root, bg=self.BG, padx=18, pady=12)
        main.pack(fill="both", expand=True)

        self.status_var = tk.StringVar(value="")
        tk.Label(main, textvariable=self.status_var, bg=self.BG, fg=self.MUTED,
                 font=("Helvetica", 9, "italic")).pack(anchor="w")

        mode_frame = tk.Frame(main, bg=self.BG)
        mode_frame.pack(fill="x", pady=(8, 4))
        tk.Label(mode_frame, text="Input mode:", bg=self.BG, fg=self.FG,
                 font=("Helvetica", 9, "bold")).pack(side="left", padx=(0, 10))
        self.mode_var = tk.StringVar(value="matchups")
        for value, label in (
            ("matchups", "Matchups (train + predict)"),
            ("predictions", "Pre-computed predictions (paste from Excel)"),
        ):
            tk.Radiobutton(
                mode_frame, text=label, value=value, variable=self.mode_var,
                bg=self.BG, fg=self.FG, selectcolor=self.BG_INPUT,
                activebackground=self.BG, activeforeground=self.ACCENT,
                font=("Helvetica", 9), command=self._on_mode_change,
            ).pack(side="left", padx=4)

        self.input_label_var = tk.StringVar()
        tk.Label(main, textvariable=self.input_label_var, bg=self.BG, fg=self.FG,
                 font=("Helvetica", 9, "bold")).pack(anchor="w", pady=(8, 4))
        input_wrap = tk.Frame(main, bg=self.BG_HEADER, padx=2, pady=2)
        input_wrap.pack(fill="both", expand=False, pady=4)
        self.fight_input = tk.Text(
            input_wrap, height=14, font=("Courier New", 10), bg=self.BG_INPUT,
            fg=self.FG, insertbackground=self.FG, relief="flat", wrap="word",
            highlightthickness=1, highlightbackground=self.ACCENT, highlightcolor=self.ACCENT,
        )
        input_sb = tk.Scrollbar(
            input_wrap, command=self.fight_input.yview, bg=self.BG_HEADER,
            troughcolor=self.BG_INPUT, activebackground=self.ACCENT,
        )
        self.fight_input.configure(yscrollcommand=input_sb.set)
        self.fight_input.pack(side="left", fill="both", expand=True)
        input_sb.pack(side="right", fill="y")

        ctrl = tk.Frame(main, bg=self.BG)
        ctrl.pack(fill="x", pady=(10, 6))
        tk.Label(ctrl, text="Trials:", bg=self.BG, fg=self.FG,
                 font=("Helvetica", 10, "bold")).pack(side="left", padx=(0, 6))
        self.trials_var = tk.StringVar(value="10000")
        tk.Entry(ctrl, textvariable=self.trials_var, width=10, font=("Courier New", 10),
                 bg=self.BG_INPUT, fg=self.FG, insertbackground=self.FG, relief="flat",
                 highlightthickness=1, highlightbackground=self.ACCENT,
                 highlightcolor=self.ACCENT).pack(side="left", padx=(0, 18))

        tk.Label(ctrl, text="Seed:", bg=self.BG, fg=self.FG,
                 font=("Helvetica", 10, "bold")).pack(side="left", padx=(0, 6))
        self.seed_var = tk.StringVar(value="42")
        tk.Entry(ctrl, textvariable=self.seed_var, width=8, font=("Courier New", 10),
                 bg=self.BG_INPUT, fg=self.FG, insertbackground=self.FG, relief="flat",
                 highlightthickness=1, highlightbackground=self.ACCENT,
                 highlightcolor=self.ACCENT).pack(side="left")

        tk.Button(ctrl, text="Clear", command=self._clear, font=("Helvetica", 10, "bold"),
                  bg="#202020", fg=self.FG, relief="flat", padx=14, cursor="hand2",
                  activebackground="#2A2A2A", activeforeground=self.FG).pack(side="left", padx=(18, 4))
        self.run_btn = tk.Button(ctrl, text="Run Simulation", command=self._run,
                                 font=("Helvetica", 11, "bold"), bg=self.ACCENT, fg=self.FG,
                                 relief="flat", padx=24, cursor="hand2",
                                 activebackground="#B40A0F", activeforeground=self.FG)
        self.run_btn.pack(side="right", padx=4)

        tk.Label(main, text="Results", bg=self.BG, fg=self.FG,
                 font=("Helvetica", 9, "bold")).pack(anchor="w", pady=(8, 4))
        out_wrap = tk.Frame(main, bg=self.BG_HEADER, padx=2, pady=2)
        out_wrap.pack(fill="both", expand=True, pady=4)
        self.output = tk.Text(
            out_wrap, font=("Courier New", 9), bg=self.BG_INPUT, fg=self.FG,
            insertbackground=self.FG, relief="flat", wrap="none",
            highlightthickness=1, highlightbackground=self.ACCENT, highlightcolor=self.ACCENT,
        )
        out_sb = tk.Scrollbar(
            out_wrap, command=self.output.yview, bg=self.BG_HEADER,
            troughcolor=self.BG_INPUT, activebackground=self.ACCENT,
        )
        self.output.configure(yscrollcommand=out_sb.set, state="disabled")
        self.output.pack(side="left", fill="both", expand=True)
        out_sb.pack(side="right", fill="y")

        self._on_mode_change()

    def _on_mode_change(self):
        if self.mode_var.get() == "matchups":
            self.input_label_var.set(
                "Matchups  (one per line: Fighter A,Fighter B,Weight Class,Gender,Rounds)"
            )
            self.status_var.set(
                "Paste matchups, choose trials, then Run. First run trains the model."
            )
        else:
            self.input_label_var.set(
                "Predictions  (paste from UFC_Predictions.xlsx — "
                "Red, Blue, Weight Class, Winner, Win %, Method, Method %, "
                "DEC %, KO %, SUB %  [+ optional Rounds column; defaults to 3])"
            )
            self.status_var.set(
                "Paste pre-computed predictions from Excel. "
                "Append a Rounds column for 5-round main events."
            )

    def _clear(self):
        self.fight_input.delete("1.0", tk.END)
        self._set_output("")

    def _set_output(self, text):
        self.output.config(state="normal")
        self.output.delete("1.0", tk.END)
        self.output.insert("1.0", text)
        self.output.config(state="disabled")

    def _parse_lines(self, text):
        rows = []
        for ln in text.splitlines():
            ln = ln.strip()
            if not ln:
                continue
            parts = [p.strip() for p in ln.split(",")]
            if len(parts) < 2:
                continue
            a, b = parts[0], parts[1]
            wc = parts[2] if len(parts) > 2 else ""
            g = parts[3] if len(parts) > 3 else ""
            try:
                rounds = int(parts[4]) if len(parts) > 4 else 3
            except ValueError:
                rounds = 3
            rows.append((a, b, wc, g, rounds))
        return rows

    @staticmethod
    def _parse_pct(s):
        """Parse a percentage cell: '58.9%', '58.9', or '0.589' all → 0.589."""
        s = (s or "").strip().rstrip("%").strip()
        if not s:
            raise ValueError("empty percentage cell")
        v = float(s)
        return v / 100.0 if v > 1.0 else v

    def _parse_predictions(self, text):
        """Parse Excel-pasted predictions into card dicts.

        Expected columns (tab- or comma-separated, in this order):
          Red Corner, Blue Corner, Weight Class, Winner, Win %,
          Method, Method %, DEC %, (T)KO %, SUB %, [Rounds optional]

        If a Rounds column isn't present, defaults to 3. Skips header row.
        Returns (card, errors).
        """
        card = []
        errors = []
        for raw_idx, raw in enumerate(text.splitlines(), start=1):
            ln = raw.rstrip("\r\n").strip()
            if not ln:
                continue
            sep = "\t" if "\t" in ln else ","
            parts = [p.strip() for p in ln.split(sep)]
            if parts and parts[0].lower() in ("red corner", "red", "fighter a"):
                continue
            if len(parts) < 10:
                errors.append(f"line {raw_idx}: expected 10 columns, got {len(parts)}")
                continue
            try:
                red, blue, wc = parts[0], parts[1], parts[2]
                winner = parts[3]
                win_pct = self._parse_pct(parts[4])
                dec_pct = self._parse_pct(parts[7])
                ko_pct = self._parse_pct(parts[8])
                sub_pct = self._parse_pct(parts[9])
            except ValueError as e:
                errors.append(f"line {raw_idx}: {e}")
                continue
            try:
                rounds = int(parts[10]) if len(parts) > 10 and parts[10] else 3
            except ValueError:
                rounds = 3

            wn = winner.casefold()
            if wn == red.casefold():
                prob_a = win_pct
            elif wn == blue.casefold():
                prob_a = 1.0 - win_pct
            else:
                errors.append(
                    f"line {raw_idx}: winner '{winner}' matches neither '{red}' nor '{blue}'"
                )
                continue

            method_probs = _normalize_method_probs({
                "Decision": dec_pct, "KO/TKO": ko_pct, "Submission": sub_pct,
            })
            card.append({
                "name_a": red,
                "name_b": blue,
                "prob_a": prob_a,
                "prob_b": 1.0 - prob_a,
                "weight_class": wc,
                "method_probs": method_probs,
                "rounds": rounds,
            })
        return card, errors

    def _attach_props(self, card):
        """Annotate card items with `rates_a`, `rates_b`, ensure `rounds` set.
        Returns list of warnings about fighters not found in the rate library."""
        warnings = []
        if self.prop_library is None:
            return warnings
        for entry in card:
            if "rounds" not in entry:
                entry["rounds"] = 3
            proj = project_matchup_rates(
                entry["name_a"], entry["name_b"], self.prop_library, self._prop_lookup_cache,
            )
            entry["rates_a"] = proj["rates_a"]
            entry["rates_b"] = proj["rates_b"]
            if not proj["a_found"]:
                warnings.append(f"{entry['name_a']} not in stats library — used population rates")
            if not proj["b_found"]:
                warnings.append(f"{entry['name_b']} not in stats library — used population rates")
        return warnings

    def _run(self):
        if self._busy:
            return
        text = self.fight_input.get("1.0", tk.END).strip()
        if not text:
            self.status_var.set("Enter at least one matchup.")
            return
        try:
            n_trials = max(1, int(self.trials_var.get().strip()))
        except ValueError:
            self.status_var.set("Trials must be an integer.")
            return
        try:
            seed = int(self.seed_var.get().strip())
        except ValueError:
            seed = 42

        self._busy = True
        self.run_btn.config(state="disabled")
        self.status_var.set("Working…")

        mode = self.mode_var.get()

        def _do():
            try:
                skipped = []
                if mode == "matchups":
                    if self.pipeline.model is None:
                        self._set_status("Training model (one-time)…")
                        self.pipeline.train()

                    rows = self._parse_lines(text)
                    self._set_status(f"Predicting {len(rows)} matchup(s)…")
                    card = []
                    for a, b, wc, g, rounds in rows:
                        if self.pipeline.is_debutant(a) or self.pipeline.is_debutant(b):
                            skipped.append(f"{a} vs {b} (debutant)")
                            continue
                        p = self.pipeline.predict_matchup(a, b, wc, g, rounds)
                        p["method_probs"] = _normalize_method_probs(p.get("method_probs", {}))
                        p["rounds"] = rounds
                        card.append(p)
                else:
                    self._set_status("Parsing pasted predictions…")
                    card, parse_errors = self._parse_predictions(text)
                    skipped.extend(parse_errors)

                if not card:
                    self._set_status("No usable rows.")
                    msg = "Nothing to simulate."
                    if skipped:
                        msg += "\n\nSkipped/errors:\n  " + "\n  ".join(skipped)
                    self.root.after(0, lambda m=msg: self._set_output(m))
                    return

                self._set_status("Attaching prop rates…")
                prop_warnings = self._attach_props(card)
                skipped.extend(prop_warnings)

                self._set_status(f"Simulating {n_trials:,} trials over {len(card)} fights…")
                agg, _winners, _methods, fight_summary = simulate_card(
                    card, n_trials, seed=seed,
                    prop_library=self.prop_library,
                    lookup_cache=self._prop_lookup_cache,
                )
                report = format_report(card, agg, fight_summary=fight_summary)
                if skipped:
                    report += "\n\nNotes:\n  " + "\n  ".join(skipped)

                try:
                    export_to_excel(EXCEL_OUT_PATH, card, agg, fight_summary)
                    report += f"\n\nSaved: {os.path.basename(EXCEL_OUT_PATH)}"
                except Exception as e:
                    report += f"\n\nExcel export failed: {e}"

                self.root.after(0, lambda: self._set_output(report))
                self._set_status(
                    f"Done. {n_trials:,} trials × {len(card)} fights "
                    f"({len(skipped)} skipped)."
                )
            except Exception as e:
                err = str(e)
                self._set_status(f"Simulation failed: {err}")
                self.root.after(0, lambda em=err: messagebox.showerror("Monte Carlo Error", em))
            finally:
                self._busy = False
                self.root.after(0, lambda: self.run_btn.config(state="normal"))

        threading.Thread(target=_do, daemon=True).start()

    def _set_status(self, msg):
        self.root.after(0, lambda: self.status_var.set(msg))


def main():
    if not os.path.exists(DATA_PATH):
        raise FileNotFoundError(f"Missing dataset: {DATA_PATH}")
    pipeline = UFCSuperModelPipeline(DATA_PATH)
    print("Building prop-stats library from pure_fight_data.csv…")
    prop_library = build_fighter_stats(DATA_PATH)
    print(f"  → {len(prop_library) - 1} fighters indexed.")
    root = tk.Tk()
    MonteCarloGUI(root, pipeline, prop_library=prop_library)
    root.mainloop()


if __name__ == "__main__":
    main()
