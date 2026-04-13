import os
from collections import defaultdict
from datetime import datetime

import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False


class MatchupDashboard:
    BG = "#0A0A0A"
    BG_HEADER = "#111111"
    BG_CARD = "#141414"
    BG_INPUT = "#1A1A1A"
    FG = "#F5F5F5"
    MUTED = "#BEBEBE"
    ACCENT = "#D20A11"
    CYAN = "#3BE8FF"

    def __init__(self, root, data_file=None):
        self.root = root
        self.root.title("UFC Matchup Dashboard")
        self.root.geometry("1360x920")
        self.root.minsize(1180, 780)
        self.root.configure(bg=self.BG)

        if data_file is None:
            data_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pure_fight_data.csv")

        self.df = self._load_data(data_file)
        self.fighter_logs = self._build_fighter_logs(self.df)
        self.fighter_names = sorted(self.fighter_logs.keys())
        self.matchups = []
        self._build_ui()

    def _load_data(self, data_file):
        if not os.path.exists(data_file):
            raise FileNotFoundError(f"Missing dataset: {data_file}")
        df = pd.read_csv(data_file)
        df["event_date"] = pd.to_datetime(df["event_date"], errors="coerce")
        return df.sort_values("event_date").reset_index(drop=True)

    @staticmethod
    def _safe(v, default=0.0):
        if pd.isna(v):
            return default
        try:
            return float(v)
        except Exception:
            return default

    @staticmethod
    def _text(v, default=""):
        if pd.isna(v):
            return default
        return str(v)

    @staticmethod
    def _norm_method(method):
        m = str(method).lower()
        if "sub" in m:
            return "Submission"
        if "ko" in m or "tko" in m or "doctor" in m or "corner" in m or "dq" in m:
            return "KO/TKO"
        if "decision" in m or "split" in m or "unanimous" in m or "majority" in m:
            return "Decision"
        return "Other"

    def _build_fighter_logs(self, df):
        logs = defaultdict(list)
        for _, row in df.iterrows():
            winner = self._text(row.get("winner"))
            red_result, blue_result = "Draw", "Draw"
            if winner == "Red":
                red_result, blue_result = "Win", "Loss"
            elif winner == "Blue":
                red_result, blue_result = "Loss", "Win"

            red_name = self._text(row.get("r_fighter"))
            blue_name = self._text(row.get("b_fighter"))
            if not red_name or not blue_name:
                continue

            common = {
                "date": row.get("event_date"),
                "method": self._text(row.get("method"), "N/A"),
                "fight_time_sec": self._safe(row.get("total_fight_time_sec"), self._safe(row.get("time_sec"), 0.0)),
            }
            red_entry = {
                **common,
                "opponent": blue_name,
                "result": red_result,
                "dob": self._text(row.get("r_date_of_birth"), ""),
                "age_at_event": self._safe(row.get("r_age_at_event"), float("nan")),
                "sig_landed": self._safe(row.get("r_sig_str"), 0.0),
                "sig_att": self._safe(row.get("r_sig_str_att"), 0.0),
                "opp_sig_landed": self._safe(row.get("b_sig_str"), 0.0),
                "opp_sig_att": self._safe(row.get("b_sig_str_att"), 0.0),
                "td_landed": self._safe(row.get("r_td"), 0.0),
                "td_att": self._safe(row.get("r_td_att"), 0.0),
                "opp_td_landed": self._safe(row.get("b_td"), 0.0),
                "opp_td_att": self._safe(row.get("b_td_att"), 0.0),
                "sub_att": self._safe(row.get("r_sub_att"), 0.0),
            }
            blue_entry = {
                **common,
                "opponent": red_name,
                "result": blue_result,
                "dob": self._text(row.get("b_date_of_birth"), ""),
                "age_at_event": self._safe(row.get("b_age_at_event"), float("nan")),
                "sig_landed": self._safe(row.get("b_sig_str"), 0.0),
                "sig_att": self._safe(row.get("b_sig_str_att"), 0.0),
                "opp_sig_landed": self._safe(row.get("r_sig_str"), 0.0),
                "opp_sig_att": self._safe(row.get("r_sig_str_att"), 0.0),
                "td_landed": self._safe(row.get("b_td"), 0.0),
                "td_att": self._safe(row.get("b_td_att"), 0.0),
                "opp_td_landed": self._safe(row.get("r_td"), 0.0),
                "opp_td_att": self._safe(row.get("r_td_att"), 0.0),
                "sub_att": self._safe(row.get("b_sub_att"), 0.0),
            }
            logs[red_name].append(red_entry)
            logs[blue_name].append(blue_entry)

        for name in logs:
            logs[name] = sorted(logs[name], key=lambda x: pd.Timestamp.min if pd.isna(x["date"]) else x["date"])
        return logs

    def _fighter_stats(self, fighter):
        fights = self.fighter_logs.get(fighter, [])
        if not fights:
            return None
        total = len(fights)
        wins = sum(1 for f in fights if f["result"] == "Win")
        losses = sum(1 for f in fights if f["result"] == "Loss")
        draws = sum(1 for f in fights if f["result"] == "Draw")

        total_sec = sum(max(f["fight_time_sec"], 1.0) for f in fights)
        total_min = max(total_sec / 60.0, 1e-6)
        sig_landed = sum(f["sig_landed"] for f in fights)
        sig_att = sum(f["sig_att"] for f in fights)
        opp_sig_landed = sum(f["opp_sig_landed"] for f in fights)
        opp_sig_att = sum(f["opp_sig_att"] for f in fights)
        td_landed = sum(f["td_landed"] for f in fights)
        td_att = sum(f["td_att"] for f in fights)
        opp_td_landed = sum(f["opp_td_landed"] for f in fights)
        opp_td_att = sum(f["opp_td_att"] for f in fights)
        sub_att = sum(f["sub_att"] for f in fights)

        latest = fights[-1]
        current_age = float("nan")
        dob = pd.to_datetime(latest.get("dob"), errors="coerce")
        if pd.notna(dob):
            current_age = (datetime.now() - dob.to_pydatetime()).days / 365.25
        elif pd.notna(latest.get("age_at_event")) and pd.notna(latest.get("date")):
            current_age = float(latest["age_at_event"]) + (datetime.now() - latest["date"].to_pydatetime()).days / 365.25

        wins_method = {"KO/TKO": 0, "Submission": 0, "Decision": 0}
        losses_method = {"KO/TKO": 0, "Submission": 0, "Decision": 0}
        for fight in fights:
            method = self._norm_method(fight["method"])
            if method in wins_method:
                if fight["result"] == "Win":
                    wins_method[method] += 1
                elif fight["result"] == "Loss":
                    losses_method[method] += 1

        return {
            "wins": wins,
            "losses": losses,
            "draws": draws,
            "win_rate": wins / total if total else 0.0,
            "current_age": current_age,
            "win_loss_ratio": wins / max(losses, 1),
            "pro_SLpM": sig_landed / total_min,
            "pro_sig_str_acc": sig_landed / max(sig_att, 1e-6),
            "pro_SApM": opp_sig_landed / total_min,
            "pro_str_def": 1.0 - (opp_sig_landed / max(opp_sig_att, 1e-6)),
            "pro_td_avg": td_landed * 15.0 / total_min,
            "pro_td_acc": td_landed / max(td_att, 1e-6),
            "pro_td_def": 1.0 - (opp_td_landed / max(opp_td_att, 1e-6)),
            "pro_sub_avg": sub_att * 15.0 / total_min,
            "wins_by_method": wins_method,
            "losses_by_method": losses_method,
            "history": fights[-12:],
        }

    @staticmethod
    def _fmt_num(v, d=2):
        return "N/A" if pd.isna(v) else f"{float(v):.{d}f}"

    @staticmethod
    def _fmt_pct(v, d=1):
        return "N/A" if pd.isna(v) else f"{float(v) * 100:.{d}f}%"

    def _comparison_rows(self, red, blue):
        def diff(key):
            if pd.isna(red[key]) or pd.isna(blue[key]):
                return float("nan")
            return float(red[key]) - float(blue[key])

        return [
            ("r_current_age / b_current_age", self._fmt_num(red["current_age"], 1), self._fmt_num(blue["current_age"], 1), self._fmt_num(diff("current_age"), 1)),
            ("r_wins / b_wins", str(red["wins"]), str(blue["wins"]), self._fmt_num(diff("wins"), 0)),
            ("r_losses / b_losses", str(red["losses"]), str(blue["losses"]), self._fmt_num(diff("losses"), 0)),
            ("r_draws / b_draws", str(red["draws"]), str(blue["draws"]), self._fmt_num(diff("draws"), 0)),
            ("r_win_loss_ratio / b_win_loss_ratio", self._fmt_num(red["win_loss_ratio"]), self._fmt_num(blue["win_loss_ratio"]), self._fmt_num(diff("win_loss_ratio"))),
            ("r_pro_SLpM / b_pro_SLpM", self._fmt_num(red["pro_SLpM"]), self._fmt_num(blue["pro_SLpM"]), self._fmt_num(diff("pro_SLpM"))),
            ("r_pro_sig_str_acc / b_pro_sig_str_acc", self._fmt_pct(red["pro_sig_str_acc"]), self._fmt_pct(blue["pro_sig_str_acc"]), self._fmt_pct(diff("pro_sig_str_acc"))),
            ("r_pro_SApM / b_pro_SApM", self._fmt_num(red["pro_SApM"]), self._fmt_num(blue["pro_SApM"]), self._fmt_num(diff("pro_SApM"))),
            ("r_pro_str_def / b_pro_str_def", self._fmt_pct(red["pro_str_def"]), self._fmt_pct(blue["pro_str_def"]), self._fmt_pct(diff("pro_str_def"))),
            ("r_pro_td_avg / b_pro_td_avg", self._fmt_num(red["pro_td_avg"]), self._fmt_num(blue["pro_td_avg"]), self._fmt_num(diff("pro_td_avg"))),
            ("r_pro_td_acc / b_pro_td_acc", self._fmt_pct(red["pro_td_acc"]), self._fmt_pct(blue["pro_td_acc"]), self._fmt_pct(diff("pro_td_acc"))),
            ("r_pro_td_def / b_pro_td_def", self._fmt_pct(red["pro_td_def"]), self._fmt_pct(blue["pro_td_def"]), self._fmt_pct(diff("pro_td_def"))),
            ("r_pro_sub_avg / b_pro_sub_avg", self._fmt_num(red["pro_sub_avg"]), self._fmt_num(blue["pro_sub_avg"]), self._fmt_num(diff("pro_sub_avg"))),
        ]

    def _build_ui(self):
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure("Fut.TCombobox", fieldbackground=self.BG_INPUT, background=self.BG_INPUT, foreground=self.FG)

        tk.Frame(self.root, bg=self.ACCENT, height=6).pack(fill="x")
        header = tk.Frame(self.root, bg=self.BG_HEADER, pady=14)
        header.pack(fill="x")
        tk.Label(header, text="UFC", font=("Helvetica", 34, "bold"), bg=self.BG_HEADER, fg=self.ACCENT).pack()
        tk.Label(header, text="MATCHUP DASHBOARD", font=("Helvetica", 11, "bold"), bg=self.BG_HEADER, fg=self.FG).pack()
        tk.Label(header, text="Leak-safe manual stat reconstruction from pure_fight_data.csv", font=("Helvetica", 9, "italic"), bg=self.BG_HEADER, fg=self.MUTED).pack(pady=(2, 0))

        main = tk.Frame(self.root, bg=self.BG, padx=18, pady=12)
        main.pack(fill="both", expand=True)
        top = tk.Frame(main, bg=self.BG)
        top.pack(fill="x", pady=(0, 10))

        tk.Label(top, text="Red Corner", bg=self.BG, fg=self.FG, font=("Helvetica", 10, "bold")).grid(row=0, column=0, sticky="w")
        self.red_var = tk.StringVar()
        self.red_combo = ttk.Combobox(top, textvariable=self.red_var, values=self.fighter_names, width=36, style="Fut.TCombobox")
        self.red_combo.grid(row=1, column=0, padx=(0, 12), pady=4)
        self.red_combo.bind("<KeyRelease>", lambda _e: self._filter_combo(self.red_combo, self.red_var))

        tk.Label(top, text="Blue Corner", bg=self.BG, fg=self.FG, font=("Helvetica", 10, "bold")).grid(row=0, column=1, sticky="w")
        self.blue_var = tk.StringVar()
        self.blue_combo = ttk.Combobox(top, textvariable=self.blue_var, values=self.fighter_names, width=36, style="Fut.TCombobox")
        self.blue_combo.grid(row=1, column=1, padx=(0, 12), pady=4)
        self.blue_combo.bind("<KeyRelease>", lambda _e: self._filter_combo(self.blue_combo, self.blue_var))

        btns = tk.Frame(top, bg=self.BG)
        btns.grid(row=1, column=2, sticky="e")
        tk.Button(btns, text="Add Matchup", command=self._add_matchup, bg=self.ACCENT, fg=self.FG, font=("Helvetica", 10, "bold"), relief="flat", padx=12, pady=6).pack(side="left", padx=(0, 8))
        tk.Button(btns, text="Generate Excel", command=self.generate_excel, bg=self.CYAN, fg="#0A0A0A", font=("Helvetica", 10, "bold"), relief="flat", padx=12, pady=6).pack(side="left", padx=(0, 8))
        tk.Button(btns, text="Clear Queue", command=self._clear_queue, bg="#262626", fg=self.FG, font=("Helvetica", 10, "bold"), relief="flat", padx=12, pady=6).pack(side="left")

        mid = tk.Frame(main, bg=self.BG)
        mid.pack(fill="both", expand=True)
        left = tk.Frame(mid, bg=self.BG_CARD, highlightthickness=1, highlightbackground=self.ACCENT)
        left.pack(side="left", fill="both", expand=True, padx=(0, 8))
        tk.Label(left, text="Matchup Queue", bg=self.BG_CARD, fg=self.CYAN, font=("Helvetica", 11, "bold")).pack(anchor="w", padx=10, pady=(8, 4))
        self.queue_list = tk.Listbox(left, bg=self.BG_INPUT, fg=self.FG, selectbackground=self.ACCENT, selectforeground=self.FG, font=("Consolas", 10), relief="flat")
        self.queue_list.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        self.queue_list.bind("<Double-Button-1>", lambda _e: self._remove_selected())

        right = tk.Frame(mid, bg=self.BG_CARD, highlightthickness=1, highlightbackground=self.CYAN)
        right.pack(side="left", fill="both", expand=True, padx=(8, 0))
        tk.Label(right, text="Batch Input", bg=self.BG_CARD, fg=self.CYAN, font=("Helvetica", 11, "bold")).pack(anchor="w", padx=10, pady=(8, 4))
        tk.Label(right, text="One per line: Fighter A,Fighter B[,WeightClass,Gender,Rounds]", bg=self.BG_CARD, fg=self.MUTED, font=("Helvetica", 9)).pack(anchor="w", padx=10)
        self.bulk_text = tk.Text(right, bg=self.BG_INPUT, fg=self.FG, insertbackground=self.FG, font=("Consolas", 10), relief="flat")
        self.bulk_text.pack(fill="both", expand=True, padx=10, pady=(8, 8))
        bar = tk.Frame(right, bg=self.BG_CARD)
        bar.pack(fill="x", padx=10, pady=(0, 10))
        tk.Button(bar, text="Parse & Add", command=self._parse_bulk, bg="#222222", fg=self.FG, font=("Helvetica", 10, "bold"), relief="flat", padx=10, pady=5).pack(side="left")
        tk.Button(bar, text="Remove Selected", command=self._remove_selected, bg="#2D2D2D", fg=self.FG, font=("Helvetica", 10, "bold"), relief="flat", padx=10, pady=5).pack(side="left", padx=(8, 0))

        self.status_var = tk.StringVar(value=f"Loaded {len(self.df)} fights and {len(self.fighter_names)} fighters.")
        tk.Label(main, textvariable=self.status_var, bg=self.BG, fg=self.MUTED, font=("Helvetica", 9, "italic")).pack(anchor="w", pady=(8, 0))

    def _filter_combo(self, combo, var):
        q = var.get().strip().lower()
        combo["values"] = [n for n in self.fighter_names if q in n.lower()][:30] if len(q) >= 2 else []

    def _refresh_queue(self):
        self.queue_list.delete(0, tk.END)
        for i, m in enumerate(self.matchups, 1):
            self.queue_list.insert(tk.END, f"{i:>2}. {m['red']}  vs  {m['blue']}")

    def _add_matchup(self):
        red, blue = self.red_var.get().strip(), self.blue_var.get().strip()
        if not red or not blue:
            messagebox.showwarning("Missing Fighters", "Select both fighters.")
            return
        if red == blue:
            messagebox.showwarning("Invalid Matchup", "A fighter cannot fight themselves.")
            return
        if red not in self.fighter_logs or blue not in self.fighter_logs:
            messagebox.showwarning("Unknown Fighter", "One or both fighters are not in the dataset.")
            return
        self.matchups.append({"red": red, "blue": blue})
        self._refresh_queue()
        self.red_var.set("")
        self.blue_var.set("")
        self.status_var.set(f"Added matchup: {red} vs {blue} | queue={len(self.matchups)}")

    def _remove_selected(self):
        sel = self.queue_list.curselection()
        if not sel:
            return
        removed = self.matchups.pop(sel[0])
        self._refresh_queue()
        self.status_var.set(f"Removed matchup: {removed['red']} vs {removed['blue']}")

    def _clear_queue(self):
        if self.matchups and messagebox.askyesno("Confirm", "Clear all queued matchups?"):
            self.matchups.clear()
            self._refresh_queue()
            self.status_var.set("Queue cleared.")

    def _parse_bulk(self):
        txt = self.bulk_text.get("1.0", tk.END).strip()
        if not txt:
            return
        added, skipped = 0, 0
        for line in [ln.strip() for ln in txt.splitlines() if ln.strip()]:
            parts = [p.strip() for p in line.split(",")]
            if len(parts) < 2:
                skipped += 1
                continue
            red, blue = parts[0], parts[1]
            if red == blue or red not in self.fighter_logs or blue not in self.fighter_logs:
                skipped += 1
                continue
            self.matchups.append({"red": red, "blue": blue})
            added += 1
        self._refresh_queue()
        self.status_var.set(f"Bulk import complete: added={added}, skipped={skipped}, queue={len(self.matchups)}")
        if added:
            self.bulk_text.delete("1.0", tk.END)

    def generate_excel(self):
        """Compatibility wrapper using legacy method name."""
        return self._generate_excel()

    def _generate_excel(self):
        if not self.matchups:
            messagebox.showwarning("Warning", "No matchups in queue. Please add matchups first.")
            return
        if not HAS_OPENPYXL:
            messagebox.showerror("Error", "openpyxl not installed. Install with: pip install openpyxl")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile="UFC_Comparisons.xlsx",
        )
        if not path:
            return

        try:
            self.status_var.set("Generating Excel workbook...")
            self.root.update()

            wb = Workbook()
            wb.remove(wb.active)

            for i, matchup in enumerate(self.matchups, 1):
                fighter1 = matchup["red"]
                fighter2 = matchup["blue"]
                self.status_var.set(
                    f"Processing matchup {i}/{len(self.matchups)}: {fighter1} vs {fighter2}"
                )
                self.root.update()

                fighter1_stats = self._get_fighter_data(fighter1)
                fighter2_stats = self._get_fighter_data(fighter2)
                if not fighter1_stats or not fighter2_stats:
                    messagebox.showwarning(
                        "Warning", f"Could not find data for matchup: {fighter1} vs {fighter2}"
                    )
                    continue

                sheet_name = f"{fighter1[:15]} vs {fighter2[:15]}"[:31]
                ws = wb.create_sheet(title=sheet_name)
                self.write_comparison_to_sheet(ws, fighter1_stats, fighter2_stats)

            wb.save(path)
            self.status_var.set(f"Excel workbook saved: {path}")
            messagebox.showinfo(
                "Success",
                f"Generated {len(wb.sheetnames)} comparison sheets!\n\nSaved to: {path}",
            )
        except Exception as e:
            messagebox.showerror("Error", f"Error generating Excel: {e}")
            self.status_var.set("Error occurred")

    def write_comparison_to_sheet(self, ws, fighter1_stats, fighter2_stats):
        """Compatibility wrapper using legacy method name."""
        return self._write_comparison_to_sheet(ws, fighter1_stats, fighter2_stats)

    def _write_comparison_to_sheet(self, ws, fighter1_stats, fighter2_stats):
        current_row = 1

        ws.merge_cells(f"A{current_row}:C{current_row}")
        title_cell = ws[f"A{current_row}"]
        title_cell.value = "UFC FIGHTER COMPARISON"
        title_cell.font = Font(size=16, bold=True, color="FFFFFF")
        title_cell.alignment = Alignment(horizontal="left")
        title_cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        current_row += 2

        def add_section(title, data, start_row):
            ws.merge_cells(f"A{start_row}:C{start_row}")
            header_cell = ws[f"A{start_row}"]
            header_cell.value = title
            header_cell.font = Font(size=12, bold=True, color="FFFFFF")
            header_cell.fill = PatternFill(
                start_color="34495e", end_color="34495e", fill_type="solid"
            )
            header_cell.alignment = Alignment(horizontal="left")
            start_row += 1

            ws[f"A{start_row}"] = "Metric"
            ws[f"B{start_row}"] = fighter1_stats["name"]
            ws[f"C{start_row}"] = fighter2_stats["name"]
            for col in ["A", "B", "C"]:
                cell = ws[f"{col}{start_row}"]
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="bdc3c7", end_color="bdc3c7", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="left")
            start_row += 1

            for row_data in data:
                ws[f"A{start_row}"] = row_data[0]
                ws[f"B{start_row}"] = row_data[1]
                ws[f"C{start_row}"] = row_data[2]
                for col in ["A", "B", "C"]:
                    ws[f"{col}{start_row}"].alignment = Alignment(horizontal="left")
                start_row += 1

            return start_row + 1

        basic_data = [
            ["Name", fighter1_stats["name"], fighter2_stats["name"]],
            ["Total Fights", fighter1_stats["total_fights"], fighter2_stats["total_fights"]],
            ["Height", fighter1_stats["height"], fighter2_stats["height"]],
            ["Reach", fighter1_stats["reach"], fighter2_stats["reach"]],
            ["Stance", fighter1_stats["stance"], fighter2_stats["stance"]],
            ["Age", fighter1_stats["current_age"], fighter2_stats["current_age"]],
            ["Weight", fighter1_stats["weight"], fighter2_stats["weight"]],
        ]
        current_row = add_section("BASIC INFORMATION", basic_data, current_row)

        record_data = [
            ["Total Fights", fighter1_stats["total_fights"], fighter2_stats["total_fights"]],
            ["Wins", fighter1_stats["wins"], fighter2_stats["wins"]],
            ["Losses", fighter1_stats["losses"], fighter2_stats["losses"]],
            [
                "Win Rate",
                f"{fighter1_stats['win_rate'] * 100:.1f}%",
                f"{fighter2_stats['win_rate'] * 100:.1f}%",
            ],
        ]
        current_row = add_section("FIGHT RECORD", record_data, current_row)

        method_data = [
            [
                "KO/TKO",
                f"{fighter1_stats['wins_by_method']['KO/TKO']}-{fighter1_stats['losses_by_method']['KO/TKO']}",
                f"{fighter2_stats['wins_by_method']['KO/TKO']}-{fighter2_stats['losses_by_method']['KO/TKO']}",
            ],
            [
                "Submission",
                f"{fighter1_stats['wins_by_method']['Submission']}-{fighter1_stats['losses_by_method']['Submission']}",
                f"{fighter2_stats['wins_by_method']['Submission']}-{fighter2_stats['losses_by_method']['Submission']}",
            ],
            [
                "Decision",
                f"{fighter1_stats['wins_by_method']['Decision']}-{fighter1_stats['losses_by_method']['Decision']}",
                f"{fighter2_stats['wins_by_method']['Decision']}-{fighter2_stats['losses_by_method']['Decision']}",
            ],
        ]
        current_row = add_section("WINS & LOSSES BY METHOD (W-L)", method_data, current_row)

        f1_time_min = int(fighter1_stats["avg_fight_time_seconds"] // 60)
        f1_time_sec = int(fighter1_stats["avg_fight_time_seconds"] % 60)
        f2_time_min = int(fighter2_stats["avg_fight_time_seconds"] // 60)
        f2_time_sec = int(fighter2_stats["avg_fight_time_seconds"] % 60)

        career_data = [
            ["Avg Fight Time", f"{f1_time_min}:{f1_time_sec:02d}", f"{f2_time_min}:{f2_time_sec:02d}"],
            [
                "SLpM (Sig Strikes/Min)",
                f"{fighter1_stats['career_averages']['slpm']:.2f}",
                f"{fighter2_stats['career_averages']['slpm']:.2f}",
            ],
            [
                "Sig Strike Accuracy (%)",
                f"{fighter1_stats['career_averages']['sig_str_accuracy'] * 100:.1f}%",
                f"{fighter2_stats['career_averages']['sig_str_accuracy'] * 100:.1f}%",
            ],
            [
                "SApM (Sig Absorbed/Min)",
                f"{fighter1_stats['career_averages']['sapm']:.2f}",
                f"{fighter2_stats['career_averages']['sapm']:.2f}",
            ],
            [
                "Strike Defense (%)",
                f"{fighter1_stats['career_averages']['str_defense'] * 100:.1f}%",
                f"{fighter2_stats['career_averages']['str_defense'] * 100:.1f}%",
            ],
            [
                "TD per 15min",
                f"{fighter1_stats['career_averages']['td_avg_per_15min']:.2f}",
                f"{fighter2_stats['career_averages']['td_avg_per_15min']:.2f}",
            ],
            [
                "TD Accuracy (%)",
                f"{fighter1_stats['career_averages']['td_accuracy'] * 100:.1f}%",
                f"{fighter2_stats['career_averages']['td_accuracy'] * 100:.1f}%",
            ],
            [
                "TD Defense (%)",
                f"{fighter1_stats['career_averages']['td_defense'] * 100:.1f}%",
                f"{fighter2_stats['career_averages']['td_defense'] * 100:.1f}%",
            ],
            [
                "Sub per 15min",
                f"{fighter1_stats['career_averages']['sub_avg_per_15min']:.2f}",
                f"{fighter2_stats['career_averages']['sub_avg_per_15min']:.2f}",
            ],
        ]
        current_row = add_section("CAREER AVERAGES", career_data, current_row)

        strike_dist_data = [
            [
                "Head (Avg Landed/Absorbed)",
                f"{fighter1_stats['strike_distribution_avg_per_fight']['head_avg']:.1f} / {fighter1_stats['strike_distribution_absorbed_avg_per_fight']['head_absorbed_avg']:.1f}",
                f"{fighter2_stats['strike_distribution_avg_per_fight']['head_avg']:.1f} / {fighter2_stats['strike_distribution_absorbed_avg_per_fight']['head_absorbed_avg']:.1f}",
            ],
            [
                "Body (Avg Landed/Absorbed)",
                f"{fighter1_stats['strike_distribution_avg_per_fight']['body_avg']:.1f} / {fighter1_stats['strike_distribution_absorbed_avg_per_fight']['body_absorbed_avg']:.1f}",
                f"{fighter2_stats['strike_distribution_avg_per_fight']['body_avg']:.1f} / {fighter2_stats['strike_distribution_absorbed_avg_per_fight']['body_absorbed_avg']:.1f}",
            ],
            [
                "Leg (Avg Landed/Absorbed)",
                f"{fighter1_stats['strike_distribution_avg_per_fight']['leg_avg']:.1f} / {fighter1_stats['strike_distribution_absorbed_avg_per_fight']['leg_absorbed_avg']:.1f}",
                f"{fighter2_stats['strike_distribution_avg_per_fight']['leg_avg']:.1f} / {fighter2_stats['strike_distribution_absorbed_avg_per_fight']['leg_absorbed_avg']:.1f}",
            ],
            [
                "Distance (Avg Landed/Absorbed)",
                f"{fighter1_stats['strike_distribution_avg_per_fight']['distance_avg']:.1f} / {fighter1_stats['strike_distribution_absorbed_avg_per_fight']['distance_absorbed_avg']:.1f}",
                f"{fighter2_stats['strike_distribution_avg_per_fight']['distance_avg']:.1f} / {fighter2_stats['strike_distribution_absorbed_avg_per_fight']['distance_absorbed_avg']:.1f}",
            ],
            [
                "Clinch (Avg Landed/Absorbed)",
                f"{fighter1_stats['strike_distribution_avg_per_fight']['clinch_avg']:.1f} / {fighter1_stats['strike_distribution_absorbed_avg_per_fight']['clinch_absorbed_avg']:.1f}",
                f"{fighter2_stats['strike_distribution_avg_per_fight']['clinch_avg']:.1f} / {fighter2_stats['strike_distribution_absorbed_avg_per_fight']['clinch_absorbed_avg']:.1f}",
            ],
            [
                "Ground (Avg Landed/Absorbed)",
                f"{fighter1_stats['strike_distribution_avg_per_fight']['ground_avg']:.1f} / {fighter1_stats['strike_distribution_absorbed_avg_per_fight']['ground_absorbed_avg']:.1f}",
                f"{fighter2_stats['strike_distribution_avg_per_fight']['ground_avg']:.1f} / {fighter2_stats['strike_distribution_absorbed_avg_per_fight']['ground_absorbed_avg']:.1f}",
            ],
        ]
        current_row = add_section("STRIKE DISTRIBUTION (PER FIGHT)", strike_dist_data, current_row)

        grappling_data = [
            [
                "Takedowns (Avg Landed/Absorbed)",
                f"{fighter1_stats['avg_takedowns_per_fight']:.2f} / {fighter1_stats['avg_takedowns_against_per_fight']:.2f}",
                f"{fighter2_stats['avg_takedowns_per_fight']:.2f} / {fighter2_stats['avg_takedowns_against_per_fight']:.2f}",
            ],
            [
                "Submissions (Avg Landed/Absorbed)",
                f"{fighter1_stats['avg_submission_attempts_per_fight']:.2f} / {fighter1_stats['avg_submission_attempts_against_per_fight']:.2f}",
                f"{fighter2_stats['avg_submission_attempts_per_fight']:.2f} / {fighter2_stats['avg_submission_attempts_against_per_fight']:.2f}",
            ],
            [
                "Reversals (Avg Landed/Absorbed)",
                f"{fighter1_stats['avg_reversals_per_fight']:.2f} / {fighter1_stats['avg_reversals_against_per_fight']:.2f}",
                f"{fighter2_stats['avg_reversals_per_fight']:.2f} / {fighter2_stats['avg_reversals_against_per_fight']:.2f}",
            ],
            [
                "Control Time (Avg For/Against sec)",
                f"{fighter1_stats['avg_control_time_per_fight']:.1f} / {fighter1_stats['avg_control_time_against_per_fight']:.1f}",
                f"{fighter2_stats['avg_control_time_per_fight']:.1f} / {fighter2_stats['avg_control_time_against_per_fight']:.1f}",
            ],
        ]
        current_row = add_section("GRAPPLING STATISTICS (PER FIGHT)", grappling_data, current_row)

        knockdown_data = [
            ["Total Landed", f"{fighter1_stats['total_knockdowns_landed']:.0f}", f"{fighter2_stats['total_knockdowns_landed']:.0f}"],
            ["Avg Landed", f"{fighter1_stats['avg_knockdowns_per_fight']:.2f}", f"{fighter2_stats['avg_knockdowns_per_fight']:.2f}"],
            ["Total Against", f"{fighter1_stats['total_knockdowns_against']:.0f}", f"{fighter2_stats['total_knockdowns_against']:.0f}"],
            ["Avg Against", f"{fighter1_stats['avg_knockdowns_against_per_fight']:.2f}", f"{fighter2_stats['avg_knockdowns_against_per_fight']:.2f}"],
        ]
        current_row = add_section("KNOCKDOWN STATISTICS", knockdown_data, current_row)

        current_row += 1

        def add_fight_history(fighter_name, fight_history, start_row):
            ws.merge_cells(f"A{start_row}:E{start_row}")
            header_cell = ws[f"A{start_row}"]
            header_cell.value = f"UFC FIGHT HISTORY - {fighter_name}"
            header_cell.font = Font(size=12, bold=True, color="FFFFFF")
            header_cell.fill = PatternFill(
                start_color="34495e", end_color="34495e", fill_type="solid"
            )
            header_cell.alignment = Alignment(horizontal="left")
            start_row += 1

            ws[f"A{start_row}"] = "Date"
            ws[f"B{start_row}"] = "Opponent"
            ws[f"C{start_row}"] = "Result"
            ws[f"D{start_row}"] = "Method"
            ws[f"E{start_row}"] = "Round"
            for col in ["A", "B", "C", "D", "E"]:
                cell = ws[f"{col}{start_row}"]
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="bdc3c7", end_color="bdc3c7", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="left")
            start_row += 1

            for fight in reversed(fight_history):
                date_str = fight["date"].strftime("%Y-%m-%d") if pd.notna(fight["date"]) else "N/A"
                ws[f"A{start_row}"] = date_str
                ws[f"B{start_row}"] = fight["opponent"]
                ws[f"C{start_row}"] = fight["result"]
                ws[f"D{start_row}"] = fight["method"]
                ws[f"E{start_row}"] = f"R{fight['round']}"

                for col in ["A", "B", "C", "D", "E"]:
                    ws[f"{col}{start_row}"].alignment = Alignment(horizontal="left")

                result_cell = ws[f"C{start_row}"]
                if fight["result"] == "Win":
                    result_cell.fill = PatternFill(
                        start_color="d4edda", end_color="d4edda", fill_type="solid"
                    )
                elif fight["result"] == "Loss":
                    result_cell.fill = PatternFill(
                        start_color="f8d7da", end_color="f8d7da", fill_type="solid"
                    )
                start_row += 1

            return start_row + 1

        current_row = add_fight_history(
            fighter1_stats["name"], fighter1_stats["fight_history"], current_row
        )
        current_row += 1
        add_fight_history(fighter2_stats["name"], fighter2_stats["fight_history"], current_row)

        for col in ["A", "B", "C", "D", "E"]:
            max_length = 0
            column = ws[col]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col].width = adjusted_width

    def _get_fighter_data(self, fighter_name):
        red_fights = self.df[self.df["r_fighter"].str.contains(fighter_name, case=False, na=False)]
        blue_fights = self.df[self.df["b_fighter"].str.contains(fighter_name, case=False, na=False)]
        if red_fights.empty and blue_fights.empty:
            return None

        fighter_fights = []
        for _, fight in red_fights.iterrows():
            fighter_fights.append(
                {
                    "date": fight["event_date"],
                    "opponent": fight["b_fighter"],
                    "result": "Win"
                    if fight["winner"] == "Red"
                    else "Loss" if fight["winner"] == "Blue" else "Draw",
                    "method": fight["method"],
                    "round": fight["finish_round"],
                    "total_time": fight["total_fight_time_sec"],
                    "is_red": True,
                    "height": fight["r_height"],
                    "reach": fight["r_reach"],
                    "stance": fight["r_stance"],
                    "weight": fight["r_weight"],
                    "age_at_fight": fight["r_age_at_event"],
                    "date_of_birth": fight["r_date_of_birth"],
                    "sig_str_landed": fight["r_sig_str"],
                    "sig_str_attempted": fight["r_sig_str_att"],
                    "sig_str_absorbed": fight["b_sig_str"],
                    "sig_str_absorbed_attempted": fight["b_sig_str_att"],
                    "knockdowns": fight["r_kd"],
                    "head_pct": fight["r_head"],
                    "body_pct": fight["r_body"],
                    "leg_pct": fight["r_leg"],
                    "distance_pct": fight["r_distance"],
                    "clinch_pct": fight["r_clinch"],
                    "ground_pct": fight["r_ground"],
                    "takedowns_landed": fight["r_td"],
                    "takedowns_attempted": fight["r_td_att"],
                    "takedowns_absorbed": fight["b_td"],
                    "takedowns_absorbed_attempted": fight["b_td_att"],
                    "submission_attempts": fight["r_sub_att"],
                    "reversals": fight["r_rev"],
                    "control_time": fight["r_ctrl_sec"],
                }
            )

        for _, fight in blue_fights.iterrows():
            fighter_fights.append(
                {
                    "date": fight["event_date"],
                    "opponent": fight["r_fighter"],
                    "result": "Win"
                    if fight["winner"] == "Blue"
                    else "Loss" if fight["winner"] == "Red" else "Draw",
                    "method": fight["method"],
                    "round": fight["finish_round"],
                    "total_time": fight["total_fight_time_sec"],
                    "is_red": False,
                    "height": fight["b_height"],
                    "reach": fight["b_reach"],
                    "stance": fight["b_stance"],
                    "weight": fight["b_weight"],
                    "age_at_fight": fight["b_age_at_event"],
                    "date_of_birth": fight["b_date_of_birth"],
                    "sig_str_landed": fight["b_sig_str"],
                    "sig_str_attempted": fight["b_sig_str_att"],
                    "sig_str_absorbed": fight["r_sig_str"],
                    "sig_str_absorbed_attempted": fight["r_sig_str_att"],
                    "knockdowns": fight["b_kd"],
                    "head_pct": fight["b_head"],
                    "body_pct": fight["b_body"],
                    "leg_pct": fight["b_leg"],
                    "distance_pct": fight["b_distance"],
                    "clinch_pct": fight["b_clinch"],
                    "ground_pct": fight["b_ground"],
                    "takedowns_landed": fight["b_td"],
                    "takedowns_attempted": fight["b_td_att"],
                    "takedowns_absorbed": fight["r_td"],
                    "takedowns_absorbed_attempted": fight["r_td_att"],
                    "submission_attempts": fight["b_sub_att"],
                    "reversals": fight["b_rev"],
                    "control_time": fight["b_ctrl_sec"],
                }
            )

        fighter_df = pd.DataFrame(fighter_fights).sort_values("date").reset_index(drop=True)
        if fighter_df.empty:
            return None
        return self._calculate_career_stats(fighter_df, fighter_name)

    def _calculate_career_stats(self, fighter_df, fighter_name):
        stats = {
            "name": fighter_name,
            "total_fights": len(fighter_df),
            "wins": len(fighter_df[fighter_df["result"] == "Win"]),
            "losses": len(fighter_df[fighter_df["result"] == "Loss"]),
            "draws": len(fighter_df[fighter_df["result"] == "Draw"]),
        }

        latest_fight = fighter_df.iloc[-1]
        stats["height"] = latest_fight["height"]
        stats["reach"] = latest_fight["reach"]
        stats["stance"] = latest_fight["stance"]
        stats["weight"] = latest_fight["weight"]

        if pd.notna(latest_fight["date_of_birth"]):
            try:
                dob = pd.to_datetime(latest_fight["date_of_birth"])
                stats["current_age"] = (datetime.now() - dob).days // 365
            except (ValueError, TypeError):
                stats["current_age"] = latest_fight["age_at_fight"] + (
                    datetime.now().year - latest_fight["date"].year
                )
        else:
            stats["current_age"] = latest_fight["age_at_fight"] + (
                datetime.now().year - latest_fight["date"].year
            )

        wins_df = fighter_df[fighter_df["result"] == "Win"]
        losses_df = fighter_df[fighter_df["result"] == "Loss"]

        def normalize_method(m):
            m = str(m) if pd.notna(m) else ""
            if "Decision" in m:
                return "Decision"
            if "KO/TKO" in m or "TKO" in m:
                return "KO/TKO"
            if "Submission" in m:
                return "Submission"
            return m

        stats["wins_by_method"] = wins_df["method"].apply(normalize_method).value_counts().to_dict()
        stats["losses_by_method"] = losses_df["method"].apply(normalize_method).value_counts().to_dict()
        for method in ["KO/TKO", "Submission", "Decision"]:
            if method not in stats["wins_by_method"]:
                stats["wins_by_method"][method] = 0
            if method not in stats["losses_by_method"]:
                stats["losses_by_method"][method] = 0

        stats["strike_distribution_avg_per_fight"] = {
            "head_avg": (fighter_df["sig_str_landed"] * fighter_df["head_pct"]).sum() / stats["total_fights"],
            "body_avg": (fighter_df["sig_str_landed"] * fighter_df["body_pct"]).sum() / stats["total_fights"],
            "leg_avg": (fighter_df["sig_str_landed"] * fighter_df["leg_pct"]).sum() / stats["total_fights"],
            "distance_avg": (fighter_df["sig_str_landed"] * fighter_df["distance_pct"]).sum() / stats["total_fights"],
            "clinch_avg": (fighter_df["sig_str_landed"] * fighter_df["clinch_pct"]).sum() / stats["total_fights"],
            "ground_avg": (fighter_df["sig_str_landed"] * fighter_df["ground_pct"]).sum() / stats["total_fights"],
        }

        absorbed_dist = {
            "head_absorbed": 0,
            "body_absorbed": 0,
            "leg_absorbed": 0,
            "distance_absorbed": 0,
            "clinch_absorbed": 0,
            "ground_absorbed": 0,
        }
        for _, fight in fighter_df.iterrows():
            if fight["is_red"]:
                opponent_fight = self.df[
                    (self.df["r_fighter"] == fighter_name)
                    & (self.df["b_fighter"] == fight["opponent"])
                    & (self.df["event_date"] == fight["date"])
                ]
                if not opponent_fight.empty:
                    opp_sig_str = opponent_fight.iloc[0]["b_sig_str"]
                    absorbed_dist["head_absorbed"] += opp_sig_str * opponent_fight.iloc[0]["b_head"]
                    absorbed_dist["body_absorbed"] += opp_sig_str * opponent_fight.iloc[0]["b_body"]
                    absorbed_dist["leg_absorbed"] += opp_sig_str * opponent_fight.iloc[0]["b_leg"]
                    absorbed_dist["distance_absorbed"] += opp_sig_str * opponent_fight.iloc[0]["b_distance"]
                    absorbed_dist["clinch_absorbed"] += opp_sig_str * opponent_fight.iloc[0]["b_clinch"]
                    absorbed_dist["ground_absorbed"] += opp_sig_str * opponent_fight.iloc[0]["b_ground"]
            else:
                opponent_fight = self.df[
                    (self.df["b_fighter"] == fighter_name)
                    & (self.df["r_fighter"] == fight["opponent"])
                    & (self.df["event_date"] == fight["date"])
                ]
                if not opponent_fight.empty:
                    opp_sig_str = opponent_fight.iloc[0]["r_sig_str"]
                    absorbed_dist["head_absorbed"] += opp_sig_str * opponent_fight.iloc[0]["r_head"]
                    absorbed_dist["body_absorbed"] += opp_sig_str * opponent_fight.iloc[0]["r_body"]
                    absorbed_dist["leg_absorbed"] += opp_sig_str * opponent_fight.iloc[0]["r_leg"]
                    absorbed_dist["distance_absorbed"] += opp_sig_str * opponent_fight.iloc[0]["r_distance"]
                    absorbed_dist["clinch_absorbed"] += opp_sig_str * opponent_fight.iloc[0]["r_clinch"]
                    absorbed_dist["ground_absorbed"] += opp_sig_str * opponent_fight.iloc[0]["r_ground"]

        stats["strike_distribution_absorbed_avg_per_fight"] = {
            "head_absorbed_avg": absorbed_dist["head_absorbed"] / stats["total_fights"],
            "body_absorbed_avg": absorbed_dist["body_absorbed"] / stats["total_fights"],
            "leg_absorbed_avg": absorbed_dist["leg_absorbed"] / stats["total_fights"],
            "distance_absorbed_avg": absorbed_dist["distance_absorbed"] / stats["total_fights"],
            "clinch_absorbed_avg": absorbed_dist["clinch_absorbed"] / stats["total_fights"],
            "ground_absorbed_avg": absorbed_dist["ground_absorbed"] / stats["total_fights"],
        }

        # Rebuild dropped leak-prone pro_* features from full career totals.
        total_fight_minutes = max(fighter_df["total_time"].sum() / 60.0, 1e-6)
        total_sig_landed = fighter_df["sig_str_landed"].sum()
        total_sig_attempted = fighter_df["sig_str_attempted"].sum()
        total_sig_absorbed = fighter_df["sig_str_absorbed"].sum()
        total_sig_absorbed_attempted = fighter_df["sig_str_absorbed_attempted"].sum()
        total_td_landed = fighter_df["takedowns_landed"].sum()
        total_td_attempted = fighter_df["takedowns_attempted"].sum()
        total_td_absorbed = fighter_df["takedowns_absorbed"].sum()
        total_td_absorbed_attempted = fighter_df["takedowns_absorbed_attempted"].sum()
        total_sub_attempts = fighter_df["submission_attempts"].sum()
        stats["career_averages"] = {
            "slpm": total_sig_landed / total_fight_minutes,
            "sig_str_accuracy": total_sig_landed / max(total_sig_attempted, 1e-6),
            "sapm": total_sig_absorbed / total_fight_minutes,
            "str_defense": 1.0 - (total_sig_absorbed / max(total_sig_absorbed_attempted, 1e-6)),
            "td_avg_per_15min": total_td_landed * 15.0 / total_fight_minutes,
            "td_accuracy": total_td_landed / max(total_td_attempted, 1e-6),
            "td_defense": 1.0 - (total_td_absorbed / max(total_td_absorbed_attempted, 1e-6)),
            "sub_avg_per_15min": total_sub_attempts * 15.0 / total_fight_minutes,
        }

        stats["avg_takedowns_per_fight"] = fighter_df["takedowns_landed"].sum() / stats["total_fights"]
        stats["avg_submission_attempts_per_fight"] = (
            fighter_df["submission_attempts"].sum() / stats["total_fights"]
        )
        stats["avg_reversals_per_fight"] = fighter_df["reversals"].sum() / stats["total_fights"]
        stats["avg_control_time_per_fight"] = fighter_df["control_time"].sum() / stats["total_fights"]

        total_takedowns_against = 0
        total_submissions_against = 0
        total_reversals_against = 0
        total_control_time_against = 0
        total_knockdowns_against = 0
        for _, fight in fighter_df.iterrows():
            if fight["is_red"]:
                opponent_fight = self.df[
                    (self.df["r_fighter"] == fighter_name)
                    & (self.df["b_fighter"] == fight["opponent"])
                    & (self.df["event_date"] == fight["date"])
                ]
                if not opponent_fight.empty:
                    total_takedowns_against += opponent_fight.iloc[0]["b_td"]
                    total_submissions_against += opponent_fight.iloc[0]["b_sub_att"]
                    total_reversals_against += opponent_fight.iloc[0]["b_rev"]
                    total_control_time_against += opponent_fight.iloc[0]["b_ctrl_sec"]
                    total_knockdowns_against += opponent_fight.iloc[0]["b_kd"]
            else:
                opponent_fight = self.df[
                    (self.df["b_fighter"] == fighter_name)
                    & (self.df["r_fighter"] == fight["opponent"])
                    & (self.df["event_date"] == fight["date"])
                ]
                if not opponent_fight.empty:
                    total_takedowns_against += opponent_fight.iloc[0]["r_td"]
                    total_submissions_against += opponent_fight.iloc[0]["r_sub_att"]
                    total_reversals_against += opponent_fight.iloc[0]["r_rev"]
                    total_control_time_against += opponent_fight.iloc[0]["r_ctrl_sec"]
                    total_knockdowns_against += opponent_fight.iloc[0]["r_kd"]

        stats["avg_takedowns_against_per_fight"] = total_takedowns_against / stats["total_fights"]
        stats["avg_submission_attempts_against_per_fight"] = (
            total_submissions_against / stats["total_fights"]
        )
        stats["avg_reversals_against_per_fight"] = total_reversals_against / stats["total_fights"]
        stats["avg_control_time_against_per_fight"] = total_control_time_against / stats["total_fights"]

        total_knockdowns_landed = fighter_df["knockdowns"].sum()
        stats["total_knockdowns_landed"] = total_knockdowns_landed
        stats["avg_knockdowns_per_fight"] = total_knockdowns_landed / stats["total_fights"]
        stats["total_knockdowns_against"] = total_knockdowns_against
        stats["avg_knockdowns_against_per_fight"] = total_knockdowns_against / stats["total_fights"]

        stats["win_rate"] = stats["wins"] / stats["total_fights"] if stats["total_fights"] > 0 else 0
        stats["avg_fight_time_seconds"] = fighter_df["total_time"].sum() / stats["total_fights"]
        stats["fight_history"] = fighter_df.to_dict("records")

        return stats


def main():
    root = tk.Tk()
    MatchupDashboard(root)
    root.mainloop()


if __name__ == "__main__":
    main()
