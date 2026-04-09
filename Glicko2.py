"""
Standalone Glicko-2 rating system for UFC fighters.

Processes all historical fights from pure_fight_data.csv chronologically,
then predicts win probabilities for new matchups via a tkinter GUI.
Exports predictions and per-weight-class top 25 rankings to Excel.

Matchup input formats (one per line):
  Fighter A,Fighter B,Weight Class,Gender,Rounds
  Fighter A,Fighter B
"""

import csv
import math
import os
import threading
import tkinter as tk
from tkinter import messagebox
from collections import defaultdict
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# ─── Glicko-2 parameters (same as Model.py) ─────────────────────────────────

MU_0 = 1500.0
PHI_0 = 200.0
SIGMA_0 = 0.06
TAU = 0.5
SCALE = 173.7178
CONVERGENCE = 1e-6

# A fighter is "active" if they fought within this many days of the latest event
ACTIVE_DAYS = 730  # ~2 years


# ─── Glicko-2 core ──────────────────────────────────────────────────────────

def _g(phi):
    return 1.0 / math.sqrt(1.0 + 3.0 * phi ** 2 / math.pi ** 2)


def _E(mu, mu_j, phi_j):
    return 1.0 / (1.0 + math.exp(-_g(phi_j) * (mu - mu_j)))


def glicko2_update(rating, opponents):
    mu, phi, sigma = rating
    mu_s = (mu - MU_0) / SCALE
    phi_s = phi / SCALE

    if not opponents:
        phi_star = math.sqrt(phi_s ** 2 + sigma ** 2)
        return (mu, phi_star * SCALE, sigma)

    v_inv = 0.0
    delta_sum = 0.0
    for opp_r, opp_rd, score in opponents:
        mu_j = (opp_r - MU_0) / SCALE
        phi_j = opp_rd / SCALE
        g_j = _g(phi_j)
        E_j = _E(mu_s, mu_j, phi_j)
        v_inv += g_j ** 2 * E_j * (1.0 - E_j)
        delta_sum += g_j * (score - E_j)

    v = 1.0 / v_inv if v_inv > 0 else 1e6
    delta = v * delta_sum

    a = math.log(sigma ** 2)

    def f(x):
        ex = math.exp(x)
        num = ex * (delta ** 2 - phi_s ** 2 - v - ex)
        den = 2.0 * (phi_s ** 2 + v + ex) ** 2
        return num / den - (x - a) / (TAU ** 2)

    A = a
    if delta ** 2 > phi_s ** 2 + v:
        B = math.log(delta ** 2 - phi_s ** 2 - v)
    else:
        k = 1
        while f(a - k * TAU) < 0:
            k += 1
        B = a - k * TAU

    fA, fB = f(A), f(B)
    for _ in range(100):
        C = A + (A - B) * fA / (fB - fA)
        fC = f(C)
        if fC * fB < 0:
            A, fA = B, fB
        else:
            fA /= 2.0
        B, fB = C, fC
        if abs(B - A) < CONVERGENCE:
            break

    new_sigma = math.exp(A / 2.0)
    phi_star = math.sqrt(phi_s ** 2 + new_sigma ** 2)
    new_phi_s = 1.0 / math.sqrt(1.0 / phi_star ** 2 + 1.0 / v)
    new_mu_s = mu_s + new_phi_s ** 2 * delta_sum

    return (
        new_mu_s * SCALE + MU_0,
        new_phi_s * SCALE,
        new_sigma,
    )


def win_probability(rating_a, rating_b):
    mu_a = (rating_a[0] - MU_0) / SCALE
    mu_b = (rating_b[0] - MU_0) / SCALE
    phi_a = rating_a[1] / SCALE
    phi_b = rating_b[1] / SCALE
    combined_phi = math.sqrt(phi_a ** 2 + phi_b ** 2)
    return _E(mu_a, mu_b, combined_phi)


# ─── Data loading ────────────────────────────────────────────────────────────

def build_ratings(csv_path):
    ratings = {}
    fight_counts = defaultdict(int)
    records = defaultdict(lambda: [0, 0, 0])
    last_fight_date = {}
    fighter_division = {}  # fighter -> (weight_class, gender)

    def get(fighter):
        if fighter not in ratings:
            ratings[fighter] = (MU_0, PHI_0, SIGMA_0)
        return ratings[fighter]

    fights = []
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            fights.append(row)

    fights.sort(key=lambda r: datetime.strptime(r["event_date"], "%m/%d/%Y"))

    for row in fights:
        r_name = row["r_name"].strip()
        b_name = row["b_name"].strip()
        winner = row["winner"].strip()
        event_date = datetime.strptime(row["event_date"], "%m/%d/%Y")
        weight_class = row["weight_class"].strip()
        gender = row["gender"].strip()

        r_rating = get(r_name)
        b_rating = get(b_name)

        if winner == "Red":
            r_score, b_score = 1.0, 0.0
            records[r_name][0] += 1
            records[b_name][1] += 1
        elif winner == "Blue":
            r_score, b_score = 0.0, 1.0
            records[b_name][0] += 1
            records[r_name][1] += 1
        else:
            r_score, b_score = 0.5, 0.5
            records[r_name][2] += 1
            records[b_name][2] += 1

        ratings[r_name] = glicko2_update(r_rating, [(b_rating[0], b_rating[1], r_score)])
        ratings[b_name] = glicko2_update(b_rating, [(r_rating[0], r_rating[1], b_score)])

        fight_counts[r_name] += 1
        fight_counts[b_name] += 1

        last_fight_date[r_name] = event_date
        last_fight_date[b_name] = event_date

        # Track most recent division (skip Catch Weight / Open Weight)
        if weight_class not in ("Catch Weight", "Open Weight"):
            division = _division_key(weight_class, gender)
            fighter_division[r_name] = division
            fighter_division[b_name] = division

    return ratings, fight_counts, records, last_fight_date, fighter_division


def _division_key(weight_class, gender):
    """Normalize division name so women's divisions are always prefixed."""
    if gender == "Women" and not weight_class.startswith("Women's"):
        return f"Women's {weight_class}"
    return weight_class


# Canonical ordering for division sheets
DIVISION_ORDER = [
    "Heavyweight",
    "Light Heavyweight",
    "Middleweight",
    "Welterweight",
    "Lightweight",
    "Featherweight",
    "Bantamweight",
    "Flyweight",
    "Women's Featherweight",
    "Women's Bantamweight",
    "Women's Flyweight",
    "Women's Strawweight",
]


def fuzzy_find(name, ratings):
    if name in ratings:
        return name
    lower = name.lower()
    for key in ratings:
        if key.lower() == lower:
            return key
    matches = [k for k in ratings if lower in k.lower()]
    if len(matches) == 1:
        return matches[0]
    return None


def format_record(records, name):
    w, l, d = records.get(name, [0, 0, 0])
    return f"{w}-{l}-{d}" if d > 0 else f"{w}-{l}"


# ─── Excel export ────────────────────────────────────────────────────────────

def _auto_width(ws):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3


def export_to_excel(output_path, predictions, ratings, records,
                    last_fight_date, fighter_division):
    wb = Workbook()

    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True, size=11, color="000000")
    header_align = Alignment(horizontal="left", vertical="center")
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    pct_fmt = "0.0%"

    # ── Sheet 1: Predictions ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Predictions"

    headers = ["Fighter A", "Fighter B", "Rating A", "Rating B",
               "Record A", "Record B", "Fighter A %", "Fighter B %",
               "Predicted Winner", "Win %"]

    for col_idx, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    for row_idx, pred in enumerate(predictions, 2):
        winner = pred["name_a"] if pred["prob_a"] >= pred["prob_b"] else pred["name_b"]
        win_pct = max(pred["prob_a"], pred["prob_b"])

        row_data = [
            pred["name_a"], pred["name_b"],
            round(pred["rating_a"]), round(pred["rating_b"]),
            pred["rec_a"], pred["rec_b"],
            pred["prob_a"], pred["prob_b"],
            winner, win_pct,
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border
            if col_idx in (7, 8, 10):
                cell.number_format = pct_fmt

    _auto_width(ws)

    # ── Per-division ranking sheets ──────────────────────────────────────────
    latest_date = max(last_fight_date.values()) if last_fight_date else datetime.now()
    active_cutoff = latest_date - timedelta(days=ACTIVE_DAYS)

    # Group active fighters by division
    division_fighters = defaultdict(list)
    for fighter, division in fighter_division.items():
        if last_fight_date.get(fighter, datetime.min) >= active_cutoff:
            mu, phi, sigma = ratings[fighter]
            rec = format_record(records, fighter)
            division_fighters[division].append((fighter, mu, phi, rec))

    # Sort each division by rating descending
    for div in division_fighters:
        division_fighters[div].sort(key=lambda x: x[1], reverse=True)

    rank_headers = ["#", "Fighter", "Rating", "RD", "Record"]

    for div_name in DIVISION_ORDER:
        if div_name not in division_fighters:
            continue

        fighters = division_fighters[div_name][:25]
        # Sheet name max 31 chars
        sheet_name = div_name[:31]
        ws_div = wb.create_sheet(title=sheet_name)

        for col_idx, hdr in enumerate(rank_headers, 1):
            cell = ws_div.cell(row=1, column=col_idx, value=hdr)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = border

        for i, (fighter, mu, phi, rec) in enumerate(fighters, 1):
            row_data = [i, fighter, round(mu), round(phi), rec]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws_div.cell(row=i + 1, column=col_idx, value=val)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = border

        _auto_width(ws_div)

    wb.save(output_path)


# ─── GUI ─────────────────────────────────────────────────────────────────────

BG = "#1a1a2e"
BG_HEADER = "#16213e"
BG_INPUT = "#0f3460"
FG = "#e0e0e0"
ACCENT = "#e94560"
ACCENT2 = "#533483"
MUTED = "#a0c4ff"
GREEN = "#4ade80"
BAR_A = "#e94560"
BAR_B = "#3b82f6"


class Glicko2GUI:
    def __init__(self, root):
        self.root = root
        self.root.title("UFC GLICKO-2 RATINGS")
        self.root.geometry("920x760")
        self.root.resizable(True, True)
        self.root.configure(bg=BG)

        self.ratings = None
        self.fight_counts = None
        self.records = None
        self.last_fight_date = None
        self.fighter_division = None

        self._build_ui()
        self._load_ratings()

    def _build_ui(self):
        # Title bar
        title_frame = tk.Frame(self.root, bg=BG_HEADER, pady=12)
        title_frame.pack(fill="x")
        tk.Label(
            title_frame,
            text="UFC GLICKO-2 RATINGS",
            font=("Helvetica", 20, "bold"),
            fg=ACCENT,
            bg=BG_HEADER,
        ).pack()

        main = tk.Frame(self.root, bg=BG, padx=16, pady=10)
        main.pack(fill="both", expand=True)

        # Status label
        self.status_var = tk.StringVar(value="Loading ratings...")
        self.status_label = tk.Label(
            main,
            textvariable=self.status_var,
            bg=BG,
            fg=MUTED,
            font=("Helvetica", 9, "italic"),
        )
        self.status_label.pack(anchor="w", pady=(0, 4))

        # Input label
        tk.Label(
            main,
            text="Enter fights (one per line: Fighter A, Fighter B)",
            bg=BG,
            fg=MUTED,
            font=("Helvetica", 9, "italic"),
        ).pack(anchor="w", pady=(4, 2))

        # Fight input
        self.fight_input = tk.Text(
            main,
            height=10,
            font=("Courier New", 10),
            bg=BG_INPUT,
            fg="white",
            insertbackground="white",
            relief="flat",
            wrap="word",
        )
        self.fight_input.pack(fill="x", pady=4)

        # Buttons
        btn_frame = tk.Frame(main, bg=BG)
        btn_frame.pack(fill="x", pady=6)

        tk.Button(
            btn_frame,
            text="Load from Fights.csv",
            command=self._load_fights_csv,
            font=("Helvetica", 10, "bold"),
            bg=ACCENT2,
            fg="white",
            relief="flat",
            padx=14,
            cursor="hand2",
        ).pack(side="left", padx=4)

        tk.Button(
            btn_frame,
            text="Clear",
            command=self._clear,
            font=("Helvetica", 10, "bold"),
            bg="#444466",
            fg="white",
            relief="flat",
            padx=14,
            cursor="hand2",
        ).pack(side="left", padx=4)

        self.predict_btn = tk.Button(
            btn_frame,
            text="Predict",
            command=self._predict,
            font=("Helvetica", 11, "bold"),
            bg=ACCENT,
            fg="white",
            relief="flat",
            padx=20,
            cursor="hand2",
            state="disabled",
        )
        self.predict_btn.pack(side="right", padx=4)

        # Results area
        results_frame = tk.Frame(main, bg=BG)
        results_frame.pack(fill="both", expand=True, pady=(6, 0))

        scrollbar = tk.Scrollbar(results_frame)
        scrollbar.pack(side="right", fill="y")

        self.results_canvas = tk.Canvas(
            results_frame,
            bg=BG,
            highlightthickness=0,
            yscrollcommand=scrollbar.set,
        )
        self.results_canvas.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=self.results_canvas.yview)

        self.results_inner = tk.Frame(self.results_canvas, bg=BG)
        self.results_canvas.create_window((0, 0), window=self.results_inner, anchor="nw")
        self.results_inner.bind(
            "<Configure>",
            lambda _: self.results_canvas.configure(
                scrollregion=self.results_canvas.bbox("all")),
        )
        self.results_canvas.bind_all(
            "<MouseWheel>",
            lambda e: self.results_canvas.yview_scroll(-1 * (e.delta // 120), "units"),
        )

    def _load_ratings(self):
        def _do():
            data_path = os.path.join(SCRIPT_DIR, "pure_fight_data.csv")
            if not os.path.exists(data_path):
                self.status_var.set(f"Error: {data_path} not found")
                return
            (self.ratings, self.fight_counts, self.records,
             self.last_fight_date, self.fighter_division) = build_ratings(data_path)
            n_fights = sum(self.fight_counts.values()) // 2
            self.status_var.set(
                f"Loaded {n_fights} fights, {len(self.ratings)} fighters rated."
            )
            self.predict_btn.config(state="normal")

        threading.Thread(target=_do, daemon=True).start()

    def _load_fights_csv(self):
        fights_path = os.path.join(SCRIPT_DIR, "Fights.csv")
        if not os.path.exists(fights_path):
            self.status_var.set("Fights.csv not found in script directory.")
            return
        self.fight_input.delete("1.0", tk.END)
        with open(fights_path, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader, None)
            for row in reader:
                if not row or not row[0].strip():
                    continue
                a = row[0].strip()
                b = row[1].strip() if len(row) > 1 else ""
                if a and b:
                    self.fight_input.insert(tk.END, f"{a},{b}\n")

    def _clear(self):
        self.fight_input.delete("1.0", tk.END)
        for w in self.results_inner.winfo_children():
            w.destroy()

    def _predict(self):
        if self.ratings is None:
            return

        text = self.fight_input.get("1.0", tk.END).strip()
        if not text:
            self.status_var.set("Enter at least one matchup.")
            return

        # Clear previous results
        for w in self.results_inner.winfo_children():
            w.destroy()

        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
        warnings = []
        predictions = []

        for line in lines:
            parts = [p.strip() for p in line.split(",")]
            if len(parts) < 2:
                continue
            a_name, b_name = parts[0], parts[1]

            a_key = fuzzy_find(a_name, self.ratings)
            b_key = fuzzy_find(b_name, self.ratings)

            a_rating = self.ratings[a_key] if a_key else (MU_0, PHI_0, SIGMA_0)
            b_rating = self.ratings[b_key] if b_key else (MU_0, PHI_0, SIGMA_0)
            a_display = a_key or a_name
            b_display = b_key or b_name
            a_rec = format_record(self.records, a_key) if a_key else "0-0"
            b_rec = format_record(self.records, b_key) if b_key else "0-0"

            if not a_key:
                warnings.append(f"{a_name}: not found, using default 1500")
            if not b_key:
                warnings.append(f"{b_name}: not found, using default 1500")

            prob_a = win_probability(a_rating, b_rating)
            prob_b = 1.0 - prob_a

            predictions.append({
                "name_a": a_display, "name_b": b_display,
                "rating_a": a_rating[0], "rating_b": b_rating[0],
                "rec_a": a_rec, "rec_b": b_rec,
                "prob_a": prob_a, "prob_b": prob_b,
            })

            self._add_matchup_card(
                a_display, b_display,
                a_rating[0], b_rating[0],
                a_rec, b_rec,
                prob_a, prob_b,
            )

        # Export to Excel
        output_path = os.path.join(SCRIPT_DIR, "Glicko2_Predictions.xlsx")
        try:
            export_to_excel(
                output_path, predictions, self.ratings, self.records,
                self.last_fight_date, self.fighter_division,
            )
            export_msg = f"Saved to {os.path.basename(output_path)}"
        except Exception as e:
            export_msg = f"Excel export failed: {e}"
            messagebox.showerror("Export Error", str(e))

        if warnings:
            self.status_var.set(
                f"{len(predictions)} matchups | {export_msg} | "
                f"Warnings: {'; '.join(warnings[:3])}"
            )
        else:
            self.status_var.set(f"{len(predictions)} matchups predicted. {export_msg}")

    def _add_matchup_card(self, name_a, name_b, rating_a, rating_b,
                          rec_a, rec_b, prob_a, prob_b):
        card = tk.Frame(self.results_inner, bg=BG_HEADER, pady=8, padx=12)
        card.pack(fill="x", pady=4, padx=4)

        # Fighter names + ratings row
        names_frame = tk.Frame(card, bg=BG_HEADER)
        names_frame.pack(fill="x")

        a_is_fav = prob_a >= prob_b
        a_color = GREEN if a_is_fav else FG
        b_color = GREEN if not a_is_fav else FG

        tk.Label(
            names_frame,
            text=f"{name_a}  ({rating_a:.0f}, {rec_a})",
            font=("Helvetica", 11, "bold"),
            fg=a_color,
            bg=BG_HEADER,
            anchor="w",
        ).pack(side="left")

        tk.Label(
            names_frame,
            text="vs",
            font=("Helvetica", 10),
            fg="#666",
            bg=BG_HEADER,
        ).pack(side="left", padx=10)

        tk.Label(
            names_frame,
            text=f"{name_b}  ({rating_b:.0f}, {rec_b})",
            font=("Helvetica", 11, "bold"),
            fg=b_color,
            bg=BG_HEADER,
            anchor="e",
        ).pack(side="left")

        # Probability bar
        bar_frame = tk.Frame(card, bg=BG_HEADER, pady=4)
        bar_frame.pack(fill="x")

        BAR_WIDTH = 500
        BAR_HEIGHT = 24

        bar_canvas = tk.Canvas(
            bar_frame,
            width=BAR_WIDTH,
            height=BAR_HEIGHT,
            bg=BG_HEADER,
            highlightthickness=0,
        )
        bar_canvas.pack(anchor="w")

        a_width = max(1, int(prob_a * BAR_WIDTH))
        b_width = BAR_WIDTH - a_width

        bar_canvas.create_rectangle(0, 0, a_width, BAR_HEIGHT, fill=BAR_A, outline="")
        bar_canvas.create_rectangle(a_width, 0, BAR_WIDTH, BAR_HEIGHT, fill=BAR_B, outline="")

        if prob_a >= 0.12:
            bar_canvas.create_text(
                a_width // 2, BAR_HEIGHT // 2,
                text=f"{prob_a:.1%}",
                fill="white",
                font=("Helvetica", 9, "bold"),
            )
        if prob_b >= 0.12:
            bar_canvas.create_text(
                a_width + b_width // 2, BAR_HEIGHT // 2,
                text=f"{prob_b:.1%}",
                fill="white",
                font=("Helvetica", 9, "bold"),
            )

        # Pick summary
        if a_is_fav:
            pick = f"{name_a} ({prob_a:.1%})"
        else:
            pick = f"{name_b} ({prob_b:.1%})"

        tk.Label(
            card,
            text=f"Pick: {pick}",
            font=("Helvetica", 10),
            fg=GREEN,
            bg=BG_HEADER,
            anchor="w",
        ).pack(anchor="w")


def main():
    root = tk.Tk()
    Glicko2GUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
