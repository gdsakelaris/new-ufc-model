"""
UFC ML Fight Predictor.

Builds chronological pre-fight features from pure_fight_data.csv, trains a
time-aware ensemble with leak-safe train/validation/test evaluation,
probability calibration, then predicts upcoming fights via
tkinter and exports results to Excel.
"""

import csv, math, os, threading, warnings, logging, random, re
import tkinter as tk
from tkinter import messagebox
from collections import defaultdict
from datetime import datetime

import numpy as np
import pandas as pd
import lightgbm as lgb
import xgboost as xgb
import catboost as cb
import optuna
from sklearn.model_selection import TimeSeriesSplit
from sklearn.metrics import log_loss, accuracy_score, brier_score_loss
from sklearn.linear_model import LogisticRegression
from sklearn.preprocessing import StandardScaler
from sklearn.impute import SimpleImputer
from sklearn.ensemble import (
    HistGradientBoostingClassifier, RandomForestClassifier,
    ExtraTreesClassifier, AdaBoostClassifier,
)
from sklearn.neural_network import MLPClassifier
from sklearn.svm import SVC
from pytorch_tabnet.tab_model import TabNetClassifier
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
optuna.logging.set_verbosity(optuna.logging.WARNING)
logging.getLogger("lightgbm").setLevel(logging.WARNING)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
RANDOM_SEED = 42
MISSINGNESS_THRESHOLD = 0.35
EARLY_STOPPING_ROUNDS = 50

os.environ["PYTHONHASHSEED"] = str(RANDOM_SEED)
random.seed(RANDOM_SEED)
np.random.seed(RANDOM_SEED)

# ─── Glicko-2 constants ────────────────────────────────────────────────────────
MU_0 = 1500.0
PHI_0 = 200.0
SIGMA_0 = 0.06
TAU = 0.5
SCALE = 173.7178
CONVERGENCE = 1e-6

# ─── Weight class ordinal mapping ─────────────────────────────────────────────
# Ordinal is only a coarse size proxy now; the exact class is carried by one-hot
# features below, so every class gets a unique code to avoid collisions.
WEIGHT_CLASS_ORDINAL = {
    "Women's Strawweight": 1, "Women's Flyweight": 2, "Women's Bantamweight": 3,
    "Women's Featherweight": 4, "Flyweight": 5, "Bantamweight": 6,
    "Featherweight": 7, "Lightweight": 8, "Welterweight": 9,
    "Middleweight": 10, "Light Heavyweight": 11, "Heavyweight": 12,
    "Catch Weight": 13, "Open Weight": 14,
}

ACTIVE_ENSEMBLE_MODELS = {"ExtraTrees", "MLP", "TabNet", "LogReg"}


def _weight_class_feature_name(weight_class):
    slug = re.sub(r"[^a-z0-9]+", "_", str(weight_class).lower()).strip("_")
    return f"wc_{slug or 'unknown'}"


WEIGHT_CLASS_FEATURES = {
    weight_class: _weight_class_feature_name(weight_class)
    for weight_class in WEIGHT_CLASS_ORDINAL
}
UNKNOWN_WEIGHT_CLASS_FEATURE = "wc_unknown"

# ─── Glicko-2 core ─────────────────────────────────────────────────────────────

def _g(phi):
    return 1.0 / math.sqrt(1.0 + 3.0 * phi**2 / math.pi**2)

def _E(mu, mu_j, phi_j):
    return 1.0 / (1.0 + math.exp(-_g(phi_j) * (mu - mu_j)))

def glicko2_update(rating, opponents):
    mu, phi, sigma = rating
    mu_s = (mu - MU_0) / SCALE
    phi_s = phi / SCALE
    if not opponents:
        phi_star = math.sqrt(phi_s**2 + sigma**2)
        return (mu, phi_star * SCALE, sigma)
    v_inv = 0.0
    delta_sum = 0.0
    for opp_r, opp_rd, score in opponents:
        mu_j = (opp_r - MU_0) / SCALE
        phi_j = opp_rd / SCALE
        g_j = _g(phi_j)
        E_j = _E(mu_s, mu_j, phi_j)
        v_inv += g_j**2 * E_j * (1.0 - E_j)
        delta_sum += g_j * (score - E_j)
    v = 1.0 / v_inv if v_inv > 0 else 1e6
    delta = v * delta_sum
    a = math.log(sigma**2)
    def f(x):
        ex = math.exp(x)
        num = ex * (delta**2 - phi_s**2 - v - ex)
        den = 2.0 * (phi_s**2 + v + ex)**2
        return num / den - (x - a) / (TAU**2)
    A = a
    if delta**2 > phi_s**2 + v:
        B = math.log(delta**2 - phi_s**2 - v)
    else:
        k = 1
        while f(a - k * TAU) < 0:
            k += 1
        B = a - k * TAU
    fA, fB = f(A), f(B)
    for _ in range(100):
        C = A + (A - B) * fA / (fB - fA)
        fC = f(C)
        if fC * fB < 0:
            A, fA = B, fB
        else:
            fA /= 2.0
        B, fB = C, fC
        if abs(B - A) < CONVERGENCE:
            break
    new_sigma = math.exp(A / 2.0)
    phi_star = math.sqrt(phi_s**2 + new_sigma**2)
    new_phi_s = 1.0 / math.sqrt(1.0 / phi_star**2 + 1.0 / v)
    new_mu_s = mu_s + new_phi_s**2 * delta_sum
    return (new_mu_s * SCALE + MU_0, new_phi_s * SCALE, new_sigma)

# ─── Helpers ────────────────────────────────────────────────────────────────────

def _isnan(v):
    if v is None:
        return True
    try:
        return math.isnan(float(v))
    except (TypeError, ValueError):
        return True

def _safe_sum(values):
    return sum(v for v in values if not _isnan(v))

def _safe_mean(values):
    clean = [v for v in values if not _isnan(v)]
    return sum(clean) / len(clean) if clean else float("nan")

def _safe_div(a, b, default=0.0):
    return a / b if b and b > 0 else default


def _num_or(value, default=0.0):
    return default if _isnan(value) else float(value)


def _abs_gap(a, b, default=float("nan")):
    if _isnan(a) or _isnan(b):
        return default
    return abs(float(a) - float(b))


def _extract_profile_from_row(row, prefix):
    return {
        "height": row.get(f"{prefix}_height", float("nan")),
        "reach": row.get(f"{prefix}_reach", float("nan")),
        "ape_index": row.get(f"{prefix}_ape_index", float("nan")),
        "weight": row.get(f"{prefix}_weight", float("nan")),
        "age": row.get(f"{prefix}_age_at_event", float("nan")),
        "stance": row.get(f"{prefix}_stance", ""),
    }


def _weighted_rate(numerator, denominator, prior=0.5, strength=25.0):
    """Bayesian-smoothed rate using equivalent sample-size `strength`."""
    if denominator is None or denominator <= 0:
        return prior
    return (numerator + prior * strength) / (denominator + strength)


def _bayes_shrink(empirical_value, sample_size, prior=0.5, strength=8.0):
    """Shrink noisy low-sample statistics toward a global prior."""
    if _isnan(empirical_value):
        empirical_value = prior
    sample_size = max(float(sample_size or 0.0), 0.0)
    return (empirical_value * sample_size + prior * strength) / (sample_size + strength)


def _time_weight(fight_date, current_date, half_life_days=730):
    """Exponential decay weight: half-life of ~2 years."""
    if current_date is None or fight_date is None:
        return 1.0
    try:
        days = (current_date - fight_date).days
        return 2.0 ** (-max(days, 0) / half_life_days)
    except (TypeError, AttributeError):
        return 1.0


def _pick_missing_cols(X, threshold=MISSINGNESS_THRESHOLD):
    return [c for c in X.columns if float(X[c].isna().mean()) >= threshold]


def _add_missingness_indicators(X, missing_cols):
    X2 = X.copy()
    for col in missing_cols:
        if col not in X2.columns:
            X2[col] = np.nan
        X2[f"miss_{col}"] = X2[col].isna().astype(float)
    return X2


def _expected_calibration_error(y_true, probs, n_bins=10):
    y = np.asarray(y_true, dtype=float)
    p = np.asarray(probs, dtype=float)
    bins = np.linspace(0.0, 1.0, n_bins + 1)
    inds = np.digitize(p, bins) - 1
    ece = 0.0
    for b in range(n_bins):
        mask = inds == b
        if not np.any(mask):
            continue
        conf = p[mask].mean()
        acc = y[mask].mean()
        ece += (mask.mean()) * abs(acc - conf)
    return float(ece)


# Per-round stat names to track
RD_STATS = [
    "sig_str", "sig_str_att", "kd", "td", "td_att", "sub_att", "ctrl_sec",
    "head", "body", "leg", "distance", "clinch", "ground",
]

# ─── Fight record extraction ───────────────────────────────────────────────────

def extract_fight_record(row, prefix, opp, result, opp_glicko_mu=MU_0):
    """Build a per-fighter fight record dict from a DataFrame row."""
    def g(col):
        v = row.get(col)
        if v is None or (isinstance(v, float) and math.isnan(v)):
            return float("nan")
        return v

    rec = {
        "date": row["event_date"],
        "fight_time": g(f"total_fight_time_sec") or 0,
        "result": result,
        "opp_glicko": opp_glicko_mu,
        "method": row.get("method", ""),
        "is_title": g("is_title_bout") or 0,
        "finish_round": g("finish_round") or 0,
        # Offense
        "sig_str": g(f"{prefix}_sig_str") or 0,
        "sig_str_att": g(f"{prefix}_sig_str_att") or 0,
        "sig_str_acc": g(f"{prefix}_sig_str_acc"),
        "str": g(f"{prefix}_str") or 0,
        "str_att": g(f"{prefix}_str_att") or 0,
        "str_acc": g(f"{prefix}_str_acc"),
        "kd": g(f"{prefix}_kd") or 0,
        "td": g(f"{prefix}_td") or 0,
        "td_att": g(f"{prefix}_td_att") or 0,
        "td_acc": g(f"{prefix}_td_acc"),
        "sub_att": g(f"{prefix}_sub_att") or 0,
        "rev": g(f"{prefix}_rev") or 0,
        "ctrl_sec": g(f"{prefix}_ctrl_sec") or 0,
        # Targeting (fight-total %)
        "head_pct": g(f"{prefix}_head"),
        "body_pct": g(f"{prefix}_body"),
        "leg_pct": g(f"{prefix}_leg"),
        # Positioning (fight-total %)
        "distance_pct": g(f"{prefix}_distance"),
        "clinch_pct": g(f"{prefix}_clinch"),
        "ground_pct": g(f"{prefix}_ground"),
        # Defense (opponent stats)
        "opp_sig_str": g(f"{opp}_sig_str") or 0,
        "opp_sig_str_att": g(f"{opp}_sig_str_att") or 0,
        "opp_sig_str_acc": g(f"{opp}_sig_str_acc"),
        "opp_str": g(f"{opp}_str") or 0,
        "opp_kd": g(f"{opp}_kd") or 0,
        "opp_td": g(f"{opp}_td") or 0,
        "opp_td_att": g(f"{opp}_td_att") or 0,
        "opp_sub_att": g(f"{opp}_sub_att") or 0,
        "opp_ctrl_sec": g(f"{opp}_ctrl_sec") or 0,
        # Physical
        "height": g(f"{prefix}_height"),
        "reach": g(f"{prefix}_reach"),
        "ape_index": g(f"{prefix}_ape_index"),
        "weight": g(f"{prefix}_weight"),
        "age": g(f"{prefix}_age_at_event"),
        "stance": row.get(f"{prefix}_stance", ""),
    }

    # Per-round stats
    for rd in range(1, 6):
        for stat in RD_STATS:
            rec[f"rd{rd}_{stat}"] = g(f"{prefix}_rd{rd}_{stat}")

    return rec


# ─── Per-fighter feature computation ───────────────────────────────────────────

_FIGHTER_FEAT_KEYS = None  # populated on first call


def _make_synthetic_fight_record(current_date=None, profile=None):
    profile = profile or {}
    rec = {
        "date": current_date or pd.Timestamp("2000-01-01"),
        "fight_time": 900.0,
        "result": "W",
        "opp_glicko": MU_0,
        "method": "Decision",
        "is_title": 0,
        "finish_round": 3,
        "sig_str": 0.0,
        "sig_str_att": 0.0,
        "sig_str_acc": float("nan"),
        "str": 0.0,
        "str_att": 0.0,
        "str_acc": float("nan"),
        "kd": 0.0,
        "td": 0.0,
        "td_att": 0.0,
        "td_acc": float("nan"),
        "sub_att": 0.0,
        "rev": 0.0,
        "ctrl_sec": 0.0,
        "head_pct": float("nan"),
        "body_pct": float("nan"),
        "leg_pct": float("nan"),
        "distance_pct": float("nan"),
        "clinch_pct": float("nan"),
        "ground_pct": float("nan"),
        "opp_sig_str": 0.0,
        "opp_sig_str_att": 0.0,
        "opp_sig_str_acc": float("nan"),
        "opp_str": 0.0,
        "opp_kd": 0.0,
        "opp_td": 0.0,
        "opp_td_att": 0.0,
        "opp_sub_att": 0.0,
        "opp_ctrl_sec": 0.0,
        "height": profile.get("height", 70.0),
        "reach": profile.get("reach", 72.0),
        "ape_index": profile.get("ape_index", 2.0),
        "weight": profile.get("weight", 170.0),
        "age": profile.get("age", 30.0),
        "stance": profile.get("stance", "Orthodox"),
    }
    for rd in range(1, 6):
        for stat in RD_STATS:
            rec[f"rd{rd}_{stat}"] = 0.0
    return rec


def _ensure_fighter_feature_keys(current_date=None):
    global _FIGHTER_FEAT_KEYS
    if _FIGHTER_FEAT_KEYS is not None:
        return
    dummy_history = [_make_synthetic_fight_record(current_date=current_date)]
    compute_fighter_features(dummy_history, (MU_0, PHI_0, SIGMA_0), [], current_date)


def compute_fighter_features(history, glicko, opp_glickos, current_date, fallback_profile=None):
    """Compute ~130 features from a fighter's fight history."""
    global _FIGHTER_FEAT_KEYS
    fallback_profile = fallback_profile or {}

    n = len(history)
    if n == 0:
        _ensure_fighter_feature_keys(current_date)
        feats = {k: float("nan") for k in _FIGHTER_FEAT_KEYS}
        stance = fallback_profile.get("stance", "") or ""
        feats.update({
            "height": _num_or(fallback_profile.get("height"), float("nan")),
            "reach": _num_or(fallback_profile.get("reach"), float("nan")),
            "ape_index": _num_or(fallback_profile.get("ape_index"), float("nan")),
            "weight": _num_or(fallback_profile.get("weight"), float("nan")),
            "age": _num_or(fallback_profile.get("age"), float("nan")),
            "num_fights": 0.0,
            "total_time_min": 0.0,
            "avg_time_min": 0.0,
            "title_bout_pct": 0.0,
            "win_rate": 0.5,
            "ko_win_pct": 0.25,
            "sub_win_pct": 0.12,
            "dec_win_pct": 0.13,
            "finish_rate": 0.5,
            "ko_loss_pct": 0.2,
            "been_finished_pct": 0.5,
            "last3_win_rate": 0.5,
            "last5_win_rate": 0.5,
            "win_streak": 0.0,
            "loss_streak": 0.0,
            "days_inactive": 365.0,
            "glicko_mu": glicko[0],
            "glicko_phi": glicko[1],
            "fights_per_year": 0.0,
            "avg_opp_glicko": MU_0,
            "td_win_rate": 0.5,
            "quality_win_rate": 0.5,
            "best_win_glicko": MU_0,
            "worst_loss_glicko": MU_0,
            "style_striking": 0.5,
            "style_wrestling": 0.25,
            "style_submission": 0.15,
            "style_clinch_ground": 0.32,
            "style_finishing": 0.5,
            "is_orthodox": 1.0 if stance == "Orthodox" else 0.0,
            "is_southpaw": 1.0 if stance == "Southpaw" else 0.0,
            "is_switch": 1.0 if stance == "Switch" else 0.0,
        })
        return feats

    feats = {}

    total_time = _safe_sum(h["fight_time"] for h in history)
    total_time_min = total_time / 60.0 if total_time > 0 else 1.0
    total_time_15 = total_time / 900.0 if total_time > 0 else 1.0

    total_sig_land = _safe_sum(h["sig_str"] for h in history)
    total_sig_att = _safe_sum(h["sig_str_att"] for h in history)
    total_str_land = _safe_sum(h["str"] for h in history)
    total_str_att = _safe_sum(h["str_att"] for h in history)
    total_td_land = _safe_sum(h["td"] for h in history)
    total_td_att = _safe_sum(h["td_att"] for h in history)

    # ── Striking offense (per minute) ──
    feats["sig_str_pm"] = total_sig_land / total_time_min
    feats["sig_str_att_pm"] = total_sig_att / total_time_min
    feats["str_pm"] = total_str_land / total_time_min
    feats["str_att_pm"] = total_str_att / total_time_min
    feats["kd_pm"] = _safe_sum(h["kd"] for h in history) / total_time_min

    # ── Accuracy (attempt-weighted + Bayesian shrinkage) ──
    feats["sig_str_acc"] = _bayes_shrink(
        _weighted_rate(total_sig_land, total_sig_att, prior=0.45, strength=30),
        n, prior=0.45, strength=8,
    )
    feats["str_acc"] = _bayes_shrink(
        _weighted_rate(total_str_land, total_str_att, prior=0.55, strength=30),
        n, prior=0.55, strength=8,
    )
    feats["td_acc"] = _bayes_shrink(
        _weighted_rate(total_td_land, total_td_att, prior=0.35, strength=20),
        n, prior=0.35, strength=8,
    )

    # ── Grappling (per 15 min) ──
    feats["td_p15"] = _safe_sum(h["td"] for h in history) / total_time_15
    feats["td_att_p15"] = _safe_sum(h["td_att"] for h in history) / total_time_15
    feats["sub_att_p15"] = _safe_sum(h["sub_att"] for h in history) / total_time_15
    feats["rev_p15"] = _safe_sum(h["rev"] for h in history) / total_time_15
    feats["ctrl_pct"] = _safe_sum(h["ctrl_sec"] for h in history) / total_time if total_time > 0 else 0

    # ── Targeting (attempt/volume-weighted + shrinkage) ──
    sig_for_target = _safe_sum(
        h["sig_str"] for h in history
        if not _isnan(h["head_pct"]) and not _isnan(h["sig_str"])
    )
    head_num = _safe_sum(
        h["head_pct"] * h["sig_str"]
        for h in history
        if not _isnan(h["head_pct"]) and not _isnan(h["sig_str"])
    )
    body_num = _safe_sum(
        h["body_pct"] * h["sig_str"]
        for h in history
        if not _isnan(h["body_pct"]) and not _isnan(h["sig_str"])
    )
    leg_num = _safe_sum(
        h["leg_pct"] * h["sig_str"]
        for h in history
        if not _isnan(h["leg_pct"]) and not _isnan(h["sig_str"])
    )
    feats["head_pct"] = _bayes_shrink(
        _weighted_rate(head_num, sig_for_target, prior=0.62, strength=40),
        n, prior=0.62, strength=8,
    )
    feats["body_pct"] = _bayes_shrink(
        _weighted_rate(body_num, sig_for_target, prior=0.22, strength=40),
        n, prior=0.22, strength=8,
    )
    feats["leg_pct"] = _bayes_shrink(
        _weighted_rate(leg_num, sig_for_target, prior=0.16, strength=40),
        n, prior=0.16, strength=8,
    )

    # ── Positioning (volume-weighted + shrinkage) ──
    sig_for_pos = _safe_sum(
        h["sig_str"] for h in history
        if not _isnan(h["distance_pct"]) and not _isnan(h["sig_str"])
    )
    dist_num = _safe_sum(
        h["distance_pct"] * h["sig_str"]
        for h in history
        if not _isnan(h["distance_pct"]) and not _isnan(h["sig_str"])
    )
    clinch_num = _safe_sum(
        h["clinch_pct"] * h["sig_str"]
        for h in history
        if not _isnan(h["clinch_pct"]) and not _isnan(h["sig_str"])
    )
    ground_num = _safe_sum(
        h["ground_pct"] * h["sig_str"]
        for h in history
        if not _isnan(h["ground_pct"]) and not _isnan(h["sig_str"])
    )
    feats["distance_pct"] = _bayes_shrink(
        _weighted_rate(dist_num, sig_for_pos, prior=0.68, strength=40),
        n, prior=0.68, strength=8,
    )
    feats["clinch_pct"] = _bayes_shrink(
        _weighted_rate(clinch_num, sig_for_pos, prior=0.14, strength=40),
        n, prior=0.14, strength=8,
    )
    feats["ground_pct"] = _bayes_shrink(
        _weighted_rate(ground_num, sig_for_pos, prior=0.18, strength=40),
        n, prior=0.18, strength=8,
    )

    # ── Defense (opponent per-minute) ──
    feats["def_sig_str_pm"] = _safe_sum(h["opp_sig_str"] for h in history) / total_time_min
    feats["def_str_pm"] = _safe_sum(h["opp_str"] for h in history) / total_time_min
    feats["def_kd_pm"] = _safe_sum(h["opp_kd"] for h in history) / total_time_min
    feats["def_td_p15"] = _safe_sum(h["opp_td"] for h in history) / total_time_15
    feats["def_sub_att_p15"] = _safe_sum(h["opp_sub_att"] for h in history) / total_time_15
    feats["def_ctrl_pct"] = _safe_sum(h["opp_ctrl_sec"] for h in history) / total_time if total_time > 0 else 0
    opp_sig_att = _safe_sum(h["opp_sig_str_att"] for h in history)
    opp_sig_land = _safe_sum(h["opp_sig_str"] for h in history)
    feats["def_sig_str_acc"] = _bayes_shrink(
        _weighted_rate(opp_sig_land, opp_sig_att, prior=0.45, strength=30),
        n, prior=0.45, strength=8,
    )

    # ── Differentials ──
    feats["net_sig_str_pm"] = feats["sig_str_pm"] - feats["def_sig_str_pm"]
    feats["net_kd_pm"] = feats["kd_pm"] - feats["def_kd_pm"]
    feats["net_td_p15"] = feats["td_p15"] - feats["def_td_p15"]
    feats["net_ctrl_pct"] = feats["ctrl_pct"] - feats["def_ctrl_pct"]

    # ── Defense rates ──
    feats["sig_str_defense_rate"] = _bayes_shrink(
        1.0 - _weighted_rate(opp_sig_land, opp_sig_att, prior=0.45, strength=30),
        n, prior=0.55, strength=8,
    )
    opp_td_att = _safe_sum(h["opp_td_att"] for h in history)
    opp_td_land = _safe_sum(h["opp_td"] for h in history)
    feats["td_defense_rate"] = _bayes_shrink(
        1.0 - _weighted_rate(opp_td_land, opp_td_att, prior=0.35, strength=20),
        n, prior=0.65, strength=8,
    )

    # ── Physical (most recent) ──
    latest = history[-1]
    feats["height"] = latest["height"] if not _isnan(latest["height"]) else _num_or(fallback_profile.get("height"), float("nan"))
    feats["reach"] = latest["reach"] if not _isnan(latest["reach"]) else _num_or(fallback_profile.get("reach"), float("nan"))
    feats["ape_index"] = latest["ape_index"] if not _isnan(latest["ape_index"]) else _num_or(fallback_profile.get("ape_index"), float("nan"))
    feats["weight"] = latest["weight"] if not _isnan(latest["weight"]) else _num_or(fallback_profile.get("weight"), float("nan"))
    feats["age"] = latest["age"] if not _isnan(latest["age"]) else _num_or(fallback_profile.get("age"), float("nan"))

    # ── Experience ──
    feats["num_fights"] = n
    feats["total_time_min"] = total_time_min
    feats["avg_time_min"] = total_time_min / n
    feats["title_bout_pct"] = sum(1 for h in history if h["is_title"]) / n

    # ── Record ──
    wins = sum(1 for h in history if h["result"] == "W")
    losses = sum(1 for h in history if h["result"] == "L")
    feats["win_rate"] = _bayes_shrink(_safe_div(wins, n, 0.5), n, prior=0.5, strength=10)
    ko_w = sum(1 for h in history if h["result"] == "W" and "KO" in str(h["method"]))
    sub_w = sum(1 for h in history if h["result"] == "W" and "Sub" in str(h["method"]))
    dec_w = sum(1 for h in history if h["result"] == "W" and "Dec" in str(h["method"]))
    feats["ko_win_pct"] = _bayes_shrink(_safe_div(ko_w, n, 0.25), n, prior=0.25, strength=10)
    feats["sub_win_pct"] = _bayes_shrink(_safe_div(sub_w, n, 0.12), n, prior=0.12, strength=10)
    feats["dec_win_pct"] = _bayes_shrink(_safe_div(dec_w, n, 0.13), n, prior=0.13, strength=10)
    feats["finish_rate"] = _safe_div(ko_w + sub_w, max(wins, 1))
    ko_l = sum(1 for h in history if h["result"] == "L" and "KO" in str(h["method"]))
    sub_l = sum(1 for h in history if h["result"] == "L" and "Sub" in str(h["method"]))
    feats["ko_loss_pct"] = _bayes_shrink(_safe_div(ko_l, n, 0.2), n, prior=0.2, strength=10)
    feats["been_finished_pct"] = _safe_div(ko_l + sub_l, max(losses, 1))

    # ── Form ──
    last3 = history[-3:]
    last5 = history[-5:]
    last3_wr = sum(1 for h in last3 if h["result"] == "W") / len(last3)
    last5_wr = sum(1 for h in last5 if h["result"] == "W") / len(last5)
    feats["last3_win_rate"] = _bayes_shrink(last3_wr, len(last3), prior=0.5, strength=6)
    feats["last5_win_rate"] = _bayes_shrink(last5_wr, len(last5), prior=0.5, strength=6)
    win_streak = 0
    for h in reversed(history):
        if h["result"] == "W":
            win_streak += 1
        else:
            break
    loss_streak = 0
    for h in reversed(history):
        if h["result"] == "L":
            loss_streak += 1
        else:
            break
    feats["win_streak"] = win_streak
    feats["loss_streak"] = loss_streak
    if current_date is not None and not _isnan(history[-1]["date"]):
        feats["days_inactive"] = (current_date - history[-1]["date"]).days
    else:
        feats["days_inactive"] = 365

    # ── Glicko-2 ──
    feats["glicko_mu"] = glicko[0]
    feats["glicko_phi"] = glicko[1]

    # ── Round 1 stats (career averages) ──
    for stat in ["sig_str", "sig_str_att", "kd", "td", "td_att", "sub_att",
                 "ctrl_sec", "head", "body", "leg", "distance", "clinch", "ground"]:
        vals = [h[f"rd1_{stat}"] for h in history if not _isnan(h.get(f"rd1_{stat}"))]
        feats[f"rd1_{stat}"] = _safe_mean(vals)

    # ── Round 2 stats ──
    for stat in ["sig_str", "sig_str_att", "kd", "td", "td_att", "sub_att",
                 "ctrl_sec", "head", "body", "leg"]:
        vals = [h[f"rd2_{stat}"] for h in history if not _isnan(h.get(f"rd2_{stat}"))]
        feats[f"rd2_{stat}"] = _safe_mean(vals)

    # ── Round 3 stats ──
    for stat in ["sig_str", "kd", "td", "ctrl_sec", "head", "body", "leg"]:
        vals = [h[f"rd3_{stat}"] for h in history if not _isnan(h.get(f"rd3_{stat}"))]
        feats[f"rd3_{stat}"] = _safe_mean(vals)

    # ── Championship rounds (4+5 combined) ──
    for stat in ["sig_str", "kd", "td", "ctrl_sec"]:
        vals = []
        for rd in [4, 5]:
            vals.extend(h[f"rd{rd}_{stat}"] for h in history
                        if not _isnan(h.get(f"rd{rd}_{stat}")))
        feats[f"champ_{stat}"] = _safe_mean(vals)

    # ── Late vs early ──
    rd1_ss = [h["rd1_sig_str"] for h in history if not _isnan(h.get("rd1_sig_str"))]
    rd3_ss = [h["rd3_sig_str"] for h in history if not _isnan(h.get("rd3_sig_str"))]
    feats["late_vs_early_sig_str"] = _safe_mean(rd3_ss) - _safe_mean(rd1_ss) \
        if rd1_ss and rd3_ss else float("nan")

    # ── Recent-window rate stats (last 1, 3, 5 fights) ──
    for prefix_tag, window in [("r1f", 1), ("r3", 3), ("r5", 5)]:
        recent = history[-window:]
        rt = _safe_sum(h["fight_time"] for h in recent)
        rt_min = rt / 60.0 if rt > 0 else 1.0
        rt_15 = rt / 900.0 if rt > 0 else 1.0
        feats[f"{prefix_tag}_sig_str_pm"] = _safe_sum(h["sig_str"] for h in recent) / rt_min
        feats[f"{prefix_tag}_kd_pm"] = _safe_sum(h["kd"] for h in recent) / rt_min
        feats[f"{prefix_tag}_td_p15"] = _safe_sum(h["td"] for h in recent) / rt_15
        feats[f"{prefix_tag}_ctrl_pct"] = _safe_sum(h["ctrl_sec"] for h in recent) / rt if rt > 0 else 0
        feats[f"{prefix_tag}_def_sig_str_pm"] = _safe_sum(h["opp_sig_str"] for h in recent) / rt_min
        rec_sig_land = _safe_sum(h["sig_str"] for h in recent)
        rec_sig_att = _safe_sum(h["sig_str_att"] for h in recent)
        rec_td_land = _safe_sum(h["td"] for h in recent)
        rec_td_att = _safe_sum(h["td_att"] for h in recent)
        feats[f"{prefix_tag}_sig_str_acc"] = _bayes_shrink(
            _weighted_rate(rec_sig_land, rec_sig_att, prior=0.45, strength=20),
            len(recent), prior=0.45, strength=6,
        )
        feats[f"{prefix_tag}_td_acc"] = _bayes_shrink(
            _weighted_rate(rec_td_land, rec_td_att, prior=0.35, strength=20),
            len(recent), prior=0.35, strength=6,
        )
        feats[f"{prefix_tag}_win"] = _bayes_shrink(
            _safe_div(sum(1 for h in recent if h["result"] == "W"), len(recent), 0.5),
            len(recent), prior=0.5, strength=6,
        )

    # ── Exponentially weighted moving average (alpha=0.3, more recent = higher) ──
    alpha = 0.3
    ewm_keys = [
        ("sig_str", "fight_time", 60),
        ("kd", "fight_time", 60),
        ("td", "fight_time", 900),
        ("opp_sig_str", "fight_time", 60),
    ]
    for stat_key, time_key, divisor in ewm_keys:
        wsum, wtot = 0.0, 0.0
        for i, h in enumerate(history):
            w = (1 - alpha) ** (n - 1 - i)
            ft = h[time_key] / divisor if h[time_key] > 0 else 1.0
            wsum += w * (h[stat_key] / ft)
            wtot += w
        tag = stat_key.replace("opp_sig_str", "def_sig_str")
        feats[f"ewm_{tag}_pm"] = wsum / wtot if wtot > 0 else 0

    # EWM ctrl_pct
    wsum, wtot = 0.0, 0.0
    for i, h in enumerate(history):
        w = (1 - alpha) ** (n - 1 - i)
        ft = h["fight_time"] if h["fight_time"] > 0 else 1.0
        wsum += w * (h["ctrl_sec"] / ft)
        wtot += w
    feats["ewm_ctrl_pct"] = wsum / wtot if wtot > 0 else 0

    # EWM accuracy
    for acc_key in ["sig_str_acc", "td_acc"]:
        wsum, wtot = 0.0, 0.0
        for i, h in enumerate(history):
            if _isnan(h[acc_key]):
                continue
            w = (1 - alpha) ** (n - 1 - i)
            wsum += w * h[acc_key]
            wtot += w
        feats[f"ewm_{acc_key}"] = wsum / wtot if wtot > 0 else float("nan")

    # EWM win
    wsum, wtot = 0.0, 0.0
    for i, h in enumerate(history):
        w = (1 - alpha) ** (n - 1 - i)
        wsum += w * (1.0 if h["result"] == "W" else 0.0)
        wtot += w
    feats["ewm_win"] = wsum / wtot if wtot > 0 else 0.5

    # ── Variance (per-fight rates std dev) ──
    if n >= 2:
        per_fight_sig = []
        per_fight_kd = []
        per_fight_td = []
        per_fight_ctrl = []
        per_fight_def_sig = []
        for h in history:
            ft_min = h["fight_time"] / 60.0 if h["fight_time"] > 0 else 1.0
            ft_15 = h["fight_time"] / 900.0 if h["fight_time"] > 0 else 1.0
            ft = h["fight_time"] if h["fight_time"] > 0 else 1.0
            per_fight_sig.append(h["sig_str"] / ft_min)
            per_fight_kd.append(h["kd"] / ft_min)
            per_fight_td.append(h["td"] / ft_15)
            per_fight_ctrl.append(h["ctrl_sec"] / ft)
            per_fight_def_sig.append(h["opp_sig_str"] / ft_min)
        feats["std_sig_str_pm"] = float(np.std(per_fight_sig))
        feats["std_kd_pm"] = float(np.std(per_fight_kd))
        feats["std_td_p15"] = float(np.std(per_fight_td))
        feats["std_ctrl_pct"] = float(np.std(per_fight_ctrl))
        feats["std_def_sig_str_pm"] = float(np.std(per_fight_def_sig))
    else:
        feats["std_sig_str_pm"] = 0.0
        feats["std_kd_pm"] = 0.0
        feats["std_td_p15"] = 0.0
        feats["std_ctrl_pct"] = 0.0
        feats["std_def_sig_str_pm"] = 0.0

    # ── Derived ──
    total_sig = _safe_sum(h["sig_str"] for h in history)
    total_str = _safe_sum(h["str"] for h in history)
    feats["power_ratio"] = _safe_div(_safe_sum(h["kd"] for h in history), max(total_sig, 1))
    feats["striking_efficiency"] = _safe_div(total_sig, max(total_str, 1))
    feats["grappling_rate_p15"] = (
        _safe_sum(h["td"] + h["sub_att"] for h in history) / total_time_15
    )
    feats["def_grappling_rate_p15"] = (
        _safe_sum(h["opp_td"] + h["opp_sub_att"] for h in history) / total_time_15
    )
    feats["damage_ratio"] = _safe_div(
        _safe_sum(h["opp_sig_str"] for h in history),
        max(_safe_sum(h["sig_str"] for h in history), 1),
    )

    # Fights per year
    if n > 1:
        career_days = (history[-1]["date"] - history[0]["date"]).days
        feats["fights_per_year"] = n / max(career_days / 365.25, 0.5)
    else:
        feats["fights_per_year"] = 1.0

    # Stance
    stance = latest.get("stance", "") or fallback_profile.get("stance", "")
    feats["is_orthodox"] = 1 if stance == "Orthodox" else 0
    feats["is_southpaw"] = 1 if stance == "Southpaw" else 0
    feats["is_switch"] = 1 if stance == "Switch" else 0

    # Average opponent Glicko
    feats["avg_opp_glicko"] = _safe_mean(opp_glickos) if opp_glickos else MU_0

    # ── Time-decayed career stats (half-life ~2 years) ──
    tw = [_time_weight(h["date"], current_date) for h in history]
    tw_total = sum(tw) or 1.0
    tw_time = sum(w * h["fight_time"] for w, h in zip(tw, history))
    tw_time_min = tw_time / 60.0 if tw_time > 0 else 1.0
    tw_time_15 = tw_time / 900.0 if tw_time > 0 else 1.0
    feats["td_sig_str_pm"] = sum(w * h["sig_str"] for w, h in zip(tw, history)) / tw_time_min
    feats["td_kd_pm"] = sum(w * h["kd"] for w, h in zip(tw, history)) / tw_time_min
    feats["td_td_p15"] = sum(w * h["td"] for w, h in zip(tw, history)) / tw_time_15
    feats["td_ctrl_pct"] = (
        sum(w * h["ctrl_sec"] for w, h in zip(tw, history)) / tw_time if tw_time > 0 else 0
    )
    feats["td_def_sig_str_pm"] = sum(w * h["opp_sig_str"] for w, h in zip(tw, history)) / tw_time_min
    tw_sig_land = sum(w * h["sig_str"] for w, h in zip(tw, history))
    tw_sig_att = sum(w * h["sig_str_att"] for w, h in zip(tw, history))
    feats["td_sig_str_acc"] = _weighted_rate(tw_sig_land, tw_sig_att, prior=0.45, strength=20)
    tw_td_land = sum(w * h["td"] for w, h in zip(tw, history))
    tw_td_att = sum(w * h["td_att"] for w, h in zip(tw, history))
    feats["td_td_acc"] = _weighted_rate(tw_td_land, tw_td_att, prior=0.35, strength=20)
    feats["td_win_rate"] = sum(
        w * (1.0 if h["result"] == "W" else 0.0) for w, h in zip(tw, history)
    ) / tw_total

    # ── Opponent-quality-adjusted features ──
    opp_g = [h.get("opp_glicko", MU_0) for h in history]
    opp_g_w = [max(g_val / MU_0, 0.1) for g_val in opp_g]
    ogw_total = sum(opp_g_w) or 1.0
    feats["quality_win_rate"] = sum(
        w * (1.0 if h["result"] == "W" else 0.0)
        for w, h in zip(opp_g_w, history)
    ) / ogw_total
    wins_opp_g = [og for h, og in zip(history, opp_g) if h["result"] == "W"]
    feats["best_win_glicko"] = max(wins_opp_g) if wins_opp_g else MU_0
    losses_opp_g = [og for h, og in zip(history, opp_g) if h["result"] == "L"]
    feats["worst_loss_glicko"] = min(losses_opp_g) if losses_opp_g else MU_0
    feats["quality_sig_str_pm"] = sum(
        w * h["sig_str"] for w, h in zip(opp_g_w, history)
    ) / (sum(w * (h["fight_time"] / 60.0 if h["fight_time"] > 0 else 1.0)
             for w, h in zip(opp_g_w, history)) or 1.0)

    # ── Style profile (continuous features for matchup interactions) ──
    total_offensive = feats["sig_str_pm"] + feats["td_p15"] + feats["sub_att_p15"] + 0.01
    feats["style_striking"] = feats["sig_str_pm"] / total_offensive
    feats["style_wrestling"] = feats["td_p15"] / total_offensive
    feats["style_submission"] = feats["sub_att_p15"] / total_offensive
    feats["style_clinch_ground"] = feats.get("clinch_pct", 0.14) + feats.get("ground_pct", 0.18)
    feats["style_finishing"] = feats.get("finish_rate", 0.5)

    if _FIGHTER_FEAT_KEYS is None:
        _set_fighter_keys(list(feats.keys()))

    return feats


def _set_fighter_keys(keys):
    global _FIGHTER_FEAT_KEYS
    _FIGHTER_FEAT_KEYS = keys


# ─── Matchup feature computation ──────────────────────────────────────────────

def compute_matchup_features(a_feats, b_feats, is_title=0, total_rounds=3, weight_class=""):
    """Difference features (A minus B) plus interaction features."""
    features = {}
    for key in a_feats:
        a_val = a_feats[key] if not _isnan(a_feats[key]) else float("nan")
        b_val = b_feats[key] if not _isnan(b_feats[key]) else float("nan")
        try:
            features[f"d_{key}"] = float(a_val) - float(b_val)
        except (TypeError, ValueError):
            features[f"d_{key}"] = float("nan")

    a_age = a_feats.get("age", float("nan"))
    b_age = b_feats.get("age", float("nan"))
    a_reach = a_feats.get("reach", float("nan"))
    b_reach = b_feats.get("reach", float("nan"))
    a_height = a_feats.get("height", float("nan"))
    b_height = b_feats.get("height", float("nan"))
    a_mu = _num_or(a_feats.get("glicko_mu"), MU_0)
    b_mu = _num_or(b_feats.get("glicko_mu"), MU_0)
    a_phi = _num_or(a_feats.get("glicko_phi"), PHI_0)
    b_phi = _num_or(b_feats.get("glicko_phi"), PHI_0)
    a_n = max(_num_or(a_feats.get("num_fights"), 0.0), 0.0)
    b_n = max(_num_or(b_feats.get("num_fights"), 0.0), 0.0)
    a_debut = a_n < 1.0
    b_debut = b_n < 1.0
    a_inexperienced = a_n < 3.0
    b_inexperienced = b_n < 3.0

    # Absolute / interaction features
    features["is_title"] = is_title
    features["total_rounds"] = total_rounds
    features["exp_sum"] = a_n + b_n
    features["age_sum"] = _num_or(a_age, 30.0) + _num_or(b_age, 30.0)
    features["glicko_mu_sum"] = a_mu + b_mu
    features["abs_age_gap"] = _abs_gap(a_age, b_age)
    features["abs_reach_gap"] = _abs_gap(a_reach, b_reach)
    features["abs_height_gap"] = _abs_gap(a_height, b_height)
    features["abs_glicko_gap"] = abs(a_mu - b_mu)
    features["abs_glicko_phi_gap"] = abs(a_phi - b_phi)
    features["abs_exp_gap"] = abs(a_n - b_n)
    features["min_num_fights"] = min(a_n, b_n)
    features["max_num_fights"] = max(a_n, b_n)
    features["experience_ratio"] = min(a_n, b_n) / max(max(a_n, b_n), 1.0)
    features["max_glicko_phi"] = max(a_phi, b_phi)
    features["min_glicko_phi"] = min(a_phi, b_phi)
    features["avg_glicko_phi"] = (a_phi + b_phi) / 2.0
    features["both_debut"] = float(a_debut and b_debut)
    features["one_debut"] = float(a_debut ^ b_debut)
    features["both_inexperienced"] = float(a_inexperienced and b_inexperienced)
    features["one_inexperienced"] = float(a_inexperienced ^ b_inexperienced)
    features["info_asymmetry"] = abs(a_n - b_n) / max(a_n + b_n, 1.0)

    # Weight class
    features["weight_class_ord"] = WEIGHT_CLASS_ORDINAL.get(weight_class, 0)
    for class_name, feature_name in WEIGHT_CLASS_FEATURES.items():
        features[feature_name] = float(weight_class == class_name)
    features[UNKNOWN_WEIGHT_CLASS_FEATURE] = float(weight_class not in WEIGHT_CLASS_ORDINAL)

    # Stance matchup interactions (d_ prefix for correct augmentation negation)
    a_ortho = _num_or(a_feats.get("is_orthodox"), 0.0)
    a_south = _num_or(a_feats.get("is_southpaw"), 0.0)
    b_ortho = _num_or(b_feats.get("is_orthodox"), 0.0)
    b_south = _num_or(b_feats.get("is_southpaw"), 0.0)
    features["d_ortho_vs_south"] = float(a_ortho * b_south) - float(b_ortho * a_south)
    features["same_stance"] = float(
        a_ortho == b_ortho and a_south == b_south and (a_ortho or a_south)
    )
    # Style cluster matchup interactions
    a_stk = _num_or(a_feats.get("style_striking"), 0.5)
    a_wrs = _num_or(a_feats.get("style_wrestling"), 0.15)
    a_sub = _num_or(a_feats.get("style_submission"), 0.1)
    a_cg = _num_or(a_feats.get("style_clinch_ground"), 0.32)
    b_stk = _num_or(b_feats.get("style_striking"), 0.5)
    b_wrs = _num_or(b_feats.get("style_wrestling"), 0.15)
    b_sub = _num_or(b_feats.get("style_submission"), 0.1)
    b_cg = _num_or(b_feats.get("style_clinch_ground"), 0.32)
    features["d_striker_vs_grappler"] = a_stk * (b_wrs + b_sub) - b_stk * (a_wrs + a_sub)
    features["style_mismatch"] = a_stk * (b_wrs + b_sub) + b_stk * (a_wrs + a_sub)
    features["style_distance"] = math.sqrt(
        (a_stk - b_stk)**2 + (a_wrs - b_wrs)**2 + (a_sub - b_sub)**2 + (a_cg - b_cg)**2
    )
    return features


# ─── Training data builder ────────────────────────────────────────────────────

def build_training_data(csv_path, progress_cb=None):
    """Process fights chronologically, build features, return X, y and state."""
    df = pd.read_csv(csv_path)
    df["event_date"] = pd.to_datetime(df["event_date"], format="%m/%d/%Y")
    df = df.sort_values("event_date").reset_index(drop=True)
    _ensure_fighter_feature_keys(df["event_date"].iloc[0] if len(df) else None)

    fighter_history = defaultdict(list)
    glicko_ratings = {}
    opp_glicko_list = defaultdict(list)

    rows_X = []
    rows_y = []
    total = len(df)

    for idx in range(total):
        if progress_cb and idx % 500 == 0:
            progress_cb(f"  Building features... fight {idx+1}/{total}")

        row = df.iloc[idx]
        r_name = row["r_name"]
        b_name = row["b_name"]

        # Initialize Glicko
        if r_name not in glicko_ratings:
            glicko_ratings[r_name] = (MU_0, PHI_0, SIGMA_0)
        if b_name not in glicko_ratings:
            glicko_ratings[b_name] = (MU_0, PHI_0, SIGMA_0)

        r_glicko = glicko_ratings[r_name]
        b_glicko = glicko_ratings[b_name]

        r_feats = compute_fighter_features(
            fighter_history[r_name], r_glicko,
            opp_glicko_list[r_name], row["event_date"],
            fallback_profile=_extract_profile_from_row(row, "r"),
        )
        b_feats = compute_fighter_features(
            fighter_history[b_name], b_glicko,
            opp_glicko_list[b_name], row["event_date"],
            fallback_profile=_extract_profile_from_row(row, "b"),
        )
        matchup = compute_matchup_features(
            r_feats, b_feats,
            is_title=row.get("is_title_bout", 0),
            total_rounds=row.get("total_rounds", 3),
            weight_class=row.get("weight_class", ""),
        )
        if row["winner"] in ("Red", "Blue"):
            rows_X.append(matchup)
            rows_y.append(1.0 if row["winner"] == "Red" else 0.0)

        # Determine result
        winner = row["winner"]
        if winner == "Red":
            r_res, b_res = "W", "L"
            r_sc, b_sc = 1.0, 0.0
        elif winner == "Blue":
            r_res, b_res = "L", "W"
            r_sc, b_sc = 0.0, 1.0
        else:
            r_res, b_res = "D", "D"
            r_sc, b_sc = 0.5, 0.5

        # Update histories
        fighter_history[r_name].append(extract_fight_record(row, "r", "b", r_res, b_glicko[0]))
        fighter_history[b_name].append(extract_fight_record(row, "b", "r", b_res, r_glicko[0]))

        # Track opponent Glicko
        opp_glicko_list[r_name].append(b_glicko[0])
        opp_glicko_list[b_name].append(r_glicko[0])

        # Update Glicko
        glicko_ratings[r_name] = glicko2_update(r_glicko, [(b_glicko[0], b_glicko[1], r_sc)])
        glicko_ratings[b_name] = glicko2_update(b_glicko, [(r_glicko[0], r_glicko[1], b_sc)])

    X = pd.DataFrame(rows_X)
    y = pd.Series(rows_y)

    if progress_cb:
        progress_cb(f"  Built {len(X)} training samples with {X.shape[1]} features.")

    return X, y, fighter_history, glicko_ratings, opp_glicko_list


# ─── Optuna tuning + ensemble ─────────────────────────────────────────────────

NEEDS_SCALE = {"LogReg", "MLP"}


def _clip_probs(probs, eps=1e-6):
    return np.clip(np.asarray(probs, dtype=float), eps, 1.0 - eps)


def _make_study():
    return optuna.create_study(
        direction="minimize",
        sampler=optuna.samplers.TPESampler(seed=RANDOM_SEED),
    )


def _build_catboost_params(params):
    p = {
        "iterations": params["iterations"],
        "learning_rate": params["lr"],
        "depth": params["depth"],
        "l2_leaf_reg": params["l2_leaf_reg"],
        "random_strength": params["random_strength"],
        "border_count": params.get("border_count", 128),
        "bootstrap_type": params.get("bootstrap_type", "Bayesian"),
        "loss_function": "Logloss",
        "eval_metric": "Logloss",
        "random_seed": RANDOM_SEED,
        "verbose": 0,
        "allow_writing_files": False,
    }
    if p["bootstrap_type"] == "Bayesian":
        p["bagging_temperature"] = params.get("bagging_temperature", 0.5)
    else:
        p["subsample"] = params.get("subsample", 0.8)
    return p


def _augment_swap(X, y):
    """Double the dataset by interleaving swapped-corner mirrors.

    For every row the d_* (difference) columns are negated and the label
    is flipped.  Interleaving (orig_0, swap_0, orig_1, swap_1 ...) keeps
    the chronological ordering intact for TimeSeriesSplit.
    """
    d_cols = [c for c in X.columns if c.startswith("d_")]
    X_swap = X.copy()
    X_swap[d_cols] = -X_swap[d_cols]
    y_swap = 1.0 - y

    n = len(X)
    X_aug = pd.DataFrame(
        np.empty((n * 2, X.shape[1]), dtype=np.float64), columns=X.columns,
    )
    y_aug = pd.Series(np.empty(n * 2, dtype=np.float64))
    X_aug.iloc[0::2] = X.values
    X_aug.iloc[1::2] = X_swap.values
    y_aug.iloc[0::2] = y.values
    y_aug.iloc[1::2] = y_swap.values
    return X_aug.reset_index(drop=True), y_aug.reset_index(drop=True)


def _swap_features(X):
    """Return a copy with all d_* columns negated (corner swap)."""
    X2 = X.copy()
    d_cols = [c for c in X2.columns if c.startswith("d_")]
    X2[d_cols] = -X2[d_cols]
    return X2


class EnsembleModel:
    def __init__(self, models, imputer, scaler, feat_cols, meta_model,
                 model_order, decision_threshold=0.5, missing_cols=None,
                 calibrator=None):
        self.models = models
        self.imputer = imputer
        self.scaler = scaler
        self.feat_cols = feat_cols
        self.meta_model = meta_model
        self.model_order = list(model_order)
        self.decision_threshold = float(decision_threshold)
        self.missing_cols = missing_cols or []
        self.calibrator = calibrator

    def _prepare_input(self, X_raw):
        X = _add_missingness_indicators(X_raw, self.missing_cols)
        X = X.reindex(columns=self.feat_cols)
        X_imp = pd.DataFrame(self.imputer.transform(X), columns=self.feat_cols)
        X_sc = pd.DataFrame(self.scaler.transform(X_imp), columns=self.feat_cols)
        return X_imp, X_sc

    def _predict_all(self, X_imp, X_sc):
        probas = {}
        for name, model in self.models.items():
            X_in = X_sc if name in NEEDS_SCALE else X_imp
            if isinstance(model, TabNetClassifier):
                X_in = X_in.values if hasattr(X_in, 'values') else np.asarray(X_in)
            probas[name] = _clip_probs(model.predict_proba(X_in)[:, 1])
        return probas

    def predict_proba_single(self, features_dict):
        """Symmetric prediction through the stacked meta-model."""
        X = pd.DataFrame([features_dict])
        X_imp, X_sc = self._prepare_input(X)
        p_orig = self._predict_all(X_imp, X_sc)

        X_sw = _swap_features(X)
        X_sw_imp, X_sw_sc = self._prepare_input(X_sw)
        p_swap = self._predict_all(X_sw_imp, X_sw_sc)

        meta_row = []
        for name in self.model_order:
            p_fwd = p_orig[name][0]
            p_rev = p_swap[name][0]
            meta_row.append((p_fwd + (1.0 - p_rev)) / 2.0)

        p_raw = float(self.meta_model.predict_proba(np.asarray([meta_row], dtype=float))[:, 1][0])

        if self.calibrator is not None:
            p_raw = float(self.calibrator.predict_proba(np.array([[p_raw]]))[:, 1][0])

        return float(np.clip(p_raw, 1e-6, 1 - 1e-6))


def _symmetric_probabilities(name, model, X_forward, X_swapped):
    p_fwd = _predict_model(name, model, X_forward)
    p_rev = _predict_model(name, model, X_swapped)
    return _clip_probs((p_fwd + (1.0 - p_rev)) / 2.0)


def _fit_regularized_stacker(X_train_meta, y_train_meta, X_val_meta=None, y_val_meta=None):
    c_grid = [0.03, 0.1, 0.3, 1.0, 3.0, 10.0]
    y_train_np = np.asarray(y_train_meta).astype(int)
    if X_val_meta is None or y_val_meta is None or len(X_val_meta) == 0:
        model = LogisticRegression(max_iter=2000, C=1.0, random_state=RANDOM_SEED)
        model.fit(X_train_meta, y_train_np)
        return model, 1.0, float("nan")

    best_model = None
    best_c = None
    best_score = float("inf")
    y_val_np = np.asarray(y_val_meta).astype(int)
    for c in c_grid:
        model = LogisticRegression(max_iter=2000, C=c, random_state=RANDOM_SEED)
        model.fit(X_train_meta, y_train_np)
        val_probs = _clip_probs(model.predict_proba(X_val_meta)[:, 1])
        score = log_loss(y_val_np, val_probs)
        if score < best_score:
            best_model = model
            best_c = c
            best_score = score
    return best_model, best_c, best_score


def _tune_pick_threshold(probs, y_true):
    probs = _clip_probs(probs)
    y_true = np.asarray(y_true).astype(int)
    candidates = np.unique(
        np.concatenate([
            np.linspace(0.30, 0.70, 41),
            np.round(probs, 4),
            np.array([0.5]),
        ])
    )
    best_threshold = 0.5
    best_acc = -1.0
    best_margin = float("inf")
    for threshold in candidates:
        preds = (probs >= threshold).astype(int)
        acc = accuracy_score(y_true, preds)
        margin = abs(threshold - 0.5)
        if acc > best_acc + 1e-12 or (abs(acc - best_acc) <= 1e-12 and margin < best_margin):
            best_threshold = float(threshold)
            best_acc = float(acc)
            best_margin = margin
    return best_threshold, best_acc


def _extract_stacker_shares(meta_model, model_order):
    coefs = getattr(meta_model, "coef_", np.zeros((1, len(model_order))))[0]
    total = float(np.abs(coefs).sum())
    if total <= 0:
        share = 1.0 / max(len(model_order), 1)
        return {name: share for name in model_order}, {name: 0.0 for name in model_order}
    return (
        {name: abs(float(coefs[i])) / total for i, name in enumerate(model_order)},
        {name: float(coefs[i]) for i, name in enumerate(model_order)},
    )


def _build_oof_base_probas(X_raw, y_raw, specs, progress_cb=None, stage_name="OOF"):
    oof = {name: np.full(len(X_raw), np.nan) for name, _ in specs}
    splitter = TimeSeriesSplit(n_splits=5)
    for fold_idx, (tr, val) in enumerate(splitter.split(X_raw), start=1):
        if progress_cb:
            progress_cb(f"    {stage_name} fold {fold_idx}/5...")
        X_fold_train = X_raw.iloc[tr]
        y_fold_train = y_raw.iloc[tr]
        X_fold_val = X_raw.iloc[val]
        X_fold_train_aug, y_fold_aug = _augment_swap(X_fold_train, y_fold_train)
        X_fold_val_swap = _swap_features(X_fold_val)

        scaler = StandardScaler()
        X_fold_sc = pd.DataFrame(scaler.fit_transform(X_fold_train_aug), columns=X_raw.columns)
        X_fold_val_sc = pd.DataFrame(scaler.transform(X_fold_val), columns=X_raw.columns)
        X_fold_val_swap_sc = pd.DataFrame(scaler.transform(X_fold_val_swap), columns=X_raw.columns)

        for name, make_model in specs:
            model = make_model()
            X_fit = X_fold_sc if name in NEEDS_SCALE else X_fold_train_aug
            _fit_model(name, model, X_fit, y_fold_aug)
            X_eval = X_fold_val_sc if name in NEEDS_SCALE else X_fold_val
            X_eval_swap = X_fold_val_swap_sc if name in NEEDS_SCALE else X_fold_val_swap
            oof[name][val] = _symmetric_probabilities(name, model, X_eval, X_eval_swap)

    return pd.DataFrame(oof)


def _fit_model(*args):
    """Backward-compatible model fit helper.

    Supports both:
      _fit_model(model, X, y)
      _fit_model(name, model, X, y)
    """
    if len(args) == 4:
        _, model, X, y = args
    else:
        model, X, y = args
    if isinstance(model, TabNetClassifier):
        X_np = X.values if hasattr(X, 'values') else np.asarray(X)
        y_np = np.asarray(y).astype(int)
        model.fit(X_np, y_np)
    else:
        model.fit(X, y)
    return model


def _predict_model(*args):
    """Backward-compatible proba helper.

    Supports both:
      _predict_model(model, X)
      _predict_model(name, model, X)
    """
    if len(args) == 3:
        _, model, X = args
    else:
        model, X = args
    if isinstance(model, TabNetClassifier):
        X_in = X.values if hasattr(X, 'values') else np.asarray(X)
        return _clip_probs(model.predict_proba(X_in)[:, 1])
    return _clip_probs(model.predict_proba(X)[:, 1])


def _fit_preprocessors(X_fit_raw, X_eval_raw, needs_scale=False):
    imputer = SimpleImputer(strategy="median")
    X_fit_imp = pd.DataFrame(imputer.fit_transform(X_fit_raw), columns=X_fit_raw.columns)
    X_eval_imp = pd.DataFrame(imputer.transform(X_eval_raw), columns=X_fit_raw.columns)

    scaler = StandardScaler()
    scaler.fit(X_fit_imp)
    X_fit_sc = pd.DataFrame(scaler.transform(X_fit_imp), columns=X_fit_raw.columns)
    X_eval_sc = pd.DataFrame(scaler.transform(X_eval_imp), columns=X_fit_raw.columns)

    if needs_scale:
        return X_fit_sc, X_eval_sc, imputer, scaler
    return X_fit_imp, X_eval_imp, imputer, scaler


def _transform_with_preprocessors(X_raw, imputer, scaler, feat_cols, needs_scale=False):
    X_imp = pd.DataFrame(imputer.transform(X_raw), columns=feat_cols)
    if needs_scale:
        return pd.DataFrame(scaler.transform(X_imp), columns=feat_cols)
    return X_imp


def _fit_platt_calibrator(probs, y_true):
    cal = LogisticRegression(max_iter=2000, C=1.0, random_state=RANDOM_SEED)
    cal.fit(_clip_probs(probs).reshape(-1, 1), np.asarray(y_true).astype(int))
    return cal


def _trial_bar(model_name, n_trials, status_var=None):
    """Return an Optuna callback that prints a single updating progress bar."""
    def callback(study, trial):
        done = trial.number + 1
        pct = done / n_trials
        filled = int(30 * pct)
        bar = "#" * filled + "-" * (30 - filled)
        best = study.best_value
        msg = f"  {model_name}  [{bar}] {done}/{n_trials}  best={best:.4f}"
        print(f"\r{msg}", end="", flush=True)
        if status_var:
            status_var.set(msg)
        if done == n_trials:
            print()
    return callback


def tune_and_train(X, y, n_trials=50, progress_cb=None, status_var=None):
    """Tune boosted trees, stack time-aware base models, and retrain for production."""
    n = len(X)
    test_size = int(n * 0.20)
    val_size = int(n * 0.15)
    train_size = n - val_size - test_size

    X_train_raw = X.iloc[:train_size]
    X_val_raw = X.iloc[train_size:train_size + val_size]
    X_test_raw = X.iloc[train_size + val_size:]
    y_train_raw = y.iloc[:train_size].reset_index(drop=True)
    y_val_raw = y.iloc[train_size:train_size + val_size].reset_index(drop=True)
    y_test_raw = y.iloc[train_size + val_size:].reset_index(drop=True)

    if progress_cb:
        progress_cb("")
        progress_cb(f"  Train: {train_size} | Val: {val_size} (stack/threshold tuning) | Test: {test_size} (final eval)")
        progress_cb("")

    missing_cols = _pick_missing_cols(X_train_raw)
    X_train_raw = _add_missingness_indicators(X_train_raw, missing_cols)
    X_val_raw = _add_missingness_indicators(X_val_raw, missing_cols)
    X_test_raw = _add_missingness_indicators(X_test_raw, missing_cols)
    X_full_raw = _add_missingness_indicators(X, missing_cols)

    if missing_cols and progress_cb:
        progress_cb(f"  Missingness indicators enabled for {len(missing_cols)} sparse features")

    # Impute on train, transform val and test
    imputer = SimpleImputer(strategy="median")
    X_train_orig = pd.DataFrame(imputer.fit_transform(X_train_raw), columns=X_train_raw.columns)
    X_val = pd.DataFrame(imputer.transform(X_val_raw), columns=X_train_raw.columns)
    X_test = pd.DataFrame(imputer.transform(X_test_raw), columns=X_train_raw.columns)

    selected_cols = [c for c in X_train_orig.columns if X_train_orig[c].nunique(dropna=False) > 1]
    dropped_cols = len(X_train_orig.columns) - len(selected_cols)
    if dropped_cols and progress_cb:
        progress_cb(f"  Feature pruning: dropped {dropped_cols} constant features")
    X_train_orig = X_train_orig[selected_cols]
    X_val = X_val[selected_cols]
    X_test = X_test[selected_cols]
    X_full_raw = X_full_raw[selected_cols]
    feature_cols = selected_cols

    tscv = TimeSeriesSplit(n_splits=5)

    def _median_best_rounds(study, param_name):
        rounds = study.best_trial.user_attrs.get("best_iterations", [])
        if not rounds:
            return int(study.best_params[param_name])
        return max(25, int(round(float(np.median(rounds)))))

    def _symmetric_fold_logloss(name, model, X_fold_val, y_fold_val):
        X_fold_swap = _swap_features(X_fold_val)
        probs = _symmetric_probabilities(name, model, X_fold_val, X_fold_swap)
        return log_loss(y_fold_val, probs)

    active_models = set(ACTIVE_ENSEMBLE_MODELS)

    # ═══════════════════════════════════════════════════════════════════════
    #  Optuna tuning — active models only
    # ═══════════════════════════════════════════════════════════════════════

    # ── LightGBM ──
    if "LightGBM" in active_models:
        if progress_cb:
            progress_cb(f"  --- Tuning LightGBM ({n_trials} trials) ---")

        def lgb_obj(trial):
            p = {
                "n_estimators": trial.suggest_int("n_estimators", 100, 800),
                "learning_rate": trial.suggest_float("lr", 0.01, 0.3, log=True),
                "max_depth": trial.suggest_int("max_depth", 3, 8),
                "num_leaves": trial.suggest_int("num_leaves", 15, 127),
                "min_child_samples": trial.suggest_int("min_child_samples", 5, 100),
                "subsample": trial.suggest_float("subsample", 0.5, 1.0),
                "colsample_bytree": trial.suggest_float("colsample_bytree", 0.3, 1.0),
                "reg_alpha": trial.suggest_float("reg_alpha", 1e-8, 10.0, log=True),
                "reg_lambda": trial.suggest_float("reg_lambda", 1e-8, 10.0, log=True),
                "random_state": RANDOM_SEED,
                "n_jobs": -1,
                "verbose": -1,
            }
            scores = []
            best_iterations = []
            for tr, val in tscv.split(X_train_orig):
                X_fold_train = X_train_orig.iloc[tr]
                y_fold_train = y_train_raw.iloc[tr]
                X_fold_val = X_train_orig.iloc[val]
                y_fold_val = y_train_raw.iloc[val]
                X_fold, y_fold = _augment_swap(X_fold_train, y_fold_train)
                X_eval_aug, y_eval_aug = _augment_swap(X_fold_val, y_fold_val)
                m = lgb.LGBMClassifier(**p)
                m.fit(
                    X_fold, y_fold,
                    eval_set=[(X_eval_aug, y_eval_aug)],
                    eval_metric="logloss",
                    callbacks=[lgb.early_stopping(EARLY_STOPPING_ROUNDS, verbose=False)],
                )
                best_iterations.append(getattr(m, "best_iteration_", p["n_estimators"]) or p["n_estimators"])
                scores.append(_symmetric_fold_logloss("LightGBM", m, X_fold_val, y_fold_val))
            trial.set_user_attr("best_iterations", best_iterations)
            return np.mean(scores)

        lgb_study = _make_study()
        lgb_study.optimize(lgb_obj, n_trials=n_trials, show_progress_bar=False,
                           callbacks=[_trial_bar("LightGBM", n_trials, status_var)])
        lp = lgb_study.best_params
        lp["n_estimators"] = min(lp["n_estimators"], _median_best_rounds(lgb_study, "n_estimators"))
        if progress_cb:
            progress_cb(f"  Best CV log-loss: {lgb_study.best_value:.4f}")
            progress_cb("")
    else:
        lp = None

    # ── XGBoost ──
    if "XGBoost" in active_models:
        if progress_cb:
            progress_cb(f"  --- Tuning XGBoost ({n_trials} trials) ---")

        def xgb_obj(trial):
            p = {
                "n_estimators": trial.suggest_int("n_estimators", 100, 800),
                "learning_rate": trial.suggest_float("lr", 0.01, 0.3, log=True),
                "max_depth": trial.suggest_int("max_depth", 3, 8),
                "min_child_weight": trial.suggest_int("min_child_weight", 1, 20),
                "subsample": trial.suggest_float("subsample", 0.5, 1.0),
                "colsample_bytree": trial.suggest_float("colsample_bytree", 0.3, 1.0),
                "reg_alpha": trial.suggest_float("reg_alpha", 1e-8, 10.0, log=True),
                "reg_lambda": trial.suggest_float("reg_lambda", 1e-8, 10.0, log=True),
                "eval_metric": "logloss",
                "random_state": RANDOM_SEED,
                "n_jobs": -1,
                "verbosity": 0,
                "early_stopping_rounds": EARLY_STOPPING_ROUNDS,
            }
            scores = []
            best_iterations = []
            for tr, val in tscv.split(X_train_orig):
                X_fold_train = X_train_orig.iloc[tr]
                y_fold_train = y_train_raw.iloc[tr]
                X_fold_val = X_train_orig.iloc[val]
                y_fold_val = y_train_raw.iloc[val]
                X_fold, y_fold = _augment_swap(X_fold_train, y_fold_train)
                X_eval_aug, y_eval_aug = _augment_swap(X_fold_val, y_fold_val)
                m = xgb.XGBClassifier(**p)
                m.fit(X_fold, y_fold, eval_set=[(X_eval_aug, y_eval_aug)], verbose=False)
                best_iter = getattr(m, "best_iteration", None)
                best_iterations.append((best_iter + 1) if best_iter is not None else p["n_estimators"])
                scores.append(_symmetric_fold_logloss("XGBoost", m, X_fold_val, y_fold_val))
            trial.set_user_attr("best_iterations", best_iterations)
            return np.mean(scores)

        xgb_study = _make_study()
        xgb_study.optimize(xgb_obj, n_trials=n_trials, show_progress_bar=False,
                           callbacks=[_trial_bar("XGBoost ", n_trials, status_var)])
        xp = xgb_study.best_params
        xp["n_estimators"] = min(xp["n_estimators"], _median_best_rounds(xgb_study, "n_estimators"))
        if progress_cb:
            progress_cb(f"  Best CV log-loss: {xgb_study.best_value:.4f}")
            progress_cb("")
    else:
        xp = None

    # ── CatBoost ──
    if "CatBoost" in active_models:
        if progress_cb:
            progress_cb(f"  --- Tuning CatBoost ({n_trials} trials) ---")

        def cb_obj(trial):
            p = {
                "iterations": trial.suggest_int("iterations", 100, 800),
                "lr": trial.suggest_float("lr", 0.01, 0.3, log=True),
                "depth": trial.suggest_int("depth", 3, 8),
                "l2_leaf_reg": trial.suggest_float("l2_leaf_reg", 1e-3, 10.0, log=True),
                "random_strength": trial.suggest_float("random_strength", 0.01, 3.0, log=True),
                "border_count": trial.suggest_int("border_count", 32, 255),
                "bootstrap_type": trial.suggest_categorical("bootstrap_type", ["Bayesian", "Bernoulli"]),
            }
            if p["bootstrap_type"] == "Bayesian":
                p["bagging_temperature"] = trial.suggest_float("bagging_temperature", 0.0, 2.0)
            else:
                p["subsample"] = trial.suggest_float("subsample", 0.6, 1.0)
            scores = []
            best_iterations = []
            for tr, val in tscv.split(X_train_orig):
                X_fold_train = X_train_orig.iloc[tr]
                y_fold_train = y_train_raw.iloc[tr]
                X_fold_val = X_train_orig.iloc[val]
                y_fold_val = y_train_raw.iloc[val]
                X_fold, y_fold = _augment_swap(X_fold_train, y_fold_train)
                X_eval_aug, y_eval_aug = _augment_swap(X_fold_val, y_fold_val)
                m = cb.CatBoostClassifier(**_build_catboost_params(p))
                m.fit(
                    X_fold, y_fold,
                    eval_set=(X_eval_aug, y_eval_aug),
                    verbose=0, use_best_model=True,
                    early_stopping_rounds=EARLY_STOPPING_ROUNDS,
                )
                best_iter = m.get_best_iteration()
                best_iterations.append((best_iter + 1) if best_iter is not None and best_iter >= 0 else p["iterations"])
                scores.append(_symmetric_fold_logloss("CatBoost", m, X_fold_val, y_fold_val))
            trial.set_user_attr("best_iterations", best_iterations)
            return np.mean(scores)

        cb_study = _make_study()
        cb_study.optimize(cb_obj, n_trials=n_trials, show_progress_bar=False,
                          callbacks=[_trial_bar("CatBoost", n_trials, status_var)])
        if cb_study.best_value < 0.693:
            cp = cb_study.best_params
            cp["iterations"] = min(cp["iterations"], _median_best_rounds(cb_study, "iterations"))
        else:
            cp = {"iterations": 500, "lr": 0.05, "depth": 6, "l2_leaf_reg": 3.0,
                  "random_strength": 1.0, "bagging_temperature": 0.5,
                  "border_count": 128, "bootstrap_type": "Bayesian"}
        if progress_cb:
            progress_cb(f"  Best CV log-loss: {cb_study.best_value:.4f}"
                        + (" (using defaults)" if cb_study.best_value >= 0.693 else ""))
            progress_cb("")
    else:
        cp = None

    n_sec = max(n_trials * 2 // 3, 20)

    # ── ExtraTrees ──
    if "ExtraTrees" in active_models:
        if progress_cb:
            progress_cb(f"  --- Tuning ExtraTrees ({n_sec} trials) ---")

        def et_obj(trial):
            p = {
                "n_estimators": trial.suggest_int("n_estimators", 200, 1000),
                "max_depth": trial.suggest_int("max_depth", 6, 20),
                "min_samples_leaf": trial.suggest_int("min_samples_leaf", 2, 30),
                "min_samples_split": trial.suggest_int("min_samples_split", 2, 20),
                "max_features": trial.suggest_float("max_features", 0.3, 1.0),
                "random_state": RANDOM_SEED, "n_jobs": -1,
            }
            scores = []
            for tr, val in tscv.split(X_train_orig):
                X_fold_train = X_train_orig.iloc[tr]
                y_fold_train = y_train_raw.iloc[tr]
                X_fold_val = X_train_orig.iloc[val]
                y_fold_val = y_train_raw.iloc[val]
                X_fold, y_fold = _augment_swap(X_fold_train, y_fold_train)
                m = ExtraTreesClassifier(**p)
                m.fit(X_fold, y_fold)
                scores.append(_symmetric_fold_logloss("ExtraTrees", m, X_fold_val, y_fold_val))
            return np.mean(scores)

        et_study = _make_study()
        et_study.optimize(et_obj, n_trials=n_sec, show_progress_bar=False,
                          callbacks=[_trial_bar("ExtraTrees", n_sec, status_var)])
        if et_study.best_value < 0.693:
            ep = et_study.best_params
        else:
            ep = {"n_estimators": 500, "max_depth": 12, "min_samples_leaf": 5,
                  "min_samples_split": 2, "max_features": 0.7}
        if progress_cb:
            progress_cb(f"  Best CV log-loss: {et_study.best_value:.4f}"
                        + (" (using defaults)" if et_study.best_value >= 0.693 else ""))
            progress_cb("")
    else:
        ep = None

    # ── HistGBM ──
    if "HistGBM" in active_models:
        if progress_cb:
            progress_cb(f"  --- Tuning HistGBM ({n_sec} trials) ---")

        def hgbm_obj(trial):
            p = {
                "max_iter": trial.suggest_int("max_iter", 200, 800),
                "learning_rate": trial.suggest_float("lr", 0.01, 0.3, log=True),
                "max_depth": trial.suggest_int("max_depth", 3, 10),
                "max_leaf_nodes": trial.suggest_int("max_leaf_nodes", 15, 127),
                "min_samples_leaf": trial.suggest_int("min_samples_leaf", 5, 50),
                "l2_regularization": trial.suggest_float("l2_reg", 1e-8, 10.0, log=True),
                "random_state": RANDOM_SEED,
            }
            scores = []
            for tr, val in tscv.split(X_train_orig):
                X_fold_train = X_train_orig.iloc[tr]
                y_fold_train = y_train_raw.iloc[tr]
                X_fold_val = X_train_orig.iloc[val]
                y_fold_val = y_train_raw.iloc[val]
                X_fold, y_fold = _augment_swap(X_fold_train, y_fold_train)
                m = HistGradientBoostingClassifier(**p)
                m.fit(X_fold, y_fold)
                scores.append(_symmetric_fold_logloss("HistGBM", m, X_fold_val, y_fold_val))
            return np.mean(scores)

        hgbm_study = _make_study()
        hgbm_study.optimize(hgbm_obj, n_trials=n_sec, show_progress_bar=False,
                            callbacks=[_trial_bar("HistGBM ", n_sec, status_var)])
        if hgbm_study.best_value < 0.693:
            hgp = hgbm_study.best_params
        else:
            hgp = {"max_iter": 500, "lr": 0.05, "max_depth": 6,
                   "max_leaf_nodes": 31, "min_samples_leaf": 20, "l2_reg": 1.0}
        if progress_cb:
            progress_cb(f"  Best CV log-loss: {hgbm_study.best_value:.4f}"
                        + (" (using defaults)" if hgbm_study.best_value >= 0.693 else ""))
            progress_cb("")
    else:
        hgp = None

    # ── RandomForest ──
    if "RandForest" in active_models:
        if progress_cb:
            progress_cb(f"  --- Tuning RandomForest ({n_sec} trials) ---")

        def rf_obj(trial):
            p = {
                "n_estimators": trial.suggest_int("n_estimators", 200, 1000),
                "max_depth": trial.suggest_int("max_depth", 6, 20),
                "min_samples_leaf": trial.suggest_int("min_samples_leaf", 2, 30),
                "min_samples_split": trial.suggest_int("min_samples_split", 2, 20),
                "max_features": trial.suggest_float("max_features", 0.3, 1.0),
                "random_state": RANDOM_SEED, "n_jobs": -1,
            }
            scores = []
            for tr, val in tscv.split(X_train_orig):
                X_fold_train = X_train_orig.iloc[tr]
                y_fold_train = y_train_raw.iloc[tr]
                X_fold_val = X_train_orig.iloc[val]
                y_fold_val = y_train_raw.iloc[val]
                X_fold, y_fold = _augment_swap(X_fold_train, y_fold_train)
                m = RandomForestClassifier(**p)
                m.fit(X_fold, y_fold)
                scores.append(_symmetric_fold_logloss("RandForest", m, X_fold_val, y_fold_val))
            return np.mean(scores)

        rf_study = _make_study()
        rf_study.optimize(rf_obj, n_trials=n_sec, show_progress_bar=False,
                          callbacks=[_trial_bar("RandForst", n_sec, status_var)])
        if rf_study.best_value < 0.693:
            rp = rf_study.best_params
        else:
            rp = {"n_estimators": 500, "max_depth": 12, "min_samples_leaf": 5,
                  "min_samples_split": 2, "max_features": 0.7}
        if progress_cb:
            progress_cb(f"  Best CV log-loss: {rf_study.best_value:.4f}"
                        + (" (using defaults)" if rf_study.best_value >= 0.693 else ""))
            progress_cb("")
    else:
        rp = None

    # ═══════════════════════════════════════════════════════════════════════
    #  Build model specs and time-aware OOF stack features
    # ═══════════════════════════════════════════════════════════════════════

    def _make_specs():
        specs = []
        if "LightGBM" in active_models and lp is not None:
            specs.append(("LightGBM", lambda: lgb.LGBMClassifier(
                n_estimators=lp["n_estimators"], learning_rate=lp["lr"],
                max_depth=lp["max_depth"], num_leaves=lp["num_leaves"],
                min_child_samples=lp["min_child_samples"],
                subsample=lp["subsample"], colsample_bytree=lp["colsample_bytree"],
                reg_alpha=lp["reg_alpha"], reg_lambda=lp["reg_lambda"],
                random_state=RANDOM_SEED, n_jobs=-1, verbose=-1)))
        if "XGBoost" in active_models and xp is not None:
            specs.append(("XGBoost", lambda: xgb.XGBClassifier(
                n_estimators=xp["n_estimators"], learning_rate=xp["lr"],
                max_depth=xp["max_depth"], min_child_weight=xp["min_child_weight"],
                subsample=xp["subsample"], colsample_bytree=xp["colsample_bytree"],
                reg_alpha=xp["reg_alpha"], reg_lambda=xp["reg_lambda"],
                eval_metric="logloss", random_state=RANDOM_SEED,
                n_jobs=-1, verbosity=0)))
        if "CatBoost" in active_models and cp is not None:
            specs.append(("CatBoost", lambda: cb.CatBoostClassifier(**_build_catboost_params(cp))))
        if "HistGBM" in active_models and hgp is not None:
            specs.append(("HistGBM", lambda: HistGradientBoostingClassifier(
                max_iter=hgp["max_iter"], learning_rate=hgp["lr"],
                max_depth=hgp["max_depth"], max_leaf_nodes=hgp["max_leaf_nodes"],
                min_samples_leaf=hgp["min_samples_leaf"],
                l2_regularization=hgp["l2_reg"],
                random_state=RANDOM_SEED)))
        if "RandForest" in active_models and rp is not None:
            specs.append(("RandForest", lambda: RandomForestClassifier(
                n_estimators=rp["n_estimators"], max_depth=rp["max_depth"],
                min_samples_leaf=rp["min_samples_leaf"],
                min_samples_split=rp["min_samples_split"],
                max_features=rp["max_features"],
                random_state=RANDOM_SEED, n_jobs=-1)))
        if "ExtraTrees" in active_models and ep is not None:
            specs.append(("ExtraTrees", lambda: ExtraTreesClassifier(
                n_estimators=ep["n_estimators"], max_depth=ep["max_depth"],
                min_samples_leaf=ep["min_samples_leaf"],
                min_samples_split=ep["min_samples_split"],
                max_features=ep["max_features"],
                random_state=RANDOM_SEED, n_jobs=-1)))
        if "MLP" in active_models:
            specs.append(("MLP", lambda: MLPClassifier(
                hidden_layer_sizes=(128, 64, 32), max_iter=500,
                early_stopping=True, random_state=RANDOM_SEED, learning_rate="adaptive")))
        if "TabNet" in active_models:
            specs.append(("TabNet", lambda: TabNetClassifier(verbose=0, seed=RANDOM_SEED)))
        if "SVM" in active_models:
            specs.append(("SVM", lambda: SVC(
                kernel="rbf", probability=True, C=1.0, random_state=RANDOM_SEED)))
        if "AdaBoost" in active_models:
            specs.append(("AdaBoost", lambda: AdaBoostClassifier(
                n_estimators=200, learning_rate=0.1, random_state=RANDOM_SEED)))
        if "LogReg" in active_models:
            specs.append(("LogReg", lambda: LogisticRegression(max_iter=2000, C=1.0, random_state=RANDOM_SEED)))
        return specs

    specs = _make_specs()
    model_order = [name for name, _ in specs]
    X_dev = pd.concat([X_train_orig, X_val], ignore_index=True)
    y_dev = pd.concat([y_train_raw, y_val_raw], ignore_index=True)

    if progress_cb:
        progress_cb("  --- Building OOF stack features on development split ---")
    dev_oof = _build_oof_base_probas(X_dev, y_dev, specs, progress_cb=progress_cb, stage_name="Stacking")
    valid_meta_mask = ~dev_oof.isna().any(axis=1)
    dev_indices = np.arange(len(X_dev))
    meta_train_mask = valid_meta_mask & (dev_indices < len(X_train_orig))
    meta_val_mask = valid_meta_mask & (dev_indices >= len(X_train_orig))
    if meta_train_mask.sum() == 0 or meta_val_mask.sum() == 0:
        valid_indices = np.where(valid_meta_mask)[0]
        split_at = max(1, int(len(valid_indices) * 0.75))
        meta_train_mask = np.zeros(len(X_dev), dtype=bool)
        meta_val_mask = np.zeros(len(X_dev), dtype=bool)
        meta_train_mask[valid_indices[:split_at]] = True
        meta_val_mask[valid_indices[split_at:]] = True

    X_meta_train = dev_oof.loc[meta_train_mask, model_order]
    y_meta_train = y_dev.loc[meta_train_mask]
    X_meta_val = dev_oof.loc[meta_val_mask, model_order]
    y_meta_val = y_dev.loc[meta_val_mask]

    stacker, stacker_c, val_ll = _fit_regularized_stacker(
        X_meta_train, y_meta_train, X_meta_val, y_meta_val
    )
    val_stack = _clip_probs(stacker.predict_proba(X_meta_val)[:, 1])
    decision_threshold, val_acc = _tune_pick_threshold(val_stack, y_meta_val)

    if progress_cb:
        progress_cb(f"  Stacker ready. | Val acc: {val_acc:.1%} | Val log-loss: {val_ll:.4f} | Threshold: {decision_threshold:.3f}")
        progress_cb("")

    # ═══════════════════════════════════════════════════════════════════════
    #  Train all base models on development split, evaluate on test
    # ═══════════════════════════════════════════════════════════════════════
    if progress_cb:
        progress_cb("  --- Training all base models on development split ---")

    X_dev_aug, y_dev_aug = _augment_swap(X_dev, y_dev)
    scaler_eval = StandardScaler()
    X_dev_sc = pd.DataFrame(scaler_eval.fit_transform(X_dev_aug), columns=feature_cols)
    X_test_sc = pd.DataFrame(scaler_eval.transform(X_test), columns=feature_cols)
    X_test_swap = _swap_features(X_test)
    X_test_swap_sc = pd.DataFrame(scaler_eval.transform(X_test_swap), columns=feature_cols)

    test_probas = {}
    for name, make_model in specs:
        if progress_cb:
            progress_cb(f"    {name}...")
        model = make_model()
        X_fit = X_dev_sc if name in NEEDS_SCALE else X_dev_aug
        _fit_model(name, model, X_fit, y_dev_aug)
        X_eval = X_test_sc if name in NEEDS_SCALE else X_test
        X_eval_swap = X_test_swap_sc if name in NEEDS_SCALE else X_test_swap
        test_probas[name] = _symmetric_probabilities(name, model, X_eval, X_eval_swap)

    test_meta = pd.DataFrame(test_probas)[model_order]
    ensemble_probs = _clip_probs(stacker.predict_proba(test_meta)[:, 1])
    ensemble_ll = log_loss(y_test_raw, ensemble_probs)
    ensemble_acc = accuracy_score(y_test_raw, (ensemble_probs >= decision_threshold).astype(int))
    test_brier = brier_score_loss(y_test_raw, ensemble_probs)
    test_ece = _expected_calibration_error(y_test_raw, ensemble_probs)
    stack_shares, _ = _extract_stacker_shares(stacker, model_order)

    if progress_cb:
        progress_cb("")
        progress_cb(f"  --- Test Set Results ({test_size} fights, never seen during tuning/stacking) ---")
        progress_cb(f"  {'Model':<14} {'Accuracy':>10} {'Log-Loss':>10} {'Stack':>8}")
        progress_cb(f"  {'-'*46}")
        for name in model_order:
            acc = accuracy_score(y_test_raw, (test_probas[name] >= 0.5).astype(int))
            ll = log_loss(y_test_raw, test_probas[name])
            progress_cb(f"  {name:<14} {acc:>9.1%} {ll:>10.4f} {stack_shares[name]:>7.0%}")
        progress_cb(f"  {'-'*46}")
        progress_cb(f"  {'Stacked':<14} {ensemble_acc:>9.1%} {ensemble_ll:>10.4f}  thr={decision_threshold:.3f}")
        progress_cb(f"  Val log-loss: {val_ll:.4f} | Test Brier: {test_brier:.4f} | Test ECE: {test_ece:.4f}")
        progress_cb("")

    # ═══════════════════════════════════════════════════════════════════════
    #  Retrain all models on ALL data for production predictions
    # ═══════════════════════════════════════════════════════════════════════
    if progress_cb:
        progress_cb("  --- Retraining all models on full data ---")

    imputer_full = SimpleImputer(strategy="median")
    X_all_orig = pd.DataFrame(imputer_full.fit_transform(X_full_raw), columns=feature_cols)
    y_all_orig = y.reset_index(drop=True)
    X_all, y_all = _augment_swap(X_all_orig, y_all_orig)
    scaler_full = StandardScaler()
    X_all_sc = pd.DataFrame(scaler_full.fit_transform(X_all), columns=feature_cols)

    final_models = {}
    for name, make_model in specs:
        if progress_cb:
            progress_cb(f"    {name}...")
        model = make_model()
        X_fit = X_all_sc if name in NEEDS_SCALE else X_all
        _fit_model(name, model, X_fit, y_all)
        final_models[name] = model

    final_stacker = LogisticRegression(max_iter=2000, C=stacker_c or 1.0, random_state=RANDOM_SEED)
    final_stacker.fit(dev_oof.loc[valid_meta_mask, model_order], y_dev.loc[valid_meta_mask].astype(int))

    if progress_cb:
        progress_cb("")
        progress_cb(f"  Model ready. | Test accuracy: {ensemble_acc:.1%} | Test log-loss: {ensemble_ll:.4f}")
        progress_cb("")

    return EnsembleModel(
        final_models, imputer_full, scaler_full, feature_cols, final_stacker,
        model_order, decision_threshold=decision_threshold,
        missing_cols=missing_cols, calibrator=None,
    )


# ─── Fuzzy name matching ──────────────────────────────────────────────────────

def fuzzy_find(name, fighter_history):
    if name in fighter_history and fighter_history[name]:
        return name
    lower = name.lower()
    for key in fighter_history:
        if key.lower() == lower and fighter_history[key]:
            return key
    matches = [k for k in fighter_history if lower in k.lower() and fighter_history[k]]
    if len(matches) == 1:
        return matches[0]
    return None


# ─── Excel export ─────────────────────────────────────────────────────────────

def _auto_width(ws):
    for col_idx in range(1, ws.max_column + 1):
        mx = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    mx = max(mx, len(str(cell.value or "")))
                except Exception:
                    pass
        ws.column_dimensions[get_column_letter(col_idx)].width = mx + 3


def export_to_excel(output_path, predictions):
    wb = Workbook()
    hdr_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    hdr_font = Font(bold=True, size=11, color="000000")
    hdr_align = Alignment(horizontal="left", vertical="center")
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    pct_fmt = "0.0%"

    ws = wb.active
    ws.title = "Predictions"

    headers = ["Red Corner", "Blue Corner", "Weight Class", "Predicted Winner", "Win%"]
    for ci, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = border

    for ri, p in enumerate(predictions, 2):
        winner = p.get("predicted_winner") or (p["name_a"] if p["prob_a"] >= 0.5 else p["name_b"])
        win_pct = p["prob_a"] if winner == p["name_a"] else (1.0 - p["prob_a"])
        data = [
            p["name_a"], p["name_b"], p.get("weight_class", ""),
            winner, win_pct,
        ]
        for ci, val in enumerate(data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border
            if ci == 5:
                cell.number_format = pct_fmt

    _auto_width(ws)
    wb.save(output_path)


# ─── GUI ──────────────────────────────────────────────────────────────────────

BG = "#1a1a2e"
BG_HEADER = "#16213e"
BG_INPUT = "#0f3460"
FG = "#e0e0e0"
ACCENT = "#e94560"
ACCENT2 = "#533483"
MUTED = "#a0c4ff"
GREEN = "#4ade80"
BAR_A = "#e94560"
BAR_B = "#3b82f6"


class MLPredictorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("UFC ML FIGHT PREDICTOR")
        self.root.geometry("960x780")
        self.root.resizable(True, True)
        self.root.configure(bg=BG)

        self.model = None
        self.fighter_history = None
        self.glicko_ratings = None
        self.opp_glicko_list = None
        self._training = False

        self._build_ui()

    # ── UI ──

    def _build_ui(self):
        title_frame = tk.Frame(self.root, bg=BG_HEADER, pady=12)
        title_frame.pack(fill="x")
        tk.Label(
            title_frame, text="UFC ML FIGHT PREDICTOR",
            font=("Helvetica", 20, "bold"), fg=ACCENT, bg=BG_HEADER,
        ).pack()

        main = tk.Frame(self.root, bg=BG, padx=16, pady=10)
        main.pack(fill="both", expand=True)

        self.status_var = tk.StringVar(value="Enter matchups and click Predict.")
        self.status_label = tk.Label(
            main, textvariable=self.status_var, bg=BG, fg=MUTED,
            font=("Helvetica", 9, "italic"),
        )
        self.status_label.pack(anchor="w", pady=(0, 4))

        tk.Label(
            main,
            text="Enter matchups (Red Corner,BLue Corner,Weight Class,Gender,Rounds):",
            bg=BG, fg=MUTED, font=("Helvetica", 9, "italic"),
        ).pack(anchor="w", pady=(4, 2))

        self.fight_input = tk.Text(
            main, height=10, font=("Courier New", 10),
            bg=BG_INPUT, fg="white", insertbackground="white",
            relief="flat", wrap="word",
        )
        self.fight_input.pack(fill="x", pady=4)

        btn_frame = tk.Frame(main, bg=BG)
        btn_frame.pack(fill="x", pady=6)

        tk.Button(
            btn_frame, text="Load from Fights.csv",
            command=self._load_fights_csv,
            font=("Helvetica", 10, "bold"), bg=ACCENT2, fg="white",
            relief="flat", padx=14, cursor="hand2",
        ).pack(side="left", padx=4)

        tk.Button(
            btn_frame, text="Clear", command=self._clear,
            font=("Helvetica", 10, "bold"), bg="#444466", fg="white",
            relief="flat", padx=14, cursor="hand2",
        ).pack(side="left", padx=4)

        self.predict_btn = tk.Button(
            btn_frame, text="Predict", command=self._predict,
            font=("Helvetica", 11, "bold"), bg=ACCENT, fg="white",
            relief="flat", padx=20, cursor="hand2",
        )
        self.predict_btn.pack(side="right", padx=4)

        results_frame = tk.Frame(main, bg=BG)
        results_frame.pack(fill="both", expand=True, pady=(6, 0))

        scrollbar = tk.Scrollbar(results_frame)
        scrollbar.pack(side="right", fill="y")

        self.results_canvas = tk.Canvas(
            results_frame, bg=BG, highlightthickness=0,
            yscrollcommand=scrollbar.set,
        )
        self.results_canvas.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=self.results_canvas.yview)

        self.results_inner = tk.Frame(self.results_canvas, bg=BG)
        self.results_canvas.create_window((0, 0), window=self.results_inner, anchor="nw")
        self.results_inner.bind(
            "<Configure>",
            lambda _: self.results_canvas.configure(
                scrollregion=self.results_canvas.bbox("all")),
        )
        self.results_canvas.bind_all(
            "<MouseWheel>",
            lambda e: self.results_canvas.yview_scroll(-1 * (e.delta // 120), "units"),
        )

    # ── Helpers ──

    def _log(self, msg):
        """Print to terminal AND update GUI status bar."""
        print(msg, flush=True)
        self.status_var.set(msg)

    # ── Actions ──

    def _load_fights_csv(self):
        fights_path = os.path.join(SCRIPT_DIR, "Fights.csv")
        if not os.path.exists(fights_path):
            self._log("Fights.csv not found in script directory.")
            return
        self.fight_input.delete("1.0", tk.END)
        with open(fights_path, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader, None)  # skip header
            for row in reader:
                if not row or not row[0].strip():
                    continue
                self.fight_input.insert(tk.END, ",".join(c.strip() for c in row) + "\n")

    def _clear(self):
        self.fight_input.delete("1.0", tk.END)
        for w in self.results_inner.winfo_children():
            w.destroy()

    def _predict(self):
        text = self.fight_input.get("1.0", tk.END).strip()
        if not text:
            self._log("Enter at least one matchup.")
            return
        if self._training:
            return

        self._training = True
        self.predict_btn.config(state="disabled")

        # Clear previous results
        for w in self.results_inner.winfo_children():
            w.destroy()

        def _do():
            try:
                self._log("")
                self._log("  ==========================================")
                self._log("       UFC ML FIGHT PREDICTOR")
                self._log("  ==========================================")
                self._log("")

                # ── Step 1: Build training data ──
                data_path = os.path.join(SCRIPT_DIR, "pure_fight_data.csv")
                if not os.path.exists(data_path):
                    self._log(f"  Error: {data_path} not found")
                    return

                self._log("  Loading data...")
                X, y, fh, gr, og = build_training_data(data_path, self._log)
                self.fighter_history = fh
                self.glicko_ratings = gr
                self.opp_glicko_list = og

                # ── Step 2: Tune and train ──
                self.model = tune_and_train(X, y, n_trials=50, progress_cb=self._log,
                                           status_var=self.status_var)

                # ── Step 3: Predict matchups ──
                self._log("  Predicting matchups...")

                lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
                warnings_list = []
                predictions = []

                for line in lines:
                    parts = [p.strip() for p in line.split(",")]
                    if len(parts) < 2:
                        continue
                    a_name, b_name = parts[0], parts[1]
                    weight_class = parts[2] if len(parts) > 2 else ""
                    rounds_sched = int(parts[4]) if len(parts) > 4 else 3
                    is_title = 1 if rounds_sched == 5 else 0

                    a_key = fuzzy_find(a_name, self.fighter_history)
                    b_key = fuzzy_find(b_name, self.fighter_history)

                    if not a_key:
                        warnings_list.append(f"{a_name}: not found, using defaults")
                        self._log(f"    WARNING: {a_name} not found, using default ratings")
                    if not b_key:
                        warnings_list.append(f"{b_name}: not found, using defaults")
                        self._log(f"    WARNING: {b_name} not found, using default ratings")

                    a_display = a_key or a_name
                    b_display = b_key or b_name

                    a_hist = self.fighter_history.get(a_key, []) if a_key else []
                    b_hist = self.fighter_history.get(b_key, []) if b_key else []
                    a_glicko = self.glicko_ratings.get(a_key, (MU_0, PHI_0, SIGMA_0)) if a_key else (MU_0, PHI_0, SIGMA_0)
                    b_glicko = self.glicko_ratings.get(b_key, (MU_0, PHI_0, SIGMA_0)) if b_key else (MU_0, PHI_0, SIGMA_0)
                    a_opp_g = self.opp_glicko_list.get(a_key, []) if a_key else []
                    b_opp_g = self.opp_glicko_list.get(b_key, []) if b_key else []

                    now = datetime.now()
                    a_feats = compute_fighter_features(a_hist, a_glicko, a_opp_g, now)
                    b_feats = compute_fighter_features(b_hist, b_glicko, b_opp_g, now)
                    matchup = compute_matchup_features(a_feats, b_feats, is_title, rounds_sched, weight_class)

                    prob_a = self.model.predict_proba_single(matchup)
                    prob_b = 1.0 - prob_a
                    pick_a = prob_a >= self.model.decision_threshold
                    predicted_winner = a_display if pick_a else b_display

                    predictions.append({
                        "name_a": a_display, "name_b": b_display,
                        "prob_a": prob_a,
                        "predicted_winner": predicted_winner,
                        "weight_class": weight_class, "rounds": rounds_sched,
                    })

                    self._add_matchup_card(a_display, b_display, weight_class, prob_a, prob_b, pick_a)

                # ── Step 4: Export ──
                output_path = os.path.join(SCRIPT_DIR, "ML_Predictions.xlsx")
                try:
                    export_to_excel(output_path, predictions)
                    export_msg = f"Saved to {os.path.basename(output_path)}"
                except Exception as e:
                    export_msg = f"Export failed: {e}"
                    messagebox.showerror("Export Error", str(e))

                self._log("")
                self._log(f"  {len(predictions)} matchups predicted. {export_msg}")
                if warnings_list:
                    self._log(f"  Warnings: {'; '.join(warnings_list[:3])}")
                self._log("")

            except Exception as e:
                self._log(f"Error: {e}")
                import traceback; traceback.print_exc()
            finally:
                self._training = False
                self.predict_btn.config(state="normal")

        threading.Thread(target=_do, daemon=True).start()

    def _add_matchup_card(self, name_a, name_b, weight_class, prob_a, prob_b, pick_a):
        card = tk.Frame(self.results_inner, bg=BG_HEADER, pady=8, padx=12)
        card.pack(fill="x", pady=4, padx=4)

        names_frame = tk.Frame(card, bg=BG_HEADER)
        names_frame.pack(fill="x")

        a_color = GREEN if pick_a else FG
        b_color = GREEN if not pick_a else FG

        tk.Label(
            names_frame, text=name_a,
            font=("Helvetica", 11, "bold"), fg=a_color, bg=BG_HEADER, anchor="w",
        ).pack(side="left")

        tk.Label(
            names_frame, text="vs",
            font=("Helvetica", 10), fg="#666", bg=BG_HEADER,
        ).pack(side="left", padx=10)

        tk.Label(
            names_frame, text=name_b,
            font=("Helvetica", 11, "bold"), fg=b_color, bg=BG_HEADER, anchor="e",
        ).pack(side="left")

        if weight_class:
            tk.Label(
                names_frame, text=f"  [{weight_class}]",
                font=("Helvetica", 9), fg=MUTED, bg=BG_HEADER,
            ).pack(side="left", padx=(10, 0))

        # Probability bar
        bar_frame = tk.Frame(card, bg=BG_HEADER, pady=4)
        bar_frame.pack(fill="x")

        BAR_W, BAR_H = 500, 24
        bar_c = tk.Canvas(bar_frame, width=BAR_W, height=BAR_H, bg=BG_HEADER, highlightthickness=0)
        bar_c.pack(anchor="w")

        a_w = max(1, int(prob_a * BAR_W))
        bar_c.create_rectangle(0, 0, a_w, BAR_H, fill=BAR_A, outline="")
        bar_c.create_rectangle(a_w, 0, BAR_W, BAR_H, fill=BAR_B, outline="")

        if prob_a >= 0.12:
            bar_c.create_text(a_w // 2, BAR_H // 2, text=f"{prob_a:.1%}",
                              fill="white", font=("Helvetica", 9, "bold"))
        if prob_b >= 0.12:
            bar_c.create_text(a_w + (BAR_W - a_w) // 2, BAR_H // 2, text=f"{prob_b:.1%}",
                              fill="white", font=("Helvetica", 9, "bold"))

        pick = f"{name_a} ({prob_a:.1%})" if pick_a else f"{name_b} ({prob_b:.1%})"
        tk.Label(
            card, text=f"Pick: {pick}",
            font=("Helvetica", 10), fg=GREEN, bg=BG_HEADER, anchor="w",
        ).pack(anchor="w")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    root = tk.Tk()
    MLPredictorGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
